# -*- coding: utf-8 -*-
"""
auditor_comments.py — build the '검토의견' sheet (added LAST): one row per
심의위원 (auditor) per file, with the auditor's review-comment text, best-effort
분야 + 채택/반영/미반영 counts (from the 총괄표). File-level columns are merged
cells spanning each file's rows. Re-reads text from simui; idempotent.
"""
import os, re
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

BASE = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(BASE, "simui_2022-2024.xlsx")
LIMIT = 32767

SPLIT = re.compile(r"[□○▣\-]?\s*위\s*원\s*명\s*[:：]")           # member-section marker
HDR = re.compile(r"^\s*(?:분\s*야|연\s*번|구\s*분).{0,70}?미\s*반\s*영", re.S)  # table-header boilerplate

def clean(v):
    return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

def field_of(body):
    m = re.match(r"\s*([가-힣][가-힣 ]{1,11}?)\s*\d", body)
    return re.sub(r"\s+", "", m.group(1)) if m else ""

def auditors(text):
    """Merge per-page repeats: one entry per distinct 위원, comment text joined."""
    mks = list(SPLIT.finditer(text))
    merged = OrderedDict()
    for i, mk in enumerate(mks):
        s = mk.end()
        e = mks[i + 1].start() if i + 1 < len(mks) else len(text)
        seg = text[s:e]
        nm = re.match(r"\s*([가-힣]{2,3})", seg)
        if not nm:
            continue
        body = seg
        f = re.search(r"작\s*성\s*자|확\s*인\s*자|※", body)
        if f:
            body = body[:f.start()]
        body = re.sub(r"^\s*[가-힣]{2,3}\s*위?원?", "", body)   # drop the name itself
        h = HDR.match(body)
        if h:
            body = body[h.end():]                              # drop table header
        body = re.sub(r"\n{2,}", "\n", body).strip()
        merged.setdefault(nm.group(1), []).append(body)
    return [(n, " \n".join(b).strip()) for n, b in merged.items()]

def summary_counts(text):
    """name -> [채택, 반영, 미반영] best-effort from the 총괄표."""
    m = re.search(r"총\s*괄\s*표", text)
    if not m:
        return {}
    reg = text[m.end():m.end() + 1500]
    cut = re.search(r"위\s*원\s*명|조치결과\s*보고|< *목", reg)
    if cut:
        reg = reg[:cut.start()]
    out = {}
    for nm in re.finditer(r"([가-힣]{2,3})\s*위원", reg):
        if nm.group(1) in ("심의", "분야"):
            continue
        out[nm.group(1)] = [0 if x == "-" else int(x)
                            for x in re.findall(r"(\d+|-)", reg[nm.end():nm.end() + 25])[:3]]
    return out

wb = load_workbook(XL)
src = wb["simui"]
col = {c.value: i + 1 for i, c in enumerate(src[1])}
FILECOLS = ["source_filename", "year", "category_auto", "발주기관", "안건명"]
AUDCOLS = ["심의위원", "분야", "채택의견", "반영", "미반영", "검토의견"]
HEADERS = FILECOLS + AUDCOLS

groups = []  # (file_meta, [ (name, field, chae, ban, mi, body), ... ])
for r in range(2, src.max_row + 1):
    meta = {c: (src.cell(r, col[c]).value or "") for c in FILECOLS}
    text = src.cell(r, col["full_text"]).value or ""
    auds = auditors(text)
    cnts = summary_counts(text)
    if not auds:
        groups.append((meta, [("(미파싱)", "", "", "", "", "")]))
        continue
    rws = []
    for name, body in auds:
        c = cnts.get(name, [])
        body = body[:LIMIT - 1] if len(body) > LIMIT else body
        rws.append((name, field_of(body),
                    c[0] if len(c) >= 1 else "",
                    c[1] if len(c) >= 2 else "",
                    c[2] if len(c) >= 3 else "",
                    body))
    groups.append((meta, rws))

if "검토의견" in wb.sheetnames:
    del wb["검토의견"]
ws = wb.create_sheet("검토의견")  # appended as the LAST sheet
ws.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    ws.cell(1, c).font = Font(bold=True)
file_align = Alignment(vertical="center", wrap_text=True)
comment_align = Alignment(vertical="top", wrap_text=True)

rownum = 2
for meta, rws in groups:
    start = rownum
    for (name, field, chae, ban, mi, body) in rws:
        vals = [meta[c] for c in FILECOLS] + [name, field, chae, ban, mi, body]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(rownum, j, clean(v))
            cell.alignment = comment_align if j == len(HEADERS) else file_align
        rownum += 1
    end = rownum - 1
    if end > start:
        for j in range(1, len(FILECOLS) + 1):
            ws.merge_cells(start_row=start, start_column=j, end_row=end, end_column=j)

W = {"source_filename": 34, "year": 6, "category_auto": 16, "발주기관": 16, "안건명": 30,
     "심의위원": 9, "분야": 12, "채택의견": 8, "반영": 7, "미반영": 8, "검토의견": 95}
for j, h in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(j)].width = W.get(h, 12)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{rownum-1}"
wb.save(XL)

n_parsed = sum(1 for _, rws in groups if not (len(rws) == 1 and rws[0][0] == "(미파싱)"))
print(f"sheets: {wb.sheetnames}")
print(f"검토의견 rows: {rownum-2}  | files: {len(groups)}  parsed: {n_parsed}  미파싱: {len(groups)-n_parsed}")
