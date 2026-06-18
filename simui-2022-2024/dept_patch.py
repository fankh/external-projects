# -*- coding: utf-8 -*-
"""
dept_patch.py — extract the ordering agency (발주기관) from each document header
(with body/whole-head fallback) into the 발주기관부서 column, then (re)build a
'dept_summary' sheet: agency counts + agency × category crosstab.
Reads text from the existing workbook (no PDF/OCR). category_final untouched.
"""
import os, re
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

BASE = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(BASE, "simui_2022-2024.xlsx")

BLACK = re.compile(r"(반영|미반영|여부|채택|의견|조치|심의|발주|부서|합계|총괄표|분야|위원|성명|작성|확인|보고서|안건|일시|일자|목\s*차|^계$|^건$)")
# Agency-name suffixes. 실/국 require >=3 chars before them so person names
# (e.g. 최영국) don't match; structural suffixes (본부/공사/공단/구청/...) are exact.
SUFFIX = re.compile(r"(본부|공사|공단|구청|시청|군청|사업소|여가센터|물재생센터|환경센터|센터|사업단|관리단|건설단|추진단|[가-힣]{3,6}실|[가-힣]{3,6}국)")
DISTRICT = re.compile(r"^([가-힣]{1,4}구)\s*([가-힣]{2,10}(?:과|국|부|센터))?$")  # 은평구 공원녹지과
GENERIC = {"건설공사", "정비공사", "개선공사", "보수공사", "신설공사", "증설공사",
           "설치공사", "복구공사", "확장공사", "조성공사", "포장공사", "공사", "회의실", "사무실"}
SPLIT = re.compile(r"^(.*?(?:도시기반시설본부|[가-힣]{0,8}본부|[가-힣]{1,4}구청|시청|군청|[가-힣]{2,8}공사|[가-힣]{2,8}공단|[가-힣]{0,6}사업소|[가-힣]{0,6}여가센터|[가-힣]{0,6}물재생센터|추진단|관리단|사업단|[가-힣]{3,6}실|[가-힣]{3,6}국))(.*)$")

def strip_pre(org):
    # drop the city-government prefix, but keep agency names like 서울시설공단
    return re.sub(r"^서울특별시\s*", "", re.sub(r"^서울시(?=\s|$)\s*", "", org)).strip()

def base_dept(org):
    org = strip_pre(org)
    m = SPLIT.match(org)
    if m:
        return re.sub(r"\s+", "", m.group(1)), m.group(2).strip()
    return re.sub(r"\s+", "", org), ""

def positional(lines):
    start = 0
    for i, l in enumerate(lines):
        if "심의일" in l:
            start = i + 1; break
    if not start:
        for i, l in enumerate(lines):
            if "안건명" in l:
                start = i + 1; break
    scan = lines[start:start + 10] if start else lines[:10]
    dept = ""
    for l in scan:
        m = re.search(r"\(([가-힣]{2,10}(?:과|팀|부|국|실|센터|관|소))\)", l)
        if m and not dept:
            dept = m.group(1)
        core = re.sub(r"\(.*?\)", "", l).strip()
        if not core or BLACK.search(core) or re.search(r"[:：0-9]", core):
            continue
        dm = DISTRICT.match(core)
        if dm:
            if dm.group(2) and not dept:
                dept = dm.group(2)
            return dm.group(1) + "청", dept
        if SUFFIX.search(core) and 3 <= len(core) <= 22 and core not in GENERIC:
            return core, dept
    return "", ""

def fallback(head):
    for m in SUFFIX.finditer(head):
        tok = re.sub(r"\s+", "", head[max(0, m.start() - 12):m.end()])
        tm = re.search(r"([가-힣]{2,14}" + SUFFIX.pattern + r")$", tok)
        if tm and not BLACK.search(tm.group(1)) and tm.group(1) not in GENERIC:
            return tm.group(1), ""
    return "", ""

def extract(text):
    if not text:
        return "", ""
    head = text[:900]
    lines = [re.sub(r"\s+", " ", l).strip() for l in head.splitlines() if l.strip()]
    org, dept = positional(lines)
    if not org:
        org, dept = fallback(head)
    if not org:
        return "", ""
    b, d2 = base_dept(org)
    return b, (dept or d2)

wb = load_workbook(XL)
ws = wb["simui"]
# Migrate the combined 발주기관부서 column into separate 발주기관 + 발주부서 (idempotent).
hdr = [c.value for c in ws[1]]
if "발주기관부서" in hdr and "발주부서" not in hdr:
    c0 = hdr.index("발주기관부서") + 1
    ws.insert_cols(c0 + 1)
    ws.cell(1, c0).value = "발주기관"
    ws.cell(1, c0 + 1).value = "발주부서"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.column_dimensions[get_column_letter(c0)].width = 22
    ws.column_dimensions[get_column_letter(c0 + 1)].width = 16
col = {c.value: i + 1 for i, c in enumerate(ws[1])}

# Curated fixed list of real Seoul agencies — used to (a) validate/clean the
# header extraction and (b) recover 미상 docs by matching known names in the
# document's intro region (precise; avoids deep cross-references).
GOOD_CORP = {"서울교통공사", "서울주택도시공사", "서울에너지공사", "서울농수산식품공사",
             "한국환경공단", "서울시설공단", "서울물재생시설공단"}
GU = [g + "구청" for g in "종로 중 용산 성동 광진 동대문 중랑 성북 강북 도봉 노원 은평 서대문 마포 양천 강서 구로 금천 영등포 동작 관악 서초 강남 송파 강동".split()]
CORE = ["도시기반시설본부", "안전총괄실", "재난안전관리실", "물순환안전국", "푸른도시여가국",
        "정원도시국", "푸른도시국", "도시교통실", "도시계획국", "균형발전본부", "미래한강본부",
        "한강사업본부", "상수도사업본부", "서울아리수본부", "중랑물재생센터", "난지물재생센터",
        "탄천물재생센터", "서남물재생센터", "서부도로사업소", "동부도로사업소"]
CURATED = set(GU) | set(CORE) | GOOD_CORP
AGENCY_TAIL = re.compile(r"(구청|본부|실|국|사업소|센터|추진단|관리단|사업단)$")
RECOVER_WIN = 2500

def valid_agency(a):
    """Accept only real agency forms; reject generic '○○공사' and truncated junk."""
    if not a:
        return False
    if a.endswith(("공사", "공단")):
        return a in GOOD_CORP
    if not (len(a) >= 4 or a.endswith("구청")) or len(a) > 11:
        return False
    return bool(AGENCY_TAIL.search(a))

# pass 1: header extraction + garbage cleanup
recs = []  # [row, text, agency, dept]
for row in range(2, ws.max_row + 1):
    text = ws.cell(row, col["full_text"]).value or ""
    agency, dept = extract(text)
    if not valid_agency(agency):
        agency, dept = "", ""
    recs.append([row, text, agency, dept])

# pass 2: recover 미상 from the intro region against the cleaned + curated whitelist
whitelist = sorted({r[2] for r in recs if r[2]} | CURATED, key=len, reverse=True)
whitelist = [w for w in whitelist if len(w) >= 4 or w.endswith("구청")]
recovered = 0
for r in recs:
    if r[2]:
        continue
    region = r[1][:RECOVER_WIN]
    best = None
    for w in whitelist:
        i = region.find(w)
        if i >= 0 and (best is None or (i, -len(w)) < best[:2]):
            best = (i, -len(w), w)
    if best and valid_agency(best[2]):
        r[2] = best[2]; recovered += 1

agency_count = Counter()
agency_cat = defaultdict(Counter)
agency_year = defaultdict(Counter)
populated = 0
for row, text, agency, dept in recs:
    ws.cell(row, col["발주기관"]).value = agency
    ws.cell(row, col["발주부서"]).value = dept
    key = agency if agency else "(미상 / unidentified)"
    agency_count[key] += 1
    agency_year[key][ws.cell(row, col["year"]).value] += 1
    if agency:
        populated += 1
        agency_cat[key][ws.cell(row, col["category_auto"]).value] += 1

if "dept_summary" in wb.sheetnames:
    del wb["dept_summary"]
ds = wb.create_sheet("dept_summary", 1)
bold = Font(bold=True)
r = 1
ds.cell(r, 1, "발주기관 (ordering agency)").font = bold
ds.cell(r, 2, "count").font = bold
r += 1
for ag, n in sorted(agency_count.items(), key=lambda kv: (kv[0].startswith("(미상"), -kv[1])):
    ds.cell(r, 1, ag); ds.cell(r, 2, n); r += 1
ds.cell(r, 1, "합계 / Total").font = bold
ds.cell(r, 2, sum(agency_count.values())).font = bold
r += 2

cats = ["용역발주심의", "설계심의", "변경설계심의", "건설사업관리계획심의", "정밀안전진단심의",
        "공사기간적정성심의", "입찰안내서심의", "사업수행능력평가기준심의", "기타"]
top = [a for a, n in agency_count.most_common() if not a.startswith("(미상") and n >= 5]
ds.cell(r, 1, "발주기관 × category (count ≥ 5)").font = bold
for j, c in enumerate(cats, 2):
    ds.cell(r, j, c).font = bold
ds.cell(r, len(cats) + 2, "Total").font = bold
r += 1
for ag in top:
    ds.cell(r, 1, ag)
    for j, c in enumerate(cats, 2):
        ds.cell(r, j, agency_cat[ag].get(c, 0))
    ds.cell(r, len(cats) + 2, sum(agency_cat[ag].values()))
    r += 1
r += 1

# 발주기관 × year pivot (top agencies with count >= 5, plus 미상, plus grand total)
years = sorted({y for c in agency_year.values() for y in c})
ds.cell(r, 1, "발주기관 × year (count ≥ 5 + 미상)").font = bold
for j, y in enumerate(years, 2):
    ds.cell(r, j, y).font = bold
ds.cell(r, len(years) + 2, "Total").font = bold
r += 1
for ag in top + ["(미상 / unidentified)"]:
    ds.cell(r, 1, ag)
    for j, y in enumerate(years, 2):
        ds.cell(r, j, agency_year[ag].get(y, 0))
    ds.cell(r, len(years) + 2, sum(agency_year[ag].values()))
    r += 1
ds.cell(r, 1, "합계 / Total").font = bold
for j, y in enumerate(years, 2):
    ds.cell(r, j, sum(agency_year[a].get(y, 0) for a in agency_year)).font = bold
ds.cell(r, len(years) + 2, sum(sum(c.values()) for c in agency_year.values())).font = bold
r += 1

ds.column_dimensions["A"].width = 30
for j in range(2, len(cats) + 3):
    ds.column_dimensions[get_column_letter(j)].width = 13

wb.save(XL)
print(f"발주기관부서 populated: {populated}/{ws.max_row-1} ({100*populated//(ws.max_row-1)}%)  blank: {ws.max_row-1-populated}")
print(f"distinct agencies: {len([a for a in agency_count if not a.startswith('(미상')])}")
print("top 12:", [(a, n) for a, n in agency_count.most_common(12)])
