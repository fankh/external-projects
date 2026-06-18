# -*- coding: utf-8 -*-
"""
simui_extract.py — Extract text from all 심의 조치결과 PDFs into one Excel sheet.

For each PDF: use the embedded text layer where present; OCR (Tesseract kor+eng)
any page that has no text layer (handles fully-scanned and mixed PDFs). Then derive
filename metadata + an auto category, best-effort parse a few header fields, and
write one row per PDF to a single-sheet workbook.

Run:  set PYTHONUTF8=1 && python simui_extract.py
"""
import os, re, sys, tempfile
from collections import Counter

import fitz  # PyMuPDF
import pytesseract
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

BASE      = os.path.dirname(os.path.abspath(__file__))
PDF_DIR   = os.path.join(BASE, "pdfs")
OVERFLOW  = os.path.join(BASE, "text_overflow")
TESSDATA  = os.path.join(BASE, "tessdata")
OUT_XLSX  = os.path.join(BASE, "simui_2022-2024.xlsx")
TESS_EXE  = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

os.makedirs(OVERFLOW, exist_ok=True)
pytesseract.pytesseract.tesseract_cmd = TESS_EXE
os.environ["TESSDATA_PREFIX"] = TESSDATA
fitz.TOOLS.mupdf_display_errors(False)  # silence non-fatal MuPDF warnings

EXCEL_CELL_LIMIT = 32767
PER_PAGE_OCR_THRESHOLD = 15   # a page with fewer real chars is treated as an image -> OCR
OCR_DPI = 300

# Ordered keyword -> canonical category (most specific first).
CATEGORY_RULES = [
    ("변경설계", "변경설계심의"),
    ("용역발주", "용역발주심의"),
    ("건설사업관리계획", "건설사업관리계획심의"),
    ("정밀안전진단", "정밀안전진단심의"),
    ("안전진단", "정밀안전진단심의"),
    ("공사기간", "공사기간적정성심의"),
    ("입찰안내서", "입찰안내서심의"),
    ("설계", "설계심의"),
]

def categorize(name: str) -> str:
    for kw, cat in CATEGORY_RULES:
        if kw in name:
            return cat
    return "기타"

def simui_type_raw(base: str) -> str:
    seg = next((s.strip() for s in re.split(r"[-_]", base) if "심의" in s), None)
    if seg:
        return seg
    m = re.search(r"([가-힣]+\s*심의)", base)
    return m.group(1).strip() if m else ""

def parse_session(base: str):
    m = re.search(r"제\s*(\d+)\s*차", base)
    return (m.group(0).replace(" ", ""), int(m.group(1))) if m else ("", None)

def parse_fields(text: str):
    head = text[:1500]
    def grab(pat):
        m = re.search(pat, head)
        return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""
    agenda = grab(r"안\s*건\s*명\s*[:：]\s*([^\n]+)")
    date   = grab(r"심\s*의\s*일\s*[시자]\s*[:：]?\s*([0-9]{4}[^\n]{0,15})")
    org    = grab(r"\n\s*([가-힣]{2,}(?:공사|공단|본부|센터|사업소|구청|시청|부))\s*[\n(]")
    return agenda, date, org

def ocr_page(page) -> str:
    pix = page.get_pixmap(dpi=OCR_DPI)
    tmp = os.path.join(tempfile.gettempdir(), "simui_ocr_page.png")
    pix.save(tmp)
    try:
        return pytesseract.image_to_string(tmp, lang="kor+eng")
    finally:
        try: os.remove(tmp)
        except OSError: pass

def extract(path: str):
    has_text = ocr_used = False
    parts, err, npages = [], "", 0
    try:
        d = fitz.open(path)
        npages = d.page_count
        for page in d:
            t = page.get_text()
            if len(t.strip()) >= PER_PAGE_OCR_THRESHOLD:
                has_text = True
                parts.append(t)
            else:
                try:
                    parts.append(ocr_page(page)); ocr_used = True
                except Exception as e:
                    err = f"ocr_err:{e}"
        d.close()
    except Exception as e:
        err = f"open_err:{e}"
    return "\n".join(parts).strip(), npages, has_text, ocr_used, err

def safe_name(s: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣._-]", "_", s)[:120]

def clean(v):
    return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

def main():
    pdfs = []
    for year in sorted(os.listdir(PDF_DIR)):
        ydir = os.path.join(PDF_DIR, year)
        if not os.path.isdir(ydir):
            continue
        for f in sorted(os.listdir(ydir)):
            if f.lower().endswith(".pdf"):
                pdfs.append((year, os.path.join(ydir, f), f))
    print(f"total PDFs: {len(pdfs)}", flush=True)

    rows = []
    for idx, (year, path, fname) in enumerate(pdfs, 1):
        base = os.path.splitext(fname)[0]
        text, npages, has_text, ocr_used, err = extract(path)
        sraw, snum = parse_session(base)
        agenda, date, org = parse_fields(text)
        char_count = len(text)
        truncated = char_count > EXCEL_CELL_LIMIT
        cell = text
        if truncated:
            ov = os.path.join(OVERFLOW, f"{year}_{safe_name(base)}.txt")
            with open(ov, "w", encoding="utf-8") as fo:
                fo.write(text)
            cell = text[:EXCEL_CELL_LIMIT - 80] + f"\n…[truncated; full text in text_overflow\\{os.path.basename(ov)}]"
        rows.append(dict(
            year=year, session_no=snum, session_no_raw=sraw,
            source_filename=fname, source_relpath=os.path.relpath(path, BASE),
            file_size_kb=round(os.path.getsize(path) / 1024, 1), page_count=npages,
            simui_type_raw=simui_type_raw(base), category_auto=categorize(base), category_final="",
            안건명=agenda, 심의일시=date, 발주기관부서=org,
            has_text_layer=has_text, ocr_used=ocr_used, char_count=char_count,
            text_truncated=truncated, needs_review=(char_count < 30), extract_error=err,
            full_text=cell,
        ))
        if idx % 20 == 0 or idx == len(pdfs):
            print(f"  {idx}/{len(pdfs)}  {'OCR' if ocr_used else 'txt'} {char_count:>6}c  {fname[:42]}", flush=True)

    cols = ["year", "session_no", "session_no_raw", "source_filename", "source_relpath",
            "file_size_kb", "page_count", "simui_type_raw", "category_auto", "category_final",
            "안건명", "심의일시", "발주기관부서", "has_text_layer", "ocr_used", "char_count",
            "text_truncated", "needs_review", "extract_error", "full_text"]
    wb = Workbook(); ws = wb.active; ws.title = "simui"
    ws.append(cols)
    for r in rows:
        ws.append([clean(r.get(c, "")) for c in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    widths = {"year": 6, "session_no": 9, "session_no_raw": 11, "source_filename": 42,
              "source_relpath": 26, "category_auto": 18, "category_final": 14, "안건명": 36,
              "심의일시": 16, "발주기관부서": 18, "char_count": 9, "full_text": 90}
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 12)
    wb.save(OUT_XLSX)

    n = len(rows)
    print(f"\nSAVED {OUT_XLSX}", flush=True)
    print(f"rows={n}  ocr_used={sum(r['ocr_used'] for r in rows)}  "
          f"text_layer={sum(r['has_text_layer'] for r in rows)}  "
          f"needs_review={sum(r['needs_review'] for r in rows)}  "
          f"truncated={sum(r['text_truncated'] for r in rows)}", flush=True)
    print("category_auto:", dict(Counter(r["category_auto"] for r in rows).most_common()), flush=True)
    print("by year:", dict(Counter(r["year"] for r in rows)), flush=True)

if __name__ == "__main__":
    main()
