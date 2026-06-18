# -*- coding: utf-8 -*-
"""
category_patch.py — refine category_auto (shrink the 기타 bucket) and add a
'summary' sheet to simui_2022-2024.xlsx. Re-derives category_auto from the
filename only (no PDF/OCR), so it is fast and safe; category_final is untouched.
"""
import os
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

BASE = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(BASE, "simui_2022-2024.xlsx")

# Ordered keyword -> canonical category (first match wins; most specific first).
# Additions vs v1: 용역별주(typo), 사업수행능력평가기준(new), 건설사업관리(broad),
# 정밀안전진(catches 정밀안전진단 + typo), 공기적정성(= 공사기간 적정성).
CATEGORY_RULES = [
    ("변경설계", "변경설계심의"),
    ("용역발주", "용역발주심의"),
    ("용역별주", "용역발주심의"),
    ("사업수행능력평가기준", "사업수행능력평가기준심의"),
    ("건설사업관리계획", "건설사업관리계획심의"),
    ("건설사업관리", "건설사업관리계획심의"),
    ("정밀안전진", "정밀안전진단심의"),
    ("안전진단", "정밀안전진단심의"),
    ("공사기간", "공사기간적정성심의"),
    ("공기적정성", "공사기간적정성심의"),
    ("입찰안내서", "입찰안내서심의"),
    ("설계", "설계심의"),
]

def categorize(name: str) -> str:
    for kw, cat in CATEGORY_RULES:
        if kw in name:
            return cat
    return "기타"

wb = load_workbook(XL)
ws = wb["simui"]
hdr = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(hdr)}

old = Counter(); new = Counter()
by_year = defaultdict(Counter)         # year -> category counts
flags = Counter()
for row in range(2, ws.max_row + 1):
    fname = ws.cell(row, col["source_filename"]).value or ""
    year = ws.cell(row, col["year"]).value
    old[ws.cell(row, col["category_auto"]).value] += 1
    cat = categorize(fname)
    ws.cell(row, col["category_auto"]).value = cat
    new[cat] += 1
    by_year[year][cat] += 1
    for f in ("has_text_layer", "ocr_used", "needs_review", "text_truncated"):
        if str(ws.cell(row, col[f]).value).upper() == "TRUE":
            flags[f] += 1

# ---- build summary sheet ----
if "summary" in wb.sheetnames:
    del wb["summary"]
sm = wb.create_sheet("summary", 0)  # put first
bold = Font(bold=True)
r = 1
def put(row, c1, c2=None, header=False):
    sm.cell(row, 1, c1).font = bold if header else Font()
    if c2 is not None:
        sm.cell(row, 2, c2)

cats = [c for c, _ in sorted(new.items(), key=lambda kv: -kv[1])]
years = sorted(by_year)

put(r, "category_auto", "count", header=True); r += 1
for c in cats:
    put(r, c, new[c]); r += 1
put(r, "합계 / Total", sum(new.values())); r += 2

put(r, "by year", "count", header=True); r += 1
for y in years:
    put(r, y, sum(by_year[y].values())); r += 1
r += 1

put(r, "quality flags", "count", header=True); r += 1
for f, label in [("has_text_layer", "has_text_layer"), ("ocr_used", "ocr_used"),
                 ("needs_review", "needs_review"), ("text_truncated", "text_truncated")]:
    put(r, label, flags[f]); r += 1
r += 1

# category_auto x year crosstab
sm.cell(r, 1, "category_auto × year").font = bold
for j, y in enumerate(years, 2):
    sm.cell(r, j, y).font = bold
sm.cell(r, len(years) + 2, "Total").font = bold
r += 1
for c in cats:
    sm.cell(r, 1, c)
    for j, y in enumerate(years, 2):
        sm.cell(r, j, by_year[y].get(c, 0))
    sm.cell(r, len(years) + 2, new[c])
    r += 1
sm.cell(r, 1, "합계 / Total").font = bold
for j, y in enumerate(years, 2):
    sm.cell(r, j, sum(by_year[y].values())).font = bold
sm.cell(r, len(years) + 2, sum(new.values())).font = bold
r += 1

sm.column_dimensions["A"].width = 26
for j in range(2, len(years) + 3):
    sm.column_dimensions[get_column_letter(j)].width = 10

wb.save(XL)
print("OLD:", dict(old.most_common()))
print("NEW:", dict(new.most_common()))
print("기타:", old.get("기타", 0), "->", new.get("기타", 0))
print("summary sheet added; flags:", dict(flags))
