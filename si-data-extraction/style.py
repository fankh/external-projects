# -*- coding: utf-8 -*-
"""
style.py — apply production styling to the workbook: header fills, thin borders,
category color-coding, warning highlights, section/total styling, banding.
Run LAST (after localize.py). Safe to re-run.
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(BASE, "simui_2022-2024.xlsx")

def fill(hex_): return PatternFill("solid", fgColor=hex_)
HEADER_FILL = fill("1F4E78")      # dark navy
SECTION_FILL = fill("4472C4")     # medium blue
TOTAL_FILL = fill("D9E1F2")       # light blue
WARN_FILL = fill("FFC7CE")        # light red
FILEGRP_FILL = fill("F2F2F2")     # light gray
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
_thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

CAT_FILL = {
    "용역발주심의": "DDEBF7", "설계심의": "E2EFDA", "정밀안전진단심의": "FCE4EC",
    "건설사업관리계획심의": "EDE7F6", "공사기간적정성심의": "FFF2CC", "입찰안내서심의": "D5F0EF",
    "변경설계심의": "FFF8E1", "사업수행능력평가기준심의": "E8EAF6", "기타": "F2F2F2",
}
SECTION_TITLES = {"분류(자동)", "발주기관", "연도별", "추출품질", "분류(자동) × 연도",
                  "발주기관 × 분류 (5건 이상)", "발주기관 × 연도 (5건 이상 + 미상)"}
CENTER_COLS = {"연도", "회차", "회차표기", "파일크기(KB)", "페이지수", "글자수", "본문텍스트유무",
               "OCR사용", "본문잘림", "검토필요", "분류(자동)", "분류(확정)", "심의일시",
               "채택의견", "반영", "미반영", "심의위원", "분야"}

def style_header(ws):
    for c in ws[1]:
        c.fill = HEADER_FILL; c.font = WHITE_BOLD; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[1].height = 24

def borders(ws):
    for row in ws.iter_rows():
        for c in row:
            c.border = BORDER

wb = load_workbook(XL)

# ---- 심의데이터 ----
ws = wb["심의데이터"]
hdr = [c.value for c in ws[1]]
idx = {v: i + 1 for i, v in enumerate(hdr)}
style_header(ws); borders(ws)
ci, ri, ei = idx.get("분류(자동)"), idx.get("검토필요"), idx.get("추출오류")
ctr_idx = [idx[h] for h in CENTER_COLS if h in idx]
for r in range(2, ws.max_row + 1):
    for j in ctr_idx:
        ws.cell(r, j).alignment = CTR
    if ci:
        v = ws.cell(r, ci).value
        if v in CAT_FILL:
            ws.cell(r, ci).fill = fill(CAT_FILL[v])
    if ri and str(ws.cell(r, ri).value).strip() == "예":
        ws.cell(r, ri).fill = WARN_FILL
    if ei and str(ws.cell(r, ei).value or "").strip():
        ws.cell(r, ei).fill = WARN_FILL

# ---- 요약 & 발주기관요약 ----
for sn in ("요약", "발주기관요약"):
    ws = wb[sn]; borders(ws)
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a in SECTION_TITLES:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value is not None or c <= 2:
                    cell.fill = SECTION_FILL; cell.font = WHITE_BOLD; cell.alignment = CTR
        elif a == "합계":
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).font = BOLD
        elif a in CAT_FILL:
            ws.cell(r, 1).fill = fill(CAT_FILL[a])

# ---- 검토의견 ----
ws = wb["검토의견"]
hdr = [c.value for c in ws[1]]
cidx = {v: i + 1 for i, v in enumerate(hdr)}
style_header(ws); borders(ws)
ctr_idx = [cidx[h] for h in CENTER_COLS if h in cidx]
for r in range(2, ws.max_row + 1):
    for c in range(1, 6):                       # file-metadata columns
        ws.cell(r, c).fill = FILEGRP_FILL
    for j in ctr_idx:
        a = ws.cell(r, j).alignment
        ws.cell(r, j).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb.save(XL)
print("styled sheets:", wb.sheetnames)
