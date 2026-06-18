# -*- coding: utf-8 -*-
"""
localize.py — final localization pass: turn the workbook's machine-style English
(column headers, TRUE/FALSE flags, sheet names, section labels) into production
Korean. Idempotent; run LAST (after simui_extract / category_patch / dept_patch /
auditor_comments, which use English column names internally).
"""
import os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(BASE, "simui_2022-2024.xlsx")

# column headers (simui + 검토의견 row 1) -> Korean
HEADER_MAP = {
    "year": "연도", "session_no": "회차", "session_no_raw": "회차표기",
    "source_filename": "파일명", "source_relpath": "파일경로", "file_size_kb": "파일크기(KB)",
    "page_count": "페이지수", "simui_type_raw": "심의유형(원문)",
    "category_auto": "분류(자동)", "category_final": "분류(확정)",
    "has_text_layer": "본문텍스트유무", "ocr_used": "OCR사용", "char_count": "글자수",
    "text_truncated": "본문잘림", "needs_review": "검토필요", "extract_error": "추출오류",
    "full_text": "본문",
}
# section labels / values in 요약 & 발주기관요약 sheets -> Korean
LABEL_MAP = {
    "category_auto": "분류(자동)", "count": "건수", "합계 / Total": "합계", "by year": "연도별",
    "quality flags": "추출품질", "category_auto × year": "분류(자동) × 연도", "Total": "합계",
    "has_text_layer": "본문텍스트유무", "ocr_used": "OCR사용", "needs_review": "검토필요",
    "text_truncated": "본문잘림",
    "발주기관 (ordering agency)": "발주기관",
    "발주기관 × category (count ≥ 5)": "발주기관 × 분류 (5건 이상)",
    "발주기관 × year (count ≥ 5 + 미상)": "발주기관 × 연도 (5건 이상 + 미상)",
    "(미상 / unidentified)": "(미상)",
}
BOOL_COLS = {"본문텍스트유무", "OCR사용", "본문잘림", "검토필요"}   # after header rename
SHEET_MAP = {"summary": "요약", "dept_summary": "발주기관요약", "simui": "심의데이터"}

wb = load_workbook(XL)

# 1) headers on the data sheets
for sn in ("simui", "검토의견"):
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    for c in ws[1]:
        if c.value in HEADER_MAP:
            c.value = HEADER_MAP[c.value]
    # TRUE/FALSE -> 예/아니오 on the flag columns (data sheet only)
    if sn == "simui":
        idx = {c.value: c.column for c in ws[1]}
        for kr in BOOL_COLS:
            j = idx.get(kr)
            if not j:
                continue
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, j).value
                if v is True or (isinstance(v, str) and v.strip().upper() == "TRUE"):
                    ws.cell(r, j).value = "예"
                elif v is False or (isinstance(v, str) and v.strip().upper() == "FALSE"):
                    ws.cell(r, j).value = "아니오"

# 2) labels/values on the summary sheets
for sn in ("summary", "dept_summary"):
    if sn not in wb.sheetnames:
        continue
    for row in wb[sn].iter_rows():
        for c in row:
            if isinstance(c.value, str):
                if c.value in LABEL_MAP:
                    c.value = LABEL_MAP[c.value]
                elif "(미상 / unidentified)" in c.value:
                    c.value = c.value.replace("(미상 / unidentified)", "(미상)")

# 3) sheet tab names (last, so earlier lookups by English name still work)
for old, new in SHEET_MAP.items():
    if old in wb.sheetnames and new not in wb.sheetnames:
        wb[old].title = new

wb.save(XL)
print("sheets:", wb.sheetnames)
