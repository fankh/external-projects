# -*- coding: utf-8 -*-
"""
patch_ocr.py — quality pass over simui_2022-2024.xlsx:
  (1) re-OCR the worst scanned docs (Hangul ratio < 0.25, i.e. rotated/garbled)
      with Tesseract --psm 1 (auto orientation) and keep the result if it improves;
  (2) annotate the 2 DRM-protected files that no parser can open;
  (3) flag any still-bad OCR rows as needs_review.
Edits the workbook in place (preserves formatting).
"""
import os, re, tempfile
import fitz, pytesseract
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

BASE = os.path.dirname(os.path.abspath(__file__))
XL   = os.path.join(BASE, "simui_2022-2024.xlsx")
OVER = os.path.join(BASE, "text_overflow")
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
os.environ["TESSDATA_PREFIX"] = os.path.join(BASE, "tessdata")
fitz.TOOLS.mupdf_display_errors(False)
LIMIT = 32767

def hr(t):
    t = t or ""
    ns = len(re.findall(r"\S", t))
    return (len(re.findall(r"[가-힣]", t)) / ns) if ns else 0.0

def clean(v):
    return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

def safe_name(s):
    return re.sub(r"[^0-9A-Za-z가-힣._-]", "_", s)[:120]

def ocr_doc_psm1(path):
    d = fitz.open(path); parts = []
    for page in d:
        pix = page.get_pixmap(dpi=300)
        tmp = os.path.join(tempfile.gettempdir(), "patch_ocr.png"); pix.save(tmp)
        try:
            parts.append(pytesseract.image_to_string(tmp, lang="kor+eng", config="--psm 1"))
        finally:
            try: os.remove(tmp)
            except OSError: pass
    d.close()
    return "\n".join(parts).strip()

wb = load_workbook(XL)
ws = wb.active
hdr = [c.value for c in ws[1]]
col = {name: i + 1 for i, name in enumerate(hdr)}   # 1-based

DRM = ("제190차-신이촌", "제245-1-1차-증산3교")
reocr = improved = drm = flagged = 0

for row in range(2, ws.max_row + 1):
    fname = ws.cell(row, col["source_filename"]).value or ""
    # (2) DRM annotation
    if any(k in fname for k in DRM):
        ws.cell(row, col["extract_error"]).value = "DRM-protected (\\x9b DRMONE header) — unreadable by any PDF parser; re-export/decrypt at source"
        ws.cell(row, col["needs_review"]).value = True
        drm += 1
        continue
    if str(ws.cell(row, col["ocr_used"]).value).upper() != "TRUE":
        continue
    cur = ws.cell(row, col["full_text"]).value or ""
    old = hr(cur)
    if old >= 0.25:
        continue
    # (1) re-OCR with orientation auto-detect
    relpath = ws.cell(row, col["source_relpath"]).value
    path = os.path.join(BASE, relpath)
    reocr += 1
    try:
        new = ocr_doc_psm1(path)
    except Exception as e:
        ws.cell(row, col["extract_error"]).value = f"reocr_err:{e}"
        new = ""
    if hr(new) > old + 0.05 and len(new) > 30:
        improved += 1
        cc = len(new)
        cell = new
        if cc > LIMIT:
            year = ws.cell(row, col["year"]).value
            ov = os.path.join(OVER, f"{year}_{safe_name(os.path.splitext(fname)[0])}.txt")
            open(ov, "w", encoding="utf-8").write(new)
            cell = new[:LIMIT - 80] + f"\n…[truncated; full text in text_overflow\\{os.path.basename(ov)}]"
            ws.cell(row, col["text_truncated"]).value = True
        ws.cell(row, col["full_text"]).value = clean(cell)
        ws.cell(row, col["char_count"]).value = cc
        ws.cell(row, col["needs_review"]).value = (hr(new) < 0.15)
    else:
        # no improvement -> flag the unreliable OCR
        ws.cell(row, col["needs_review"]).value = True
        flagged += 1

wb.save(XL)
print(f"re-OCR attempted={reocr} improved={improved} flagged_low_quality={flagged} drm_annotated={drm}", flush=True)
