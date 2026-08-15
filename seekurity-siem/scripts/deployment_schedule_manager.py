"""
Seekurity SIEM Deployment Schedule Manager

SIEM 솔루션 구축 일정 관리를 위한 Excel 파일 생성

Usage:
    python deployment_schedule_manager.py

Output:
    SeekuritySIEM_Deployment_Schedule.xlsx
"""

from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION
# ============================================================


class Config:
    """Configuration for deployment schedule"""

    OUTPUT_FILE = "SeekuritySIEM_Deployment_Schedule.xlsx"

    # Colors
    COLORS = {
        "title": "1B2838",
        "phase1": "2E5090",  # 준비 단계
        "phase2": "3D7EAA",  # 설계 단계
        "phase3": "5B9BD5",  # 구축 단계
        "phase4": "70AD47",  # 테스트 단계
        "phase5": "FFC000",  # 안정화 단계
        "header": "B4C6E7",
        "row_even": "FFFFFF",
        "row_odd": "F8F9FA",
        "border": "D0D0D0",
        "complete": "C6EFCE",  # 완료
        "in_progress": "FFEB9C",  # 진행중
        "delayed": "FFC7CE",  # 지연
        "pending": "F2F2F2",  # 대기
    }

    # Status options
    STATUS_OPTIONS = ["대기", "진행중", "완료", "지연", "보류"]

    # Priority options
    PRIORITY_OPTIONS = ["높음", "중간", "낮음"]

    # Deployment phases and tasks
    PHASES = [
        {
            "name": "1. 준비 단계",
            "color": "phase1",
            "tasks": [
                {"task": "프로젝트 킥오프 미팅", "duration": 1, "owner": "PM"},
                {"task": "프로젝트 범위 확정", "duration": 2, "owner": "PM"},
                {"task": "현행 시스템 분석", "duration": 5, "owner": "Engineer"},
                {"task": "Log Source 목록 확정", "duration": 3, "owner": "Engineer"},
                {"task": "네트워크 구성도 확보", "duration": 2, "owner": "고객사"},
                {"task": "방화벽 정책 요청", "duration": 3, "owner": "고객사"},
                {"task": "서버 자원 할당", "duration": 3, "owner": "고객사"},
            ],
        },
        {
            "name": "2. 설계 단계",
            "color": "phase2",
            "tasks": [
                {"task": "SIEM Architecture 설계", "duration": 5, "owner": "Engineer"},
                {"task": "Log 수집 설계", "duration": 3, "owner": "Engineer"},
                {"task": "Parser 설계", "duration": 5, "owner": "Engineer"},
                {"task": "Use Case 설계", "duration": 5, "owner": "Engineer"},
                {"task": "Dashboard 설계", "duration": 3, "owner": "Engineer"},
                {"task": "Alert Rule 설계", "duration": 3, "owner": "Engineer"},
                {"task": "설계 검토 및 승인", "duration": 2, "owner": "PM"},
            ],
        },
        {
            "name": "3. 구축 단계",
            "color": "phase3",
            "tasks": [
                {"task": "Seekurity SIEM 설치", "duration": 2, "owner": "Engineer"},
                {"task": "Collector 구성", "duration": 2, "owner": "Engineer"},
                {"task": "Network Security 연동", "duration": 5, "owner": "Engineer"},
                {"task": "Endpoint Security 연동", "duration": 3, "owner": "Engineer"},
                {"task": "Data & Application 연동", "duration": 5, "owner": "Engineer"},
                {"task": "Parser 개발 및 적용", "duration": 7, "owner": "Engineer"},
                {"task": "Use Case 구현", "duration": 5, "owner": "Engineer"},
                {"task": "Dashboard 구성", "duration": 3, "owner": "Engineer"},
                {"task": "Alert Rule 설정", "duration": 3, "owner": "Engineer"},
            ],
        },
        {
            "name": "4. 테스트 단계",
            "color": "phase4",
            "tasks": [
                {"task": "Log 수집 검증", "duration": 3, "owner": "Engineer"},
                {"task": "Parser 검증", "duration": 3, "owner": "Engineer"},
                {"task": "Use Case 테스트", "duration": 3, "owner": "Engineer"},
                {"task": "Alert 발생 테스트", "duration": 2, "owner": "Engineer"},
                {"task": "성능 테스트", "duration": 3, "owner": "Engineer"},
                {"task": "사용자 교육", "duration": 2, "owner": "Engineer"},
                {"task": "기능 확인 (FVT)", "duration": 3, "owner": "고객사"},
            ],
        },
        {
            "name": "5. 안정화 단계",
            "color": "phase5",
            "tasks": [
                {"task": "운영 모니터링", "duration": 10, "owner": "Engineer"},
                {"task": "이슈 대응 및 튜닝", "duration": 10, "owner": "Engineer"},
                {"task": "운영 문서 작성", "duration": 5, "owner": "Engineer"},
                {"task": "운영 인계", "duration": 2, "owner": "Engineer"},
                {"task": "프로젝트 종료 보고", "duration": 1, "owner": "PM"},
            ],
        },
    ]


# ============================================================
# STYLES
# ============================================================


class Styles:
    """Pre-defined styles"""

    def __init__(self):
        c = Config.COLORS

        # Fills
        self.title_fill = PatternFill(
            start_color=c["title"], end_color=c["title"], fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color=c["header"], end_color=c["header"], fill_type="solid"
        )
        self.row_even = PatternFill(
            start_color=c["row_even"], end_color=c["row_even"], fill_type="solid"
        )
        self.row_odd = PatternFill(
            start_color=c["row_odd"], end_color=c["row_odd"], fill_type="solid"
        )

        # Phase fills
        self.phase_fills = {
            "phase1": PatternFill(
                start_color=c["phase1"], end_color=c["phase1"], fill_type="solid"
            ),
            "phase2": PatternFill(
                start_color=c["phase2"], end_color=c["phase2"], fill_type="solid"
            ),
            "phase3": PatternFill(
                start_color=c["phase3"], end_color=c["phase3"], fill_type="solid"
            ),
            "phase4": PatternFill(
                start_color=c["phase4"], end_color=c["phase4"], fill_type="solid"
            ),
            "phase5": PatternFill(
                start_color=c["phase5"], end_color=c["phase5"], fill_type="solid"
            ),
        }

        # Status fills
        self.status_fills = {
            "완료": PatternFill(
                start_color=c["complete"], end_color=c["complete"], fill_type="solid"
            ),
            "진행중": PatternFill(
                start_color=c["in_progress"],
                end_color=c["in_progress"],
                fill_type="solid",
            ),
            "지연": PatternFill(
                start_color=c["delayed"], end_color=c["delayed"], fill_type="solid"
            ),
            "대기": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
            "보류": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
        }

        # Fonts
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.phase_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_font = Font(bold=True, size=10, color="1B2838")
        self.data_font = Font(size=9)

        # Borders
        self.thin_border = Border(
            left=Side(style="thin", color=c["border"]),
            right=Side(style="thin", color=c["border"]),
            top=Side(style="thin", color=c["border"]),
            bottom=Side(style="thin", color=c["border"]),
        )
        self.header_border = Border(
            left=Side(style="thin", color="8EA9DB"),
            right=Side(style="thin", color="8EA9DB"),
            top=Side(style="medium", color="2E5090"),
            bottom=Side(style="medium", color="2E5090"),
        )

        # Alignments
        self.center = Alignment(horizontal="center", vertical="center")
        self.center_wrap = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.left = Alignment(vertical="center")
        self.left_wrap = Alignment(vertical="center", wrap_text=True)


# ============================================================
# SCHEDULE BUILDER
# ============================================================


class ScheduleBuilder:
    """Build deployment schedule Excel"""

    def __init__(self):
        self.styles = Styles()
        self.wb = Workbook()
        self.start_date = datetime.now().replace(
            hour=0, minute=0, second=0, microsecond=0
        )

    def create_schedule_sheet(self):
        """Create main schedule sheet"""
        ws = self.wb.active
        ws.title = "Deployment Schedule"

        # Row 1: Title
        ws.merge_cells("A1:L1")
        cell = ws["A1"]
        cell.value = "Seekurity SIEM 구축 일정표"
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

        # Row 2: Project info
        ws.merge_cells("A2:L2")
        cell = ws["A2"]
        cell.value = f"프로젝트: KOVAN SIEM 구축 | 시작일: {self.start_date.strftime('%Y-%m-%d')} | 작성일: {datetime.now().strftime('%Y-%m-%d')}"
        cell.fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )
        cell.font = Font(size=10)
        cell.alignment = self.styles.center

        # Row 3: Headers
        headers = [
            ("No.", 5),
            ("Phase", 12),
            ("Task", 40),
            ("담당자", 10),
            ("우선순위", 10),
            ("예상 기간", 10),
            ("시작일", 12),
            ("종료일", 12),
            ("진행률", 10),
            ("Status", 10),
            ("이슈/비고", 30),
            ("산출물", 20),
        ]

        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center_wrap
            cell.border = self.styles.header_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        row_idx = 4
        task_no = 1
        current_date = self.start_date

        for phase in Config.PHASES:
            phase_start_row = row_idx
            phase_color = phase["color"]

            for task_idx, task in enumerate(phase["tasks"]):
                is_odd = (row_idx - 4) % 2 == 1
                base_fill = self.styles.row_odd if is_odd else self.styles.row_even

                # Calculate dates
                end_date = current_date + timedelta(days=task["duration"] - 1)

                # No.
                cell = ws.cell(row=row_idx, column=1, value=task_no)
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # Phase (only first task shows phase name)
                cell = ws.cell(row=row_idx, column=2)
                if task_idx == 0:
                    cell.value = phase["name"]
                cell.fill = self.styles.phase_fills[phase_color]
                cell.font = (
                    self.styles.phase_font if task_idx == 0 else self.styles.data_font
                )
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center_wrap

                # Task
                cell = ws.cell(row=row_idx, column=3, value=task["task"])
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.left

                # 담당자
                cell = ws.cell(row=row_idx, column=4, value=task["owner"])
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 우선순위
                cell = ws.cell(row=row_idx, column=5, value="중간")
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 예상 기간
                cell = ws.cell(row=row_idx, column=6, value=f"{task['duration']}일")
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 시작일
                cell = ws.cell(
                    row=row_idx, column=7, value=current_date.strftime("%Y-%m-%d")
                )
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 종료일
                cell = ws.cell(
                    row=row_idx, column=8, value=end_date.strftime("%Y-%m-%d")
                )
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 진행률
                cell = ws.cell(row=row_idx, column=9, value="0%")
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # Status
                cell = ws.cell(row=row_idx, column=10, value="대기")
                cell.fill = self.styles.status_fills["대기"]
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.center

                # 이슈/비고
                cell = ws.cell(row=row_idx, column=11, value="")
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.left_wrap

                # 산출물
                cell = ws.cell(row=row_idx, column=12, value="")
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.left

                # Move to next date (tasks are sequential)
                current_date = end_date + timedelta(days=1)
                row_idx += 1
                task_no += 1

            # Merge phase cells
            if len(phase["tasks"]) > 1:
                ws.merge_cells(f"B{phase_start_row}:B{row_idx - 1}")

        # Add data validations
        last_row = row_idx - 1

        # Status dropdown
        status_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.STATUS_OPTIONS)}"',
            allow_blank=False,
        )
        status_dv.prompt = "Status 선택"
        ws.add_data_validation(status_dv)
        status_dv.add(f"J4:J{last_row}")

        # Priority dropdown
        priority_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.PRIORITY_OPTIONS)}"',
            allow_blank=False,
        )
        priority_dv.prompt = "우선순위 선택"
        ws.add_data_validation(priority_dv)
        priority_dv.add(f"E4:E{last_row}")

        # Progress dropdown (0% to 100%)
        progress_options = ",".join([f"{i}%" for i in range(0, 101, 10)])
        progress_dv = DataValidation(
            type="list",
            formula1=f'"{progress_options}"',
            allow_blank=False,
        )
        ws.add_data_validation(progress_dv)
        progress_dv.add(f"I4:I{last_row}")

        # Row heights
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 30

        # Freeze panes
        ws.freeze_panes = "C4"

        return self

    def create_summary_sheet(self):
        """Create summary dashboard sheet"""
        ws = self.wb.create_sheet(title="Summary")

        # Title
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value = "프로젝트 현황 요약"
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

        # Phase summary table
        ws["A3"] = "Phase"
        ws["B3"] = "Task 수"
        ws["C3"] = "예상 기간"
        ws["D3"] = "Status"

        for col in range(1, 5):
            cell = ws.cell(row=3, column=col)
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center
            cell.border = self.styles.header_border

        row = 4
        total_tasks = 0
        total_days = 0

        for phase in Config.PHASES:
            task_count = len(phase["tasks"])
            phase_days = sum(t["duration"] for t in phase["tasks"])
            total_tasks += task_count
            total_days += phase_days

            ws.cell(
                row=row, column=1, value=phase["name"]
            ).fill = self.styles.phase_fills[phase["color"]]
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=10)
            ws.cell(row=row, column=1).border = self.styles.thin_border
            ws.cell(row=row, column=1).alignment = self.styles.left

            ws.cell(
                row=row, column=2, value=task_count
            ).border = self.styles.thin_border
            ws.cell(row=row, column=2).alignment = self.styles.center

            ws.cell(
                row=row, column=3, value=f"{phase_days}일"
            ).border = self.styles.thin_border
            ws.cell(row=row, column=3).alignment = self.styles.center

            ws.cell(row=row, column=4, value="대기").border = self.styles.thin_border
            ws.cell(row=row, column=4).alignment = self.styles.center
            ws.cell(row=row, column=4).fill = self.styles.status_fills["대기"]

            row += 1

        # Total row
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=1).border = self.styles.thin_border
        ws.cell(row=row, column=2, value=total_tasks).font = Font(bold=True)
        ws.cell(row=row, column=2).border = self.styles.thin_border
        ws.cell(row=row, column=2).alignment = self.styles.center
        ws.cell(row=row, column=3, value=f"{total_days}일").font = Font(bold=True)
        ws.cell(row=row, column=3).border = self.styles.thin_border
        ws.cell(row=row, column=3).alignment = self.styles.center
        ws.cell(row=row, column=4).border = self.styles.thin_border

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

        # Status legend
        ws["A" + str(row + 3)] = "Status Legend"
        ws["A" + str(row + 3)].font = Font(bold=True)

        legends = [
            ("대기", "pending"),
            ("진행중", "in_progress"),
            ("완료", "complete"),
            ("지연", "delayed"),
        ]

        for idx, (status, color_key) in enumerate(legends):
            r = row + 4 + idx
            ws.cell(row=r, column=1, value=status).border = self.styles.thin_border
            ws.cell(row=r, column=2).fill = self.styles.status_fills[status]
            ws.cell(row=r, column=2).border = self.styles.thin_border

        ws.row_dimensions[1].height = 30

        return self

    def create_milestone_sheet(self):
        """Create milestone sheet"""
        ws = self.wb.create_sheet(title="Milestones")

        # Title
        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value = "주요 Milestone"
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

        # Headers
        headers = ["No.", "Milestone", "예정일", "완료일", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center
            cell.border = self.styles.header_border

        # Milestones
        milestones = [
            "프로젝트 킥오프",
            "설계 완료 및 승인",
            "SIEM 설치 완료",
            "Log Source 연동 완료",
            "테스트 완료 (FVT)",
            "안정화 완료",
            "프로젝트 종료",
        ]

        current_date = self.start_date
        phase_durations = []
        for phase in Config.PHASES:
            phase_durations.append(sum(t["duration"] for t in phase["tasks"]))

        milestone_dates = [
            self.start_date,  # 킥오프
            self.start_date + timedelta(days=sum(phase_durations[:2])),  # 설계 완료
            self.start_date + timedelta(days=sum(phase_durations[:2]) + 4),  # SIEM 설치
            self.start_date + timedelta(days=sum(phase_durations[:3])),  # 연동 완료
            self.start_date + timedelta(days=sum(phase_durations[:4])),  # 테스트 완료
            self.start_date
            + timedelta(days=sum(phase_durations[:5]) - 3),  # 안정화 완료
            self.start_date + timedelta(days=sum(phase_durations)),  # 종료
        ]

        for row, (milestone, date) in enumerate(
            zip(milestones, milestone_dates), start=3
        ):
            is_odd = (row - 3) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row, column=1, value=row - 2).border = self.styles.thin_border
            ws.cell(row=row, column=1).alignment = self.styles.center
            ws.cell(row=row, column=1).fill = base_fill

            ws.cell(row=row, column=2, value=milestone).border = self.styles.thin_border
            ws.cell(row=row, column=2).fill = base_fill

            ws.cell(
                row=row, column=3, value=date.strftime("%Y-%m-%d")
            ).border = self.styles.thin_border
            ws.cell(row=row, column=3).alignment = self.styles.center
            ws.cell(row=row, column=3).fill = base_fill

            ws.cell(row=row, column=4, value="").border = self.styles.thin_border
            ws.cell(row=row, column=4).alignment = self.styles.center
            ws.cell(row=row, column=4).fill = base_fill

            ws.cell(row=row, column=5, value="대기").border = self.styles.thin_border
            ws.cell(row=row, column=5).alignment = self.styles.center
            ws.cell(row=row, column=5).fill = self.styles.status_fills["대기"]

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12

        ws.row_dimensions[1].height = 30

        # Status dropdown
        status_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.STATUS_OPTIONS)}"',
            allow_blank=False,
        )
        ws.add_data_validation(status_dv)
        status_dv.add(f"E3:E{2 + len(milestones)}")

        return self

    def create_resource_sheet(self):
        """Create resource allocation sheet"""
        ws = self.wb.create_sheet(title="Resources")

        # Title
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = "투입 인력 현황"
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

        # Headers
        headers = ["No.", "Role", "담당자명", "소속", "연락처", "비고"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center
            cell.border = self.styles.header_border

        # Default roles
        roles = [
            ("PM", "프로젝트 관리"),
            ("SIEM Engineer", "SIEM 구축"),
            ("SIEM Engineer", "Log 연동"),
            ("네트워크 담당자", "방화벽/VPN 설정"),
            ("보안 담당자", "정책 검토"),
            ("시스템 담당자", "서버 관리"),
        ]

        for row, (role, note) in enumerate(roles, start=3):
            is_odd = (row - 3) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row, column=1, value=row - 2).border = self.styles.thin_border
            ws.cell(row=row, column=1).alignment = self.styles.center
            ws.cell(row=row, column=1).fill = base_fill

            ws.cell(row=row, column=2, value=role).border = self.styles.thin_border
            ws.cell(row=row, column=2).fill = base_fill

            ws.cell(row=row, column=3, value="").border = self.styles.thin_border
            ws.cell(row=row, column=3).fill = base_fill

            ws.cell(row=row, column=4, value="").border = self.styles.thin_border
            ws.cell(row=row, column=4).fill = base_fill

            ws.cell(row=row, column=5, value="").border = self.styles.thin_border
            ws.cell(row=row, column=5).fill = base_fill

            ws.cell(row=row, column=6, value=note).border = self.styles.thin_border
            ws.cell(row=row, column=6).fill = base_fill

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 20

        ws.row_dimensions[1].height = 30

        return self

    def save(self, output_file: str = None):
        """Save workbook"""
        output = output_file or Config.OUTPUT_FILE
        self.wb.save(output)
        print(f"Saved: {output}")
        return self


# ============================================================
# MAIN
# ============================================================


def main():
    print("Creating Seekurity SIEM Deployment Schedule...")
    print()

    builder = ScheduleBuilder()
    builder.create_schedule_sheet()
    builder.create_summary_sheet()
    builder.create_milestone_sheet()
    builder.create_resource_sheet()
    builder.save()

    # Summary
    total_tasks = sum(len(p["tasks"]) for p in Config.PHASES)
    total_days = sum(sum(t["duration"] for t in p["tasks"]) for p in Config.PHASES)

    print()
    print("=" * 50)
    print("Deployment Schedule Created")
    print("=" * 50)
    print(f"Output: {Config.OUTPUT_FILE}")
    print()
    print("Sheets:")
    print("  - Deployment Schedule (main schedule)")
    print("  - Summary (phase overview)")
    print("  - Milestones (key milestones)")
    print("  - Resources (team allocation)")
    print()
    print(f"Total Tasks: {total_tasks}")
    print(f"Total Duration: {total_days} days")
    print()
    print("Phases:")
    for phase in Config.PHASES:
        days = sum(t["duration"] for t in phase["tasks"])
        print(f"  - {phase['name']}: {len(phase['tasks'])} tasks, {days} days")


if __name__ == "__main__":
    main()
