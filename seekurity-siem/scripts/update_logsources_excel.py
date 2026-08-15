"""
Seekurity SIEM Log Sources Excel 업데이트 스크립트

기존 Excel 파일의 헤더를 수정하고 새 컬럼을 추가합니다.
- IP Address를 필수 영역으로 이동
- Manager → 담당자로 변경
- 제조사명, 장비명, 버전 컬럼 추가
"""

import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


def update_logsources_excel(input_file: str, output_file: str = None):
    """
    Log Sources Excel 파일 헤더 업데이트 - 새 Workbook 생성 방식

    Args:
        input_file: 원본 Excel 파일 경로
        output_file: 출력 파일 경로 (None이면 원본 파일명_updated.xlsx)
    """

    # 출력 파일명 설정
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d")
        output_file = input_file.replace(".xlsx", f"_updated_{timestamp}.xlsx")

    # 원본 파일 백업
    backup_file = input_file.replace(".xlsx", "_backup.xlsx")
    shutil.copy(input_file, backup_file)
    print(f"Backup created: {backup_file}")

    # 원본 Excel 파일 로드 (data_only=True로 수식 결과값 읽기)
    wb_data = load_workbook(input_file, data_only=True)
    ws_data = wb_data["LogSources"]
    ws_ref = wb_data["Reference"]

    # 스타일 정의
    header_fill_mandatory = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_fill_additional = PatternFill(
        start_color="70AD47", end_color="70AD47", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 기존 데이터 읽기 (Row 4부터 데이터)
    existing_data = []
    for row_idx in range(4, ws_data.max_row + 1):
        system_type = ws_data.cell(row=row_idx, column=1).value
        system_name = ws_data.cell(row=row_idx, column=2).value
        protocol = ws_data.cell(row=row_idx, column=3).value
        log_source_name = ws_data.cell(row=row_idx, column=4).value
        description = ws_data.cell(row=row_idx, column=5).value
        ip_address = ws_data.cell(row=row_idx, column=6).value
        manager = ws_data.cell(row=row_idx, column=7).value

        if system_name is not None:
            existing_data.append(
                {
                    "system_type": system_type if system_type else "",
                    "system_name": system_name,
                    "protocol": protocol,
                    "log_source_name": log_source_name,
                    "description": description,
                    "ip_address": ip_address,
                    "manager": manager,
                }
            )

    print(f"Found {len(existing_data)} data rows")

    # Reference 데이터 읽기
    reference_data = []
    for row_idx in range(2, ws_ref.max_row + 1):
        sys_name = ws_ref.cell(row=row_idx, column=1).value
        sys_type = ws_ref.cell(row=row_idx, column=2).value
        if sys_name:
            reference_data.append({"system_name": sys_name, "system_type": sys_type})

    # 새 Workbook 생성
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "LogSources"

    # Row 1: 영역 구분 헤더
    ws_new.cell(row=1, column=1, value="필수")
    ws_new.cell(row=1, column=6, value="추가")

    for col_idx in range(1, 11):
        cell = ws_new.cell(row=1, column=col_idx)
        if col_idx <= 5:
            cell.fill = header_fill_mandatory
        else:
            cell.fill = header_fill_additional
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws_new.merge_cells("A1:E1")
    ws_new.merge_cells("F1:J1")

    # Row 2: 컬럼 헤더
    headers = [
        "System Type (수정 금지)",  # A
        "System Name",  # B
        "Protocol",  # C
        "Log Source Name",  # D
        "IP Address",  # E (이동됨)
        "Log Source Description",  # F
        "제조사명",  # G (신규)
        "장비명",  # H (신규)
        "버전",  # I (신규)
        "담당자",  # J (Manager에서 변경)
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_new.cell(row=2, column=col_idx, value=header)
        if col_idx <= 5:
            cell.fill = header_fill_mandatory
        else:
            cell.fill = header_fill_additional
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 데이터 작성 (Row 3부터)
    for row_idx, data in enumerate(existing_data, start=3):
        ws_new.cell(row=row_idx, column=1, value=data["system_type"])
        ws_new.cell(row=row_idx, column=2, value=data["system_name"])
        ws_new.cell(row=row_idx, column=3, value=data["protocol"])
        ws_new.cell(row=row_idx, column=4, value=data["log_source_name"])
        ws_new.cell(row=row_idx, column=5, value=data["ip_address"])
        ws_new.cell(row=row_idx, column=6, value=data["description"])
        ws_new.cell(row=row_idx, column=7, value=None)  # 제조사명
        ws_new.cell(row=row_idx, column=8, value=None)  # 장비명
        ws_new.cell(row=row_idx, column=9, value=None)  # 버전
        ws_new.cell(row=row_idx, column=10, value=data["manager"])

        for col_idx in range(1, 11):
            cell = ws_new.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 컬럼 너비 조정
    column_widths = {
        "A": 25,
        "B": 35,
        "C": 12,
        "D": 30,
        "E": 15,
        "F": 25,
        "G": 15,
        "H": 20,
        "I": 12,
        "J": 12,
    }
    for col_letter, width in column_widths.items():
        ws_new.column_dimensions[col_letter].width = width

    # 행 높이 조정
    ws_new.row_dimensions[1].height = 25
    ws_new.row_dimensions[2].height = 30

    # Reference 시트 생성
    ws_ref_new = wb_new.create_sheet(title="Reference")
    ws_ref_new.cell(row=1, column=1, value="System Name")
    ws_ref_new.cell(row=1, column=2, value="System Type")

    for col_idx in range(1, 3):
        cell = ws_ref_new.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

    for row_idx, ref in enumerate(reference_data, start=2):
        ws_ref_new.cell(row=row_idx, column=1, value=ref["system_name"])
        ws_ref_new.cell(row=row_idx, column=2, value=ref["system_type"])

    ws_ref_new.column_dimensions["A"].width = 40
    ws_ref_new.column_dimensions["B"].width = 20

    # 파일 저장
    wb_new.save(output_file)
    print(f"Updated file saved: {output_file}")

    return output_file


if __name__ == "__main__":
    input_file = "SeekuritySIEM_Logsources_KOVAN.xlsx"
    output_file = update_logsources_excel(
        input_file, "SeekuritySIEM_Logsources_KOVAN_v2.xlsx"
    )
    print(f"\nDone! Check the output file: {output_file}")
