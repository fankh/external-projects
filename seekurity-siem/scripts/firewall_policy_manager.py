"""
Seekurity SIEM Firewall Policy Manager

방화벽 정책 요청서 Excel 파일 생성

Usage:
    python firewall_policy_manager.py create
    python firewall_policy_manager.py create --logsource ../output/LogSources.xlsx

Output:
    SeekuritySIEM_Firewall_Policy.xlsx
"""

import argparse
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION
# ============================================================


class Config:
    """Configuration for firewall policy"""

    OUTPUT_FILE = "SeekuritySIEM_Firewall_Policy.xlsx"

    # Colors
    COLORS = {
        "title": "1B2838",
        "header": "2E5090",
        "subheader": "B4C6E7",
        "section": "E7E6E6",
        "row_even": "FFFFFF",
        "row_odd": "F8F9FA",
        "border": "D0D0D0",
        "inbound": "E2EFDA",  # Light green
        "outbound": "FCE4D6",  # Light orange
        "bidirect": "DDEBF7",  # Light blue
        "approved": "C6EFCE",
        "pending": "FFEB9C",
        "rejected": "FFC7CE",
    }

    # Direction options
    DIRECTION_OPTIONS = ["Inbound", "Outbound", "Bidirectional"]

    # Protocol options
    PROTOCOL_OPTIONS = ["TCP", "UDP", "TCP/UDP", "ICMP", "ANY"]

    # Status options
    STATUS_OPTIONS = ["요청", "승인", "반려", "완료", "보류"]

    # SIEM Server info (to be filled)
    SIEM_SERVERS = [
        {
            "name": "SIEM Manager",
            "ip": "x.x.x.x",
            "role": "Seekurity SIEM Manager / OpenSearch",
        },
        {
            "name": "SIEM Collector #1",
            "ip": "x.x.x.x",
            "role": "Log Collector (Forwarder)",
        },
        {
            "name": "SIEM Collector #2",
            "ip": "x.x.x.x",
            "role": "Log Collector (Failover)",
        },
    ]

    # Standard firewall policies for SIEM
    STANDARD_POLICIES = [
        # Log 수집 (SS-Syslog-Receiver)
        {
            "category": "Log 수집",
            "desc": "Syslog 수집 (UDP)",
            "src": "Log Source (ALL)",
            "src_ip": "ANY",
            "dst": "SIEM Collector",
            "dst_ip": "SIEM_COLLECTOR_IP",
            "port": "514",
            "protocol": "UDP",
            "direction": "Inbound",
            "note": "SS-Syslog-Receiver",
        },
        # 웹 접속 (Nginx)
        {
            "category": "웹 접속",
            "desc": "Nginx HTTPS",
            "src": "관리자 PC",
            "src_ip": "ADMIN_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "443",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "웹 접속 (리버스 프록시)",
        },
        {
            "category": "웹 접속",
            "desc": "Nginx HTTP Redirect",
            "src": "관리자 PC",
            "src_ip": "ADMIN_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "80",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "HTTPS 리다이렉트",
        },
        # 내부 서비스 (Cluster 통신)
        {
            "category": "내부 서비스",
            "desc": "SS-API",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "23001",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "REST API 서버",
        },
        {
            "category": "내부 서비스",
            "desc": "SS-Console",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "23002",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "웹 UI 서버",
        },
        {
            "category": "내부 서비스",
            "desc": "OpenSearch API",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "19200",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "검색 엔진 API",
        },
        {
            "category": "내부 서비스",
            "desc": "OpenSearch Cluster",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Nodes",
            "dst_ip": "SIEM_ALL_IP",
            "port": "19300",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "노드 간 통신",
        },
        {
            "category": "내부 서비스",
            "desc": "PostgreSQL",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "15432",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "데이터베이스",
        },
        {
            "category": "내부 서비스",
            "desc": "Kafka",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "19092",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "메시지 브로커",
        },
        {
            "category": "내부 서비스",
            "desc": "Zookeeper",
            "src": "SIEM Nodes",
            "src_ip": "SIEM_ALL_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "12181",
            "protocol": "TCP",
            "direction": "Bidirectional",
            "note": "Kafka 코디네이터",
        },
        # 관리 접속
        {
            "category": "관리 접속",
            "desc": "SSH 관리",
            "src": "관리자 PC",
            "src_ip": "ADMIN_IP",
            "dst": "SIEM Servers",
            "dst_ip": "SIEM_ALL_IP",
            "port": "22",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "서버 관리",
        },
        # Engineer PC (구축 엔지니어)
        {
            "category": "엔지니어 접속",
            "desc": "엔지니어 SSH 접속",
            "src": "엔지니어 PC",
            "src_ip": "ENGINEER_IP",
            "dst": "SIEM Servers",
            "dst_ip": "SIEM_ALL_IP",
            "port": "22",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "구축/운영 엔지니어",
        },
        {
            "category": "엔지니어 접속",
            "desc": "엔지니어 웹 접속 (HTTPS)",
            "src": "엔지니어 PC",
            "src_ip": "ENGINEER_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "443",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "Nginx HTTPS",
        },
        {
            "category": "엔지니어 접속",
            "desc": "엔지니어 웹 접속 (HTTP)",
            "src": "엔지니어 PC",
            "src_ip": "ENGINEER_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "80",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "HTTPS 리다이렉트",
        },
        # API 연동
        {
            "category": "API 연동",
            "desc": "SS-API 외부 연동",
            "src": "연동 서버",
            "src_ip": "API_CLIENT_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "23001",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "REST API 서버",
        },
        {
            "category": "API 연동",
            "desc": "OpenSearch API 외부",
            "src": "연동 서버",
            "src_ip": "API_CLIENT_IP",
            "dst": "SIEM Manager",
            "dst_ip": "SIEM_MANAGER_IP",
            "port": "19200",
            "protocol": "TCP",
            "direction": "Inbound",
            "note": "검색 엔진 API",
        },
        # External (Outbound)
        {
            "category": "외부 통신",
            "desc": "OS 패키지 업데이트",
            "src": "SIEM Servers",
            "src_ip": "SIEM_ALL_IP",
            "dst": "Internet",
            "dst_ip": "ANY",
            "port": "80,443",
            "protocol": "TCP",
            "direction": "Outbound",
            "note": "yum/apt repository",
        },
        {
            "category": "외부 통신",
            "desc": "Threat Intelligence",
            "src": "SIEM Manager",
            "src_ip": "SIEM_MANAGER_IP",
            "dst": "Internet",
            "dst_ip": "ANY",
            "port": "443",
            "protocol": "TCP",
            "direction": "Outbound",
            "note": "CVE/IOC 업데이트",
        },
        {
            "category": "외부 통신",
            "desc": "NTP 시간 동기화",
            "src": "SIEM Servers",
            "src_ip": "SIEM_ALL_IP",
            "dst": "NTP Server",
            "dst_ip": "NTP_SERVER_IP",
            "port": "123",
            "protocol": "UDP",
            "direction": "Outbound",
            "note": "시간 동기화 필수",
        },
        {
            "category": "외부 통신",
            "desc": "DNS 조회",
            "src": "SIEM Servers",
            "src_ip": "SIEM_ALL_IP",
            "dst": "DNS Server",
            "dst_ip": "DNS_SERVER_IP",
            "port": "53",
            "protocol": "TCP/UDP",
            "direction": "Outbound",
            "note": "DNS Resolution",
        },
        # Notification
        {
            "category": "알림 연동",
            "desc": "SMTP 메일 발송",
            "src": "SIEM Manager",
            "src_ip": "SIEM_MANAGER_IP",
            "dst": "Mail Server",
            "dst_ip": "MAIL_SERVER_IP",
            "port": "25,587",
            "protocol": "TCP",
            "direction": "Outbound",
            "note": "알림 메일 발송",
        },
        {
            "category": "알림 연동",
            "desc": "Slack/Teams Webhook",
            "src": "SIEM Manager",
            "src_ip": "SIEM_MANAGER_IP",
            "dst": "Internet",
            "dst_ip": "ANY",
            "port": "443",
            "protocol": "TCP",
            "direction": "Outbound",
            "note": "메신저 알림",
        },
    ]

    # Log source specific policies (to be generated from log sources)
    LOG_SOURCE_POLICIES = [
        {
            "category": "Firewall",
            "desc": "방화벽 Syslog",
            "src": "Firewall",
            "port": "514",
            "protocol": "UDP",
        },
        {
            "category": "IPS/IDS",
            "desc": "IPS Syslog",
            "src": "IPS/IDS",
            "port": "514",
            "protocol": "UDP",
        },
        {
            "category": "VPN",
            "desc": "VPN Syslog",
            "src": "VPN",
            "port": "514",
            "protocol": "UDP",
        },
        {
            "category": "WAF",
            "desc": "WAF Syslog",
            "src": "WAF",
            "port": "514",
            "protocol": "UDP",
        },
        {
            "category": "Server",
            "desc": "Windows Event (Agent)",
            "src": "Windows Server",
            "port": "1514",
            "protocol": "TCP",
        },
        {
            "category": "Server",
            "desc": "Linux Syslog (Agent)",
            "src": "Linux Server",
            "port": "1514",
            "protocol": "TCP",
        },
        {
            "category": "Database",
            "desc": "DB Audit Log",
            "src": "Database",
            "port": "514",
            "protocol": "TCP",
        },
        {
            "category": "EDR",
            "desc": "EDR API 연동",
            "src": "SIEM",
            "port": "443",
            "protocol": "TCP",
        },
    ]


# ============================================================
# STYLES
# ============================================================


class Styles:
    """Pre-defined styles"""

    def __init__(self):
        c = Config.COLORS

        self.title_fill = PatternFill(
            start_color=c["title"], end_color=c["title"], fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color=c["header"], end_color=c["header"], fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color=c["subheader"], end_color=c["subheader"], fill_type="solid"
        )
        self.section_fill = PatternFill(
            start_color=c["section"], end_color=c["section"], fill_type="solid"
        )
        self.row_even = PatternFill(
            start_color=c["row_even"], end_color=c["row_even"], fill_type="solid"
        )
        self.row_odd = PatternFill(
            start_color=c["row_odd"], end_color=c["row_odd"], fill_type="solid"
        )

        self.direction_fills = {
            "Inbound": PatternFill(
                start_color=c["inbound"], end_color=c["inbound"], fill_type="solid"
            ),
            "Outbound": PatternFill(
                start_color=c["outbound"], end_color=c["outbound"], fill_type="solid"
            ),
            "Bidirectional": PatternFill(
                start_color=c["bidirect"], end_color=c["bidirect"], fill_type="solid"
            ),
        }

        self.status_fills = {
            "승인": PatternFill(
                start_color=c["approved"], end_color=c["approved"], fill_type="solid"
            ),
            "요청": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
            "반려": PatternFill(
                start_color=c["rejected"], end_color=c["rejected"], fill_type="solid"
            ),
            "완료": PatternFill(
                start_color=c["approved"], end_color=c["approved"], fill_type="solid"
            ),
            "보류": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
        }

        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.subheader_font = Font(bold=True, size=10, color="1B2838")
        self.data_font = Font(size=9)
        self.section_font = Font(bold=True, size=10)

        self.thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        self.header_border = Border(
            left=Side(style="thin", color="8EA9DB"),
            right=Side(style="thin", color="8EA9DB"),
            top=Side(style="medium", color="2E5090"),
            bottom=Side(style="medium", color="2E5090"),
        )

        self.center = Alignment(horizontal="center", vertical="center")
        self.center_wrap = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.left = Alignment(vertical="center")
        self.left_wrap = Alignment(vertical="center", wrap_text=True)


# ============================================================
# FIREWALL POLICY BUILDER
# ============================================================


class FirewallPolicyBuilder:
    """Build firewall policy Excel"""

    def __init__(self):
        self.styles = Styles()
        self.wb = Workbook()

    def _create_title_row(self, ws, title: str, merge_range: str):
        """Create title row"""
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(":")[0]]
        cell.value = title
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

    def _apply_header_style(self, ws, row: int, columns: List[tuple]):
        """Apply header style"""
        for col, (header, width) in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.styles.subheader_fill
            cell.font = self.styles.subheader_font
            cell.alignment = self.styles.center_wrap
            cell.border = self.styles.header_border
            ws.column_dimensions[get_column_letter(col)].width = width

    def create_policy_request_sheet(self):
        """Create main policy request sheet"""
        ws = self.wb.active
        ws.title = "방화벽 정책 요청서"

        self._create_title_row(ws, "Seekurity SIEM 방화벽 정책 요청서", "A1:L1")

        # Project info (editable - separate cells)
        ws.merge_cells("A2:B2")
        ws["A2"] = "프로젝트:"
        ws["A2"].fill = self.styles.section_fill
        ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
        ws["A2"].font = Font(bold=True, size=10)

        ws.merge_cells("C2:D2")
        ws["C2"] = "[고객사명] SIEM 구축"
        ws["C2"].fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
        ws["C2"].border = self.styles.thin_border

        ws["E2"] = "요청일:"
        ws["E2"].fill = self.styles.section_fill
        ws["E2"].alignment = Alignment(horizontal="right", vertical="center")
        ws["E2"].font = Font(bold=True, size=10)

        ws.merge_cells("F2:G2")
        ws["F2"] = datetime.now().strftime("%Y-%m-%d")
        ws["F2"].fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F2"].border = self.styles.thin_border

        ws["H2"] = "요청자:"
        ws["H2"].fill = self.styles.section_fill
        ws["H2"].alignment = Alignment(horizontal="right", vertical="center")
        ws["H2"].font = Font(bold=True, size=10)

        ws.merge_cells("I2:L2")
        ws["I2"] = ""
        ws["I2"].fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        ws["I2"].alignment = Alignment(horizontal="left", vertical="center")
        ws["I2"].border = self.styles.thin_border

        # Headers
        headers = [
            ("No.", 5),
            ("Category", 12),
            ("설명", 25),
            ("Source", 15),
            ("Source IP", 15),
            ("Destination", 15),
            ("Destination IP", 15),
            ("Port", 12),
            ("Protocol", 10),
            ("Direction", 12),
            ("Status", 10),
            ("비고", 20),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data - Standard policies (uniform white background for easy copy/paste)
        row_idx = 4
        prev_category = None
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        for idx, policy in enumerate(Config.STANDARD_POLICIES, start=1):
            # No.
            ws.cell(row=row_idx, column=1, value=idx).fill = white_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

            # Category
            cell = ws.cell(
                row=row_idx,
                column=2,
                value=policy["category"] if policy["category"] != prev_category else "",
            )
            cell.fill = (
                self.styles.section_fill
                if policy["category"] != prev_category
                else white_fill
            )
            cell.font = (
                self.styles.section_font
                if policy["category"] != prev_category
                else self.styles.data_font
            )
            cell.border = self.styles.thin_border
            prev_category = policy["category"]

            # Description
            ws.cell(row=row_idx, column=3, value=policy["desc"]).fill = white_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).font = self.styles.data_font

            # Source
            ws.cell(row=row_idx, column=4, value=policy["src"]).fill = white_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border
            ws.cell(row=row_idx, column=4).font = self.styles.data_font

            # Source IP
            ws.cell(row=row_idx, column=5, value=policy["src_ip"]).fill = white_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border
            ws.cell(row=row_idx, column=5).font = self.styles.data_font
            ws.cell(row=row_idx, column=5).alignment = self.styles.center

            # Destination
            ws.cell(row=row_idx, column=6, value=policy["dst"]).fill = white_fill
            ws.cell(row=row_idx, column=6).border = self.styles.thin_border
            ws.cell(row=row_idx, column=6).font = self.styles.data_font

            # Destination IP
            ws.cell(row=row_idx, column=7, value=policy["dst_ip"]).fill = white_fill
            ws.cell(row=row_idx, column=7).border = self.styles.thin_border
            ws.cell(row=row_idx, column=7).font = self.styles.data_font
            ws.cell(row=row_idx, column=7).alignment = self.styles.center

            # Port
            ws.cell(row=row_idx, column=8, value=policy["port"]).fill = white_fill
            ws.cell(row=row_idx, column=8).border = self.styles.thin_border
            ws.cell(row=row_idx, column=8).font = self.styles.data_font
            ws.cell(row=row_idx, column=8).alignment = self.styles.center

            # Protocol
            ws.cell(row=row_idx, column=9, value=policy["protocol"]).fill = white_fill
            ws.cell(row=row_idx, column=9).border = self.styles.thin_border
            ws.cell(row=row_idx, column=9).font = self.styles.data_font
            ws.cell(row=row_idx, column=9).alignment = self.styles.center

            # Direction
            direction = policy["direction"]
            ws.cell(
                row=row_idx, column=10, value=direction
            ).fill = self.styles.direction_fills.get(direction, white_fill)
            ws.cell(row=row_idx, column=10).border = self.styles.thin_border
            ws.cell(row=row_idx, column=10).font = self.styles.data_font
            ws.cell(row=row_idx, column=10).alignment = self.styles.center

            # Status
            ws.cell(
                row=row_idx, column=11, value="요청"
            ).fill = self.styles.status_fills["요청"]
            ws.cell(row=row_idx, column=11).border = self.styles.thin_border
            ws.cell(row=row_idx, column=11).font = self.styles.data_font
            ws.cell(row=row_idx, column=11).alignment = self.styles.center

            # Note
            ws.cell(row=row_idx, column=12, value=policy["note"]).fill = white_fill
            ws.cell(row=row_idx, column=12).border = self.styles.thin_border
            ws.cell(row=row_idx, column=12).font = self.styles.data_font

            row_idx += 1

        # Add 20 empty rows for additional entries
        for i in range(20):
            for col in range(1, 13):
                cell = ws.cell(row=row_idx, column=col, value="")
                cell.fill = white_fill
                cell.border = self.styles.thin_border
                cell.font = self.styles.data_font
                if col in [1, 5, 7, 8, 9, 10, 11]:
                    cell.alignment = self.styles.center
            row_idx += 1

        # Data validations
        last_row = row_idx - 1

        direction_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.DIRECTION_OPTIONS)}"',
            allow_blank=False,
        )
        ws.add_data_validation(direction_dv)
        direction_dv.add(f"J4:J{last_row + 20}")

        protocol_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.PROTOCOL_OPTIONS)}"',
            allow_blank=False,
        )
        ws.add_data_validation(protocol_dv)
        protocol_dv.add(f"I4:I{last_row + 20}")

        status_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(Config.STATUS_OPTIONS)}"',
            allow_blank=False,
        )
        ws.add_data_validation(status_dv)
        status_dv.add(f"K4:K{last_row + 20}")

        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "D4"

        return self

    def create_server_info_sheet(self):
        """Create SIEM server info sheet"""
        ws = self.wb.create_sheet(title="SIEM 서버 정보")

        self._create_title_row(ws, "SIEM 서버 정보", "A1:D1")

        # Headers
        headers = [("No.", 5), ("서버명", 18), ("IP Address", 18), ("역할", 35)]
        self._apply_header_style(ws, 3, headers)

        # Data
        for idx, server in enumerate(Config.SIEM_SERVERS, start=1):
            row_idx = idx + 3
            is_odd = (idx - 1) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row_idx, column=1, value=idx).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

            ws.cell(row=row_idx, column=2, value=server["name"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border

            ws.cell(row=row_idx, column=3, value=server["ip"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center

            ws.cell(row=row_idx, column=4, value=server["role"]).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border

        # Add empty rows for additional servers
        for i in range(4, 10):
            row_idx = i + 3
            is_odd = (i - 1) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = base_fill
                ws.cell(row=row_idx, column=col).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def create_port_reference_sheet(self):
        """Create port reference sheet"""
        ws = self.wb.create_sheet(title="Port Reference")

        self._create_title_row(ws, "SIEM 표준 Port Reference", "A1:E1")

        # Headers
        headers = [
            ("Service", 25),
            ("Port", 10),
            ("Protocol", 10),
            ("Direction", 12),
            ("설명", 40),
        ]
        self._apply_header_style(ws, 3, headers)

        # Port reference data - Seekurity SIEM 실제 Port
        ports = [
            {
                "service": "Nginx (HTTPS)",
                "port": "443",
                "protocol": "TCP",
                "direction": "Inbound",
                "desc": "웹 접속 (리버스 프록시)",
            },
            {
                "service": "Nginx (HTTP)",
                "port": "80",
                "protocol": "TCP",
                "direction": "Inbound",
                "desc": "HTTPS 리다이렉트",
            },
            {
                "service": "SS-API",
                "port": "23001",
                "protocol": "TCP",
                "direction": "Inbound",
                "desc": "REST API 서버",
            },
            {
                "service": "SS-Console",
                "port": "23002",
                "protocol": "TCP",
                "direction": "Inbound",
                "desc": "웹 UI 서버",
            },
            {
                "service": "SS-Syslog-Receiver",
                "port": "514",
                "protocol": "UDP",
                "direction": "Inbound",
                "desc": "Syslog 수신",
            },
            {
                "service": "OpenSearch API",
                "port": "19200",
                "protocol": "TCP",
                "direction": "Bidirectional",
                "desc": "검색 엔진 API",
            },
            {
                "service": "OpenSearch Transport",
                "port": "19300",
                "protocol": "TCP",
                "direction": "Bidirectional",
                "desc": "노드 간 통신",
            },
            {
                "service": "PostgreSQL",
                "port": "15432",
                "protocol": "TCP",
                "direction": "Bidirectional",
                "desc": "데이터베이스",
            },
            {
                "service": "Kafka",
                "port": "19092",
                "protocol": "TCP",
                "direction": "Bidirectional",
                "desc": "메시지 브로커",
            },
            {
                "service": "Zookeeper",
                "port": "12181",
                "protocol": "TCP",
                "direction": "Bidirectional",
                "desc": "Kafka 코디네이터",
            },
            {
                "service": "SSH",
                "port": "22",
                "protocol": "TCP",
                "direction": "Inbound",
                "desc": "서버 관리",
            },
            {
                "service": "NTP",
                "port": "123",
                "protocol": "UDP",
                "direction": "Outbound",
                "desc": "시간 동기화",
            },
            {
                "service": "DNS",
                "port": "53",
                "protocol": "TCP/UDP",
                "direction": "Outbound",
                "desc": "DNS Resolution",
            },
            {
                "service": "SMTP",
                "port": "25,587",
                "protocol": "TCP",
                "direction": "Outbound",
                "desc": "메일 발송",
            },
        ]

        for idx, port in enumerate(ports, start=1):
            row_idx = idx + 3
            is_odd = (idx - 1) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row_idx, column=1, value=port["service"]).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border

            ws.cell(row=row_idx, column=2, value=port["port"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border
            ws.cell(row=row_idx, column=2).alignment = self.styles.center

            ws.cell(row=row_idx, column=3, value=port["protocol"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center

            direction = port["direction"]
            ws.cell(
                row=row_idx, column=4, value=direction
            ).fill = self.styles.direction_fills.get(direction, base_fill)
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border
            ws.cell(row=row_idx, column=4).alignment = self.styles.center

            ws.cell(row=row_idx, column=5, value=port["desc"]).fill = base_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def create_checklist_sheet(self):
        """Create policy verification checklist"""
        ws = self.wb.create_sheet(title="검증 Checklist")

        self._create_title_row(ws, "방화벽 정책 검증 Checklist", "A1:D1")

        # Headers
        headers = [("No.", 5), ("검증 항목", 45), ("결과", 10), ("비고", 30)]
        self._apply_header_style(ws, 3, headers)

        # Checklist items (Seekurity SIEM 실제 Port)
        checklist = [
            "Log Source → Collector Syslog 수신 (514/UDP)",
            "관리자 PC → Nginx HTTPS (443/TCP)",
            "관리자 PC → Nginx HTTP (80/TCP)",
            "엔지니어 PC → SIEM Servers SSH (22/TCP)",
            "엔지니어 PC → SIEM Manager 웹 (443/TCP)",
            "SIEM 내부 → SS-API (23001/TCP)",
            "SIEM 내부 → SS-Console (23002/TCP)",
            "SIEM 내부 → OpenSearch API (19200/TCP)",
            "SIEM 내부 → OpenSearch Cluster (19300/TCP)",
            "SIEM 내부 → PostgreSQL (15432/TCP)",
            "SIEM 내부 → Kafka (19092/TCP)",
            "SIEM 내부 → Zookeeper (12181/TCP)",
            "SIEM → NTP 서버 통신 (123/UDP)",
            "SIEM → DNS 서버 통신 (53/TCP,UDP)",
            "SIEM → 메일 서버 통신 (25,587/TCP)",
            "SIEM → 외부 업데이트 (80,443/TCP)",
        ]

        for idx, item in enumerate(checklist, start=1):
            row_idx = idx + 3
            is_odd = (idx - 1) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row_idx, column=1, value=idx).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

            ws.cell(row=row_idx, column=2, value=item).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border

            ws.cell(row=row_idx, column=3, value="").fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center

            ws.cell(row=row_idx, column=4, value="").fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border

        # Result dropdown
        result_dv = DataValidation(
            type="list", formula1='"Pass,Fail,N/A"', allow_blank=True
        )
        ws.add_data_validation(result_dv)
        result_dv.add(f"C4:C{3 + len(checklist)}")

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def build_all(self):
        """Build all sheets"""
        self.create_policy_request_sheet()
        self.create_server_info_sheet()
        self.create_port_reference_sheet()
        self.create_checklist_sheet()
        return self

    def save(self, output_file: str = None):
        """Save workbook"""
        output = output_file or Config.OUTPUT_FILE
        self.wb.save(output)
        print(f"Saved: {output}")
        return self


# ============================================================
# MAIN
# ============================================================


def create_firewall_policy(output_file: str = None):
    """Create firewall policy Excel"""
    print("Creating Seekurity SIEM Firewall Policy Request...")
    print()

    builder = FirewallPolicyBuilder()
    builder.build_all()
    builder.save(output_file)

    print()
    print("=" * 50)
    print("Firewall Policy Request Created")
    print("=" * 50)
    print(f"Output: {output_file or Config.OUTPUT_FILE}")
    print()
    print("Sheets created:")
    print("  - 방화벽 정책 요청서 (main request)")
    print("  - SIEM 서버 정보 (server list)")
    print("  - Port Reference (standard ports)")
    print("  - 검증 Checklist (verification)")
    print()
    print(f"Total policies: {len(Config.STANDARD_POLICIES)}")


def main():
    parser = argparse.ArgumentParser(
        description="Seekurity SIEM Firewall Policy Manager"
    )
    subparsers = parser.add_subparsers(dest="command", help="Commands")

    # Create command
    create_parser = subparsers.add_parser(
        "create", help="Create firewall policy request"
    )
    create_parser.add_argument("--output", "-o", help="Output file")
    create_parser.add_argument("--logsource", "-l", help="Log source file to import")

    args = parser.parse_args()

    if args.command == "create":
        create_firewall_policy(args.output)
    else:
        create_firewall_policy()


if __name__ == "__main__":
    main()
