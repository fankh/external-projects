"""
Seekurity SIEM Log Source Manager

Log Source Excel 파일 생성 및 관리를 위한 통합 스크립트

Usage:
    python logsource_manager.py [command] [options]

Commands:
    create      Create new Excel from original file
    update      Update existing Excel with new data
    export      Export to CSV format
    validate    Validate Excel data

Examples:
    python logsource_manager.py create
    python logsource_manager.py create --input original.xlsx --output new.xlsx
    python logsource_manager.py validate --input file.xlsx
"""

import argparse
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION
# ============================================================


class Config:
    """Configuration for Excel styling and structure"""

    # File settings (override with --input/--output arguments)
    DEFAULT_INPUT = "logsources_input.xlsx"
    DEFAULT_OUTPUT = "logsources_output.xlsx"

    # Column structure
    HEADERS = [
        "System Type",
        "System Name",
        "Protocol",
        "Log Source Name",
        "IP Address",
        "Description",
        "제조사명",
        "장비명",
        "버전",
        "담당자",
    ]

    COLUMN_WIDTHS = {
        "A": 20,  # System Type
        "B": 38,  # System Name
        "C": 12,  # Protocol
        "D": 28,  # Log Source Name
        "E": 15,  # IP Address
        "F": 22,  # Description
        "G": 14,  # 제조사명
        "H": 18,  # 장비명
        "I": 10,  # 버전
        "J": 10,  # 담당자
    }

    # Protocol options for dropdown
    PROTOCOLS = ["Syslog", "API", "File", "Agent"]

    # Colors (Modern palette)
    COLORS = {
        "title": "1B2838",  # Dark navy
        "mandatory": "2E5090",  # Medium blue
        "additional": "3D7EAA",  # Light blue
        "header": "B4C6E7",  # Light blue-gray
        "header_text": "1B2838",  # Dark text
        "protected": "F2F2F2",  # Light gray
        "row_even": "FFFFFF",  # White
        "row_odd": "F8F9FA",  # Very light gray
        "border": "D0D0D0",  # Light gray border
        "border_header": "8EA9DB",  # Blue border
        # System type category colors
        "network": "E2EFDA",  # Light green
        "endpoint": "FCE4D6",  # Light orange
        "data_app": "DDEBF7",  # Light blue
    }


# ============================================================
# STYLES
# ============================================================


class Styles:
    """Pre-defined styles for Excel cells"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self._init_styles()

    def _init_styles(self):
        c = self.config.COLORS

        # Fills
        self.title_fill = PatternFill(
            start_color=c["title"], end_color=c["title"], fill_type="solid"
        )
        self.mandatory_fill = PatternFill(
            start_color=c["mandatory"], end_color=c["mandatory"], fill_type="solid"
        )
        self.additional_fill = PatternFill(
            start_color=c["additional"], end_color=c["additional"], fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color=c["header"], end_color=c["header"], fill_type="solid"
        )
        self.protected_fill = PatternFill(
            start_color=c["protected"], end_color=c["protected"], fill_type="solid"
        )
        self.row_even_fill = PatternFill(
            start_color=c["row_even"], end_color=c["row_even"], fill_type="solid"
        )
        self.row_odd_fill = PatternFill(
            start_color=c["row_odd"], end_color=c["row_odd"], fill_type="solid"
        )

        # System type fills
        self.system_type_fills = {
            "Network Security": PatternFill(
                start_color=c["network"], end_color=c["network"], fill_type="solid"
            ),
            "Endpoint Security": PatternFill(
                start_color=c["endpoint"], end_color=c["endpoint"], fill_type="solid"
            ),
            "Data & Application": PatternFill(
                start_color=c["data_app"], end_color=c["data_app"], fill_type="solid"
            ),
        }

        # Fonts
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.category_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_font = Font(bold=True, size=10, color=c["header_text"])
        self.data_font = Font(size=9)

        # Borders
        self.thin_border = Border(
            left=Side(style="thin", color=c["border"]),
            right=Side(style="thin", color=c["border"]),
            top=Side(style="thin", color=c["border"]),
            bottom=Side(style="thin", color=c["border"]),
        )
        self.header_border = Border(
            left=Side(style="thin", color=c["border_header"]),
            right=Side(style="thin", color=c["border_header"]),
            top=Side(style="medium", color=c["mandatory"]),
            bottom=Side(style="medium", color=c["mandatory"]),
        )
        self.title_border = Border(bottom=Side(style="medium", color=c["additional"]))

        # Alignments
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.center_wrap_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.left_align = Alignment(vertical="center")

        # Protection
        self.locked = Protection(locked=True)
        self.unlocked = Protection(locked=False)


# ============================================================
# DATA READER
# ============================================================


class DataReader:
    """Read data from original Excel file"""

    def __init__(self, input_file: str):
        self.input_file = input_file
        self.wb = None
        self.data = []
        self.reference_data = []

    def load(self):
        """Load workbook with data_only=True to get calculated values"""
        print(f"Loading: {self.input_file}")
        self.wb = load_workbook(self.input_file, data_only=True)
        return self

    def read_logsources(self, sheet_name: str = "LogSources", start_row: int = 4):
        """Read log source data from sheet"""
        ws = self.wb[sheet_name]
        self.data = []

        for row_idx in range(start_row, ws.max_row + 1):
            system_name = ws.cell(row=row_idx, column=2).value
            if system_name:
                self.data.append(
                    {
                        "system_type": ws.cell(row=row_idx, column=1).value or "",
                        "system_name": system_name,
                        "protocol": ws.cell(row=row_idx, column=3).value,
                        "log_source_name": ws.cell(row=row_idx, column=4).value,
                        "description": ws.cell(row=row_idx, column=5).value,
                        "ip_address": ws.cell(row=row_idx, column=6).value,
                        "manager": ws.cell(row=row_idx, column=7).value,
                    }
                )

        print(f"Found {len(self.data)} log sources")
        return self.data

    def read_reference(self, sheet_name: str = "Reference", start_row: int = 2):
        """Read reference data from sheet"""
        ws = self.wb[sheet_name]
        self.reference_data = []

        for row_idx in range(start_row, ws.max_row + 1):
            sys_name = ws.cell(row=row_idx, column=1).value
            if sys_name:
                self.reference_data.append(
                    {
                        "system_name": sys_name,
                        "system_type": ws.cell(row=row_idx, column=2).value,
                    }
                )

        print(f"Found {len(self.reference_data)} reference items")
        return self.reference_data


# ============================================================
# EXCEL BUILDER
# ============================================================


class ExcelBuilder:
    """Build Excel workbook with styling"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self.styles = Styles(self.config)
        self.wb = Workbook()
        self.ws = None

    def create_logsources_sheet(self, data: list):
        """Create main LogSources sheet"""
        self.ws = self.wb.active
        self.ws.title = "LogSources"

        self._create_title_row()
        self._create_category_row()
        self._create_header_row()
        self._create_data_rows(data)
        self._add_data_validations(len(data))
        self._set_column_widths()
        self._set_row_heights()
        self._set_freeze_panes()
        self._set_protection()

        return self

    def _create_title_row(self):
        """Row 1: Title"""
        self.ws.merge_cells("A1:J1")
        cell = self.ws["A1"]
        cell.value = "Seekurity SIEM Log Source 관리"
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center_align
        cell.protection = self.styles.locked

        for col in range(1, 11):
            self.ws.cell(row=1, column=col).border = self.styles.title_border

    def _create_category_row(self):
        """Row 2: Category (필수/추가)"""
        self.ws.merge_cells("A2:E2")
        self.ws.merge_cells("F2:J2")

        cell = self.ws["A2"]
        cell.value = "필수 항목"
        cell.fill = self.styles.mandatory_fill
        cell.font = self.styles.category_font
        cell.alignment = self.styles.center_align
        cell.protection = self.styles.locked

        cell = self.ws["F2"]
        cell.value = "추가 정보"
        cell.fill = self.styles.additional_fill
        cell.font = self.styles.category_font
        cell.alignment = self.styles.center_align
        cell.protection = self.styles.locked

        for col in range(1, 11):
            c = self.ws.cell(row=2, column=col)
            c.fill = (
                self.styles.mandatory_fill if col <= 5 else self.styles.additional_fill
            )

    def _create_header_row(self):
        """Row 3: Column headers"""
        for col, header in enumerate(self.config.HEADERS, start=1):
            cell = self.ws.cell(row=3, column=col, value=header)
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center_wrap_align
            cell.border = self.styles.header_border
            cell.protection = self.styles.locked

    def _create_data_rows(self, data: list):
        """Data rows starting from row 4"""
        for idx, row_data in enumerate(data):
            row_idx = idx + 4
            is_odd = idx % 2 == 1
            base_fill = (
                self.styles.row_odd_fill if is_odd else self.styles.row_even_fill
            )

            # Column A: System Type (locked, colored by category)
            cell = self.ws.cell(row=row_idx, column=1, value=row_data["system_type"])
            cell.fill = self.styles.system_type_fills.get(
                row_data["system_type"], self.styles.protected_fill
            )
            cell.font = self.styles.data_font
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.center_align
            cell.protection = self.styles.locked

            # Column B: System Name (unlocked)
            cell = self.ws.cell(row=row_idx, column=2, value=row_data["system_name"])
            cell.fill = base_fill
            cell.font = self.styles.data_font
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.left_align
            cell.protection = self.styles.unlocked

            # Column C: Protocol (unlocked)
            cell = self.ws.cell(row=row_idx, column=3, value=row_data["protocol"])
            cell.fill = base_fill
            cell.font = self.styles.data_font
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.center_align
            cell.protection = self.styles.unlocked

            # Columns D-J (unlocked)
            values = [
                row_data["log_source_name"],
                row_data["ip_address"],
                row_data["description"],
                None,  # 제조사명
                None,  # 장비명
                None,  # 버전
                row_data["manager"],
            ]
            for col, val in enumerate(values, start=4):
                cell = self.ws.cell(row=row_idx, column=col, value=val)
                cell.fill = base_fill
                cell.font = self.styles.data_font
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.left_align
                cell.protection = self.styles.unlocked

    def _add_data_validations(self, data_count: int):
        """Add dropdown validations"""
        last_row = 3 + data_count

        # Protocol dropdown
        protocol_list = ",".join(self.config.PROTOCOLS)
        protocol_dv = DataValidation(
            type="list",
            formula1=f'"{protocol_list}"',
            allow_blank=True,
        )
        protocol_dv.prompt = "Protocol 선택"
        protocol_dv.promptTitle = "Protocol"
        self.ws.add_data_validation(protocol_dv)
        protocol_dv.add(f"C4:C{last_row}")

        # System Name dropdown (from Reference sheet)
        system_name_dv = DataValidation(
            type="list",
            formula1="Reference!$A$2:$A$13",
            allow_blank=True,
        )
        system_name_dv.prompt = "System Name 선택"
        system_name_dv.promptTitle = "System Name"
        self.ws.add_data_validation(system_name_dv)
        system_name_dv.add(f"B4:B{last_row}")

    def _set_column_widths(self):
        """Set column widths"""
        for col, width in self.config.COLUMN_WIDTHS.items():
            self.ws.column_dimensions[col].width = width

    def _set_row_heights(self):
        """Set row heights"""
        self.ws.row_dimensions[1].height = 35  # Title
        self.ws.row_dimensions[2].height = 22  # Category
        self.ws.row_dimensions[3].height = 28  # Headers

    def _set_freeze_panes(self):
        """Freeze header rows"""
        self.ws.freeze_panes = "A4"

    def _set_protection(self):
        """Enable sheet protection with formatting allowed"""
        self.ws.protection.sheet = True
        self.ws.protection.formatCells = False
        self.ws.protection.formatColumns = False
        self.ws.protection.formatRows = False
        self.ws.protection.enable()

    def create_reference_sheet(self, reference_data: list):
        """Create Reference sheet"""
        ws_ref = self.wb.create_sheet(title="Reference")

        # Title
        ws_ref.merge_cells("A1:B1")
        cell = ws_ref["A1"]
        cell.value = "System Type Reference"
        cell.fill = self.styles.title_fill
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = self.styles.center_align

        # Headers
        ws_ref["A2"] = "System Name"
        ws_ref["B2"] = "System Type"
        for col in ["A", "B"]:
            cell = ws_ref[f"{col}2"]
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center_align
            cell.border = self.styles.header_border

        # Data
        for row_idx, ref in enumerate(reference_data, start=3):
            cell_name = ws_ref.cell(row=row_idx, column=1, value=ref["system_name"])
            cell_type = ws_ref.cell(row=row_idx, column=2, value=ref["system_type"])

            fill = self.styles.system_type_fills.get(
                ref["system_type"], self.styles.row_even_fill
            )
            cell_name.fill = fill
            cell_type.fill = fill
            cell_name.font = Font(size=10)
            cell_type.font = Font(size=10)
            cell_name.border = self.styles.thin_border
            cell_type.border = self.styles.thin_border
            cell_type.alignment = self.styles.center_align

        ws_ref.column_dimensions["A"].width = 45
        ws_ref.column_dimensions["B"].width = 22
        ws_ref.row_dimensions[1].height = 30
        ws_ref.row_dimensions[2].height = 25

        return self

    def create_legend_sheet(self):
        """Create Legend sheet for color reference"""
        ws_legend = self.wb.create_sheet(title="Legend")

        # Title
        ws_legend.merge_cells("A1:B1")
        cell = ws_legend["A1"]
        cell.value = "Color Legend"
        cell.fill = self.styles.title_fill
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = self.styles.center_align

        # Headers
        ws_legend["A2"] = "System Type"
        ws_legend["B2"] = "Color"
        for col in ["A", "B"]:
            cell = ws_legend[f"{col}2"]
            cell.fill = self.styles.header_fill
            cell.font = self.styles.header_font
            cell.alignment = self.styles.center_align
            cell.border = self.styles.header_border

        # Legend items
        legends = [
            ("Network Security", self.styles.system_type_fills["Network Security"]),
            ("Endpoint Security", self.styles.system_type_fills["Endpoint Security"]),
            ("Data & Application", self.styles.system_type_fills["Data & Application"]),
        ]

        for row_idx, (name, fill) in enumerate(legends, start=3):
            cell = ws_legend.cell(row=row_idx, column=1, value=name)
            cell.border = self.styles.thin_border
            cell.font = Font(size=10)
            cell = ws_legend.cell(row=row_idx, column=2, value="")
            cell.fill = fill
            cell.border = self.styles.thin_border

        ws_legend.column_dimensions["A"].width = 25
        ws_legend.column_dimensions["B"].width = 15
        ws_legend.row_dimensions[1].height = 30

        return self

    def save(self, output_file: str):
        """Save workbook to file"""
        self.wb.save(output_file)
        print(f"Saved: {output_file}")
        return self


# ============================================================
# MAIN FUNCTIONS
# ============================================================


def create_excel(input_file: str, output_file: str):
    """Create new Excel from original file"""

    # Read data
    reader = DataReader(input_file)
    reader.load()
    data = reader.read_logsources()
    reference_data = reader.read_reference()

    # Build Excel
    builder = ExcelBuilder()
    builder.create_logsources_sheet(data)
    builder.create_reference_sheet(reference_data)
    builder.create_legend_sheet()
    builder.save(output_file)


def create_template(output_file: str):
    """Create empty template Excel without input file"""

    # Sample data for template
    sample_data = [
        {
            "system_type": "Network Security",
            "system_name": "Firewall",
            "protocol": "Syslog",
            "log_source_name": "[샘플] Firewall-01",
            "ip_address": "x.x.x.x",
            "description": "방화벽 로그",
            "manufacturer": "",
            "model": "",
            "version": "",
            "manager": "",
        },
        {
            "system_type": "Network Security",
            "system_name": "IPS/IDS",
            "protocol": "Syslog",
            "log_source_name": "[샘플] IPS-01",
            "ip_address": "x.x.x.x",
            "description": "IPS 로그",
            "manufacturer": "",
            "model": "",
            "version": "",
            "manager": "",
        },
        {
            "system_type": "Endpoint Security",
            "system_name": "Windows Server",
            "protocol": "Agent",
            "log_source_name": "[샘플] WinServer-01",
            "ip_address": "x.x.x.x",
            "description": "Windows Event Log",
            "manufacturer": "",
            "model": "",
            "version": "",
            "manager": "",
        },
        {
            "system_type": "Endpoint Security",
            "system_name": "Linux Server",
            "protocol": "Syslog",
            "log_source_name": "[샘플] LinuxServer-01",
            "ip_address": "x.x.x.x",
            "description": "Linux Syslog",
            "manufacturer": "",
            "model": "",
            "version": "",
            "manager": "",
        },
        {
            "system_type": "Data & Application",
            "system_name": "Database",
            "protocol": "Syslog",
            "log_source_name": "[샘플] DB-01",
            "ip_address": "x.x.x.x",
            "description": "Database Audit Log",
            "manufacturer": "",
            "model": "",
            "version": "",
            "manager": "",
        },
    ]

    # Reference data (list of dicts format)
    reference_data = [
        {"system_name": "Firewall", "system_type": "Network Security"},
        {"system_name": "IPS/IDS", "system_type": "Network Security"},
        {"system_name": "VPN", "system_type": "Network Security"},
        {"system_name": "WAF", "system_type": "Network Security"},
        {"system_name": "NAC", "system_type": "Network Security"},
        {"system_name": "Proxy", "system_type": "Network Security"},
        {"system_name": "Windows Server", "system_type": "Endpoint Security"},
        {"system_name": "Linux Server", "system_type": "Endpoint Security"},
        {"system_name": "EDR", "system_type": "Endpoint Security"},
        {"system_name": "AV", "system_type": "Endpoint Security"},
        {"system_name": "Database", "system_type": "Data & Application"},
        {"system_name": "Web Server", "system_type": "Data & Application"},
        {"system_name": "Application", "system_type": "Data & Application"},
        {"system_name": "Mail Server", "system_type": "Data & Application"},
    ]

    # Build Excel
    builder = ExcelBuilder()
    builder.create_logsources_sheet(sample_data)
    builder.create_reference_sheet(reference_data)
    builder.create_legend_sheet()
    builder.save(output_file)

    # Summary
    print()
    print("=" * 50)
    print("Template Creation Complete")
    print("=" * 50)
    print(f"Output: {output_file}")
    print()
    print("Sheets created:")
    print("  - LogSources (sample data)")
    print("  - Reference (system type lookup)")
    print("  - Legend (color guide)")
    print()
    print("Note: Replace [샘플] entries with actual log sources")


def validate_excel(input_file: str):
    """Validate Excel file data"""
    reader = DataReader(input_file)
    reader.load()
    data = reader.read_logsources()

    print()
    print("=" * 50)
    print("Validation Report")
    print("=" * 50)

    # Check for missing values
    missing_ip = [d for d in data if not d["ip_address"]]
    missing_manager = [d for d in data if not d["manager"]]

    print(f"Total rows: {len(data)}")
    print(f"Missing IP Address: {len(missing_ip)}")
    print(f"Missing 담당자: {len(missing_manager)}")

    if missing_ip:
        print()
        print("Rows missing IP Address:")
        for d in missing_ip[:10]:
            print(f"  - {d['log_source_name']}")
        if len(missing_ip) > 10:
            print(f"  ... and {len(missing_ip) - 10} more")


def main():
    parser = argparse.ArgumentParser(
        description="Seekurity SIEM Log Source Manager",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    subparsers = parser.add_subparsers(dest="command", help="Commands")

    # Create command
    create_parser = subparsers.add_parser("create", help="Create new Excel")
    create_parser.add_argument(
        "--input", "-i", default=Config.DEFAULT_INPUT, help="Input Excel file"
    )
    create_parser.add_argument(
        "--output", "-o", default=Config.DEFAULT_OUTPUT, help="Output Excel file"
    )

    # Validate command
    validate_parser = subparsers.add_parser("validate", help="Validate Excel data")
    validate_parser.add_argument(
        "--input", "-i", default=Config.DEFAULT_OUTPUT, help="Excel file to validate"
    )

    # Template command
    template_parser = subparsers.add_parser("template", help="Create empty template")
    template_parser.add_argument(
        "--output",
        "-o",
        default="SeekuritySIEM_Logsources_Template.xlsx",
        help="Output Excel file",
    )

    args = parser.parse_args()

    if args.command == "create":
        create_excel(args.input, args.output)
    elif args.command == "validate":
        validate_excel(args.input)
    elif args.command == "template":
        create_template(args.output)
    else:
        # Default: template
        create_template("SeekuritySIEM_Logsources_Template.xlsx")


if __name__ == "__main__":
    main()
