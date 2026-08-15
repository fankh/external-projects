"""
Seekurity SIEM Log Sources - Improved Design
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation


def create_improved_excel():
    # Load original data
    wb_data = load_workbook("SeekuritySIEM_Logsources_KOVAN.xlsx", data_only=True)
    ws_data = wb_data["LogSources"]
    ws_ref = wb_data["Reference"]

    # Read existing data
    existing_data = []
    for row_idx in range(4, ws_data.max_row + 1):
        system_name = ws_data.cell(row=row_idx, column=2).value
        if system_name:
            existing_data.append(
                {
                    "system_type": ws_data.cell(row=row_idx, column=1).value or "",
                    "system_name": system_name,
                    "protocol": ws_data.cell(row=row_idx, column=3).value,
                    "log_source_name": ws_data.cell(row=row_idx, column=4).value,
                    "description": ws_data.cell(row=row_idx, column=5).value,
                    "ip_address": ws_data.cell(row=row_idx, column=6).value,
                    "manager": ws_data.cell(row=row_idx, column=7).value,
                }
            )

    # Read reference data
    reference_data = []
    for row_idx in range(2, ws_ref.max_row + 1):
        sys_name = ws_ref.cell(row=row_idx, column=1).value
        if sys_name:
            reference_data.append(
                {
                    "system_name": sys_name,
                    "system_type": ws_ref.cell(row=row_idx, column=2).value,
                }
            )

    print(f"Data rows: {len(existing_data)}")

    # Create NEW workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "LogSources"

    # === MODERN COLOR PALETTE ===
    title_fill = PatternFill(
        start_color="1B2838", end_color="1B2838", fill_type="solid"
    )
    title_font = Font(bold=True, size=16, color="FFFFFF")

    mandatory_fill = PatternFill(
        start_color="2E5090", end_color="2E5090", fill_type="solid"
    )
    additional_fill = PatternFill(
        start_color="3D7EAA", end_color="3D7EAA", fill_type="solid"
    )
    category_font = Font(bold=True, size=11, color="FFFFFF")

    header_fill = PatternFill(
        start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"
    )
    header_font = Font(bold=True, size=10, color="1B2838")

    # System type colors
    network_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    endpoint_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    data_app_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )

    protected_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    row_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_odd = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    header_border = Border(
        left=Side(style="thin", color="8EA9DB"),
        right=Side(style="thin", color="8EA9DB"),
        top=Side(style="medium", color="2E5090"),
        bottom=Side(style="medium", color="2E5090"),
    )

    locked = Protection(locked=True)
    unlocked = Protection(locked=False)

    # === ROW 1: Title ===
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Seekurity SIEM Log Source 관리"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.protection = locked
    for col in range(1, 11):
        ws.cell(row=1, column=col).border = Border(
            bottom=Side(style="medium", color="3D7EAA")
        )

    # === ROW 2: Category ===
    ws.merge_cells("A2:E2")
    ws.merge_cells("F2:J2")

    cell = ws["A2"]
    cell.value = "필수 항목"
    cell.fill = mandatory_fill
    cell.font = category_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.protection = locked

    cell = ws["F2"]
    cell.value = "추가 정보"
    cell.fill = additional_fill
    cell.font = category_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.protection = locked

    for col in range(1, 11):
        c = ws.cell(row=2, column=col)
        if col <= 5:
            c.fill = mandatory_fill
        else:
            c.fill = additional_fill

    # === ROW 3: Headers ===
    headers = [
        "System Type",
        "System Name",
        "Protocol",
        "Log Source Name",
        "IP Address",
        "Description",
        "제조사명",
        "장비명",
        "버전",
        "담당자",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = header_border
        cell.protection = locked

    # === DATA ROWS ===
    system_type_colors = {
        "Network Security": network_fill,
        "Endpoint Security": endpoint_fill,
        "Data & Application": data_app_fill,
    }

    for idx, data in enumerate(existing_data):
        row_idx = idx + 4
        is_odd = idx % 2 == 1
        base_fill = row_odd if is_odd else row_even

        # Column A: System Type - colored by category
        cell = ws.cell(row=row_idx, column=1, value=data["system_type"])
        cell.fill = system_type_colors.get(data["system_type"], protected_fill)
        cell.font = Font(size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = locked

        # Column B: System Name
        cell = ws.cell(row=row_idx, column=2, value=data["system_name"])
        cell.fill = base_fill
        cell.font = Font(size=9)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        cell.protection = unlocked

        # Column C: Protocol
        cell = ws.cell(row=row_idx, column=3, value=data["protocol"])
        cell.fill = base_fill
        cell.font = Font(size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = unlocked

        # Columns D-J
        values = [
            data["log_source_name"],
            data["ip_address"],
            data["description"],
            None,
            None,
            None,
            data["manager"],
        ]
        for col, val in enumerate(values, start=4):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = base_fill
            cell.font = Font(size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            cell.protection = unlocked

    # === DATA VALIDATION ===
    protocol_dv = DataValidation(
        type="list", formula1='"Syslog,API,File,Agent"', allow_blank=True
    )
    protocol_dv.prompt = "Protocol 선택"
    protocol_dv.promptTitle = "Protocol"
    ws.add_data_validation(protocol_dv)
    protocol_dv.add(f"C4:C{3 + len(existing_data)}")

    system_name_dv = DataValidation(
        type="list", formula1="Reference!$A$2:$A$13", allow_blank=True
    )
    system_name_dv.prompt = "System Name 선택"
    system_name_dv.promptTitle = "System Name"
    ws.add_data_validation(system_name_dv)
    system_name_dv.add(f"B4:B{3 + len(existing_data)}")

    # === COLUMN WIDTHS ===
    widths = {
        "A": 20,
        "B": 38,
        "C": 12,
        "D": 28,
        "E": 15,
        "F": 22,
        "G": 14,
        "H": 18,
        "I": 10,
        "J": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # === ROW HEIGHTS ===
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 28

    # === FREEZE PANES ===
    ws.freeze_panes = "A4"

    # === SHEET PROTECTION ===
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.enable()

    # ========================================
    # === REFERENCE SHEET ===
    # ========================================
    ws_ref_new = wb.create_sheet(title="Reference")

    # Title
    ws_ref_new.merge_cells("A1:B1")
    cell = ws_ref_new["A1"]
    cell.value = "System Type Reference"
    cell.fill = title_fill
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    ws_ref_new["A2"] = "System Name"
    ws_ref_new["B2"] = "System Type"
    for col in ["A", "B"]:
        cell = ws_ref_new[f"{col}2"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    # Data with category colors
    for row_idx, ref in enumerate(reference_data, start=3):
        cell_name = ws_ref_new.cell(row=row_idx, column=1, value=ref["system_name"])
        cell_type = ws_ref_new.cell(row=row_idx, column=2, value=ref["system_type"])

        fill = system_type_colors.get(ref["system_type"], row_even)
        cell_name.fill = fill
        cell_type.fill = fill
        cell_name.font = Font(size=10)
        cell_type.font = Font(size=10)
        cell_name.border = thin_border
        cell_type.border = thin_border
        cell_type.alignment = Alignment(horizontal="center", vertical="center")

    ws_ref_new.column_dimensions["A"].width = 45
    ws_ref_new.column_dimensions["B"].width = 22
    ws_ref_new.row_dimensions[1].height = 30
    ws_ref_new.row_dimensions[2].height = 25

    # ========================================
    # === LEGEND SHEET ===
    # ========================================
    ws_legend = wb.create_sheet(title="Legend")

    ws_legend.merge_cells("A1:B1")
    cell = ws_legend["A1"]
    cell.value = "Color Legend"
    cell.fill = title_fill
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    legends = [
        ("Network Security", network_fill),
        ("Endpoint Security", endpoint_fill),
        ("Data & Application", data_app_fill),
    ]

    ws_legend["A2"] = "System Type"
    ws_legend["B2"] = "Color"
    ws_legend["A2"].fill = header_fill
    ws_legend["B2"].fill = header_fill
    ws_legend["A2"].font = header_font
    ws_legend["B2"].font = header_font
    ws_legend["A2"].alignment = Alignment(horizontal="center")
    ws_legend["B2"].alignment = Alignment(horizontal="center")
    ws_legend["A2"].border = header_border
    ws_legend["B2"].border = header_border

    for row_idx, (name, fill) in enumerate(legends, start=3):
        cell = ws_legend.cell(row=row_idx, column=1, value=name)
        cell.border = thin_border
        cell.font = Font(size=10)
        cell = ws_legend.cell(row=row_idx, column=2, value="")
        cell.fill = fill
        cell.border = thin_border

    ws_legend.column_dimensions["A"].width = 25
    ws_legend.column_dimensions["B"].width = 15
    ws_legend.row_dimensions[1].height = 30

    # Save
    wb.save("SeekuritySIEM_Logsources_KOVAN_v2.xlsx")
    print("Saved: SeekuritySIEM_Logsources_KOVAN_v2.xlsx")
    print()
    print("Design improvements:")
    print("- Modern color palette (dark navy title)")
    print("- System Type color-coded by category:")
    print("  * Network Security: Light Green")
    print("  * Endpoint Security: Light Orange")
    print("  * Data & Application: Light Blue")
    print("- Alternating row colors for readability")
    print("- Improved Reference sheet with colors")
    print("- New Legend sheet added")


if __name__ == "__main__":
    create_improved_excel()
