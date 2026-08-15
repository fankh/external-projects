"""
Seekurity SIEM SOW (Statement of Work) Manager

SOW Excel 파일 생성 및 일정표 업데이트

Usage:
    python sow_manager.py create              # Create new SOW
    python sow_manager.py create --project "Project Name" --customer "Customer"
    python sow_manager.py update-schedule     # Update schedule from SOW
    python sow_manager.py sync                # Sync SOW and Schedule

Output:
    SeekuritySIEM_SOW.xlsx
"""

import argparse
from datetime import datetime, timedelta
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION
# ============================================================


class SOWConfig:
    """SOW Configuration"""

    OUTPUT_FILE = "SeekuritySIEM_SOW.xlsx"

    # Default project info (override with --project/--customer arguments)
    PROJECT_NAME = "Seekurity SIEM 구축 프로젝트"
    DOCUMENT_TYPE = "작업 명세서 (Statement of Work)"
    VERSION = "v1.0"
    CUSTOMER = "[고객사명]"
    PARTNER = "브로드밴드시큐리티"

    # Colors
    COLORS = {
        "title": "1B2838",
        "header": "2E5090",
        "subheader": "B4C6E7",
        "row_even": "FFFFFF",
        "row_odd": "F8F9FA",
        "border": "D0D0D0",
        "section": "E7E6E6",
        # Phase colors
        "phase_analysis": "5B9BD5",
        "phase_deploy": "70AD47",
        "phase_integrate": "FFC000",
        "phase_config": "ED7D31",
        "phase_stabilize": "9E480E",
        "phase_close": "7030A0",
        # Status colors
        "complete": "C6EFCE",
        "in_progress": "FFEB9C",
        "delayed": "FFC7CE",
        "pending": "F2F2F2",
        # Risk colors
        "high": "FF6B6B",
        "medium": "FFE066",
        "low": "69DB7C",
    }

    # SOW Summary items
    SOW_SUMMARY = [
        {
            "category": "1. 프로젝트 개요",
            "item": "프로젝트명",
            "detail": "SIEM 구축 및 보안 관제 고도화 프로젝트",
            "note": "",
        },
        {
            "category": "1. 프로젝트 개요",
            "item": "수행 기간",
            "detail": "착수일로부터 4주 (안정화 기간 포함)",
            "note": "상세 일정 시트 참조",
        },
        {
            "category": "2. 작업 범위",
            "item": "인프라 구축",
            "detail": "SIEM 서버(S/W) 설치 및 클러스터링 구성",
            "note": "",
        },
        {
            "category": "2. 작업 범위",
            "item": "로그 연동",
            "detail": "방화벽, IPS, 서버(Win/Linux) 등 주요 자산 로그 연동",
            "note": "자산 목록 시트 참조",
        },
        {
            "category": "2. 작업 범위",
            "item": "탐지 시나리오",
            "detail": "핵심 탐지 시나리오 10종 개발 및 오탐 최적화(Tuning)",
            "note": "탐지 룰 시트 참조",
        },
        {
            "category": "2. 작업 범위",
            "item": "시각화/보고서",
            "detail": "통합 관제 대시보드 및 일/주/월간 보고서 자동화",
            "note": "",
        },
        {
            "category": "3. 산출물",
            "item": "기획/설계",
            "detail": "프로젝트 수행 계획서, 로그 연동 설계서",
            "note": "",
        },
        {
            "category": "3. 산출물",
            "item": "기술 문서",
            "detail": "탐지 시나리오 정의서, 운영/관리 매뉴얼",
            "note": "",
        },
        {
            "category": "3. 산출물",
            "item": "최종",
            "detail": "최종 완료 보고회 자료, 프로젝트 종료 보고서",
            "note": "산출물 시트 참조",
        },
        {
            "category": "4. R&R",
            "item": "수행사(Seekers)",
            "detail": "솔루션 설치, 파싱 규칙 생성, 탐지 룰 개발, 기술 교육",
            "note": "",
        },
        {
            "category": "4. R&R",
            "item": "발주사(고객)",
            "detail": "로그 수집 권한 부여, 네트워크 정책 오픈, 인프라 환경 제공",
            "note": "R&R 시트 참조",
        },
        {
            "category": "5. 특약 사항",
            "item": "유지보수",
            "detail": "무상 유지보수 기간 및 범위는 [별첨 5]에 의거함",
            "note": "유지보수 시트 참조",
        },
    ]

    # Schedule phases and tasks
    SCHEDULE = [
        {
            "phase": "분석/설계",
            "task": "현황 분석 및 요구사항 정의",
            "duration": 2,
            "owner": "PL",
            "note": "인터뷰 진행",
        },
        {
            "phase": "분석/설계",
            "task": "로그 연동 아키텍처 설계",
            "duration": 4,
            "owner": "PL",
            "note": "설계서 작성",
        },
        {
            "phase": "구축(Deployment)",
            "task": "H/W 랙 마운트 및 케이블링",
            "duration": 1,
            "owner": "엔지니어",
            "note": "IDC 방문",
        },
        {
            "phase": "구축(Deployment)",
            "task": "OS 설치 및 보안 설정(Hardening)",
            "duration": 1,
            "owner": "엔지니어",
            "note": "Rocky Linux",
        },
        {
            "phase": "구축(Deployment)",
            "task": "SIEM 엔진 설치 (Seekurity SIEM/OpenSearch)",
            "duration": 1,
            "owner": "엔지니어",
            "note": "Cluster 구성",
        },
        {
            "phase": "구축(Deployment)",
            "task": "수집기(Forwarder) 구성",
            "duration": 5,
            "owner": "엔지니어",
            "note": "LB/Failover",
        },
        {
            "phase": "연동(Integration)",
            "task": "네트워크 방화벽 정책 신청/확인",
            "duration": 1,
            "owner": "고객/엔지니어",
            "note": "통신 테스트",
        },
        {
            "phase": "연동(Integration)",
            "task": "서버 에이전트 배포 (Windows/Linux)",
            "duration": 2,
            "owner": "엔지니어",
            "note": "자동 배포 스크립트",
        },
        {
            "phase": "연동(Integration)",
            "task": "네트워크 장비 Syslog 연동 (FW/IPS)",
            "duration": 5,
            "owner": "엔지니어",
            "note": "표준 포트(514)",
        },
        {
            "phase": "연동(Integration)",
            "task": "로그 파싱 및 정규화(Normalization)",
            "duration": 4,
            "owner": "개발자",
            "note": "필드 매핑",
        },
        {
            "phase": "설정(Configuration)",
            "task": "대시보드 및 보고서 커스터마이징",
            "duration": 3,
            "owner": "개발자",
            "note": "요구사항 반영",
        },
        {
            "phase": "설정(Configuration)",
            "task": "탐지 룰 설정 및 오탐 튜닝",
            "duration": 9,
            "owner": "보안관제역",
            "note": "시나리오 적용",
        },
        {
            "phase": "안정화",
            "task": "통합 테스트 및 사용자 교육",
            "duration": 2,
            "owner": "PM",
            "note": "운영 이관",
        },
        {
            "phase": "종료",
            "task": "최종 보고 및 산출물 이관",
            "duration": 1,
            "owner": "PM",
            "note": "프로젝트 종료",
        },
    ]

    # Detection scenarios
    DETECTION_RULES = [
        {
            "no": 1,
            "category": "Access",
            "name": "Brute Force Attack (Successful)",
            "risk": "High",
            "log_source": "Server(Win/Linux)",
            "logic": "동일 계정으로 1분 내 5회 실패 후 1회 성공",
            "mitre": "T1110",
            "validation": "Hydra 또는 스크립트 이용 로그인 시도",
            "response": "계정 비활성화 및 접근 IP 차단",
        },
        {
            "no": 2,
            "category": "Malware",
            "name": "Malware Detected on Endpoint",
            "risk": "Critical",
            "log_source": "EDR/AV",
            "logic": "악성코드 탐지 이벤트 발생",
            "mitre": "T1204",
            "validation": "EICAR 테스트 파일 실행",
            "response": "파일 격리, 침해 장비 포렌식",
        },
        {
            "no": 3,
            "category": "Web Attack",
            "name": "SQL Injection Attempt",
            "risk": "High",
            "log_source": "WAF/WebServer",
            "logic": "SQL 인젝션 패턴 탐지 (UNION, SELECT 등)",
            "mitre": "T1190",
            "validation": "SQLMap 도구 이용 테스트",
            "response": "공격 IP 차단 및 웹 취약점 점검",
        },
        {
            "no": 4,
            "category": "Network",
            "name": "Lateral Movement Detection",
            "risk": "High",
            "log_source": "Firewall/IDS",
            "logic": "내부 → 내부 비정상 포트 스캔 (10+ ports)",
            "mitre": "T1021",
            "validation": "Nmap 내부 스캔 시뮬레이션",
            "response": "이상 탐지 장비 네트워크 격리",
        },
        {
            "no": 5,
            "category": "Execution",
            "name": "Suspicious Command Execution",
            "risk": "Medium",
            "log_source": "Server(Win/Linux)",
            "logic": "PowerShell -enc 또는 bash -c 등 인코딩 명령 실행",
            "mitre": "T1059",
            "validation": "인코딩된 명령어 실행 테스트",
            "response": "명령어 실행 이력 조사",
        },
        {
            "no": 6,
            "category": "Exfiltration",
            "name": "Large Data Transfer",
            "risk": "High",
            "log_source": "Firewall/Proxy",
            "logic": "단일 세션 100MB 이상 외부 전송",
            "mitre": "T1048",
            "validation": "대용량 파일 외부 업로드 테스트",
            "response": "통신 세션 강제 종료 및 데이터 확인",
        },
        {
            "no": 7,
            "category": "C2",
            "name": "C2 Beaconing Activity",
            "risk": "Critical",
            "log_source": "Firewall/DNS",
            "logic": "동일 외부 IP/도메인 주기적 연결 (5분 간격 10회+)",
            "mitre": "T1071",
            "validation": "Cobalt Strike 시뮬레이션",
            "response": "장비 네트워크 즉시 차단",
        },
        {
            "no": 8,
            "category": "Access",
            "name": "Off-hours Login",
            "risk": "Medium",
            "log_source": "Server/AD",
            "logic": "업무 외 시간(22시~06시) 로그인 성공",
            "mitre": "T1078",
            "validation": "야간 시간대 로그인 테스트",
            "response": "로그인 사용자 확인 및 소명 요청",
        },
        {
            "no": 9,
            "category": "Web Attack",
            "name": "Web Shell Upload Attempt",
            "risk": "Critical",
            "log_source": "WAF/WebServer",
            "logic": "PHP/JSP/ASP 확장자 파일 업로드 시도",
            "mitre": "T1505.003",
            "validation": "웹쉘 파일 업로드 테스트",
            "response": "웹 서비스 필터링 설정 강화",
        },
        {
            "no": 10,
            "category": "DoS",
            "name": "DDoS Attack Indication",
            "risk": "High",
            "log_source": "Firewall/IPS",
            "logic": "동일 소스에서 1분간 1000+ 연결 시도",
            "mitre": "T1498",
            "validation": "트래픽 생성기로 부하 테스트",
            "response": "DDoS 방어 장비 임계치 확인",
        },
    ]

    # Deliverables
    DELIVERABLES = [
        {
            "phase": "착수",
            "name": "프로젝트 수행 계획서",
            "format": "PPT/PDF",
            "due": "착수 후 1주 이내",
            "note": "",
        },
        {
            "phase": "분석",
            "name": "요구사항 정의서",
            "format": "XLSX/HWP",
            "due": "분석 단계 완료 시",
            "note": "",
        },
        {
            "phase": "설계",
            "name": "로그 연동 설계서",
            "format": "XLSX",
            "due": "설계 단계 완료 시",
            "note": "포트 및 파싱 내역",
        },
        {
            "phase": "구축",
            "name": "설치 완료 보고서",
            "format": "DOCX",
            "due": "설치 완료 시",
            "note": "스크린샷 포함",
        },
        {
            "phase": "구축",
            "name": "탐지 시나리오 정의서",
            "format": "XLSX",
            "due": "튜닝 완료 시",
            "note": "임계치 설정 근거",
        },
        {
            "phase": "완료",
            "name": "운영자 매뉴얼",
            "format": "PDF",
            "due": "프로젝트 종료 시",
            "note": "",
        },
        {
            "phase": "완료",
            "name": "완료 보고서",
            "format": "PPT/PDF",
            "due": "프로젝트 종료 시",
            "note": "",
        },
    ]

    # R&R (Roles and Responsibilities)
    RNR_TEMPLATE = [
        {
            "org": "브로드밴드시큐리티",
            "name": "",
            "title": "PM",
            "role": "프로젝트 총괄/진척 관리",
            "email": "",
            "phone": "",
        },
        {
            "org": "씨커스",
            "name": "",
            "title": "PL",
            "role": "아키텍처 설계/기술 리딩",
            "email": "",
            "phone": "",
        },
        {
            "org": "씨커스",
            "name": "",
            "title": "엔지니어",
            "role": "솔루션 구축/로그 연동",
            "email": "",
            "phone": "",
        },
        {
            "org": "고객사 (발주)",
            "name": "",
            "title": "팀장",
            "role": "프로젝트 의사결정",
            "email": "",
            "phone": "",
        },
        {
            "org": "고객사 (발주)",
            "name": "",
            "title": "담당자",
            "role": "현업 요구사항 조율/지원",
            "email": "",
            "phone": "",
        },
    ]

    # Maintenance terms
    MAINTENANCE = [
        {
            "category": "기본 사항",
            "item": "대상 품목",
            "detail": "SIEM Appliance 및 수집 Agent S/W",
            "sla": "-",
        },
        {
            "category": "기본 사항",
            "item": "기간",
            "detail": "검수 완료일로부터 12개월",
            "sla": "365일 지원",
        },
        {
            "category": "기술 지원",
            "item": "장애 지원",
            "detail": "서비스 중단 시 긴급 복구 지원",
            "sla": "접수 후 12시간 내 도착, 24시간 내 복구",
        },
        {
            "category": "기술 지원",
            "item": "정기 점검",
            "detail": "시스템 상태 점검 및 리포트 제공",
            "sla": "월 1회 방문 점검",
        },
        {
            "category": "기술 지원",
            "item": "패치 지원",
            "detail": "보안 패치 및 마이너 업그레이드",
            "sla": "분기 1회 또는 긴급 시",
        },
        {
            "category": "운영 지원",
            "item": "룰 튜닝",
            "detail": "신규 패턴 업데이트 및 오탐 예외 처리",
            "sla": "요청 시 (연 3건 이내)",
        },
    ]

    # Status options
    STATUS_OPTIONS = ["대기", "진행중", "완료", "지연", "보류"]
    RISK_OPTIONS = ["Critical", "High", "Medium", "Low"]


# ============================================================
# STYLES
# ============================================================


class Styles:
    """Pre-defined styles for SOW"""

    def __init__(self):
        c = SOWConfig.COLORS

        # Fills
        self.title_fill = PatternFill(
            start_color=c["title"], end_color=c["title"], fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color=c["header"], end_color=c["header"], fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color=c["subheader"], end_color=c["subheader"], fill_type="solid"
        )
        self.section_fill = PatternFill(
            start_color=c["section"], end_color=c["section"], fill_type="solid"
        )
        self.row_even = PatternFill(
            start_color=c["row_even"], end_color=c["row_even"], fill_type="solid"
        )
        self.row_odd = PatternFill(
            start_color=c["row_odd"], end_color=c["row_odd"], fill_type="solid"
        )

        # Phase fills
        self.phase_fills = {
            "분석/설계": PatternFill(
                start_color=c["phase_analysis"],
                end_color=c["phase_analysis"],
                fill_type="solid",
            ),
            "구축(Deployment)": PatternFill(
                start_color=c["phase_deploy"],
                end_color=c["phase_deploy"],
                fill_type="solid",
            ),
            "연동(Integration)": PatternFill(
                start_color=c["phase_integrate"],
                end_color=c["phase_integrate"],
                fill_type="solid",
            ),
            "설정(Configuration)": PatternFill(
                start_color=c["phase_config"],
                end_color=c["phase_config"],
                fill_type="solid",
            ),
            "안정화": PatternFill(
                start_color=c["phase_stabilize"],
                end_color=c["phase_stabilize"],
                fill_type="solid",
            ),
            "종료": PatternFill(
                start_color=c["phase_close"],
                end_color=c["phase_close"],
                fill_type="solid",
            ),
        }

        # Status fills
        self.status_fills = {
            "완료": PatternFill(
                start_color=c["complete"], end_color=c["complete"], fill_type="solid"
            ),
            "진행중": PatternFill(
                start_color=c["in_progress"],
                end_color=c["in_progress"],
                fill_type="solid",
            ),
            "지연": PatternFill(
                start_color=c["delayed"], end_color=c["delayed"], fill_type="solid"
            ),
            "대기": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
            "보류": PatternFill(
                start_color=c["pending"], end_color=c["pending"], fill_type="solid"
            ),
        }

        # Risk fills
        self.risk_fills = {
            "Critical": PatternFill(
                start_color=c["high"], end_color=c["high"], fill_type="solid"
            ),
            "High": PatternFill(
                start_color=c["high"], end_color=c["high"], fill_type="solid"
            ),
            "Medium": PatternFill(
                start_color=c["medium"], end_color=c["medium"], fill_type="solid"
            ),
            "Low": PatternFill(
                start_color=c["low"], end_color=c["low"], fill_type="solid"
            ),
        }

        # Fonts
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.subheader_font = Font(bold=True, size=10, color="1B2838")
        self.data_font = Font(size=9)
        self.section_font = Font(bold=True, size=10)

        # Borders
        self.thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        self.header_border = Border(
            left=Side(style="thin", color="8EA9DB"),
            right=Side(style="thin", color="8EA9DB"),
            top=Side(style="medium", color="2E5090"),
            bottom=Side(style="medium", color="2E5090"),
        )

        # Alignments
        self.center = Alignment(horizontal="center", vertical="center")
        self.center_wrap = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.left = Alignment(vertical="center")
        self.left_wrap = Alignment(vertical="center", wrap_text=True)


# ============================================================
# SOW BUILDER
# ============================================================


class SOWBuilder:
    """Build SOW Excel workbook"""

    def __init__(
        self,
        project_name: str = None,
        customer: str = None,
        start_date: datetime = None,
    ):
        self.styles = Styles()
        self.wb = Workbook()
        self.project_name = project_name or SOWConfig.PROJECT_NAME
        self.customer = customer or SOWConfig.CUSTOMER
        self.start_date = start_date or datetime.now().replace(
            hour=0, minute=0, second=0, microsecond=0
        )

    def _create_title_row(self, ws, title: str, merge_range: str):
        """Create title row"""
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(":")[0]]
        cell.value = title
        cell.fill = self.styles.title_fill
        cell.font = self.styles.title_font
        cell.alignment = self.styles.center

    def _apply_header_style(self, ws, row: int, columns: List[tuple]):
        """Apply header style to row"""
        for col, (header, width) in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.styles.subheader_fill
            cell.font = self.styles.subheader_font
            cell.alignment = self.styles.center_wrap
            cell.border = self.styles.header_border
            ws.column_dimensions[get_column_letter(col)].width = width

    def create_cover_sheet(self):
        """Create cover page sheet"""
        ws = self.wb.active
        ws.title = "0.Cover"

        self._create_title_row(ws, "0.Cover Page", "A1:B1")

        # Info rows
        info = [
            ("구분", "내용"),
            ("Project Name", self.project_name),
            ("Document Type", SOWConfig.DOCUMENT_TYPE),
            ("Version", SOWConfig.VERSION),
            ("Date", datetime.now().strftime("%Y-%m-%d")),
            ("Company (Customer)", self.customer),
            ("Company (Partner)", SOWConfig.PARTNER),
            ("Status", "-"),
        ]

        for row_idx, (label, value) in enumerate(info, start=3):
            ws.cell(row=row_idx, column=1, value=label).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).fill = self.styles.section_fill
            ws.cell(row=row_idx, column=1).font = self.styles.section_font

            ws.cell(row=row_idx, column=2, value=value).border = self.styles.thin_border

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 35
        ws.row_dimensions[1].height = 30

        return self

    def create_summary_sheet(self):
        """Create SOW summary sheet"""
        ws = self.wb.create_sheet(title="SOW(요약)")

        self._create_title_row(ws, "SOW(요약)", "A1:D1")

        # Headers
        headers = [
            ("구분", 15),
            ("항목", 15),
            ("세부 내용(Scope & Details)", 55),
            ("비고", 18),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data
        prev_category = None
        for row_idx, item in enumerate(SOWConfig.SOW_SUMMARY, start=4):
            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            # Category (merge if same)
            cell = ws.cell(
                row=row_idx,
                column=1,
                value=item["category"] if item["category"] != prev_category else "",
            )
            cell.fill = (
                self.styles.section_fill
                if item["category"] != prev_category
                else base_fill
            )
            cell.font = (
                self.styles.section_font
                if item["category"] != prev_category
                else self.styles.data_font
            )
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.left_wrap
            prev_category = item["category"]

            # Item
            ws.cell(row=row_idx, column=2, value=item["item"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border
            ws.cell(row=row_idx, column=2).font = self.styles.data_font

            # Detail
            ws.cell(row=row_idx, column=3, value=item["detail"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).font = self.styles.data_font
            ws.cell(row=row_idx, column=3).alignment = self.styles.left_wrap

            # Note
            ws.cell(row=row_idx, column=4, value=item["note"]).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border
            ws.cell(row=row_idx, column=4).font = self.styles.data_font

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def create_schedule_sheet(self):
        """Create schedule sheet"""
        ws = self.wb.create_sheet(title="1.일정표(Schedule)")

        self._create_title_row(ws, "1.일정표(Schedule)", "A1:F1")

        # Headers
        headers = [
            ("단계", 18),
            ("Task", 40),
            ("시작일", 14),
            ("종료일", 14),
            ("담당자", 12),
            ("비고", 18),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data
        current_date = self.start_date
        prev_phase = None
        phase_start_row = 4

        for row_idx, task in enumerate(SOWConfig.SCHEDULE, start=4):
            end_date = current_date + timedelta(days=task["duration"] - 1)

            # Phase
            phase = task["phase"]
            cell = ws.cell(
                row=row_idx, column=1, value=phase if phase != prev_phase else ""
            )
            if phase != prev_phase:
                cell.fill = self.styles.phase_fills.get(phase, self.styles.section_fill)
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                if prev_phase and row_idx > 4:
                    # Merge previous phase cells
                    ws.merge_cells(f"A{phase_start_row}:A{row_idx - 1}")
                phase_start_row = row_idx
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.center_wrap

            prev_phase = phase

            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            # Task
            ws.cell(row=row_idx, column=2, value=task["task"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border
            ws.cell(row=row_idx, column=2).font = self.styles.data_font

            # Start date
            ws.cell(
                row=row_idx, column=3, value=current_date.strftime("%Y-%m-%d")
            ).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center
            ws.cell(row=row_idx, column=3).font = self.styles.data_font

            # End date
            ws.cell(
                row=row_idx, column=4, value=end_date.strftime("%Y-%m-%d")
            ).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border
            ws.cell(row=row_idx, column=4).alignment = self.styles.center
            ws.cell(row=row_idx, column=4).font = self.styles.data_font

            # Owner
            ws.cell(row=row_idx, column=5, value=task["owner"]).fill = base_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border
            ws.cell(row=row_idx, column=5).alignment = self.styles.center
            ws.cell(row=row_idx, column=5).font = self.styles.data_font

            # Note
            ws.cell(row=row_idx, column=6, value=task["note"]).fill = base_fill
            ws.cell(row=row_idx, column=6).border = self.styles.thin_border
            ws.cell(row=row_idx, column=6).font = self.styles.data_font

            current_date = end_date + timedelta(days=1)

        # Merge last phase
        ws.merge_cells(f"A{phase_start_row}:A{row_idx}")

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25
        ws.freeze_panes = "B4"

        return self

    def create_asset_sheet(self, assets: List[Dict] = None):
        """Create asset list sheet"""
        ws = self.wb.create_sheet(title="2.자산목록(Asset)")

        self._create_title_row(ws, "2.자산목록(Asset)", "A1:H1")

        # Headers
        headers = [
            ("연번", 6),
            ("호스트명", 15),
            ("IP 주소", 14),
            ("자산 유형", 35),
            ("Log Source", 20),
            ("OS/Model", 15),
            ("로그 수집 방식", 12),
            ("비고", 15),
        ]
        self._apply_header_style(ws, 3, headers)

        # Default placeholder data
        if not assets:
            for row_idx in range(4, 14):
                is_odd = (row_idx - 4) % 2 == 1
                base_fill = self.styles.row_odd if is_odd else self.styles.row_even

                ws.cell(row=row_idx, column=1, value=row_idx - 3).fill = base_fill
                ws.cell(row=row_idx, column=1).border = self.styles.thin_border
                ws.cell(row=row_idx, column=1).alignment = self.styles.center

                for col in range(2, 9):
                    ws.cell(
                        row=row_idx,
                        column=col,
                        value="조사 후 기재" if col == 2 and row_idx == 4 else "",
                    ).fill = base_fill
                    ws.cell(row=row_idx, column=col).border = self.styles.thin_border

        # Data validation for 로그 수집 방식
        log_method_dv = DataValidation(
            type="list", formula1='"Syslog,Agent,API,File"', allow_blank=True
        )
        ws.add_data_validation(log_method_dv)
        log_method_dv.add("G4:G100")

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25
        ws.freeze_panes = "C4"

        return self

    def create_detection_rules_sheet(self):
        """Create detection rules sheet"""
        ws = self.wb.create_sheet(title="2.탐지시나리오(Rule)")

        self._create_title_row(ws, "2.탐지시나리오(Rule)", "A1:I1")

        # Headers
        headers = [
            ("No", 5),
            ("분류(Category)", 12),
            ("탐지 시나리오명", 28),
            ("위험도", 8),
            ("로그 소스", 16),
            ("탐지 로직(Logic/Threshold)", 32),
            ("MITRE ATT&CK", 12),
            ("검증 방법(Validation)", 25),
            ("대응 방안", 25),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data
        for row_idx, rule in enumerate(SOWConfig.DETECTION_RULES, start=4):
            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            # No
            ws.cell(row=row_idx, column=1, value=rule["no"]).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

            # Category
            ws.cell(row=row_idx, column=2, value=rule["category"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border
            ws.cell(row=row_idx, column=2).alignment = self.styles.center

            # Name
            ws.cell(row=row_idx, column=3, value=rule["name"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.left_wrap

            # Risk
            ws.cell(
                row=row_idx, column=4, value=rule["risk"]
            ).fill = self.styles.risk_fills.get(rule["risk"], base_fill)
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border
            ws.cell(row=row_idx, column=4).alignment = self.styles.center
            ws.cell(row=row_idx, column=4).font = Font(bold=True, size=9)

            # Log source
            ws.cell(row=row_idx, column=5, value=rule["log_source"]).fill = base_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border
            ws.cell(row=row_idx, column=5).alignment = self.styles.center

            # Logic
            ws.cell(row=row_idx, column=6, value=rule["logic"]).fill = base_fill
            ws.cell(row=row_idx, column=6).border = self.styles.thin_border
            ws.cell(row=row_idx, column=6).alignment = self.styles.left_wrap

            # MITRE
            ws.cell(row=row_idx, column=7, value=rule["mitre"]).fill = base_fill
            ws.cell(row=row_idx, column=7).border = self.styles.thin_border
            ws.cell(row=row_idx, column=7).alignment = self.styles.center

            # Validation
            ws.cell(row=row_idx, column=8, value=rule["validation"]).fill = base_fill
            ws.cell(row=row_idx, column=8).border = self.styles.thin_border
            ws.cell(row=row_idx, column=8).alignment = self.styles.left_wrap

            # Response
            ws.cell(row=row_idx, column=9, value=rule["response"]).fill = base_fill
            ws.cell(row=row_idx, column=9).border = self.styles.thin_border
            ws.cell(row=row_idx, column=9).alignment = self.styles.left_wrap

        # Risk dropdown
        risk_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(SOWConfig.RISK_OPTIONS)}"',
            allow_blank=False,
        )
        ws.add_data_validation(risk_dv)
        risk_dv.add("D4:D50")

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25
        ws.freeze_panes = "D4"

        return self

    def create_deliverables_sheet(self):
        """Create deliverables sheet"""
        ws = self.wb.create_sheet(title="3.산출물(Docs)")

        self._create_title_row(ws, "3.산출물(Docs)", "A1:E1")

        # Headers
        headers = [
            ("단계", 10),
            ("산출물명", 22),
            ("제출 형식", 12),
            ("제출 시기", 18),
            ("비고", 18),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data
        for row_idx, doc in enumerate(SOWConfig.DELIVERABLES, start=4):
            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row_idx, column=1, value=doc["phase"]).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border
            ws.cell(row=row_idx, column=1).alignment = self.styles.center

            ws.cell(row=row_idx, column=2, value=doc["name"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border

            ws.cell(row=row_idx, column=3, value=doc["format"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center

            ws.cell(row=row_idx, column=4, value=doc["due"]).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border

            ws.cell(row=row_idx, column=5, value=doc["note"]).fill = base_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def create_rnr_sheet(self):
        """Create R&R (Roles and Responsibilities) sheet"""
        ws = self.wb.create_sheet(title="4.인력(R&R)")

        self._create_title_row(ws, "4.인력(R&R)", "A1:F1")

        # Headers
        headers = [
            ("소속", 15),
            ("성명", 10),
            ("직급", 10),
            ("역할", 22),
            ("이메일", 25),
            ("연락처", 15),
        ]
        self._apply_header_style(ws, 3, headers)

        # Data
        for row_idx, person in enumerate(SOWConfig.RNR_TEMPLATE, start=4):
            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            ws.cell(row=row_idx, column=1, value=person["org"]).fill = base_fill
            ws.cell(row=row_idx, column=1).border = self.styles.thin_border

            ws.cell(row=row_idx, column=2, value=person["name"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border

            ws.cell(row=row_idx, column=3, value=person["title"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.center

            ws.cell(row=row_idx, column=4, value=person["role"]).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border

            ws.cell(row=row_idx, column=5, value=person["email"]).fill = base_fill
            ws.cell(row=row_idx, column=5).border = self.styles.thin_border

            ws.cell(row=row_idx, column=6, value=person["phone"]).fill = base_fill
            ws.cell(row=row_idx, column=6).border = self.styles.thin_border

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def create_maintenance_sheet(self):
        """Create maintenance terms sheet"""
        ws = self.wb.create_sheet(title="별첨5.유지보수")

        self._create_title_row(ws, "별첨5.유지보수", "A1:D1")

        # Headers
        headers = [("구분", 12), ("항목", 12), ("내용", 40), ("SLA 수준 / 주기", 28)]
        self._apply_header_style(ws, 3, headers)

        # Data
        prev_category = None
        for row_idx, item in enumerate(SOWConfig.MAINTENANCE, start=4):
            is_odd = (row_idx - 4) % 2 == 1
            base_fill = self.styles.row_odd if is_odd else self.styles.row_even

            # Category
            cell = ws.cell(
                row=row_idx,
                column=1,
                value=item["category"] if item["category"] != prev_category else "",
            )
            cell.fill = (
                self.styles.section_fill
                if item["category"] != prev_category
                else base_fill
            )
            cell.font = (
                self.styles.section_font
                if item["category"] != prev_category
                else self.styles.data_font
            )
            cell.border = self.styles.thin_border
            prev_category = item["category"]

            ws.cell(row=row_idx, column=2, value=item["item"]).fill = base_fill
            ws.cell(row=row_idx, column=2).border = self.styles.thin_border

            ws.cell(row=row_idx, column=3, value=item["detail"]).fill = base_fill
            ws.cell(row=row_idx, column=3).border = self.styles.thin_border
            ws.cell(row=row_idx, column=3).alignment = self.styles.left_wrap

            ws.cell(row=row_idx, column=4, value=item["sla"]).fill = base_fill
            ws.cell(row=row_idx, column=4).border = self.styles.thin_border

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 25

        return self

    def build_all(self):
        """Build all sheets"""
        self.create_cover_sheet()
        self.create_summary_sheet()
        self.create_schedule_sheet()
        self.create_asset_sheet()
        self.create_detection_rules_sheet()
        self.create_deliverables_sheet()
        self.create_rnr_sheet()
        self.create_maintenance_sheet()
        return self

    def save(self, output_file: str = None):
        """Save workbook"""
        output = output_file or SOWConfig.OUTPUT_FILE
        self.wb.save(output)
        print(f"Saved: {output}")
        return self


# ============================================================
# SCHEDULE UPDATER
# ============================================================


class ScheduleUpdater:
    """Update deployment schedule from SOW"""

    def __init__(self, sow_file: str, schedule_file: str):
        self.sow_file = sow_file
        self.schedule_file = schedule_file

    def sync_schedule(self):
        """Sync schedule from SOW to deployment schedule"""
        # Load SOW
        sow_wb = load_workbook(self.sow_file, data_only=True)
        sow_schedule = sow_wb["1.일정표(Schedule)"]

        # Read SOW schedule data
        tasks = []
        for row_idx in range(4, sow_schedule.max_row + 1):
            phase = sow_schedule.cell(row=row_idx, column=1).value
            task = sow_schedule.cell(row=row_idx, column=2).value
            start = sow_schedule.cell(row=row_idx, column=3).value
            end = sow_schedule.cell(row=row_idx, column=4).value
            owner = sow_schedule.cell(row=row_idx, column=5).value
            note = sow_schedule.cell(row=row_idx, column=6).value

            if task:
                tasks.append(
                    {
                        "phase": phase,
                        "task": task,
                        "start": start,
                        "end": end,
                        "owner": owner,
                        "note": note,
                    }
                )

        print(f"Read {len(tasks)} tasks from SOW")

        # Update deployment schedule
        # (Implementation depends on the deployment schedule format)

        return tasks


# ============================================================
# MAIN
# ============================================================


def create_sow(project_name: str = None, customer: str = None, output_file: str = None):
    """Create new SOW Excel"""
    print("Creating Seekurity SIEM SOW...")
    print()

    builder = SOWBuilder(project_name=project_name, customer=customer)
    builder.build_all()
    builder.save(output_file)

    print()
    print("=" * 50)
    print("SOW Creation Complete")
    print("=" * 50)
    print(f"Output: {output_file or SOWConfig.OUTPUT_FILE}")
    print()
    print("Sheets created:")
    print("  - 0.Cover (표지)")
    print("  - SOW(요약)")
    print("  - 1.일정표(Schedule)")
    print("  - 2.자산목록(Asset)")
    print("  - 2.탐지시나리오(Rule)")
    print("  - 3.산출물(Docs)")
    print("  - 4.인력(R&R)")
    print("  - 별첨5.유지보수")


def main():
    parser = argparse.ArgumentParser(description="Seekurity SIEM SOW Manager")
    subparsers = parser.add_subparsers(dest="command", help="Commands")

    # Create command
    create_parser = subparsers.add_parser("create", help="Create new SOW")
    create_parser.add_argument("--project", "-p", help="Project name")
    create_parser.add_argument("--customer", "-c", help="Customer name")
    create_parser.add_argument("--output", "-o", help="Output file")

    # Update schedule command
    update_parser = subparsers.add_parser(
        "update-schedule", help="Update schedule from SOW"
    )
    update_parser.add_argument(
        "--sow", "-s", default=SOWConfig.OUTPUT_FILE, help="SOW file"
    )
    update_parser.add_argument(
        "--schedule",
        default="SeekuritySIEM_Deployment_Schedule.xlsx",
        help="Schedule file",
    )

    args = parser.parse_args()

    if args.command == "create":
        create_sow(args.project, args.customer, args.output)
    elif args.command == "update-schedule":
        updater = ScheduleUpdater(args.sow, args.schedule)
        updater.sync_schedule()
    else:
        # Default: create
        create_sow()


if __name__ == "__main__":
    main()
