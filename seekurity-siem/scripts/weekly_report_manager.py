#!/usr/bin/env python3
"""주간보고 Excel 생성기 (프로젝트 공용).

프로젝트별 주차 데이터(JSON)를 읽어 주간보고 Excel 을 만든다.
데이터와 서식을 분리했으므로 어느 프로젝트에서든 그대로 쓸 수 있다.

    python scripts/weekly_report_manager.py init aig          # 이번 주 JSON 뼈대 생성
    python scripts/weekly_report_manager.py generate aig      # JSON -> Excel
    python scripts/weekly_report_manager.py generate aig --week 2026-W34

입력  projects/<project>/01.사업관리/주간보고/데이터/<YYYY-Www>.json
출력  projects/<project>/01.사업관리/주간보고/<PROJECT>_주간보고_<YYYYMMDD>.xlsx
"""
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
PROJECTS = ROOT / "projects"

# 한국 SI 표준 산출물 구조 — 주간보고는 사업관리 단계에 둔다
REPORT_DIR = "01.사업관리/주간보고"
DATA_DIR = REPORT_DIR + "/데이터"

COLORS = {
    "title": "1F3864",
    "section": "2E5C8A",
    "header": "D6E4F0",
    "border": "B4C6E7",
    "muted": "808080",
}

STATUS_FILL = {
    "완료": "C6EFCE",
    "진행중": "FFEB9C",
    "지연": "FFC7CE",
    "대기": "EDEDED",
    "해결": "C6EFCE",
    "조치중": "FFEB9C",
    "미해결": "FFC7CE",
}

PRIORITY_FILL = {
    "높음": "FFC7CE",
    "중간": "FFEB9C",
    "낮음": "EDEDED",
}


# ------------------------------------------------------------------ 데이터


def iso_week(date):
    y, w, _ = date.isocalendar()
    return "%d-W%02d" % (y, w)


def week_range(week):
    """'2026-W34' -> (월요일, 일요일)"""
    year, wk = week.split("-W")
    monday = dt.date.fromisocalendar(int(year), int(wk), 1)
    return monday, monday + dt.timedelta(days=6)


def data_path(project, week):
    return PROJECTS / project / DATA_DIR / (week + ".json")


def skeleton(project, week):
    start, end = week_range(week)
    return {
        "project": project.upper(),
        "title": project.upper() + " Seekurity SIEM 구축",
        "week": week,
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "author": "",
        "customer_manager": "",
        "overall_progress": 0,
        "summary": "",
        "this_week": [
            {
                "category": "",
                "task": "",
                "owner": "",
                "planned": "",
                "done": "",
                "progress": 0,
                "status": "완료",
                "note": "",
            }
        ],
        "next_week": [
            {
                "category": "",
                "task": "",
                "owner": "",
                "start": "",
                "end": "",
                "priority": "중간",
                "note": "",
            }
        ],
        "issues": [
            {
                "category": "",
                "issue": "",
                "impact": "",
                "action": "",
                "owner": "",
                "due": "",
                "status": "미해결",
            }
        ],
    }


# ------------------------------------------------------------------ 서식


class Sheet:
    """주간보고 시트 조립기. 섹션을 순서대로 붙여 나간다."""

    def __init__(self, ws, width):
        self.ws = ws
        self.width = width
        self.row = 1
        side = Side(style="thin", color=COLORS["border"])
        self.thin = Border(left=side, right=side, top=side, bottom=side)

    def title(self, text, subtitle=""):
        ws = self.ws
        ws.merge_cells(
            start_row=self.row, start_column=1, end_row=self.row, end_column=self.width
        )
        c = ws.cell(self.row, 1, text)
        c.font = Font(bold=True, size=16, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COLORS["title"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[self.row].height = 30
        self.row += 1
        if subtitle:
            ws.merge_cells(
                start_row=self.row,
                start_column=1,
                end_row=self.row,
                end_column=self.width,
            )
            c = ws.cell(self.row, 1, subtitle)
            c.font = Font(size=10, color=COLORS["muted"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            self.row += 1
        self.row += 1

    def meta(self, pairs):
        """2열씩 묶어 표기하는 머리말 정보.

        항목명은 좁은 첫 열 하나에 넣으면 잘리므로 두 열에 걸쳐 표기한다.
        """
        ws = self.ws
        half = self.width // 2
        key_span = 2
        for i in range(0, len(pairs), 2):
            group = pairs[i : i + 2]
            for j, (k, v) in enumerate(group):
                col = 1 + j * half
                # 마지막 묶음의 값은 시트 오른쪽 끝까지 채운다
                last = j == len(group) - 1
                val_end = self.width if last else col + half - 1

                ws.merge_cells(
                    start_row=self.row,
                    start_column=col,
                    end_row=self.row,
                    end_column=col + key_span - 1,
                )
                kc = ws.cell(self.row, col, k)
                kc.font = Font(bold=True, size=9)
                kc.alignment = Alignment(horizontal="center", vertical="center")
                for c in range(col, col + key_span):
                    cc = ws.cell(self.row, c)
                    cc.fill = PatternFill("solid", start_color=COLORS["header"])
                    cc.border = self.thin

                ws.merge_cells(
                    start_row=self.row,
                    start_column=col + key_span,
                    end_row=self.row,
                    end_column=val_end,
                )
                vc = ws.cell(self.row, col + key_span, v)
                vc.font = Font(size=9)
                vc.alignment = Alignment(vertical="center", indent=1)
                for c in range(col + key_span, val_end + 1):
                    ws.cell(self.row, c).border = self.thin
            self.row += 1
        self.row += 1

    def note(self, text):
        ws = self.ws
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row + 2,
            end_column=self.width,
        )
        c = ws.cell(self.row, 1, text)
        c.font = Font(size=9)
        c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        c.border = self.thin
        self.row += 4

    def section(self, text):
        ws = self.ws
        ws.merge_cells(
            start_row=self.row, start_column=1, end_row=self.row, end_column=self.width
        )
        c = ws.cell(self.row, 1, text)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COLORS["section"])
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[self.row].height = 22
        self.row += 1

    def table(self, headers, rows, fill_cols=None):
        """fill_cols: {열번호: 값->색 매핑} 형태로 셀 배경 지정."""
        ws = self.ws
        fill_cols = fill_cols or {}
        for i, h in enumerate(headers, start=1):
            c = ws.cell(self.row, i, h)
            c.font = Font(bold=True, size=9)
            c.fill = PatternFill("solid", start_color=COLORS["header"])
            c.border = self.thin
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[self.row].height = 20
        self.row += 1

        if not rows:
            ws.merge_cells(
                start_row=self.row,
                start_column=1,
                end_row=self.row,
                end_column=self.width,
            )
            c = ws.cell(self.row, 1, "해당 없음")
            c.font = Font(size=9, color=COLORS["muted"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self.thin
            self.row += 1

        for r in rows:
            for i, v in enumerate(r, start=1):
                c = ws.cell(self.row, i, v)
                c.font = Font(size=9)
                c.border = self.thin
                wrap = i in (3, len(headers))
                c.alignment = Alignment(
                    vertical="center" if not wrap else "top", wrap_text=wrap
                )
                if i in fill_cols and v in fill_cols[i]:
                    c.fill = PatternFill("solid", start_color=fill_cols[i][v])
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if i == 1:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            self.row += 1
        self.row += 1


# ------------------------------------------------------------------ 생성


def build(data, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "주간보고"
    ws.sheet_view.showGridLines = False

    widths = [5, 14, 50, 10, 11, 11, 9, 9, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    s = Sheet(ws, width=len(widths))
    start = data["period"]["start"]
    end = data["period"]["end"]
    s.title(
        data["title"] + " 주간보고",
        "%s  (%s ~ %s)" % (data["week"], start, end),
    )

    s.meta(
        [
            ("프로젝트", data["title"]),
            ("주차", "%s  (%s ~ %s)" % (data["week"], start, end)),
            ("작성자", data.get("author", "")),
            ("고객 담당", data.get("customer_manager", "")),
            ("전체 진척률", "%s%%" % data.get("overall_progress", 0)),
            ("작성일", dt.date.today().isoformat()),
        ]
    )

    if data.get("summary"):
        s.section("금주 요약")
        s.note(data["summary"])

    s.section("1. 금주 수행 실적")
    s.table(
        ["No", "구분", "수행 내용", "담당", "계획일", "완료일", "진척", "상태", "비고"],
        [
            [
                i,
                t.get("category", ""),
                t.get("task", ""),
                t.get("owner", ""),
                t.get("planned", ""),
                t.get("done", ""),
                "%s%%" % t.get("progress", 0),
                t.get("status", ""),
                t.get("note", ""),
            ]
            for i, t in enumerate(data.get("this_week", []), 1)
        ],
        fill_cols={8: STATUS_FILL},
    )

    s.section("2. 차주 수행 계획")
    s.table(
        ["No", "구분", "수행 계획", "담당", "시작 예정", "완료 예정", "우선순위", "", "비고"],
        [
            [
                i,
                t.get("category", ""),
                t.get("task", ""),
                t.get("owner", ""),
                t.get("start", ""),
                t.get("end", ""),
                t.get("priority", ""),
                "",
                t.get("note", ""),
            ]
            for i, t in enumerate(data.get("next_week", []), 1)
        ],
        fill_cols={7: PRIORITY_FILL},
    )

    s.section("3. 이슈 및 요청사항")
    s.table(
        ["No", "구분", "이슈 내용", "영향", "조치 방안", "담당", "목표일", "상태", "비고"],
        [
            [
                i,
                t.get("category", ""),
                t.get("issue", ""),
                t.get("impact", ""),
                t.get("action", ""),
                t.get("owner", ""),
                t.get("due", ""),
                t.get("status", ""),
                t.get("note", ""),
            ]
            for i, t in enumerate(data.get("issues", []), 1)
        ],
        fill_cols={8: STATUS_FILL},
    )

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ------------------------------------------------------------------ CLI


def cmd_init(args):
    week = args.week or iso_week(dt.date.today())
    path = data_path(args.project, week)
    if path.exists() and not args.force:
        print("이미 존재: %s  (--force 로 덮어쓰기)" % path)
        return 1
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(skeleton(args.project, week), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print("생성: %s" % path)
    return 0


def cmd_generate(args):
    week = args.week or iso_week(dt.date.today())
    path = data_path(args.project, week)
    if not path.exists():
        print("데이터 없음: %s" % path)
        print("먼저 init 을 실행하세요: python scripts/weekly_report_manager.py init %s"
              % args.project)
        return 1
    data = json.loads(path.read_text(encoding="utf-8"))
    stamp = data["period"]["end"].replace("-", "")
    out = (
        PROJECTS
        / args.project
        / REPORT_DIR
        / ("%s_주간보고_%s.xlsx" % (data["project"], stamp))
    )
    build(data, out)
    print("생성: %s" % out)
    print(
        "  금주 실적 %d건 / 차주 계획 %d건 / 이슈 %d건"
        % (
            len(data.get("this_week", [])),
            len(data.get("next_week", [])),
            len(data.get("issues", [])),
        )
    )
    return 0


def main():
    p = argparse.ArgumentParser(description="주간보고 Excel 생성기 (프로젝트 공용)")
    sub = p.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("init", help="주차 데이터 JSON 뼈대 생성")
    pi.add_argument("project")
    pi.add_argument("--week", help="ISO 주차 (예: 2026-W34). 미지정 시 이번 주")
    pi.add_argument("--force", action="store_true")
    pi.set_defaults(func=cmd_init)

    pg = sub.add_parser("generate", help="JSON 을 주간보고 Excel 로 변환")
    pg.add_argument("project")
    pg.add_argument("--week", help="ISO 주차 (예: 2026-W34). 미지정 시 이번 주")
    pg.set_defaults(func=cmd_generate)

    args = p.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
