# -*- coding: utf-8 -*-
"""KT DS AI 가드레일 견적서 생성 — 씨커스 표준 견적 양식(05_견적서_표준양식_씨커스.xlsx) 기준."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\repos\external-projects\ktds-kyra-guardrail\[KT DS] AI가드레일_구축_견적서_20260828.xlsx"

GULIM = "굴림체"
DARK = "FF333333"
GRAY = "FF969696"
LIME = "FF99CC00"
WHITE = "FFFFFFFF"
BAND = "FFF2F2F2"

thin = Side(style="thin")
med = Side(style="medium")
_S = {"thin": thin, "medium": med, None: Side()}


def B(l=None, r=None, t=None, b=None):
    return Border(left=_S[l], right=_S[r], top=_S[t], bottom=_S[b])


def put(ws, coord, value=None, *, sz=10, bold=False, color=None, fill=None,
        h=None, v="center", wrap=True, nf="General", bd=None, indent=0):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = Font(name=GULIM, size=sz, bold=bold, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
    c.number_format = nf
    if bd:
        c.border = bd
    return c


def est_height(cells, sz=10, line=14.0, pad=10.0, minimum=0.0):
    """(텍스트, 열너비) 목록에서 줄바꿈을 감안한 행 높이를 추정한다."""
    import math
    rows = 1
    for text, width in cells:
        if not isinstance(text, str):
            continue
        cap = max(4.0, width * 11.0 / sz - 2.0)
        n = 0
        for ln in text.splitlines() or [""]:
            w = sum(2 if ord(ch) > 0x1100 else 1 for ch in ln)
            n += max(1, int(math.ceil(w / cap)))
        rows = max(rows, n)
    return max(minimum, rows * line + pad)


# ---------------------------------------------------------------- 견적 데이터
GROUPS = [
    ("KYRA\nAI Guardrail\n플랫폼 공급", [
        ("- 가드레일 엔진 : 입력 · AI모델 동작 · 출력 3단계 다중화 파이프라인\n"
         "  (단계별 독립 판정 — 민감정보 필터링, 적대적 공격 방어, 요청 폭주 대응, 응답 결과 변형)",
         12_000_000, "SEC-006"),
        ("- 정책 엔진(OPA 기반 RBAC 5단계 + ABAC 속성 정책) 및 통합 관리 콘솔\n"
         "  (응답 범위 제한 · 권한별 기능 접근통제 · 실시간 모니터링 대시보드)",
         9_000_000, "SEC-008"),
        ("- 감사로그 · 추적성 모듈\n"
         "  (프롬프트 단위 입·출력 이력 보존, 해시 체인 무결성, Syslog / SIEM 연동)",
         7_000_000, "SEC-008"),
    ]),
    ("요건 대응\n구축 ·\n커스터마이징", [
        ("- 프롬프트 인젝션 방어 룰셋 구축 및 튜닝\n"
         "  (Jailbreak · 시스템 프롬프트 탈취 · 역할 전환 · 인코딩/다국어 우회, 2차 분류 모델 판정)",
         5_000_000, "SEC-008"),
        ("- 입·출력 콘텐츠 필터링 및 마스킹 정책 구축\n"
         "  (유해 콘텐츠 · 개인정보 · 기밀데이터 자동 검출, 검증 기반 오탐 억제, 마스킹 후 전달)",
         6_000_000, "SEC-008"),
        ("- 이상 행위 탐지 · 모니터링 체계 구축\n"
         "  (비정상 쿼리 패턴, 반복적 우회 시도, 과도한 요청(Rate Abuse) 실시간 탐지 및 관리자 경보)",
         5_000_000, "SEC-008"),
        ("- 응답 범위 제한(Topic Restriction) 및 권한별 기능 접근통제 정책 설계\n"
         "  (업무 목적 토픽 화이트리스트, 역할 × 분류등급 권한 매트릭스 수립)",
         4_000_000, "SEC-008"),
        ("- AI 플랫폼 탑재 및 다중망 배포 연동\n"
         "  (AI망 · 운영망 · 개발망 · DMZ · AI개발망 인스턴스 구성, 폐쇄망 오프라인 번들, 중앙 정책 동기화)",
         5_000_000, "SEC-008"),
    ]),
    ("시험 ·\n안정화 ·\n이관", [
        ("- 통합 · 보안 시험 및 성능 튜닝\n"
         "  (인젝션 공격 · 민감정보 유출 · 오탐 · 부하 시나리오 측정, 검사 구간 지연 최적화)",
         5_000_000, "공통"),
        ("- 운영 매뉴얼 · 보안정책서 작성, 운영자 교육 및 기술 이전",
         5_000_000, "공통"),
    ]),
]

DIRECT = sum(a for _, items in GROUPS for _, a, _ in items)
OVERHEAD = 3_780_000
TECHFEE = 3_220_000
TOTAL = DIRECT + OVERHEAD + TECHFEE
assert DIRECT == 63_000_000 and TOTAL == 70_000_000, (DIRECT, TOTAL)

wb = openpyxl.Workbook()

# ==================================================================== 견적서
ws = wb.active
ws.title = "견적서"
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = 9
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

for col, w in {"A": 1.77, "B": 5.55, "C": 9.0, "D": 9.43, "E": 60.5,
               "F": 13.32, "G": 10.55, "H": 14.5, "I": 14.5, "J": 12.0,
               "K": 1.77}.items():
    ws.column_dimensions[col].width = w

ws.merge_cells("B1:J1")
put(ws, "B1", "견    적    서", sz=34, h="center", wrap=False)
ws.row_dimensions[1].height = 43.5
ws.row_dimensions[2].height = 31.5
ws.row_dimensions[3].height = 31.5

ws.merge_cells("B4:F5")
put(ws, "B4", "케이티디에스(주) 귀중", sz=18, bold=True, color="FF000000", h="left")
ws.row_dimensions[4].height = 31.5
ws.row_dimensions[5].height = 13.5

LEFT = [
    "견  적   명 : AI 가드레일 구축 (공공사업 RFI 요건 SEC-006 · SEC-008 대응)",
    "참       조 : AX솔루션개발팀 박준현 님",
    "견 적 일 자 : 2026년 8월 28일",
    "유 효 기 간 : 견적일로부터 30일",
    "견 적 번 호 : SKS-20260828-01",
]
RIGHT = [
    "주식회사 씨커스(Seekers Inc.)",
    "대 표 이 사 : 최  경  호 (직인생략)",
    "전       화 : 02) 2039-8160",
    "팩       스 : 02) 2039-8161",
    "주       소 : 경상북도 포항시 북구 흥해읍 융합기술로66, ",
    "              포항지식산업센터 408~409호",
]
for i, txt in enumerate(LEFT):
    put(ws, "B%d" % (6 + i), txt, sz=12, color="FF000000", h="left", wrap=False)
for i, txt in enumerate(RIGHT):
    put(ws, "G%d" % (6 + i), txt, sz=10.5, color="FF000000", h="left", wrap=False)
ws.row_dimensions[6].height = 23.25
for r in range(7, 14):
    ws.row_dimensions[r].height = 14.25

put(ws, "B12", "아래와 같이 견적합니다.", sz=12, color="FF000000", h="left", wrap=False)

ws.merge_cells("B14:J14")
put(ws, "B14", "최종 견적 금액 : 일금 칠천만원 정(70,000,000원) (VAT 별도)",
    sz=14, bold=True, h="left", wrap=False)
ws.row_dimensions[14].height = 36.0

# 표 머리
ws.merge_cells("C15:D15")
for col in "BCDEFGHIJ":
    put(ws, "%s15" % col, None, sz=11, bold=True, color=WHITE, fill=DARK, h="center",
        bd=B("medium" if col == "B" else "thin",
             "medium" if col == "J" else "thin", "medium", "medium"))
for col, label in (("B", "SEQ"), ("C", "구분"), ("E", "세부내용"), ("F", "단가"),
                   ("G", "모듈(식)"), ("H", "금액"), ("I", "할인금액"), ("J", "비고")):
    ws["%s15" % col].value = label
ws.row_dimensions[15].height = 39.95

# 과업명 밴드
ws.merge_cells("B16:J16")
for col in "BCDEFGHIJ":
    put(ws, "%s16" % col, None, sz=12, bold=True, color="FF808080", fill=LIME, h="center",
        bd=B("medium" if col == "B" else None,
             "medium" if col == "J" else None, "medium", "medium"))
ws["B16"].value = "AI 가드레일 구축 — KYRA AI Guardrail 기반 (다중화 가드레일 · AI 플랫폼 탑재)"
ws.row_dimensions[16].height = 39.95

row = 17
group_ranges = []
for seq, (gname, items) in enumerate(GROUPS, start=1):
    start = row
    for detail, amount, note in items:
        put(ws, "B%d" % row, None, sz=10, color="FF000000", fill=WHITE, h="center",
            bd=B("medium", "thin", "thin", "thin"))
        put(ws, "C%d" % row, None, sz=10, bold=True, fill=WHITE, h="center",
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "D%d" % row, None, sz=10, fill=WHITE, h="center",
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "E%d" % row, detail, sz=10, fill=WHITE, h="left", indent=1,
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "F%d" % row, amount, sz=10, fill=WHITE, h="right", nf="#,##0",
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "G%d" % row, 1, sz=10, fill=WHITE, h="center",
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "H%d" % row, "=F%d*G%d" % (row, row), sz=10, fill=WHITE, h="right",
            nf="#,##0", bd=B("thin", "thin", "thin", "thin"))
        put(ws, "I%d" % row, 0, sz=10, fill=WHITE, h="right", nf="#,##0",
            bd=B("thin", "thin", "thin", "thin"))
        put(ws, "J%d" % row, note, sz=10, fill=WHITE, h="center",
            bd=B("thin", "medium", "thin", "thin"))
        ws.row_dimensions[row].height = est_height([(detail, 60.5)], minimum=42.0)
        row += 1
    end = row - 1
    ws.merge_cells("B%d:B%d" % (start, end))
    ws["B%d" % start].value = seq
    ws.merge_cells("C%d:D%d" % (start, end))
    ws["C%d" % start].value = gname
    group_ranges.append((start, end))

item_first, item_last = group_ranges[0][0], group_ranges[-1][1]


def subtotal_row(r, label, formula, top="thin", bottom="thin"):
    ws.merge_cells("B%d:E%d" % (r, r))
    for col in "BCDE":
        put(ws, "%s%d" % (col, r), None, sz=11, bold=True, color="FF000000", fill=BAND,
            h="center", bd=B("medium" if col == "B" else None, None, top, bottom))
    ws["B%d" % r].value = label
    for col, val, nf in (("F", "-", "General"), ("G", "-", "General"),
                         ("H", formula, "#,##0"), ("I", 0, "#,##0"),
                         ("J", None, "General")):
        put(ws, "%s%d" % (col, r), val, sz=11, bold=True, fill=BAND,
            h="right" if col in "HI" else "center", nf=nf,
            bd=B("thin", "medium" if col == "J" else "thin", top, bottom))
    ws.row_dimensions[r].height = 28.0


r_sub1 = row
subtotal_row(r_sub1, "소계", "=SUM(H%d:H%d)" % (item_first, item_last))
row += 1

for seq, (name, detail, val, note) in enumerate(
        [("재경비", "- 재경비 (일반관리비, 직접인건비의 6%)", OVERHEAD, "6%"),
         ("기술료", "- 기술료 (이윤 상당, 10% 이내)", TECHFEE, "4.82%")],
        start=len(GROUPS) + 1):
    put(ws, "B%d" % row, seq, sz=10, color="FF000000", fill=WHITE, h="center",
        bd=B("medium", "thin", "thin", "thin"))
    ws.merge_cells("C%d:D%d" % (row, row))
    for col in "CD":
        put(ws, "%s%d" % (col, row), None, sz=10, bold=True, fill=WHITE, h="center",
            bd=B("thin", "thin", "thin", "thin"))
    ws["C%d" % row].value = name
    put(ws, "E%d" % row, detail, sz=10, fill=WHITE, h="left", indent=1,
        bd=B("thin", "thin", "thin", "thin"))
    put(ws, "F%d" % row, "-", sz=10, fill=WHITE, h="center",
        bd=B("thin", "thin", "thin", "thin"))
    put(ws, "G%d" % row, "-", sz=10, fill=WHITE, h="center",
        bd=B("thin", "thin", "thin", "thin"))
    put(ws, "H%d" % row, val, sz=10, fill=WHITE, h="right", nf="#,##0",
        bd=B("thin", "thin", "thin", "thin"))
    put(ws, "I%d" % row, 0, sz=10, fill=WHITE, h="right", nf="#,##0",
        bd=B("thin", "thin", "thin", "thin"))
    put(ws, "J%d" % row, note, sz=10, fill=WHITE, h="center",
        bd=B("thin", "medium", "thin", "thin"))
    ws.row_dimensions[row].height = 26.0
    row += 1

r_sub2 = row
subtotal_row(r_sub2, "소계", "=SUM(H%d:H%d)" % (r_sub2 - 2, r_sub2 - 1))
row += 1

# 합계
r_tot = row
ws.merge_cells("B%d:E%d" % (r_tot, r_tot))
for col in "BCDE":
    put(ws, "%s%d" % (col, r_tot), None, sz=11, color=WHITE, fill=GRAY, h="center",
        bd=B("medium" if col == "B" else None, None, "medium", "medium"))
ws["B%d" % r_tot].value = "합계"
for col, val, nf in (("F", "-", "General"), ("G", "-", "General"),
                     ("H", "=H%d+H%d" % (r_sub1, r_sub2), "#,##0"),
                     ("I", "=I%d+I%d" % (r_sub1, r_sub2), "#,##0"),
                     ("J", None, "General")):
    put(ws, "%s%d" % (col, r_tot), val, sz=11,
        color="FF000000" if col in "HI" else WHITE, fill=GRAY,
        h="right" if col in "HI" else "center", nf=nf,
        bd=B("thin", "medium" if col == "J" else "thin", "medium", "medium"))
ws.row_dimensions[r_tot].height = 30.0
row += 1

# 특별 견적 금액
r_sp = row
for pair in ("B%d:E%d", "F%d:F%d", "G%d:G%d", "H%d:H%d", "I%d:I%d"):
    ws.merge_cells(pair % (r_sp, r_sp + 1) if pair.startswith("B")
                   else pair % (r_sp, r_sp + 1))
for rr in (r_sp, r_sp + 1):
    for col in "BCDEFGHIJ":
        put(ws, "%s%d" % (col, rr), None, sz=11, bold=True, color=WHITE, fill=DARK,
            h="center",
            bd=B("medium" if col == "B" else "thin",
                 "medium" if col == "J" else "thin",
                 "medium" if rr == r_sp else None,
                 "medium" if rr == r_sp + 1 else None))
    ws.row_dimensions[rr].height = 20.1
ws["B%d" % r_sp].value = "특별 견적 금액"
put(ws, "H%d" % r_sp, "=H%d" % r_tot, sz=11, bold=True, color=WHITE, fill=DARK,
    h="right", nf="#,##0", bd=B("thin", "thin", "medium", "medium"))
put(ws, "I%d" % r_sp, "=H%d" % r_tot, sz=11, bold=True, color=WHITE, fill=DARK,
    h="right", nf="#,##0", bd=B("thin", "thin", "medium", "medium"))
ws["J%d" % r_sp].value = "*할인률"
put(ws, "J%d" % (r_sp + 1), "=100%%-(I%d/H%d)" % (r_sp, r_sp), sz=11, color=WHITE,
    fill=DARK, h="center", nf="0.00%", bd=B("thin", "medium", None, "medium"))
row += 2

# 비고
r_note = row
ws.merge_cells("B%d:J%d" % (r_note, r_note + 1))
for rr in (r_note, r_note + 1):
    for col in "BCDEFGHIJ":
        put(ws, "%s%d" % (col, rr), None, sz=11, h="left",
            bd=B("medium" if col == "B" else None,
                 "medium" if col == "J" else None,
                 "medium" if rr == r_note else None,
                 "medium" if rr == r_note + 1 else None))
ws["B%d" % r_note].value = (
    "※ 비 고\n"
    "   - 본 견적은 공공사업 RFI 요건 SEC-006(AI시스템 가드레일 다중화 구성) 및 "
    "SEC-008(AI 플랫폼 내 가드레일 탑재) 대응 범위에 대한 견적입니다.\n"
    "   - 공급 방식은 당사 KYRA AI Guardrail 제품을 도입하고 상기 요건에 맞추어 "
    "커스터마이징·연동하는 구성이며, 요건별 대응 내역은 「요건대응표」 시트를 참조하여 주시기 바랍니다.\n"
    "   - 재경비(일반관리비 6%)·기술료(이윤 10% 이내)는 국가계약법 시행규칙 제8조 "
    "원가계산 기준에 따라 산정하였습니다.\n"
    "   - 구축 부분의 인력 투입 산출 내역은 「산출근거」 시트에 정리하였으며, 단가는 "
    "SW기술자 직무별 평균임금 체계를 따른 당사 표준 단가 기준입니다.\n"
    "   - 상기 금액은 부가가치세(VAT) 별도 금액이며, 계약 조건은 계약서에 따릅니다.\n"
    "   - 서버·스토리지 등 하드웨어, 상용 LLM API 사용료, 타사 솔루션 라이선스 비용은 "
    "본 견적에 포함되어 있지 않습니다."
)
ws.row_dimensions[r_note].height = 132.0
ws.row_dimensions[r_note + 1].height = 12.0
ws.print_area = "A1:K%d" % (r_note + 1)

# ================================================================ 요건대응표
ws2 = wb.create_sheet("요건대응표")
ws2.sheet_view.showGridLines = False
ws2.page_setup.orientation = "landscape"
ws2.page_setup.paperSize = 9
ws2.sheet_properties.pageSetUpPr.fitToPage = True
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 1
ws2.print_title_rows = "4:4"
for col, w in {"A": 1.77, "B": 11.0, "C": 46.0, "D": 62.0, "E": 16.0, "F": 13.0,
               "G": 1.77}.items():
    ws2.column_dimensions[col].width = w

ws2.merge_cells("B1:F1")
put(ws2, "B1", "RFI 요건 대응표 — AI 가드레일 (SEC-006 · SEC-008)", sz=20, bold=True,
    h="center", wrap=False)
ws2.row_dimensions[1].height = 40.0

ws2.merge_cells("B2:F2")
put(ws2, "B2", "KYRA AI Guardrail 기반 · 주식회사 씨커스 · 2026년 8월 28일 "
               "(견적번호 SKS-20260828-01)", sz=11, h="center", wrap=False)
ws2.row_dimensions[2].height = 22.0

REQ_HDR = ["요건번호", "RFI 요건", "KYRA AI Guardrail 대응 내용", "제공 방식", "견적 항목"]
for i, label in enumerate(REQ_HDR):
    col = chr(ord("B") + i)
    put(ws2, "%s4" % col, label, sz=11, bold=True, color=WHITE, fill=DARK, h="center",
        bd=B("medium" if col == "B" else "thin",
             "medium" if col == "F" else "thin", "medium", "medium"))
ws2.row_dimensions[4].height = 32.0

REQS = [
    ("SEC-006",
     "AI시스템 가드레일 다중화 구성\n"
     "(사용자의 입력, AI모델 동작, 출력 결과 등 각 동작 단계별로 민감정보 필터링, "
     "적대적 공격 방어, 서비스 거부 공격 대응, 응답 결과 변형 등을 수행할 수 있는 "
     "복수의 가드레일을 배치)",
     "입력 · 모델 동작 · 출력 3단계에 각각 독립 판정기를 배치하는 다중 가드레일 "
     "파이프라인으로 제공합니다.\n"
     "  · 입력 단계 — 민감정보 탐지·마스킹, 적대적 프롬프트 탐지, 요청 쿼터 및 "
     "Rate Limit 기반 서비스 거부 공격 대응\n"
     "  · 모델 동작 단계 — 도구 호출 화이트리스트, 참조 문서 권한 검증, 시스템 프롬프트 보호\n"
     "  · 출력 단계 — 유해 콘텐츠·기밀정보 재검사, 응답 재작성 및 마스킹(응답 결과 변형)\n"
     "각 단계는 독립적으로 판정하며 어느 한 단계를 우회하더라도 후속 단계에서 차단됩니다.",
     "제품 + 정책 구성", "1-1 / 2-1~2-3"),
    ("SEC-008",
     "AI 가드레일을 AI 플랫폼 내 탑재하며 AI 플랫폼 전체"
     "(AI망, 운영망, 개발망, DMZ 및 AI개발망)에 구동되어야 함",
     "컨테이너 기반 배포 패키지로 AI 플랫폼 내부에 탑재하며, 망별로 가드레일 인스턴스를 "
     "구성하고 중앙 정책 저장소와 동기화합니다.\n"
     "폐쇄망 환경을 위한 오프라인 이미지 번들과 설치·기동 스크립트를 함께 제공하여 "
     "외부 네트워크 없이 설치가 가능합니다.",
     "구축 · 배포 연동", "2-5"),
    ("SEC-008",
     "[프롬프트 인젝션 방지]",
     "1차 룰셋으로 Jailbreak, 시스템 프롬프트 탈취, 역할 전환, 인코딩 우회, 다국어 우회 "
     "패턴을 탐지하고, 1차 통과분에 대해 분류 모델 기반 2차 판정을 수행합니다.\n"
     "판정 결과는 위험 스코어와 차단 사유로 기록되어 오탐 튜닝 근거로 활용됩니다.",
     "제품 + 룰셋 튜닝", "1-1 / 2-1"),
    ("SEC-008",
     "[입·출력 콘텐츠 필터링]\n"
     "유해 콘텐츠, 개인정보, 기밀 데이터가 AI 응답에 포함되지 않도록 출력값에 대한 "
     "자동 검열 및 마스킹 처리 기능 제공",
     "주민등록번호·계좌번호·카드번호·연락처·이메일 등을 정규식과 검증(체크섬)으로 "
     "판별하여 오탐을 억제하고, 문서 분류등급이 부여된 기밀 데이터와 유해 콘텐츠를 "
     "분류기로 탐지합니다.\n"
     "탐지 시 차단 또는 마스킹 후 전달을 정책으로 선택할 수 있으며, 입력과 출력 양방향에 "
     "동일하게 적용됩니다.",
     "제품 + 규칙 커스터마이징", "1-1 / 2-2"),
    ("SEC-008",
     "[이상 행위 탐지 및 모니터링]\n"
     "비정상적인 쿼리 패턴, 반복적 우회 시도, 과도한 요청(Rate Abuse) 등을 실시간으로 "
     "탐지하고 관리자에게 경보를 발행하는 모니터링 체계 운영",
     "사용자·에이전트 단위로 행위 프로파일을 유지하여 평소와 다른 쿼리 패턴을 탐지하고, "
     "반복 차단 시도를 누적 스코어링하여 임계치 초과 시 세션을 제한합니다.\n"
     "과도한 요청은 단위 시간당 호출량 임계치로 차단하며, 실시간 대시보드와 함께 "
     "메일 · Syslog · SIEM 경보를 발행합니다.",
     "제품 + 임계치 튜닝", "1-2 / 2-3"),
    ("SEC-008",
     "[응답 범위 제한 및 접근 통제]\n"
     "AI 모델의 응답 범위를 업무 목적에 한정하는 토픽 제한(Topic Restriction) 정책을 "
     "적용하고, 사용자 권한에 따른 기능 접근통제 시행",
     "업무 목적 토픽 화이트리스트와 금지 토픽 정책을 적용하여 응답 범위를 한정합니다.\n"
     "접근통제는 OPA 정책 엔진 기반의 RBAC 5단계 역할 체계와 ABAC 속성 정책"
     "(문서 분류등급 · 사용자 역할 · 소속 · 동작 유형 · 시간대)을 함께 평가하는 "
     "이중 구조이며, 정책 변경은 재기동 없이 즉시 반영됩니다.",
     "제품 + 정책 설계", "1-2 / 2-4"),
    ("SEC-008",
     "[감사 로그 및 추적성 확보]\n"
     "모든 입·출력 이력을 보존하여 보안 사고발생 시 원인 추적 및 감사가 가능하도록 "
     "로그 관리 체계 구축",
     "프롬프트 단위로 요청·응답·판정 결과(허용 / 마스킹 / 차단)와 사유를 건별로 "
     "기록하고, 해시 체인으로 위·변조를 검증합니다.\n"
     "감사 조회 콘솔에서 사용자·기간·판정별 추적이 가능하며, Syslog 전송으로 기관 SIEM에 "
     "연동하여 장기 보존과 상관분석을 지원합니다.",
     "제품 + 연동", "1-3"),
]

r = 5
for no, req, ans, mode, item in REQS:
    vals = [no, req, ans, mode, item]
    for i, v in enumerate(vals):
        col = chr(ord("B") + i)
        put(ws2, "%s%d" % (col, r), v, sz=10,
            bold=(i == 0), fill=WHITE,
            h="left" if i in (1, 2) else "center", v="center", indent=1 if i in (1, 2) else 0,
            bd=B("medium" if col == "B" else "thin",
                 "medium" if col == "F" else "thin", "thin", "thin"))
    ws2.row_dimensions[r].height = est_height([(req, 46.0), (ans, 62.0)], minimum=64.0)
    r += 1

ws2.merge_cells("B%d:F%d" % (r + 1, r + 1))
put(ws2, "B%d" % (r + 1),
    "※ 상기 대응 내용은 KYRA AI Guardrail 제품에 기 구현된 기능과, 본 견적에 포함된 "
    "요건 대응 커스터마이징 범위를 함께 표기한 것입니다. "
    "「견적 항목」 열은 견적서 시트의 세부 항목 순번을 의미합니다.",
    sz=10, h="left", bd=B("medium", "medium", "medium", "medium"))
ws2.row_dimensions[r + 1].height = 36.0
ws2.print_area = "A1:G%d" % (r + 1)

# ================================================================== 산출근거
ws3 = wb.create_sheet("산출근거")
ws3.sheet_view.showGridLines = False
ws3.page_setup.orientation = "landscape"
ws3.page_setup.paperSize = 9
ws3.sheet_properties.pageSetUpPr.fitToPage = True
ws3.page_setup.fitToWidth = 1
ws3.page_setup.fitToHeight = 1
for col, w in {"A": 1.77, "B": 5.5, "C": 20.0, "D": 52.0, "E": 8.0, "F": 14.0,
               "G": 10.0, "H": 16.0, "I": 1.77}.items():
    ws3.column_dimensions[col].width = w

ws3.merge_cells("B1:H1")
put(ws3, "B1", "산 출 근 거", sz=24, bold=True, h="center", wrap=False)
ws3.row_dimensions[1].height = 42.0

ws3.merge_cells("B2:H2")
put(ws3, "B2", "AI 가드레일 구축 (견적번호 SKS-20260828-01) · 주식회사 씨커스",
    sz=11, h="center", wrap=False)
ws3.row_dimensions[2].height = 22.0

# 1. 총괄
ws3.merge_cells("B4:H4")
put(ws3, "B4", "1. 총괄 (VAT 별도)", sz=13, bold=True, h="left",
    bd=B(None, None, None, "medium"))
ws3.row_dimensions[4].height = 28.0

TOTALS = [
    ("구분", "산출 기준", "금액", True),
    ("제품 공급", "KYRA AI Guardrail 플랫폼 (가드레일 엔진 · 정책 엔진/관리 콘솔 · 감사로그 모듈)",
     28_000_000, False),
    ("구축 · 커스터마이징 · 이관", "요건 대응 구축 및 시험·안정화·이관 — 인력 투입 4.0 M/M (아래 2항)",
     35_000_000, False),
    ("직접비 소계", "", 63_000_000, False),
    ("재경비", "일반관리비 6%", 3_780_000, False),
    ("기술료", "이윤 상당 (10% 이내, 4.82%)", 3_220_000, False),
    ("합계", "", 70_000_000, False),
]
r = 5
for name, basis, amount, is_hdr in TOTALS:
    emphasis = name in ("직접비 소계", "합계")
    fill = DARK if is_hdr else (BAND if emphasis else WHITE)
    color = WHITE if is_hdr else "FF000000"
    ws3.merge_cells("D%d:G%d" % (r, r))
    ws3.merge_cells("B%d:C%d" % (r, r))
    for col in "BCDEFGH":
        put(ws3, "%s%d" % (col, r), None, sz=11, bold=is_hdr or emphasis, color=color,
            fill=fill, h="center",
            bd=B("medium" if col == "B" else "thin",
                 "medium" if col == "H" else "thin",
                 "medium" if is_hdr else "thin",
                 "medium" if (is_hdr or name == "합계") else "thin"))
    put(ws3, "B%d" % r, name, sz=11, bold=is_hdr or emphasis, color=color, fill=fill,
        h="center", bd=B("medium", "thin", "medium" if is_hdr else "thin",
                         "medium" if (is_hdr or name == "합계") else "thin"))
    put(ws3, "D%d" % r, basis if not is_hdr else "산출 기준", sz=11, bold=is_hdr,
        color=color, fill=fill, h="left", indent=1,
        bd=B("thin", "thin", "medium" if is_hdr else "thin",
             "medium" if (is_hdr or name == "합계") else "thin"))
    put(ws3, "H%d" % r, amount if not is_hdr else "금액", sz=11,
        bold=is_hdr or emphasis, color=color, fill=fill, h="right",
        nf="General" if is_hdr else "#,##0",
        bd=B("thin", "medium", "medium" if is_hdr else "thin",
             "medium" if (is_hdr or name == "합계") else "thin"))
    ws3.row_dimensions[r].height = 30.0
    r += 1

# 2. 인력 투입 산출
r += 1
ws3.merge_cells("B%d:H%d" % (r, r))
put(ws3, "B%d" % r, "2. 구축 · 커스터마이징 · 이관 인력 투입 산출", sz=13, bold=True,
    h="left", bd=B(None, None, None, "medium"))
ws3.row_dimensions[r].height = 28.0
r += 1

MAN_HDR = ["SEQ", "직무", "담당 업무", "인원", "단가(월)", "투입월(M/M)", "금액"]
for i, label in enumerate(MAN_HDR):
    col = chr(ord("B") + i)
    put(ws3, "%s%d" % (col, r), label, sz=11, bold=True, color=WHITE, fill=DARK,
        h="center", bd=B("medium" if col == "B" else "thin",
                         "medium" if col == "H" else "thin", "medium", "medium"))
ws3.row_dimensions[r].height = 32.0
man_hdr_row = r
r += 1

MANPOWER = [
    ("정보보안전문가", "가드레일 정책 설계, 프롬프트 인젝션 방어 룰셋, DLP 규칙 및 "
     "접근통제 정책 수립, 보안 시험", 1, 10_411_680, 1.5),
    ("응용SW개발자", "AI 플랫폼 탑재 연동, 다중망 배포 구성, 경보·SIEM 연동 개발", 1, 7_754_124, 1.0),
    ("시스템SW개발자", "폐쇄망 패키징, 감사로그 파이프라인 및 모니터링 지표 구성", 1, 5_840_196, 1.0),
    ("IT품질관리자", "통합·성능 시험, 오탐 튜닝, 운영 매뉴얼·보안정책서 작성, 운영자 교육",
     1, 11_042_071, 0.5),
]
man_first = r
for seq, (job, duty, cnt, rate, mm) in enumerate(MANPOWER, start=1):
    vals = [seq, job, duty, cnt, rate, mm, "=ROUND(E%d*F%d*G%d,0)" % (r, r, r)]
    for i, v in enumerate(vals):
        col = chr(ord("B") + i)
        nf = "#,##0" if col in ("F", "H") else ("0.0" if col == "G" else "General")
        put(ws3, "%s%d" % (col, r), v, sz=10, fill=WHITE,
            h="left" if col == "D" else ("right" if col in ("F", "H") else "center"),
            indent=1 if col == "D" else 0, nf=nf,
            bd=B("medium" if col == "B" else "thin",
                 "medium" if col == "H" else "thin", "thin", "thin"))
    ws3.row_dimensions[r].height = 40.0
    r += 1
man_last = r - 1

ws3.merge_cells("B%d:D%d" % (r, r))
for col in "BCDEFGH":
    put(ws3, "%s%d" % (col, r), None, sz=11, bold=True, fill=BAND, h="center",
        bd=B("medium" if col == "B" else "thin",
             "medium" if col == "H" else "thin", "thin", "medium"))
ws3["B%d" % r].value = "인력 투입 산출 소계"
put(ws3, "E%d" % r, "-", sz=11, bold=True, fill=BAND, h="center",
    bd=B("thin", "thin", "thin", "medium"))
put(ws3, "F%d" % r, "-", sz=11, bold=True, fill=BAND, h="center",
    bd=B("thin", "thin", "thin", "medium"))
put(ws3, "G%d" % r, "=SUM(G%d:G%d)" % (man_first, man_last), sz=11, bold=True,
    fill=BAND, h="center", nf="0.0", bd=B("thin", "thin", "thin", "medium"))
put(ws3, "H%d" % r, "=SUM(H%d:H%d)" % (man_first, man_last), sz=11, bold=True,
    fill=BAND, h="right", nf="#,##0", bd=B("thin", "medium", "thin", "medium"))
ws3.row_dimensions[r].height = 30.0
r += 2

ws3.merge_cells("B%d:H%d" % (r, r + 1))
for rr in (r, r + 1):
    for col in "BCDEFGH":
        put(ws3, "%s%d" % (col, rr), None, sz=10, h="left",
            bd=B("medium" if col == "B" else None,
                 "medium" if col == "H" else None,
                 "medium" if rr == r else None,
                 "medium" if rr == r + 1 else None))
ws3["B%d" % r].value = (
    "※ 산출 기준\n"
    "   - 단가는 SW기술자 직무별 평균임금 체계를 따른 당사 표준 단가 기준입니다.\n"
    "   - 인력 투입 산출 합계는 34,732,876원이며, 견적서에는 협의 정액 기준으로 "
    "35,000,000원을 반영하였습니다.\n"
    "   - 제품 공급가 28,000,000원은 KYRA AI Guardrail 플랫폼의 공급 대가로, "
    "인력 투입 산출과 중복되지 않습니다.\n"
    "   - 재경비 및 기술료는 국가계약법 시행규칙 제8조 원가계산 기준에 따라 "
    "직접비를 기준으로 산정하였습니다.\n"
    "   - 상기 금액은 부가가치세(VAT) 별도 금액입니다."
)
ws3.row_dimensions[r].height = 108.0
ws3.row_dimensions[r + 1].height = 12.0
ws3.print_area = "A1:I%d" % (r + 1)

wb.save(OUT)
print("saved:", OUT)
print("DIRECT=%d OVERHEAD=%d TECHFEE=%d TOTAL=%d" % (DIRECT, OVERHEAD, TECHFEE, TOTAL))
print("manpower sum =", sum(c * rate * mm for _, _, c, rate, mm in MANPOWER))
print("techfee ratio = %.4f%%" % (TECHFEE / (DIRECT + OVERHEAD) * 100))
