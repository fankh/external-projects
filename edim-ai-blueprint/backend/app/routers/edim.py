"""EDIM /api/v1 — 실 PostgreSQL 기반 (OpenAPI 스펙 경로 준수).

인증: POST /auth/login → HMAC 토큰. 데이터 엔드포인트는 개발 단계라 토큰 검증을
강제하지 않는다 (게이트웨이/JWT 전환 시 Depends 로 교체).
"""
from __future__ import annotations

import asyncio
import hashlib
import hmac
import io
import json
import os
import re
import time
from contextvars import ContextVar
from datetime import date
from typing import Any

from fastapi import APIRouter, HTTPException, Request, Response, UploadFile
from fastapi import Depends, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.db import db_ok, get_pool
from app.services import storage
from app.services.edim_seed import TENANT
from app.services.macro_engine import Evaluator, MacroError

SECRET = os.getenv("EDIM_SECRET", "edim-dev-secret").encode()

# 공개 경로 — 그 외 전부 Bearer 토큰 필수 (P0-1)
PUBLIC_SUFFIXES = ("/health", "/auth/login")


LEVEL_RANK = {"GENERAL": 0, "SETUP": 1, "ADMIN": 2, "PLATFORM": 3}

TOKEN_TTL = 8 * 3600        # 발급 수명
RENEW_WINDOW = 30 * 60      # 만료 30분 전부터 응답 헤더로 재발급 (B8 — 하드컷 제거)


# 1.2 — 멀티테넌시: 요청별 테넌트 컨텍스트 (미설정 시 환경 기본 테넌트로 폴백).
# 세션 토큰이 테넌트를 담고, ASGI 미들웨어가 요청 문맥에 심는다 (동기 엔드포인트 스레드로 전파됨).
_TENANT_CTX: ContextVar[int | None] = ContextVar("edim_tenant_id", default=None)


def _issue_token(login: str, ttl: int = TOKEN_TTL, tenant_id: int | None = None) -> str:
    """세션 토큰 — `login.exp.tenantId.sig` (구형 3세그먼트도 검증 호환)."""
    exp = int(time.time()) + ttl
    tid = tenant_id if tenant_id is not None else (_TENANT_CTX.get() or 0)
    payload = f"{login}.{exp}.{tid}"
    sig = hmac.new(SECRET, payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}.{sig}"


def _parse_token(token: str) -> tuple[str, int, int | None] | None:
    """(login, exp, tenantId|None) — 서명·만료 검증 통과 시에만 반환.

    신형 `login.exp.tenantId.sig` 를 먼저 시도하고, 실패하면 구형 `login.exp.sig` 로 재해석한다.
    (login 에 '.' 이 포함될 수 있어 — park.f — 세그먼트 수만으로는 구분되지 않는다.)
    """
    def _try(n: int, with_tenant: bool) -> tuple[str, int, int | None] | None:
        parts = token.rsplit(".", n)
        if len(parts) != n + 1:
            return None
        if with_tenant:
            login, exp, tid_s, sig = parts
            if not tid_s.isdigit():
                return None
            payload, tid = f"{login}.{exp}.{tid_s}", int(tid_s)
        else:
            login, exp, sig = parts
            payload, tid = f"{login}.{exp}", None
        if not exp.isdigit() or int(exp) < time.time():
            return None
        expected = hmac.new(SECRET, payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, sig):
            return None
        return login, int(exp), (tid or None)

    return _try(3, True) or _try(2, False)


def require_auth(request: Request, response: Response) -> None:
    # /i18n/{locale} 번들은 로그인 화면에서도 필요 — 공개. 단 /i18n/data* (데이터 번역 관리·오버레이)는 인증 필요.
    if request.url.path.endswith(PUBLIC_SUFFIXES) or (
            "/i18n/" in request.url.path and "/i18n/data" not in request.url.path):
        return
    auth = request.headers.get("authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, detail="인증 필요 — 로그인 토큰이 없습니다")
    parsed = _parse_token(auth[7:])
    if not parsed:
        raise HTTPException(401, detail="토큰 무효 — 다시 로그인하십시오 (서명·형식·만료)")
    login, exp, tok_tid = parsed
    # 사용자 컨텍스트 (RBAC·알림 대상) — SYS-005. 테넌트는 토큰 기준 (1.2 멀티테넌시)
    with _conn() as conn, conn.cursor() as cur:
        tid = tok_tid or _tenant_id(cur)
        cur.execute(
            """SELECT user_id, user_level FROM sys_user
               WHERE tenant_id=%s AND login_id=%s AND status='ACTIVE'""", (tid, login))
        row = cur.fetchone()
    if not row:
        raise HTTPException(401, detail="비활성/미존재 사용자")
    request.state.login = login
    request.state.user_id = row[0]
    request.state.level = row[1]
    request.state.tenant_id = tid
    # 슬라이딩 갱신 — 잔여 30분 미만이면 새 토큰을 헤더로 전달 (프론트가 교체)
    if int(exp) - time.time() < RENEW_WINDOW:
        response.headers["X-EDIM-Token"] = _issue_token(login, tenant_id=tid)


class TenantContextMiddleware:
    """1.2 — 토큰의 테넌트를 요청 문맥(ContextVar)에 심는 ASGI 미들웨어.

    같은 태스크에서 downstream 을 호출하므로 동기 엔드포인트(스레드풀)까지 값이 전파된다.
    서명·만료 검증을 통과한 토큰만 반영하며, 나머지는 환경 기본 테넌트로 폴백한다.
    """

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope.get("type") != "http":
            await self.app(scope, receive, send)
            return
        tid = None
        for k, v in scope.get("headers", []):
            if k == b"authorization" and v.startswith(b"Bearer "):
                parsed = _parse_token(v[7:].decode("latin-1"))
                if parsed:
                    tid = parsed[2]
                break
        token = _TENANT_CTX.set(tid)
        try:
            await self.app(scope, receive, send)
        finally:
            _TENANT_CTX.reset(token)


def min_level(name: str):
    """엔드포인트 최소 권한 — 권한승인정의서 매트릭스 축약 (쓰기=SETUP+, 관리자 작업=ADMIN)."""
    def dep(request: Request) -> None:
        lvl = getattr(request.state, "level", None)
        if lvl is None:
            raise HTTPException(401, detail="인증 필요")
        if LEVEL_RANK[lvl] < LEVEL_RANK[name]:
            raise HTTPException(
                403, detail=f"권한 부족 — {name} 이상 필요 (현재 {lvl}, SYS-005)")
    return dep


SETUP = Depends(min_level("SETUP"))
ADMIN = Depends(min_level("ADMIN"))


def _platform_guard(request: Request) -> None:
    """1.3 — 플랫폼 계층 (요구 #5 2계층 권한 1단계).

    고객사 프로비저닝은 EDIM 운영 테넌트(환경 기본)의 ADMIN 만 수행한다.
    고객사 ADMIN 은 자기 테넌트 안에서만 관리자다.
    """
    lvl = getattr(request.state, "level", None)
    if lvl is None:
        raise HTTPException(401, detail="인증 필요")
    if LEVEL_RANK[lvl] < LEVEL_RANK["ADMIN"]:
        raise HTTPException(403, detail="권한 부족 — ADMIN 이상 필요")
    with _conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT tenant_code FROM sys_tenant WHERE tenant_id=%s",
                    (getattr(request.state, "tenant_id", 0),))
        row = cur.fetchone()
    if not row or row[0] != TENANT:
        raise HTTPException(403, detail="플랫폼 운영 권한 필요 — 고객사 관리자는 자사 범위만 관리합니다")


PLATFORM = Depends(_platform_guard)


# ── 1.5 정보 접근 권한·마스킹 (요구 #4/#6) — 작업 권한(RBAC)과 분리된 열람 통제 ──
# 정보그룹: 원가·단가·견적금액·거래처(고객/공급자명). 모드: full/masked/summary/hidden/no_download.
INFO_GROUPS = {
    "cost": "원가 (실적·PCR·계산 상세)",
    "price": "단가 (구매·판매)",
    "quote": "견적 금액",
    "partner": "거래처·공급자명",
}
INFO_MODES = ("full", "masked", "summary", "hidden", "no_download")
MASK_TEXT = "••••"


def _info_mode(cur, tid: int, request: Request, group: str) -> str:
    """현재 사용자의 정보그룹 열람 모드 — 임시 접근(유효기간) > 역할 규칙 > full(기본).

    미설정이면 full 이라 기존 동작이 보존된다(도입 시 무영향).
    """
    uid = getattr(request.state, "user_id", None)
    if uid:
        cur.execute(
            """SELECT mode FROM sys_temp_access
               WHERE tenant_id=%s AND user_id=%s AND info_group=%s AND NOT revoked
                 AND valid_from <= now() AND valid_to > now()
               ORDER BY temp_access_id DESC LIMIT 1""", (tid, uid, group))
        t = cur.fetchone()
        if t:
            return t[0]
    lvl = getattr(request.state, "level", None)
    if not lvl:
        return "full"
    # 사용자 등급(암묵 역할) + 배정 역할 중 가장 제한적인 규칙 적용 (Deny 우선 — 요구 #5)
    cur.execute(
        """SELECT ia.mode FROM sys_info_access ia
           WHERE ia.tenant_id=%s AND ia.info_group=%s
             AND (ia.role_name=%s OR ia.role_name IN (
                   SELECT r.role_name FROM sys_user_role ur JOIN sys_role r ON r.role_id=ur.role_id
                   WHERE ur.user_id=%s))""", (tid, group, lvl, uid))
    modes = [r[0] for r in cur.fetchall()]
    if not modes:
        return "full"
    order = {"hidden": 0, "summary": 1, "masked": 2, "no_download": 3, "full": 4}
    return min(modes, key=lambda m: order.get(m, 4))


def _mask_num(value: float | None, mode: str) -> Any:
    """숫자 민감값 마스킹 — hidden/summary 는 값 제거, masked 는 자릿수만 노출."""
    if value is None or mode in ("full", "no_download"):
        return value
    if mode in ("hidden", "summary"):
        return None
    # masked — 상위 1자리만 남기고 절사 (규모만 파악, 실값 비노출)
    try:
        n = int(abs(float(value)))
    except (TypeError, ValueError):
        return None
    digits = len(str(n))
    return f"{str(n)[0]}{'0' * (digits - 1)}~" if digits > 1 else "~"


def _mask_text(value: str | None, mode: str) -> str | None:
    if value is None or mode in ("full", "no_download"):
        return value
    if mode == "hidden":
        return None
    return (value[:1] + MASK_TEXT) if value else value


def _assert_downloadable(cur, tid: int, request: Request, group: str) -> None:
    """no_download/hidden/summary 모드는 내려받기 차단 (요구 #6 — 조회 가능 ≠ Export 가능)."""
    mode = _info_mode(cur, tid, request, group)
    if mode != "full":
        raise HTTPException(
            403, detail=f"다운로드 권한 없음 — {INFO_GROUPS.get(group, group)} 열람 모드: {mode}")


def _notify(cur, tid: int, user_id: int, notify_type: str, title: str,
            link: str | None = None) -> None:
    cur.execute(
        """INSERT INTO sys_notification (tenant_id, user_id, notify_type, title, link_url)
           VALUES (%s,%s,%s,%s,%s)""", (tid, user_id, notify_type, title[:200], link))


def _audit(cur, tid: int, target_table: str, target_id: int, action: str,
           actor_id: int, after: dict | None = None, before: dict | None = None) -> None:
    """보안 감사 — sys_history 공통 기록 (SYS-005·B8)."""
    cur.execute(
        """INSERT INTO sys_history (tenant_id, target_table, target_id, action,
                                    actor_id, before_data, after_data)
           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
        (tid, target_table, target_id, action[:20], actor_id,
         json.dumps(before) if before else None,
         json.dumps(after) if after else None))


router = APIRouter(prefix="/api/v1", tags=["edim"], dependencies=[Depends(require_auth)])

SOURCE_LABEL = {"APPLIED": "견적적용", "PURCHASE": "구매", "STOCK": "재고", "QUOTE": "견적"}
SOURCE_PRIORITY = ["APPLIED", "PURCHASE", "STOCK", "QUOTE"]


def _conn():
    pool = get_pool()
    if pool is None or not db_ok():
        raise HTTPException(503, detail="EDIM DB unavailable")
    return pool.connection()


def _tenant_id(cur) -> int:
    """현재 요청의 테넌트 — 세션 토큰 기준(1.2), 미인증/구형 토큰은 환경 기본 테넌트."""
    ctx = _TENANT_CTX.get()
    if ctx:
        return ctx
    cur.execute("SELECT tenant_id FROM sys_tenant WHERE tenant_code=%s", (TENANT,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(503, detail="seed not applied")
    return row[0]


# ── health ──
@router.get("/health")
def health() -> dict[str, Any]:
    return {"status": "ok", "db": db_ok()}


# ── 운영 설정 노출 — 개발서버 전용 기능 게이트 ──
DEV_MODE = os.getenv("EDIM_DEV_MODE", "") == "1"


@router.get("/config")
def app_config() -> dict[str, Any]:
    """프론트 기능 게이트 — devMode 는 개발서버(EDIM_DEV_MODE=1)에서만 true."""
    return {"devMode": DEV_MODE}


# ── 개발서버 전용: 운영자 요구사항 접수 (dev_requirement — 54-테이블 설계 외 운영 도구) ──

def _require_dev_mode() -> None:
    if not DEV_MODE:
        raise HTTPException(404, detail="개발서버 전용 기능입니다")


class DevReqCreate(BaseModel):
    title: str
    content: str = ""
    category: str = "CHANGE"     # CHANGE | BUG | FEATURE
    priority: str = "P2"         # P1 | P2 | P3
    screenId: str = ""


class DevReqPatch(BaseModel):
    status: str                  # OPEN | IN_PROGRESS | DONE | REJECTED
    resolution: str = ""


@router.get("/dev/requirements")
def dev_req_list(status: str = "") -> list[dict[str, Any]]:
    """요구사항 목록 — 전 사용자 조회 가능 (최신 우선, status 필터 옵션)."""
    _require_dev_mode()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        q = """SELECT r.req_id, r.screen_id, r.category, r.title, r.content, r.priority, r.status,
                      r.requester, COALESCE(r.resolution,''),
                      to_char(r.created_at,'YYYY-MM-DD HH24:MI'),
                      to_char(r.resolved_at,'YYYY-MM-DD HH24:MI'),
                      (SELECT count(*) FROM dev_requirement_image i WHERE i.req_id=r.req_id)
               FROM dev_requirement r WHERE r.tenant_id=%s"""
        args: list[Any] = [tid]
        if status:
            q += " AND r.status=%s"
            args.append(status)
        cur.execute(q + " ORDER BY r.req_id DESC", args)
        return [
            {"reqId": r[0], "screenId": r[1] or "", "category": r[2], "title": r[3],
             "content": r[4], "priority": r[5], "status": r[6], "requester": r[7],
             "resolution": r[8], "createdAt": r[9], "resolvedAt": r[10], "imageCount": r[11]}
            for r in cur.fetchall()
        ]


@router.post("/dev/requirements", status_code=201)
def dev_req_create(request: Request, body: DevReqCreate) -> dict[str, Any]:
    """요구사항 등록 — 로그인한 운영자 누구나 (GENERAL 포함)."""
    _require_dev_mode()
    title = body.title.strip()
    if not title:
        raise HTTPException(422, detail="제목은 필수입니다")
    if body.category not in ("CHANGE", "BUG", "FEATURE"):
        raise HTTPException(422, detail=f"분류 오류: {body.category} (CHANGE|BUG|FEATURE)")
    if body.priority not in ("P1", "P2", "P3"):
        raise HTTPException(422, detail=f"우선순위 오류: {body.priority} (P1|P2|P3)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """INSERT INTO dev_requirement (tenant_id, screen_id, category, title, content,
               priority, requester) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING req_id""",
            (tid, body.screenId.strip()[:50] or None, body.category, title[:200],
             body.content.strip(), body.priority, request.state.login))
        return {"reqId": cur.fetchone()[0], "status": "OPEN"}


@router.patch("/dev/requirements/{req_id}", dependencies=[SETUP])
def dev_req_update(req_id: int, body: DevReqPatch) -> dict[str, Any]:
    """상태 변경 (SETUP+) — 처리 라운드에서 IN_PROGRESS/DONE/REJECTED 마킹."""
    _require_dev_mode()
    if body.status not in ("OPEN", "IN_PROGRESS", "DONE", "REJECTED"):
        raise HTTPException(422, detail=f"상태 오류: {body.status}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE dev_requirement
               SET status=%s, resolution=NULLIF(%s,''),
                   resolved_at=CASE WHEN %s IN ('DONE','REJECTED') THEN now() ELSE NULL END
               WHERE tenant_id=%s AND req_id=%s RETURNING req_id""",
            (body.status, body.resolution.strip(), body.status, tid, req_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"요구사항 없음: #{req_id}")
    return {"reqId": req_id, "status": body.status}


@router.delete("/dev/requirements/{req_id}", dependencies=[SETUP])
def dev_req_delete(req_id: int) -> dict[str, Any]:
    """삭제 (SETUP+) — 첨부 이미지(MinIO 객체 포함)까지 연쇄 정리."""
    _require_dev_mode()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT file_path FROM dev_requirement_image WHERE req_id=%s", (req_id,))
        keys = [r[0] for r in cur.fetchall()]
        cur.execute("DELETE FROM dev_requirement WHERE tenant_id=%s AND req_id=%s RETURNING req_id",
                    (tid, req_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"요구사항 없음: #{req_id}")
    for key in keys:  # 이미지 행은 FK CASCADE — 객체는 여기서 제거
        try:
            storage.remove_object(key)
        except RuntimeError:
            pass  # 스토리지 불가 시 orphan (버킷 정리 배치 대상)
    return {"deleted": req_id}


# ── 요구사항 첨부 이미지 — 스크린샷 붙여넣기/업로드 (MinIO dev-req/ prefix) ──

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".webp")
IMAGE_MAX_BYTES = 10 * 1024 * 1024


@router.post("/dev/requirements/{req_id}/images", status_code=201)
async def dev_req_image_upload(req_id: int, uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """이미지 첨부 — 등록자 포함 전 사용자 (OPEN 요구에 스크린샷 추가)."""
    _require_dev_mode()
    fname = (uploadedFile.filename or "image.png").replace("/", "_")
    if not fname.lower().endswith(IMAGE_EXTS):
        raise HTTPException(422, detail=f"이미지 파일만 첨부 가능: {fname} (png/jpg/gif/webp)")
    data = await uploadedFile.read()
    if len(data) > IMAGE_MAX_BYTES:
        raise HTTPException(413, detail="10MB 초과")
    if not data:
        raise HTTPException(422, detail="빈 파일")
    ctype = uploadedFile.content_type or "image/png"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM dev_requirement WHERE tenant_id=%s AND req_id=%s", (tid, req_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"요구사항 없음: #{req_id}")
        cur.execute(
            """INSERT INTO dev_requirement_image (req_id, file_name, file_path, file_size, content_type)
               VALUES (%s,%s,%s,%s,%s) RETURNING image_id""",
            (req_id, fname[:200], "", len(data), ctype[:60]))
        image_id = cur.fetchone()[0]
        key = f"dev-req/{req_id}/{image_id}_{fname}"
        cur.execute("UPDATE dev_requirement_image SET file_path=%s WHERE image_id=%s", (key, image_id))
    try:
        storage.put_object(key, data, ctype)
    except RuntimeError:
        with _conn() as conn, conn.cursor() as cur:
            cur.execute("DELETE FROM dev_requirement_image WHERE image_id=%s", (image_id,))
        raise HTTPException(503, detail="storage unavailable")
    return {"imageId": image_id, "fileName": fname, "size": len(data)}


@router.get("/dev/requirements/{req_id}/images")
def dev_req_image_list(req_id: int) -> list[dict[str, Any]]:
    _require_dev_mode()
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """SELECT image_id, file_name, file_size, content_type
               FROM dev_requirement_image WHERE req_id=%s ORDER BY image_id""", (req_id,))
        return [{"imageId": r[0], "fileName": r[1], "size": r[2], "contentType": r[3]}
                for r in cur.fetchall()]


@router.get("/dev/requirements/images/{image_id}")
def dev_req_image_get(image_id: int) -> StreamingResponse:
    """이미지 바이트 스트림 — 프론트는 authorized fetch → blob URL 로 표시."""
    _require_dev_mode()
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT file_path, content_type FROM dev_requirement_image WHERE image_id=%s", (image_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"이미지 없음: #{image_id}")
    try:
        obj = storage.get_object(row[0])
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    return StreamingResponse(obj.stream(64 * 1024), media_type=row[1])


# ── SVC-01 i18n 번들 (REQ-N-015 · SYS-021) ──
@router.get("/i18n/{locale}")
def i18n_bundle(locale: str) -> dict[str, str]:
    """UI 리소스 번들 — KO 는 프론트 기본 문자열(폴백)이므로 빈 사전."""
    if locale == "ko":
        return {}
    if locale not in ("en", "ja", "zh"):
        raise HTTPException(404, detail=f"unsupported locale: {locale}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT field, text FROM sys_translation
               WHERE tenant_id=%s AND locale=%s AND entity_type='UI'""", (tid, locale))
        return dict(cur.fetchall())


# ── 데이터 콘텐츠 i18n — 마스터 데이터(품명·거래처·문서명) 번역 트랙 (sys_translation 재사용) ──
# entity_type → (source table, id 컬럼, 원문 컬럼, field 라벨)
DATA_I18N_SOURCES: dict[str, tuple[str, str, str, str]] = {
    "COMPANY": ("com_company", "company_id", "company_name", "name"),
    "PRODUCT": ("product_code", "product_code_id", "code_name", "codeName"),
    "DOCUMENT": ("doc_control", "doc_control_id", "title", "title"),
}


def _data_i18n_src(entity_type: str) -> tuple[str, str, str, str]:
    src = DATA_I18N_SOURCES.get(entity_type.strip().upper())
    if not src:
        raise HTTPException(404, detail=f"지원하지 않는 entity_type: {entity_type} "
                            f"(COMPANY/PRODUCT/DOCUMENT)")
    return src


@router.get("/i18n/data/{entity_type}")
def i18n_data_list(entity_type: str, locale: str) -> list[dict[str, Any]]:
    """마스터 데이터 번역 관리 — 원문 + 현재 번역(빈 값 = 미번역). SETUP 화면용."""
    if locale not in ("en", "ja", "zh"):
        raise HTTPException(422, detail=f"locale 오류: {locale} (en/ja/zh)")
    et = entity_type.strip().upper()
    table, id_col, name_col, field = _data_i18n_src(et)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT s.{id_col}, s.{name_col}, COALESCE(t.text,'')
                FROM {table} s
                LEFT JOIN sys_translation t
                  ON t.tenant_id=s.tenant_id AND t.entity_type=%s AND t.entity_id=s.{id_col}
                     AND t.field=%s AND t.locale=%s
                WHERE s.tenant_id=%s ORDER BY s.{id_col}""",
            (et, field, locale, tid))
        return [{"entityId": r[0], "source": r[1] or "", "value": r[2]} for r in cur.fetchall()]


@router.get("/i18n/data/{entity_type}/map")
def i18n_data_map(entity_type: str, locale: str) -> dict[str, str]:
    """번역 오버레이 맵 — {entityId(str): 번역문}. 비어있지 않은 것만. 목록 렌더 오버레이용."""
    if locale not in ("en", "ja", "zh"):
        return {}
    et = entity_type.strip().upper()
    _table, _id_col, _name_col, field = _data_i18n_src(et)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT entity_id, text FROM sys_translation
               WHERE tenant_id=%s AND entity_type=%s AND field=%s AND locale=%s AND text<>''""",
            (tid, et, field, locale))
        return {str(r[0]): r[1] for r in cur.fetchall()}


class DataTransUpsert(BaseModel):
    entityType: str
    entityId: int
    locale: str
    value: str


@router.put("/i18n/data", dependencies=[SETUP])
def i18n_data_upsert(request: Request, body: DataTransUpsert) -> dict[str, Any]:
    """마스터 데이터 번역 저장/삭제 (빈 값 = 삭제). sys_translation upsert."""
    et = body.entityType.strip().upper()
    _table, _id_col, _name_col, field = _data_i18n_src(et)
    if body.locale not in ("en", "ja", "zh"):
        raise HTTPException(422, detail=f"locale 오류: {body.locale}")
    val = body.value.strip()[:1000]
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if not val:
            cur.execute(
                """DELETE FROM sys_translation WHERE tenant_id=%s AND locale=%s
                   AND entity_type=%s AND entity_id=%s AND field=%s""",
                (tid, body.locale, et, body.entityId, field))
            _audit(cur, tid, "sys_translation", body.entityId, "DATA_I18N_DEL",
                   request.state.user_id, {"entityType": et, "locale": body.locale})
            return {"entityType": et, "entityId": body.entityId, "locale": body.locale, "deleted": True}
        cur.execute(
            """INSERT INTO sys_translation (tenant_id, locale, entity_type, entity_id, field, text,
               created_by, updated_by, updated_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,now())
               ON CONFLICT (tenant_id, locale, entity_type, entity_id, field)
               DO UPDATE SET text=EXCLUDED.text, updated_by=EXCLUDED.updated_by, updated_at=now()""",
            (tid, body.locale, et, body.entityId, field, val, request.state.login, request.state.login))
        _audit(cur, tid, "sys_translation", body.entityId, "DATA_I18N", request.state.user_id,
               {"entityType": et, "locale": body.locale, "value": val})
    return {"entityType": et, "entityId": body.entityId, "locale": body.locale, "value": val}


@router.get("/i18n/data/{entity_type}/export.xlsx")
def i18n_data_export(entity_type: str) -> "Response":
    """데이터 번역 일괄 Export (메뉴정의서 다국어 P2) — 원문 + en/ja/zh 3로케일 한 시트 (미번역=빈 칸)."""
    et = entity_type.strip().upper()
    table, id_col, name_col, field = _data_i18n_src(et)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(f"SELECT {id_col}, {name_col} FROM {table} WHERE tenant_id=%s ORDER BY {id_col}", (tid,))
        src = cur.fetchall()
        cur.execute(
            """SELECT entity_id, locale, text FROM sys_translation
               WHERE tenant_id=%s AND entity_type=%s AND field=%s""", (tid, et, field))
        tr = {(r[0], r[1]): r[2] for r in cur.fetchall()}
    rows = [[r[0], r[1] or "", tr.get((r[0], "en"), ""), tr.get((r[0], "ja"), ""), tr.get((r[0], "zh"), "")]
            for r in src]
    return _xlsx_response(f"번역-{et}", ["ID", "원문", "en", "ja", "zh"], rows, f"i18n_{et.lower()}")


@router.post("/i18n/data/{entity_type}/import-excel", dependencies=[SETUP])
async def i18n_data_import(entity_type: str, request: Request,
                           uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """데이터 번역 일괄 Import — 헤더 ID·en/ja/zh. 비어있지 않은 셀만 업서트 (빈 칸=변경 없음, 삭제는 화면에서)."""
    et = entity_type.strip().upper()
    table, id_col, _name_col, field = _data_i18n_src(et)
    ws, idx = _load_ws(await uploadedFile.read(), ["ID"])
    locs = [lc for lc in ("en", "ja", "zh") if lc in idx]
    if not locs:
        raise HTTPException(422, detail="로케일 컬럼(en/ja/zh)이 없습니다")
    upserted, rejected = 0, []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(f"SELECT {id_col} FROM {table} WHERE tenant_id=%s", (tid,))
        valid = {r[0] for r in cur.fetchall()}
        for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            raw = row[idx["ID"]].value
            try:
                eid = int(raw)
            except (TypeError, ValueError):
                continue
            if eid not in valid:
                rejected.append(f"{r_i}행: 미존재 ID {eid}")
                continue
            for lc in locs:
                v = row[idx[lc]].value
                val = str(v).strip()[:1000] if v is not None else ""
                if not val:
                    continue
                cur.execute(
                    """INSERT INTO sys_translation (tenant_id, locale, entity_type, entity_id, field, text,
                       created_by, updated_by, updated_at)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,now())
                       ON CONFLICT (tenant_id, locale, entity_type, entity_id, field)
                       DO UPDATE SET text=EXCLUDED.text, updated_by=EXCLUDED.updated_by, updated_at=now()""",
                    (tid, lc, et, eid, field, val, request.state.login, request.state.login))
                upserted += 1
        _audit(cur, tid, "sys_translation", 0, "DATA_I18N_IMPORT", request.state.user_id,
               {"entityType": et, "upserted": upserted, "rejected": len(rejected)})
    return {"upserted": upserted, "rejected": rejected, "rejectedCount": len(rejected)}


# ── SVC-01 Auth ──
MAX_LOGIN_FAILS = 5   # 연속 실패 → 자동 LOCKED (B8, SEC-002)

# C10 — 로그인 레이트리밋 (무차별 대입 완화; 계정 잠금 이전 단계 속도 제한, 인메모리)
from collections import deque  # noqa: E402

LOGIN_RATE_MAX = 30       # WINDOW 초 내 최대 시도 (계정 잠금 5회가 1차 방어, 레이트리밋은 2차)
LOGIN_RATE_WINDOW = 60
_login_hits: dict[str, deque] = {}


def _login_rate_check(key: str) -> None:
    now = time.time()
    dq = _login_hits.setdefault(key, deque(maxlen=64))
    while dq and now - dq[0] > LOGIN_RATE_WINDOW:
        dq.popleft()
    if len(dq) >= LOGIN_RATE_MAX:
        raise HTTPException(429, detail="로그인 시도 과다 — 잠시 후 다시 시도하십시오 (rate limit)")
    dq.append(now)


class LoginRequest(BaseModel):
    userId: str
    password: str
    ttlSeconds: int | None = None   # 토큰 수명 단축 (갱신 검증용, 60s~8h 클램프)
    otp: str = ""                   # 트리아지 #10 — MFA 활성 사용자 2단계 코드
    tenantCode: str = ""            # 1.2 — 동일 사번이 여러 고객사에 있을 때 소속 지정


def _totp_code(secret_b32: str, at: float | None = None, step: int = 30) -> str:
    """RFC 6238 TOTP (SHA1·6자리·30s) — 표준 인증 앱 호환."""
    import base64
    import hmac as _hmac
    import struct
    import time as _time
    key = base64.b32decode(secret_b32)
    counter = int((at if at is not None else _time.time()) // step)
    h = _hmac.new(key, struct.pack(">Q", counter), hashlib.sha1).digest()
    o = h[-1] & 0x0F
    return f"{(int.from_bytes(h[o:o + 4], 'big') & 0x7FFFFFFF) % 1_000_000:06d}"


def _totp_verify(secret_b32: str, code: str) -> bool:
    import time as _time
    c = (code or "").strip()
    return len(c) == 6 and any(_totp_code(secret_b32, _time.time() + d * 30) == c for d in (-1, 0, 1))


@router.post("/auth/login")
def login(body: LoginRequest) -> dict[str, Any]:
    login_id = body.userId.strip()
    _login_rate_check(login_id or "-")   # C10 — 속도 제한 (429)
    with _conn() as conn, conn.cursor() as cur:
        # 1.2 — 소속 테넌트 해석: tenantCode 지정 > 사번 유일 소속 > 다중 소속이면 지정 요구
        tcode = (body.tenantCode or "").strip()
        if tcode:
            cur.execute("SELECT tenant_id FROM sys_tenant WHERE tenant_code=%s", (tcode,))
            trow = cur.fetchone()
            if not trow:
                raise HTTPException(401, detail="사번 또는 비밀번호가 올바르지 않습니다")
            tid = trow[0]
        else:
            cur.execute(
                """SELECT u.tenant_id, t.tenant_code FROM sys_user u
                   JOIN sys_tenant t ON t.tenant_id=u.tenant_id
                   WHERE u.login_id=%s AND u.status='ACTIVE'""", (login_id,))
            owners = cur.fetchall()
            if len(owners) > 1:
                raise HTTPException(
                    409, detail="여러 고객사에 동일 사번이 있습니다 — 테넌트 코드를 지정하십시오: "
                                + ", ".join(o[1] for o in owners))
            tid = owners[0][0] if owners else _tenant_id(cur)
        # 1.3 — 계약 게이트: 중지/해지 고객사는 로그인 차단
        cur.execute("SELECT status, tenant_name FROM sys_tenant WHERE tenant_id=%s", (tid,))
        trow0 = cur.fetchone()
        if trow0 and (trow0[0] or "ACTIVE").upper() != "ACTIVE":
            raise HTTPException(
                403, detail=f"고객사 이용이 중지되었습니다 ({trow0[1]} — {trow0[0]}). 담당자에게 문의하십시오")
        cur.execute(
            """SELECT user_id, user_name, department, user_level, password_hash, status,
                      totp_secret, mfa_enabled
               FROM sys_user WHERE tenant_id=%s AND login_id=%s""", (tid, login_id))
        row = cur.fetchone()
        if not row:
            # 미존재 사용자 — sys_history.actor_id FK 로 기록 불가, 정보 노출 방지 위해 동일 메시지
            raise HTTPException(401, detail="사번 또는 비밀번호가 올바르지 않습니다")
        uid, status = row[0], row[5]
        if status == "LOCKED":
            _audit(cur, tid, "sys_user", uid, "LOGIN_DENY", uid, {"reason": "LOCKED"})
            raise HTTPException(403, detail="계정 잠금(LOCKED) — 관리자에게 잠금 해제를 요청하십시오")
        if status != "ACTIVE":
            raise HTTPException(401, detail="사번 또는 비밀번호가 올바르지 않습니다")
        if hashlib.sha256(body.password.encode()).hexdigest() != row[4]:
            _audit(cur, tid, "sys_user", uid, "LOGIN_FAIL", uid, {"login": login_id})
            # 마지막 성공/잠금해제 이후 연속 실패 수 — 도달 시 자동 잠금
            cur.execute(
                """SELECT count(*) FROM sys_history
                   WHERE tenant_id=%s AND target_table='sys_user' AND target_id=%s
                     AND action='LOGIN_FAIL'
                     AND acted_at > COALESCE(
                       (SELECT max(acted_at) FROM sys_history
                        WHERE tenant_id=%s AND target_table='sys_user' AND target_id=%s
                          AND action IN ('LOGIN_OK','UNLOCK')), '-infinity')""",
                (tid, uid, tid, uid))
            fails = cur.fetchone()[0]
            if fails >= MAX_LOGIN_FAILS:
                cur.execute(
                    "UPDATE sys_user SET status='LOCKED', updated_at=now() WHERE user_id=%s", (uid,))
                _audit(cur, tid, "sys_user", uid, "LOCK", uid,
                       {"reason": f"로그인 {fails}회 연속 실패 자동 잠금"})
                raise HTTPException(
                    403, detail=f"로그인 {MAX_LOGIN_FAILS}회 실패 — 계정이 잠겼습니다 (관리자 잠금 해제 필요)")
            raise HTTPException(
                401, detail=f"사번 또는 비밀번호가 올바르지 않습니다 (실패 {fails}/{MAX_LOGIN_FAILS})")
        # 트리아지 #10 — MFA 활성 사용자 2단계 (비활성 사용자·기존 흐름 무영향)
        if row[7]:
            if not body.otp.strip():
                return {"mfaRequired": True}
            if not _totp_verify(row[6] or "", body.otp):
                _audit(cur, tid, "sys_user", uid, "LOGIN_MFA_FAIL", uid, {"login": login_id})
                raise HTTPException(401, detail="OTP 코드가 올바르지 않습니다 (인증 앱 확인)")
        _audit(cur, tid, "sys_user", uid, "LOGIN_OK", uid, {"login": login_id})
        cur.execute("SELECT tenant_code, tenant_name FROM sys_tenant WHERE tenant_id=%s", (tid,))
        trow = cur.fetchone() or (TENANT, TENANT)
    ttl = max(60, min(int(body.ttlSeconds or TOKEN_TTL), TOKEN_TTL))
    return {
        "token": _issue_token(login_id, ttl, tenant_id=tid),
        "user": {
            "userId": login_id, "name": row[1], "department": row[2] or "",
            "userLevel": row[3], "tenantId": trow[0], "tenantName": trow[1],
        },
    }


class PasswordChangeRequest(BaseModel):
    currentPassword: str
    newPassword: str


class MfaCode(BaseModel):
    code: str = ""


@router.get("/users/me/mfa")
def mfa_status(request: Request) -> dict[str, Any]:
    """MFA 상태 (트리아지 #10) — enabled / pending(시크릿 발급됨·미활성)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT totp_secret, mfa_enabled FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, request.state.login))
        row = cur.fetchone()
    return {"enabled": bool(row and row[1]), "pending": bool(row and row[0] and not row[1])}


@router.post("/users/me/mfa/setup")
def mfa_setup(request: Request) -> dict[str, Any]:
    """MFA 설정 시작 — TOTP 시크릿 발급 (활성화는 enable 에서 코드 검증 후, 잠금 방지 2단계)."""
    import base64
    import secrets as _secrets
    secret = base64.b32encode(_secrets.token_bytes(20)).decode()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_user SET totp_secret=%s, mfa_enabled=false, updated_at=now()
               WHERE tenant_id=%s AND login_id=%s RETURNING user_id""",
            (secret, tid, request.state.login))
        uid = cur.fetchone()[0]
        _audit(cur, tid, "sys_user", uid, "MFA_SETUP", uid, {})
    return {"secret": secret, "issuer": "EDIM", "account": request.state.login,
            "note": "인증 앱(TOTP)에 시크릿 수동 등록 후 enable 로 코드 검증"}


@router.post("/users/me/mfa/enable")
def mfa_enable(request: Request, body: MfaCode) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id, totp_secret FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, request.state.login))
        uid, secret = cur.fetchone()
        if not secret:
            raise HTTPException(422, detail="setup 먼저 실행하십시오")
        if not _totp_verify(secret, body.code):
            raise HTTPException(422, detail="OTP 코드가 올바르지 않습니다 — 인증 앱 시간 확인")
        cur.execute("UPDATE sys_user SET mfa_enabled=true, updated_at=now() WHERE user_id=%s", (uid,))
        _audit(cur, tid, "sys_user", uid, "MFA_ENABLE", uid, {})
    return {"enabled": True}


@router.post("/users/me/mfa/disable")
def mfa_disable(request: Request, body: MfaCode) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id, totp_secret, mfa_enabled FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, request.state.login))
        uid, secret, enabled = cur.fetchone()
        if enabled and not _totp_verify(secret or "", body.code):
            raise HTTPException(422, detail="해제하려면 현재 OTP 코드가 필요합니다")
        cur.execute("UPDATE sys_user SET totp_secret=NULL, mfa_enabled=false, updated_at=now() "
                    "WHERE user_id=%s", (uid,))
        _audit(cur, tid, "sys_user", uid, "MFA_DISABLE", uid, {})
    return {"enabled": False}


@router.put("/users/me/password")
def change_my_password(request: Request, body: PasswordChangeRequest) -> dict[str, Any]:
    """비밀번호 변경 (B8, SEC-001) — 현재 비밀번호 검증 후 교체 + 감사."""
    new = body.newPassword
    if len(new) < 4:
        raise HTTPException(422, detail="새 비밀번호는 4자 이상이어야 합니다")
    if new == body.currentPassword:
        raise HTTPException(422, detail="새 비밀번호가 현재 비밀번호와 같습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        uid = request.state.user_id
        cur.execute("SELECT password_hash FROM sys_user WHERE user_id=%s", (uid,))
        row = cur.fetchone()
        if hashlib.sha256(body.currentPassword.encode()).hexdigest() != row[0]:
            _audit(cur, tid, "sys_user", uid, "PW_CHANGE_FAIL", uid,
                   {"reason": "현재 비밀번호 불일치"})
            raise HTTPException(403, detail="현재 비밀번호가 올바르지 않습니다")
        cur.execute(
            """UPDATE sys_user SET password_hash=%s, updated_by=%s, updated_at=now()
               WHERE user_id=%s""",
            (hashlib.sha256(new.encode()).hexdigest(), request.state.login, uid))
        _audit(cur, tid, "sys_user", uid, "PW_CHANGE", uid)   # 비밀번호 자체는 기록하지 않음
    return {"login": request.state.login, "changed": True}


# ── SVC-03 Code ──
@router.get("/codes/groups/{group}/slots")
def group_slots(group: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT ci.item_slot, ci.item_name,
                      COALESCE(array_agg(civ.value_code ORDER BY civ.sort_order)
                               FILTER (WHERE civ.approval_status='APPROVED'), '{}'),
                      COALESCE(array_agg(civ.value_code ORDER BY civ.sort_order)
                               FILTER (WHERE civ.value_code IS NOT NULL), '{}'),
                      COALESCE(bool_or(civ.approval_status <> 'APPROVED'), false)
               FROM code_group cg
               JOIN code_item ci ON ci.group_id=cg.group_id
               LEFT JOIN code_item_value civ ON civ.item_id=ci.item_id
               WHERE cg.tenant_id=%s AND cg.group_code=%s
               GROUP BY ci.item_slot, ci.item_name, ci.sort_order
               ORDER BY ci.item_slot""", (tid, group))
        return [
            {"slot": r[0], "label": r[1], "values": list(r[2]),
             "allValues": list(r[3]), "count": len(r[3]),
             "status": "PENDING" if r[4] else "APPROVED", "approved": not r[4]}
            for r in cur.fetchall()
        ]


class GroupCreate(BaseModel):
    groupCode: str
    groupName: str
    groupType: str = "SPECIFICATION"
    hierarchyAddress: str = ""


@router.get("/codes/groups")
def list_groups() -> list[dict[str, Any]]:
    """코드 그룹 목록 (C2 — S-1-1 상위 그룹 관리)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT cg.group_code, cg.group_name, cg.group_type, cg.hierarchy_address,
                      cg.approval_status, count(ci.item_id)
               FROM code_group cg LEFT JOIN code_item ci ON ci.group_id=cg.group_id
               WHERE cg.tenant_id=%s GROUP BY cg.group_id ORDER BY cg.group_code""", (tid,))
        return [{"groupCode": r[0], "groupName": r[1], "groupType": r[2],
                 "hierarchyAddress": r[3], "status": r[4], "slotCount": r[5]}
                for r in cur.fetchall()]


@router.post("/codes/groups", status_code=201, dependencies=[SETUP])
def create_group(request: Request, body: GroupCreate) -> dict[str, Any]:
    """코드 그룹 등록 (C2) — DRAFT 로 생성."""
    gt = body.groupType.strip().upper()
    if gt not in ("SPECIFICATION", "RAW_MATERIAL", "GPI", "PRODUCT"):
        raise HTTPException(422, detail="그룹 유형 오류 (SPECIFICATION/RAW_MATERIAL/GPI/PRODUCT)")
    code = body.groupCode.strip().upper()[:20]
    if not code or not body.groupName.strip():
        raise HTTPException(422, detail="필수 — 그룹 코드·이름")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM code_group WHERE tenant_id=%s AND group_code=%s", (tid, code))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 그룹 코드 {code}")
        addr = body.hierarchyAddress.strip() or f"/C/{code}"
        # 0.9 — 고아 주소 쓰기 차단 (#25 계열): 트리 노드 주소이거나 그 하위(leaf 확장)여야 함.
        # 0.8 감사에서 구세대 자유형 주소 8건 수리 — 재발을 쓰기 시점에 봉쇄.
        cur.execute(
            """SELECT 1 FROM sys_hierarchy WHERE tenant_id=%s
               AND (%s = address OR %s LIKE address || '/%%') LIMIT 1""", (tid, addr, addr))
        if not cur.fetchone():
            raise HTTPException(
                422, detail=f"Hierarchy 주소가 트리에 없음: {addr} — M-3-1 노드(또는 그 하위) 주소를 사용하십시오")
        cur.execute(
            """INSERT INTO code_group (tenant_id, group_code, group_name, group_type,
               hierarchy_address, approval_status, created_by)
               VALUES (%s,%s,%s,%s,%s,'DRAFT',%s) RETURNING group_id""",
            (tid, code, body.groupName.strip()[:100], gt, addr, str(request.state.user_id)))
        gid = cur.fetchone()[0]
        _audit(cur, tid, "code_group", gid, "CREATE", request.state.user_id, {"groupCode": code})
    return {"groupCode": code, "status": "DRAFT"}


# ── G3 제품 코드 마스터 CRUD (수동 생성/수정/비활성) ──
# product_code 참조 테이블 (삭제 가드) — (table, fk_column)
_PC_REFS = [
    ("code_relationship", "mother_code_id"), ("code_relationship", "child_code_id"),
    ("cst_price", "product_code_id"), ("cpq_selection", "product_code_id"),
    ("cpq_selection_item", "child_code_id"), ("dwg_dimension", "product_code_id"),
    ("erp_work_process", "product_code_id"),
    ("prt_part", "product_code_id"), ("prt_supplier_code_map", "product_code_id"),
    ("arrangement_component", "product_code_id"),
]
# product_code_item(조합 구성) 은 소유 자식 행 — 외부 참조가 아니므로 삭제 시 함께 정리한다(#28).
_PC_OWNED = ["product_code_item"]


@router.get("/codes/products")
def list_products(status: str = "") -> list[dict[str, Any]]:
    """제품 코드 마스터 목록 — main_code·code_name·그룹·상태·참조 수."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if status.strip() and status != "ALL":
            clause = " AND pc.approval_status=%s"
            params.append(status.strip())
        cur.execute(
            f"""SELECT pc.product_code_id, pc.main_code, pc.code_name, cg.group_code,
                       pc.approval_status, to_char(pc.created_at,'YYYY-MM-DD'), pc.origin,
                       (SELECT count(*) FROM code_relationship cr
                          WHERE cr.mother_code_id=pc.product_code_id OR cr.child_code_id=pc.product_code_id)
                       + (SELECT count(*) FROM cst_price p WHERE p.product_code_id=pc.product_code_id)
                       + (SELECT count(*) FROM cpq_selection s WHERE s.product_code_id=pc.product_code_id)
                FROM product_code pc JOIN code_group cg ON cg.group_id=pc.group_id
                WHERE pc.tenant_id=%s{clause} ORDER BY pc.main_code""", tuple(params))
        return [{"productCodeId": r[0], "mainCode": r[1], "codeName": r[2], "groupCode": r[3],
                 "status": r[4], "createdAt": r[5], "origin": r[6] or "MANUAL",
                 "refs": int(r[7])} for r in cur.fetchall()]


class ProductCreate(BaseModel):
    mainCode: str
    codeName: str
    groupCode: str


@router.post("/codes/products", status_code=201, dependencies=[SETUP])
def create_product(request: Request, body: ProductCreate) -> dict[str, Any]:
    """제품 코드 수동 생성 — DRAFT (중복 main_code 409·그룹 없음 422).

    #28 불변식: Slot(code_item) 이 정의된 그룹은 자유텍스트 생성을 거부한다.
    그런 그룹의 제품 코드는 승인된 Sub Code 조합(POST /codes/products/build)으로만 만들어진다.
    Slot 이 없는 그룹(마스터성·임시 코드군)은 종전대로 수기 등록을 허용한다."""
    main = body.mainCode.strip()[:50]
    if not main or not body.codeName.strip():
        raise HTTPException(422, detail="필수 — 코드·코드명")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT group_id, COALESCE(hierarchy_address,'') FROM code_group WHERE tenant_id=%s AND group_code=%s",
                    (tid, body.groupCode.strip()))
        grp = cur.fetchone()
        if not grp:
            raise HTTPException(422, detail=f"그룹 없음: {body.groupCode}")
        cur.execute("SELECT count(*) FROM code_item WHERE group_id=%s", (grp[0],))
        if int(cur.fetchone()[0]) > 0:
            raise HTTPException(
                422,
                detail=f"자유텍스트 등록 불가 — 그룹 {body.groupCode.strip()} 은 Sub Code Slot 이 정의되어 "
                       "있습니다. 승인된 Sub Code 조합으로 생성하십시오 (제품 코드 조합 생성, #28)")
        cur.execute("SELECT 1 FROM product_code WHERE tenant_id=%s AND main_code=%s", (tid, main))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 코드 {main}")
        addr = f"{grp[1]}/{main}" if grp[1] else f"/C/{body.groupCode.strip()}/{main}"
        cur.execute(
            """INSERT INTO product_code (tenant_id, main_code, group_id, code_name, hierarchy_address, approval_status, created_by)
               VALUES (%s,%s,%s,%s,%s,'DRAFT',%s) RETURNING product_code_id""",
            (tid, main, grp[0], body.codeName.strip()[:200], addr[:200], str(request.state.user_id)))
        pid = cur.fetchone()[0]
        _audit(cur, tid, "product_code", pid, "CREATE", request.state.user_id, {"mainCode": main})
    return {"productCodeId": pid, "mainCode": main, "status": "DRAFT"}


# ── #28 Product Code Builder — 승인된 Sub Code 조합으로만 생성 (핵심 불변식) ──
def _combo_hash(group_code: str, pairs: list[tuple[str, str]]) -> str:
    """조합 해시 — 그룹 + (slot, valueCode) 정렬쌍의 정규화 SHA-256.
    Snapshot(_snapshot_checksum)과 같은 규약: 키 순서·공백 무관, 조합이 같으면 해시가 같다."""
    return _snapshot_checksum({"group": group_code, "slots": sorted(pairs)})


def _compose_main_code(group_code: str, pairs: list[tuple[str, str]]) -> str:
    """조합 → 제품 코드 문자열. 현행 코드 관례(`KDP 1-21-13-15`)와 동일한 표기."""
    return f"{group_code} {'-'.join(v for _, v in sorted(pairs))}"


@router.get("/codes/products/builder")
def product_builder(group: str) -> dict[str, Any]:
    """조합 생성용 슬롯·선택지 — 그룹의 Slot 별 **승인된** Sub Code 값만 노출.

    승인된 값이 하나도 없는 Slot 은 blocked 로 표시한다(그 그룹은 조합 생성 불가 — 정직한 사유 제공)."""
    g = group.strip()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT group_id, group_name FROM code_group WHERE tenant_id=%s AND group_code=%s", (tid, g))
        grp = cur.fetchone()
        if not grp:
            raise HTTPException(404, detail=f"그룹 없음: {g}")
        cur.execute(
            """SELECT ci.item_id, ci.item_slot, ci.item_name, civ.value_id, civ.value_code,
                      COALESCE(civ.value_name,''), civ.revision_no, civ.approval_status
               FROM code_item ci
               LEFT JOIN code_item_value civ ON civ.item_id=ci.item_id
               WHERE ci.group_id=%s ORDER BY ci.item_slot, civ.sort_order, civ.value_code""", (grp[0],))
        slots: dict[str, dict[str, Any]] = {}
        for item_id, slot, name, vid, vcode, vname, rev, status in cur.fetchall():
            s = slots.setdefault(slot, {"slot": slot, "itemId": item_id, "label": name,
                                        "values": [], "pending": 0})
            if vid is None:
                continue
            if status == "APPROVED":
                s["values"].append({"valueId": vid, "valueCode": vcode, "valueName": vname,
                                    "revisionNo": int(rev or 1)})
            else:
                s["pending"] += 1
        rows = [dict(v, blocked=not v["values"]) for v in slots.values()]
        return {"groupCode": g, "groupName": grp[1], "slots": rows,
                "buildable": bool(rows) and all(not r["blocked"] for r in rows)}


class ProductBuild(BaseModel):
    groupCode: str
    codeName: str = ""
    selections: dict[str, str]      # {slot: valueCode}


@router.post("/codes/products/build", status_code=201, dependencies=[SETUP])
def build_product(request: Request, body: ProductBuild) -> dict[str, Any]:
    """승인된 Sub Code 조합으로 제품 코드 생성 (#28).

    - 그룹의 모든 Slot 에 선택이 있어야 한다 (누락 Slot 명시 422)
    - 선택된 값은 그 Slot 의 **APPROVED** 값이어야 한다 (미승인 값 명시 422)
    - main_code 는 조합에서 파생 — 자유텍스트 없음
    - 같은 조합은 재생성 불가 (409, 기존 코드 회신)"""
    g = body.groupCode.strip()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT group_id, COALESCE(hierarchy_address,'') FROM code_group "
                    "WHERE tenant_id=%s AND group_code=%s", (tid, g))
        grp = cur.fetchone()
        if not grp:
            raise HTTPException(422, detail=f"그룹 없음: {g}")
        cur.execute(
            """SELECT ci.item_slot, ci.item_id, civ.value_id, civ.value_code, civ.revision_no,
                      civ.approval_status
               FROM code_item ci LEFT JOIN code_item_value civ ON civ.item_id=ci.item_id
               WHERE ci.group_id=%s""", (grp[0],))
        catalog: dict[str, dict[str, Any]] = {}
        for slot, item_id, vid, vcode, rev, status in cur.fetchall():
            c = catalog.setdefault(slot, {"itemId": item_id, "values": {}})
            if vid is not None:
                c["values"][vcode] = {"valueId": vid, "revisionNo": int(rev or 1), "status": status}
        if not catalog:
            raise HTTPException(422, detail=f"그룹 {g} 에 Sub Code Slot 이 없습니다 — S-1-1 에서 먼저 등록하십시오")

        picked = {k.strip().upper(): str(v).strip() for k, v in body.selections.items() if str(v).strip()}
        unknown = sorted(set(picked) - set(catalog))
        if unknown:
            raise HTTPException(422, detail=f"그룹 {g} 에 없는 Slot: {', '.join(unknown)}")
        missing = sorted(set(catalog) - set(picked))
        if missing:
            raise HTTPException(422, detail=f"모든 Slot 을 선택해야 합니다 — 누락: {', '.join(missing)}")

        pairs: list[tuple[str, str]] = []
        items: list[dict[str, Any]] = []
        for slot in sorted(catalog):
            vcode = picked[slot]
            v = catalog[slot]["values"].get(vcode)
            if not v:
                raise HTTPException(422, detail=f"Slot {slot} 에 없는 값: {vcode}")
            if v["status"] != "APPROVED":
                raise HTTPException(
                    422, detail=f"승인되지 않은 Sub Code — {slot}={vcode} ({v['status']}). "
                                "승인 후 조합할 수 있습니다 (#28)")
            pairs.append((slot, vcode))
            items.append({"slot": slot, "itemId": catalog[slot]["itemId"], "valueId": v["valueId"],
                          "valueCode": vcode, "revisionNo": v["revisionNo"]})

        chash = _combo_hash(g, pairs)
        cur.execute("SELECT main_code FROM product_code WHERE tenant_id=%s AND combo_hash=%s", (tid, chash))
        dup = cur.fetchone()
        if dup:
            raise HTTPException(409, detail=f"동일 조합의 제품 코드가 이미 있습니다: {dup[0]}")
        main = _compose_main_code(g, pairs)
        if len(main) > 50:
            # 잘라내면 서로 다른 조합이 같은 코드가 된다 — 조용한 절단 대신 정직한 거부.
            raise HTTPException(
                422, detail=f"파생 코드가 50자를 초과합니다({len(main)}자): {main} — Slot 값 표기를 줄이십시오")
        cur.execute("SELECT 1 FROM product_code WHERE tenant_id=%s AND main_code=%s", (tid, main))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 코드 {main}")

        combo = {"group": g, "slots": [{"slot": s, "valueCode": vc,
                                        "revisionNo": next(i["revisionNo"] for i in items if i["slot"] == s)}
                                       for s, vc in sorted(pairs)]}
        addr = (f"{grp[1]}/{main}" if grp[1] else f"/C/{g}/{main}")[:200]
        cur.execute(
            """INSERT INTO product_code (tenant_id, main_code, group_id, code_name, hierarchy_address,
               approval_status, combo, combo_hash, origin, created_by)
               VALUES (%s,%s,%s,%s,%s,'DRAFT',%s,%s,'COMPOSED',%s) RETURNING product_code_id""",
            (tid, main, grp[0], (body.codeName.strip() or main)[:200], addr,
             json.dumps(combo, ensure_ascii=False), chash, str(request.state.user_id)))
        pid = cur.fetchone()[0]
        for n, it in enumerate(items):
            cur.execute(
                """INSERT INTO product_code_item (product_code_id, item_slot, source_item_id,
                   is_required, sort_order, value_id, value_code, revision_no)
                   VALUES (%s,%s,%s,true,%s,%s,%s,%s)""",
                (pid, it["slot"], it["itemId"], n, it["valueId"], it["valueCode"], it["revisionNo"]))
        _audit(cur, tid, "product_code", pid, "COMPOSE", request.state.user_id,
               {"mainCode": main, "group": g, "comboHash": chash[:12],
                "slots": {s: vc for s, vc in pairs}})
    return {"productCodeId": pid, "mainCode": main, "status": "DRAFT",
            "comboHash": chash, "origin": "COMPOSED", "slots": items}


@router.get("/codes/products/{product_code_id}/composition")
def product_composition(product_code_id: int) -> dict[str, Any]:
    """제품 코드 조합 상세 (#28) — 고정된 Slot·값·Revision + 해시 무결성/Revision drift.

    조합 자체는 불변이다. 원본 Sub Code 값이 이후 개정되면 revDrift 로 드러날 뿐 조합은 바뀌지 않는다."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT pc.main_code, pc.code_name, pc.approval_status, cg.group_code,
                      pc.combo, pc.combo_hash, pc.origin
               FROM product_code pc JOIN code_group cg ON cg.group_id=pc.group_id
               WHERE pc.tenant_id=%s AND pc.product_code_id=%s""", (tid, product_code_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"코드 없음: {product_code_id}")
        cur.execute(
            """SELECT pci.item_slot, ci.item_name, pci.value_code, pci.revision_no,
                      civ.revision_no, civ.approval_status, COALESCE(civ.value_name,'')
               FROM product_code_item pci
               JOIN code_item ci ON ci.item_id=pci.source_item_id
               LEFT JOIN code_item_value civ ON civ.value_id=pci.value_id
               WHERE pci.product_code_id=%s ORDER BY pci.sort_order, pci.item_slot""",
            (product_code_id,))
        slots = [
            {"slot": r[0], "label": r[1], "valueCode": r[2], "valueName": r[6],
             "boundRevision": r[3], "currentRevision": r[4], "currentStatus": r[5],
             "revDrift": bool(r[3] and r[4] and int(r[3]) != int(r[4]))}
            for r in cur.fetchall()
        ]
        chash = row[5]
        intact = None
        if chash and row[4]:
            pairs = [(s["slot"], s["valueCode"]) for s in slots if s["valueCode"]]
            intact = _combo_hash(row[3], pairs) == chash
        return {"productCodeId": product_code_id, "mainCode": row[0], "codeName": row[1],
                "status": row[2], "groupCode": row[3], "origin": row[6],
                "comboHash": chash, "intact": intact, "slots": slots,
                "drift": [s["slot"] for s in slots if s["revDrift"]]}


class ProductPatch(BaseModel):
    codeName: str | None = None
    status: str | None = None   # DRAFT | APPROVED | INACTIVE


@router.patch("/codes/products/{product_code_id}", dependencies=[SETUP])
def patch_product(product_code_id: int, request: Request, body: ProductPatch) -> dict[str, Any]:
    """제품 코드 수정 — 코드명·상태(DRAFT/APPROVED/INACTIVE=비활성)."""
    sets: list[str] = []
    params: list[Any] = []
    if body.codeName is not None and body.codeName.strip():
        sets.append("code_name=%s"); params.append(body.codeName.strip()[:200])
    if body.status is not None:
        st = body.status.strip().upper()
        if st not in ("DRAFT", "APPROVED", "INACTIVE"):
            raise HTTPException(422, detail="상태는 DRAFT/APPROVED/INACTIVE")
        sets.append("approval_status=%s"); params.append(st)
    if not sets:
        raise HTTPException(422, detail="변경할 값이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params += [tid, product_code_id]
        cur.execute(
            f"UPDATE product_code SET {', '.join(sets)} WHERE tenant_id=%s AND product_code_id=%s "
            "RETURNING main_code, code_name, approval_status", tuple(params))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail="코드 없음")
        _audit(cur, tid, "product_code", product_code_id, "UPDATE", request.state.user_id,
               {"codeName": body.codeName, "status": body.status})
    return {"mainCode": row[0], "codeName": row[1], "status": row[2]}


@router.delete("/codes/products/{product_code_id}", dependencies=[SETUP])
def delete_product(product_code_id: int, request: Request) -> dict[str, Any]:
    """제품 코드 삭제 — 참조 있으면 409(비활성 권장)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for table, col in _PC_REFS:
            cur.execute(f"SELECT 1 FROM {table} WHERE {col}=%s LIMIT 1", (product_code_id,))
            if cur.fetchone():
                raise HTTPException(409, detail=f"참조 있어 삭제 불가({table}) — 비활성(INACTIVE) 처리 권장")
        for owned in _PC_OWNED:
            cur.execute(f"DELETE FROM {owned} WHERE product_code_id=%s", (product_code_id,))
        cur.execute("DELETE FROM product_code WHERE tenant_id=%s AND product_code_id=%s RETURNING main_code",
                    (tid, product_code_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail="코드 없음")
        _audit(cur, tid, "product_code", product_code_id, "DELETE", request.state.user_id, {"mainCode": row[0]})
    return {"deleted": product_code_id, "mainCode": row[0]}


class ProductBatch(BaseModel):
    ids: list[int]
    action: str            # STATUS | DELETE
    status: str = ""       # action=STATUS 시 DRAFT/APPROVED/INACTIVE


@router.post("/codes/products/batch", dependencies=[SETUP])
def batch_product(request: Request, body: ProductBatch) -> dict[str, Any]:
    """제품 코드 일괄 작업 (그리드 다중 선택) — 상태 변경 또는 삭제.

    부분 실패 허용: 항목별 처리, 삭제는 참조 있으면 skip(사유 반환). 최대 200건."""
    ids = list(dict.fromkeys(body.ids))[:200]
    if not ids:
        raise HTTPException(422, detail="대상이 없습니다")
    action = body.action.strip().upper()
    if action not in ("STATUS", "DELETE"):
        raise HTTPException(422, detail="action 은 STATUS 또는 DELETE")
    st = body.status.strip().upper()
    if action == "STATUS" and st not in ("DRAFT", "APPROVED", "INACTIVE"):
        raise HTTPException(422, detail="상태는 DRAFT/APPROVED/INACTIVE")
    done = 0
    skipped: list[dict[str, str]] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for pid in ids:
            cur.execute("SELECT main_code FROM product_code WHERE tenant_id=%s AND product_code_id=%s",
                        (tid, pid))
            row = cur.fetchone()
            if not row:
                skipped.append({"id": str(pid), "reason": "없음"}); continue
            main_code = row[0]
            if action == "DELETE":
                ref = None
                for table, col in _PC_REFS:
                    cur.execute(f"SELECT 1 FROM {table} WHERE {col}=%s LIMIT 1", (pid,))
                    if cur.fetchone():
                        ref = table; break
                if ref:
                    skipped.append({"id": str(pid), "code": main_code, "reason": f"참조({ref})"}); continue
                for owned in _PC_OWNED:
                    cur.execute(f"DELETE FROM {owned} WHERE product_code_id=%s", (pid,))
                cur.execute("DELETE FROM product_code WHERE tenant_id=%s AND product_code_id=%s", (tid, pid))
                _audit(cur, tid, "product_code", pid, "DELETE", request.state.user_id,
                       {"mainCode": main_code, "batch": True})
            else:
                cur.execute("UPDATE product_code SET approval_status=%s "
                            "WHERE tenant_id=%s AND product_code_id=%s", (st, tid, pid))
                _audit(cur, tid, "product_code", pid, "UPDATE", request.state.user_id,
                       {"status": st, "batch": True})
            done += 1
    return {"action": action, "status": st or None, "done": done,
            "skipped": skipped, "requested": len(ids)}


@router.get("/codes/groups/{group}/export.xlsx")
def export_group_xlsx(group: str) -> Response:
    """그룹 code_item Excel 내보내기 (C2 — Slot·Item Name·Sort·Values)."""
    from openpyxl import Workbook
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT group_id FROM code_group WHERE tenant_id=%s AND group_code=%s", (tid, group))
        g = cur.fetchone()
        if not g:
            raise HTTPException(404, detail=f"그룹 없음: {group}")
        cur.execute(
            """SELECT ci.item_slot, ci.item_name, ci.sort_order,
                      string_agg(civ.value_code, ',' ORDER BY civ.sort_order)
               FROM code_item ci LEFT JOIN code_item_value civ ON civ.item_id=ci.item_id
               WHERE ci.group_id=%s GROUP BY ci.item_id ORDER BY ci.item_slot""", (g[0],))
        rows = cur.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = group[:31]
    ws.append(["Slot", "Item Name", "Sort", "Values"])
    for r in rows:
        ws.append([r[0], r[1], r[2], r[3] or ""])
    buf = io.BytesIO()
    wb.save(buf)
    from urllib.parse import quote
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(group)}.xlsx"})


# ── D8 XLSX export 전면 — 주요 대장 그리드 Excel 내보내기 ──
def _xlsx_response(sheet: str, headers: list[str], rows: list[list[Any]], filename: str) -> Response:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    from urllib.parse import quote
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}.xlsx",
                 "X-Row-Count": str(len(rows))})


@router.get("/prices/export.xlsx")
def export_prices_xlsx(request: Request) -> Response:
    """단가 대장 XLSX (D8). 1.5 — 단가 열람 제한 사용자는 다운로드 금지(403)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        _assert_downloadable(cur, tid, request, "price")
        cur.execute(
            """SELECT pc.main_code, pc.code_name, COALESCE(cc.company_name,'-'),
                      p.price, p.price_source, p.valid_from, p.valid_to
               FROM cst_price p JOIN product_code pc ON pc.product_code_id=p.product_code_id
               LEFT JOIN com_company cc ON cc.company_id=p.supplier_id
               WHERE p.tenant_id=%s ORDER BY pc.main_code, p.valid_from DESC""", (tid,))
        rows = [[r[0], r[1], r[2], float(r[3]), SOURCE_LABEL.get(r[4], r[4]),
                 r[5].isoformat(), r[6].isoformat() if r[6] else ""] for r in cur.fetchall()]
    return _xlsx_response("단가", ["코드", "품명", "공급처", "단가", "출처", "적용일", "만료일"], rows, "prices")


@router.get("/parts/export.xlsx")
def export_parts_xlsx() -> Response:
    """부품 대장 XLSX (D8)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.part_no, p.part_name, COALESCE(p.specification,''),
                      COALESCE(m.material_code,''), COALESCE(c.company_name,''),
                      COALESCE(pc.main_code,''), p.unit, p.weight, p.is_standard
               FROM prt_part p
               LEFT JOIN mat_material m ON m.material_id=p.material_id
               LEFT JOIN com_company c ON c.company_id=p.supplier_id
               LEFT JOIN product_code pc ON pc.product_code_id=p.product_code_id
               WHERE p.tenant_id=%s ORDER BY p.part_no""", (tid,))
        rows = [[r[0], r[1], r[2], r[3], r[4], r[5], r[6],
                 float(r[7]) if r[7] is not None else "", "표준" if r[8] else "사양"]
                for r in cur.fetchall()]
    return _xlsx_response("부품", ["부품번호", "품명", "사양", "재질", "공급처", "제품코드", "단위", "중량", "구분"], rows, "parts")


@router.get("/drawings/export.xlsx")
def export_drawings_xlsx() -> Response:
    """도면 대장 XLSX (D8)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT d.drawing_no, d.drawing_name, d.drawing_type, d.dwg_kind,
                      d.current_rev, d.status,
                      (SELECT COUNT(*) FROM dwg_revision r WHERE r.drawing_id=d.drawing_id)
               FROM dwg_drawing d WHERE d.tenant_id=%s ORDER BY d.drawing_no""", (tid,))
        rows = [[r[0], r[1], r[2], r[3], r[4], r[5], r[6]] for r in cur.fetchall()]
    return _xlsx_response("도면", ["도면번호", "도면명", "유형", "종류", "Rev", "상태", "개정수"], rows, "drawings")


@router.get("/erp/warehouses/export.xlsx")
def export_warehouses_xlsx() -> Response:
    """창고·저장위치 XLSX (D8)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT location_code, location_name, location_type,
                      COALESCE(hazard_allowed,''), COALESCE(inspection_cycle,''), COALESCE(remarks,'')
               FROM erp_warehouse WHERE tenant_id=%s ORDER BY location_code""", (tid,))
        rows = [[r[0], r[1], r[2], r[3], r[4], r[5]] for r in cur.fetchall()]
    return _xlsx_response("창고", ["위치코드", "위치명", "유형", "위험물", "점검주기", "비고"], rows, "warehouses")


@router.get("/companies/export.xlsx")
def export_companies_xlsx() -> Response:
    """공급처·거래처 XLSX (D8)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT company_name, company_type, COALESCE(nation,''),
                      COALESCE(evaluation_grade,''), COALESCE(payment_terms,''), COALESCE(remarks,'')
               FROM com_company WHERE tenant_id=%s ORDER BY company_id""", (tid,))
        rows = [[r[0], r[1], r[2], r[3], r[4], r[5]] for r in cur.fetchall()]
    return _xlsx_response("거래처", ["업체명", "구분", "국가", "평가등급", "결제조건", "비고"], rows, "companies")


@router.get("/history/export.xlsx")
def export_history_xlsx(limit: int = 1000, fromDate: str = "", toDate: str = "",
                        user: str = "", action: str = "", target: str = "") -> Response:
    """감사 로그 XLSX (D8/D9) — 감사 조회와 동일 필터(기간/사용자/작업/대상) 적용."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        where, params = _audit_where(tid, fromDate, toDate, user, action, target)
        cur.execute(
            f"""SELECT to_char(h.acted_at,'YYYY-MM-DD HH24:MI:SS'),
                       h.target_table||' #'||h.target_id, h.action, u.user_name, u.login_id
                FROM sys_history h JOIN sys_user u ON u.user_id=h.actor_id
                WHERE {where} ORDER BY h.history_id DESC LIMIT %s""",
            (*params, max(1, min(limit, 10000))))
        rows = [[r[0], r[1], r[2], r[3], r[4]] for r in cur.fetchall()]
    return _xlsx_response("감사로그", ["일시", "대상", "작업", "수행자", "사번"], rows, "audit")


@router.post("/codes/groups/{group}/import-excel", dependencies=[SETUP])
async def import_group_excel(group: str, request: Request,
                             uploadedFile: UploadFile = File(...),
                             dryRun: bool = False) -> dict[str, Any]:
    """code_item 일괄 Import (C2) — Slot·Item Name 헤더, 행 단위 거부 리포트 (Slot 중복=갱신).
    dryRun=true (트리아지 #32 Diff Review): 반영 없이 insert/update/거부 미리보기만 반환."""
    import openpyxl
    data = await uploadedFile.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(422, detail="Excel 파일이 아닙니다 (.xlsx)")
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    if "slot" not in header or "item name" not in header:
        raise HTTPException(422, detail="헤더 필요 — Slot, Item Name")
    si, ni = header.index("slot"), header.index("item name")
    inserted = updated = 0
    rejected: list[str] = []
    diff: list[dict[str, Any]] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT group_id FROM code_group WHERE tenant_id=%s AND group_code=%s", (tid, group))
        g = cur.fetchone()
        if not g:
            raise HTTPException(404, detail=f"그룹 없음: {group}")
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            slot = str(row[si].value).strip().upper()[:5] if si < len(row) and row[si].value is not None else ""
            nm = str(row[ni].value).strip()[:100] if ni < len(row) and row[ni].value is not None else ""
            if not slot or not nm:
                rejected.append(f"{i}행: Slot·Item Name 필수")
                continue
            cur.execute("SELECT item_id, item_name FROM code_item WHERE group_id=%s AND item_slot=%s",
                        (g[0], slot))
            ex = cur.fetchone()
            if ex:
                if dryRun:
                    diff.append({"row": i, "slot": slot, "action": "update", "before": ex[1], "after": nm})
                else:
                    cur.execute("UPDATE code_item SET item_name=%s, updated_at=now() WHERE item_id=%s",
                                (nm, ex[0]))
                updated += 1
            else:
                if dryRun:
                    diff.append({"row": i, "slot": slot, "action": "insert", "after": nm})
                else:
                    cur.execute(
                        """INSERT INTO code_item (tenant_id, group_id, item_slot, item_name, sort_order)
                           VALUES (%s,%s,%s,%s,%s)""", (tid, g[0], slot, nm, inserted + updated))
                inserted += 1
        if dryRun:
            return {"dryRun": True, "inserted": inserted, "updated": updated,
                    "rejected": rejected, "diff": diff[:50]}
        _audit(cur, tid, "code_group", g[0], "IMPORT", request.state.user_id,
               {"group": group, "inserted": inserted, "updated": updated, "rejected": len(rejected)})
    return {"inserted": inserted, "updated": updated, "rejected": rejected}


class ExpandRequest(BaseModel):
    rootCode: str = "KDCR 3-13"
    slotValues: dict[str, str]


def _expand_rows(cur, tid: int, root_code: str, slot_values: dict[str, str]) -> list[tuple]:
    cur.execute(
            """
            WITH RECURSIVE bom AS (
              SELECT pc.product_code_id, pc.main_code, pc.code_name,
                     1::numeric AS quantity, 0 AS lvl,
                     %s::jsonb AS slots, pc.main_code::text AS path, ''::text AS ord,
                     NULL::bigint AS rel_id, NULL::int AS rel_rev,
                     ARRAY[pc.product_code_id] AS seen
              FROM product_code pc
              WHERE pc.tenant_id=%s AND pc.main_code=%s
              UNION ALL
              SELECT c.product_code_id, c.main_code, c.code_name,
                     (b.quantity * r.quantity)::numeric, b.lvl + 1,
                     COALESCE((SELECT jsonb_object_agg(sm.child_slot,
                         CASE WHEN sm.mother_slot IS NOT NULL
                              THEN b.slots ->> sm.mother_slot
                              ELSE sm.fixed_value END)
                       FROM code_relationship_slot_map sm
                       WHERE sm.rel_id = r.rel_id), '{}'::jsonb),
                     (b.path || ' > ' || c.main_code)::text,
                     (b.ord || lpad(r.sort_order::text, 4, '0'))::text,
                     r.rel_id, r.revision_no,
                     b.seen || c.product_code_id
              FROM bom b
              JOIN code_relationship r
                ON r.mother_code_id = b.product_code_id AND r.approval_status = 'APPROVED'
               AND r.tenant_id = %s
              JOIN product_code c
                ON c.product_code_id = r.child_code_id AND c.tenant_id = %s
              WHERE b.lvl < 10
                AND NOT (c.product_code_id = ANY(b.seen))   -- 순환 가드 (where-used 와 동일 규약)
            )
            SELECT b.main_code, b.code_name, b.quantity, b.lvl, b.slots, b.path,
                   (SELECT p.price FROM cst_price p
                    WHERE p.product_code_id = b.product_code_id
                      AND p.valid_from <= CURRENT_DATE
                      AND (p.valid_to IS NULL OR p.valid_to >= CURRENT_DATE)
                    ORDER BY array_position(%s::text[], p.price_source), p.valid_from DESC
                    LIMIT 1) AS price,
                   b.rel_id, b.rel_rev
            FROM bom b WHERE b.lvl > 0 ORDER BY b.ord
            """,
            (json.dumps(slot_values), tid, root_code, tid, tid, SOURCE_PRIORITY))
    return cur.fetchall()


def _rel_basis(rows: list[tuple]) -> dict[str, Any]:
    """전개 근거 (#40) — 이 BOM 이 사용한 (관계 rel_id, Revision) 집합 + 체크섬.

    같은 근거 = 같은 BOM 이므로, 근거 체크섬이 같으면 재실행 결과가 같음을 보증한다.
    반대로 체크섬이 달라졌다면 무엇이 바뀌었는지 edges 비교로 정확히 짚을 수 있다."""
    edges = sorted({(int(r[7]), int(r[8] or 1)) for r in rows if r[7] is not None})
    payload = [{"relId": e[0], "revisionNo": e[1]} for e in edges]
    return {"edges": payload, "checksum": _snapshot_checksum({"edges": payload})}


def _resolved(main: str, slots: dict[str, str]) -> str:
    parts = [v for _, v in sorted(slots.items()) if v]
    return f"{main}-{'-'.join(parts)}" if parts else main


@router.post("/codes/products/expand")
def expand(body: ExpandRequest) -> dict[str, Any]:
    """BOM 재귀 전개 — code_relationship + slot_map (verify_runtime.sql T1 과 동일 로직)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        rows = _expand_rows(cur, tid, body.rootCode, body.slotValues)
    if not rows:
        raise HTTPException(404, detail=f"root code not found: {body.rootCode}")

    resolved = _resolved
    sv = body.slotValues
    finished = f"{body.rootCode}-{sv.get('B', '?')}-{sv.get('E', '?')}"
    return {
        "finishedGoodsCode": finished,
        "items": [
            {
                "level": r[3], "mainCode": r[0],
                "resolvedCode": resolved(r[0], r[4] or {}),
                "name": r[1], "quantity": float(r[2]),
                "priceK": round(float(r[6]) / 1000) if r[6] is not None else None,
                "path": r[5],
            }
            for r in rows
        ],
    }


# ── B10 — C-1 마감: 견적 미리보기 PDF · 사양 Excel Import (CPQ-002) ──

class QuotePreview(BaseModel):
    rootCode: str = "KDCR 3-13"
    slotValues: dict[str, str] = {}


@router.post("/cpq/quote-preview.pdf", dependencies=[SETUP])
def quote_preview(body: QuotePreview) -> Any:
    """견적 미리보기 — 현재 슬롯 선택으로 BOM 전개+단가 후 견적서 PDF 즉석 렌더 (영속 없음)."""
    from types import SimpleNamespace

    from fastapi.responses import Response

    from ..services import run_pipeline as rp
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        rows = _expand_rows(cur, tid, body.rootCode, body.slotValues)
    if not rows:
        raise HTTPException(404, detail=f"root code not found: {body.rootCode}")
    items = [{
        "resolvedCode": _resolved(r[0], r[4] or {}), "name": r[1],
        "quantity": float(r[2]),
        "priceK": round(float(r[6]) / 1000) if r[6] is not None else None,
    } for r in rows]
    total_k = sum((i["priceK"] or 0) * i["quantity"] for i in items)
    ns = SimpleNamespace(items=items, total_k=total_k, files=[])
    rp.step_quotation(ns, "PS-61313-5 · 미리보기")
    return Response(content=ns.files[-1][3], media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=\"quote-preview.pdf\""})


@router.post("/cpq/spec-import", dependencies=[SETUP])
async def spec_import(uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """사양 Excel Import — 헤더 Slot·Value 2열 → slotValues (CPQ-002)."""
    import openpyxl
    data = await uploadedFile.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(422, detail="Excel 파일이 아닙니다 (.xlsx)")
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if "Slot" not in header or "Value" not in header:
        raise HTTPException(422, detail="헤더 불일치 — 필요 열: Slot·Value")
    si, vi = header.index("Slot"), header.index("Value")
    sv: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        s, v = row[si].value, row[vi].value
        if s is None or v is None:
            continue
        sv[str(s).strip().upper()[:5]] = str(v).strip()
    if not sv:
        raise HTTPException(422, detail="유효한 Slot·Value 행 없음")
    return {"slotValues": sv}


# ── B13 — Arrangement Set-Up (M-4-2) · Templet 관리 (S-2-3) ──

@router.get("/arrangements")
def arrangements() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT a.arrangement_code, a.arrangement_name, a.product_family,
                      COALESCE(a.direction_option,''), COALESCE(a.install_option,''),
                      a.approval_status,
                      (SELECT count(*) FROM arrangement_component c
                        WHERE c.arrangement_id=a.arrangement_id)
               FROM arrangement_code a WHERE a.tenant_id=%s ORDER BY a.arrangement_id""", (tid,))
        return [
            {"code": r[0], "name": r[1], "family": r[2], "direction": r[3],
             "install": r[4], "status": r[5], "components": int(r[6])}
            for r in cur.fetchall()
        ]


@router.get("/arrangements/{code}/components")
def arrangement_components(code: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT COALESCE(c.position_key,''), pc.main_code, pc.code_name, c.quantity,
                      c.component_id, c.pos_x, c.pos_y, c.width, c.height, c.sort_order
               FROM arrangement_component c
               JOIN arrangement_code a ON a.arrangement_id=c.arrangement_id
               JOIN product_code pc ON pc.product_code_id=c.product_code_id
               WHERE a.tenant_id=%s AND a.arrangement_code=%s ORDER BY c.sort_order""",
            (tid, code))
        rows = cur.fetchall()
    return [
        {"position": r[0], "code": r[1], "name": r[2], "quantity": float(r[3]),
         "componentId": r[4],
         # C11 — geometry (미지정 시 sort_order 기반 2열 자동 배치 폴백)
         "x": r[5] if r[5] is not None else 20 + (r[9] % 2) * 150,
         "y": r[6] if r[6] is not None else 20 + (r[9] // 2) * 70,
         "w": r[7] if r[7] is not None else 130,
         "h": r[8] if r[8] is not None else 56}
        for r in rows
    ]


class GeometryPatch(BaseModel):
    x: int
    y: int
    w: int = 130
    h: int = 56


@router.patch("/arrangements/{code}/components/{component_id}/geometry", dependencies=[SETUP])
def set_component_geometry(code: str, component_id: int, body: GeometryPatch) -> dict[str, Any]:
    """구성도 블록 좌표 저장 (C11) — 드래그 배치 영속."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE arrangement_component c SET pos_x=%s, pos_y=%s, width=%s, height=%s
               FROM arrangement_code a
               WHERE c.arrangement_id=a.arrangement_id AND a.tenant_id=%s
                 AND a.arrangement_code=%s AND c.component_id=%s
               RETURNING c.component_id""",
            (body.x, body.y, max(40, body.w), max(30, body.h), tid, code, component_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"구성 컴포넌트 없음: {component_id}")
    return {"componentId": component_id, "x": body.x, "y": body.y, "w": body.w, "h": body.h}


class ArrangementCreate(BaseModel):
    code: str
    name: str
    family: str = "AHU"
    direction: str = ""
    install: str = ""


@router.post("/arrangements", status_code=201, dependencies=[SETUP])
def create_arrangement(request: Request, body: ArrangementCreate) -> dict[str, Any]:
    if not body.code.strip() or not body.name.strip():
        raise HTTPException(422, detail="필수(노란 셀) — Code·명칭")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM arrangement_code WHERE tenant_id=%s AND arrangement_code=%s",
            (tid, body.code.strip()))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — Arrangement {body.code} 이미 등록됨")
        cur.execute(
            """INSERT INTO arrangement_code (tenant_id, arrangement_code, arrangement_name,
               product_family, direction_option, install_option)
               VALUES (%s,%s,%s,%s,NULLIF(%s,''),NULLIF(%s,'')) RETURNING arrangement_id""",
            (tid, body.code.strip(), body.name.strip(), body.family.strip()[:50],
             body.direction.strip(), body.install.strip()))
        arr_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'arrangement_code',%s,'CREATE','승인',%s,%s)""",
            (tid, arr_id, request.state.user_id,
             f"Arrangement 등록 — {body.code} {body.name}"[:200]))
    return {"arrangementId": arr_id, "status": "DRAFT"}


class ComponentAdd(BaseModel):
    productCode: str
    position: str = ""
    quantity: float = 1


@router.post("/arrangements/{code}/components", status_code=201, dependencies=[SETUP])
def add_arrangement_component(code: str, body: ComponentAdd) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT arrangement_id FROM arrangement_code WHERE tenant_id=%s AND arrangement_code=%s",
            (tid, code))
        a = cur.fetchone()
        if not a:
            raise HTTPException(404, detail=f"Arrangement 없음: {code}")
        cur.execute(
            "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
            (tid, body.productCode.strip()))
        pc = cur.fetchone()
        if not pc:
            raise HTTPException(404, detail=f"코드 없음: {body.productCode}")
        cur.execute(
            """INSERT INTO arrangement_component (arrangement_id, product_code_id,
               position_key, quantity, sort_order)
               VALUES (%s,%s,NULLIF(%s,''),%s,
                       (SELECT COALESCE(max(sort_order),0)+1 FROM arrangement_component
                         WHERE arrangement_id=%s))
               RETURNING component_id""",
            (a[0], pc[0], body.position.strip()[:30], body.quantity, a[0]))
        cid = cur.fetchone()[0]
    return {"componentId": cid}


@router.get("/toolbox/templets")
def templets() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT templet_name, templet_type, definition, approval_status, is_system
               FROM tbx_templet WHERE tenant_id=%s ORDER BY templet_id""", (tid,))
        return [
            {"name": r[0], "templetType": r[1], "definition": r[2],
             "status": r[3], "system": bool(r[4])}
            for r in cur.fetchall()
        ]


class TempletSave(BaseModel):
    templetType: str = "COMMAND"
    definition: dict[str, Any] = {}


@router.put("/toolbox/templets/{name}", dependencies=[SETUP])
def save_templet(name: str, body: TempletSave) -> dict[str, Any]:
    """Templet upsert — 저장 시 DRAFT 로 회귀 (승인은 POST /approvals)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE tbx_templet SET definition=%s, templet_type=%s,
               approval_status='DRAFT', updated_at=now()
               WHERE tenant_id=%s AND templet_name=%s RETURNING templet_id""",
            (json.dumps(body.definition), body.templetType.strip()[:30], tid, name))
        row = cur.fetchone()
        if not row:
            cur.execute(
                """INSERT INTO tbx_templet (tenant_id, templet_name, templet_type, definition)
                   VALUES (%s,%s,%s,%s) RETURNING templet_id""",
                (tid, name, body.templetType.strip()[:30] or "COMMAND",
                 json.dumps(body.definition)))
            row = cur.fetchone()
    return {"templetId": row[0], "status": "DRAFT"}


# ── B13-2 — Variant·Constant (S-1-2) · Raw Material·GPI (M-3-2) · Quality (M-4-5) ──

@router.get("/codes/values")
def code_values(group: str = "KOF") -> list[dict[str, Any]]:
    """그룹 슬롯별 값 목록 — code_item + code_item_value (S-1-2)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT i.item_slot, i.item_name, v.value_code, COALESCE(v.value_name,''),
                      v.approval_status, v.value_id
               FROM code_item i
               JOIN code_group g ON g.group_id=i.group_id
               LEFT JOIN code_item_value v ON v.item_id=i.item_id
               WHERE i.tenant_id=%s AND g.group_code=%s
               ORDER BY i.item_slot, v.sort_order""", (tid, group))
        return [
            {"slot": r[0], "itemName": r[1], "valueCode": r[2] or "", "valueName": r[3],
             "status": r[4] or "-", "valueId": r[5]}
            for r in cur.fetchall()
        ]


class ValueAdd(BaseModel):
    group: str = "KOF"
    slot: str
    valueCode: str
    valueName: str = ""


@router.post("/codes/values", status_code=201, dependencies=[SETUP])
def add_code_value(request: Request, body: ValueAdd) -> dict[str, Any]:
    """슬롯 값 추가 — PENDING + 승인 요청 자동 (add_item 패턴)."""
    if not body.slot.strip() or not body.valueCode.strip():
        raise HTTPException(422, detail="필수(노란 셀) — Slot·값 코드")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT i.item_id FROM code_item i
               JOIN code_group g ON g.group_id=i.group_id
               WHERE i.tenant_id=%s AND g.group_code=%s AND i.item_slot=%s""",
            (tid, body.group, body.slot.strip().upper()))
        item = cur.fetchone()
        if not item:
            raise HTTPException(404, detail=f"Slot 없음: {body.group}/{body.slot}")
        cur.execute(
            "SELECT 1 FROM code_item_value WHERE item_id=%s AND value_code=%s",
            (item[0], body.valueCode.strip()[:30]))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 값 {body.valueCode} 이미 등록됨")
        cur.execute(
            """INSERT INTO code_item_value (tenant_id, item_id, value_code, value_name,
               approval_status, sort_order)
               VALUES (%s,%s,%s,NULLIF(%s,''),'PENDING',
                       (SELECT COALESCE(max(sort_order),0)+1 FROM code_item_value WHERE item_id=%s))
               RETURNING value_id""",
            (tid, item[0], body.valueCode.strip()[:30], body.valueName.strip()[:100], item[0]))
        vid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'code_item_value',%s,'CREATE','승인',%s,%s)""",
            (tid, vid, request.state.user_id,
             f"슬롯 값 등록 — {body.group}/{body.slot} {body.valueCode}"[:200]))
    return {"valueId": vid, "status": "PENDING"}


@router.get("/materials")
def materials() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT material_code, material_name, material_type,
                      density, COALESCE(standard,''), COALESCE(hazard_class,'')
               FROM mat_material WHERE tenant_id=%s ORDER BY material_id""", (tid,))
        return [
            {"code": r[0], "name": r[1], "materialType": r[2],
             "density": float(r[3]) if r[3] is not None else None,
             "standard": r[4], "hazard": r[5]}
            for r in cur.fetchall()
        ]


class MaterialCreate(BaseModel):
    code: str
    name: str
    materialType: str = "STEEL"
    density: float | None = None
    standard: str = ""
    hazard: str = ""


@router.post("/materials", status_code=201, dependencies=[SETUP])
def create_material(body: MaterialCreate) -> dict[str, Any]:
    if not body.code.strip() or not body.name.strip():
        raise HTTPException(422, detail="필수(노란 셀) — 재질 코드·명")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM mat_material WHERE tenant_id=%s AND material_code=%s",
            (tid, body.code.strip()))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 재질 {body.code} 이미 등록됨")
        cur.execute(
            """INSERT INTO mat_material (tenant_id, material_code, material_name,
               material_type, density, standard, hazard_class)
               VALUES (%s,%s,%s,%s,%s,NULLIF(%s,''),NULLIF(%s,'')) RETURNING material_id""",
            (tid, body.code.strip()[:30], body.name.strip()[:100],
             body.materialType.strip()[:20], body.density,
             body.standard.strip()[:30], body.hazard.strip()[:30]))
        mid = cur.fetchone()[0]
    return {"materialId": mid}


@router.get("/drawings/{drawing_no}/verifications")
def drawing_verifications(drawing_no: str) -> list[dict[str, Any]]:
    """검증 Macro 규칙 — dwg_verification (Quality M-4-5)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT v.rule_name, m.macro_name, v.warning_message, v.is_active,
                      v.verification_id
               FROM dwg_verification v
               JOIN dwg_drawing d ON d.drawing_id=v.drawing_id
               JOIN tbx_macro m ON m.macro_id=v.macro_id
               WHERE v.tenant_id=%s AND d.drawing_no=%s ORDER BY v.verification_id""",
            (tid, drawing_no))
        return [
            {"rule": r[0], "macro": r[1], "warning": r[2], "active": bool(r[3]),
             "verificationId": r[4]}
            for r in cur.fetchall()
        ]


class VerificationAdd(BaseModel):
    ruleName: str
    macroName: str
    warning: str


@router.post("/drawings/{drawing_no}/verifications", status_code=201, dependencies=[SETUP])
def add_drawing_verification(drawing_no: str, body: VerificationAdd) -> dict[str, Any]:
    if not body.ruleName.strip() or not body.warning.strip():
        raise HTTPException(422, detail="필수(노란 셀) — 규칙명·경고 문구")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT drawing_id FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s LIMIT 1",
            (tid, drawing_no))
        d = cur.fetchone()
        if not d:
            raise HTTPException(404, detail=f"도면 없음: {drawing_no}")
        cur.execute(
            "SELECT macro_id FROM tbx_macro WHERE tenant_id=%s AND macro_name=%s",
            (tid, body.macroName.strip()))
        m = cur.fetchone()
        if not m:
            raise HTTPException(404, detail=f"Macro 없음: {body.macroName}")
        cur.execute(
            """INSERT INTO dwg_verification (tenant_id, drawing_id, rule_name, macro_id,
               warning_message) VALUES (%s,%s,%s,%s,%s) RETURNING verification_id""",
            (tid, d[0], body.ruleName.strip()[:100], m[0], body.warning.strip()[:500]))
        vid = cur.fetchone()[0]
    return {"verificationId": vid}


class VerifyRunRequest(BaseModel):
    measurements: dict[str, float] = {}   # 규칙 Macro 입력 변수 (측정값)


@router.post("/drawings/{drawing_no}/verify")
def verify_drawing(drawing_no: str, body: VerifyRunRequest) -> dict[str, Any]:
    """검증 규칙 자동 판정 (D4) — 측정값을 활성 규칙 Macro 로 평가해 합/부 제안.
    규칙 값이 참(≠0)=통과, 거짓/오류=위반(warning). 전건 통과=합격 제안, 아니면 불합격."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT v.rule_name, m.macro_name, m.macro_expr, v.warning_message
               FROM dwg_verification v
               JOIN dwg_drawing d ON d.drawing_id=v.drawing_id
               JOIN tbx_macro m ON m.macro_id=v.macro_id
               WHERE v.tenant_id=%s AND d.drawing_no=%s AND v.is_active=true
               ORDER BY v.verification_id""", (tid, drawing_no))
        rules = cur.fetchall()
        if not rules:
            raise HTTPException(404, detail=f"활성 검증 규칙 없음: {drawing_no}")
        resolver = _make_table_resolver(cur, tid)
        results = []
        for rule_name, macro_name, expr, warn in rules:
            ev = Evaluator(dict(body.measurements), resolver)
            try:
                val = ev.run(expr or "0")
                passed = bool(val) and float(val) != 0.0
                res = {"rule": rule_name, "macro": macro_name,
                       "value": float(val) if isinstance(val, (int, float)) else None,
                       "pass": passed, "warning": None if passed else warn}
            except MacroError as e:
                res = {"rule": rule_name, "macro": macro_name, "value": None,
                       "pass": False, "warning": f"{warn} (평가 오류: {e})"}
            results.append(res)
    fails = sum(1 for r in results if not r["pass"])
    suggestion = "합격" if fails == 0 else "불합격"
    return {"drawingNo": drawing_no, "evaluated": len(results),
            "pass": len(results) - fails, "fail": fails,
            "suggestion": suggestion, "results": results}


# ── B14 — 마스터 데이터 (com_company) · RBAC 동적화 (sys_role) · Hierarchy ──

@router.get("/companies")
def companies(active_only: bool = False) -> list[dict[str, Any]]:
    """거래처 목록. active_only=true 면 비활성 제외 (고객/공급처 선택 리스트용)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT company_name, company_type, COALESCE(nation,''),
                      COALESCE(evaluation_grade,''), COALESCE(payment_terms,''),
                      company_id, COALESCE(remarks,''), COALESCE(is_active, true)
               FROM com_company WHERE tenant_id=%s
                 AND (%s = false OR COALESCE(is_active, true) = true)
               ORDER BY company_id""", (tid, active_only))
        return [
            {"name": r[0], "companyType": r[1], "nation": r[2], "grade": r[3], "terms": r[4],
             "companyId": r[5], "remarks": r[6], "isActive": r[7]}
            for r in cur.fetchall()
        ]


class CompanyCreate(BaseModel):
    name: str
    companyType: str = "SUPPLIER"
    nation: str = ""
    grade: str = ""
    terms: str = ""


@router.post("/companies", status_code=201, dependencies=[SETUP])
def create_company(body: CompanyCreate) -> dict[str, Any]:
    if not body.name.strip():
        raise HTTPException(422, detail="필수(노란 셀) — 업체명")
    ctype = body.companyType.strip().upper()
    if ctype not in ("CUSTOMER", "SUPPLIER", "PARTNER", "BANK"):
        raise HTTPException(422, detail=f"유형 오류: {body.companyType}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM com_company WHERE tenant_id=%s AND company_name=%s",
            (tid, body.name.strip()))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 업체 {body.name} 이미 등록됨")
        cur.execute(
            """INSERT INTO com_company (tenant_id, company_name, company_type, nation,
               evaluation_grade, payment_terms)
               VALUES (%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,'')) RETURNING company_id""",
            (tid, body.name.strip()[:200], ctype, body.nation.strip()[:50],
             body.grade.strip()[:10], body.terms.strip()[:200]))
        cid = cur.fetchone()[0]
    return {"companyId": cid}


def _load_ws(data: bytes, required: list[str]) -> tuple[Any, dict[str, int]]:
    """xlsx 로드 + 헤더 검증 → (worksheet, 열인덱스맵)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(422, detail="Excel 파일이 아닙니다 (.xlsx)")
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    miss = [h for h in required if h not in header]
    if miss:
        raise HTTPException(422, detail=f"헤더 누락: {miss} (필요: {required})")
    return ws, {h: header.index(h) for h in header if h}


@router.post("/companies/import-excel", dependencies=[SETUP])
async def import_companies_excel(uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """거래처 대량 등록 — 헤더: 업체명·유형·국가·결제조건 (중복명 거부). 유형 미지정=SUPPLIER."""
    ws, idx = _load_ws(await uploadedFile.read(), ["업체명"])
    inserted, rejected = 0, []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            def cell(col: str) -> str:
                i = idx.get(col)
                v = row[i].value if i is not None else None
                return str(v).strip() if v is not None else ""
            name = cell("업체명")
            if not name:
                continue
            ctype = (cell("유형") or "SUPPLIER").upper()
            if ctype not in ("CUSTOMER", "SUPPLIER", "PARTNER", "BANK"):
                rejected.append(f"{r_i}행 {name}: 유형 오류({ctype})"); continue
            cur.execute("SELECT 1 FROM com_company WHERE tenant_id=%s AND company_name=%s", (tid, name[:200]))
            if cur.fetchone():
                rejected.append(f"{r_i}행 {name}: 중복"); continue
            cur.execute(
                """INSERT INTO com_company (tenant_id, company_name, company_type, nation, payment_terms)
                   VALUES (%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''))""",
                (tid, name[:200], ctype, cell("국가")[:50], cell("결제조건")[:200]))
            inserted += 1
    return {"inserted": inserted, "rejected": rejected, "rejectedCount": len(rejected)}


@router.get("/roles")
def roles() -> list[dict[str, Any]]:
    """역할 + 권한 매트릭스 — sys_role/sys_role_permission (M-14-6 실데이터)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.role_id, r.role_name, COALESCE(r.description,'')
               FROM sys_role r WHERE r.tenant_id=%s ORDER BY r.role_id""", (tid,))
        out = [{"name": r[1], "description": r[2], "_id": r[0]} for r in cur.fetchall()]
        for role in out:
            cur.execute(
                """SELECT resource_key, action FROM sys_role_permission
                   WHERE role_id=%s ORDER BY resource_key""", (role["_id"],))
            perms: dict[str, str] = {}
            for key, action in cur.fetchall():
                # WRITE 가 READ 를 포함 — 셀 표기는 최상위 권한
                if perms.get(key) != "WRITE":
                    perms[key] = action
            role["permissions"] = perms
            del role["_id"]
    return out


class RolePermissions(BaseModel):
    permissions: dict[str, str] = {}   # resource_key(screenId) -> 'READ'|'WRITE'|'NONE'


@router.put("/roles/{role_name}/permissions", dependencies=[SETUP])
def set_role_permissions(role_name: str, request: Request, body: RolePermissions) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT role_id FROM sys_role WHERE tenant_id=%s AND role_name=%s",
            (tid, role_name))
        role = cur.fetchone()
        if not role:
            raise HTTPException(404, detail=f"역할 없음: {role_name}")
        n = 0
        for key, action in body.permissions.items():
            act = action.strip().upper()
            cur.execute(
                "DELETE FROM sys_role_permission WHERE role_id=%s AND resource_key=%s",
                (role[0], key[:200]))
            if act in ("READ", "WRITE"):
                cur.execute(
                    """INSERT INTO sys_role_permission (role_id, resource_type, resource_key, action)
                       VALUES (%s,'MENU',%s,%s)""", (role[0], key[:200], act))
            n += 1
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'sys_role_permission',%s,'PERM_CHANGE',%s,%s)""",
            (tid, role[0], request.state.user_id, json.dumps(body.permissions)))
    return {"role": role_name, "updated": n}


_BUILTIN_ROLES = {"PLATFORM", "ADMIN", "SETUP", "GENERAL"}


class RoleCreate(BaseModel):
    name: str
    description: str = ""


@router.post("/roles", status_code=201, dependencies=[ADMIN])
def create_role(request: Request, body: RoleCreate) -> dict[str, Any]:
    """역할 생성 (커스텀) — 예약(내장) 이름·중복 409."""
    name = body.name.strip().upper()[:50]
    if not name:
        raise HTTPException(422, detail="필수 — 역할명")
    if name in _BUILTIN_ROLES:
        raise HTTPException(409, detail=f"예약된 내장 역할: {name}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM sys_role WHERE tenant_id=%s AND role_name=%s", (tid, name))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 역할 {name}")
        cur.execute(
            "INSERT INTO sys_role (tenant_id, role_name, description, created_by) "
            "VALUES (%s,%s,NULLIF(%s,''),%s) RETURNING role_id",
            (tid, name, body.description.strip()[:200], str(request.state.user_id)))
        rid = cur.fetchone()[0]
        _audit(cur, tid, "sys_role", rid, "ROLE_CREATE", request.state.user_id, {"name": name})
    return {"roleId": rid, "name": name}


@router.delete("/roles/{role_name}", dependencies=[ADMIN])
def delete_role(role_name: str, request: Request) -> dict[str, Any]:
    """역할 삭제 — 내장 역할·사용자 배정 시 409. 권한 매트릭스는 연쇄 정리."""
    name = role_name.strip().upper()
    if name in _BUILTIN_ROLES:
        raise HTTPException(409, detail=f"내장 역할은 삭제 불가: {name}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT role_id FROM sys_role WHERE tenant_id=%s AND role_name=%s", (tid, name))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"역할 없음: {role_name}")
        rid = row[0]
        cur.execute("SELECT count(*) FROM sys_user_role WHERE role_id=%s", (rid,))
        if cur.fetchone()[0] > 0:
            raise HTTPException(409, detail="사용자에 배정된 역할 — 배정 해제 후 삭제")
        cur.execute("DELETE FROM sys_role_permission WHERE role_id=%s", (rid,))
        cur.execute("DELETE FROM sys_role WHERE tenant_id=%s AND role_id=%s", (tid, rid))
        _audit(cur, tid, "sys_role", rid, "ROLE_DELETE", request.state.user_id, {"name": name})
    return {"deleted": name}


@router.get("/hierarchy")
def hierarchy(treeType: str = "PRODUCT") -> list[dict[str, Any]]:
    """Hierarchy 주소 체계 — tbx_macro/tbl 의 hierarchy_address 원천 (M-3-1)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT hierarchy_id, parent_id, node_name, COALESCE(symbol,''),
                      address, approval_status,
                      COALESCE(remark,''), COALESCE(color,''), is_locked
               FROM sys_hierarchy WHERE tenant_id=%s AND tree_type=%s
               ORDER BY sort_order, hierarchy_id""", (tid, treeType.strip().upper()))
        return [
            {"id": r[0], "parentId": r[1], "name": r[2], "symbol": r[3],
             "address": r[4], "status": r[5],
             "remark": r[6], "color": r[7], "locked": r[8]}
            for r in cur.fetchall()
        ]


# ── SVC-05 Table ──
@router.get("/tables")
def tables_list() -> list[dict[str, Any]]:
    """데이터 Table 목록 (U13 우측 공용 패널·M-3-7 콤보) — 이름·유형·행 수."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT t.table_name, t.table_type, COALESCE(t.department,''),
                      COALESCE(t.description,''), count(r.row_id)
               FROM tbl_data_table t LEFT JOIN tbl_data_row r ON r.table_id=t.table_id
               WHERE t.tenant_id=%s
               GROUP BY t.table_id ORDER BY t.table_name""", (tid,))
        return [{"name": r[0], "type": r[1], "department": r[2],
                 "description": r[3], "rows": r[4]} for r in cur.fetchall()]


@router.get("/tables/tech-data")
def tech_data(airflow: float = 0, pressure: float = 0) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.row_key, r.row_values
               FROM tbl_data_table t JOIN tbl_data_row r ON r.table_id=t.table_id
               WHERE t.tenant_id=%s AND t.table_name='FanTechData'
               ORDER BY r.row_key_num""", (tid,))
        rows = [{"model": r[0], **r[1]} for r in cur.fetchall()]
    rows.sort(key=lambda x: abs(x["pd"] - airflow) + abs(x["pt"] - pressure))
    return rows


# ── SVC-05 Table CRUD + Excel Import (TBL-001~006) ──
def _table_id(cur, tid: int, name: str) -> tuple[int, list[str]]:
    cur.execute(
        """SELECT table_id, column_def FROM tbl_data_table
           WHERE tenant_id=%s AND table_name=%s""", (tid, name))
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"table not found: {name}")
    return row[0], list(row[1].get("columns", []))


@router.get("/tables/{name}")
def get_table(name: str) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, columns = _table_id(cur, tid, name)
        cur.execute(
            """SELECT row_key, row_values FROM tbl_data_row
               WHERE table_id=%s ORDER BY row_key_num NULLS LAST, row_key""", (table_id,))
        rows = [{"key": r[0], "values": r[1]} for r in cur.fetchall()]
    return {"name": name, "columns": columns, "rows": rows}


@router.get("/tables/{name}/export.xlsx")
def export_table_xlsx(name: str) -> Response:
    """Table XLSX 내보내기 (F4 — M-3-7 '⬆ Export' 실배선, D8 XLSX 트랙 1호)."""
    from openpyxl import Workbook
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, columns = _table_id(cur, tid, name)
        cur.execute(
            """SELECT row_key, row_values FROM tbl_data_row
               WHERE table_id=%s ORDER BY row_key_num NULLS LAST, row_key""", (table_id,))
        rows = cur.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]
    ws.append(["Key", *columns])
    for key, values in rows:
        ws.append([key, *[values.get(c) for c in columns]])
    buf = io.BytesIO()
    wb.save(buf)
    from urllib.parse import quote
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(name)}.xlsx"})


class TableRowRequest(BaseModel):
    key: str
    values: dict[str, float | int | None] = {}
    baseValues: dict[str, float | int | None] | None = None   # D9 확대 — 편집 시작 시점 스냅샷 (불일치 409)


@router.post("/tables/{name}/rows", status_code=201, dependencies=[SETUP])
def add_table_row(name: str, body: TableRowRequest) -> dict[str, Any]:
    key = body.key.strip()
    if not key:
        raise HTTPException(422, detail="Key 필수")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, _ = _table_id(cur, tid, name)
        cur.execute("SELECT 1 FROM tbl_data_row WHERE table_id=%s AND row_key=%s",
                    (table_id, key))
        if cur.fetchone():
            raise HTTPException(409, detail=f"Key 중복: {key}")
        num = float(key) if key.replace(".", "", 1).isdigit() else None
        cur.execute(
            """INSERT INTO tbl_data_row (table_id, row_key, row_key_num, row_values)
               VALUES (%s,%s,%s,%s)""",
            (table_id, key, num, json.dumps({k: v for k, v in body.values.items() if v is not None})))
    return {"key": key}


@router.put("/tables/{name}/rows/{key}", dependencies=[SETUP])
def update_table_row(name: str, key: str, body: TableRowRequest) -> dict[str, Any]:
    """행 저장 — baseValues(편집 시작 스냅샷) 전달 시 낙관적 잠금: 타인 선수정이면 409 (D9 확대)."""
    values = {k: v for k, v in body.values.items() if v is not None}
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, _ = _table_id(cur, tid, name)
        if body.baseValues is not None:
            cur.execute("SELECT row_values FROM tbl_data_row WHERE table_id=%s AND row_key=%s",
                        (table_id, key))
            cur_row = cur.fetchone()
            if not cur_row:
                raise HTTPException(404, detail=f"row not found: {key}")
            base = {k: v for k, v in body.baseValues.items() if v is not None}
            current = {k: v for k, v in (cur_row[0] or {}).items() if v is not None}
            if {k: float(v) for k, v in current.items()} != {k: float(v) for k, v in base.items()}:
                raise HTTPException(409, detail="동시 수정 충돌 — 다른 사용자가 먼저 수정했습니다 (재조회 후 재시도)")
        cur.execute(
            """UPDATE tbl_data_row SET row_values=%s
               WHERE table_id=%s AND row_key=%s RETURNING row_id""",
            (json.dumps(values), table_id, key))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"row not found: {key}")
    return {"key": key, "values": values}


@router.delete("/tables/{name}/rows/{key}", dependencies=[SETUP])
def delete_table_row(name: str, key: str) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, _ = _table_id(cur, tid, name)
        cur.execute("DELETE FROM tbl_data_row WHERE table_id=%s AND row_key=%s RETURNING row_id",
                    (table_id, key))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"row not found: {key}")
    return {"deleted": key}


@router.post("/tables/{name}/import-excel", dependencies=[SETUP])
async def import_excel(name: str, uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """정형 양식(1행 헤더 = Key + 열 이름) — Key 중복은 갱신, 수치 아닌 셀은 무시."""
    import openpyxl
    data = await uploadedFile.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(422, detail="Excel 파일이 아닙니다 (.xlsx)")
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    inserted = updated = 0
    rejected: list[str] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        table_id, columns = _table_id(cur, tid, name)
        col_idx = {c: header.index(c) for c in columns if c in header}
        if not col_idx:
            raise HTTPException(422, detail=f"헤더 불일치 — 필요 열: {columns}")
        for row in ws.iter_rows(min_row=2):
            key = str(row[0].value).strip() if row[0].value is not None else ""
            if not key:
                continue
            values: dict[str, float] = {}
            for col, idx in col_idx.items():
                v = row[idx].value
                if isinstance(v, (int, float)):
                    values[col] = v
                elif v is not None:
                    rejected.append(f"{key}.{col}: 수치 아님 ({v})")
            num = float(key) if key.replace(".", "", 1).isdigit() else None
            cur.execute(
                """INSERT INTO tbl_data_row (table_id, row_key, row_key_num, row_values)
                   VALUES (%s,%s,%s,%s)
                   ON CONFLICT (table_id, row_key)
                   DO UPDATE SET row_values = tbl_data_row.row_values || EXCLUDED.row_values
                   RETURNING (xmax = 0)""",
                (table_id, key, num, json.dumps(values)))
            if cur.fetchone()[0]:
                inserted += 1
            else:
                updated += 1
    return {"inserted": inserted, "updated": updated, "rejected": rejected}


# ── AI-04/06 — Prompt→Macro · UI 초안 (ANTHROPIC_API_KEY 없으면 sample 모드) ──
class AiMacroRequest(BaseModel):
    prompt: str


@router.post("/ai/macro-generate", dependencies=[SETUP])
def ai_macro_generate(body: AiMacroRequest) -> dict[str, Any]:
    from app.services.ai_assist import generate_macro
    if not body.prompt.strip():
        raise HTTPException(422, detail="Prompt 를 입력하십시오")
    return generate_macro(body.prompt.strip())


class AiChatRequest(BaseModel):
    question: str


@router.post("/ai/chat")
def ai_chat(body: AiChatRequest) -> dict[str, Any]:
    """U28 (s27 노트 'AI 질의 응답 — 내부 자료 검색·응답용') 1단계.
    항시: 키워드 기반 내부 자산 검색(코드·문서·Table·Macro·부품) → 근거 목록.
    live(키+크레딧): 검색 근거를 컨텍스트로 Claude 요약 답변. 폴백은 mode='search'."""
    q = body.question.strip()
    if not q:
        raise HTTPException(422, detail="질문을 입력하십시오")
    terms = [t for t in re.split(r"\s+", q) if len(t) >= 2][:5] or [q]
    like = [f"%{t}%" for t in terms]
    refs: list[dict[str, Any]] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)

        def _q(sql: str, kind: str, href_tpl: str) -> None:
            for t in like:
                cur.execute(sql, (tid, t, t))
                for code, title in cur.fetchall():
                    ref = {"kind": kind, "code": str(code), "title": str(title or ""),
                           "href": href_tpl.format(code=str(code))}
                    if ref not in refs:
                        refs.append(ref)

        _q("""SELECT main_code, code_name FROM product_code
              WHERE tenant_id=%s AND (main_code ILIKE %s OR code_name ILIKE %s) LIMIT 5""",
           "제품 코드", "/detail/code?code={code}")
        _q("""SELECT doc_no, title FROM doc_control
              WHERE tenant_id=%s AND (doc_no ILIKE %s OR title ILIKE %s) LIMIT 5""",
           "문서", "/cpq/documents")
        _q("""SELECT table_name, COALESCE(description,'') FROM tbl_data_table
              WHERE tenant_id=%s AND (table_name ILIKE %s OR description ILIKE %s) LIMIT 5""",
           "데이터 Table", "/code/datatable?name={code}")
        _q("""SELECT macro_name, COALESCE(description_text,'') FROM tbx_macro
              WHERE tenant_id=%s AND (macro_name ILIKE %s OR description_text ILIKE %s) LIMIT 5""",
           "Macro", "/toolbox/macros")
        _q("""SELECT part_no, part_name FROM prt_part
              WHERE tenant_id=%s AND (part_no ILIKE %s OR part_name ILIKE %s) LIMIT 5""",
           "부품", "/plm/parts")
        # U10 선행 — 도면 대장 검색 (도면번호·도면명, AI 도면 학습 트랙의 검색 기반)
        _q("""SELECT drawing_no, drawing_name FROM dwg_drawing
              WHERE tenant_id=%s AND (drawing_no ILIKE %s OR drawing_name ILIKE %s) LIMIT 5""",
           "도면", "/plm/drawings")
        # U10 2단계 — 도면 내용(추출 텍스트) 검색: 파일 뷰어 딥링크
        for t in like:
            cur.execute(
                """SELECT f.file_id, f.file_name FROM dwg_text_index ti
                   JOIN dwg_file f ON f.file_id = ti.file_id
                   WHERE ti.tenant_id=%s AND ti.content ILIKE %s LIMIT 4""", (tid, t))
            for fid, fname in cur.fetchall():
                ref = {"kind": "도면 내용", "code": str(fname), "title": f"file {fid} — 추출 텍스트 일치",
                       "href": f"/detail/cad-viewer?fileId={fid}"}
                if ref not in refs:
                    refs.append(ref)
    refs = refs[:18]

    from app.services.ai_assist import _client
    client = _client()
    if client is not None:
        try:
            ctx = "\n".join(f"- [{r['kind']}] {r['code']} : {r['title']}" for r in refs) or "(일치 자료 없음)"
            msg = client.messages.create(
                model=settings.anthropic_model_id, max_tokens=600,
                system=("EDIM(제조 CPQ/PLM/ERP 플랫폼) 내부 자료 안내 도우미. 아래 검색된 내부 자산 목록만 근거로 "
                        "한국어 2~4문장으로 답하고, 목록에 없는 내용은 추정하지 말 것.\n검색 결과:\n" + ctx),
                messages=[{"role": "user", "content": q}],
            )
            return {"mode": "live", "answer": msg.content[0].text[:1200], "refs": refs}
        except Exception as e:  # noqa: BLE001
            return {"mode": "error", "error": str(e)[:200],
                    "answer": f"검색 일치 {len(refs)}건 — 아래 근거를 확인하십시오. (AI 합성은 크레딧 준비 후)", "refs": refs}
    return {"mode": "search", "answer": f"검색 일치 {len(refs)}건 — 아래 근거를 확인하십시오. (AI 합성은 키 설정 후)", "refs": refs}


class AiUiRequest(BaseModel):
    description: str


@router.post("/ai/ui-suggest", dependencies=[SETUP])
def ai_ui_suggest(body: AiUiRequest) -> dict[str, Any]:
    from app.services.ai_assist import suggest_ui
    if not body.description.strip():
        raise HTTPException(422, detail="설명을 입력하십시오")
    return suggest_ui(body.description.strip())


# ── ENG-01 Macro 실행 엔진 v1 ──
def _make_table_resolver(cur, tid: int):
    cache: dict[str, list[tuple[float | None, str, dict]]] = {}

    def load(table: str):
        if table in cache:
            return cache[table]
        cur.execute(
            """SELECT r.row_key_num, r.row_key, r.row_values
               FROM tbl_data_table t JOIN tbl_data_row r ON r.table_id=t.table_id
               WHERE t.tenant_id=%s AND t.table_name=%s
               ORDER BY r.row_key_num NULLS LAST""", (tid, table))
        rows = [(float(r[0]) if r[0] is not None else None, r[1], r[2]) for r in cur.fetchall()]
        if not rows:
            raise MacroError(f"Table 없음 또는 빈 Table: {table}")
        cache[table] = rows
        return rows

    def resolve(table: str, col: str, key, agg: str) -> float:
        rows = load(table)
        if agg == "GET":
            for num, rkey, values in rows:
                if (isinstance(key, float) and num is not None and abs(num - key) < 1e-9) \
                        or rkey == str(key) or (isinstance(key, float) and rkey == f"{key:g}"):
                    if col not in values:
                        raise MacroError(f"{table}[{rkey}] 에 열 {col} 없음")
                    return float(values[col])
            key_disp = f"{key:g}" if isinstance(key, float) else str(key)
            raise MacroError(f"{table}: key {key_disp} 없음")
        lo, hi = key
        vals = [float(v[col]) for num, _, v in rows
                if num is not None and lo <= num <= hi and col in v]
        if not vals:
            raise MacroError(f"{table}({col},{lo:g}:{hi:g}) 범위에 값 없음")
        if agg == "SUM":
            return sum(vals)
        if agg == "AVG":
            return sum(vals) / len(vals)
        if agg == "MIN":
            return min(vals)
        if agg == "MAX":
            return max(vals)
        return float(len(vals))

    return resolve


class EvaluateRequest(BaseModel):
    formula: str
    variables: dict[str, float] = {}


@router.post("/macros/evaluate")
def evaluate_macro(body: EvaluateRequest) -> dict[str, Any]:
    """Excel 호환 Macro 평가 — Table 참조는 실 tbl_data_row (TBX-011)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        ev = Evaluator(body.variables, _make_table_resolver(cur, tid))
        try:
            value = ev.run(body.formula)
        except MacroError as e:
            return {"ok": False, "error": str(e), "trace": ev.trace}
    return {"ok": True, "value": value, "trace": ev.trace}


# ── SVC-12 파일 업/다운로드 (MinIO 프록시) ──
@router.post("/files/upload", status_code=201, dependencies=[SETUP])
async def upload_file(
    request: Request,
    uploadedFile: UploadFile = File(...),
    folder: str = Form("RECEIVED"),
    project: str = Form("PS-61313-5"),
) -> dict[str, Any]:
    if folder not in ("DWG", "PRICE", "DATA", "BOM", "RECEIVED"):
        raise HTTPException(422, detail=f"folder 오류: {folder}")
    # 실행 파일 차단 (B15 — 업로드 에러 케이스)
    blocked = (".exe", ".bat", ".cmd", ".msi", ".scr", ".ps1", ".sh", ".dll", ".com")
    if (uploadedFile.filename or "").lower().endswith(blocked):
        raise HTTPException(422, detail=f"허용되지 않는 파일 형식: {uploadedFile.filename}")
    data = await uploadedFile.read()
    if len(data) > 100 * 1024 * 1024:
        raise HTTPException(413, detail="100MB 초과")
    fname = (uploadedFile.filename or "file").replace("/", "_")
    key = f"{project}/{folder}/{fname}"
    try:
        storage.put_object(key, data, uploadedFile.content_type or "application/octet-stream")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT project_id FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, project))
        prj = cur.fetchone()
        cur.execute(
            """INSERT INTO dwg_file (tenant_id, project_id, folder, file_name, file_type,
               file_path, file_size, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING file_id""",
            (tid, prj[0] if prj else None, folder, fname,
             (fname.rsplit(".", 1)[-1] if "." in fname else "BIN").upper()[:10],
             key, len(data), request.state.login))
        file_id = cur.fetchone()[0]
    return {"fileId": file_id, "key": key, "size": len(data)}


@router.get("/files/download/{file_id}")
def download_file(file_id: int) -> StreamingResponse:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
            (tid, file_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"file not found: {file_id}")
    try:
        obj = storage.get_object(row[0])
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    from urllib.parse import quote
    return StreamingResponse(
        obj.stream(64 * 1024), media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(row[1])}"})


def _stamp_pdf_confidential(data: bytes) -> bytes:
    """기존 PDF 에 CONFIDENTIAL 전면 대각 워터마크 오버레이 (B4 잔여 — 고객 전달 파일별 실적용, DOC-002)."""
    import io as _io
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as rl_canvas
    from app.services.run_pipeline import _draw_watermark
    reader = PdfReader(_io.BytesIO(data))
    writer = PdfWriter()
    overlays: dict[tuple[int, int], Any] = {}
    for page in reader.pages:
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        key = (round(w), round(h))
        if key not in overlays:
            ob = _io.BytesIO()
            c = rl_canvas.Canvas(ob, pagesize=(w, h))
            _draw_watermark(c, w, h, "CONFIDENTIAL")
            c.showPage()
            c.save()
            overlays[key] = PdfReader(_io.BytesIO(ob.getvalue())).pages[0]
        page.merge_page(overlays[key])
        writer.add_page(page)
    out = _io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _zip_files(rows: list[tuple], prefix_by_folder: bool = True,
               extra: dict[str, bytes] | None = None,
               stamp_pdf: bool = False) -> tuple[bytes, int]:
    """(file_path, file_name, folder[, ...]) 행 → (ZIP 바이트, 워터마크 적용 수). MinIO 수집, arcname 중복 회피.
    stamp_pdf=True 면 .pdf 파일에 CONFIDENTIAL 워터마크 적용 (실패 시 원본 유지)."""
    import io
    import zipfile
    buf = io.BytesIO()
    used: dict[str, int] = {}
    stamped = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in (extra or {}).items():
            zf.writestr(name, data)
        for r in rows:
            path, name, folder = r[0], r[1], r[2]
            try:
                obj = storage.get_object(path)
                data = obj.read()
                obj.close()
                obj.release_conn()
            except Exception:
                continue   # 누락 객체는 건너뜀 (부분 다운로드)
            if stamp_pdf and name.lower().endswith(".pdf"):
                try:
                    data = _stamp_pdf_confidential(data)
                    stamped += 1
                except Exception:
                    pass   # 손상 PDF 등 — 원본 그대로 포함 (부분 실패 허용)
            arc = f"{folder}/{name}" if prefix_by_folder else name
            n = used.get(arc, 0)
            used[arc] = n + 1
            if n:
                base, dot, ext = name.rpartition(".")
                arc = f"{folder}/{n}_{name}" if prefix_by_folder else f"{n}_{name}"
            zf.writestr(arc, data)
    return buf.getvalue(), stamped


@router.get("/files/zip")
def files_zip(project: str = "PS-61313-5", folder: str = "") -> StreamingResponse:
    """폴더 파일 일괄 ZIP 다운로드 (E2) — 선택 폴더(또는 전체)의 dwg_file 을 MinIO 에서 수집."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        q = ("""SELECT f.file_path, f.file_name, f.folder FROM dwg_file f
                JOIN prj_project p ON p.project_id=f.project_id AND p.tenant_id=%s
                WHERE p.project_no=%s""")
        params: list[Any] = [tid, project]
        if folder.strip():
            q += " AND f.folder=%s"
            params.append(folder.strip())
        q += " ORDER BY f.folder, f.file_id"
        cur.execute(q, tuple(params))
        rows = cur.fetchall()
    if not rows:
        raise HTTPException(404, detail="다운로드할 파일이 없습니다")
    blob, _ = _zip_files(rows)
    from urllib.parse import quote
    zipname = f"{project}{('_' + folder.strip()) if folder.strip() else ''}.zip"
    return StreamingResponse(iter([blob]), media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(zipname)}",
                                      "X-File-Count": str(len(rows))})


# 고객 전달 제외 폴더 — 내부 접수자료(RECEIVED)는 산출물이 아니므로 미포함
DELIVER_FOLDERS = ["DWG", "PRICE", "DATA", "BOM"]


@router.get("/files/export-package")
def files_export_package(project: str = "PS-61313-5") -> StreamingResponse:
    """고객 전달용 내보내기 (E2) — 산출물 폴더만 ZIP + 전달 매니페스트. 내부 접수자료 제외."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT f.file_path, f.file_name, f.folder, to_char(f.created_at,'YYYY-MM-DD')
               FROM dwg_file f
               JOIN prj_project p ON p.project_id=f.project_id AND p.tenant_id=%s
               WHERE p.project_no=%s AND f.folder = ANY(%s)
               ORDER BY f.folder, f.file_id""",
            (tid, project, DELIVER_FOLDERS))
        rows = cur.fetchall()
        cur.execute("SELECT project_name FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, project))
        pn = cur.fetchone()
    if not rows:
        raise HTTPException(404, detail="전달할 산출물이 없습니다 (DWG/PRICE/DATA/BOM)")
    manifest = [
        "EDIM 고객 전달 패키지 (Customer Delivery Package)",
        f"프로젝트: {project}" + (f" — {pn[0]}" if pn else ""),
        f"파일 수: {len(rows)}건",
        "포함: 산출물(DWG/PRICE/DATA/BOM)  ·  제외: 내부 접수자료(RECEIVED)·S-1/S-2 등급 문서",
        "PDF 산출물: CONFIDENTIAL 전면 워터마크 적용본 (DOC-002)",
        "-" * 56,
    ]
    manifest += [f"[{r[2]}] {r[1]}  ({r[3]})" for r in rows]
    extra = {"전달목록.txt": ("\n".join(manifest)).encode("utf-8")}
    blob, stamped = _zip_files(rows, extra=extra, stamp_pdf=True)
    from urllib.parse import quote
    zipname = f"{project}_고객전달.zip"
    return StreamingResponse(iter([blob]), media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(zipname)}",
                                      "X-File-Count": str(len(rows)),
                                      "X-Watermarked": str(stamped)})


@router.delete("/files/{file_id}", dependencies=[SETUP])
def delete_file(file_id: int, request: Request) -> dict[str, Any]:
    """파일 삭제 — Run 산출물·견적서가 참조 중이면 409 (MinIO 객체 + dwg_file 행 제거)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
            (tid, file_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"file not found: {file_id}")
        cur.execute("SELECT 1 FROM cpq_output WHERE file_id=%s LIMIT 1", (file_id,))
        if cur.fetchone():
            raise HTTPException(409, detail="Run 산출물이 참조하는 파일은 삭제 불가")
        cur.execute("SELECT 1 FROM cst_quotation WHERE doc_file_id=%s LIMIT 1", (file_id,))
        if cur.fetchone():
            raise HTTPException(409, detail="견적서가 참조하는 파일은 삭제 불가")
        cur.execute("DELETE FROM dwg_file WHERE tenant_id=%s AND file_id=%s", (tid, file_id))
        # 동일 key 를 공유하는 다른 행(재업로드 덮어쓰기)이 남아 있으면 객체는 보존
        cur.execute("SELECT 1 FROM dwg_file WHERE tenant_id=%s AND file_path=%s LIMIT 1",
                    (tid, row[0]))
        key_shared = cur.fetchone() is not None
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_file',%s,'DELETE',%s,%s)""",
            (tid, file_id, request.state.user_id, json.dumps({"fileName": row[1]})))
    if not key_shared:
        try:
            storage.remove_object(row[0])
        except RuntimeError:
            pass  # 스토리지 불가 시 DB 정리만 — 객체는 orphan (버킷 정리 배치 대상)
    return {"deleted": file_id}


class FileGcRequest(BaseModel):
    apply: bool = False    # false=dry-run(목록만), true=실제 삭제
    prefix: str = ""       # 특정 프로젝트/폴더 접두어로 한정 (선택)


@router.post("/files/gc", dependencies=[ADMIN])
def files_gc(request: Request, body: FileGcRequest) -> dict[str, Any]:
    """MinIO 객체 GC — dwg_file 미참조 orphan 정리 (E2/E3 잔여). 기본 dry-run, apply=true 실삭제."""
    try:
        keys = storage.list_object_keys(body.prefix.strip())
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        # 버킷은 테넌트 공유 — MinIO 키를 저장하는 모든 테이블을 참조 집합으로 합집합
        referenced: set[str] = set()
        cur.execute("SELECT DISTINCT file_path FROM dwg_file WHERE file_path IS NOT NULL")
        referenced.update(r[0] for r in cur.fetchall())
        cur.execute("SELECT DISTINCT file_path FROM dev_requirement_image WHERE file_path IS NOT NULL")
        referenced.update(r[0] for r in cur.fetchall())
        # 시스템 샘플/시드 오브젝트는 보호 (GC 대상 제외)
        PROTECTED_PREFIX = ("sample_data/", "sample_")
        orphans = [k for k in keys
                   if k not in referenced and not k.startswith(PROTECTED_PREFIX)]
        removed: list[str] = []
        if body.apply:
            for k in orphans:
                try:
                    storage.remove_object(k)
                    removed.append(k)
                except RuntimeError:
                    break
            if removed:
                _audit(cur, tid, "dwg_file", 0, "MINIO_GC", request.state.user_id,
                       {"removed": len(removed), "prefix": body.prefix})
    return {"totalObjects": len(keys), "referenced": len(keys) - len(orphans),
            "orphans": len(orphans), "removed": len(removed), "applied": body.apply,
            "sampleOrphans": orphans[:20]}


# ── INT-04 CAD 호환 — DXF 뷰/Import/Export (DWG 는 ODA 플러그블) ──
def _parse_cad_bytes(data: bytes, file_name: str) -> dict[str, Any]:
    """DXF(직접)/DWG(ODA 변환 후) → 정규화 DrawingDocument(JSON)."""
    import tempfile
    from pathlib import Path

    from app.services.dwg_converter import get_configured_dwg_converter
    from app.services.dxf_importer import convert_dxf_to_drawing_document

    ext = Path(file_name).suffix.lower()
    if ext not in (".dxf", ".dwg"):
        raise HTTPException(422, detail=f"지원하지 않는 CAD 형식: {ext} (.dxf/.dwg)")
    with tempfile.TemporaryDirectory(prefix="edim_cad_") as tmp:
        src = Path(tmp) / Path(file_name).name
        src.write_bytes(data)
        dxf_path = str(src)
        source_format = "dxf"
        if ext == ".dwg":
            dxf_path = get_configured_dwg_converter().convert(str(src))  # 미설정 → 501
            source_format = "dwg"
        doc = convert_dxf_to_drawing_document(dxf_path, Path(file_name).stem, source_format)
    return doc.model_dump()


@router.get("/cad/view/{file_id}")
def cad_view(file_id: int) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
            (tid, file_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"file not found: {file_id}")
    try:
        obj = storage.get_object(row[0])
        data = obj.read()
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    return {"fileId": file_id, "document": _parse_cad_bytes(data, row[1])}


def _build_cad_plot_pdf(doc: dict[str, Any], scale: float, paper: str, orient: str) -> bytes:
    """정규화 DrawingDocument → 축척(1:scale) 벡터 PDF (reportlab). 단위=mm 가정."""
    import io as _io
    import math as _m

    from reportlab.lib.pagesizes import A3, A4, landscape, portrait
    from reportlab.pdfgen import canvas

    pg = {"A4": A4, "A3": A3}.get(paper.upper(), A4)
    pg = landscape(pg) if orient == "landscape" else portrait(pg)
    W, H = pg
    buf = _io.BytesIO()
    c = canvas.Canvas(buf, pagesize=pg)

    ents = doc.get("entities", [])
    b = doc.get("bounds", {}) or {}
    minx, miny = b.get("minX", 0.0), b.get("minY", 0.0)
    maxx, maxy = b.get("maxX", 0.0), b.get("maxY", 0.0)
    k = (72.0 / 25.4) / scale                       # 도면 단위(mm) → pt
    ox = W / 2 - (minx + maxx) / 2 * k              # 도면 중심 → 용지 중심
    oy = H / 2 - (miny + maxy) / 2 * k

    def X(x: float) -> float:
        return ox + x * k

    def Y(y: float) -> float:
        return oy + y * k

    c.setLineWidth(0.3)
    c.setStrokeColorRGB(0.16, 0.23, 0.33)
    for e in ents:
        t = e.get("entityType")
        try:
            if t == "line":
                c.line(X(e["startPoint"]["x"]), Y(e["startPoint"]["y"]),
                       X(e["endPoint"]["x"]), Y(e["endPoint"]["y"]))
            elif t == "circle":
                c.circle(X(e["centerPoint"]["x"]), Y(e["centerPoint"]["y"]), e["radius"] * k, stroke=1, fill=0)
            elif t == "arc":
                cx, cy, r = e["centerPoint"]["x"], e["centerPoint"]["y"], e["radius"]
                a0, a1 = e["startAngleDegrees"], e["endAngleDegrees"]
                if a1 < a0:
                    a1 += 360
                steps = max(6, int((a1 - a0) / 6))
                path = c.beginPath()
                for i in range(steps + 1):
                    ang = _m.radians(a0 + (a1 - a0) * i / steps)
                    px, py = X(cx + r * _m.cos(ang)), Y(cy + r * _m.sin(ang))
                    (path.moveTo if i == 0 else path.lineTo)(px, py)
                c.drawPath(path)
            elif t == "polyline":
                vs = e.get("vertexPoints", [])
                if len(vs) >= 2:
                    path = c.beginPath()
                    path.moveTo(X(vs[0]["x"]), Y(vs[0]["y"]))
                    for v in vs[1:]:
                        path.lineTo(X(v["x"]), Y(v["y"]))
                    if e.get("isClosed"):
                        path.close()
                    c.drawPath(path)
            elif t == "text":
                ip = e["insertionPoint"]
                c.setFont("Helvetica", max(4.0, (e.get("textHeight") or 10) * k))
                c.drawString(X(ip["x"]), Y(ip["y"]), str(e.get("textContent") or ""))
        except (KeyError, TypeError):
            continue

    c.setLineWidth(0.6)
    c.setStrokeColorRGB(0.16, 0.23, 0.33)
    c.rect(10, 10, W - 20, H - 20)
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.16, 0.23, 0.33)
    c.drawString(16, 15, f"SCALE 1:{int(scale)}  ·  {doc.get('drawingName', '')}  ·  units {doc.get('units', '')}  ·  {paper.upper()} {orient}")
    c.showPage()
    c.save()
    return buf.getvalue()


@router.get("/cad/view/{file_id}/plot.pdf")
def cad_plot(file_id: int, scale: float = 100, paper: str = "A4",
             orient: str = "landscape") -> StreamingResponse:
    """축척 인쇄(plot to scale) — DXF 를 1:scale 벡터 PDF 로 출력."""
    import io as _io
    from pathlib import Path

    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
            (tid, file_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"file not found: {file_id}")
    try:
        data = storage.get_object(row[0]).read()
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    if scale <= 0:
        scale = 100
    doc = _parse_cad_bytes(data, row[1])
    pdf = _build_cad_plot_pdf(doc, scale, paper, orient)
    stem = Path(row[1]).stem
    return StreamingResponse(
        _io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{stem}_1-{int(scale)}.pdf"'})


class CadEditOp(BaseModel):
    op: str            # 'move' | 'delete' | 'copy' | 'rotate' | 'mirror' | 'add' | 'trim'
    entityId: str = ""
    boundaryId: str = ""  # trim/연장 — 경계선 엔티티(교점 계산)
    dx: float = 0.0
    dy: float = 0.0
    angle: float = 0.0    # rotate — 도(°), 반시계
    axis: str = "y"       # mirror — 'y'=수직축(좌우반전) · 'x'=수평축(상하반전)
    # add(자유 작도) — line/circle/rect/dim/block
    entityType: str = ""  # 'line' | 'circle' | 'rect' | 'dim' | 'block'
    layer: str = "0"
    x1: float = 0.0
    y1: float = 0.0
    x2: float = 0.0
    y2: float = 0.0
    radius: float = 0.0
    text: str = ""        # dim=자동거리 · block=라벨


class CadEditRequest(BaseModel):
    ops: list[CadEditOp]


# importer 와 동일한 인식 타입 집합 — entityId(e1..) 순번 규칙 일치 보장
_CAD_EDIT_TYPES = ("LINE", "LWPOLYLINE", "POLYLINE", "CIRCLE", "ARC", "TEXT", "MTEXT")


@router.post("/cad/view/{file_id}/edit", dependencies=[SETUP])
def cad_edit(file_id: int, request: Request, body: CadEditRequest) -> dict[str, Any]:
    """G1 엔티티 편집 — 이동/삭제 → DXF 재저장(MinIO 덮어쓰기) 후 재파싱 문서 반환.
    entityId = 인식 엔티티(LINE/POLYLINE/CIRCLE/ARC/TEXT) 모델스페이스 순번(e1..) — importer 규칙과 동일."""
    import math
    import tempfile
    from pathlib import Path

    import ezdxf
    from ezdxf import recover
    from ezdxf.bbox import extents
    from ezdxf.math import Matrix44

    def _pivot(entity: Any) -> tuple[float, float] | None:
        bb = extents([entity])
        if not bb.has_data:
            return None
        return bb.center.x, bb.center.y

    def _line_isect(a: Any, b: Any) -> tuple[float, float] | None:
        """두 LINE 무한연장 교점 (평행/동일 시 None)."""
        x1, y1 = a.dxf.start.x, a.dxf.start.y
        x2, y2 = a.dxf.end.x, a.dxf.end.y
        x3, y3 = b.dxf.start.x, b.dxf.start.y
        x4, y4 = b.dxf.end.x, b.dxf.end.y
        d = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)
        if abs(d) < 1e-9:
            return None
        px = ((x1 * y2 - y1 * x2) * (x3 - x4) - (x1 - x2) * (x3 * y4 - y3 * x4)) / d
        py = ((x1 * y2 - y1 * x2) * (y3 - y4) - (y1 - y2) * (x3 * y4 - y3 * x4)) / d
        return px, py

    if not body.ops:
        raise HTTPException(422, detail="편집 작업이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
            (tid, file_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"file not found: {file_id}")
    file_path, file_name = row
    if Path(file_name).suffix.lower() != ".dxf":
        raise HTTPException(501, detail="편집은 DXF 만 지원 (DWG=ODA 변환 대기)")
    try:
        data = storage.get_object(file_path).read()
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")

    with tempfile.TemporaryDirectory(prefix="edim_cadedit_") as tmp:
        src = Path(tmp) / Path(file_name).name
        src.write_bytes(data)
        try:
            doc = ezdxf.readfile(str(src))
        except ezdxf.DXFStructureError:
            doc, _auditor = recover.readfile(str(src))
        msp = doc.modelspace()
        id_map: dict[str, Any] = {}
        seq = 0
        for entity in msp:
            if entity.dxftype() in _CAD_EDIT_TYPES:
                seq += 1
                id_map[f"e{seq}"] = entity
        applied = 0
        for op in body.ops:
            if op.op == "add":   # 자유 작도 — entityId 불요
                lyr = op.layer or "0"
                if lyr not in doc.layers:
                    try:
                        doc.layers.new(lyr)
                    except Exception:
                        pass
                attr = {"layer": lyr}
                try:
                    if op.entityType == "line":
                        msp.add_line((op.x1, op.y1), (op.x2, op.y2), dxfattribs=attr)
                        applied += 1
                    elif op.entityType == "circle":
                        r = op.radius or math.hypot(op.x2 - op.x1, op.y2 - op.y1)
                        if r > 0:
                            msp.add_circle((op.x1, op.y1), r, dxfattribs=attr)
                            applied += 1
                    elif op.entityType == "rect":
                        pts = [(op.x1, op.y1), (op.x2, op.y1), (op.x2, op.y2), (op.x1, op.y2)]
                        msp.add_lwpolyline(pts, close=True, dxfattribs=attr)
                        applied += 1
                    elif op.entityType == "dim":   # 치수 — 치수선 + 거리 텍스트(중점)
                        dlyr = {"layer": lyr if lyr != "0" else "DIM"}
                        if dlyr["layer"] not in doc.layers:
                            try:
                                doc.layers.new(dlyr["layer"])
                            except Exception:
                                pass
                        msp.add_line((op.x1, op.y1), (op.x2, op.y2), dxfattribs=dlyr)
                        d = math.hypot(op.x2 - op.x1, op.y2 - op.y1)
                        th = max(1.0, d * 0.06)
                        te = msp.add_text(f"{d:.0f}", height=th, dxfattribs=dlyr)
                        te.set_placement(((op.x1 + op.x2) / 2, (op.y1 + op.y2) / 2 + th * 0.3))
                        applied += 1
                    elif op.entityType == "block":   # 블록 삽입 — 라벨 박스(사각+텍스트)
                        blyr = {"layer": lyr if lyr != "0" else "BLOCK"}
                        if blyr["layer"] not in doc.layers:
                            try:
                                doc.layers.new(blyr["layer"])
                            except Exception:
                                pass
                        pts = [(op.x1, op.y1), (op.x2, op.y1), (op.x2, op.y2), (op.x1, op.y2)]
                        msp.add_lwpolyline(pts, close=True, dxfattribs=blyr)
                        label = (op.text or "BLOCK")[:40]
                        bh = max(1.0, abs(op.y2 - op.y1) * 0.25)
                        tb = msp.add_text(label, height=bh, dxfattribs=blyr)
                        tb.set_placement((min(op.x1, op.x2) + abs(op.x2 - op.x1) * 0.1,
                                          (op.y1 + op.y2) / 2))
                        applied += 1
                except Exception:
                    pass
                continue
            if op.op == "trim":   # 트림/연장 — 대상 선의 가까운 끝점을 경계선 교점으로 이동
                tgt = id_map.get(op.entityId)
                bnd = id_map.get(op.boundaryId)
                if tgt is None or bnd is None or tgt.dxftype() != "LINE" or bnd.dxftype() != "LINE":
                    continue
                inter = _line_isect(tgt, bnd)
                if inter is None:
                    continue
                s, e2 = tgt.dxf.start, tgt.dxf.end
                ds = math.hypot(s.x - op.x1, s.y - op.y1)
                de = math.hypot(e2.x - op.x1, e2.y - op.y1)
                if ds <= de:
                    tgt.dxf.start = (inter[0], inter[1], 0)
                else:
                    tgt.dxf.end = (inter[0], inter[1], 0)
                applied += 1
                continue
            ent = id_map.get(op.entityId)
            if ent is None:
                continue
            if op.op == "delete":
                msp.delete_entity(ent)
                applied += 1
            elif op.op == "move":
                try:
                    ent.translate(op.dx, op.dy, 0)   # 도면 좌표 델타
                    applied += 1
                except (AttributeError, TypeError):
                    try:
                        ent.transform(Matrix44.translate(op.dx, op.dy, 0))
                        applied += 1
                    except Exception:
                        pass
            elif op.op == "copy":
                try:
                    new = ent.copy()
                    msp.add_entity(new)
                    new.translate(op.dx, op.dy, 0)
                    applied += 1
                except Exception:
                    pass
            elif op.op == "rotate":
                piv = _pivot(ent)
                if piv:
                    cx, cy = piv
                    m = Matrix44.chain(
                        Matrix44.translate(-cx, -cy, 0),
                        Matrix44.z_rotate(math.radians(op.angle)),
                        Matrix44.translate(cx, cy, 0))
                    try:
                        ent.transform(m)
                        applied += 1
                    except Exception:
                        pass
            elif op.op == "mirror":
                piv = _pivot(ent)
                if piv:
                    cx, cy = piv
                    sx, sy = (-1, 1) if op.axis == "y" else (1, -1)
                    m = Matrix44.chain(
                        Matrix44.translate(-cx, -cy, 0),
                        Matrix44.scale(sx, sy, 1),
                        Matrix44.translate(cx, cy, 0))
                    try:
                        ent.transform(m)
                        applied += 1
                    except Exception:
                        pass
        if applied == 0:
            raise HTTPException(422, detail="적용 가능한 편집 대상이 없습니다 (entityId 확인)")
        out = Path(tmp) / "edited.dxf"
        doc.saveas(str(out))
        new_bytes = out.read_bytes()

    try:
        storage.put_object(file_path, new_bytes, "application/dxf")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("UPDATE dwg_file SET file_size=%s WHERE tenant_id=%s AND file_id=%s",
                    (len(new_bytes), tid, file_id))
        _audit(cur, tid, "dwg_file", file_id, "CAD_EDIT", request.state.user_id,
               {"ops": len(body.ops), "applied": applied})
    return {"fileId": file_id, "applied": applied,
            "document": _parse_cad_bytes(new_bytes, file_name)}


# ── U2 — Block 단위 저장·상위 호출 (s63: "저장 > Block 단위로 저장 (상위 호출 시 Block 상태)") ──
class BlockCreate(BaseModel):
    name: str
    entityIds: list[str] = []


class BlockInsertReq(BaseModel):
    name: str
    x: float = 0.0
    y: float = 0.0


def _load_dwg_dxf(file_id: int) -> tuple[bytes, str, str]:
    from pathlib import Path
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT file_path, file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s",
                    (tid, file_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"file not found: {file_id}")
    file_path, file_name = row
    if Path(file_name).suffix.lower() != ".dxf":
        raise HTTPException(501, detail="Block 은 DXF 만 지원")
    try:
        data = storage.get_object(file_path).read()
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    return data, file_path, file_name


def _store_dwg_dxf(file_id: int, file_path: str, new_bytes: bytes, user_id: str, action: str, detail: dict) -> None:
    try:
        storage.put_object(file_path, new_bytes, "application/dxf")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("UPDATE dwg_file SET file_size=%s WHERE tenant_id=%s AND file_id=%s",
                    (len(new_bytes), tid, file_id))
        _audit(cur, tid, "dwg_file", file_id, action, user_id, detail)


def _open_dxf(tmp: str, file_name: str, data: bytes):
    from pathlib import Path
    import ezdxf
    from ezdxf import recover
    src = Path(tmp) / Path(file_name).name
    src.write_bytes(data)
    try:
        return ezdxf.readfile(str(src))
    except ezdxf.DXFStructureError:
        doc, _a = recover.readfile(str(src))
        return doc


def _extract_dxf_texts(data: bytes, file_name: str) -> tuple[str, int]:
    """U10 선행 — DXF 에서 TEXT/MTEXT·레이어명 추출 (검색 인덱스용)."""
    import tempfile
    doc_texts: list[str] = []
    count = 0
    with tempfile.TemporaryDirectory(prefix="edim_idx_") as tmp:
        doc = _open_dxf(tmp, file_name, data)
        for e in doc.modelspace():
            count += 1
            t = e.dxftype()
            if t == "TEXT":
                doc_texts.append(str(e.dxf.text))
            elif t == "MTEXT":
                doc_texts.append(str(e.text))
        doc_texts.extend(str(l.dxf.name) for l in doc.layers)
    content = " ".join(x.strip() for x in doc_texts if x and x.strip())[:20000]
    return content, count


@router.get("/cad/view/{file_id}/related-codes")
def cad_related_codes(file_id: int) -> dict[str, Any]:
    """U10 3단계 — 도면-코드 자동 연결 (s25 '부품명 규칙 기반 분류'):
    파일명 + 인덱스 텍스트에서 제품 코드(main_code) 매칭."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT file_name FROM dwg_file WHERE tenant_id=%s AND file_id=%s", (tid, file_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"file not found: {file_id}")
        fname = str(row[0])
        cur.execute("SELECT COALESCE(content,'') FROM dwg_text_index WHERE tenant_id=%s AND file_id=%s",
                    (tid, file_id))
        idx = cur.fetchone()
        content = str(idx[0]) if idx else ""
        haystack = f"{fname} {content}".upper()
        compact = haystack.replace(" ", "").replace("-", "").replace("_", "")
        cur.execute("SELECT main_code, code_name FROM product_code WHERE tenant_id=%s", (tid,))
        related = []
        for code, name in cur.fetchall():
            c = str(code)
            key = c.upper().replace(" ", "").replace("-", "")
            if len(key) >= 4 and key in compact:
                related.append({"code": c, "name": str(name or ""),
                                "href": f"/detail/code?code={c}"})
    return {"fileId": file_id, "indexed": bool(content), "codes": related[:8]}


@router.post("/cad/view/{file_id}/index", dependencies=[SETUP])
def cad_text_index(file_id: int) -> dict[str, Any]:
    """도면 텍스트 인덱싱 — TEXT/MTEXT·레이어명 추출 → dwg_text_index upsert."""
    data, _fp, file_name = _load_dwg_dxf(file_id)
    content, count = _extract_dxf_texts(data, file_name)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """INSERT INTO dwg_text_index (file_id, tenant_id, content, entity_count, indexed_at)
               VALUES (%s,%s,%s,%s,now())
               ON CONFLICT (file_id) DO UPDATE SET content=EXCLUDED.content,
                   entity_count=EXCLUDED.entity_count, indexed_at=now()""",
            (file_id, tid, content, count))
    return {"fileId": file_id, "chars": len(content), "entities": count,
            "preview": content[:120]}


@router.post("/cad/index-all", dependencies=[SETUP])
def cad_index_all() -> dict[str, Any]:
    """도면 텍스트 일괄 인덱싱 — DXF 파일 전량 (최대 50건/회)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT file_id, file_name FROM dwg_file
               WHERE tenant_id=%s AND lower(file_name) LIKE '%%.dxf'
               ORDER BY file_id LIMIT 50""", (tid,))
        targets = cur.fetchall()
    done, failed = [], []
    for fid, _fn in targets:
        try:
            r = cad_text_index(fid)
            done.append({"fileId": fid, "chars": r["chars"]})
        except HTTPException:
            failed.append(fid)
    return {"indexed": len(done), "failed": failed, "files": done[:10]}


@router.get("/cad/view/{file_id}/blocks")
def cad_blocks(file_id: int) -> dict[str, Any]:
    """사용자 Block 목록 (*레이아웃 제외)."""
    import tempfile
    data, _fp, file_name = _load_dwg_dxf(file_id)
    with tempfile.TemporaryDirectory(prefix="edim_blk_") as tmp:
        doc = _open_dxf(tmp, file_name, data)
        out = [{"name": b.name, "entities": len(list(b))}
               for b in doc.blocks if not b.name.startswith(("*", "_"))]
    return {"blocks": out}


@router.post("/cad/view/{file_id}/blocks", dependencies=[SETUP])
def cad_block_create(file_id: int, request: Request, body: BlockCreate) -> dict[str, Any]:
    """선택 엔티티 → 명명 Block 등록 (원자 엔티티는 Block 참조로 대체 — 원위치 INSERT)."""
    import re as _re
    import tempfile
    from pathlib import Path

    from ezdxf.bbox import extents
    from ezdxf.math import Matrix44

    name = body.name.strip()
    if not _re.fullmatch(r"[A-Za-z0-9_\-]{1,30}", name):
        raise HTTPException(422, detail="Block 이름: 영문/숫자/_- 1~30자")
    if not body.entityIds:
        raise HTTPException(422, detail="선택 엔티티가 없습니다")
    data, file_path, file_name = _load_dwg_dxf(file_id)
    with tempfile.TemporaryDirectory(prefix="edim_blk_") as tmp:
        doc = _open_dxf(tmp, file_name, data)
        if name in doc.blocks:
            raise HTTPException(409, detail=f"Block 중복: {name}")
        msp = doc.modelspace()
        id_map: dict[str, Any] = {}
        seq = 0
        for entity in msp:
            if entity.dxftype() in _CAD_EDIT_TYPES:
                seq += 1
                id_map[f"e{seq}"] = entity
        picked = [id_map[i] for i in body.entityIds if i in id_map]
        if not picked:
            raise HTTPException(422, detail="적용 가능한 엔티티가 없습니다 (entityId 확인)")
        bb = extents(picked)
        bx, by = (bb.extmin.x, bb.extmin.y) if bb.has_data else (0.0, 0.0)
        blk = doc.blocks.new(name=name)
        for ent in picked:
            c = ent.copy()
            try:
                c.transform(Matrix44.translate(-bx, -by, 0))
            except Exception:
                pass
            blk.add_entity(c)
            msp.delete_entity(ent)
        msp.add_blockref(name, insert=(bx, by))
        out = Path(tmp) / "blocked.dxf"
        doc.saveas(str(out))
        new_bytes = out.read_bytes()
    _store_dwg_dxf(file_id, file_path, new_bytes, request.state.user_id, "BLOCK_REG",
                   {"name": name, "entities": len(picked)})
    return {"fileId": file_id, "name": name, "entities": len(picked),
            "document": _parse_cad_bytes(new_bytes, file_name)}


@router.post("/cad/view/{file_id}/blocks/insert", dependencies=[SETUP])
def cad_block_insert(file_id: int, request: Request, body: BlockInsertReq) -> dict[str, Any]:
    """상위 호출 — 등록 Block 을 지정 좌표에 INSERT."""
    import tempfile
    from pathlib import Path

    name = body.name.strip()
    data, file_path, file_name = _load_dwg_dxf(file_id)
    with tempfile.TemporaryDirectory(prefix="edim_blk_") as tmp:
        doc = _open_dxf(tmp, file_name, data)
        if name not in doc.blocks or name.startswith(("*", "_")):
            raise HTTPException(404, detail=f"Block 없음: {name}")
        doc.modelspace().add_blockref(name, insert=(body.x, body.y))
        out = Path(tmp) / "inserted.dxf"
        doc.saveas(str(out))
        new_bytes = out.read_bytes()
    _store_dwg_dxf(file_id, file_path, new_bytes, request.state.user_id, "BLOCK_INSERT",
                   {"name": name, "x": body.x, "y": body.y})
    return {"fileId": file_id, "name": name,
            "document": _parse_cad_bytes(new_bytes, file_name)}


@router.post("/cad/import", status_code=201, dependencies=[SETUP])
async def cad_import(
    uploadedFile: UploadFile = File(...),
    project: str = Form("PS-61313-5"),
) -> dict[str, Any]:
    data = await uploadedFile.read()
    if len(data) > 100 * 1024 * 1024:
        raise HTTPException(413, detail="100MB 초과")
    fname = (uploadedFile.filename or "drawing.dxf").replace("/", "_")
    document = _parse_cad_bytes(data, fname)   # 파싱 실패 시 저장하지 않음
    key = f"{project}/DWG/{fname}"
    try:
        storage.put_object(key, data, uploadedFile.content_type or "application/dxf")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT project_id FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, project))
        prj = cur.fetchone()
        cur.execute(
            """INSERT INTO dwg_file (tenant_id, project_id, folder, file_name, file_type,
               file_path, file_size)
               VALUES (%s,%s,'DWG',%s,%s,%s,%s) RETURNING file_id""",
            (tid, prj[0] if prj else None, fname,
             fname.rsplit(".", 1)[-1].upper()[:10], key, len(data)))
        file_id = cur.fetchone()[0]
    return {"fileId": file_id, "document": document}


@router.get("/cad/arrangement")
def cad_arrangement() -> dict[str, Any]:
    """C-1 구성 캔버스의 CAD 정본 — 실 DXF 작도 후 정규화 문서로 반환."""
    from app.services.run_pipeline import build_arrangement_dxf
    data = build_arrangement_dxf()
    return {"document": _parse_cad_bytes(data, "AHU5_arrangement.dxf")}


@router.get("/cad/arrangement.dxf")
def cad_arrangement_dxf() -> StreamingResponse:
    import io as _io

    from app.services.run_pipeline import build_arrangement_dxf
    return StreamingResponse(
        _io.BytesIO(build_arrangement_dxf()), media_type="application/dxf",
        headers={"Content-Disposition": "attachment; filename=AHU5_arrangement.dxf"})


@router.get("/cad/duct-layout")
def cad_duct_layout(diffusers: int = 3, floor: str = "3F") -> dict[str, Any]:
    """건축설비 Duct 자동 배치 CAD 정본 (M-4-3) — 실 DXF 작도 후 정규화 문서 반환."""
    from app.services.run_pipeline import build_duct_layout_dxf
    data = build_duct_layout_dxf(max(1, min(diffusers, 12)), floor.strip()[:10] or "3F")
    return {"document": _parse_cad_bytes(data, "duct_layout.dxf")}


class DuctLayoutSaveRequest(BaseModel):
    diffusers: int = 3
    floor: str = "3F"
    project: str = "PS-61313-5"


@router.post("/cad/duct-layout/save", dependencies=[SETUP])
def cad_duct_layout_save(request: Request, body: DuctLayoutSaveRequest) -> dict[str, Any]:
    """U8 — Duct 배치 편집 대상화: 자동 배치 DXF 를 dwg_file(MinIO) 실체화(같은 이름=덮어쓰기).
    반환 fileId 로 /cad/view/{id}/edit 수동 조정(이동/삭제/분할=trim/합체) 가능."""
    from app.services.run_pipeline import build_duct_layout_dxf

    floor = body.floor.strip()[:10] or "3F"
    data = build_duct_layout_dxf(max(1, min(body.diffusers, 12)), floor)
    fname = f"duct_{floor}.dxf"
    key = f"{body.project}/DWG/{fname}"
    try:
        storage.put_object(key, data, "application/dxf")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT project_id FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, body.project))
        prj = cur.fetchone()
        cur.execute(
            "SELECT file_id FROM dwg_file WHERE tenant_id=%s AND folder='DWG' AND file_name=%s",
            (tid, fname))
        row = cur.fetchone()
        if row:
            file_id = row[0]
            cur.execute("UPDATE dwg_file SET file_path=%s, file_size=%s WHERE file_id=%s",
                        (key, len(data), file_id))
        else:
            cur.execute(
                """INSERT INTO dwg_file (tenant_id, project_id, folder, file_name, file_type,
                   file_path, file_size) VALUES (%s,%s,'DWG',%s,'DXF',%s,%s) RETURNING file_id""",
                (tid, prj[0] if prj else None, fname, key, len(data)))
            file_id = cur.fetchone()[0]
        _audit(cur, tid, "dwg_file", file_id, "DUCT_EDIT_MATERIALIZE", request.state.user_id,
               {"floor": floor, "diffusers": body.diffusers})
    return {"fileId": file_id, "document": _parse_cad_bytes(data, fname)}


class CadBlock(BaseModel):
    id: str = ""
    name: str = ""
    sub: str = ""
    x: float
    y: float
    w: float
    h: float
    dashed: bool = False


class CadDim(BaseModel):
    x: float
    y: float
    w: float
    label: str = ""


class CadLabel(BaseModel):
    x: float
    y: float
    text: str = ""


class BlockDocRequest(BaseModel):
    name: str = "Block Diagram"
    blocks: list[CadBlock] = []
    dims: list[CadDim] = []
    labels: list[CadLabel] = []


def _block_dxf_bytes(body: BlockDocRequest) -> bytes:
    from app.services.run_pipeline import build_blocks_dxf
    if not body.blocks:
        raise HTTPException(422, detail="blocks 가 비어 있습니다")
    return build_blocks_dxf(
        [b.model_dump() for b in body.blocks[:500]],
        [d.model_dump() for d in body.dims[:200]],
        [lb.model_dump() for lb in body.labels[:200]],
        body.name.strip()[:80] or "Block Diagram")


@router.post("/cad/from-blocks")
def cad_from_blocks(body: BlockDocRequest) -> dict[str, Any]:
    """블록 캔버스(Cvs 모델) → 정규화 DrawingDocument (엔진 통합 — CadSvg 실엔진 렌더 공용)."""
    return {"document": _parse_cad_bytes(_block_dxf_bytes(body), "block_diagram.dxf")}


@router.post("/cad/from-blocks.dxf")
def cad_from_blocks_dxf(body: BlockDocRequest) -> StreamingResponse:
    """블록 캔버스 → DXF 다운로드 (블록 다이어그램 DXF 익스포트)."""
    import io as _io
    return StreamingResponse(
        _io.BytesIO(_block_dxf_bytes(body)), media_type="application/dxf",
        headers={"Content-Disposition": "attachment; filename=block_diagram.dxf"})


class PartDrawingRequest(BaseModel):
    dims: dict[str, float] = {}


@router.post("/cad/part-drawing")
def cad_part_drawing(body: PartDrawingRequest) -> dict[str, Any]:
    """Design Editor CAD 모드 — 현재 치수로 부품도 작도 후 정규화 문서 반환.
    dims 미지정 시 dwg_dimension 엔진 평가값 사용."""
    from app.services import run_pipeline as rp
    dims = {k: float(v) for k, v in body.dims.items()}
    if not dims:
        r = rp.PipelineResult()
        with _conn() as conn, conn.cursor() as cur:
            tid = _tenant_id(cur)
            rp.step_dims(cur, tid, _make_table_resolver(cur, tid), r)
        dims = r.dims
    data = rp.build_part_dxf(dims)
    return {"dims": dims, "document": _parse_cad_bytes(data, "KDCR3-13_part.dxf")}


class PartDrawingSaveRequest(BaseModel):
    dims: dict[str, float] = {}
    name: str = "part_edit.dxf"
    project: str = "PS-61313-5"


@router.post("/cad/part-drawing/save", dependencies=[SETUP])
def cad_part_drawing_save(request: Request, body: PartDrawingSaveRequest) -> dict[str, Any]:
    """Design Editor 편집 대상화 — 현재 치수 부품도를 dwg_file(MinIO)로 실체화(같은 이름=덮어쓰기).
    반환 fileId 로 /cad/view/{id}/edit 편집·영속 가능."""
    from app.services import run_pipeline as rp

    dims = {k: float(v) for k, v in body.dims.items()}
    if not dims:
        r = rp.PipelineResult()
        with _conn() as conn, conn.cursor() as cur:
            tid = _tenant_id(cur)
            rp.step_dims(cur, tid, _make_table_resolver(cur, tid), r)
        dims = r.dims
    data = rp.build_part_dxf(dims)
    fname = (body.name or "part_edit.dxf").replace("/", "_")
    if not fname.lower().endswith(".dxf"):
        fname += ".dxf"
    key = f"{body.project}/DWG/{fname}"
    try:
        storage.put_object(key, data, "application/dxf")
    except RuntimeError:
        raise HTTPException(503, detail="storage unavailable")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT project_id FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, body.project))
        prj = cur.fetchone()
        cur.execute(
            "SELECT file_id FROM dwg_file WHERE tenant_id=%s AND folder='DWG' AND file_name=%s",
            (tid, fname))
        row = cur.fetchone()
        if row:
            file_id = row[0]
            cur.execute("UPDATE dwg_file SET file_path=%s, file_size=%s WHERE file_id=%s",
                        (key, len(data), file_id))
        else:
            cur.execute(
                """INSERT INTO dwg_file (tenant_id, project_id, folder, file_name, file_type,
                   file_path, file_size) VALUES (%s,%s,'DWG',%s,'DXF',%s,%s) RETURNING file_id""",
                (tid, prj[0] if prj else None, fname, key, len(data)))
            file_id = cur.fetchone()[0]
    return {"fileId": file_id, "document": _parse_cad_bytes(data, fname)}


class CadExportRequest(BaseModel):
    dims: dict[str, float] = {}


@router.post("/cad/export-dxf", dependencies=[SETUP])
def cad_export(body: CadExportRequest) -> StreamingResponse:
    """현재 치수(미지정 시 dwg_dimension 평가값)로 제작 DXF 생성 — 순수 다운로드."""
    import io as _io

    from app.services import run_pipeline as rp

    r = rp.PipelineResult()
    if body.dims:
        r.dims = {k: float(v) for k, v in body.dims.items()}
    else:
        with _conn() as conn, conn.cursor() as cur:
            tid = _tenant_id(cur)
            rp.step_dims(cur, tid, _make_table_resolver(cur, tid), r)
    rp.step_drawing(r)
    _, fname, _, data = r.files[0]
    return StreamingResponse(
        _io.BytesIO(data), media_type="application/dxf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── SVC-08 Cost ──
@router.get("/prices/resolve")
def resolve_price(code: str, at: str | None = None) -> dict[str, Any]:
    ref = at or date.today().isoformat()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT pc.main_code, pc.code_name, p.price_source, p.price,
                      p.valid_from, p.valid_to, cc.company_name
               FROM cst_price p
               JOIN product_code pc ON pc.product_code_id=p.product_code_id
               LEFT JOIN com_company cc ON cc.company_id=p.supplier_id
               WHERE p.tenant_id=%s AND upper(pc.main_code)=upper(%s)
                 AND p.valid_from <= %s::date
                 AND (p.valid_to IS NULL OR p.valid_to >= %s::date)
               ORDER BY array_position(%s::text[], p.price_source), p.valid_from DESC
               LIMIT 1""", (tid, code, ref, ref, SOURCE_PRIORITY))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"단가 없음: {code} @ {ref}")
    return {
        "code": row[0], "name": row[1], "source": SOURCE_LABEL[row[2]],
        "price": float(row[3]), "from": row[4].isoformat(),
        "to": row[5].isoformat() if row[5] else None,
        "supplier": row[6] or "-", "active": True,
    }


# ── SVC-10 Approval (승인함 M-15-2) ──
def _apply_eco(cur, tid: int, eco_id: int, approve: bool,
               actor_login: str, actor_id: int) -> None:
    """설계변경(ECO) 승인 결과 적용 — 승인 시 상태 APPLIED + Rev-up(DRAWING) + ECN 통지."""
    cur.execute(
        "SELECT eco_no, title, target_type, target_no, status, impact_data FROM eco_change "
        "WHERE tenant_id=%s AND eco_id=%s", (tid, eco_id))
    e = cur.fetchone()
    if not e:
        return
    eco_no, title, ttype, tno, _status, impact = e
    if not approve:
        cur.execute("UPDATE eco_change SET status='REJECTED' WHERE eco_id=%s", (eco_id,))
        return
    rev_from = rev_to = None
    superseded = None                        # D5 — 대체 도면 (신·구)
    new_dwg = (impact or {}).get("supersededBy") if isinstance(impact, dict) else None
    if ttype == "DRAWING":
        cur.execute(
            "SELECT drawing_id, current_rev FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s",
            (tid, tno))
        d = cur.fetchone()
        if d and new_dwg:
            # 도면 대체(Supersedure) — 구도면 → 신도면 (Rev-up 대신 대체 이력)
            old_id = d[0]
            cur.execute("SELECT drawing_id FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s",
                        (tid, str(new_dwg)))
            nd = cur.fetchone()
            cur.execute("SELECT 1 FROM dwg_supersedure WHERE old_drawing_id=%s", (old_id,))
            if nd and nd[0] != old_id and not cur.fetchone():
                cur.execute(
                    """INSERT INTO dwg_supersedure (tenant_id, old_drawing_id, new_drawing_id,
                       reason, superseded_date, created_by)
                       VALUES (%s,%s,%s,%s,CURRENT_DATE,%s)""",
                    (tid, old_id, nd[0], f"ECO {eco_no} — {title[:400]}", actor_login))
                superseded = str(new_dwg)
                cur.execute(
                    """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
                       VALUES (%s,'dwg_drawing',%s,'SUPERSEDE',%s,%s)""",
                    (tid, old_id, actor_id, json.dumps({"old": tno, "new": new_dwg, "eco": eco_no})))
        elif d:
            did, cur_rev = d
            rev_from = cur_rev
            rev_to = _rev_next(cur_rev)
            cur.execute(
                """INSERT INTO dwg_revision (drawing_id, rev_no, rev_date, rev_reason, revised_by)
                   VALUES (%s,%s,CURRENT_DATE,%s,%s)""",
                (did, rev_to, f"ECO {eco_no} — {title[:400]}", actor_login))
            cur.execute(
                "UPDATE dwg_drawing SET current_rev=%s, updated_by=%s, updated_at=now() "
                "WHERE drawing_id=%s", (rev_to, actor_login, did))
            cur.execute(
                """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
                   VALUES (%s,'dwg_drawing',%s,'REV_UP',%s,%s)""",
                (tid, did, actor_id, json.dumps({"from": cur_rev, "to": rev_to, "eco": eco_no})))
    cur.execute(
        "UPDATE eco_change SET status='APPLIED', rev_from=%s, rev_to=%s, applied_at=now() "
        "WHERE eco_id=%s", (rev_from, rev_to, eco_id))
    # 변경 통지(ECN) — 영향 부서(SETUP/ADMIN) 알림
    ecn = f"설계변경 통지(ECN) — {eco_no} {title[:50]}"
    if rev_to:
        ecn += f" (Rev {rev_from}→{rev_to} 적용)"
    if superseded:
        ecn += f" (도면 대체: {tno}→{superseded})"
    cur.execute(
        "SELECT user_id FROM sys_user WHERE tenant_id=%s AND user_level IN ('SETUP','ADMIN')", (tid,))
    for (uid,) in cur.fetchall():
        _notify(cur, tid, uid, "TASK_ASSIGNED", ecn, "/plm")


@router.get("/approvals/inbox")
def approvals_inbox() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT a.approval_id, a.target_table, a.request_type, a.step,
                      u.user_name, to_char(a.requested_at,'MM-DD'), a.comment,
                      COALESCE(pc.main_code, dc.doc_no,
                               ec.eco_no||' '||ec.title, a.target_table||'#'||a.target_id),
                      u.login_id
               FROM sys_approval_request a
               JOIN sys_user u ON u.user_id=a.requester_id
               LEFT JOIN product_code pc
                 ON a.target_table='product_code' AND pc.product_code_id=a.target_id
               LEFT JOIN doc_control dc
                 ON a.target_table='doc_control' AND dc.doc_control_id=a.target_id
               LEFT JOIN eco_change ec
                 ON a.target_table='eco_change' AND ec.eco_id=a.target_id
               WHERE a.tenant_id=%s AND a.result IS NULL
               ORDER BY a.approval_id""", (tid,))
        type_label = {"product_code": "Code", "doc_control": "문서",
                      "code_item": "Code", "tbx_macro": "Macro",
                      "code_relationship": "관계", "eco_change": "설계변경"}
        return [
            {"id": r[0], "assetType": type_label.get(r[1], r[1]),
             "target": r[6] or r[7], "reqKind": r[2], "requester": r[4],
             "reqDate": r[5], "stage": r[3], "tested": r[1] == "tbx_macro",
             "requesterLogin": r[8]}   # F3 — '내 요청' 필터용
            for r in cur.fetchall()
        ]


class DecideRequest(BaseModel):
    approve: bool
    comment: str = ""


def _apply_decision(cur, tid: int, approval_id: int, approve: bool, comment: str,
                    actor_login: str, actor_id: int) -> str | None:
    """단건 승인 결정 적용 — 상태 전이 + 요청자 알림 + 자산 전이 + 이력.
    처리 가능한 요청이 아니면 None (이미 결정됨/미존재). decide·decide_batch 공용."""
    result = "APPROVED" if approve else "REJECTED"
    cur.execute(
        """UPDATE sys_approval_request
           SET result=%s, comment=NULLIF(%s,'') , decided_at=now()
           WHERE tenant_id=%s AND approval_id=%s AND result IS NULL
           RETURNING target_table, target_id, requester_id, comment""",
        (result, comment.strip(), tid, approval_id))
    row = cur.fetchone()
    if not row:
        return None
    # 요청자 알림 (SVC-13)
    _notify(cur, tid, row[2], "APPROVAL_RESULT",
            f"{'승인' if approve else '반려'} — {row[0]} #{row[1]}"
            + (f" ({comment.strip()})" if comment.strip() else ""),
            "/common")
    # 대상 자산 상태 전이 + 이력
    if row[0] == "product_code":
        cur.execute(
            "UPDATE product_code SET approval_status=%s WHERE product_code_id=%s",
            (result, row[1]))
    elif row[0] == "code_item":
        # #28 — Sub Code 승인 결정이 값에 전혀 반영되지 않던 결함 수정.
        # S-1-1 은 항목(code_item) 단위로 요청하므로 그 하위 값 전체를 전이한다.
        # (Revision 은 값 내용이 바뀔 때 PATCH /codes/values 에서 올린다 — 승인은 상태만 전이)
        cur.execute(
            "UPDATE code_item_value SET approval_status=%s, updated_at=now() "
            "WHERE tenant_id=%s AND item_id=%s", (result, tid, row[1]))
    elif row[0] == "code_item_value":
        cur.execute(
            "UPDATE code_item_value SET approval_status=%s, updated_at=now() "
            "WHERE tenant_id=%s AND value_id=%s", (result, tid, row[1]))
    elif row[0] == "code_relationship":
        # F4 — Mother 코드의 관계 세트 전이 (Running Test 통과 승인, CODE-009)
        # #40 — 승인 라운드마다 관계 Revision 증가: 과거 Run 은 자기 근거 Revision 을 그대로 보존하고,
        # 근거가 움직였다는 사실은 bom-basis 대조에서 드러난다(BOM 자체는 Snapshot 이 불변 보존).
        cur.execute(
            f"""UPDATE code_relationship SET approval_status=%s
                {', revision_no = revision_no + 1' if approve else ''}
               WHERE tenant_id=%s AND mother_code_id=%s""",
            (result, tid, row[1]))
    elif row[0] == "eco_change":
        # D5 — 설계변경 승인: 승인 시 Rev-up 자동 적용(DRAWING) + 변경 통지(ECN)
        _apply_eco(cur, tid, row[1], approve, actor_login, actor_id)
    elif row[0] == "tbx_macro":
        # 권한승인정의서 상태기계 #7 — 승인=APPROVED, 반려=DRAFT 복귀 (재작성 후 재요청)
        cur.execute("UPDATE tbx_macro SET status=%s, updated_at=now() "
                    "WHERE tenant_id=%s AND macro_id=%s",
                    ("APPROVED" if approve else "DRAFT", tid, row[1]))
    elif row[0] == "erp_handoff":
        # 트리아지 #46 — Handoff 상태기계: 승인=approved (수신 대기), 반려=rejected
        cur.execute("UPDATE erp_handoff SET status=%s, decided_at=now() "
                    "WHERE tenant_id=%s AND handoff_id=%s AND status='approval_requested'",
                    ("approved" if approve else "rejected", tid, row[1]))
    cur.execute(
        """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
           VALUES (%s,%s,%s,%s,%s,%s)""",
        (tid, row[0], row[1], result, actor_id, json.dumps({"comment": comment})))
    return result


@router.post("/approvals/{approval_id}/decide", dependencies=[SETUP])
def decide(approval_id: int, request: Request, body: DecideRequest) -> dict[str, Any]:
    if not body.approve and not body.comment.strip():
        raise HTTPException(422, detail="반려는 코멘트 필수")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        result = _apply_decision(cur, tid, approval_id, body.approve, body.comment,
                                 request.state.login, request.state.user_id)
        if result is None:
            raise HTTPException(404, detail=f"처리 가능한 요청 없음: {approval_id}")
    return {"approvalId": approval_id, "result": result}


class DecideBatchRequest(BaseModel):
    approvalIds: list[int]
    approve: bool
    comment: str = ""


@router.post("/approvals/decide-batch", dependencies=[SETUP])
def decide_batch(request: Request, body: DecideBatchRequest) -> dict[str, Any]:
    """승인함 일괄 승인/반려 (D8) — 다중 선택 처리. 이미 결정된 건은 건너뜀."""
    if not body.approvalIds:
        raise HTTPException(422, detail="선택된 요청이 없습니다")
    if not body.approve and not body.comment.strip():
        raise HTTPException(422, detail="일괄 반려는 코멘트 필수")
    if len(body.approvalIds) > 200:
        raise HTTPException(422, detail="한 번에 최대 200건")
    done, skipped = [], []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for aid in body.approvalIds:
            r = _apply_decision(cur, tid, aid, body.approve, body.comment,
                                request.state.login, request.state.user_id)
            (done if r is not None else skipped).append(aid)
    return {"result": "APPROVED" if body.approve else "REJECTED",
            "processed": len(done), "skipped": len(skipped),
            "processedIds": done, "skippedIds": skipped}


# ── SVC-13 알림 ──
TYPE_PRIORITY = {"ESCALATION": "HIGH", "DEADLINE_WARN": "HIGH",
                 "APPROVAL_REQUEST": "MED", "APPROVAL_RESULT": "MED", "TASK_ASSIGNED": "MED",
                 "ANNOUNCE": "MED"}


@router.get("/notifications")
def notifications(request: Request, limit: int = 20, type: str = "",
                  unreadOnly: bool = False) -> list[dict[str, Any]]:
    """알림 목록 (C4/E6 — 우선순위 파생·유형 필터·활성/보관함). ORDER: 미읽음→우선순위→최신.
    unreadOnly=true = 활성(미읽음만), false = 보관함(읽음 포함, 소실 방지)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid, request.state.user_id]
        type_clause = ""
        if type.strip():
            type_clause = " AND notify_type=%s"
            params.append(type.strip())
        if unreadOnly:
            type_clause += " AND is_read=false"
        params.append(max(1, min(limit, 100)))
        cur.execute(
            f"""SELECT notification_id, notify_type, title, link_url, is_read,
                      to_char(created_at,'MM-DD HH24:MI')
               FROM sys_notification
               WHERE tenant_id=%s AND user_id=%s{type_clause}
               ORDER BY is_read,
                        CASE notify_type WHEN 'ESCALATION' THEN 0 WHEN 'DEADLINE_WARN' THEN 0
                             WHEN 'APPROVAL_REQUEST' THEN 1 WHEN 'APPROVAL_RESULT' THEN 1
                             WHEN 'TASK_ASSIGNED' THEN 1 ELSE 2 END,
                        notification_id DESC LIMIT %s""", tuple(params))
        return [
            {"id": r[0], "type": r[1], "title": r[2], "link": r[3],
             "read": r[4], "at": r[5], "priority": TYPE_PRIORITY.get(r[1], "LOW")}
            for r in cur.fetchall()
        ]


class AnnounceCreate(BaseModel):
    title: str
    link: str = ""


@router.post("/notifications/announce", status_code=201, dependencies=[ADMIN])
def notifications_announce(request: Request, body: AnnounceCreate) -> dict[str, Any]:
    """공지 발송 (메뉴정의서 공통/알림 P2) — 전 활성 사용자 인앱 알림 (type=ANNOUNCE)."""
    title = body.title.strip()
    if not title:
        raise HTTPException(422, detail="공지 제목을 입력하십시오")
    link = body.link.strip()
    if link and not link.startswith("/"):
        raise HTTPException(422, detail="링크는 내부 경로(/...)만 허용")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id FROM sys_user WHERE tenant_id=%s AND status='ACTIVE'", (tid,))
        uids = [r[0] for r in cur.fetchall()]
        for uid in uids:
            _notify(cur, tid, uid, "ANNOUNCE", title, link or None)
        _audit(cur, tid, "sys_notification", 0, "ANNOUNCE", request.state.user_id,
               {"title": title[:100], "recipients": len(uids)})
    return {"sent": len(uids), "title": title}


@router.get("/notifications/digest")
def notifications_digest(request: Request) -> dict[str, Any]:
    """알림 다이제스트 (C4) — 미읽음 유형별 요약 + 지연 이벤트 수 (로그인 시 요약)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT notify_type, count(*) FROM sys_notification
               WHERE tenant_id=%s AND user_id=%s AND is_read=false GROUP BY notify_type""",
            (tid, request.state.user_id))
        by_type = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute(
            """SELECT count(*) FROM erp_process_event
               WHERE tenant_id=%s AND status<>'DONE' AND due_date < CURRENT_DATE""", (tid,))
        overdue = cur.fetchone()[0]
    high = sum(v for k, v in by_type.items() if TYPE_PRIORITY.get(k) == "HIGH")
    return {"unread": sum(by_type.values()), "byType": by_type, "overdue": overdue, "high": high}


@router.post("/notifications/{notification_id}/read")
def read_notification(notification_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_notification SET is_read=true
               WHERE tenant_id=%s AND user_id=%s AND notification_id=%s""",
            (tid, request.state.user_id, notification_id))
    return {"id": notification_id, "read": True}


@router.post("/notifications/read-all")
def read_all_notifications(request: Request) -> dict[str, Any]:
    """모두 읽음 (B6)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_notification SET is_read=true
               WHERE tenant_id=%s AND user_id=%s AND is_read=false""",
            (tid, request.state.user_id))
        n = cur.rowcount
    return {"read": n}


# ── SVC-11 Documents (문서함 M-5-4) ──
STATUS_LABEL = {"SET_UP": "Set-up", "CHECK": "Check", "APPROVE": "Approve 대기", "ACCEPTED": "Accepted"}


DOC_SORT = {"docNo": "d.doc_no", "title": "d.title", "date": "d.created_at",
            "status": "d.released_status", "grade": "d.management_grade"}


@router.get("/documents")
def documents(sort: str = "", dir: str = "asc") -> list[dict[str, Any]]:
    # F8 — 서버 정렬 (화이트리스트; 기본 = 등록순)
    order = "d.doc_control_id"
    if sort in DOC_SORT:
        order = f"{DOC_SORT[sort]} {'DESC' if dir == 'desc' else 'ASC'}"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT d.doc_no, d.title, d.person, to_char(d.created_at,'MM-DD'),
                      d.released_status, u.user_name,
                      to_char(d.approval_date,'MM-DD'), d.version, d.management_grade,
                      d.doc_type
               FROM doc_control d LEFT JOIN sys_user u ON u.user_id=d.approver_id
               WHERE d.tenant_id=%s ORDER BY {order}""", (tid,))
        return [
            {"docNo": r[0], "title": r[1], "person": r[2], "date": r[3],
             "status": STATUS_LABEL[r[4]], "approver": r[5] or "-",
             "appDate": r[6] or "-", "version": r[7], "grade": r[8], "docType": r[9]}
            for r in cur.fetchall()
        ]


@router.delete("/documents/{doc_no}", dependencies=[SETUP])
def document_delete(doc_no: str, request: Request) -> dict[str, Any]:
    """문서 삭제 — SET_UP(작성 중) 상태 한정 (승인/발행 문서 보호). 최신 행 기준."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT doc_control_id, released_status FROM doc_control
               WHERE tenant_id=%s AND doc_no=%s ORDER BY doc_control_id DESC LIMIT 1""",
            (tid, doc_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"문서 없음: {doc_no}")
        if row[1] != "SET_UP":
            raise HTTPException(409, detail=f"SET_UP 문서만 삭제 가능 (현재 {row[1]})")
        cur.execute("DELETE FROM doc_control WHERE doc_control_id=%s", (row[0],))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'doc_control',%s,'DELETE',%s,%s)""",
            (tid, row[0], request.state.user_id, json.dumps({"docNo": doc_no})))
    return {"deleted": doc_no}


# ── B4 — 문서 등록 + Grade 워터마크 PDF 렌더 (SVC-11 · DOC-002) ──

class DocCreate(BaseModel):
    docNo: str
    title: str
    docType: str = "TECH_DOC"
    grade: str = "S-3"


@router.post("/documents", status_code=201, dependencies=[SETUP])
def create_document(request: Request, body: DocCreate) -> dict[str, Any]:
    if not body.docNo.strip() or not body.title.strip():
        raise HTTPException(422, detail="필수(노란 셀) — DOC No·제목")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM doc_control WHERE tenant_id=%s AND doc_no=%s", (tid, body.docNo.strip()))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — DOC No {body.docNo} 이미 등록됨")
        cur.execute("SELECT user_name FROM sys_user WHERE user_id=%s", (request.state.user_id,))
        person = (cur.fetchone() or ["-"])[0]
        cur.execute(
            """INSERT INTO doc_control (tenant_id, doc_no, title, doc_type, released_status,
               version, person, management_grade)
               VALUES (%s,%s,%s,%s,'SET_UP','KD-0.1',%s,%s) RETURNING doc_control_id""",
            (tid, body.docNo.strip(), body.title.strip(), body.docType.strip()[:20],
             person, body.grade.strip()[:10]))
        doc_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'doc_control',%s,'CREATE','승인',%s,%s)""",
            (tid, doc_id, request.state.user_id, f"문서 등록 — {body.docNo} {body.title}"[:200]))
    return {"docId": doc_id, "status": "Set-up"}


class RegisterOutput(BaseModel):
    fileName: str
    folder: str = "DWG"
    fileType: str = "PDF"


@router.post("/documents/register-output", dependencies=[SETUP])
def register_output(request: Request, body: RegisterOutput) -> dict[str, Any]:
    """Run 산출물을 doc_control 정본으로 find-or-create (승인 상태 영속 — G3-a).

    파일명 기준 멱등: 이미 등록됐으면 기존 doc_no·released_status 반환(재진입 시 상태 유지),
    없으면 RUN 유형으로 채번·등록. 이후 PATCH /documents/{no}/status 로 실 상태 전이.
    """
    title = body.fileName.strip()[:200]
    if not title:
        raise HTTPException(422, detail="필수 — 파일명")
    dt = "RUN"
    grade = "S-2" if title.lower().endswith(".pdf") or "견적" in title else "S-3"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT doc_no, released_status FROM doc_control
               WHERE tenant_id=%s AND doc_type=%s AND title=%s
               ORDER BY doc_control_id DESC LIMIT 1""", (tid, dt, title))
        row = cur.fetchone()
        if row:
            return {"docNo": row[0], "status": row[1], "created": False}
        cur.execute("SELECT count(*) FROM doc_control WHERE tenant_id=%s AND doc_type=%s", (tid, dt))
        seq = cur.fetchone()[0] + 1
        while True:
            doc_no = f"{dt}-{seq:04d}"
            cur.execute("SELECT 1 FROM doc_control WHERE tenant_id=%s AND doc_no=%s", (tid, doc_no))
            if not cur.fetchone():
                break
            seq += 1
        cur.execute("SELECT user_name FROM sys_user WHERE user_id=%s", (request.state.user_id,))
        person = (cur.fetchone() or ["-"])[0]
        cur.execute(
            """INSERT INTO doc_control (tenant_id, doc_no, title, doc_type, ref_type,
               released_status, version, person, management_grade)
               VALUES (%s,%s,%s,%s,'RUN_OUTPUT','SET_UP','KD-0.1',%s,%s) RETURNING doc_control_id""",
            (tid, doc_no, title, dt, person, grade))
        doc_id = cur.fetchone()[0]
        _audit(cur, tid, "doc_control", doc_id, "CREATE", request.state.user_id,
               {"docNo": doc_no, "title": title, "source": "run-output"})
    return {"docNo": doc_no, "status": "SET_UP", "created": True}


# D9 — 문서 Grade 열람 통제: 미달 레벨은 열람 차단 (워터마크 위 단계 강제)
DOC_GRADE_POLICY = {"S-1": ("ADMIN", "PLATFORM"), "S-2": ("SETUP", "ADMIN", "PLATFORM")}


@router.get("/documents/{doc_no}/render.pdf")
def render_document(doc_no: str, request: Request) -> Any:
    """문서 PDF 실렌더 — Grade 열람 통제(S-1=ADMIN+, S-2=SETUP+) + S-1/S-2 CONFIDENTIAL 워터마크 (DOC-002/DOC-004)."""
    from ..services import run_pipeline as rp
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT title, doc_type, released_status, version, person, management_grade,
                      to_char(created_at,'YYYY-MM-DD'), doc_control_id
               FROM doc_control WHERE tenant_id=%s AND doc_no=%s
               ORDER BY doc_control_id DESC LIMIT 1""", (tid, doc_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"문서 없음: {doc_no}")
        title, dtype, status, ver, person, grade, cdate, doc_id = row
        # 열람 enforcement — 등급 미달 레벨은 403 + 거부 감사
        allowed = DOC_GRADE_POLICY.get(grade or "")
        if allowed and request.state.level not in allowed:
            _audit(cur, tid, "doc_control", doc_id, "READ_DENY", request.state.user_id,
                   {"docNo": doc_no, "grade": grade, "level": request.state.level})
            raise HTTPException(
                403, detail=f"열람 권한 부족 — {grade} 등급 문서는 {'/'.join(allowed)} 만 열람 가능 "
                            f"(현재 등급: {request.state.level})")
        # 권한승인정의서 문서보안등급 — S-1 은 열람 성공도 감사 (워터마크+열람 로그)
        if grade == "S-1":
            _audit(cur, tid, "doc_control", doc_id, "DOC_READ", request.state.user_id,
                   {"docNo": doc_no, "grade": grade})
    watermark = grade in ("S-1", "S-2")
    pdf = rp.build_doc_pdf(
        doc_no=doc_no, title=title, doc_type=dtype, status=STATUS_LABEL.get(status, status),
        version=ver, person=person or "-", grade=grade or "-", created=cdate,
        confidential=watermark)
    from fastapi.responses import Response
    return Response(content=pdf, media_type="application/pdf", headers={
        "Content-Disposition": f"inline; filename=\"{doc_no}.pdf\"",
    })


class RenderRequest(BaseModel):
    title: str
    subtitle: str = ""
    lines: list[str] = []
    confidential: bool = False
    # U6 출력 옵션 (슬라이드 50) — 미전달 시 기존 동작
    paper: str = "A4"          # A4 | A3 | LETTER
    landscapeMode: bool = False
    marginMm: float = 17.6
    fontPt: float = 9.5
    grayscale: bool = False
    footerText: str = ""


@router.post("/render/pdf", dependencies=[SETUP])
def render_generic_pdf(body: RenderRequest) -> Any:
    """범용 PDF 렌더 (SVC-11·U6) — Print Set-up Test 출력·Doc Templet Print. 용지·여백·글꼴·색상 옵션."""
    from fastapi.responses import Response

    from ..services import run_pipeline as rp
    pdf = rp.build_lines_pdf(
        title=body.title.strip() or "EDIM 출력", subtitle=body.subtitle.strip(),
        lines=body.lines, confidential=body.confidential,
        paper=body.paper, land=body.landscapeMode, margin_mm=body.marginMm,
        font_pt=body.fontPt, grayscale=body.grayscale, footer_text=body.footerText.strip())
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=\"edim-print.pdf\""})


@router.post("/render/xlsx", dependencies=[SETUP])
def render_generic_xlsx(body: RenderRequest) -> Any:
    """범용 Office(xlsx) 내보내기 (U6, 슬라이드 50 'File 내보내기') — 렌더 라인을 시트로."""
    rows: list[list[Any]] = [[ln] for ln in body.lines]
    if body.subtitle.strip():
        rows.insert(0, [body.subtitle.strip()])
    return _xlsx_response((body.title.strip() or "EDIM 출력")[:31],
                          [body.title.strip() or "EDIM 출력"], rows, "edim-print")


# ── SVC-01 Users (M-14-6) ──
ROLE_LABEL ={"PLATFORM": "Platform", "ADMIN": "관리자", "SETUP": "설계 Set-up", "GENERAL": "일반"}


@router.get("/users", dependencies=[SETUP])
def users() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT login_id, user_name, department, user_level, status, email
               FROM sys_user WHERE tenant_id=%s ORDER BY user_id""", (tid,))
        return [
            {"login": r[0], "name": r[1], "dept": r[2] or "-", "level": r[3],
             "role": (r[2] or "") + " " + ROLE_LABEL.get(r[3], r[3]), "status": r[4],
             "email": r[5] or ""}
            for r in cur.fetchall()
        ]


@router.post("/users/{login}/unlock", dependencies=[ADMIN])
def unlock_user(login: str, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_user SET status='ACTIVE', updated_by=%s, updated_at=now()
               WHERE tenant_id=%s AND login_id=%s AND status='LOCKED'
               RETURNING user_id""", (request.state.login, tid, login))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"LOCKED 계정 아님: {login}")
        _audit(cur, tid, "sys_user", row[0], "UNLOCK", request.state.user_id,
               {"login": login})   # 잠금 해제 = 실패 카운터 초기화 기준점 (B8)
    return {"login": login, "status": "ACTIVE"}


class LevelChangeRequest(BaseModel):
    level: str


@router.patch("/users/{login}/level", dependencies=[ADMIN])
def change_user_level(login: str, request: Request, body: LevelChangeRequest) -> dict[str, Any]:
    """권한 레벨 변경 (B8 감사 확장) — before/after 를 sys_history 에 기록."""
    level = body.level.strip().upper()
    if level not in LEVEL_RANK:
        raise HTTPException(422, detail=f"레벨은 {'/'.join(LEVEL_RANK)} 중 하나여야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT user_id, user_level FROM sys_user WHERE tenant_id=%s AND login_id=%s",
            (tid, login))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        if row[1] == level:
            raise HTTPException(409, detail=f"{login} 은 이미 {level} 입니다")
        cur.execute(
            """UPDATE sys_user SET user_level=%s, updated_by=%s, updated_at=now()
               WHERE user_id=%s""", (level, request.state.login, row[0]))
        _audit(cur, tid, "sys_user", row[0], "LEVEL_CHANGE", request.state.user_id,
               after={"level": level}, before={"level": row[1]})
    return {"login": login, "level": level}


# ── F2 — 사용자 등록·프로필 수정·삭제 (M-14-6, SYS-005) ──

_LOGIN_RE = re.compile(r"^[a-z0-9][a-z0-9._-]{2,49}$")


class UserCreate(BaseModel):
    login: str
    name: str
    department: str = ""
    email: str = ""
    level: str = "GENERAL"
    initialPassword: str


@router.post("/users", status_code=201, dependencies=[ADMIN])
def create_user(body: UserCreate, request: Request) -> dict[str, Any]:
    """사용자 신규 등록 (F2) — 초기 비밀번호는 관리자가 지정, USER_CREATE 감사."""
    login = body.login.strip().lower()
    if not _LOGIN_RE.fullmatch(login):
        raise HTTPException(422, detail="login 은 소문자·숫자·._- 3~50자 (첫 글자는 영숫자)")
    if not body.name.strip():
        raise HTTPException(422, detail="이름을 입력하십시오")
    level = body.level.strip().upper()
    if level not in ("GENERAL", "SETUP", "ADMIN"):
        raise HTTPException(422, detail="레벨은 GENERAL/SETUP/ADMIN 중 하나 (PLATFORM 은 테넌트 관리 전용)")
    if len(body.initialPassword) < 4:
        raise HTTPException(422, detail="초기 비밀번호는 4자 이상이어야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM sys_user WHERE tenant_id=%s AND login_id=%s", (tid, login))
        if cur.fetchone():
            raise HTTPException(409, detail=f"이미 존재하는 login: {login}")
        cur.execute(
            """INSERT INTO sys_user (tenant_id, login_id, user_name, email, password_hash,
               department, user_level, status, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,'ACTIVE',%s) RETURNING user_id""",
            (tid, login, body.name.strip(), body.email.strip() or None,
             hashlib.sha256(body.initialPassword.encode()).hexdigest(),
             body.department.strip() or None, level, request.state.login))
        uid = cur.fetchone()[0]
        _audit(cur, tid, "sys_user", uid, "USER_CREATE", request.state.user_id,
               after={"login": login, "name": body.name.strip(), "level": level})
    return {"login": login, "name": body.name.strip(), "dept": body.department.strip() or "-",
            "level": level, "status": "ACTIVE"}


class UserPatch(BaseModel):
    name: str | None = None
    department: str | None = None
    email: str | None = None


@router.patch("/users/{login}", dependencies=[ADMIN])
def patch_user(login: str, body: UserPatch, request: Request) -> dict[str, Any]:
    """사용자 프로필 수정 (F2) — 이름·부서·이메일 (레벨은 /level, 상태는 /active·/unlock)."""
    fields = {k: v for k, v in (("user_name", body.name), ("department", body.department),
                                ("email", body.email)) if v is not None}
    if not fields:
        raise HTTPException(422, detail="수정할 필드가 없습니다 (name/department/email)")
    if "user_name" in fields and not fields["user_name"].strip():
        raise HTTPException(422, detail="이름은 비울 수 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT user_id, user_name, department, email FROM sys_user
               WHERE tenant_id=%s AND login_id=%s""", (tid, login))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        before = {"name": row[1], "department": row[2], "email": row[3]}
        sets = ", ".join(f"{k}=%s" for k in fields)
        cur.execute(
            f"UPDATE sys_user SET {sets}, updated_by=%s, updated_at=now() WHERE user_id=%s",
            (*[v.strip() or None for v in fields.values()], request.state.login, row[0]))
        _audit(cur, tid, "sys_user", row[0], "USER_UPDATE", request.state.user_id,
               after={k: (v.strip() or None) for k, v in
                      (("name", body.name), ("department", body.department),
                       ("email", body.email)) if v is not None},
               before={k: before[k] for k in
                       ("name", "department", "email")
                       if (k == "name" and body.name is not None)
                       or (k == "department" and body.department is not None)
                       or (k == "email" and body.email is not None)})
    return {"login": login, "updated": sorted(fields)}


@router.delete("/users/{login}", dependencies=[ADMIN])
def delete_user(login: str, request: Request) -> dict[str, Any]:
    """사용자 삭제 (F2) — 업무 이력이 있으면 409 (비활성화 사용), 무참조만 하드 삭제."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, login))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        uid = row[0]
        if uid == request.state.user_id:
            raise HTTPException(422, detail="본인 계정은 삭제할 수 없습니다")
        for tbl, col, label in (
            ("sys_history", "actor_id", "감사 이력"),
            ("sys_approval_request", "requester_id", "승인 요청"),
            ("sys_approval_request", "approver_id", "승인 처리"),
            ("sys_notification", "user_id", "알림"),
            ("prj_project", "manager_id", "프로젝트 담당"),
            ("erp_process_event", "assignee_id", "업무 이벤트"),
        ):
            cur.execute(f"SELECT count(*) FROM {tbl} WHERE {col}=%s", (uid,))
            c = cur.fetchone()[0]
            if c:
                raise HTTPException(
                    409, detail=f"참조 존재 — {label} {c}건 (삭제 대신 비활성화를 사용하십시오)")
        cur.execute("DELETE FROM sys_user_role WHERE user_id=%s", (uid,))
        cur.execute("DELETE FROM sys_user WHERE user_id=%s", (uid,))
        _audit(cur, tid, "sys_user", uid, "USER_DELETE", request.state.user_id,
               after={"login": login})
    return {"deleted": login}


# ── SVC-09 ERP 이벤트 (업무함 M-15-3 · Dashboard 경고) ──
@router.get("/erp/events")
def erp_events() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT e.event_id, d.proc_code, d.proc_name, p.project_no,
                      COALESCE(u.user_name,'-'), to_char(e.due_date,'MM-DD'),
                      e.status, (e.due_date < CURRENT_DATE AND e.status <> 'DONE')
               FROM erp_process_event e
               JOIN erp_process_def d ON d.proc_def_id=e.proc_def_id
               JOIN prj_project p ON p.project_id=e.project_id
               LEFT JOIN sys_user u ON u.user_id=e.assignee_id
               WHERE e.tenant_id=%s ORDER BY e.event_id""", (tid,))
        return [
            {"eventId": r[0], "code": r[1], "procName": r[2], "project": r[3],
             "owner": r[4], "deadline": r[5] or "-",
             "delayed": bool(r[7]),
             "status": "DONE" if r[6] == "DONE" else ("지연" if r[7] else
                       ("진행" if r[6] == "IN_PROGRESS" else "TODO"))}
            for r in cur.fetchall()
        ]


class CompleteRequest(BaseModel):
    comment: str = ""


@router.post("/erp/events/{event_id}/complete")
def complete_event(event_id: int, body: CompleteRequest) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE erp_process_event SET status='DONE', done_at=now(),
               data=COALESCE(data,'{}'::jsonb) || %s::jsonb
               WHERE tenant_id=%s AND event_id=%s AND status<>'DONE'
               RETURNING event_id""",
            (json.dumps({"comment": body.comment}), tid, event_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"완료 처리 대상 아님: {event_id}")
    return {"eventId": event_id, "status": "DONE"}


# ── B6 — 이벤트 재배정·에스컬레이션 ──

class ReassignRequest(BaseModel):
    assignee: str      # login_id
    comment: str = ""


@router.patch("/erp/events/{event_id}", dependencies=[SETUP])
def reassign_event(event_id: int, request: Request, body: ReassignRequest) -> dict[str, Any]:
    """재배정 — 담당자 변경 + 새 담당자 알림 + 이력 (ERP-031)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT user_id, user_name FROM sys_user WHERE tenant_id=%s AND login_id=%s",
            (tid, body.assignee.strip()))
        u = cur.fetchone()
        if not u:
            raise HTTPException(404, detail=f"사용자 없음: {body.assignee}")
        cur.execute(
            """UPDATE erp_process_event SET assignee_id=%s,
               data=COALESCE(data,'{}'::jsonb) || %s::jsonb
               WHERE tenant_id=%s AND event_id=%s RETURNING event_id""",
            (u[0], json.dumps({"reassign": body.comment or body.assignee}), tid, event_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"이벤트 없음: {event_id}")
        _notify(cur, tid, u[0], "TASK_ASSIGNED", f"업무 재배정 — 이벤트 #{event_id}", "/common")
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'erp_process_event',%s,'REASSIGN',%s,%s)""",
            (tid, event_id, request.state.user_id,
             json.dumps({"assignee": body.assignee, "comment": body.comment})))
    return {"eventId": event_id, "assignee": u[1]}


class EscalateRequest(BaseModel):
    reason: str = ""


@router.post("/erp/events/{event_id}/escalate", dependencies=[SETUP])
def escalate_event(event_id: int, request: Request, body: EscalateRequest) -> dict[str, Any]:
    """에스컬레이션 — ADMIN 전원 알림 + 이력 (이상 경고 상위 보고)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM erp_process_event WHERE tenant_id=%s AND event_id=%s", (tid, event_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"이벤트 없음: {event_id}")
        cur.execute(
            """SELECT user_id FROM sys_user
               WHERE tenant_id=%s AND user_level IN ('ADMIN','PLATFORM')""", (tid,))
        n = 0
        for (uid,) in cur.fetchall():
            _notify(cur, tid, uid, "ESCALATION",
                    f"에스컬레이션 — 이벤트 #{event_id}: {body.reason[:60] or '지연 상위 보고'}", "/erp")
            n += 1
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'erp_process_event',%s,'ESCALATE',%s,%s)""",
            (tid, event_id, request.state.user_id, json.dumps({"reason": body.reason})))
    return {"eventId": event_id, "notified": n}


@router.post("/erp/events/escalate-overdue", dependencies=[SETUP])
def escalate_overdue(request: Request) -> dict[str, Any]:
    """지연 이벤트 자동 에스컬레이션 (C4 / JOB-05) — 기한 초과 미처리 이벤트를 ADMIN 상위 보고.

    멱등: data.autoEscalated 플래그로 재통지 방지 (스케줄러/수동 재실행 안전).
    """
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT e.event_id, e.due_date, p.project_no, d.proc_name
               FROM erp_process_event e
               JOIN prj_project p ON p.project_id=e.project_id
               JOIN erp_process_def d ON d.proc_def_id=e.proc_def_id
               WHERE e.tenant_id=%s AND e.status<>'DONE' AND e.due_date < CURRENT_DATE
                 AND COALESCE((e.data->>'autoEscalated')::boolean, false) = false""", (tid,))
        overdue = cur.fetchall()
        cur.execute(
            "SELECT user_id FROM sys_user WHERE tenant_id=%s AND user_level IN ('ADMIN','PLATFORM')", (tid,))
        admins = [r[0] for r in cur.fetchall()]
        escalated = []
        for ev_id, due, proj, pname in overdue:
            for uid in admins:
                _notify(cur, tid, uid, "ESCALATION",
                        f"지연 자동 에스컬레이션 — {proj} {pname} (기한 {due} 초과)", "/erp")
            cur.execute(
                """UPDATE erp_process_event
                   SET data = COALESCE(data,'{}'::jsonb) || '{"autoEscalated":true}'::jsonb,
                       updated_at=now() WHERE event_id=%s""", (ev_id,))
            _audit(cur, tid, "erp_process_event", ev_id, "AUTO_ESCALATE", request.state.user_id,
                   {"project": proj, "process": pname, "due": str(due)})
            escalated.append({"eventId": ev_id, "project": proj, "process": pname, "due": str(due)})
    return {"escalated": len(escalated), "admins": len(admins), "events": escalated}


# ── SVC-09 Project (S-3-5) ──
STAGE_MAP = {"기술 제안": "TECH_PROPOSAL", "견적": "QUOTE", "협의": "NEGOTIATION",
             "계약": "CONTRACT", "계약 변경": "CONTRACT_CHANGE", "종료": "CLOSED"}
STAGE_LABEL = {v: k for k, v in STAGE_MAP.items()}


# D9 — 동시 편집 보호: 낙관적 잠금 버전 토큰 (updated_at, 없으면 created_at)
_VER_FMT = "YYYY-MM-DD HH24:MI:SS.US"


# ── U9 Project 중심 대화 (SYS-018·M-15-5) — 코멘트 스레드 ──

class CommentCreate(BaseModel):
    body: str


@router.get("/projects/{project_no}/comments")
def project_comments(project_no: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT comment_id, author, body, to_char(created_at,'MM-DD HH24:MI')
               FROM sys_project_comment WHERE tenant_id=%s AND project_no=%s
               ORDER BY comment_id DESC LIMIT 100""", (tid, project_no.strip()))
        return [{"id": r[0], "author": r[1], "body": r[2], "at": r[3]} for r in cur.fetchall()]


@router.post("/projects/{project_no}/comments", status_code=201)
def project_comment_add(project_no: str, request: Request, body: CommentCreate) -> dict[str, Any]:
    """프로젝트 대화 등록 (U9) — 이력 관리(수정 불가·본인/ADMIN 삭제만)."""
    text = body.body.strip()
    if not text:
        raise HTTPException(422, detail="내용을 입력하십시오")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM prj_project WHERE tenant_id=%s AND project_no=%s", (tid, project_no.strip()))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"프로젝트 없음: {project_no}")
        cur.execute(
            """INSERT INTO sys_project_comment (tenant_id, project_no, author, body)
               VALUES (%s,%s,%s,%s) RETURNING comment_id""",
            (tid, project_no.strip(), request.state.login, text[:1000]))
        cid = cur.fetchone()[0]
    return {"id": cid}


@router.delete("/projects/comments/{comment_id}")
def project_comment_delete(comment_id: int, request: Request) -> dict[str, Any]:
    """대화 삭제 — 본인 또는 ADMIN 이상."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT author FROM sys_project_comment WHERE tenant_id=%s AND comment_id=%s", (tid, comment_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"코멘트 없음: {comment_id}")
        if row[0] != request.state.login and LEVEL_RANK.get(request.state.level, 0) < LEVEL_RANK.get("ADMIN", 99):
            raise HTTPException(403, detail="본인 또는 ADMIN 만 삭제할 수 있습니다")
        cur.execute("DELETE FROM sys_project_comment WHERE tenant_id=%s AND comment_id=%s", (tid, comment_id))
    return {"deleted": True}


@router.get("/projects/{project_no}")
def get_project(project_no: str) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT project_no, project_name, project_type, sales_stage, client_contact,
                       to_char(COALESCE(updated_at, created_at),'{_VER_FMT}')
               FROM prj_project WHERE tenant_id=%s AND project_no=%s""", (tid, project_no))
        r = cur.fetchone()
    if not r:
        raise HTTPException(404, detail=f"project not found: {project_no}")
    return {"projectNo": r[0], "projectName": r[1], "projectType": r[2] or "Client",
            "stage": STAGE_LABEL.get(r[3], r[3]), "clientContact": r[4] or "",
            "updatedAt": r[5]}   # D9 — 낙관적 잠금 버전 토큰


class StagePatch(BaseModel):
    stage: str | None = None            # 영업 단계 (미지정 시 메타만 수정)
    projectName: str | None = None      # 프로젝트명 수정
    client: str | None = None           # 고객사명 수정 (없으면 CUSTOMER 자동 생성, "" = 연결 해제)
    dueDate: str | None = None          # 납기 YYYY-MM-DD ("" = 해제)
    baseUpdatedAt: str = ""             # D9 — 조회 시점 버전 (불일치 시 409)


@router.patch("/projects/{project_no}", dependencies=[SETUP])
def patch_project(project_no: str, request: Request, body: StagePatch) -> dict[str, Any]:
    """프로젝트 단계·메타(명/고객/납기) 수정 — D9 낙관적 잠금·감사."""
    code = None
    if body.stage is not None:
        code = STAGE_MAP.get(body.stage)
        if not code:
            raise HTTPException(422, detail=f"알 수 없는 영업 단계: {body.stage}")
    if body.projectName is not None and not body.projectName.strip():
        raise HTTPException(422, detail="Project명은 비울 수 없습니다")
    due = None
    if body.dueDate is not None and body.dueDate.strip():
        try:
            due = date.fromisoformat(body.dueDate.strip())
        except ValueError:
            raise HTTPException(422, detail=f"납기 날짜 형식 오류: {body.dueDate} (YYYY-MM-DD)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT project_id, to_char(COALESCE(updated_at, created_at),'{_VER_FMT}')
               FROM prj_project WHERE tenant_id=%s AND project_no=%s""", (tid, project_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"project not found: {project_no}")
        # D9 — 낙관적 잠금: 조회 이후 다른 사용자가 먼저 수정했으면 409
        if body.baseUpdatedAt.strip() and body.baseUpdatedAt.strip() != row[1]:
            raise HTTPException(
                409, detail="다른 사용자가 먼저 수정했습니다 — 재조회 후 다시 시도하십시오 (동시 편집 충돌)")

        sets: dict[str, Any] = {}
        after: dict[str, Any] = {}
        if code is not None:
            sets["sales_stage"] = code
            after["stage"] = code
        if body.projectName is not None:
            sets["project_name"] = body.projectName.strip()
            after["projectName"] = body.projectName.strip()
        if body.dueDate is not None:
            sets["due_date"] = due               # "" → NULL 로 해제
            after["dueDate"] = due.isoformat() if due else None
        if body.client is not None:
            client = body.client.strip()
            customer_id = None
            if client:
                cur.execute(
                    "SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s",
                    (tid, client))
                c = cur.fetchone()
                if c:
                    customer_id = c[0]
                else:
                    cur.execute(
                        """INSERT INTO com_company (tenant_id, company_type, company_name)
                           VALUES (%s,'CUSTOMER',%s) RETURNING company_id""", (tid, client))
                    customer_id = cur.fetchone()[0]
            sets["customer_id"] = customer_id     # "" → NULL 로 해제
            after["client"] = client

        if not sets:
            raise HTTPException(422, detail="수정할 항목이 없습니다")

        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE prj_project SET {assign}, updated_by=%s, updated_at=now() "
            "WHERE tenant_id=%s AND project_no=%s",
            (*sets.values(), request.state.login, tid, project_no))
        action = "STAGE" if set(sets) == {"sales_stage"} else "PROJECT_UPDATE"
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'prj_project',%s,%s,%s,%s)""",
            (tid, row[0], action, request.state.user_id, json.dumps(after)))
    return {"projectNo": project_no, "stage": body.stage, "updated": sorted(after)}


@router.get("/projects")
def list_projects() -> list[dict[str, Any]]:
    """프로젝트 목록 (F1) — S-3-5 대장·타이틀바 컨텍스트 공용."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.project_no, p.project_name, p.project_type, p.sales_stage,
                      p.client_contact, p.status, COALESCE(p.note,''),
                      to_char(p.created_at,'YYYY-MM-DD'), COALESCE(c.company_name,''),
                      to_char(p.due_date,'YYYY-MM-DD')
               FROM prj_project p
               LEFT JOIN com_company c ON c.company_id=p.customer_id
               WHERE p.tenant_id=%s ORDER BY p.project_id DESC""", (tid,))
        rows = cur.fetchall()
    return [{"projectNo": r[0], "projectName": r[1], "projectType": r[2] or "Client",
             "stage": STAGE_LABEL.get(r[3], r[3]), "clientContact": r[4] or "",
             "status": r[5], "item": r[6], "registeredAt": r[7], "client": r[8],
             "dueDate": r[9] or ""} for r in rows]


class ProjectCreate(BaseModel):
    projectName: str
    projectType: str = "Client"
    item: str = ""
    client: str = ""
    clientContact: str = ""


@router.post("/projects", status_code=201, dependencies=[SETUP])
def create_project(body: ProjectCreate, request: Request) -> dict[str, Any]:
    """프로젝트 신규 등록 (F1) — PS 자동 채번. item 은 note 컬럼에 보관 (스키마 무변경)."""
    name = body.projectName.strip()
    if not name:
        raise HTTPException(422, detail="Project명을 입력하십시오")
    if body.projectType not in ("Client", "Stock", "R&D"):
        raise HTTPException(422, detail=f"알 수 없는 Type: {body.projectType}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        # PS 자동 채번 — 'PS-<숫자>' 단일 세그먼트 최대치 + 1 (시드 PS-612 → PS-613 …)
        cur.execute(
            r"""SELECT COALESCE(MAX((substring(project_no FROM '^PS-(\d+)$'))::int), 600)
               FROM prj_project WHERE tenant_id=%s AND project_no ~ '^PS-\d+$'""", (tid,))
        seq = (cur.fetchone()[0] or 600) + 1
        project_no = f"PS-{seq}"
        for _ in range(50):
            cur.execute("SELECT 1 FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                        (tid, project_no))
            if not cur.fetchone():
                break
            seq += 1
            project_no = f"PS-{seq}"
        customer_id = None
        client = body.client.strip()
        if client:
            # 고객사 자동 연결 — 없으면 CUSTOMER 로 생성 (단가 공급처 패턴 재사용)
            cur.execute(
                """SELECT company_id FROM com_company
                   WHERE tenant_id=%s AND company_name=%s""", (tid, client))
            c = cur.fetchone()
            if c:
                customer_id = c[0]
            else:
                cur.execute(
                    """INSERT INTO com_company (tenant_id, company_type, company_name)
                       VALUES (%s,'CUSTOMER',%s) RETURNING company_id""", (tid, client))
                customer_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO prj_project (tenant_id, project_no, project_name, project_type,
               customer_id, client_contact, note, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING project_id""",
            (tid, project_no, name, body.projectType, customer_id,
             body.clientContact.strip(), body.item.strip(), request.state.login))
        pid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'prj_project',%s,'PROJECT_CREATE',%s,%s)""",
            (tid, pid, request.state.user_id,
             json.dumps({"projectNo": project_no, "name": name, "type": body.projectType})))
    return {"projectNo": project_no, "projectName": name, "projectType": body.projectType,
            "stage": "기술 제안", "item": body.item, "client": client,
            "clientContact": body.clientContact,
            "status": "IN_PROGRESS", "registeredAt": date.today().isoformat()}


@router.delete("/projects/{project_no}", dependencies=[SETUP])
def delete_project(project_no: str, request: Request) -> dict[str, Any]:
    """프로젝트 삭제 (F1) — 기술 제안 단계 + 무참조만 (참조 존재 시 409)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("""SELECT project_id, sales_stage FROM prj_project
                       WHERE tenant_id=%s AND project_no=%s""", (tid, project_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"project not found: {project_no}")
        pid, stage = row
        if stage != "TECH_PROPOSAL":
            raise HTTPException(
                409, detail=f"기술 제안 단계만 삭제 가능 (현재 {STAGE_LABEL.get(stage, stage)})")
        for tbl, label in (("cpq_selection", "견적 선택"), ("cst_quotation", "견적서"),
                           ("dwg_file", "파일"), ("dwg_drawing", "도면"),
                           ("erp_process_event", "이벤트")):
            cur.execute(
                f"SELECT count(*) FROM {tbl} WHERE tenant_id=%s AND project_id=%s", (tid, pid))
            n = cur.fetchone()[0]
            if n:
                raise HTTPException(409, detail=f"참조 존재 — {label} {n}건 (삭제 불가)")
        cur.execute("DELETE FROM prj_project WHERE project_id=%s", (pid,))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'prj_project',%s,'PROJECT_DELETE',%s,%s)""",
            (tid, pid, request.state.user_id, json.dumps({"projectNo": project_no})))
    return {"deleted": project_no}


# ── SVC-08 단가 대장 ──
PRICE_SORT = {"code": "pc.main_code", "price": "p.price", "from": "p.valid_from",
              "supplier": "cc.company_name"}


@router.get("/prices")
def prices(request: Request, sort: str = "", dir: str = "asc") -> list[dict[str, Any]]:
    # F8 — 대량 대장 서버 정렬 (화이트리스트 컬럼만; 기본 = 코드·적용일 역순)
    order = "pc.main_code, p.valid_from DESC"
    if sort in PRICE_SORT:
        order = f"{PRICE_SORT[sort]} {'DESC' if dir == 'desc' else 'ASC'}"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT pc.main_code, pc.code_name, COALESCE(cc.company_name,'-'),
                      p.price, p.price_source, p.valid_from, p.valid_to,
                      (p.valid_from <= CURRENT_DATE AND
                       (p.valid_to IS NULL OR p.valid_to >= CURRENT_DATE)), p.price_id
               FROM cst_price p
               JOIN product_code pc ON pc.product_code_id=p.product_code_id
               LEFT JOIN com_company cc ON cc.company_id=p.supplier_id
               WHERE p.tenant_id=%s AND pc.main_code IN ('FDV-480','KDC-1','EWT-3')
               ORDER BY {order}""", (tid,))
        rows = cur.fetchall()
        # 1.5 — 정보 접근 권한: 단가·거래처명 마스킹 (모드 미설정 시 full)
        pm = _info_mode(cur, tid, request, "price")
        qm = _info_mode(cur, tid, request, "partner")
        if pm != "full" or qm != "full":
            _audit(cur, tid, "cst_price", 0, "MASKED_READ", request.state.user_id,
                   {"price": pm, "partner": qm})
        return [
            {"code": r[0], "name": r[1], "supplier": _mask_text(r[2], qm),
             "price": _mask_num(float(r[3]), pm),
             "source": SOURCE_LABEL[r[4]], "from": r[5].isoformat(),
             "to": r[6].isoformat() if r[6] else None, "active": bool(r[7]), "priceId": r[8],
             "maskMode": pm}
            for r in rows
        ]


# ── B3 — 단가 등록 (M-12-5 ＋ 단가 등록) ──

SOURCE_CODE = {"견적적용": "APPLIED", "구매": "PURCHASE", "재고": "STOCK", "견적": "QUOTE"}


class PriceCreate(BaseModel):
    code: str
    supplier: str = "-"
    price: float
    source: str = "QUOTE"        # enum 또는 한글 라벨
    validFrom: str               # YYYY-MM-DD
    validTo: str | None = None


@router.post("/prices", status_code=201, dependencies=[SETUP])
def create_price(body: PriceCreate) -> dict[str, Any]:
    src = SOURCE_CODE.get(body.source.strip(), body.source.strip().upper())
    if src not in ("APPLIED", "PURCHASE", "STOCK", "QUOTE"):
        raise HTTPException(422, detail=f"단가 Table 구분 오류: {body.source}")
    if body.price <= 0:
        raise HTTPException(422, detail="단가는 0 보다 커야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
            (tid, body.code.strip()))
        pc = cur.fetchone()
        if not pc:
            raise HTTPException(404, detail=f"코드 없음: {body.code}")
        supplier_id = None
        name = body.supplier.strip()
        if name and name != "-":
            cur.execute(
                "SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s",
                (tid, name))
            row = cur.fetchone()
            if row:
                supplier_id = row[0]
            else:
                cur.execute(
                    """INSERT INTO com_company (tenant_id, company_name, company_type)
                       VALUES (%s,%s,'SUPPLIER') RETURNING company_id""", (tid, name))
                supplier_id = cur.fetchone()[0]
        try:
            cur.execute(
                """INSERT INTO cst_price (tenant_id, product_code_id, supplier_id,
                   price, price_source, valid_from, valid_to)
                   VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING price_id""",
                (tid, pc[0], supplier_id, body.price, src,
                 body.validFrom, body.validTo))
        except Exception as e:  # EXCLUDE 기간 중복 (DB v0.5)
            if "exclusion" in str(e).lower() or "overlap" in str(e).lower():
                raise HTTPException(409, detail="기간 중복 — 동일 Table·Code 의 유효기간이 겹칩니다 (EXCLUDE)") from e
            raise
        price_id = cur.fetchone()[0]
    return {"priceId": price_id, "source": src}


@router.post("/prices/import-excel", dependencies=[SETUP])
async def import_prices_excel(uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """단가 Excel Import — 헤더: Code·공급처·단가·Table·적용시작·적용종료 (행 단위 등록)."""
    import openpyxl
    data = await uploadedFile.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(422, detail="Excel 파일이 아닙니다 (.xlsx)")
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = ["Code", "단가", "Table", "적용시작"]
    if any(h not in header for h in required):
        raise HTTPException(422, detail=f"헤더 불일치 — 필요 열: {required} (+공급처·적용종료 선택)")
    idx = {h: header.index(h) for h in header if h}
    inserted = 0
    rejected: list[str] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            def cell(col: str) -> str:
                i = idx.get(col)
                v = row[i].value if i is not None else None
                return str(v).strip() if v is not None else ""
            code = cell("Code")
            if not code:
                continue
            try:
                price = float(cell("단가"))
                src = SOURCE_CODE.get(cell("Table"), cell("Table").upper())
                if src not in ("APPLIED", "PURCHASE", "STOCK", "QUOTE"):
                    raise ValueError(f"Table 구분 오류: {cell('Table')}")
                cur.execute(
                    "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
                    (tid, code))
                pc = cur.fetchone()
                if not pc:
                    raise ValueError(f"코드 없음: {code}")
                supplier_id = None
                name = cell("공급처")
                if name and name != "-":
                    cur.execute(
                        "SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s",
                        (tid, name))
                    found = cur.fetchone()
                    if found:
                        supplier_id = found[0]
                    else:
                        cur.execute(
                            """INSERT INTO com_company (tenant_id, company_name, company_type)
                               VALUES (%s,%s,'SUPPLIER') RETURNING company_id""", (tid, name))
                        supplier_id = cur.fetchone()[0]
                # autocommit — 실패 행만 개별 거부하고 계속 (EXCLUDE 등)
                try:
                    cur.execute(
                        """INSERT INTO cst_price (tenant_id, product_code_id, supplier_id,
                           price, price_source, valid_from, valid_to)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (tid, pc[0], supplier_id, price, src,
                         cell("적용시작"), cell("적용종료") or None))
                    inserted += 1
                except Exception as e:  # noqa: BLE001
                    if "exclusion" in str(e).lower() or "overlap" in str(e).lower():
                        raise ValueError("기간 중복 (EXCLUDE)") from e
                    raise ValueError(str(e)[:80]) from e
            except ValueError as e:
                rejected.append(f"{r_i}행 {code}: {e}")
    return {"inserted": inserted, "rejected": rejected}


# ── B5 — 통합 검색 (⌘K · M-15-x) ──

@router.get("/search")
def global_search(q: str, request: Request) -> dict[str, list[dict[str, Any]]]:
    """통합 검색 (B5 + F6 확장) — 코드·문서·파일 + 부품·공급처·창고·매크로·프로젝트
    (+사용자는 SETUP 이상). 화면 검색은 프론트 레지스트리에서 병합."""
    term = q.strip()
    empty: dict[str, list[dict[str, Any]]] = {
        "codes": [], "docs": [], "files": [], "parts": [], "companies": [],
        "warehouses": [], "macros": [], "projects": [], "users": []}
    if len(term) < 2:
        return empty
    like = f"%{term}%"
    out = dict(empty)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT main_code, code_name FROM product_code
               WHERE tenant_id=%s AND (main_code ILIKE %s OR code_name ILIKE %s)
               ORDER BY main_code LIMIT 8""", (tid, like, like))
        out["codes"] = [{"code": r[0], "name": r[1]} for r in cur.fetchall()]
        cur.execute(
            """SELECT DISTINCT doc_no, title, management_grade FROM doc_control
               WHERE tenant_id=%s AND (doc_no ILIKE %s OR title ILIKE %s)
               ORDER BY doc_no LIMIT 8""", (tid, like, like))
        out["docs"] = [{"docNo": r[0], "title": r[1], "grade": r[2]} for r in cur.fetchall()]
        cur.execute(
            """SELECT file_id, file_name, file_type FROM dwg_file
               WHERE tenant_id=%s AND file_name ILIKE %s
               ORDER BY file_id DESC LIMIT 8""", (tid, like))
        out["files"] = [{"fileId": r[0], "name": r[1], "type": r[2]} for r in cur.fetchall()]
        # ── F6 확장 그룹 ──
        cur.execute(
            """SELECT part_no, part_name FROM prt_part
               WHERE tenant_id=%s AND (part_no ILIKE %s OR part_name ILIKE %s)
               ORDER BY part_no LIMIT 8""", (tid, like, like))
        out["parts"] = [{"partNo": r[0], "name": r[1]} for r in cur.fetchall()]
        cur.execute(
            """SELECT company_id, company_name, company_type FROM com_company
               WHERE tenant_id=%s AND company_name ILIKE %s
               ORDER BY company_name LIMIT 8""", (tid, like))
        out["companies"] = [{"companyId": r[0], "name": r[1], "companyType": r[2]}
                            for r in cur.fetchall()]
        cur.execute(
            """SELECT location_code, location_name, location_type FROM erp_warehouse
               WHERE tenant_id=%s AND (location_code ILIKE %s OR location_name ILIKE %s)
               ORDER BY location_code LIMIT 8""", (tid, like, like))
        out["warehouses"] = [{"code": r[0], "name": r[1], "locationType": r[2]}
                             for r in cur.fetchall()]
        cur.execute(
            """SELECT DISTINCT ON (macro_name) macro_name, apply_type, status FROM tbx_macro
               WHERE tenant_id=%s AND macro_name ILIKE %s
               ORDER BY macro_name, version DESC LIMIT 8""", (tid, like))
        out["macros"] = [{"name": r[0], "applyType": r[1], "status": r[2]}
                         for r in cur.fetchall()]
        cur.execute(
            """SELECT project_no, project_name, sales_stage FROM prj_project
               WHERE tenant_id=%s AND (project_no ILIKE %s OR project_name ILIKE %s)
               ORDER BY project_id DESC LIMIT 8""", (tid, like, like))
        out["projects"] = [{"projectNo": r[0], "name": r[1],
                            "stage": STAGE_LABEL.get(r[2], r[2])} for r in cur.fetchall()]
        # 사용자 — 읽기 자체가 SETUP 가드(M-14-6)이므로 동일 레벨에만 노출
        if LEVEL_RANK.get(getattr(request.state, "level", "GENERAL"), 0) >= LEVEL_RANK["SETUP"]:
            cur.execute(
                """SELECT login_id, user_name, user_level FROM sys_user
                   WHERE tenant_id=%s AND (login_id ILIKE %s OR user_name ILIKE %s)
                   ORDER BY login_id LIMIT 8""", (tid, like, like))
            out["users"] = [{"login": r[0], "name": r[1], "level": r[2]}
                            for r in cur.fetchall()]
    return out


# ── SYS-012 이력 (M-15-9) ──
HISTORY_SORT = {"at": "h.acted_at", "action": "h.action", "target": "h.target_table",
                "by": "u.user_name"}


@router.get("/history")
def history(limit: int = 20, sort: str = "", dir: str = "desc") -> list[dict[str, Any]]:
    # F8 — 서버 정렬 (화이트리스트; 기본 = 최신순)
    order = "h.history_id DESC"
    if sort in HISTORY_SORT:
        order = f"{HISTORY_SORT[sort]} {'DESC' if dir == 'desc' else 'ASC'}"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT to_char(h.acted_at,'MM-DD HH24:MI'), h.target_table||' #'||h.target_id,
                      h.action, u.user_name, h.history_id, h.before_data, h.after_data
               FROM sys_history h JOIN sys_user u ON u.user_id=h.actor_id
               WHERE h.tenant_id=%s ORDER BY {order} LIMIT %s""", (tid, limit))
        return [
            {"at": r[0], "target": r[1], "action": r[2], "by": r[3],
             "historyId": r[4], "before": r[5], "after": r[6]}   # F7 — diff 뷰어
            for r in cur.fetchall()
        ]


# ── D9 전용 감사 조회 (sys_history 필터) — 기간·사용자·작업 ──
def _audit_where(tid: int, fromDate: str, toDate: str, user: str,
                 action: str, target: str) -> tuple[str, list[Any]]:
    clauses, params = ["h.tenant_id=%s"], [tid]
    if fromDate.strip():
        clauses.append("h.acted_at >= %s::date")
        params.append(fromDate.strip())
    if toDate.strip():
        clauses.append("h.acted_at < (%s::date + 1)")
        params.append(toDate.strip())
    if user.strip():
        clauses.append("u.login_id = %s")
        params.append(user.strip())
    if action.strip():
        clauses.append("h.action ILIKE %s")
        params.append(f"%{action.strip()}%")
    if target.strip():
        clauses.append("h.target_table ILIKE %s")
        params.append(f"%{target.strip()}%")
    return " AND ".join(clauses), params


@router.get("/audit", dependencies=[ADMIN])
def audit_query(fromDate: str = "", toDate: str = "", user: str = "",
                action: str = "", target: str = "", limit: int = 200) -> dict[str, Any]:
    """전용 감사 조회 (D9) — 기간/사용자/작업/대상 필터. ADMIN 전용."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        where, params = _audit_where(tid, fromDate, toDate, user, action, target)
        cur.execute(
            f"""SELECT to_char(h.acted_at,'YYYY-MM-DD HH24:MI'),
                       h.target_table||' #'||h.target_id, h.action, u.user_name, u.login_id,
                       h.history_id, h.before_data, h.after_data
                FROM sys_history h JOIN sys_user u ON u.user_id=h.actor_id
                WHERE {where} ORDER BY h.history_id DESC LIMIT %s""",
            (*params, max(1, min(limit, 2000))))
        rows = [{"at": r[0], "target": r[1], "action": r[2], "by": r[3], "login": r[4],
                 "historyId": r[5], "before": r[6], "after": r[7]} for r in cur.fetchall()]
        # 필터 드롭다운용 facet (전체 tenant 기준 distinct)
        cur.execute("SELECT DISTINCT action FROM sys_history WHERE tenant_id=%s ORDER BY action", (tid,))
        actions = [r[0] for r in cur.fetchall()]
        cur.execute(
            """SELECT DISTINCT u.login_id FROM sys_history h JOIN sys_user u ON u.user_id=h.actor_id
               WHERE h.tenant_id=%s ORDER BY u.login_id""", (tid,))
        users = [r[0] for r in cur.fetchall()]
    return {"rows": rows, "actions": actions, "users": users, "count": len(rows)}


# ── 잔여 mock 실데이터화 (v4.0) — 치수 정의·Macro 목록·공정 정의·역참조 ──

@router.get("/drawings/dimensions")
def drawing_dimensions(drawing: str = "KDCR 3-13") -> list[dict[str, Any]]:
    """Design Rule 치수 Set-up — dwg_dimension + tbx_macro (Design Editor·부품 상세 실데이터)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT d.dim_label, d.dim_type, d.variant_value, m.macro_expr
               FROM dwg_dimension d
               JOIN dwg_drawing w ON w.drawing_id=d.drawing_id
               LEFT JOIN tbx_macro m ON m.macro_id=d.macro_id
               WHERE d.tenant_id=%s AND w.drawing_no=%s
               ORDER BY d.dim_label""", (tid, drawing))
        return [
            {"no": r[0],
             "value": (r[3] if r[3] else (r[2] or "")),
             "binding": "MACRO" if r[3] else "VARIANT",
             "kind": r[1]}
            for r in cur.fetchall()
        ]


@router.get("/macros")
def macros_list() -> list[dict[str, Any]]:
    """Toolbox Macro 라이브러리 — tbx_macro 실데이터 (S-2-2 좌측 목록)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT macro_name, macro_expr, status, COALESCE(hierarchy_address,''),
                      COALESCE(prompt_text,''), COALESCE(description_text,''),
                      COALESCE(code_text,''), flowchart_def, apply_type, version,
                      test_input, test_result
               FROM tbx_macro WHERE tenant_id=%s ORDER BY macro_id""", (tid,))
        return [
            {"name": r[0], "expr": r[1] or "", "status": r[2], "address": r[3],
             "prompt": r[4], "description": r[5], "codeText": r[6],
             "flowchartDef": r[7], "applyType": r[8], "version": r[9],
             "testInput": r[10], "testResult": r[11]}
            for r in cur.fetchall()
        ]


@router.get("/erp/events/{event_id}/flow")
def event_flow(event_id: int) -> dict[str, Any]:
    """이벤트 전후 공정 (E4) — erp_process_edge 기반 선행/후행 공정 실데이터."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT e.proc_def_id, d.proc_name, d.proc_code
               FROM erp_process_event e JOIN erp_process_def d ON d.proc_def_id=e.proc_def_id
               WHERE e.tenant_id=%s AND e.event_id=%s""", (tid, event_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"event not found: {event_id}")
        cur_def, cur_name, cur_code = row
        cur.execute(
            """SELECT d.proc_name FROM erp_process_edge x
               JOIN erp_process_def d ON d.proc_def_id=x.from_def_id
               WHERE x.tenant_id=%s AND x.to_def_id=%s ORDER BY d.proc_def_id""", (tid, cur_def))
        prev = [r[0] for r in cur.fetchall()]
        cur.execute(
            """SELECT d.proc_name FROM erp_process_edge x
               JOIN erp_process_def d ON d.proc_def_id=x.to_def_id
               WHERE x.tenant_id=%s AND x.from_def_id=%s ORDER BY d.proc_def_id""", (tid, cur_def))
        nxt = [r[0] for r in cur.fetchall()]
    return {"eventId": event_id, "current": cur_name, "currentCode": cur_code,
            "prev": prev, "next": nxt}


@router.get("/erp/process-defs")
def process_defs() -> dict[str, Any]:
    """공정 정의·흐름 — erp_process_def + erp_process_edge (M-14-7·Dashboard 흐름)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT proc_def_id, proc_code, proc_name, department, is_auto
               FROM erp_process_def WHERE tenant_id=%s ORDER BY proc_def_id""", (tid,))
        defs = [{"id": r[0], "code": r[1], "name": r[2], "dept": r[3], "auto": bool(r[4])}
                for r in cur.fetchall()]
        cur.execute(
            "SELECT from_def_id, to_def_id FROM erp_process_edge WHERE tenant_id=%s", (tid,))
        edges = [{"from": r[0], "to": r[1]} for r in cur.fetchall()]
    return {"defs": defs, "edges": edges}


@router.get("/codes/{code}/referencers")
def code_referencers(code: str) -> list[dict[str, Any]]:
    """Where-Used 역참조 — 이 코드를 child 로 갖는 mother (코드 상세 Referencers)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT mc.main_code, mc.code_name, r.quantity, r.approval_status
               FROM code_relationship r
               JOIN product_code cc ON cc.product_code_id=r.child_code_id
               JOIN product_code mc ON mc.product_code_id=r.mother_code_id
               WHERE r.tenant_id=%s AND cc.main_code=%s
               ORDER BY mc.main_code""", (tid, code))
        return [
            {"code": r[0], "name": r[1], "qty": float(r[2]), "status": r[3]}
            for r in cur.fetchall()
        ]


@router.get("/codes/{code}/where-used")
def code_where_used(code: str, maxLevel: int = 10) -> dict[str, Any]:
    """다단계 Where-Used 역전개 (트리아지 #34) — mother 체인 재귀 상승, 경로·레벨·순환 가드.

    referencers(1단)의 일반화 — 이 코드가 궁극적으로 어느 상위 제품까지 쓰이는지 전 경로 표시.
    """
    lv = max(1, min(int(maxLevel), 20))
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM product_code WHERE tenant_id=%s AND main_code=%s", (tid, code))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"코드 없음: {code}")
        cur.execute(
            """WITH RECURSIVE up AS (
                 SELECT r.mother_code_id, r.quantity, r.approval_status,
                        1 AS level, ARRAY[cc.main_code, mc.main_code]::text[] AS path
                 FROM code_relationship r
                 JOIN product_code cc ON cc.product_code_id=r.child_code_id
                 JOIN product_code mc ON mc.product_code_id=r.mother_code_id
                 WHERE r.tenant_id=%s AND cc.main_code=%s
                 UNION ALL
                 SELECT r.mother_code_id, r.quantity, r.approval_status,
                        up.level+1, up.path || mc.main_code::text
                 FROM up
                 JOIN code_relationship r ON r.tenant_id=%s AND r.child_code_id=up.mother_code_id
                 JOIN product_code mc ON mc.product_code_id=r.mother_code_id
                 WHERE up.level < %s AND NOT (mc.main_code = ANY(up.path))
               )
               SELECT mc.main_code, mc.code_name, up.quantity, up.approval_status,
                      up.level, array_to_string(up.path, ' › ')
               FROM up JOIN product_code mc ON mc.product_code_id=up.mother_code_id
               ORDER BY up.level, mc.main_code""",
            (tid, code, tid, lv))
        rows = [{"code": r[0], "name": r[1], "qty": float(r[2]), "status": r[3],
                 "level": int(r[4]), "path": r[5]} for r in cur.fetchall()]
    return {"code": code, "count": len(rows),
            "maxLevel": max((x["level"] for x in rows), default=0), "rows": rows}


def _code_children(cur, tid: int, code: str) -> dict[str, dict[str, Any]] | None:
    """코드의 BOM 구성(child main_code → {name, qty}). 코드 미존재 시 None."""
    cur.execute("SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
                (tid, code))
    if not cur.fetchone():
        return None
    cur.execute(
        """SELECT cc.main_code, cc.code_name, r.quantity
           FROM code_relationship r
           JOIN product_code mc ON mc.product_code_id=r.mother_code_id
           JOIN product_code cc ON cc.product_code_id=r.child_code_id
           WHERE r.tenant_id=%s AND mc.main_code=%s ORDER BY cc.main_code""", (tid, code))
    return {r[0]: {"name": r[1], "qty": float(r[2])} for r in cur.fetchall()}


@router.get("/codes/bom-compare")
def bom_compare(base: str = "", target: str = "") -> dict[str, Any]:
    """BOM 비교 (G3) — 두 코드 구성 diff: 추가/삭제/수량변경/동일."""
    if not base.strip() or not target.strip():
        raise HTTPException(422, detail="base·target 코드 필요")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        b = _code_children(cur, tid, base.strip())
        t = _code_children(cur, tid, target.strip())
    if b is None:
        raise HTTPException(404, detail=f"코드 없음: {base}")
    if t is None:
        raise HTTPException(404, detail=f"코드 없음: {target}")
    added, removed, changed, same = [], [], [], 0
    for code, tv in t.items():
        if code not in b:
            added.append({"code": code, "name": tv["name"], "qty": tv["qty"]})
        elif b[code]["qty"] != tv["qty"]:
            changed.append({"code": code, "name": tv["name"], "baseQty": b[code]["qty"], "targetQty": tv["qty"]})
        else:
            same += 1
    for code, bv in b.items():
        if code not in t:
            removed.append({"code": code, "name": bv["name"], "qty": bv["qty"]})
    return {"base": base, "target": target, "baseCount": len(b), "targetCount": len(t),
            "added": added, "removed": removed, "changed": changed, "unchanged": same,
            "identical": not added and not removed and not changed}


# ── B1 — 범용 승인 요청 (모든 화면의 승인 요청 버튼 실배선) + Macro 저장 ──

class ApprovalCreate(BaseModel):
    targetTable: str
    targetId: int = 0
    requestType: str = "UPDATE"
    label: str = ""
    targetCode: str = ""   # F4 — product code 문자열로 대상 지정 (code_relationship 등)


@router.post("/approvals", status_code=201, dependencies=[SETUP])
def create_approval(request: Request, body: ApprovalCreate) -> dict[str, Any]:
    """범용 승인 요청 — Design Editor·Macro Studio·Print Set-up·UI Designer 등."""
    tt = body.targetTable.strip()[:50] or "asset"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        actor_id = request.state.user_id
        target_id = body.targetId
        if body.targetCode.strip() and not target_id:
            # 자산 인지형 targetCode 해석 (v34.58) — 기존엔 product_code 로만 해석해
            # Macro Studio·UI Designer·Print Set-up 승인 요청이 404 로 침묵 고장 (11차 조사 발견)
            code = body.targetCode.strip()
            if tt == "tbx_macro":
                cur.execute("SELECT macro_id FROM tbx_macro WHERE tenant_id=%s AND macro_name=%s",
                            (tid, code))
                hit = cur.fetchone()
                if not hit:
                    raise HTTPException(404, detail=f"macro not found: {code}")
            elif tt == "tbx_ui_form":
                cur.execute("SELECT form_id FROM tbx_ui_form WHERE tenant_id=%s AND form_name=%s",
                            (tid, code))
                hit = cur.fetchone()
                if not hit:
                    raise HTTPException(404, detail=f"ui form not found: {code}")
            elif tt == "doc_control":
                cur.execute(
                    """SELECT doc_control_id FROM doc_control
                       WHERE tenant_id=%s AND (doc_no=%s OR title=%s)
                       ORDER BY doc_control_id DESC LIMIT 1""", (tid, code, code))
                hit = cur.fetchone()
                if not hit:
                    raise HTTPException(404, detail=f"document not found: {code}")
            else:
                # F4 — 코드 문자열 → product_code_id (Code Relationship·도면 승인 등)
                cur.execute(
                    "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
                    (tid, code))
                hit = cur.fetchone()
                if not hit:
                    raise HTTPException(404, detail=f"code not found: {code}")
            target_id = hit[0]
        # 권한승인정의서 승인상태기계 #7 — Macro 는 Test Run 통과(TESTED) 후에만 승인 요청 가능 (TBX-012)
        if tt == "tbx_macro" and target_id:
            cur.execute("SELECT test_result FROM tbx_macro WHERE tenant_id=%s AND macro_id=%s",
                        (tid, target_id))
            tr = cur.fetchone()
            tr_val = tr[0] if tr else None
            if tr_val is None or (isinstance(tr_val, dict) and tr_val.get("ok") is False):
                raise HTTPException(422, detail="Test Run 통과 후 승인 요청 가능 — Macro 는 TESTED 상태 필수 (TBX-012)")
        # uq_approval_pending — 동일 대상의 PENDING 이 이미 있으면 정직한 409
        cur.execute(
            """SELECT approval_id FROM sys_approval_request
               WHERE tenant_id=%s AND target_table=%s AND target_id=%s AND result IS NULL""",
            (tid, tt, target_id))
        dup = cur.fetchone()
        if dup:
            raise HTTPException(409, detail=f"이미 승인 대기 중 — {tt} (승인함 #{dup[0]} 처리 후 재요청)")
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,%s,%s,%s,'승인',%s,%s) RETURNING approval_id""",
            (tid, tt, target_id, body.requestType.strip()[:20] or "UPDATE",
             actor_id, body.label.strip()[:200]))
        approval_id = cur.fetchone()[0]
        # 승인권자(SETUP+) 알림 — 요청자 제외 (SVC-13)
        cur.execute(
            """SELECT user_id FROM sys_user
               WHERE tenant_id=%s AND user_level IN ('SETUP','ADMIN') AND user_id<>%s""",
            (tid, actor_id))
        for (uid,) in cur.fetchall():
            _notify(cur, tid, uid, "APPROVAL_REQUEST",
                    f"승인 요청 — {body.label.strip()[:80] or tt}", "/common")
    return {"approvalId": approval_id, "status": "PENDING"}


class MacroSave(BaseModel):
    prompt: str = ""
    expr: str = ""
    codeText: str = ""                       # B20 — Coding 패널
    flowchartDef: dict[str, Any] | None = None   # B20 — Flowchart 노드/엣지
    descriptionText: str = ""                # B20 — Description 패널
    applyType: str = ""                      # MACRO | CODING (빈 값 = 유지)
    testInput: dict[str, Any] | None = None  # B20 — 마지막 Test Run 영속
    testResult: dict[str, Any] | None = None


# 엔진 내장 함수 — 이 외의 호출형 토큰은 Table 참조로 간주 (tbx_macro_ref 추출)
ENGINE_BUILTINS = {"IF", "IFERROR", "AND", "OR", "NOT", "SUM", "MIN", "MAX", "VAR", "PREC",
                   # U27 — 공학 함수 Templet
                   "ABS", "SQRT", "ROUND", "POWER", "EXP", "LN", "LOG", "MOD",
                   "CEILING", "FLOOR", "PI", "SIN", "COS", "TAN", "RADIANS", "DEGREES", "INTERP"}


def _rebuild_macro_refs(cur, tid: int, macro_id: int, expr: str) -> int:
    """수식에서 Table 참조 추출 → tbx_macro_ref 재구성 (영향도 분석 원천)."""
    import re as _re
    cur.execute("DELETE FROM tbx_macro_ref WHERE macro_id=%s", (macro_id,))
    n = 0
    for token in set(_re.findall(r"([A-Za-z_][A-Za-z0-9_]*)\s*\(", expr or "")):
        if token.upper() in ENGINE_BUILTINS:
            continue
        cur.execute("SELECT table_id FROM tbl_data_table WHERE tenant_id=%s AND table_name=%s",
                    (tid, token))
        t = cur.fetchone()
        if t:
            cur.execute(
                """INSERT INTO tbx_macro_ref (macro_id, ref_type, ref_target_id)
                   VALUES (%s,'TABLE',%s)""", (macro_id, t[0]))
            n += 1
    return n


@router.put("/macros/{name}", dependencies=[SETUP])
def save_macro(name: str, body: MacroSave) -> dict[str, Any]:
    """Macro Studio 저장 — 4-Way Sync 전체 영속 (수식·코드·플로차트·설명 + Test 결과, B20).

    apply_type: MACRO=수식 실행 · CODING=코드 기반 (엔진 v1 미실행 — 등록·관리)."""
    at = body.applyType.strip().upper()
    if at and at not in ("MACRO", "CODING"):
        raise HTTPException(422, detail=f"apply_type 오류: {body.applyType} (MACRO|CODING)")
    if at != "CODING" and not body.expr.strip() and not at:
        raise HTTPException(422, detail="수식이 비어 있습니다")
    if at == "CODING" and not body.codeText.strip():
        raise HTTPException(422, detail="CODING 모드는 코드가 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE tbx_macro SET
               macro_expr=CASE WHEN %s<>'' THEN %s ELSE macro_expr END,
               prompt_text=CASE WHEN %s<>'' THEN %s ELSE prompt_text END,
               code_text=CASE WHEN %s<>'' THEN %s ELSE code_text END,
               description_text=CASE WHEN %s<>'' THEN %s ELSE description_text END,
               flowchart_def=COALESCE(%s, flowchart_def),
               test_input=COALESCE(%s, test_input),
               test_result=COALESCE(%s, test_result),
               apply_type=CASE WHEN %s<>'' THEN %s ELSE apply_type END,
               version=version+1, status='DRAFT', updated_at=now()
               WHERE tenant_id=%s AND macro_name=%s RETURNING macro_id, version""",
            (body.expr.strip(), body.expr.strip(),
             body.prompt.strip(), body.prompt.strip(),
             body.codeText.strip(), body.codeText.strip(),
             body.descriptionText.strip(), body.descriptionText.strip(),
             json.dumps(body.flowchartDef) if body.flowchartDef is not None else None,
             json.dumps(body.testInput) if body.testInput is not None else None,
             json.dumps(body.testResult) if body.testResult is not None else None,
             at, at, tid, name))
        row = cur.fetchone()
        if not row:
            cur.execute(
                """INSERT INTO tbx_macro (tenant_id, macro_name, macro_expr, prompt_text,
                   code_text, description_text, flowchart_def, test_input, test_result,
                   apply_type, status)
                   VALUES (%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),
                           %s,%s,%s,%s,'DRAFT') RETURNING macro_id, version""",
                (tid, name, body.expr.strip(), body.prompt.strip(), body.codeText.strip(),
                 body.descriptionText.strip(),
                 json.dumps(body.flowchartDef) if body.flowchartDef is not None else None,
                 json.dumps(body.testInput) if body.testInput is not None else None,
                 json.dumps(body.testResult) if body.testResult is not None else None,
                 at or "MACRO"))
            row = cur.fetchone()
        refs = _rebuild_macro_refs(cur, tid, row[0], body.expr)
    return {"macroId": row[0], "version": row[1], "status": "DRAFT", "refs": refs}


@router.get("/macros/functions")
def macro_functions(q: str = "") -> list[dict[str, Any]]:
    """함수 카탈로그·자연어 검색 (TBX-014 로컬) — name·설명·키워드 매칭."""
    catalog = [
        {"name": "IF", "sig": "IF(조건, 참값, 거짓값)", "desc": "조건 분기", "keywords": "조건 분기 만약 if"},
        {"name": "IFERROR", "sig": "IFERROR(식, 대체값)", "desc": "오류 시 대체값", "keywords": "오류 에러 대체 fallback"},
        {"name": "AND", "sig": "AND(a, b, …)", "desc": "논리곱", "keywords": "그리고 모두 논리"},
        {"name": "OR", "sig": "OR(a, b, …)", "desc": "논리합", "keywords": "또는 하나라도 논리"},
        {"name": "NOT", "sig": "NOT(a)", "desc": "부정", "keywords": "아니다 반전 부정"},
        {"name": "SUM", "sig": "Table(열, 시작:끝)", "desc": "범위 합계 (Table 범위 조회 기본)", "keywords": "합계 더하기 총합 범위 sum"},
        {"name": "MIN", "sig": "Table(열, 시작:끝, MIN)", "desc": "범위 최솟값", "keywords": "최소 작은 min"},
        {"name": "MAX", "sig": "Table(열, 시작:끝, MAX)", "desc": "범위 최댓값", "keywords": "최대 큰 max"},
        {"name": "VAR", "sig": "Var(이름, 기본값)", "desc": "Variant 변수 참조", "keywords": "변수 배리언트 참조 variant"},
        {"name": "PREC", "sig": "PreC(자릿수)", "desc": "정밀도(반올림 자릿수) 지정", "keywords": "반올림 정밀도 자릿수 소수"},
        {"name": "Table12", "sig": "Table12(열, 키[:키2])", "desc": "데이터 Table 단일/범위 조회", "keywords": "테이블 조회 데이터 표 참조 lookup"},
        # U27 — 공학 함수 Templet (s27 노트)
        {"name": "ABS", "sig": "ABS(x)", "desc": "절댓값", "keywords": "절대값 absolute"},
        {"name": "SQRT", "sig": "SQRT(x)", "desc": "제곱근", "keywords": "제곱근 루트 root"},
        {"name": "ROUND", "sig": "ROUND(x, 자릿수)", "desc": "반올림", "keywords": "반올림 자릿수"},
        {"name": "POWER", "sig": "POWER(x, y)", "desc": "거듭제곱 x^y", "keywords": "거듭제곱 승 지수"},
        {"name": "EXP", "sig": "EXP(x)", "desc": "지수 e^x", "keywords": "지수 exponential"},
        {"name": "LN", "sig": "LN(x)", "desc": "자연로그", "keywords": "자연로그 로그"},
        {"name": "LOG", "sig": "LOG(x[, 밑])", "desc": "로그 (기본 10)", "keywords": "로그 상용"},
        {"name": "MOD", "sig": "MOD(a, b)", "desc": "나머지", "keywords": "나머지 모듈로"},
        {"name": "CEILING", "sig": "CEILING(x[, 단위])", "desc": "단위 올림", "keywords": "올림 규격 단위"},
        {"name": "FLOOR", "sig": "FLOOR(x[, 단위])", "desc": "단위 내림", "keywords": "내림 절사"},
        {"name": "PI", "sig": "PI()", "desc": "원주율", "keywords": "원주율 파이"},
        {"name": "SIN", "sig": "SIN(rad)", "desc": "사인 (라디안)", "keywords": "사인 삼각"},
        {"name": "COS", "sig": "COS(rad)", "desc": "코사인 (라디안)", "keywords": "코사인 삼각"},
        {"name": "TAN", "sig": "TAN(rad)", "desc": "탄젠트 (라디안)", "keywords": "탄젠트 삼각"},
        {"name": "RADIANS", "sig": "RADIANS(도)", "desc": "도→라디안", "keywords": "각도 라디안 변환"},
        {"name": "DEGREES", "sig": "DEGREES(rad)", "desc": "라디안→도", "keywords": "각도 도 변환"},
        {"name": "INTERP", "sig": "INTERP(x, x1, y1, x2, y2)", "desc": "2점 선형 보간", "keywords": "보간 선형 interpolation 성능표"},
    ]
    needle = q.strip().lower()
    if not needle:
        return catalog
    return [f for f in catalog
            if needle in f["name"].lower() or needle in f["desc"].lower()
            or needle in f["keywords"].lower()]


@router.get("/macros/{name}/refs")
def macro_refs(name: str) -> list[dict[str, Any]]:
    """Macro 참조 — 이 매크로가 참조하는 Table (tbx_macro_ref)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.ref_type, t.table_name
               FROM tbx_macro_ref r
               JOIN tbx_macro m ON m.macro_id=r.macro_id AND m.tenant_id=%s
               LEFT JOIN tbl_data_table t ON t.table_id=r.ref_target_id AND r.ref_type='TABLE'
               WHERE m.macro_name=%s ORDER BY r.ref_id""", (tid, name))
        return [{"refType": r[0], "target": r[1] or "?"} for r in cur.fetchall()]


@router.get("/tables/{name}/impact")
def table_impact(name: str) -> list[dict[str, Any]]:
    """영향도 분석 — 이 Table 을 참조하는 Macro 목록 (tbx_macro_ref 역방향)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT m.macro_name, m.status, m.apply_type
               FROM tbx_macro_ref r
               JOIN tbl_data_table t ON t.table_id=r.ref_target_id AND r.ref_type='TABLE'
               JOIN tbx_macro m ON m.macro_id=r.macro_id
               WHERE t.tenant_id=%s AND t.table_name=%s ORDER BY m.macro_name""", (tid, name))
        return [{"macro": r[0], "status": r[1], "applyType": r[2]} for r in cur.fetchall()]


# ── B2 — 편집 영속화: 치수 저장 · Work Process · UI Form ──

class DimsSave(BaseModel):
    drawing: str = "KDCR 3-13"
    dims: list[dict[str, Any]] = []


@router.put("/drawings/dimensions", dependencies=[SETUP])
def save_dimensions(body: DimsSave) -> dict[str, Any]:
    """Design Editor 임시저장 F12 — VARIANT 는 variant_value, =식은 바인딩된 tbx_macro 갱신."""
    n_var = n_macro = 0
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for d in body.dims:
            label = str(d.get("no", "")).strip()[:10]
            value = str(d.get("value", "")).strip()
            if not label or not value:
                continue
            if value.startswith("="):
                cur.execute(
                    """UPDATE tbx_macro m SET macro_expr=%s, updated_at=now()
                       FROM dwg_dimension dd
                       JOIN dwg_drawing w ON w.drawing_id=dd.drawing_id
                       WHERE m.macro_id=dd.macro_id AND dd.tenant_id=%s
                         AND w.drawing_no=%s AND dd.dim_label=%s""",
                    (value, tid, body.drawing, label))
                n_macro += cur.rowcount
            else:
                try:
                    num = float(value)
                except ValueError:
                    continue
                # MACRO 바인딩 치수의 평가값은 파생값 — 덮어쓰지 않음 (ck_dim_binding)
                cur.execute(
                    """UPDATE dwg_dimension dd SET variant_value=%s, updated_at=now()
                       FROM dwg_drawing w
                       WHERE w.drawing_id=dd.drawing_id AND dd.tenant_id=%s
                         AND w.drawing_no=%s AND dd.dim_label=%s AND dd.macro_id IS NULL""",
                    (num, tid, body.drawing, label))
                n_var += cur.rowcount
    return {"variantSaved": n_var, "macroSaved": n_macro}


class WorkProcessSave(BaseModel):
    code: str = "KDCR 3-13"
    items: list[dict[str, Any]] = []   # {item, makeOrBuy}


@router.get("/erp/work-process")
def get_work_process(code: str = "KDCR 3-13") -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT wp.process_type, wp.make_or_buy
               FROM erp_work_process wp
               JOIN product_code pc ON pc.product_code_id=wp.product_code_id
               WHERE wp.tenant_id=%s AND pc.main_code=%s ORDER BY wp.seq_no""",
            (tid, code))
        return [{"item": r[0], "makeOrBuy": r[1]} for r in cur.fetchall()]


@router.put("/erp/work-process", dependencies=[SETUP])
def save_work_process(body: WorkProcessSave) -> dict[str, Any]:
    """Work Process MAKE/BUY 저장 — erp_work_process upsert (S-4-1-2 F12)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
            (tid, body.code))
        pc = cur.fetchone()
        if not pc:
            raise HTTPException(404, detail=f"코드 없음: {body.code}")
        def _num(v):
            try:
                return float(v) if v is not None and str(v).strip() != "" else None
            except (TypeError, ValueError):
                return None

        n = 0
        for i, it in enumerate(body.items):
            item = str(it.get("item", "")).strip()[:20]
            mob = str(it.get("makeOrBuy", "")).strip().upper()
            if not item or mob not in ("MAKE", "BUY"):
                continue
            # U3 공정 파라미터 (슬라이드 45) — 미전달 필드는 None(기존 값 유지 아님: 전체 행 저장 방식)
            workshop = str(it.get("workshop", "") or "").strip()[:50] or None
            warehouse = str(it.get("warehouse", "") or "").strip()[:50] or None
            if warehouse == "-":
                warehouse = None
            person = it.get("person")
            person = int(person) if isinstance(person, (int, float)) and person >= 0 else None
            skill = str(it.get("skill", "") or "").strip()[:20] or None
            work_time = _num(it.get("timeMin"))
            min_stock = _num(it.get("minStock"))
            remarks = str(it.get("remarks", "") or "").strip()[:300] or None
            cur.execute(
                """UPDATE erp_work_process SET make_or_buy=%s, seq_no=%s, workshop=%s,
                       warehouse=%s, min_stock=%s, person_count=%s, skill_grade=%s,
                       work_time=%s, remarks=%s, updated_at=now()
                   WHERE tenant_id=%s AND product_code_id=%s AND process_type=%s""",
                (mob, i, workshop, warehouse, min_stock, person, skill, work_time, remarks,
                 tid, pc[0], item))
            if cur.rowcount == 0:
                cur.execute(
                    """INSERT INTO erp_work_process (tenant_id, product_code_id, process_type,
                       seq_no, make_or_buy, workshop, warehouse, min_stock, person_count,
                       skill_grade, work_time, remarks)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (tid, pc[0], item, i, mob, workshop, warehouse, min_stock, person,
                     skill, work_time, remarks))
            n += 1
    return {"saved": n}


@router.get("/erp/work-process/materials")
def work_process_materials(code: str = "KDCR 3-13") -> list[dict[str, Any]]:
    """Work Process 자재행 (G3-c) — 도면 BOM 부품 + erp_work_process(make/buy·창고·시간) 실조인.

    자재행 = 코드(=도면) BOM 의 실 부품(부품명·공급처). make/buy 는 erp_work_process 저장값,
    없으면 is_standard 기본(표준=BUY·비표준=MAKE). 창고·최소재고·시간은 저장 시 채워짐.
    """
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, code)
        cur.execute(
            "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s", (tid, code))
        pcr = cur.fetchone()
        pcid = pcr[0] if pcr else None
        cur.execute(
            """SELECT p.part_name, COALESCE(c.company_name,''), p.is_standard,
                      COALESCE(p.specification,''),
                      wp.make_or_buy, wp.warehouse, wp.min_stock, wp.work_time,
                      wp.workshop, wp.person_count, wp.skill_grade, wp.remarks
               FROM dwg_bom b JOIN prt_part p ON p.part_id=b.part_id
               LEFT JOIN com_company c ON c.company_id=p.supplier_id
               LEFT JOIN erp_work_process wp
                 ON wp.tenant_id=%s AND wp.product_code_id=%s AND wp.process_type=p.part_name
               WHERE b.drawing_id=%s ORDER BY b.item_no""",
            (tid, pcid, did))
        out = []
        for r in cur.fetchall():
            mb = r[4] or ("BUY" if r[2] else "MAKE")
            out.append({
                "item": r[0], "warehouse": r[5] or "-",
                "minStock": float(r[6]) if r[6] is not None else 0,
                "supplier": r[1] or "-", "makeBuy": mb,
                "timeMin": (float(r[7]) if r[7] is not None else 45) if mb == "MAKE" else None,
                # U3 공정 파라미터 (슬라이드 45 — Work place·Person·Skill)
                "workshop": r[8] or "", "person": r[9], "skill": r[10] or "",
                "remarks": (r[11] or r[3]) or "",
            })
        return out


class FormSave(BaseModel):
    layout: list[dict[str, Any]] = []
    formType: str = "SCREEN"


# ── U16 — Combo Data 바인딩 옵션 (슬라이드 27, 화이트리스트 테이블 열) ──
_BIND_COLUMNS: dict[str, list[str]] = {
    "prt_part": ["part_no", "part_name", "unit"],
    "cst_quotation": ["quotation_no", "status", "currency"],
}


@router.get("/toolbox/bind-options")
def bind_options(table: str, column: str) -> dict[str, Any]:
    """UI Designer Combo Data set-up — 화이트리스트 테이블 열의 distinct 값."""
    cols = _BIND_COLUMNS.get(table)
    if not cols or column not in cols:
        raise HTTPException(422, detail=f"바인딩 불가 테이블/열: {table}.{column}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT DISTINCT {column} FROM {table}
                WHERE tenant_id=%s AND {column} IS NOT NULL ORDER BY 1 LIMIT 30""", (tid,))
        values = [r[0] for r in cur.fetchall()]
    return {"table": table, "column": column, "values": values, "tables": _BIND_COLUMNS}


@router.get("/toolbox/forms/{name}")
def get_form(name: str) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT form_id, version, layout_def FROM tbx_ui_form
               WHERE tenant_id=%s AND form_name=%s""", (tid, name))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"Form 없음: {name}")
    return {"formId": row[0], "version": row[1], "layout": row[2]}


@router.put("/toolbox/forms/{name}", dependencies=[SETUP])
def save_form(name: str, body: FormSave) -> dict[str, Any]:
    """UI Designer 레이아웃 저장 — tbx_ui_form upsert (version+1)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE tbx_ui_form SET layout_def=%s, version=version+1, updated_at=now()
               WHERE tenant_id=%s AND form_name=%s RETURNING form_id, version""",
            (json.dumps(body.layout), tid, name))
        row = cur.fetchone()
        if not row:
            cur.execute(
                """INSERT INTO tbx_ui_form (tenant_id, form_name, form_type, layout_def)
                   VALUES (%s,%s,%s,%s) RETURNING form_id, version""",
                (tid, name, body.formType.strip()[:30] or "SCREEN", json.dumps(body.layout)))
            row = cur.fetchone()
    return {"formId": row[0], "version": row[1]}


# ── SVC-03 Sub Code 항목 등록 (S-1-1 write) ──
class NewItemRequest(BaseModel):
    slot: str
    name: str
    values: list[str] = []


@router.post("/codes/groups/{group}/items", status_code=201, dependencies=[SETUP])
def add_item(group: str, request: Request, body: NewItemRequest) -> dict[str, Any]:
    slot = body.slot.strip().upper()
    if not slot or not body.name.strip():
        raise HTTPException(422, detail="필수(노란 셀) 미입력")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT group_id FROM code_group WHERE tenant_id=%s AND group_code=%s", (tid, group))
        g = cur.fetchone()
        if not g:
            raise HTTPException(404, detail=f"group not found: {group}")
        cur.execute(
            "SELECT 1 FROM code_item WHERE group_id=%s AND item_slot=%s", (g[0], slot))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — Item {slot} 이미 등록됨 (CODE-006)")
        cur.execute(
            """INSERT INTO code_item (tenant_id, group_id, item_slot, item_name)
               VALUES (%s,%s,%s,%s) RETURNING item_id""", (tid, g[0], slot, body.name.strip()))
        item_id = cur.fetchone()[0]
        for i, v in enumerate([x for x in body.values if x.strip()]):
            cur.execute(
                """INSERT INTO code_item_value (tenant_id, item_id, value_code, sort_order,
                   approval_status) VALUES (%s,%s,%s,%s,'PENDING')""",
                (tid, item_id, v.strip()[:30], i))
        actor_id = request.state.user_id
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'code_item',%s,'CREATE','승인',%s,%s)""",
            (tid, item_id, actor_id, f"{group} / {slot}: {body.name.strip()}"))
        # 승인권자(SETUP+) 알림 — 요청자 제외 (SVC-13)
        cur.execute(
            """SELECT user_id FROM sys_user
               WHERE tenant_id=%s AND user_level IN ('SETUP','ADMIN') AND user_id<>%s""",
            (tid, actor_id))
        for (uid,) in cur.fetchall():
            _notify(cur, tid, uid, "APPROVAL_REQUEST",
                    f"승인 요청 — {group}/{slot} {body.name.strip()}", "/common")
    return {"slot": slot, "status": "PENDING"}


# ── SVC-03 Code Relationship (S-1-4) ──
@router.get("/codes/relationships/{mother}/children")
def relationship_children(mother: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT c.main_code, c.code_name, r.quantity, COALESCE(r.remarks,''),
                      COALESCE(string_agg(sm.child_slot || '←' ||
                        COALESCE(sm.mother_slot, '"'||sm.fixed_value||'"'), ' · '
                        ORDER BY sm.child_slot), '-')
               FROM code_relationship r
               JOIN product_code m ON m.product_code_id=r.mother_code_id
               JOIN product_code c ON c.product_code_id=r.child_code_id
               LEFT JOIN code_relationship_slot_map sm ON sm.rel_id=r.rel_id
               WHERE r.tenant_id=%s AND m.main_code=%s AND r.approval_status='APPROVED'
               GROUP BY c.main_code, c.code_name, r.quantity, r.remarks, r.sort_order
               ORDER BY r.sort_order""", (tid, mother))
        return [
            {"code": r[0], "desc": r[1], "qty": float(r[2]), "remarks": r[3], "slotMap": r[4]}
            for r in cur.fetchall()
        ]


class RunningTestRequest(BaseModel):
    motherCode: str = "KDCR 3-13"
    slotValues: dict[str, str]


@router.post("/codes/relationships/running-test")
def running_test(body: RunningTestRequest) -> dict[str, Any]:
    """CODE-009 — Mother slot 조합으로 전량 전개 검증 (expand 재사용)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        rows = _expand_rows(cur, tid, body.motherCode, body.slotValues)
        cur.execute(
            "SELECT code_name FROM product_code WHERE tenant_id=%s AND main_code=%s",
            (tid, body.motherCode))
        mother = cur.fetchone()
    if not rows:
        raise HTTPException(404, detail=f"mother not found: {body.motherCode}")
    sv = body.slotValues
    out = [{
        "no": "Main",
        "name": f"{body.motherCode}-{sv.get('B', '?')}-{sv.get('C', '?')}-{sv.get('E', '?')}",
        "desc": mother[0] if mother else body.motherCode, "qty": 1, "remarks": "",
        "mainCode": body.motherCode,
    }]
    for i, r in enumerate(rows):
        out.append({
            "no": str(i + 1), "name": _resolved(r[0], r[4] or {}),
            "desc": r[1], "qty": float(r[2]), "remarks": "", "mainCode": r[0],
            "level": r[3], "path": r[5],
        })
    return {"passed": True, "cycleCheck": "OK", "rows": out}


# ── SVC-09 Dashboard 집계 (ERP-014) ──
@router.get("/erp/dashboard")
def dashboard() -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT count(*) FROM prj_project WHERE tenant_id=%s AND status='IN_PROGRESS'", (tid,))
        projects = cur.fetchone()[0]
        cur.execute(
            "SELECT count(*) FROM sys_approval_request WHERE tenant_id=%s AND result IS NULL", (tid,))
        pending = cur.fetchone()[0]
        cur.execute(
            """SELECT count(*) FROM erp_process_event
               WHERE tenant_id=%s AND status<>'DONE' AND due_date < CURRENT_DATE""", (tid,))
        alerts = cur.fetchone()[0]
        cur.execute(
            """SELECT d.department,
                      count(*) FILTER (WHERE e.status='TODO' AND NOT
                        (e.due_date < CURRENT_DATE)),
                      count(*) FILTER (WHERE e.status='IN_PROGRESS'),
                      count(*) FILTER (WHERE e.status='DONE'
                        AND e.done_at > now() - interval '7 days'),
                      count(*) FILTER (WHERE e.status<>'DONE' AND e.due_date < CURRENT_DATE)
               FROM erp_process_event e
               JOIN erp_process_def d ON d.proc_def_id=e.proc_def_id
               WHERE e.tenant_id=%s GROUP BY d.department ORDER BY d.department""", (tid,))
        depts = [
            {"dept": r[0], "waiting": r[1], "running": r[2], "doneWeek": r[3], "delayed": r[4]}
            for r in cur.fetchall()
        ]
    return {
        "kpis": [
            {"label": "진행 Project", "value": str(projects)},
            {"label": "승인 대기", "value": str(pending)},
            {"label": "이번 달 수주", "value": "₩ 8.4억"},   # 견적/수주 모듈 구축 전 고정값
            {"label": "이상 경고 (시간·자금)", "value": str(alerts), "err": alerts > 0},
        ],
        "deptEvents": depts,
    }


@router.get("/erp/analytics")
def analytics() -> dict[str, Any]:
    """C3 분석 — Run 통계(cpq_run) + 원가 추이(cst_calc) 누적 집계."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        # E3 — 테스트성 Run 은 업무 통계에서 제외 (is_test)
        cur.execute(
            """SELECT count(*), count(*) FILTER (WHERE status='SUCCESS'),
                      count(*) FILTER (WHERE status='FAILED'),
                      avg(EXTRACT(EPOCH FROM (finished_at-started_at))) FILTER (WHERE finished_at IS NOT NULL)
               FROM cpq_run WHERE tenant_id=%s AND NOT is_test""", (tid,))
        total, success, failed, avg_sec = cur.fetchone()
        cur.execute(
            """SELECT run_id, status, run_type,
                      EXTRACT(EPOCH FROM (finished_at-started_at)), to_char(started_at,'MM-DD HH24:MI')
               FROM cpq_run WHERE tenant_id=%s AND NOT is_test ORDER BY run_id DESC LIMIT 10""", (tid,))
        recent = [{"runId": x[0], "status": x[1], "runType": x[2],
                   "durationSec": round(float(x[3]), 1) if x[3] is not None else None, "at": x[4]}
                  for x in cur.fetchall()]
        cur.execute(
            """SELECT calc_type, COALESCE(sum(total_amount),0), count(DISTINCT run_id)
               FROM cst_calc WHERE tenant_id=%s GROUP BY calc_type""", (tid,))
        cost = {x[0]: {"total": float(x[1]), "runs": x[2]} for x in cur.fetchall()}
        # 최근 원가 있는 Run 별 3분류 (추이 — 최근 8건)
        cur.execute(
            """SELECT run_id,
                      sum(total_amount) FILTER (WHERE calc_type='MATERIAL'),
                      sum(total_amount) FILTER (WHERE calc_type='MANUFACTURING'),
                      sum(total_amount) FILTER (WHERE calc_type='DIRECT')
               FROM cst_calc WHERE tenant_id=%s GROUP BY run_id ORDER BY run_id DESC LIMIT 8""", (tid,))
        trend = [{"runId": x[0], "material": float(x[1] or 0), "manufacturing": float(x[2] or 0),
                  "direct": float(x[3] or 0)} for x in cur.fetchall()][::-1]
        # C3 — 견적(추정) vs 실적 차이 위젯 (추정=최근 Run cst_calc, 실적=cst_actual)
        try:
            _, est = _latest_cost_base(cur, tid)
        except HTTPException:
            est = {}
        cur.execute("SELECT category, COALESCE(sum(amount),0) FROM cst_actual WHERE tenant_id=%s GROUP BY category",
                    (tid,))
        act = {x[0]: float(x[1]) for x in cur.fetchall()}
        vcats = []
        for c in COST_CATEGORIES:
            e = float(est.get(c, 0)); a = act.get(c, 0.0)
            rate = ((a - e) / e) if e else (1.0 if a else 0.0)
            vcats.append({"category": c, "label": COST_CAT_LABEL[c], "estimate": e, "actual": a,
                          "variance": round(a - e, 2), "varianceRate": round(rate, 4),
                          "alert": rate > VARIANCE_ALERT_RATE})
        te = sum(x["estimate"] for x in vcats); ta = sum(x["actual"] for x in vcats)
        trate = ((ta - te) / te) if te else (1.0 if ta else 0.0)
        variance = {"categories": vcats, "totalEstimate": te, "totalActual": ta,
                    "totalVariance": round(ta - te, 2), "totalVarianceRate": round(trate, 4),
                    "alert": trate > VARIANCE_ALERT_RATE, "hasActual": ta > 0}
        # C3 잔여 — 월별 매출/기여마진 추이 (D1 수주: ORDERED 견적 계약액 + PCR 기여마진)
        cur.execute(
            """SELECT to_char(q.order_date,'YYYY-MM'),
                      COALESCE(sum(COALESCE(q.contract_amount, q.total_amount)),0),
                      sum(p.contribution_margin), count(*)
               FROM cst_quotation q JOIN cst_pcr p ON p.pcr_id=q.pcr_id
               WHERE q.tenant_id=%s AND q.status='ORDERED' AND q.order_date IS NOT NULL
               GROUP BY 1 ORDER BY 1 DESC LIMIT 12""", (tid,))
        monthly = []
        for ym, rev, margin, cnt in cur.fetchall()[::-1]:
            rev = float(rev)
            m = float(margin) if margin is not None else None
            monthly.append({"month": ym, "revenue": rev, "margin": m,
                            "marginRate": round(m / rev, 4) if (m is not None and rev) else None,
                            "orders": cnt})
    return {
        "runStats": {"total": total, "success": success, "failed": failed,
                     "successRate": round(success / total * 100, 1) if total else 0,
                     "avgDurationSec": round(float(avg_sec), 1) if avg_sec else 0},
        "recentRuns": recent, "costByType": cost, "costTrend": trend, "variance": variance,
        "monthlyOrders": monthly,
    }


# ── SVC-09 발주 품목 (M-8-2) — 단가 resolve 실연동 ──
PR_META = {
    "FDV-480": {"supplierCode": "HS-M480", "qty": 2, "requiredDate": "08-20", "delivery": "EXW"},
    "KDC-1": {"supplierCode": "JW-C001", "qty": 4, "requiredDate": "08-15", "delivery": "지정장소"},
    "EWT-3": {"supplierCode": "-", "qty": 1, "requiredDate": "-", "delivery": "-", "stockOk": True},
}


@router.get("/erp/pr-items")
def pr_items() -> list[dict[str, Any]]:
    ref = date.today().isoformat()
    out: list[dict[str, Any]] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for code, meta in PR_META.items():
            cur.execute(
                """SELECT pc.code_name, p.price, cc.company_name
                   FROM product_code pc
                   LEFT JOIN cst_price p ON p.product_code_id=pc.product_code_id
                     AND p.valid_from <= %s::date
                     AND (p.valid_to IS NULL OR p.valid_to >= %s::date)
                   LEFT JOIN com_company cc ON cc.company_id=p.supplier_id
                   WHERE pc.tenant_id=%s AND pc.main_code=%s
                   ORDER BY array_position(%s::text[], p.price_source), p.valid_from DESC
                   LIMIT 1""", (ref, ref, tid, code, SOURCE_PRIORITY))
            r = cur.fetchone()
            if not r:
                continue
            # D2 — Stock Check 실재고 기반 · ATP(가용재고 = 보유 − ACTIVE 예약) ≥ 소요수량
            cur.execute(
                "SELECT COALESCE(sum(quantity),0) FROM inv_stock WHERE tenant_id=%s AND item_code=%s",
                (tid, code))
            on_hand = float(cur.fetchone()[0])
            cur.execute(
                "SELECT COALESCE(sum(quantity),0) FROM inv_reservation "
                "WHERE tenant_id=%s AND item_code=%s AND status='ACTIVE'", (tid, code))
            reserved = float(cur.fetchone()[0])
            available = on_hand - reserved
            stock_ok = available >= meta["qty"]
            out.append({
                "code": code, "name": r[0],
                "supplierCode": meta["supplierCode"],
                "supplier": "-" if stock_ok else (r[2] or "-"),
                "qty": meta["qty"], "onHand": on_hand, "reserved": reserved, "available": available,
                "price": None if stock_ok else (float(r[1]) if r[1] is not None else None),
                "requiredDate": meta["requiredDate"], "delivery": meta["delivery"],
                "stockOk": stock_ok, "checked": not stock_ok,
            })
    return out


def _stock_price(cur, tid: int, item_code: str) -> float:
    """cst_price(STOCK, 유효기간 내) 자동 조회 — item_code → part_no 또는 product main_code 매칭.

    미등록 시 0.0 (단가 없음)."""
    cur.execute(
        """SELECT cp.price FROM cst_price cp
           JOIN prt_part p ON p.part_id=cp.part_id
           WHERE cp.tenant_id=%s AND p.part_no=%s AND cp.price_source='STOCK'
             AND cp.valid_from<=CURRENT_DATE AND (cp.valid_to IS NULL OR cp.valid_to>=CURRENT_DATE)
           ORDER BY cp.valid_from DESC LIMIT 1""", (tid, item_code))
    r = cur.fetchone()
    if r:
        return float(r[0])
    cur.execute(
        """SELECT cp.price FROM cst_price cp
           JOIN product_code pc ON pc.product_code_id=cp.product_code_id
           WHERE cp.tenant_id=%s AND pc.main_code=%s AND cp.price_source='STOCK'
             AND cp.valid_from<=CURRENT_DATE AND (cp.valid_to IS NULL OR cp.valid_to>=CURRENT_DATE)
           ORDER BY cp.valid_from DESC LIMIT 1""", (tid, item_code))
    r = cur.fetchone()
    return float(r[0]) if r else 0.0


class InboundRequest(BaseModel):
    itemCode: str
    itemName: str = ""
    locationCode: str
    quantity: float
    refNo: str = ""
    lotNo: str = ""        # 로트 번호 (배치 추적)
    serialNo: str = ""     # 시리얼 번호 (직번 단위 추적)
    unitPrice: float | None = None   # 입고 단가 (미지정 시 cst_price STOCK 자동 적재)
    expiryDate: str = ""   # 유통기한 YYYY-MM-DD (U5 — 로트 만료 경고 근거)


@router.post("/erp/stock/inbound", status_code=201, dependencies=[SETUP])
def stock_inbound(request: Request, body: InboundRequest) -> dict[str, Any]:
    """입고 처리 (D2) — PO 품목 입고 → inv_movement(IN) + inv_stock upsert. 발주→재고 고리(MI).

    로트/시리얼 지정 시 이력에 추적 차원 적재. 단가 미지정 시 cst_price(STOCK) 자동 적재 →
    이동평균 단가·평가액 산출."""
    item = body.itemCode.strip()[:50]
    loc = body.locationCode.strip()[:30]
    if not item or not loc or body.quantity <= 0:
        raise HTTPException(422, detail="필수 — 품목·위치·수량(>0)")
    if body.unitPrice is not None and body.unitPrice < 0:
        raise HTTPException(422, detail="단가는 0 이상이어야 합니다")
    lot = body.lotNo.strip()[:50] or None
    serial = body.serialNo.strip()[:80] or None
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        # 단가 결정 — 명시 지정 우선, 없으면 cst_price(STOCK) 자동 적재
        auto = body.unitPrice is None
        in_price = _stock_price(cur, tid, item) if auto else float(body.unitPrice)
        cur.execute(
            """INSERT INTO inv_movement (tenant_id, item_code, location_code, movement_type,
               quantity, ref_type, ref_no, lot_no, serial_no, expiry_date, created_by)
               VALUES (%s,%s,%s,'IN',%s,'PO',%s,%s,%s,NULLIF(%s,'')::date,%s)""",
            (tid, item, loc, body.quantity, body.refNo.strip()[:50] or None,
             lot, serial, body.expiryDate.strip()[:10], str(request.state.user_id)))
        # 이동평균 — 기존 (수량·단가) + 입고분 가중평균
        cur.execute(
            "SELECT quantity, unit_price FROM inv_stock WHERE tenant_id=%s AND item_code=%s AND location_code=%s",
            (tid, item, loc))
        prev = cur.fetchone()
        if prev:
            q0, p0 = float(prev[0]), float(prev[1])
            new_q = q0 + body.quantity
            new_price = round((q0 * p0 + body.quantity * in_price) / new_q, 2) if new_q else in_price
        else:
            new_price = round(in_price, 2)
        cur.execute(
            """INSERT INTO inv_stock (tenant_id, item_code, item_name, location_code, quantity, unit_price)
               VALUES (%s,%s,%s,%s,%s,%s)
               ON CONFLICT (tenant_id, item_code, location_code)
               DO UPDATE SET quantity = inv_stock.quantity + EXCLUDED.quantity,
                             item_name = COALESCE(EXCLUDED.item_name, inv_stock.item_name),
                             unit_price = %s, updated_at = now()
               RETURNING quantity""",
            (tid, item, body.itemName.strip()[:200] or None, loc, body.quantity, new_price, new_price))
        on_hand = float(cur.fetchone()[0])
        _audit(cur, tid, "inv_stock", 0, "INBOUND", request.state.user_id,
               {"item": item, "location": loc, "qty": body.quantity, "lot": lot, "serial": serial,
                "unitPrice": in_price, "priceAuto": auto})
    return {"itemCode": item, "locationCode": loc, "onHand": on_hand,
            "lotNo": lot, "serialNo": serial, "unitPrice": in_price,
            "avgPrice": new_price, "priceAuto": auto, "value": round(on_hand * new_price, 2)}


@router.get("/erp/stock")
def stock_list() -> list[dict[str, Any]]:
    """재고 조회 (D2) — 품목×창고 위치 현재 수량·단가(이동평균)·평가액."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT s.item_code, s.item_name, s.location_code, w.location_name,
                      s.quantity, s.unit, to_char(s.updated_at,'YYYY-MM-DD HH24:MI'), s.unit_price
               FROM inv_stock s
               LEFT JOIN erp_warehouse w ON w.tenant_id=s.tenant_id AND w.location_code=s.location_code
               WHERE s.tenant_id=%s ORDER BY s.item_code, s.location_code""", (tid,))
        return [{"itemCode": r[0], "itemName": r[1] or "-", "locationCode": r[2],
                 "locationName": r[3] or r[2], "quantity": float(r[4]), "unit": r[5],
                 "updatedAt": r[6], "unitPrice": float(r[7]),
                 "value": round(float(r[4]) * float(r[7]), 2)} for r in cur.fetchall()]


@router.get("/erp/stock/movements")
def stock_movements(item: str = "", limit: int = 30) -> list[dict[str, Any]]:
    """입출고 이력 (D2)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if item.strip():
            clause = " AND item_code=%s"
            params.append(item.strip())
        params.append(limit)
        cur.execute(
            f"""SELECT item_code, location_code, movement_type, quantity, ref_type, ref_no,
                       to_char(moved_at,'MM-DD HH24:MI'), lot_no, serial_no
                FROM inv_movement WHERE tenant_id=%s{clause}
                ORDER BY movement_id DESC LIMIT %s""", tuple(params))
        return [{"itemCode": r[0], "locationCode": r[1], "type": r[2], "quantity": float(r[3]),
                 "refType": r[4], "refNo": r[5], "at": r[6],
                 "lotNo": r[7] or "", "serialNo": r[8] or ""} for r in cur.fetchall()]


@router.get("/erp/stock/lots")
def stock_lots(item: str = "") -> list[dict[str, Any]]:
    """로트/시리얼 잔량 (D2) — 이력에서 산출: 로트·시리얼별 (IN−OUT) 잔량·위치.

    규제·직번 부품 추적성: inv_stock(품목×위치 집계)엔 없는 로트 차원."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if item.strip():
            clause = " AND item_code=%s"
            params.append(item.strip())
        cur.execute(
            f"""SELECT item_code, lot_no, serial_no,
                       max(location_code) AS location_code,
                       sum(CASE WHEN movement_type='IN' THEN quantity ELSE -quantity END) AS balance,
                       to_char(max(moved_at),'YYYY-MM-DD HH24:MI') AS last_at,
                       to_char(min(expiry_date),'YYYY-MM-DD') AS expiry,
                       CASE WHEN min(expiry_date) IS NULL THEN ''
                            WHEN min(expiry_date) < CURRENT_DATE THEN 'EXPIRED'
                            WHEN min(expiry_date) <= CURRENT_DATE + 30 THEN 'EXPIRING'
                            ELSE 'OK' END AS expiry_status
                FROM inv_movement
                WHERE tenant_id=%s AND (lot_no IS NOT NULL OR serial_no IS NOT NULL){clause}
                GROUP BY item_code, lot_no, serial_no
                ORDER BY item_code, lot_no NULLS LAST, serial_no NULLS LAST""", tuple(params))
        return [{"itemCode": r[0], "lotNo": r[1] or "", "serialNo": r[2] or "",
                 "locationCode": r[3], "balance": float(r[4]), "lastAt": r[5],
                 "expiry": r[6] or "", "expiryStatus": r[7]}
                for r in cur.fetchall()]


class LotExpirySet(BaseModel):
    itemCode: str
    lotNo: str
    expiryDate: str   # YYYY-MM-DD, '' = 해제


@router.patch("/erp/stock/lots/expiry", dependencies=[SETUP])
def set_lot_expiry(request: Request, body: LotExpirySet) -> dict[str, Any]:
    """로트 유통기한 설정/해제 (U5) — 해당 품목×로트 전 이동행에 반영."""
    item = body.itemCode.strip()[:50]
    lot = body.lotNo.strip()[:50]
    if not item or not lot:
        raise HTTPException(422, detail="itemCode·lotNo 필요")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE inv_movement SET expiry_date=NULLIF(%s,'')::date
               WHERE tenant_id=%s AND item_code=%s AND lot_no=%s""",
            (body.expiryDate.strip()[:10], tid, item, lot))
        if cur.rowcount == 0:
            raise HTTPException(404, detail=f"로트 없음: {item} / {lot}")
        _audit(cur, tid, "inv_movement", 0, "LOT_EXPIRY_SET", request.state.user_id,
               after={"item": item, "lot": lot, "expiry": body.expiryDate.strip()[:10] or None})
    return {"updated": True}


# ── U5 창고 정기점검 실적 (슬라이드 46 — 보관 품질 유지) ──

class InspectionCreate(BaseModel):
    result: str = "OK"     # OK | ISSUE
    note: str = ""


@router.get("/erp/warehouses/{code}/inspections")
def wh_inspections(code: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT inspection_id, result, COALESCE(note,''), inspected_by,
                      to_char(inspected_at,'YYYY-MM-DD HH24:MI')
               FROM erp_wh_inspection WHERE tenant_id=%s AND location_code=%s
               ORDER BY inspected_at DESC LIMIT 50""", (tid, code.strip()[:50]))
        return [{"id": r[0], "result": r[1], "note": r[2], "by": r[3], "at": r[4]}
                for r in cur.fetchall()]


@router.post("/erp/warehouses/{code}/inspections", status_code=201, dependencies=[SETUP])
def wh_inspection_add(code: str, request: Request, body: InspectionCreate) -> dict[str, Any]:
    """창고 위치 점검 실적 등록 (U5) — 검사주기 필드(B19)의 실적 기록."""
    result = body.result.strip().upper()
    if result not in ("OK", "ISSUE"):
        raise HTTPException(422, detail=f"result 오류: {body.result} (OK|ISSUE)")
    loc = code.strip()[:50]
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM erp_warehouse WHERE tenant_id=%s AND location_code=%s", (tid, loc))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"위치 없음: {loc}")
        cur.execute(
            """INSERT INTO erp_wh_inspection (tenant_id, location_code, result, note, inspected_by)
               VALUES (%s,%s,%s,NULLIF(%s,''),%s) RETURNING inspection_id""",
            (tid, loc, result, body.note.strip()[:300], request.state.login))
        iid = cur.fetchone()[0]
        _audit(cur, tid, "erp_wh_inspection", iid, "WH_INSPECT", request.state.user_id,
               after={"location": loc, "result": result})
    return {"id": iid}


# ── U5 대체 자재 (슬라이드 46 — 대체 자재 연구·적용) ──

class SubstituteAdd(BaseModel):
    substituteNo: str
    note: str = ""


@router.get("/parts/{part_no}/substitutes")
def part_substitutes(part_no: str) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT s.sub_id, p2.part_no, p2.part_name, COALESCE(s.note,''),
                      to_char(s.created_at,'YYYY-MM-DD')
               FROM prt_part_substitute s
               JOIN prt_part p ON p.part_id=s.part_id
               JOIN prt_part p2 ON p2.part_id=s.substitute_part_id
               WHERE s.tenant_id=%s AND p.part_no=%s ORDER BY s.sub_id""",
            (tid, part_no.strip()))
        return [{"id": r[0], "partNo": r[1], "partName": r[2], "note": r[3], "at": r[4]}
                for r in cur.fetchall()]


@router.post("/parts/{part_no}/substitutes", status_code=201, dependencies=[SETUP])
def part_substitute_add(part_no: str, request: Request, body: SubstituteAdd) -> dict[str, Any]:
    """대체 자재 연결 (U5) — 자기 자신 422·미존재 404·중복 409."""
    sub_no = body.substituteNo.strip()
    if not sub_no:
        raise HTTPException(422, detail="substituteNo 필요")
    if sub_no == part_no.strip():
        raise HTTPException(422, detail="자기 자신은 대체 자재로 지정할 수 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT part_id FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, part_no.strip()))
        p1 = cur.fetchone()
        cur.execute("SELECT part_id FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, sub_no))
        p2 = cur.fetchone()
        if not p1 or not p2:
            raise HTTPException(404, detail=f"부품 없음: {part_no if not p1 else sub_no}")
        cur.execute(
            "SELECT 1 FROM prt_part_substitute WHERE tenant_id=%s AND part_id=%s AND substitute_part_id=%s",
            (tid, p1[0], p2[0]))
        if cur.fetchone():
            raise HTTPException(409, detail=f"이미 등록된 대체 관계: {part_no} → {sub_no}")
        cur.execute(
            """INSERT INTO prt_part_substitute (tenant_id, part_id, substitute_part_id, note)
               VALUES (%s,%s,%s,NULLIF(%s,'')) RETURNING sub_id""",
            (tid, p1[0], p2[0], body.note.strip()[:300]))
        sid = cur.fetchone()[0]
        _audit(cur, tid, "prt_part_substitute", sid, "PART_SUBSTITUTE_ADD", request.state.user_id,
               after={"part": part_no.strip(), "substitute": sub_no})
    return {"id": sid}


@router.delete("/parts/substitutes/{sub_id}", dependencies=[SETUP])
def part_substitute_delete(sub_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("DELETE FROM prt_part_substitute WHERE tenant_id=%s AND sub_id=%s", (tid, sub_id))
        if cur.rowcount == 0:
            raise HTTPException(404, detail=f"대체 관계 없음: {sub_id}")
        _audit(cur, tid, "prt_part_substitute", sub_id, "PART_SUBSTITUTE_DEL", request.state.user_id)
    return {"deleted": True}


@router.get("/erp/stock/trace")
def stock_trace(lot: str = "", serial: str = "", item: str = "") -> list[dict[str, Any]]:
    """로트/시리얼 이력 추적 (D2, genealogy) — 지정 로트 또는 시리얼의 전 입출고 이력."""
    if not lot.strip() and not serial.strip():
        raise HTTPException(422, detail="lot 또는 serial 중 하나는 지정해야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        conds = []
        if lot.strip():
            conds.append("lot_no=%s")
            params.append(lot.strip())
        if serial.strip():
            conds.append("serial_no=%s")
            params.append(serial.strip())
        if item.strip():
            conds.append("item_code=%s")
            params.append(item.strip())
        where = " AND ".join(conds)
        cur.execute(
            f"""SELECT item_code, location_code, movement_type, quantity, ref_type, ref_no,
                       to_char(moved_at,'YYYY-MM-DD HH24:MI'), lot_no, serial_no
                FROM inv_movement WHERE tenant_id=%s AND {where}
                ORDER BY movement_id ASC""", tuple(params))
        return [{"itemCode": r[0], "locationCode": r[1], "type": r[2], "quantity": float(r[3]),
                 "refType": r[4], "refNo": r[5], "at": r[6],
                 "lotNo": r[7] or "", "serialNo": r[8] or ""} for r in cur.fetchall()]


def _atp(cur: Any, tid: int, item: str) -> tuple[float, float, float]:
    """(on_hand, reserved(ACTIVE), available) — 품목 단위 가용재고."""
    cur.execute("SELECT COALESCE(sum(quantity),0) FROM inv_stock WHERE tenant_id=%s AND item_code=%s",
                (tid, item))
    on_hand = float(cur.fetchone()[0])
    cur.execute("SELECT COALESCE(sum(quantity),0) FROM inv_reservation "
                "WHERE tenant_id=%s AND item_code=%s AND status='ACTIVE'", (tid, item))
    reserved = float(cur.fetchone()[0])
    return on_hand, reserved, on_hand - reserved


@router.get("/erp/stock/atp")
def stock_atp() -> list[dict[str, Any]]:
    """가용재고(ATP) — 품목별 보유·ACTIVE 예약·가용(보유−예약)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT x.item_code, max(x.item_name), sum(x.on_hand) AS on_hand,
                      COALESCE(sum(x.reserved),0) AS reserved
               FROM (
                 SELECT item_code, max(item_name) item_name, sum(quantity) on_hand, 0 reserved
                   FROM inv_stock WHERE tenant_id=%s GROUP BY item_code
                 UNION ALL
                 SELECT item_code, NULL, 0, sum(quantity)
                   FROM inv_reservation WHERE tenant_id=%s AND status='ACTIVE' GROUP BY item_code
               ) x GROUP BY x.item_code ORDER BY x.item_code""", (tid, tid))
        out = []
        for r in cur.fetchall():
            on_hand, reserved = float(r[2]), float(r[3])
            out.append({"itemCode": r[0], "itemName": r[1] or "-", "onHand": on_hand,
                        "reserved": reserved, "available": on_hand - reserved})
        return out


class ReserveRequest(BaseModel):
    itemCode: str
    quantity: float
    refType: str = "SO"
    refNo: str = ""


@router.post("/erp/stock/reserve", status_code=201, dependencies=[SETUP])
def stock_reserve(request: Request, body: ReserveRequest) -> dict[str, Any]:
    """재고 예약 — 가용재고(ATP) 초과 시 409. 수주/작업지시가 재고 선점."""
    item = body.itemCode.strip()[:50]
    if not item or body.quantity <= 0:
        raise HTTPException(422, detail="필수 — 품목·수량(>0)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        on_hand, reserved, available = _atp(cur, tid, item)
        if body.quantity > available:
            raise HTTPException(409, detail=f"가용재고 부족 — 요청 {body.quantity} > 가용 {available} (보유 {on_hand}·예약 {reserved})")
        cur.execute(
            """INSERT INTO inv_reservation (tenant_id, item_code, quantity, ref_type, ref_no, created_by)
               VALUES (%s,%s,%s,%s,%s,%s) RETURNING reservation_id""",
            (tid, item, body.quantity, body.refType.strip()[:20] or None,
             body.refNo.strip()[:50] or None, str(request.state.user_id)))
        rid = cur.fetchone()[0]
        _audit(cur, tid, "inv_reservation", rid, "RESERVE", request.state.user_id,
               {"item": item, "qty": body.quantity, "ref": body.refNo})
        _, _, avail_after = _atp(cur, tid, item)
    return {"reservationId": rid, "itemCode": item, "available": avail_after}


@router.get("/erp/stock/reservations")
def stock_reservations(item: str = "", status: str = "ACTIVE") -> list[dict[str, Any]]:
    """예약 목록 (기본 ACTIVE)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if status.strip() and status != "ALL":
            clause += " AND status=%s"
            params.append(status.strip())
        if item.strip():
            clause += " AND item_code=%s"
            params.append(item.strip())
        cur.execute(
            f"""SELECT reservation_id, item_code, quantity, ref_type, ref_no, status,
                       to_char(created_at,'MM-DD HH24:MI')
                FROM inv_reservation WHERE tenant_id=%s{clause}
                ORDER BY reservation_id DESC LIMIT 100""", tuple(params))
        return [{"reservationId": r[0], "itemCode": r[1], "quantity": float(r[2]),
                 "refType": r[3], "refNo": r[4], "status": r[5], "at": r[6]} for r in cur.fetchall()]


@router.post("/erp/stock/reservations/{reservation_id}/release", dependencies=[SETUP])
def stock_reservation_release(reservation_id: int, request: Request) -> dict[str, Any]:
    """예약 해제 — ACTIVE → RELEASED (가용재고 복원)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "UPDATE inv_reservation SET status='RELEASED' "
            "WHERE tenant_id=%s AND reservation_id=%s AND status='ACTIVE' RETURNING item_code",
            (tid, reservation_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(409, detail="해제 불가 — ACTIVE 예약 아님")
        _audit(cur, tid, "inv_reservation", reservation_id, "RESERVE_RELEASE", request.state.user_id, {})
        _, _, avail = _atp(cur, tid, row[0])
    return {"reservationId": reservation_id, "itemCode": row[0], "available": avail}


# ── G3 공급처 평가/등급 ──
def _supplier_grade(total: float) -> str:
    return "A" if total >= 90 else "B" if total >= 80 else "C" if total >= 70 else "D"


@router.get("/erp/suppliers/{company_id}/metrics")
def supplier_metrics(company_id: int) -> dict[str, Any]:
    """발주 이행 지표 — PO 건수·발주/입고 수량·이행률(납기 산출 힌트)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT company_name FROM com_company WHERE tenant_id=%s AND company_id=%s",
                    (tid, company_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail="company not found")
        name = row[0]
        cur.execute("SELECT count(*), count(*) FILTER (WHERE status='CLOSED') "
                    "FROM erp_po WHERE tenant_id=%s AND supplier=%s", (tid, name))
        po_count, closed = cur.fetchone()
        cur.execute(
            """SELECT COALESCE(sum(i.order_qty),0), COALESCE(sum(i.received_qty),0)
               FROM erp_po_item i JOIN erp_po p ON p.po_id=i.po_id
               WHERE p.tenant_id=%s AND p.supplier=%s""", (tid, name))
        ordered, received = cur.fetchone()
        ordered, received = float(ordered), float(received)
        fulfillment = round(received / ordered * 100, 1) if ordered > 0 else 0.0
    return {"companyId": company_id, "supplier": name, "poCount": po_count,
            "closedCount": closed, "orderedQty": ordered, "receivedQty": received,
            "fulfillmentPct": fulfillment, "suggestedDelivery": min(100.0, fulfillment)}


@router.get("/erp/suppliers/evals")
def supplier_evals(company_id: int = 0) -> list[dict[str, Any]]:
    """공급처 평가 목록 (company_id 지정 시 해당 업체만)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if company_id:
            clause = " AND e.supplier_id=%s"
            params.append(company_id)
        cur.execute(
            f"""SELECT e.eval_id, e.supplier_id, c.company_name, e.period, e.delivery,
                       e.quality, e.price, e.total, e.grade, COALESCE(e.note,''),
                       to_char(e.created_at,'YYYY-MM-DD')
                FROM com_supplier_eval e JOIN com_company c ON c.company_id=e.supplier_id
                WHERE e.tenant_id=%s{clause}
                ORDER BY e.period DESC, e.eval_id DESC""", tuple(params))
        return [{"evalId": r[0], "supplierId": r[1], "supplier": r[2], "period": r[3],
                 "delivery": float(r[4]), "quality": float(r[5]), "price": float(r[6]),
                 "total": float(r[7]), "grade": r[8], "note": r[9], "createdAt": r[10]}
                for r in cur.fetchall()]


class SupplierEvalRequest(BaseModel):
    supplierId: int
    period: str
    delivery: float = 0
    quality: float = 0
    price: float = 0
    note: str = ""


@router.post("/erp/suppliers/evals", status_code=201, dependencies=[SETUP])
def supplier_eval_create(request: Request, body: SupplierEvalRequest) -> dict[str, Any]:
    """공급처 평가 등록/갱신 — 가중 총점(납기.4·품질.4·단가.2)→등급, com_company 등급 반영."""
    period = body.period.strip()[:7]
    if not period:
        raise HTTPException(422, detail="필수 — 평가 기간(YYYY-MM)")
    for v in (body.delivery, body.quality, body.price):
        if not (0 <= v <= 100):
            raise HTTPException(422, detail="점수는 0~100")
    total = round(body.delivery * 0.4 + body.quality * 0.4 + body.price * 0.2, 2)
    grade = _supplier_grade(total)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT company_name FROM com_company WHERE tenant_id=%s AND company_id=%s",
                    (tid, body.supplierId))
        if not cur.fetchone():
            raise HTTPException(404, detail="company not found")
        cur.execute(
            """INSERT INTO com_supplier_eval
               (tenant_id, supplier_id, period, delivery, quality, price, total, grade, note, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
               ON CONFLICT (tenant_id, supplier_id, period)
               DO UPDATE SET delivery=EXCLUDED.delivery, quality=EXCLUDED.quality,
                             price=EXCLUDED.price, total=EXCLUDED.total, grade=EXCLUDED.grade,
                             note=EXCLUDED.note, created_at=now()
               RETURNING eval_id""",
            (tid, body.supplierId, period, body.delivery, body.quality, body.price,
             total, grade, body.note.strip()[:300] or None, str(request.state.user_id)))
        eval_id = cur.fetchone()[0]
        # 최신 등급을 마스터에 반영 (해당 업체의 최신 period 평가 등급)
        cur.execute(
            """UPDATE com_company SET evaluation_grade=(
                 SELECT grade FROM com_supplier_eval
                 WHERE tenant_id=%s AND supplier_id=%s ORDER BY period DESC, eval_id DESC LIMIT 1)
               WHERE tenant_id=%s AND company_id=%s""",
            (tid, body.supplierId, tid, body.supplierId))
        _audit(cur, tid, "com_supplier_eval", eval_id, "SUPP_EVAL", request.state.user_id,
               {"supplierId": body.supplierId, "period": period, "total": total, "grade": grade})
    return {"evalId": eval_id, "total": total, "grade": grade}


class WorkOrderCreate(BaseModel):
    title: str
    drawingNo: str = ""
    projectNo: str = ""
    assemblyNote: str = ""
    assignee: str = ""


WO_TRANSITIONS = {"ISSUED": ("STARTED",), "STARTED": ("DONE",), "DONE": ()}


@router.post("/erp/work-orders", status_code=201, dependencies=[SETUP])
def work_order_create(request: Request, body: WorkOrderCreate) -> dict[str, Any]:
    """작업지시 발행 (D3) — 설계 패키지(도면·BOM·공정)를 제작 지시로. WO-{seq} 채번, ISSUED."""
    if not body.title.strip():
        raise HTTPException(422, detail="필수 — 작업지시 제목")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM erp_work_order WHERE tenant_id=%s", (tid,))
        seq = cur.fetchone()[0] + 1
        while True:
            wo_no = f"WO-{seq:04d}"
            cur.execute("SELECT 1 FROM erp_work_order WHERE tenant_id=%s AND wo_no=%s", (tid, wo_no))
            if not cur.fetchone():
                break
            seq += 1
        cur.execute(
            """INSERT INTO erp_work_order (tenant_id, wo_no, drawing_no, project_no, title,
               assembly_note, assignee, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING wo_id""",
            (tid, wo_no, body.drawingNo.strip()[:50] or None, body.projectNo.strip()[:30] or None,
             body.title.strip()[:200], body.assemblyNote.strip()[:500] or None,
             body.assignee.strip()[:50] or None, str(request.state.user_id)))
        wo_id = cur.fetchone()[0]
        _audit(cur, tid, "erp_work_order", wo_id, "ISSUE", request.state.user_id,
               {"woNo": wo_no, "title": body.title})
    return {"woNo": wo_no, "status": "ISSUED"}


@router.get("/erp/work-orders")
def work_order_list() -> list[dict[str, Any]]:
    """작업지시 목록 (D3)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT wo_no, title, drawing_no, project_no, status, COALESCE(assignee,''),
                      to_char(issued_at,'MM-DD HH24:MI'), to_char(done_at,'MM-DD HH24:MI'),
                      COALESCE(assembly_note,'')
               FROM erp_work_order WHERE tenant_id=%s ORDER BY wo_id DESC""", (tid,))
        return [{"woNo": r[0], "title": r[1], "drawingNo": r[2], "projectNo": r[3],
                 "status": r[4], "assignee": r[5], "issuedAt": r[6], "doneAt": r[7],
                 "assemblyNote": r[8]} for r in cur.fetchall()]


class WoStatusPatch(BaseModel):
    status: str


@router.patch("/erp/work-orders/{wo_no}/status", dependencies=[SETUP])
def work_order_status(wo_no: str, request: Request, body: WoStatusPatch) -> dict[str, Any]:
    """작업지시 상태 전이 (D3) — ISSUED→STARTED→DONE. 완료 시 부서 후속(알림) 연동."""
    new = body.status.strip().upper()
    if new not in ("STARTED", "DONE"):
        raise HTTPException(422, detail="상태 오류 (STARTED/DONE)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT wo_id, status, title FROM erp_work_order WHERE tenant_id=%s AND wo_no=%s",
            (tid, wo_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"작업지시 없음: {wo_no}")
        wo_id, cur_status, title = row
        if new not in WO_TRANSITIONS.get(cur_status, ()):
            raise HTTPException(409, detail=f"전이 불가: {cur_status} → {new} "
                                f"(허용: {', '.join(WO_TRANSITIONS.get(cur_status, ())) or '없음'})")
        ts_col = "started_at" if new == "STARTED" else "done_at"
        cur.execute(f"UPDATE erp_work_order SET status=%s, {ts_col}=now() WHERE wo_id=%s",
                    (new, wo_id))
        if new == "DONE":
            # 부서 후속 — 생산 관리자(ADMIN/PLATFORM) 알림 (완료 통지, D3 고리)
            cur.execute(
                "SELECT user_id FROM sys_user WHERE tenant_id=%s AND user_level IN ('ADMIN','PLATFORM')",
                (tid,))
            for (uid,) in cur.fetchall():
                _notify(cur, tid, uid, "TASK_ASSIGNED", f"작업지시 완료 — {wo_no} {title[:50]}", "/erp")
        _audit(cur, tid, "erp_work_order", wo_id, "STATUS", request.state.user_id,
               {"woNo": wo_no, "from": cur_status, "to": new})
    return {"woNo": wo_no, "status": new}


# ── D4 검사·품질 기록 (qc_inspection) — 규칙이 판정이 되는 고리 ──
QC_TYPES = ("INCOMING", "PROCESS", "OUTGOING")
QC_RESULTS = ("PASS", "FAIL", "CONDITIONAL")


class QcInspectionCreate(BaseModel):
    inspType: str                 # INCOMING | PROCESS | OUTGOING
    result: str                   # PASS | FAIL | CONDITIONAL
    refNo: str = ""               # PO/WO/도면 등 참조
    itemCode: str = ""
    itemName: str = ""
    measured: str = ""            # 측정값·판정 근거
    inspector: str = ""


@router.post("/qc/inspections", status_code=201, dependencies=[SETUP])
def qc_inspection_create(request: Request, body: QcInspectionCreate) -> dict[str, Any]:
    """검사 결과 등록 (D4) — 수입·공정·출하 검사. 불합격·조건부 시 품질 관리자 알림."""
    itype = body.inspType.strip().upper()
    result = body.result.strip().upper()
    if itype not in QC_TYPES:
        raise HTTPException(422, detail=f"검사 유형 오류 ({'/'.join(QC_TYPES)})")
    if result not in QC_RESULTS:
        raise HTTPException(422, detail=f"판정 오류 ({'/'.join(QC_RESULTS)})")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM qc_inspection WHERE tenant_id=%s", (tid,))
        seq = cur.fetchone()[0] + 1
        while True:
            insp_no = f"QC-{seq:04d}"
            cur.execute("SELECT 1 FROM qc_inspection WHERE tenant_id=%s AND insp_no=%s", (tid, insp_no))
            if not cur.fetchone():
                break
            seq += 1
        cur.execute(
            """INSERT INTO qc_inspection (tenant_id, insp_no, insp_type, ref_no, item_code,
               item_name, result, measured, inspector, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING inspection_id""",
            (tid, insp_no, itype, body.refNo.strip()[:50] or None,
             body.itemCode.strip()[:50] or None, body.itemName.strip()[:200] or None,
             result, body.measured.strip()[:500] or None, body.inspector.strip()[:50] or None,
             str(request.state.user_id)))
        insp_id = cur.fetchone()[0]
        if result in ("FAIL", "CONDITIONAL"):
            # 이상 고리 — 품질 관리자(ADMIN/PLATFORM) 알림 (불합격/조건부 통지)
            tag = "불합격" if result == "FAIL" else "조건부"
            label = (body.itemName or body.itemCode or body.refNo or "").strip()
            cur.execute(
                "SELECT user_id FROM sys_user WHERE tenant_id=%s AND user_level IN ('ADMIN','PLATFORM')",
                (tid,))
            for (uid,) in cur.fetchall():
                _notify(cur, tid, uid, "TASK_ASSIGNED",
                        f"검사 {tag} — {insp_no} {label[:50]}", "/plm")
            # 이상 이벤트 승격 (D4→통합) — 불합격=HIGH, 조건부=MED
            _raise_anomaly(cur, tid, "QC", "HIGH" if result == "FAIL" else "MED",
                           f"검사 {tag} — {insp_no} {label[:60]}", insp_no, f"QC:{insp_no}",
                           {"type": itype, "result": result})
        _audit(cur, tid, "qc_inspection", insp_id, "INSPECT", request.state.user_id,
               {"inspNo": insp_no, "type": itype, "result": result})
    return {"inspNo": insp_no, "result": result}


@router.get("/qc/inspections")
def qc_inspection_list(result: str = "", inspType: str = "") -> list[dict[str, Any]]:
    """검사 기록 목록 (D4) — 결과·유형 필터."""
    clause, params = "", []
    r = result.strip().upper()
    if r in QC_RESULTS:
        clause += " AND result=%s"
        params.append(r)
    it = inspType.strip().upper()
    if it in QC_TYPES:
        clause += " AND insp_type=%s"
        params.append(it)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT insp_no, insp_type, COALESCE(ref_no,''), COALESCE(item_code,''),
                       COALESCE(item_name,''), result, COALESCE(measured,''),
                       COALESCE(inspector,''), to_char(inspected_at,'MM-DD HH24:MI')
                FROM qc_inspection WHERE tenant_id=%s{clause}
                ORDER BY inspection_id DESC""", (tid, *params))
        return [{"inspNo": x[0], "inspType": x[1], "refNo": x[2], "itemCode": x[3],
                 "itemName": x[4], "result": x[5], "measured": x[6], "inspector": x[7],
                 "inspectedAt": x[8]} for x in cur.fetchall()]


QC_RESULT_KO = {"PASS": "합격", "FAIL": "불합격", "CONDITIONAL": "조건부"}


@router.get("/qc/certificate.pdf")
def qc_certificate(refNo: str = "", item: str = "", result: str = "") -> Any:
    """QC 성적서 PDF (D4) — 검사 이력 기반. 인증서 요구 PO(ERP-017 certRequired) 연계 표기."""
    from fastapi.responses import Response

    from ..services import run_pipeline as rp
    clause, params = "", []
    if refNo.strip():
        clause += " AND ref_no=%s"
        params.append(refNo.strip())
    if item.strip():
        clause += " AND (item_code=%s OR item_name=%s)"
        params.extend([item.strip(), item.strip()])
    r = result.strip().upper()
    if r in QC_RESULTS:
        clause += " AND result=%s"
        params.append(r)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT insp_no, insp_type, COALESCE(item_name, item_code, '-'),
                       result, COALESCE(measured,''), COALESCE(inspector,'-'),
                       to_char(inspected_at,'YYYY-MM-DD'), COALESCE(ref_no,'')
                FROM qc_inspection WHERE tenant_id=%s{clause}
                ORDER BY inspection_id""", (tid, *params))
        insp = cur.fetchall()
        if not insp:
            raise HTTPException(404, detail="성적서 대상 검사 기록이 없습니다")
        # ERP-017 — refNo 가 인증서 요구 PO 면 연계 표기
        cert_required = False
        if refNo.strip():
            cur.execute(
                "SELECT 1 FROM doc_control WHERE tenant_id=%s AND doc_no=%s AND doc_type='PO' "
                "AND remarks LIKE '%%인증서:요구%%'", (tid, refNo.strip()))
            cert_required = cur.fetchone() is not None
    total = len(insp)
    passed = sum(1 for x in insp if x[3] == "PASS")
    failed = sum(1 for x in insp if x[3] == "FAIL")
    cond = total - passed - failed
    overall = "합격" if failed == 0 else "불합격"
    lines = [
        f"발행일: {insp[0][6]}   대상: {refNo or item or '전체 검사'}",
        f"종합 판정: {overall}   (합격 {passed} · 조건부 {cond} · 불합격 {failed} / 총 {total}건)",
    ]
    if cert_required:
        lines.append("※ 인증서 요구 발주(PO) 연계 — 본 성적서는 납품 인증 첨부용 (ERP-017)")
    lines.append("-" * 58)
    lines.append("검사번호   유형    대상                판정    검사자    검사일")
    type_ko = {"INCOMING": "수입", "PROCESS": "공정", "OUTGOING": "출하"}
    for x in insp:
        lines.append(
            f"{x[0]:<9} {type_ko.get(x[1], x[1]):<5} {x[2][:18]:<18} "
            f"{QC_RESULT_KO.get(x[3], x[3]):<5} {x[5][:8]:<8} {x[6]}")
        if x[4]:
            lines.append(f"           └ 측정/근거: {x[4][:70]}")
    lines += ["-" * 58, "본 성적서는 EDIM 검사 이력(qc_inspection)에 근거하여 자동 생성되었습니다."]
    pdf = rp.build_lines_pdf(title=f"QC 성적서 (Quality Certificate) — {overall}",
                             subtitle="EDIM 품질 검사 성적서 (D4)", lines=lines)
    from urllib.parse import quote
    fname = f"QC성적서_{refNo or item or 'all'}.pdf"
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename*=UTF-8''{quote(fname)}"})


# ── 이상 이벤트 통합 (sys_anomaly) — QC 불합격·원가 차이·마일스톤 지연 승격 ──
def _raise_anomaly(cur, tid: int, source: str, severity: str, title: str,
                   ref_no: str | None, dedup_key: str, detail: dict[str, Any]) -> None:
    """이상 이벤트 등록 (dedup_key 중복 시 무시 — 스캔 반복 안전)."""
    cur.execute(
        """INSERT INTO sys_anomaly (tenant_id, source, severity, title, ref_no, dedup_key, detail)
           VALUES (%s,%s,%s,%s,%s,%s,%s)
           ON CONFLICT (tenant_id, dedup_key) DO NOTHING""",
        (tid, source, severity, title[:300], (ref_no or None), dedup_key[:120], json.dumps(detail)))


@router.post("/anomalies/scan", dependencies=[SETUP])
def anomaly_scan(request: Request) -> dict[str, Any]:
    """이상 스캔 (D6/D7 승격) — 원가 차이경보·마일스톤 지연을 이상 이벤트로 등록 (dedup)."""
    before = 0
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM sys_anomaly WHERE tenant_id=%s", (tid,))
        before = cur.fetchone()[0]
        # 1) 원가 차이경보 (D6) — 분류별 실적이 추정 대비 +임계 초과
        try:
            _run_id, est = _latest_cost_base(cur, tid)
        except HTTPException:
            est = {}
        if est:
            cur.execute(
                "SELECT category, COALESCE(sum(amount),0) FROM cst_actual WHERE tenant_id=%s GROUP BY category",
                (tid,))
            act = {r[0]: float(r[1]) for r in cur.fetchall()}
            for cat in COST_CATEGORIES:
                e = float(est.get(cat, 0))
                a = act.get(cat, 0.0)
                if e and (a - e) / e > VARIANCE_ALERT_RATE:
                    rate = round((a - e) / e * 100, 1)
                    _raise_anomaly(cur, tid, "COST", "HIGH" if rate > 25 else "MED",
                                   f"원가 차이경보 — {COST_CAT_LABEL[cat]} +{rate}% (추정 {e:,.0f}→실적 {a:,.0f})",
                                   cat, f"COST:{cat}", {"estimate": e, "actual": a, "rate": rate})
        # 2) 마일스톤 지연 (D7) — 계획일 초과·미완
        cur.execute(
            """SELECT project_no, stage, to_char(planned_date,'YYYY-MM-DD'),
                      (CURRENT_DATE - planned_date)
               FROM prj_milestone
               WHERE tenant_id=%s AND actual_date IS NULL AND planned_date < CURRENT_DATE""", (tid,))
        for proj, stage, planned, days in cur.fetchall():
            _raise_anomaly(cur, tid, "MILESTONE", "HIGH" if days > 14 else "MED",
                           f"마일스톤 지연 — {proj} {MILESTONE_LABEL.get(stage, stage)} ({days}일 초과, 계획 {planned})",
                           proj, f"MILESTONE:{proj}:{stage}", {"days": days, "planned": planned})
        cur.execute("SELECT count(*) FROM sys_anomaly WHERE tenant_id=%s", (tid,))
        after = cur.fetchone()[0]
    return {"created": after - before, "total": after}


@router.get("/anomalies")
def anomaly_list(status: str = "", source: str = "") -> dict[str, Any]:
    """이상 이벤트 목록 (통합) — 상태·출처 필터 + 요약."""
    clause, params = "", []
    st = status.strip().upper()
    if st in ("OPEN", "ACK", "RESOLVED"):
        clause += " AND status=%s"
        params.append(st)
    sc = source.strip().upper()
    if sc in ("QC", "COST", "MILESTONE", "MANUAL"):
        clause += " AND source=%s"
        params.append(sc)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT anomaly_id, source, severity, title, COALESCE(ref_no,''), status,
                       to_char(created_at,'MM-DD HH24:MI'), to_char(resolved_at,'MM-DD HH24:MI')
                FROM sys_anomaly WHERE tenant_id=%s{clause}
                ORDER BY (status='RESOLVED'),
                         CASE severity WHEN 'HIGH' THEN 0 WHEN 'MED' THEN 1 ELSE 2 END,
                         anomaly_id DESC""", (tid, *params))
        rows = [{"anomalyId": r[0], "source": r[1], "severity": r[2], "title": r[3],
                 "refNo": r[4], "status": r[5], "createdAt": r[6], "resolvedAt": r[7]}
                for r in cur.fetchall()]
        cur.execute(
            "SELECT count(*) FILTER (WHERE status='OPEN'), count(*) FILTER (WHERE status='OPEN' AND severity='HIGH') "
            "FROM sys_anomaly WHERE tenant_id=%s", (tid,))
        s = cur.fetchone()
    return {"rows": rows, "open": s[0], "openHigh": s[1]}


ANOMALY_TRANSITIONS = {"OPEN": ("ACK", "RESOLVED"), "ACK": ("RESOLVED",), "RESOLVED": ()}


class AnomalyStatus(BaseModel):
    status: str


@router.patch("/anomalies/{anomaly_id}/status", dependencies=[SETUP])
def anomaly_status(anomaly_id: int, request: Request, body: AnomalyStatus) -> dict[str, Any]:
    """이상 이벤트 상태 전이 — OPEN→ACK→RESOLVED."""
    new = body.status.strip().upper()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT status FROM sys_anomaly WHERE tenant_id=%s AND anomaly_id=%s",
                    (tid, anomaly_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"이상 이벤트 없음: {anomaly_id}")
        if new not in ANOMALY_TRANSITIONS.get(row[0], ()):
            raise HTTPException(409, detail=f"전이 불가: {row[0]} → {new}")
        cur.execute(
            "UPDATE sys_anomaly SET status=%s, "
            "resolved_at=CASE WHEN %s='RESOLVED' THEN now() ELSE resolved_at END "
            "WHERE anomaly_id=%s", (new, new, anomaly_id))
        _audit(cur, tid, "sys_anomaly", anomaly_id, "ANOMALY_STATUS", request.state.user_id,
               {"from": row[0], "to": new})
    return {"anomalyId": anomaly_id, "status": new}


ESCALATE_AGE_DAYS = 3   # OPEN 이상 방치 임계 (일)


@router.post("/anomalies/escalate", dependencies=[SETUP])
def anomaly_escalate(request: Request) -> dict[str, Any]:
    """이상 자동 에스컬레이션 (C4 연계) — 미처리(OPEN) HIGH 또는 임계 방치를 관리자 ESCALATION 통지.
    이미 에스컬레이션된 건은 재통지 안 함(detail.escalated)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT anomaly_id, title, severity FROM sys_anomaly
               WHERE tenant_id=%s AND status='OPEN'
                 AND COALESCE((detail->>'escalated')::boolean, false)=false
                 AND (severity='HIGH' OR created_at < now() - (%s||' days')::interval)
               ORDER BY anomaly_id""", (tid, ESCALATE_AGE_DAYS))
        rows = cur.fetchall()
        cur.execute(
            "SELECT user_id FROM sys_user WHERE tenant_id=%s AND user_level IN ('ADMIN','PLATFORM')",
            (tid,))
        admins = [r[0] for r in cur.fetchall()]
        escalated = []
        for aid, title, sev in rows:
            for uid in admins:
                _notify(cur, tid, uid, "ESCALATION",
                        f"이상 에스컬레이션 [{sev}] — {title[:80]}", "/erp")
            cur.execute(
                "UPDATE sys_anomaly SET detail = COALESCE(detail,'{}'::jsonb) || '{\"escalated\":true}'::jsonb "
                "WHERE anomaly_id=%s", (aid,))
            _audit(cur, tid, "sys_anomaly", aid, "ANOMALY_ESCALATE", request.state.user_id,
                   {"severity": sev})
            escalated.append(aid)
    return {"escalated": len(escalated), "admins": len(admins)}


# ── D5 설계 변경 관리 (eco_change) — Rev-up 을 공식 절차로 ──
def _eco_impact(cur, tid: int, target_type: str, target_no: str) -> dict[str, Any]:
    """영향 분석 자동 첨부 — CODE=Where-Used 역참조, DRAWING=Rev 이력·현재 Rev."""
    if target_type == "CODE":
        cur.execute(
            """SELECT mc.main_code, mc.code_name, r.quantity
               FROM code_relationship r
               JOIN product_code cc ON cc.product_code_id=r.child_code_id
               JOIN product_code mc ON mc.product_code_id=r.mother_code_id
               WHERE r.tenant_id=%s AND cc.main_code=%s ORDER BY mc.main_code""",
            (tid, target_no))
        wu = [{"code": r[0], "name": r[1], "qty": float(r[2])} for r in cur.fetchall()]
        return {"kind": "CODE", "whereUsedCount": len(wu), "whereUsed": wu,
                "note": f"이 코드를 사용하는 상위(mother) 코드 {len(wu)}건 — 변경 시 동반 검토 필요"}
    # DRAWING — 현재 Rev + Rev 이력 수
    cur.execute(
        """SELECT d.current_rev,
                  (SELECT count(*) FROM dwg_revision r WHERE r.drawing_id=d.drawing_id)
           FROM dwg_drawing d WHERE d.tenant_id=%s AND d.drawing_no=%s""",
        (tid, target_no))
    d = cur.fetchone()
    if not d:
        return {"kind": "DRAWING", "found": False, "note": "도면 미등록 — 승인 시 Rev-up 대상 없음"}
    # B20 Table 영향도 — 이 도면 치수 매크로가 참조하는 데이터 Table 수 (변경 파급 범위)
    cur.execute(
        """SELECT count(DISTINCT mr.ref_target_id), count(DISTINCT d.dimension_id)
           FROM dwg_dimension d
           JOIN dwg_drawing w ON w.drawing_id=d.drawing_id
           JOIN tbx_macro_ref mr ON mr.macro_id=d.macro_id AND mr.ref_type='TABLE'
           WHERE d.tenant_id=%s AND w.drawing_no=%s""",
        (tid, target_no))
    tref = cur.fetchone()
    table_refs, dim_macros = (tref[0] or 0, tref[1] or 0) if tref else (0, 0)
    return {"kind": "DRAWING", "found": True, "currentRev": d[0], "revCount": d[1],
            "tableRefs": table_refs, "dimMacros": dim_macros,
            "note": f"현재 Rev {d[0]} · 개정 이력 {d[1]}건 · Table 영향도 {table_refs}개(치수 매크로 {dim_macros}) "
                    f"— 승인 시 Rev-up 자동 적용"}


class EcoCreate(BaseModel):
    title: str
    targetType: str               # DRAWING | CODE
    targetNo: str
    reason: str = ""
    newDrawingNo: str = ""        # D5 — 대체 도면(supersede) 지정 시 승인 후 Rev-up 대신 대체 이력


@router.post("/eco/changes", status_code=201, dependencies=[SETUP])
def eco_create(request: Request, body: EcoCreate) -> dict[str, Any]:
    """설계변경 요청(ECR) 등록 (D5) — 영향 분석 자동 첨부 + 승인 요청(sys_approval_request) 연동."""
    tt = body.targetType.strip().upper()
    if tt not in ("DRAWING", "CODE"):
        raise HTTPException(422, detail="대상 유형 오류 (DRAWING/CODE)")
    if not body.title.strip() or not body.targetNo.strip():
        raise HTTPException(422, detail="필수 — 제목·대상 번호")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        impact = _eco_impact(cur, tid, tt, body.targetNo.strip())
        # D5 — 대체 도면 지정 시 영향 분석에 병기 (승인 시 dwg_supersedure 자동 등록)
        if tt == "DRAWING" and body.newDrawingNo.strip():
            impact["supersededBy"] = body.newDrawingNo.strip()[:80]
            impact["mode"] = "SUPERSEDE"
        cur.execute("SELECT count(*) FROM eco_change WHERE tenant_id=%s", (tid,))
        seq = cur.fetchone()[0] + 1
        while True:
            eco_no = f"ECO-{seq:04d}"
            cur.execute("SELECT 1 FROM eco_change WHERE tenant_id=%s AND eco_no=%s", (tid, eco_no))
            if not cur.fetchone():
                break
            seq += 1
        cur.execute(
            """INSERT INTO eco_change (tenant_id, eco_no, title, reason, target_type,
               target_no, impact_data, status, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,'SUBMITTED',%s) RETURNING eco_id""",
            (tid, eco_no, body.title.strip()[:200], body.reason.strip()[:1000] or None,
             tt, body.targetNo.strip()[:80], json.dumps(impact), str(request.state.user_id)))
        eco_id = cur.fetchone()[0]
        # 단계 승인 — 기존 승인함(sys_approval_request) 재사용
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'eco_change',%s,'CHANGE','승인',%s,%s)""",
            (tid, eco_id, request.state.user_id, f"{eco_no} {body.title.strip()[:150]}"))
        cur.execute(
            """SELECT user_id FROM sys_user
               WHERE tenant_id=%s AND user_level IN ('SETUP','ADMIN') AND user_id<>%s""",
            (tid, request.state.user_id))
        for (uid,) in cur.fetchall():
            _notify(cur, tid, uid, "APPROVAL_REQUEST",
                    f"설계변경 승인 요청 — {eco_no} {body.title.strip()[:60]}", "/common")
        _audit(cur, tid, "eco_change", eco_id, "ECR", request.state.user_id,
               {"ecoNo": eco_no, "target": f"{tt}:{body.targetNo}"})
    return {"ecoNo": eco_no, "status": "SUBMITTED", "impact": impact}


@router.get("/eco/changes")
def eco_list() -> list[dict[str, Any]]:
    """설계변경 목록 (D5)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT eco_no, title, target_type, target_no, status,
                      COALESCE(rev_from,''), COALESCE(rev_to,''),
                      COALESCE((impact_data->>'whereUsedCount')::int,
                               (impact_data->>'revCount')::int, 0),
                      to_char(created_at,'MM-DD HH24:MI'),
                      COALESCE(reason,'')
               FROM eco_change WHERE tenant_id=%s ORDER BY eco_id DESC""", (tid,))
        return [{"ecoNo": r[0], "title": r[1], "targetType": r[2], "targetNo": r[3],
                 "status": r[4], "revFrom": r[5], "revTo": r[6], "impactCount": r[7],
                 "createdAt": r[8], "reason": r[9]} for r in cur.fetchall()]


@router.get("/eco/ledger")
def eco_ledger(status: str = "", targetType: str = "") -> dict[str, Any]:
    """변경 이력 대장 전용 뷰 (D5) — 전체 설계변경 라이프사이클 대장 + 상태 집계.

    변경유형 파생: APPLIED+rev_to=Rev-up · APPLIED+rev_null=대체(Supersede) · REJECTED=반려 · else 진행."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        conds = ["tenant_id=%s"]
        params: list[Any] = [tid]
        if status.strip():
            conds.append("status=%s"); params.append(status.strip().upper())
        if targetType.strip():
            conds.append("target_type=%s"); params.append(targetType.strip().upper())
        where = " AND ".join(conds)
        cur.execute(
            f"""SELECT eco_no, title, target_type, target_no, status,
                       COALESCE(rev_from,''), COALESCE(rev_to,''), COALESCE(reason,''),
                       created_by, to_char(created_at,'YYYY-MM-DD HH24:MI'),
                       to_char(applied_at,'YYYY-MM-DD HH24:MI'),
                       COALESCE(impact_data->>'newDrawingNo','')
                FROM eco_change WHERE {where} ORDER BY eco_id DESC""", tuple(params))
        rows = []
        for r in cur.fetchall():
            st, rev_to, new_dwg = r[4], r[6], r[11]
            if st == "REJECTED":
                ctype = "반려"
            elif st == "APPLIED" and rev_to:
                ctype = "Rev-up"
            elif st == "APPLIED" and (new_dwg or not rev_to):
                ctype = "대체"
            else:
                ctype = "진행"
            rows.append({
                "ecoNo": r[0], "title": r[1], "targetType": r[2], "targetNo": r[3],
                "status": st, "revFrom": r[5], "revTo": rev_to, "reason": r[7],
                "createdBy": r[8], "createdAt": r[9], "appliedAt": r[10] or "",
                "newDrawingNo": new_dwg, "changeType": ctype,
                "revTransition": (f"{r[5] or '—'} → {rev_to}" if rev_to else (f"→ {new_dwg}" if new_dwg else "—")),
            })
        # 상태 집계 (필터 무관 전체)
        cur.execute(
            """SELECT status, count(*) FROM eco_change WHERE tenant_id=%s GROUP BY status""", (tid,))
        by_status = {k: v for k, v in cur.fetchall()}
        summary = {
            "total": sum(by_status.values()),
            "applied": by_status.get("APPLIED", 0),
            "pending": by_status.get("SUBMITTED", 0) + by_status.get("APPROVED", 0) + by_status.get("DRAFT", 0),
            "rejected": by_status.get("REJECTED", 0),
        }
    return {"summary": summary, "rows": rows}


@router.get("/eco/changes/{eco_no}")
def eco_detail(eco_no: str) -> dict[str, Any]:
    """설계변경 상세 + 영향 분석 (D5)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT eco_no, title, COALESCE(reason,''), target_type, target_no, status,
                      COALESCE(rev_from,''), COALESCE(rev_to,''), impact_data,
                      to_char(created_at,'MM-DD HH24:MI'), to_char(applied_at,'MM-DD HH24:MI')
               FROM eco_change WHERE tenant_id=%s AND eco_no=%s""", (tid, eco_no))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"설계변경 없음: {eco_no}")
        return {"ecoNo": r[0], "title": r[1], "reason": r[2], "targetType": r[3],
                "targetNo": r[4], "status": r[5], "revFrom": r[6], "revTo": r[7],
                "impact": r[8], "createdAt": r[9], "appliedAt": r[10]}


# ── D6 원가 실적 (cst_actual) — 추정이 실적으로 검증되는 고리 ──
COST_CATEGORIES = ("MATERIAL", "MANUFACTURING", "DIRECT")
COST_CAT_LABEL = {"MATERIAL": "재료비", "MANUFACTURING": "제조비", "DIRECT": "직접경비"}
VARIANCE_ALERT_RATE = 0.10   # 실적이 추정 대비 +10% 초과 시 경보


class CostActualCreate(BaseModel):
    category: str                 # MATERIAL | MANUFACTURING | DIRECT
    itemCode: str = ""
    itemName: str = ""
    poNo: str = ""
    qty: float = 1
    unitPrice: float = 0
    projectNo: str = ""           # 프로젝트 귀속 (선택)


@router.post("/cost/actuals", status_code=201, dependencies=[SETUP])
def cost_actual_create(request: Request, body: CostActualCreate) -> dict[str, Any]:
    """구매 실적 적재 (D6) — PO 확정 단가 → 실적 원가(cst_calc 추정과 분리 기록). 프로젝트 귀속(선택)."""
    cat = body.category.strip().upper()
    if cat not in COST_CATEGORIES:
        raise HTTPException(422, detail=f"원가 분류 오류 ({'/'.join(COST_CATEGORIES)})")
    if body.qty <= 0 or body.unitPrice < 0:
        raise HTTPException(422, detail="수량·단가 확인 (수량>0, 단가≥0)")
    amount = round(body.qty * body.unitPrice, 2)
    pno = body.projectNo.strip()[:30] or None
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if pno:
            cur.execute("SELECT 1 FROM prj_project WHERE tenant_id=%s AND project_no=%s", (tid, pno))
            if not cur.fetchone():
                raise HTTPException(404, detail=f"프로젝트 없음: {pno}")
        cur.execute(
            """INSERT INTO cst_actual (tenant_id, category, item_code, item_name, po_no,
               qty, unit_price, amount, project_no, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING actual_id""",
            (tid, cat, body.itemCode.strip()[:50] or None, body.itemName.strip()[:200] or None,
             body.poNo.strip()[:50] or None, body.qty, body.unitPrice, amount, pno,
             str(request.state.user_id)))
        aid = cur.fetchone()[0]
        _audit(cur, tid, "cst_actual", aid, "ACTUAL", request.state.user_id,
               {"category": cat, "amount": amount, "poNo": body.poNo, "projectNo": pno})
    return {"actualId": aid, "category": cat, "amount": amount, "projectNo": pno}


@router.get("/cost/actuals")
def cost_actual_list(project: str = "") -> list[dict[str, Any]]:
    """실적 원가 목록 (D6). project 지정 시 해당 프로젝트 귀속분만."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if project.strip():
            clause = " AND project_no=%s"
            params.append(project.strip())
        cur.execute(
            f"""SELECT actual_id, category, COALESCE(item_code,''), COALESCE(item_name,''),
                      COALESCE(po_no,''), qty, unit_price, amount,
                      to_char(recorded_at,'MM-DD HH24:MI'), COALESCE(project_no,'')
               FROM cst_actual WHERE tenant_id=%s{clause} ORDER BY actual_id DESC""", tuple(params))
        return [{"actualId": r[0], "category": r[1], "itemCode": r[2], "itemName": r[3],
                 "poNo": r[4], "qty": float(r[5]), "unitPrice": float(r[6]),
                 "amount": float(r[7]), "recordedAt": r[8], "projectNo": r[9]}
                for r in cur.fetchall()]


@router.get("/cost/variance")
def cost_variance(project: str = "") -> dict[str, Any]:
    """견적(추정) vs 실적 차이 분석 (D6) — 분류별 추정/실적/차이·차이율 + 임계 경보.

    project 지정 시 추정(프로젝트 Run)·실적(프로젝트 귀속) 모두 프로젝트 스코프."""
    proj = project.strip()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        # 추정 — 최근 SUCCESS Run 의 cst_calc 분류 합계 (없으면 0), 프로젝트 스코프
        try:
            run_id, est = _latest_cost_base(cur, tid, proj)
        except HTTPException:
            run_id, est = None, {}
        # 실적 — cst_actual 분류 합계 (프로젝트 스코프)
        aparams: list[Any] = [tid]
        aclause = ""
        if proj:
            aclause = " AND project_no=%s"
            aparams.append(proj)
        cur.execute(
            f"SELECT category, COALESCE(sum(amount),0) FROM cst_actual WHERE tenant_id=%s{aclause} GROUP BY category",
            tuple(aparams))
        act = {r[0]: float(r[1]) for r in cur.fetchall()}
        cats = []
        for c in COST_CATEGORIES:
            e = float(est.get(c, 0))
            a = act.get(c, 0.0)
            var = round(a - e, 2)
            rate = (var / e) if e else (1.0 if a else 0.0)
            cats.append({"category": c, "label": COST_CAT_LABEL[c], "estimate": e,
                         "actual": a, "variance": var, "varianceRate": round(rate, 4),
                         "alert": rate > VARIANCE_ALERT_RATE})
        te = sum(x["estimate"] for x in cats)
        ta = sum(x["actual"] for x in cats)
        tv = round(ta - te, 2)
        trate = (tv / te) if te else (1.0 if ta else 0.0)
        return {"runId": run_id, "estimateAvailable": run_id is not None, "projectNo": proj,
                "categories": cats, "totalEstimate": te, "totalActual": ta,
                "totalVariance": tv, "totalVarianceRate": round(trate, 4),
                "alert": trate > VARIANCE_ALERT_RATE, "alertRate": VARIANCE_ALERT_RATE}


# ── D7 프로젝트 일정·마일스톤 (prj_milestone) ──
MILESTONE_STAGES = ("ORDER", "DESIGN", "PURCHASE", "PRODUCTION", "SHIPMENT")
MILESTONE_LABEL = {"ORDER": "수주", "DESIGN": "설계", "PURCHASE": "구매",
                   "PRODUCTION": "제작", "SHIPMENT": "출하"}
DUE_SOON_DAYS = 7   # 계획일 D-7 이내 = 임박


class MilestoneCreate(BaseModel):
    projectNo: str
    stage: str                    # ORDER | DESIGN | PURCHASE | PRODUCTION | SHIPMENT
    plannedDate: str              # YYYY-MM-DD
    note: str = ""


@router.post("/erp/milestones", status_code=201, dependencies=[SETUP])
def milestone_create(request: Request, body: MilestoneCreate) -> dict[str, Any]:
    """마일스톤 등록 (D7) — 프로젝트 단계별 납기 (project×stage UNIQUE upsert)."""
    stage = body.stage.strip().upper()
    if stage not in MILESTONE_STAGES:
        raise HTTPException(422, detail=f"단계 오류 ({'/'.join(MILESTONE_STAGES)})")
    if not body.projectNo.strip() or not body.plannedDate.strip():
        raise HTTPException(422, detail="필수 — 프로젝트·계획일")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, body.projectNo.strip()))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"프로젝트 없음: {body.projectNo}")
        cur.execute(
            """INSERT INTO prj_milestone (tenant_id, project_no, stage, planned_date, note, created_by)
               VALUES (%s,%s,%s,%s,%s,%s)
               ON CONFLICT (tenant_id, project_no, stage)
               DO UPDATE SET planned_date=EXCLUDED.planned_date, note=EXCLUDED.note
               RETURNING milestone_id""",
            (tid, body.projectNo.strip()[:30], stage, body.plannedDate.strip(),
             body.note.strip()[:300] or None, str(request.state.user_id)))
        mid = cur.fetchone()[0]
        _audit(cur, tid, "prj_milestone", mid, "PLAN", request.state.user_id,
               {"project": body.projectNo, "stage": stage, "planned": body.plannedDate})
    return {"milestoneId": mid, "stage": stage}


_MILESTONE_SELECT = """
    SELECT milestone_id, project_no, stage, to_char(planned_date,'YYYY-MM-DD'),
           to_char(actual_date,'YYYY-MM-DD'), status, COALESCE(note,''),
           CASE WHEN actual_date IS NOT NULL THEN 'DONE'
                WHEN planned_date < CURRENT_DATE THEN 'OVERDUE'
                WHEN planned_date <= CURRENT_DATE + %s THEN 'DUE_SOON'
                ELSE 'PENDING' END,
           (planned_date - CURRENT_DATE),
           CASE WHEN actual_date IS NOT NULL THEN 0 ELSE
             (SELECT count(*)::int FROM generate_series(
                LEAST(CURRENT_DATE, planned_date) + 1, GREATEST(CURRENT_DATE, planned_date),
                interval '1 day') gs
              WHERE extract(dow FROM gs) NOT IN (0, 6)
                AND gs::date NOT IN (SELECT holiday_date FROM cal_holiday WHERE tenant_id=%s))
             * CASE WHEN planned_date < CURRENT_DATE THEN -1 ELSE 1 END
           END
    FROM prj_milestone WHERE tenant_id=%s"""


def _milestone_row(r) -> dict[str, Any]:
    return {"milestoneId": r[0], "projectNo": r[1], "stage": r[2],
            "stageLabel": MILESTONE_LABEL.get(r[2], r[2]), "plannedDate": r[3],
            "actualDate": r[4], "status": r[5], "note": r[6],
            "delayStatus": r[7], "daysLeft": r[8], "workdaysLeft": r[9]}


@router.get("/erp/milestones")
def milestone_list(project: str = "") -> list[dict[str, Any]]:
    """마일스톤 목록 (D7) — 지연 임박/초과 자동 판정. 단계 순 정렬."""
    order = ("ORDER BY project_no, array_position(ARRAY['ORDER','DESIGN','PURCHASE',"
             "'PRODUCTION','SHIPMENT']::text[], stage)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if project.strip():
            cur.execute(_MILESTONE_SELECT + " AND project_no=%s " + order,
                        (DUE_SOON_DAYS, tid, tid, project.strip()))
        else:
            cur.execute(_MILESTONE_SELECT + " " + order, (DUE_SOON_DAYS, tid, tid))
        return [_milestone_row(r) for r in cur.fetchall()]


class MilestoneDone(BaseModel):
    actualDate: str = ""          # 생략 시 오늘


@router.patch("/erp/milestones/{milestone_id}/done", dependencies=[SETUP])
def milestone_done(milestone_id: int, request: Request, body: MilestoneDone) -> dict[str, Any]:
    """마일스톤 완료 (D7) — 실제 완료일 기록 + 진척 갱신."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        ad = body.actualDate.strip() or None
        cur.execute(
            """UPDATE prj_milestone
               SET status='DONE', actual_date=COALESCE(%s::date, CURRENT_DATE)
               WHERE tenant_id=%s AND milestone_id=%s RETURNING project_no, stage""",
            (ad, tid, milestone_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"마일스톤 없음: {milestone_id}")
        _audit(cur, tid, "prj_milestone", milestone_id, "DONE", request.state.user_id,
               {"project": row[0], "stage": row[1]})
    return {"milestoneId": milestone_id, "status": "DONE"}


@router.get("/erp/milestones/summary")
def milestone_summary() -> dict[str, Any]:
    """마일스톤 요약 (D7) — 지연/임박 카운트 + 프로젝트별 진척 롤업 (Dashboard)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT project_no,
                      count(*),
                      count(*) FILTER (WHERE actual_date IS NOT NULL),
                      count(*) FILTER (WHERE actual_date IS NULL AND planned_date < CURRENT_DATE),
                      count(*) FILTER (WHERE actual_date IS NULL AND planned_date >= CURRENT_DATE
                                       AND planned_date <= CURRENT_DATE + %s)
               FROM prj_milestone WHERE tenant_id=%s GROUP BY project_no ORDER BY project_no""",
            (DUE_SOON_DAYS, tid))
        projects = [{"projectNo": r[0], "total": r[1], "done": r[2],
                     "overdue": r[3], "dueSoon": r[4],
                     "progress": round(r[2] / r[1] * 100) if r[1] else 0}
                    for r in cur.fetchall()]
        return {"projects": projects,
                "totalOverdue": sum(p["overdue"] for p in projects),
                "totalDueSoon": sum(p["dueSoon"] for p in projects),
                "projectCount": len(projects)}


# ── G3 근무일/휴일 캘린더 (cal_holiday) — 영업일 기준 기한 계산 ──
def _holiday_set(cur, tid: int) -> set:
    cur.execute("SELECT holiday_date FROM cal_holiday WHERE tenant_id=%s", (tid,))
    return {r[0] for r in cur.fetchall()}


def _add_workdays(start: "date", n: int, holidays: set) -> "date":
    from datetime import timedelta
    d = start
    step = 1 if n >= 0 else -1
    cnt = 0
    while cnt < abs(n):
        d = d + timedelta(days=step)
        if d.weekday() < 5 and d not in holidays:
            cnt += 1
    return d


@router.get("/calendar/holidays")
def holidays_list(year: int = 0) -> list[dict[str, Any]]:
    """공휴일 목록 (year 지정 시 해당 연도)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if year:
            cur.execute("SELECT holiday_id, to_char(holiday_date,'YYYY-MM-DD'), name FROM cal_holiday "
                        "WHERE tenant_id=%s AND extract(year FROM holiday_date)=%s ORDER BY holiday_date",
                        (tid, year))
        else:
            cur.execute("SELECT holiday_id, to_char(holiday_date,'YYYY-MM-DD'), name FROM cal_holiday "
                        "WHERE tenant_id=%s ORDER BY holiday_date", (tid,))
        return [{"holidayId": r[0], "date": r[1], "name": r[2]} for r in cur.fetchall()]


class HolidayCreate(BaseModel):
    date: str
    name: str


@router.post("/calendar/holidays", status_code=201, dependencies=[SETUP])
def holiday_create(request: Request, body: HolidayCreate) -> dict[str, Any]:
    """공휴일 등록 (중복 날짜 409)."""
    if not body.date.strip() or not body.name.strip():
        raise HTTPException(422, detail="필수 — 날짜·명칭")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM cal_holiday WHERE tenant_id=%s AND holiday_date=%s::date",
                    (tid, body.date.strip()))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — {body.date} 이미 등록됨")
        cur.execute(
            "INSERT INTO cal_holiday (tenant_id, holiday_date, name, created_by) "
            "VALUES (%s,%s::date,%s,%s) RETURNING holiday_id",
            (tid, body.date.strip(), body.name.strip()[:100], str(request.state.user_id)))
        hid = cur.fetchone()[0]
        _audit(cur, tid, "cal_holiday", hid, "CREATE", request.state.user_id,
               {"date": body.date, "name": body.name})
    return {"holidayId": hid, "date": body.date}


@router.delete("/calendar/holidays/{holiday_id}", dependencies=[SETUP])
def holiday_delete(holiday_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("DELETE FROM cal_holiday WHERE tenant_id=%s AND holiday_id=%s RETURNING holiday_date",
                    (tid, holiday_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail="공휴일 없음")
        _audit(cur, tid, "cal_holiday", holiday_id, "DELETE", request.state.user_id, {})
    return {"deleted": holiday_id}


@router.get("/calendar/workdays")
def calendar_workdays(start: str, end: str) -> dict[str, Any]:
    """두 날짜 사이 영업일 수 (주말·공휴일 제외, start 제외·end 포함)."""
    from datetime import date as _date
    try:
        d1 = _date.fromisoformat(start.strip()); d2 = _date.fromisoformat(end.strip())
    except ValueError:
        raise HTTPException(422, detail="날짜 형식 YYYY-MM-DD")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        hol = _holiday_set(cur, tid)
    from datetime import timedelta
    lo, hi = (d1, d2) if d1 <= d2 else (d2, d1)
    cnt = 0
    d = lo
    while d < hi:
        d = d + timedelta(days=1)
        if d.weekday() < 5 and d not in hol:
            cnt += 1
    return {"start": start, "end": end, "workdays": cnt if d1 <= d2 else -cnt}


@router.get("/calendar/due")
def calendar_due(start: str, days: int) -> dict[str, Any]:
    """start 로부터 N 영업일 후(음수=이전) 날짜."""
    from datetime import date as _date
    try:
        d0 = _date.fromisoformat(start.strip())
    except ValueError:
        raise HTTPException(422, detail="날짜 형식 YYYY-MM-DD")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        hol = _holiday_set(cur, tid)
    due = _add_workdays(d0, days, hol)
    return {"start": start, "days": days, "due": due.isoformat()}


# ── G3 다통화/환율 + 세금 엔진 (fx_rate·tax_code) ──
def _fx_rate(cur, tid: int, currency: str) -> float | None:
    """최신 환율(외화 1 = ? KRW). KRW=1."""
    c = currency.strip().upper()
    if c == "KRW":
        return 1.0
    cur.execute("SELECT rate FROM fx_rate WHERE tenant_id=%s AND currency=%s "
                "ORDER BY valid_from DESC LIMIT 1", (tid, c))
    r = cur.fetchone()
    return float(r[0]) if r else None


@router.get("/finance/fx")
def fx_list() -> list[dict[str, Any]]:
    """환율 목록 — 통화별 최신 (기준통화 KRW=1 포함)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT DISTINCT ON (currency) fx_id, currency, rate, to_char(valid_from,'YYYY-MM-DD')
               FROM fx_rate WHERE tenant_id=%s ORDER BY currency, valid_from DESC""", (tid,))
        rows = [{"fxId": r[0], "currency": r[1], "rate": float(r[2]), "validFrom": r[3]}
                for r in cur.fetchall()]
    return [{"fxId": 0, "currency": "KRW", "rate": 1.0, "validFrom": "기준통화"}] + rows


class FxCreate(BaseModel):
    currency: str
    rate: float
    validFrom: str = ""


@router.post("/finance/fx", status_code=201, dependencies=[SETUP])
def fx_upsert(request: Request, body: FxCreate) -> dict[str, Any]:
    """환율 등록/갱신 (통화+적용일 upsert). KRW 는 기준통화라 등록 불가."""
    c = body.currency.strip().upper()[:3]
    if c == "KRW":
        raise HTTPException(422, detail="KRW 는 기준통화(고정 1)")
    if not c or body.rate <= 0:
        raise HTTPException(422, detail="필수 — 통화·환율(>0)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        vf = body.validFrom.strip() or None
        cur.execute(
            """INSERT INTO fx_rate (tenant_id, currency, rate, valid_from, created_by)
               VALUES (%s,%s,%s,COALESCE(%s::date, CURRENT_DATE),%s)
               ON CONFLICT (tenant_id, currency, valid_from)
               DO UPDATE SET rate=EXCLUDED.rate RETURNING fx_id""",
            (tid, c, body.rate, vf, str(request.state.user_id)))
        fid = cur.fetchone()[0]
        _audit(cur, tid, "fx_rate", fid, "FX_SET", request.state.user_id, {"currency": c, "rate": body.rate})
    return {"fxId": fid, "currency": c, "rate": body.rate}


@router.delete("/finance/fx/{fx_id}", dependencies=[SETUP])
def fx_delete(fx_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("DELETE FROM fx_rate WHERE tenant_id=%s AND fx_id=%s RETURNING currency", (tid, fx_id))
        if not cur.fetchone():
            raise HTTPException(404, detail="환율 없음")
        _audit(cur, tid, "fx_rate", fx_id, "DELETE", request.state.user_id, {})
    return {"deleted": fx_id}


@router.get("/finance/tax-codes")
def tax_codes() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT tax_id, code, name, rate_pct FROM tax_code WHERE tenant_id=%s ORDER BY code", (tid,))
        return [{"taxId": r[0], "code": r[1], "name": r[2], "ratePct": float(r[3])} for r in cur.fetchall()]


class TaxCreate(BaseModel):
    code: str
    name: str
    ratePct: float = 0


@router.post("/finance/tax-codes", status_code=201, dependencies=[SETUP])
def tax_code_create(request: Request, body: TaxCreate) -> dict[str, Any]:
    code = body.code.strip().upper()[:20]
    if not code or not body.name.strip():
        raise HTTPException(422, detail="필수 — 코드·명칭")
    if not (0 <= body.ratePct <= 100):
        raise HTTPException(422, detail="세율 0~100")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM tax_code WHERE tenant_id=%s AND code=%s", (tid, code))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 세금코드 {code}")
        cur.execute(
            "INSERT INTO tax_code (tenant_id, code, name, rate_pct, created_by) "
            "VALUES (%s,%s,%s,%s,%s) RETURNING tax_id",
            (tid, code, body.name.strip()[:100], body.ratePct, str(request.state.user_id)))
        tx = cur.fetchone()[0]
        _audit(cur, tid, "tax_code", tx, "CREATE", request.state.user_id, {"code": code})
    return {"taxId": tx, "code": code}


@router.delete("/finance/tax-codes/{tax_id}", dependencies=[SETUP])
def tax_code_delete(tax_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("DELETE FROM tax_code WHERE tenant_id=%s AND tax_id=%s RETURNING code", (tid, tax_id))
        if not cur.fetchone():
            raise HTTPException(404, detail="세금코드 없음")
        _audit(cur, tid, "tax_code", tax_id, "DELETE", request.state.user_id, {})
    return {"deleted": tax_id}


class QuoteCalcRequest(BaseModel):
    currency: str = "KRW"
    amount: float = 0
    taxCode: str = ""


@router.post("/finance/quote-calc")
def quote_calc(body: QuoteCalcRequest) -> dict[str, Any]:
    """세금엔진 — 통화 금액 → 세액·합계 + 기준통화(KRW) 환산."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        rate = _fx_rate(cur, tid, body.currency)
        if rate is None:
            raise HTTPException(422, detail=f"환율 미등록 통화: {body.currency}")
        pct = 0.0
        if body.taxCode.strip():
            cur.execute("SELECT rate_pct FROM tax_code WHERE tenant_id=%s AND code=%s",
                        (tid, body.taxCode.strip().upper()))
            tr = cur.fetchone()
            if not tr:
                raise HTTPException(422, detail=f"세금코드 없음: {body.taxCode}")
            pct = float(tr[0])
    cur_c = body.currency.strip().upper()
    amt = round(body.amount, 2)
    tax = round(amt * pct / 100, 2)
    total = round(amt + tax, 2)
    return {
        "currency": cur_c, "rate": rate, "taxPct": pct,
        "amount": amt, "taxAmount": tax, "total": total,
        "baseAmount": round(amt * rate, 2), "baseTax": round(tax * rate, 2),
        "baseTotal": round(total * rate, 2), "baseCurrency": "KRW",
    }


# ── D10 Head 메뉴 편집 (sys_menu_config) — 사용자별 모듈 표시 구성 ──
ALL_MODULES = ["cpq", "plm", "code", "erp", "toolbox", "common"]
MODULE_LABELS = {"cpq": "CPQ", "plm": "PLM", "code": "Code Set-up",
                 "erp": "ERP", "toolbox": "Toolbox", "common": "공통"}


def _effective_modules(stored: list[str] | None) -> tuple[list[str], bool]:
    """저장 구성 → 실효 모듈. 빈/미설정=전체. 'common' 은 항상 포함(잠금 방지)."""
    if not stored:
        return list(ALL_MODULES), False
    eff = [m for m in ALL_MODULES if m in stored or m == "common"]
    return (eff or list(ALL_MODULES)), True


@router.get("/menu/modules")
def menu_modules() -> list[dict[str, str]]:
    """구성 가능한 모듈 목록 (D10 관리 UI)."""
    return [{"id": m, "label": MODULE_LABELS[m]} for m in ALL_MODULES]


@router.get("/menu/config")
def menu_config(request: Request) -> dict[str, Any]:
    """현재 사용자의 표시 모듈 (D10) — Shell 필터용."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT modules FROM sys_menu_config WHERE tenant_id=%s AND login_id=%s",
                    (tid, request.state.login))
        row = cur.fetchone()
    eff, restricted = _effective_modules(row[0] if row else None)
    return {"modules": eff, "restricted": restricted}


@router.get("/menu/config/{login}", dependencies=[ADMIN])
def menu_config_get(login: str) -> dict[str, Any]:
    """특정 사용자의 저장 구성 (D10 관리 — ADMIN)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM sys_user WHERE tenant_id=%s AND login_id=%s", (tid, login))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        cur.execute("SELECT modules FROM sys_menu_config WHERE tenant_id=%s AND login_id=%s",
                    (tid, login))
        row = cur.fetchone()
    stored = row[0] if row else []
    eff, restricted = _effective_modules(stored)
    return {"login": login, "modules": stored, "effective": eff, "restricted": restricted}


class MenuConfigPut(BaseModel):
    modules: list[str]   # 표시 허용 모듈 목록 (빈 목록 = 제한 해제, 전체 표시)


@router.put("/menu/config/{login}", dependencies=[ADMIN])
def menu_config_put(login: str, request: Request, body: MenuConfigPut) -> dict[str, Any]:
    """사용자 표시 모듈 구성 저장 (D10 — ADMIN). 빈 목록 = 제한 해제."""
    mods = [m for m in body.modules if m in ALL_MODULES]
    invalid = [m for m in body.modules if m not in ALL_MODULES]
    if invalid:
        raise HTTPException(422, detail=f"알 수 없는 모듈: {', '.join(invalid)}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM sys_user WHERE tenant_id=%s AND login_id=%s", (tid, login))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        cur.execute(
            """INSERT INTO sys_menu_config (tenant_id, login_id, modules, updated_by)
               VALUES (%s,%s,%s,%s)
               ON CONFLICT (tenant_id, login_id)
               DO UPDATE SET modules=EXCLUDED.modules, updated_by=EXCLUDED.updated_by, updated_at=now()""",
            (tid, login, json.dumps(mods), request.state.login))
        _audit(cur, tid, "sys_menu_config", 0, "MENU_CONFIG", request.state.user_id,
               {"login": login, "modules": mods})
    eff, restricted = _effective_modules(mods)
    return {"login": login, "modules": mods, "effective": eff, "restricted": restricted}


# ── D8 사용자 환경설정 (sys_user_pref) — 즐겨찾기·컬럼 설정 등 ──
PREF_KEYS = ("favorites", "recent", "gridColumns")


class PrefPut(BaseModel):
    value: Any = None   # 임의 JSON (배열/객체) — 클라이언트 UI 상태


# ── U11 테넌트 브랜딩 (슬라이드 57 — 회사 로고 배치) ──

class BrandingPut(BaseModel):
    logoData: str = ""   # data URL (base64) — '' = 제거


# ── U30 — 테넌트 기본 좌측 메뉴 (관리자 지정, 사용자 개인 pref > 테넌트 기본 > 전체 트리) ──
class TenantLeftNavPut(BaseModel):
    value: dict[str, Any] = {}


@router.get("/tenant/leftnav")
def tenant_leftnav() -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT settings FROM sys_tenant WHERE tenant_id=%s", (tid,))
        row = cur.fetchone()
        st = row[0] if row and row[0] else {}
    v = st.get("tenantLeftNav")
    return {"value": v if isinstance(v, dict) else {}}


@router.put("/tenant/leftnav", dependencies=[ADMIN])
def tenant_leftnav_put(request: Request, body: TenantLeftNavPut) -> dict[str, Any]:
    """테넌트 기본 좌측 메뉴 — 모듈별 leaf id 목록 (관리자). {} = 기본 전체 트리 복귀."""
    v = body.value or {}
    for k, ids in v.items():
        if not isinstance(ids, list) or not all(isinstance(x, str) for x in ids):
            raise HTTPException(422, detail=f"모듈 {k}: 문자열 배열이어야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_tenant
               SET settings = COALESCE(settings,'{}'::jsonb) || jsonb_build_object('tenantLeftNav', %s::jsonb)
               WHERE tenant_id=%s""", (json.dumps(v), tid))
        _audit(cur, tid, "sys_tenant", tid, "TENANT_LEFTNAV_SET", request.state.user_id,
               {"modules": list(v.keys())})
    return {"value": v}


@router.get("/tenant/headnav")
def tenant_headnav() -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT settings FROM sys_tenant WHERE tenant_id=%s", (tid,))
        row = cur.fetchone()
        st = row[0] if row and row[0] else {}
    v = st.get("tenantHeadNav")
    return {"value": v if isinstance(v, dict) else {}}


@router.put("/tenant/headnav", dependencies=[ADMIN])
def tenant_headnav_put(request: Request, body: TenantLeftNavPut) -> dict[str, Any]:
    """U30 확장 — 테넌트 기본 헤더 드롭다운 (관리자). {} = 기본 전체 복귀."""
    v = body.value or {}
    for k, ids in v.items():
        if not isinstance(ids, list) or not all(isinstance(x, str) for x in ids):
            raise HTTPException(422, detail=f"모듈 {k}: 문자열 배열이어야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_tenant
               SET settings = COALESCE(settings,'{}'::jsonb) || jsonb_build_object('tenantHeadNav', %s::jsonb)
               WHERE tenant_id=%s""", (json.dumps(v), tid))
        _audit(cur, tid, "sys_tenant", tid, "TENANT_HEADNAV_SET", request.state.user_id,
               {"modules": list(v.keys())})
    return {"value": v}


@router.get("/tenant/branding")
def tenant_branding() -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT settings FROM sys_tenant WHERE tenant_id=%s", (tid,))
        row = cur.fetchone()
        st = row[0] if row and row[0] else {}
    return {"logoData": st.get("logoData") or None}


@router.put("/tenant/branding", dependencies=[ADMIN])
def tenant_branding_put(request: Request, body: BrandingPut) -> dict[str, Any]:
    """테넌트 로고 설정 (U11) — data URL base64, 64KB 상한. '' = 제거."""
    data = body.logoData.strip()
    if data:
        if not data.startswith("data:image/"):
            raise HTTPException(422, detail="data:image/* 형식의 data URL 이어야 합니다")
        if len(data) > 64 * 1024:
            raise HTTPException(422, detail="로고는 64KB 이하로 축소하십시오")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_tenant
               SET settings = COALESCE(settings,'{}'::jsonb) || jsonb_build_object('logoData', %s::text)
               WHERE tenant_id=%s""",
            (data or None, tid))
        _audit(cur, tid, "sys_tenant", tid, "BRANDING_SET", request.state.user_id,
               after={"logo": bool(data)})
    return {"saved": True}


# ── 2.0 좌측 패널 = 업무 프로세스 (요구 #15/#17) — 고객사가 정의하는 프로세스 트리 ──

class ProcessNodeCreate(BaseModel):
    name: str
    parentId: int | None = None
    icon: str = ""
    screenHref: str = ""
    note: str = ""


class ProcessNodePatch(BaseModel):
    name: str | None = None
    icon: str | None = None
    screenHref: str | None = None
    note: str | None = None
    parentId: int | None = None
    stepNo: int | None = None


# 기본 프로세스 — 고객사가 즉시 쓰기 시작할 수 있는 표준 흐름 (전량 편집 가능)
_PROCESS_SEED: list[tuple[str, str, str, list[tuple[str, str, str]]]] = [
    ("영업·견적", "①", "", [
        ("프로젝트 등록", "", "/erp/projects"),
        ("사양 선택 (CPQ)", "", "/cpq/selection"),
        ("견적 산출", "", "/cpq/run"),
        ("견적서 발행", "", "/erp/sales-order"),
    ]),
    ("설계·기술", "②", "", [
        ("Sub Code 관리", "", "/code/subcode"),
        ("제품 코드", "", "/code/product-codes"),
        ("BOM 관계", "", "/code/relationship"),
        ("도면 관리", "", "/plm/drawings"),
    ]),
    ("생산·구매", "③", "", [
        ("작업 지시", "", "/erp/work-order"),
        ("소요 계획", "", "/erp/mrp"),
        ("발주", "", "/erp/po"),
        ("재고", "", "/erp/inventory"),
    ]),
    ("품질·출하", "④", "", [
        ("검사", "", "/erp/quality"),
        ("ERP Handoff", "", "/erp/sales-order"),
        ("산출물 패키지", "", "/common/folder"),
    ]),
    ("공통", "⑤", "", [
        ("승인함", "", "/common/approval"),
        ("Run 이력·Snapshot", "", "/toolbox/runs"),
    ]),
]


@router.get("/process/tree")
def process_tree() -> list[dict[str, Any]]:
    """좌측 프로세스 트리 — 미정의 테넌트는 빈 배열(프런트가 메뉴 모드로 폴백)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT node_id, parent_id, name, COALESCE(icon,''), COALESCE(screen_href,''),
                      step_no, COALESCE(note,''), status
               FROM sys_process_node WHERE tenant_id=%s AND status='ACTIVE'
               ORDER BY COALESCE(parent_id,0), step_no, node_id""", (tid,))
        return [{"nodeId": r[0], "parentId": r[1], "name": r[2], "icon": r[3],
                 "screenHref": r[4], "stepNo": r[5], "note": r[6], "status": r[7]}
                for r in cur.fetchall()]


@router.post("/process/seed", dependencies=[SETUP])
def process_seed(request: Request) -> dict[str, Any]:
    """표준 프로세스 시드 — 이미 정의돼 있으면 409 (덮어쓰기 방지)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM sys_process_node WHERE tenant_id=%s", (tid,))
        if cur.fetchone()[0]:
            raise HTTPException(409, detail="이미 프로세스가 정의되어 있습니다 — 편집으로 변경하십시오")
        n = 0
        for i, (name, icon, href, children) in enumerate(_PROCESS_SEED):
            cur.execute(
                """INSERT INTO sys_process_node (tenant_id, parent_id, name, icon, screen_href,
                                                 step_no, created_by)
                   VALUES (%s,NULL,%s,%s,NULLIF(%s,''),%s,%s) RETURNING node_id""",
                (tid, name, icon, href, i, request.state.user_id))
            pid = cur.fetchone()[0]
            n += 1
            for j, (cname, cicon, chref) in enumerate(children):
                cur.execute(
                    """INSERT INTO sys_process_node (tenant_id, parent_id, name, icon, screen_href,
                                                     step_no, created_by)
                       VALUES (%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),%s,%s)""",
                    (tid, pid, cname, cicon, chref, j, request.state.user_id))
                n += 1
        _audit(cur, tid, "sys_process_node", 0, "PROCESS_SEED", request.state.user_id, {"nodes": n})
    return {"seeded": n}


@router.post("/process/nodes", status_code=201, dependencies=[SETUP])
def process_node_create(request: Request, body: ProcessNodeCreate) -> dict[str, Any]:
    name = body.name.strip()[:80]
    if not name:
        raise HTTPException(422, detail="단계 이름은 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if body.parentId:
            cur.execute("SELECT 1 FROM sys_process_node WHERE tenant_id=%s AND node_id=%s",
                        (tid, body.parentId))
            if not cur.fetchone():
                raise HTTPException(404, detail=f"상위 단계 없음: #{body.parentId}")
        cur.execute(
            """SELECT COALESCE(max(step_no),-1)+1 FROM sys_process_node
               WHERE tenant_id=%s AND parent_id IS NOT DISTINCT FROM %s""", (tid, body.parentId))
        step = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_process_node (tenant_id, parent_id, name, icon, screen_href,
                                             step_no, note, created_by)
               VALUES (%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),%s,NULLIF(%s,''),%s) RETURNING node_id""",
            (tid, body.parentId, name, body.icon.strip()[:8], body.screenHref.strip()[:200],
             step, body.note.strip()[:200], request.state.user_id))
        nid = cur.fetchone()[0]
        _audit(cur, tid, "sys_process_node", nid, "CREATE", request.state.user_id, {"name": name})
    return {"nodeId": nid, "stepNo": step}


@router.patch("/process/nodes/{node_id}", dependencies=[SETUP])
def process_node_patch(node_id: int, request: Request, body: ProcessNodePatch) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if body.parentId is not None and body.parentId == node_id:
            raise HTTPException(422, detail="자기 자신을 상위로 지정할 수 없습니다")
        # 지정된 필드만 갱신 — NULL 파라미터 타입 추론 실패(500) 회피를 위해 동적 SET 절
        sets: list[str] = []
        params: list[Any] = []
        if body.name is not None:
            sets.append("name=%s")
            params.append(body.name.strip()[:80])
        if body.icon is not None:
            sets.append("icon=NULLIF(%s,'')")
            params.append(body.icon.strip()[:8])
        if body.screenHref is not None:
            sets.append("screen_href=NULLIF(%s,'')")
            params.append(body.screenHref.strip()[:200])
        if body.note is not None:
            sets.append("note=NULLIF(%s,'')")
            params.append(body.note.strip()[:200])
        if body.parentId is not None:
            sets.append("parent_id=%s")
            params.append(body.parentId)
        if body.stepNo is not None:
            sets.append("step_no=%s")
            params.append(body.stepNo)
        if not sets:
            raise HTTPException(422, detail="변경할 항목이 없습니다")
        sets.append("updated_at=now()")
        cur.execute(
            f"""UPDATE sys_process_node SET {', '.join(sets)}
                WHERE tenant_id=%s AND node_id=%s RETURNING name, step_no""",
            (*params, tid, node_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"단계 없음: #{node_id}")
        _audit(cur, tid, "sys_process_node", node_id, "UPDATE", request.state.user_id,
               {"name": row[0], "stepNo": row[1]})
    return {"nodeId": node_id, "name": row[0], "stepNo": row[1]}


@router.post("/process/nodes/{node_id}/move", dependencies=[SETUP])
def process_node_move(node_id: int, request: Request, dir: str = "up") -> dict[str, Any]:
    """형제 간 순서 이동 — 인접 단계와 step_no 교환."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT parent_id, step_no FROM sys_process_node WHERE tenant_id=%s AND node_id=%s",
                    (tid, node_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"단계 없음: #{node_id}")
        parent, step = row
        op_cmp, order = ("<", "DESC") if dir == "up" else (">", "ASC")
        cur.execute(
            f"""SELECT node_id, step_no FROM sys_process_node
                WHERE tenant_id=%s AND parent_id IS NOT DISTINCT FROM %s AND step_no {op_cmp} %s
                ORDER BY step_no {order} LIMIT 1""", (tid, parent, step))
        nb = cur.fetchone()
        if not nb:
            return {"nodeId": node_id, "moved": False}
        cur.execute("UPDATE sys_process_node SET step_no=%s WHERE tenant_id=%s AND node_id=%s",
                    (nb[1], tid, node_id))
        cur.execute("UPDATE sys_process_node SET step_no=%s WHERE tenant_id=%s AND node_id=%s",
                    (step, tid, nb[0]))
        _audit(cur, tid, "sys_process_node", node_id, "MOVE", request.state.user_id, {"dir": dir})
    return {"nodeId": node_id, "moved": True}


@router.delete("/process/nodes/{node_id}", dependencies=[SETUP])
def process_node_delete(node_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM sys_process_node WHERE tenant_id=%s AND parent_id=%s",
                    (tid, node_id))
        kids = cur.fetchone()[0]
        cur.execute(
            "DELETE FROM sys_process_node WHERE tenant_id=%s AND node_id=%s RETURNING name",
            (tid, node_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"단계 없음: #{node_id}")
        _audit(cur, tid, "sys_process_node", node_id, "DELETE", request.state.user_id,
               {"name": row[0], "children": kids})
    return {"deleted": node_id, "children": kids}


# ── 1.5 정보 접근 권한 관리 (요구 #4/#6) — 역할×정보그룹 매트릭스·임시 접근 ──

class InfoAccessSet(BaseModel):
    roleName: str
    infoGroup: str
    mode: str


class TempAccessGrant(BaseModel):
    login: str
    infoGroup: str
    mode: str = "full"
    hours: int = 8
    reason: str = ""


@router.get("/access/info")
def info_access_list(request: Request) -> dict[str, Any]:
    """정보 접근 매트릭스 — 역할×정보그룹 모드(미설정=full) + 내 유효 모드."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT role_name, info_group, mode FROM sys_info_access WHERE tenant_id=%s", (tid,))
        rules = [{"roleName": r[0], "infoGroup": r[1], "mode": r[2]} for r in cur.fetchall()]
        cur.execute("SELECT role_name FROM sys_role WHERE tenant_id=%s ORDER BY role_id", (tid,))
        roles = [r[0] for r in cur.fetchall()]
        mine = {g: _info_mode(cur, tid, request, g) for g in INFO_GROUPS}
    return {"groups": [{"key": k, "label": v} for k, v in INFO_GROUPS.items()],
            "modes": list(INFO_MODES), "roles": roles, "rules": rules, "mine": mine}


@router.put("/access/info", dependencies=[ADMIN])
def info_access_set(request: Request, body: InfoAccessSet) -> dict[str, Any]:
    """역할별 정보 열람 모드 설정 — full 지정 시 규칙 삭제(기본값 복귀)."""
    if body.infoGroup not in INFO_GROUPS:
        raise HTTPException(422, detail=f"정보그룹 오류 ({'/'.join(INFO_GROUPS)})")
    if body.mode not in INFO_MODES:
        raise HTTPException(422, detail=f"모드 오류 ({'/'.join(INFO_MODES)})")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if body.mode == "full":
            cur.execute(
                "DELETE FROM sys_info_access WHERE tenant_id=%s AND role_name=%s AND info_group=%s",
                (tid, body.roleName, body.infoGroup))
        else:
            cur.execute(
                """INSERT INTO sys_info_access (tenant_id, role_name, info_group, mode, updated_by)
                   VALUES (%s,%s,%s,%s,%s)
                   ON CONFLICT (tenant_id, role_name, info_group)
                   DO UPDATE SET mode=EXCLUDED.mode, updated_by=EXCLUDED.updated_by, updated_at=now()""",
                (tid, body.roleName, body.infoGroup, body.mode, request.state.login))
        _audit(cur, tid, "sys_info_access", 0, "INFO_ACCESS", request.state.user_id,
               {"role": body.roleName, "group": body.infoGroup, "mode": body.mode})
    return {"roleName": body.roleName, "infoGroup": body.infoGroup, "mode": body.mode}


@router.get("/access/temp")
def temp_access_list() -> list[dict[str, Any]]:
    """임시 열람 부여 현황 (요구 #6 Temporary Access) — 유효/만료 구분."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT t.temp_access_id, u.login_id, t.info_group, t.mode, t.reason,
                      to_char(t.valid_to,'MM-DD HH24:MI'), t.revoked,
                      (NOT t.revoked AND t.valid_to > now())
               FROM sys_temp_access t JOIN sys_user u ON u.user_id=t.user_id
               WHERE t.tenant_id=%s ORDER BY t.temp_access_id DESC LIMIT 50""", (tid,))
        return [{"id": r[0], "login": r[1], "infoGroup": r[2], "mode": r[3], "reason": r[4],
                 "validTo": r[5], "revoked": r[6], "active": r[7]} for r in cur.fetchall()]


@router.post("/access/temp", status_code=201, dependencies=[ADMIN])
def temp_access_grant(request: Request, body: TempAccessGrant) -> dict[str, Any]:
    """기간 한정 임시 열람 부여 — 사유·만료 필수, 자동 만료(별도 작업 불요)."""
    if body.infoGroup not in INFO_GROUPS or body.mode not in INFO_MODES:
        raise HTTPException(422, detail="정보그룹/모드 오류")
    hours = max(1, min(int(body.hours or 8), 720))
    if not body.reason.strip():
        raise HTTPException(422, detail="접근 사유는 필수입니다 (감사 대상)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, body.login.strip()))
        u = cur.fetchone()
        if not u:
            raise HTTPException(404, detail=f"사용자 없음: {body.login}")
        cur.execute(
            """INSERT INTO sys_temp_access (tenant_id, user_id, info_group, mode, reason,
                                            granted_by, valid_to)
               VALUES (%s,%s,%s,%s,%s,%s, now() + (%s || ' hours')::interval)
               RETURNING temp_access_id, to_char(valid_to,'MM-DD HH24:MI')""",
            (tid, u[0], body.infoGroup, body.mode, body.reason.strip()[:300],
             request.state.user_id, str(hours)))
        row = cur.fetchone()
        _notify(cur, tid, u[0], "ACCESS",
                f"임시 열람 부여 — {INFO_GROUPS[body.infoGroup]} ({body.mode}, {row[1]} 까지)")
        _audit(cur, tid, "sys_temp_access", row[0], "TEMP_GRANT", request.state.user_id,
               {"login": body.login, "group": body.infoGroup, "mode": body.mode, "hours": hours})
    return {"id": row[0], "validTo": row[1]}


@router.delete("/access/temp/{grant_id}", dependencies=[ADMIN])
def temp_access_revoke(grant_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "UPDATE sys_temp_access SET revoked=true WHERE tenant_id=%s AND temp_access_id=%s "
            "RETURNING info_group", (tid, grant_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"부여 없음: #{grant_id}")
        _audit(cur, tid, "sys_temp_access", grant_id, "TEMP_REVOKE", request.state.user_id,
               {"group": row[0]})
    return {"id": grant_id, "revoked": True}


# ── 1.3 플랫폼 — 고객사(테넌트) 프로비저닝 (런치: 온보딩을 psql 없이 수행) ──

class TenantCreate(BaseModel):
    tenantCode: str
    tenantName: str
    plan: str = "SAAS"
    adminLogin: str = "admin"
    adminName: str = "관리자"
    adminPassword: str = ""


class TenantPatch(BaseModel):
    tenantName: str | None = None
    plan: str | None = None
    status: str | None = None    # ACTIVE | SUSPENDED | TERMINATED


# 신규 테넌트 최소 사용 가능 상태 — Hierarchy 루트 (주소 가드 0.9 전제)
_TENANT_SEED_NODES = [
    ("PRODUCT", "/C", "Code"), ("PRODUCT", "/M", "Macro"), ("PRODUCT", "/T", "Table"),
]


@router.get("/platform/tenants", dependencies=[PLATFORM])
def platform_tenants() -> list[dict[str, Any]]:
    """고객사 목록 (플랫폼 운영) — 사용자 수·주요 데이터량 병기."""
    with _conn() as conn, conn.cursor() as cur:
        cur.execute(
            """SELECT t.tenant_id, t.tenant_code, t.tenant_name, t.plan, t.status,
                      to_char(t.created_at,'YYYY-MM-DD'),
                      (SELECT count(*) FROM sys_user u WHERE u.tenant_id=t.tenant_id),
                      (SELECT count(*) FROM product_code p WHERE p.tenant_id=t.tenant_id),
                      (SELECT count(*) FROM prj_project j WHERE j.tenant_id=t.tenant_id)
               FROM sys_tenant t ORDER BY t.tenant_id""")
        return [{"tenantId": r[0], "tenantCode": r[1], "tenantName": r[2], "plan": r[3],
                 "status": r[4], "createdAt": r[5], "users": r[6],
                 "productCodes": r[7], "projects": r[8]} for r in cur.fetchall()]


@router.post("/platform/tenants", status_code=201, dependencies=[PLATFORM])
def platform_tenant_create(request: Request, body: TenantCreate) -> dict[str, Any]:
    """고객사 생성 — 테넌트 + 초기 관리자 + Hierarchy 루트 시드까지 한 번에 (온보딩 1스텝)."""
    code = body.tenantCode.strip()[:30]
    name = body.tenantName.strip()[:100]
    login_id = body.adminLogin.strip()[:50]
    pw = body.adminPassword.strip()
    if not code or not name or not login_id:
        raise HTTPException(422, detail="필수 — 고객사 코드·이름·관리자 사번")
    if len(pw) < 6:
        raise HTTPException(422, detail="관리자 초기 비밀번호는 6자 이상이어야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT 1 FROM sys_tenant WHERE tenant_code=%s", (code,))
        if cur.fetchone():
            raise HTTPException(409, detail=f"중복 — 고객사 코드 {code}")
        cur.execute(
            """INSERT INTO sys_tenant (tenant_code, tenant_name, plan, status, created_by)
               VALUES (%s,%s,%s,'ACTIVE',%s) RETURNING tenant_id""",
            (code, name, body.plan.strip()[:20] or "SAAS", str(request.state.user_id)))
        new_tid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_user (tenant_id, login_id, user_name, department, user_level,
                                     password_hash, status)
               VALUES (%s,%s,%s,'',%s,%s,'ACTIVE') RETURNING user_id""",
            (new_tid, login_id, body.adminName.strip()[:50] or "관리자", "ADMIN",
             hashlib.sha256(pw.encode()).hexdigest()))
        admin_uid = cur.fetchone()[0]
        for tree, addr, node in _TENANT_SEED_NODES:
            cur.execute(
                """INSERT INTO sys_hierarchy (tenant_id, parent_id, tree_type, node_name,
                                              address, approval_status)
                   VALUES (%s,NULL,%s,%s,%s,'APPROVED')""", (new_tid, tree, node, addr))
        _audit(cur, _tenant_id(cur), "sys_tenant", new_tid, "TENANT_CREATE",
               request.state.user_id, {"tenantCode": code, "admin": login_id})
    return {"tenantId": new_tid, "tenantCode": code, "adminUserId": admin_uid,
            "seededNodes": len(_TENANT_SEED_NODES)}


@router.patch("/platform/tenants/{code}", dependencies=[PLATFORM])
def platform_tenant_patch(code: str, request: Request, body: TenantPatch) -> dict[str, Any]:
    """고객사 상태·플랜 변경 — SUSPENDED/TERMINATED 는 즉시 로그인 차단(계약 게이트)."""
    st = (body.status or "").strip().upper()
    if st and st not in ("ACTIVE", "SUSPENDED", "TERMINATED"):
        raise HTTPException(422, detail="상태 오류 (ACTIVE/SUSPENDED/TERMINATED)")
    with _conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT tenant_id, tenant_code FROM sys_tenant WHERE tenant_code=%s", (code,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"고객사 없음: {code}")
        if row[1] == TENANT and st in ("SUSPENDED", "TERMINATED"):
            raise HTTPException(409, detail="플랫폼 운영 테넌트는 중지할 수 없습니다")
        cur.execute(
            """UPDATE sys_tenant SET
               tenant_name=COALESCE(NULLIF(%s,''), tenant_name),
               plan=COALESCE(NULLIF(%s,''), plan),
               status=COALESCE(NULLIF(%s,''), status),
               updated_by=%s, updated_at=now()
               WHERE tenant_id=%s RETURNING tenant_name, plan, status""",
            ((body.tenantName or "").strip(), (body.plan or "").strip(), st,
             str(request.state.user_id), row[0]))
        t = cur.fetchone()
        _audit(cur, _tenant_id(cur), "sys_tenant", row[0], "TENANT_UPDATE",
               request.state.user_id, {"status": t[2], "plan": t[1]})
    return {"tenantCode": code, "tenantName": t[0], "plan": t[1], "status": t[2]}


@router.get("/prefs/{key}")
def pref_get(key: str, request: Request) -> dict[str, Any]:
    """현재 사용자 환경설정 조회 (D8). 미설정 시 빈 값."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT pref_value FROM sys_user_pref WHERE tenant_id=%s AND login_id=%s AND pref_key=%s",
            (tid, request.state.login, key.strip()[:60]))
        row = cur.fetchone()
    return {"key": key, "value": row[0] if row else None}


@router.put("/prefs/{key}")
def pref_put(key: str, request: Request, body: PrefPut) -> dict[str, Any]:
    """환경설정 저장 (D8) — key-value(jsonb) upsert. 사용자 본인 범위."""
    k = key.strip()[:60]
    if not k:
        raise HTTPException(422, detail="pref key 필요")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """INSERT INTO sys_user_pref (tenant_id, login_id, pref_key, pref_value)
               VALUES (%s,%s,%s,%s)
               ON CONFLICT (tenant_id, login_id, pref_key)
               DO UPDATE SET pref_value=EXCLUDED.pref_value, updated_at=now()""",
            (tid, request.state.login, k, json.dumps(body.value)))
    return {"key": k, "saved": True}


# ── SVC-12 Project Folder 파일 (M-15-8) — cpq_output 실데이터 ──
OUTPUT_KIND = {"DWG": ("승인도", "ok"), "PRICE": ("견적/원가", "info"),
               "DATA": ("기술자료", "ok"), "BOM": ("BOM", "ok")}


@router.get("/files")
def project_files(project: str = "PS-61313-5", allRuns: bool = False) -> list[dict[str, Any]]:
    """Project Folder 파일 — 기본은 최신 SUCCESS Run 산출물만(최신 Rev 필터). allRuns=true=전체 Run."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        run_clause = ("" if allRuns else
                      " AND r.run_id = (SELECT max(run_id) FROM cpq_run "
                      "WHERE tenant_id=%s AND status='SUCCESS')")
        params = (tid,) if allRuns else (tid, tid)
        cur.execute(
            f"""SELECT o.output_type, o.data->>'file', o.data->>'fileType',
                       o.run_id, to_char(o.created_at,'MM-DD')
                FROM cpq_output o
                JOIN cpq_run r ON r.run_id=o.run_id AND r.tenant_id=%s{run_clause}
                ORDER BY o.run_id DESC, o.output_id""", params)
        rows = cur.fetchall()
    files = [
        {
            "name": (r[1] or "output").replace(" ", "_")[:60],
            "fileType": r[2] or "PDF",
            "kind": OUTPUT_KIND.get(r[0], (r[0], "info"))[0],
            "kindTone": OUTPUT_KIND.get(r[0], (r[0], "info"))[1],
            "run": f"#{r[3]}", "date": r[4], "folder": r[0],
        }
        for r in rows
    ]
    # 업로드 실파일 (dwg_file — MinIO 저장, fileId 로 다운로드 가능)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT f.file_id, f.file_name, f.file_type, f.folder,
                      to_char(f.uploaded_date,'MM-DD'), f.created_by
               FROM dwg_file f
               LEFT JOIN prj_project p ON p.project_id=f.project_id
               WHERE f.tenant_id=%s AND (p.project_no=%s OR f.project_id IS NULL)
               ORDER BY f.file_id DESC""", (tid, project))
        uploads = [
            {"name": r[1], "fileType": r[2],
             "kind": "접수자료" if r[3] == "RECEIVED" else "업로드", "kindTone": "info",
             "run": "-", "date": r[4], "folder": r[3], "fileId": r[0], "registrant": r[5]}
            for r in cur.fetchall()
        ]
    return files + uploads


# ── SVC-07 CPQ / ENG-02 Run — 실 파이프라인 (P3-1 · P2-4) ──
from app.services import run_pipeline as rp   # noqa: E402

RUN_TASKS = [
    "BOM 전개 (Code Relationship 재귀)",
    "치수 Macro 평가 (엔진 v1)",
    "도면 생성 (제작 DXF)",
    "원가·견적 (단가 resolve)",
    "견적서 PDF·BOM XLSX 생성",
    "서류·Project Folder 저장 (MinIO)",
]

_runs: dict[int, dict[str, Any]] = {}  # 진행 상태 (산출물·완료는 DB 영속)


class RunRequest(BaseModel):
    runType: str = "ALL"
    selectionId: int | None = None   # C1 — 지정 견적안 대상 실행 (미지정 시 최신)
    isTest: bool = False             # E3 — 테스트성 Run 표기 (스위트·Studio Test, 통계 제외)


def _out_row(folder: str, fname: str, ftype: str, file_id: int | None) -> dict[str, Any]:
    meta = {
        "DWG": ("생성", "ok", "미리보기"),
        "PRICE": ("DRAFT", "info", "QCR 발행"),
        "BOM": ("생성", "ok", "ERP 전송"),
        "DATA": ("생성", "ok", None),
    }.get(folder, ("생성", "ok", None))
    return {"folder": folder, "file": fname, "fileType": ftype,
            "status": meta[0], "statusTone": meta[1], "nextAction": meta[2],
            "fileId": file_id}


async def _advance(run_id: int, tid: int, selection_id: int,
                   slot_values: dict[str, str], project_no: str) -> None:
    state = _runs[run_id]
    r = rp.PipelineResult()

    def begin(i: int) -> float:
        state["current"] = i
        state["steps"][i]["status"] = "RUNNING"
        return time.perf_counter()

    def finish(i: int, t0: float, measured: str, warn: bool = False) -> None:
        state["steps"][i]["measured"] = measured
        state["steps"][i]["elapsed"] = f"{time.perf_counter() - t0:.2f}s"
        state["steps"][i]["status"] = "WARN" if warn else "DONE"

    def log(msg: str, level: str = "info") -> None:
        state["logs"].append({
            "time": time.strftime("%H:%M:%S"), "message": msg, "level": level})

    try:
        pool = get_pool()
        with pool.connection() as conn, conn.cursor() as cur:
            # 1. BOM
            t0 = begin(0)
            await asyncio.sleep(0.4)
            m = rp.step_bom(cur, tid, _expand_rows, "KDCR 3-13", slot_values, selection_id, r,
                            rel_basis=_rel_basis)
            finish(0, t0, m)
            log(f"BOM expand root=KDCR 3-13 … {len(r.items)} items → cpq_selection_item")

            # 2. 치수 Macro (엔진)
            t0 = begin(1)
            await asyncio.sleep(0.4)
            m = rp.step_dims(cur, tid, _make_table_resolver(cur, tid), r)
            finish(1, t0, m)
            log("dims " + " · ".join(f"{k}={v:g}" for k, v in r.dims.items()))

            # 3. 도면 (DXF)
            t0 = begin(2)
            await asyncio.sleep(0.4)
            m = rp.step_drawing(r)
            finish(2, t0, m)
            log("drawing compose → DWG/ (ezdxf R2010)")

            # 4. 원가
            t0 = begin(3)
            await asyncio.sleep(0.4)
            m = rp.step_pricing(r)
            has_warn = bool(r.warn)
            finish(3, t0, m, warn=has_warn)
            for w in r.warn:
                log(f"warn {w}", "warn")
            log(f"pricing total {r.total_k:,.0f}K · resolve {r.resolved}/{len(r.items)}")
            # 원가 상세 영속 (B18) — cst_calc 3분류 (재료비/제조비/직접경비)
            _write_cst_calc(cur, tid, run_id, r)
            log("cst_calc — MATERIAL·MANUFACTURING·DIRECT 상세 적재")

            # 5. 견적서 PDF + BOM XLSX
            t0 = begin(4)
            await asyncio.sleep(0.4)
            m1 = rp.step_quotation(r, project_no)
            m2 = rp.step_bom_xlsx(r)
            finish(4, t0, f"{m1} · {m2}")
            log("quotation PDF (워터마크) + BOM XLSX 생성")

            # 6. 저장 (MinIO + dwg_file + cpq_output)
            t0 = begin(5)
            await asyncio.sleep(0.4)
            file_ids = rp.persist_outputs(cur, tid, run_id, project_no, r)
            finish(5, t0, f"산출물 {len(file_ids)} → MinIO/{project_no}")
            log(f"outputs {len(file_ids)} → Project Folder SUCCESS")

            state["outputs"] = [
                _out_row(folder, fname, ftype, fid)
                for (folder, fname, ftype, _), fid in zip(r.files, file_ids)
            ]
            state["totalK"] = r.total_k
            cur.execute(
                """UPDATE cpq_run SET status='SUCCESS', finished_at=now(),
                   dimension_values=%s, bom_snapshot=%s, rel_basis=%s WHERE run_id=%s""",
                (json.dumps({"KDCR 3-13": r.dims}),
                 # 트리아지 #41 — BOM Snapshot: 전개 결과를 Run 에 고정 (같은 Snapshot = 같은 결과 재현)
                 json.dumps(r.items),
                 # #40 — 전개 근거(관계 Revision 집합) 고정: "같은 근거면 같은 BOM" 을 대조 가능하게
                 json.dumps(r.rel_basis) if r.rel_basis else None, run_id))
        state["status"] = "SUCCESS"
        state["current"] = len(RUN_TASKS)
    except Exception as e:  # noqa: BLE001
        state["status"] = "FAILED"
        log(f"실패: {e}", "warn")
        try:
            pool = get_pool()
            with pool.connection() as conn, conn.cursor() as cur:
                cur.execute(
                    """UPDATE cpq_run SET status='FAILED', finished_at=now(),
                       error_detail=%s WHERE run_id=%s""",
                    (json.dumps({"error": str(e)}), run_id))
        except Exception:  # noqa: BLE001
            pass


def _write_cst_calc(cur, tid: int, run_id: int, r) -> None:
    """B18 — Run 원가 상세 3분류를 cst_calc 에 영속 (detail=JSONB 라인 내역).

    MATERIAL = BOM 단가 resolve 결과 · MANUFACTURING = 조립 공수(dwg_bom 노트×표준 임율)
    · DIRECT = 운송/검사 직접경비 (재료비 비율)."""
    material_lines = [
        {"code": i["resolvedCode"], "name": i.get("name", ""), "qty": i["quantity"],
         "priceK": i["priceK"], "amount": round((i["priceK"] or 0) * 1000 * i["quantity"])}
        for i in r.items
    ]
    material_total = round(r.total_k * 1000)
    # 제조비 — dwg_bom 조립 스텝 × 표준 2h × 임율 55,000/h (CST-003 입력 전 표준치)
    cur.execute(
        """SELECT COALESCE(b.assembly_note, p.part_name)
           FROM dwg_bom b JOIN prt_part p ON p.part_id=b.part_id
           JOIN dwg_drawing d ON d.drawing_id=b.drawing_id
           WHERE d.tenant_id=%s AND d.drawing_no='KDCR 3-13'
           ORDER BY COALESCE(b.assembly_seq, 999)""", (tid,))
    steps = [row[0] for row in cur.fetchall()] or ["조립 (표준)"]
    rate = 55_000
    mfg_lines = [{"step": s, "hours": 2, "rate": rate, "amount": 2 * rate} for s in steps]
    mfg_total = sum(ln["amount"] for ln in mfg_lines)
    direct_lines = [
        {"item": "운송·포장", "basis": "재료비 3%", "amount": round(material_total * 0.03)},
        {"item": "검사·시운전", "basis": "재료비 2%", "amount": round(material_total * 0.02)},
    ]
    direct_total = sum(ln["amount"] for ln in direct_lines)
    cur.execute("DELETE FROM cst_calc WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
    for ctype, lines, total in (
        ("MATERIAL", material_lines, material_total),
        ("MANUFACTURING", mfg_lines, mfg_total),
        ("DIRECT", direct_lines, direct_total),
    ):
        cur.execute(
            """INSERT INTO cst_calc (tenant_id, run_id, calc_type, detail, total_amount)
               VALUES (%s,%s,%s,%s,%s)""",
            (tid, run_id, ctype, json.dumps({"lines": lines}), total))


@router.get("/cpq/runs/{run_id}/costs")
def run_costs(run_id: int) -> list[dict[str, Any]]:
    """Run 원가 상세 — cst_calc 3분류 (B18)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT calc_type, detail, total_amount FROM cst_calc
               WHERE tenant_id=%s AND run_id=%s ORDER BY calc_id""", (tid, run_id))
        rows = cur.fetchall()
    if not rows:
        raise HTTPException(404, detail=f"원가 상세 없음 — Run #{run_id} (v10.2 이후 실행분부터 적재)")
    return [{"calcType": r[0], "lines": r[1].get("lines", []), "total": float(r[2])} for r in rows]


# ── B21 시스템·UX 마감 — auth/me·다중 역할·Hierarchy 편집·문서 채번/전이·초대/비활성 ──

@router.get("/auth/me")
def auth_me(request: Request) -> dict[str, Any]:
    """현재 세션 사용자 — 프론트 세션 복원·역할 표시 (스펙 get_auth_me)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT login_id, user_name, department, user_level, status
               FROM sys_user WHERE tenant_id=%s AND user_id=%s""", (tid, request.state.user_id))
        row = cur.fetchone()
        cur.execute(
            """SELECT r.role_name FROM sys_user_role ur
               JOIN sys_role r ON r.role_id=ur.role_id
               WHERE ur.user_id=%s ORDER BY r.role_name""", (request.state.user_id,))
        roles = [x[0] for x in cur.fetchall()]
        cur.execute("SELECT tenant_code, tenant_name FROM sys_tenant WHERE tenant_id=%s", (tid,))
        trow = cur.fetchone() or ("", "")
    return {"login": row[0], "name": row[1], "department": row[2],
            "userLevel": row[3], "status": row[4], "roles": roles,
            "tenantCode": trow[0], "tenantName": trow[1]}


@router.get("/auth/permissions")
def auth_permissions(request: Request) -> dict[str, str]:
    """유효 권한 매트릭스 — user_level 암묵 역할 + sys_user_role 할당 역할의 합집합 (WRITE 우선)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_level FROM sys_user WHERE tenant_id=%s AND user_id=%s",
                    (tid, request.state.user_id))
        level = cur.fetchone()[0]
        cur.execute(
            """SELECT p.resource_key, p.action
               FROM sys_role_permission p JOIN sys_role r ON r.role_id=p.role_id
               WHERE r.tenant_id=%s AND (r.role_name=%s OR r.role_id IN
                 (SELECT role_id FROM sys_user_role WHERE user_id=%s))""",
            (tid, level, request.state.user_id))
        out: dict[str, str] = {}
        for key, action in cur.fetchall():
            if out.get(key) != "WRITE":   # WRITE ⊃ READ
                out[key] = action
    return out


@router.get("/users/{login}/roles")
def user_roles(login: str) -> list[str]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.role_name FROM sys_user u
               JOIN sys_user_role ur ON ur.user_id=u.user_id
               JOIN sys_role r ON r.role_id=ur.role_id
               WHERE u.tenant_id=%s AND u.login_id=%s ORDER BY r.role_name""", (tid, login))
        return [x[0] for x in cur.fetchall()]


class RolesAssign(BaseModel):
    roles: list[str]


@router.put("/users/{login}/roles", dependencies=[SETUP])
def user_roles_assign(login: str, request: Request, body: RolesAssign) -> dict[str, Any]:
    """다중 역할 할당 — sys_user_role 전체 교체 (감사 ROLE_ASSIGN)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id FROM sys_user WHERE tenant_id=%s AND login_id=%s", (tid, login))
        u = cur.fetchone()
        if not u:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        role_ids = []
        for name in body.roles:
            cur.execute("SELECT role_id FROM sys_role WHERE tenant_id=%s AND role_name=%s",
                        (tid, name.strip().upper()))
            r = cur.fetchone()
            if not r:
                raise HTTPException(422, detail=f"역할 없음: {name}")
            role_ids.append(r[0])
        cur.execute("DELETE FROM sys_user_role WHERE user_id=%s", (u[0],))
        for rid in role_ids:
            cur.execute("INSERT INTO sys_user_role (user_id, role_id) VALUES (%s,%s)", (u[0], rid))
        _audit(cur, tid, "sys_user_role", u[0], "ROLE_ASSIGN", request.state.user_id,
               {"login": login, "roles": body.roles})
    return {"login": login, "roles": body.roles}


@router.post("/users/{login}/invite", dependencies=[SETUP])
def user_invite(login: str, request: Request) -> dict[str, Any]:
    """초대 안내 — 메일 서버 미설정 환경: 인앱 알림 + 감사 (정직한 범위)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT user_id, user_name FROM sys_user WHERE tenant_id=%s AND login_id=%s",
                    (tid, login))
        u = cur.fetchone()
        if not u:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        _notify(cur, tid, u[0], "INVITE",
                f"EDIM 초대 — {u[1]}님, 시스템 사용을 시작하십시오 (edim.seekerslab.com)", "/cpq")
        _audit(cur, tid, "sys_user", u[0], "INVITE", request.state.user_id, {"login": login})
    return {"login": login, "channel": "IN_APP", "note": "메일 서버 미설정 — 인앱 알림 발송"}


class ActivePatch(BaseModel):
    active: bool


@router.patch("/users/{login}/active", dependencies=[SETUP])
def user_active(login: str, request: Request, body: ActivePatch) -> dict[str, Any]:
    """계정 비활성화/재활성 — DISABLED 는 로그인 거부 (본인 비활성화 불가)."""
    if login == request.state.login and not body.active:
        raise HTTPException(422, detail="본인 계정은 비활성화할 수 없습니다")
    new_status = "ACTIVE" if body.active else "DISABLED"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_user SET status=%s, updated_at=now()
               WHERE tenant_id=%s AND login_id=%s RETURNING user_id""",
            (new_status, tid, login))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"사용자 없음: {login}")
        _audit(cur, tid, "sys_user", row[0],
               "REACTIVATE" if body.active else "DEACTIVATE",
               request.state.user_id, {"login": login, "status": new_status})
    return {"login": login, "status": new_status}


class HierarchyNodeCreate(BaseModel):
    treeType: str = "PRODUCT"
    name: str
    symbol: str = ""
    address: str
    parentAddress: str = ""


@router.post("/hierarchy/nodes", status_code=201, dependencies=[SETUP])
def hierarchy_node_create(request: Request, body: HierarchyNodeCreate) -> dict[str, Any]:
    """Hierarchy 노드 등록 — 주소 유일(409), 상위 주소 검증 (M-3-1 편집 개방)."""
    name, addr = body.name.strip(), body.address.strip()
    if not name or not addr:
        raise HTTPException(422, detail="이름·주소는 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM sys_hierarchy WHERE tenant_id=%s AND address=%s", (tid, addr))
        if cur.fetchone():
            raise HTTPException(409, detail=f"주소 중복: {addr}")
        parent_id = None
        if body.parentAddress.strip():
            cur.execute("SELECT hierarchy_id FROM sys_hierarchy WHERE tenant_id=%s AND address=%s",
                        (tid, body.parentAddress.strip()))
            p = cur.fetchone()
            if not p:
                raise HTTPException(422, detail=f"상위 주소 없음: {body.parentAddress}")
            parent_id = p[0]
            if not addr.startswith(body.parentAddress.strip() + "/"):
                raise HTTPException(422, detail="주소는 상위 주소로 시작해야 합니다 (예: /M/ENG/새노드)")
        cur.execute(
            """INSERT INTO sys_hierarchy (tenant_id, parent_id, tree_type, node_name, symbol,
               address, approval_status) VALUES (%s,%s,%s,%s,NULLIF(%s,''),%s,'DRAFT')
               RETURNING hierarchy_id""",
            (tid, parent_id, body.treeType.strip().upper(), name[:100], body.symbol.strip()[:30], addr[:500]))
        hid = cur.fetchone()[0]
        _audit(cur, tid, "sys_hierarchy", hid, "CREATE", request.state.user_id,
               {"address": addr, "name": name})
    return {"hierarchyId": hid, "address": addr}


class HierarchyNodePatch(BaseModel):
    name: str = ""
    symbol: str = ""
    remark: str | None = None      # 트리아지 #22 — None=유지, ''=삭제
    color: str | None = None       # None=유지, ''=삭제
    locked: bool | None = None     # None=유지


@router.patch("/hierarchy/nodes/{node_id}", dependencies=[SETUP])
def hierarchy_node_patch(node_id: int, request: Request, body: HierarchyNodePatch) -> dict[str, Any]:
    """노드 속성 수정 (트리아지 #22) — 잠금 노드는 해제 요청 없이는 다른 필드 수정 409."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT is_locked FROM sys_hierarchy WHERE tenant_id=%s AND hierarchy_id=%s",
                    (tid, node_id))
        cur_row = cur.fetchone()
        if not cur_row:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
        if cur_row[0] and body.locked is not False and (
                body.name.strip() or body.symbol.strip()
                or body.remark is not None or body.color is not None):
            raise HTTPException(409, detail="잠금 노드 — 잠금 해제 후 수정하십시오 (🔒)")
        cur.execute(
            """UPDATE sys_hierarchy SET
               node_name=CASE WHEN %s<>'' THEN %s ELSE node_name END,
               symbol=CASE WHEN %s<>'' THEN %s ELSE symbol END,
               remark=CASE WHEN %s THEN NULLIF(%s,'') ELSE remark END,
               color=CASE WHEN %s THEN NULLIF(%s,'') ELSE color END,
               is_locked=COALESCE(%s, is_locked),
               updated_at=now()
               WHERE tenant_id=%s AND hierarchy_id=%s RETURNING address""",
            (body.name.strip(), body.name.strip()[:100],
             body.symbol.strip(), body.symbol.strip()[:30],
             body.remark is not None, (body.remark or "").strip()[:200],
             body.color is not None, (body.color or "").strip()[:16],
             body.locked, tid, node_id))
        row = cur.fetchone()
        _audit(cur, tid, "sys_hierarchy", node_id, "UPDATE", request.state.user_id,
               {"name": body.name, "symbol": body.symbol,
                "remark": body.remark, "color": body.color, "locked": body.locked})
    return {"hierarchyId": node_id, "address": row[0]}


# ── U18 Hierarchy 편집 심화 (슬라이드 64) — 노드 이동(주소 연쇄 재계산)·정보 ──

class HierarchyMove(BaseModel):
    targetParentId: int | None = None   # None = 루트로 이동


# 트리아지 #24/#25 — hierarchy_address 를 참조하는 자산 테이블 (영향 분석·이동 연쇄 갱신·삭제 가드)
_H_REF_TABLES: list[tuple[str, str, str]] = [
    ("code_group", "group_code", "코드 그룹"),
    ("product_code", "main_code", "제품 코드"),
    ("tbx_macro", "macro_name", "매크로"),
    ("tbl_data_table", "table_name", "데이터 테이블"),
    ("tbx_ui_form", "form_name", "UI 폼"),
]


def _hierarchy_refs(cur, tid: int, addr: str) -> list[dict[str, Any]]:
    """주소(정확 일치 또는 하위 — '/'·'.' 구분자 모두)를 참조하는 자산 집계."""
    out: list[dict[str, Any]] = []
    for tbl, code_col, label in _H_REF_TABLES:
        cur.execute(
            f"""SELECT {code_col} FROM {tbl}
                WHERE tenant_id=%s AND (hierarchy_address=%s
                      OR hierarchy_address LIKE %s OR hierarchy_address LIKE %s)
                ORDER BY {code_col}""",
            (tid, addr, addr + "/%", addr + ".%"))
        rows = [r[0] for r in cur.fetchall()]
        if rows:
            out.append({"table": tbl, "label": label, "count": len(rows), "samples": rows[:5]})
    return out


@router.get("/hierarchy/nodes/{node_id}/impact")
def hierarchy_node_impact(node_id: int) -> dict[str, Any]:
    """이동·삭제 전 영향 분석 (트리아지 #25) — 하위 노드 수 + 주소 참조 자산 5테이블 집계."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT address, node_name, tree_type FROM sys_hierarchy
               WHERE tenant_id=%s AND hierarchy_id=%s""", (tid, node_id))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
        addr, name, tree = r
        cur.execute(
            """SELECT count(*) FROM sys_hierarchy
               WHERE tenant_id=%s AND tree_type=%s AND (address LIKE %s OR address LIKE %s)""",
            (tid, tree, addr + "/%", addr + ".%"))
        desc = int(cur.fetchone()[0])
        refs = _hierarchy_refs(cur, tid, addr)
    return {"nodeId": node_id, "address": addr, "name": name, "descendants": desc,
            "referencingTotal": sum(x["count"] for x in refs), "references": refs}


@router.get("/hierarchy/validate")
def hierarchy_validate(tree: str = "PRODUCT") -> dict[str, Any]:
    """U22 — 저장 전 정합 점검 (슬라이드 57-⑧): 주소 중복·고아 노드·부모 주소 불일치·루트 형식."""
    issues: list[dict[str, Any]] = []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT hierarchy_id, parent_id, node_name, address FROM sys_hierarchy
               WHERE tenant_id=%s AND tree_type=%s ORDER BY address""", (tid, tree))
        rows = cur.fetchall()
        by_id = {r[0]: r for r in rows}
        seen: dict[str, int] = {}

        # 주소 구분자(. 또는 /) 혼용 대응 — 세그먼트 기준 비교
        def _segs(a: str) -> list[str]:
            return [x for x in re.split(r"[./]", a) if x]

        for hid, parent_id, name, addr in rows:
            if addr in seen:
                issues.append({"type": "DUP_ADDRESS", "nodeId": hid, "name": name, "address": addr,
                               "detail": f"주소 중복 — #{seen[addr]} 와 동일"})
            else:
                seen[addr] = hid
            if parent_id is not None and parent_id not in by_id:
                issues.append({"type": "ORPHAN", "nodeId": hid, "name": name, "address": addr,
                               "detail": f"고아 노드 — 부모 #{parent_id} 없음"})
            elif parent_id is not None:
                paddr = by_id[parent_id][3]
                ps, cs = _segs(paddr), _segs(addr)
                if not (len(cs) == len(ps) + 1 and cs[: len(ps)] == ps):
                    issues.append({"type": "ADDR_MISMATCH", "nodeId": hid, "name": name, "address": addr,
                                   "detail": f"부모 주소 불일치 — 부모 {paddr} 의 한 단계 하위 아님"})
            if parent_id is None and len(_segs(addr)) > 1:
                issues.append({"type": "ROOT_FORMAT", "nodeId": hid, "name": name, "address": addr,
                               "detail": "루트 노드 주소가 다중 세그먼트"})
        # 0.8 — 고아 자산 주소: 어떤 트리 노드 프리픽스에도 안 걸리는 hierarchy_address
        # (영향 분석·이동 연쇄 갱신 대상에서 누락되는 참조 — PRODUCT 점검 시 1회 보고)
        if tree == "PRODUCT":
            for tbl, code_col, label in _H_REF_TABLES:
                cur.execute(
                    f"""SELECT {code_col}, hierarchy_address FROM {tbl}
                        WHERE tenant_id=%s AND COALESCE(hierarchy_address,'')<>''
                          AND NOT EXISTS (SELECT 1 FROM sys_hierarchy h
                                          WHERE h.tenant_id=%s AND (hierarchy_address=h.address
                                                OR hierarchy_address LIKE h.address || '/%%'))
                        ORDER BY {code_col}""", (tid, tid))
                for code, addr in cur.fetchall():
                    issues.append({"type": "ORPHAN_ASSET", "nodeId": 0, "name": f"{label} {code}",
                                   "address": addr,
                                   "detail": "자산 주소가 어떤 트리 노드에도 속하지 않음 — 영향 분석·이동 연쇄 갱신 누락"})
    return {"tree": tree, "nodes": len(rows), "ok": not issues, "issues": issues}


@router.get("/hierarchy/nodes/{node_id}/info")
def hierarchy_node_info(node_id: int) -> dict[str, Any]:
    """노드 속성/정보 (U18) — 주소·심볼·상태·하위 수·이력 메타."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT h.address, h.node_name, COALESCE(h.symbol,''), h.tree_type,
                      h.approval_status, h.is_system, COALESCE(h.remarks,''),
                      h.created_by, to_char(h.created_at,'YYYY-MM-DD HH24:MI'),
                      to_char(h.updated_at,'YYYY-MM-DD HH24:MI'),
                      (SELECT count(*) FROM sys_hierarchy c
                       WHERE c.tenant_id=h.tenant_id AND c.tree_type=h.tree_type
                         AND c.address LIKE h.address || '.%%') AS descendants
               FROM sys_hierarchy h WHERE h.tenant_id=%s AND h.hierarchy_id=%s""",
            (tid, node_id))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
    return {"address": r[0], "name": r[1], "symbol": r[2], "treeType": r[3],
            "status": r[4], "isSystem": r[5], "remarks": r[6],
            "createdBy": r[7], "createdAt": r[8], "updatedAt": r[9], "descendants": int(r[10])}


@router.post("/hierarchy/nodes/{node_id}/move", dependencies=[SETUP])
def hierarchy_node_move(node_id: int, request: Request, body: HierarchyMove) -> dict[str, Any]:
    """노드 이동 (U18) — 대상 부모 하위로 재배치, 본인+하위 전체 주소 접두 연쇄 재계산.

    가드: 시스템 노드 불가·자기 자신/자기 하위로 이동 불가·트리 불일치 422.
    """
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT address, tree_type, is_system, is_locked FROM sys_hierarchy
               WHERE tenant_id=%s AND hierarchy_id=%s""", (tid, node_id))
        src = cur.fetchone()
        if not src:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
        if src[2]:
            raise HTTPException(422, detail="시스템 제공 노드는 이동할 수 없습니다")
        if src[3]:
            raise HTTPException(409, detail="잠금 노드는 이동 불가 — 잠금 해제 후 진행 (🔒)")
        old_addr, tree = src[0], src[1]
        if body.targetParentId is not None:
            cur.execute(
                """SELECT address, tree_type FROM sys_hierarchy
                   WHERE tenant_id=%s AND hierarchy_id=%s""", (tid, body.targetParentId))
            tgt = cur.fetchone()
            if not tgt:
                raise HTTPException(404, detail=f"대상 노드 없음: #{body.targetParentId}")
            if tgt[1] != tree:
                raise HTTPException(422, detail="다른 트리로는 이동할 수 없습니다")
            if tgt[0] == old_addr or tgt[0].startswith(old_addr + "."):
                raise HTTPException(422, detail="자기 자신 또는 하위로는 이동할 수 없습니다")
            parent_addr = tgt[0]
        else:
            parent_addr = ""
        # 새 주소 = 대상 부모 하위 다음 순번 (형제 마지막 세그먼트 최대+1 — 비수치 세그먼트는 무시)
        if parent_addr:
            cur.execute(
                """SELECT address FROM sys_hierarchy
                   WHERE tenant_id=%s AND tree_type=%s AND address LIKE %s AND address NOT LIKE %s""",
                (tid, tree, parent_addr + ".%", parent_addr + ".%.%"))
            segs = [a[0][len(parent_addr) + 1:] for a in cur.fetchall()]
        else:
            cur.execute(
                "SELECT address FROM sys_hierarchy WHERE tenant_id=%s AND tree_type=%s AND address NOT LIKE '%%.%%'",
                (tid, tree))
            segs = [a[0] for a in cur.fetchall()]
        next_seq = max((int(x) for x in segs if x.isdigit()), default=0) + 1
        new_addr = f"{parent_addr}.{next_seq}" if parent_addr else str(next_seq)
        # 본인 + 하위 연쇄 접두 치환
        cur.execute(
            """UPDATE sys_hierarchy
               SET address = %s || substring(address from %s), updated_at = now()
               WHERE tenant_id=%s AND tree_type=%s
                 AND (address = %s OR address LIKE %s)""",
            (new_addr, len(old_addr) + 1, tid, tree, old_addr, old_addr + ".%"))
        moved = cur.rowcount
        cur.execute(
            "UPDATE sys_hierarchy SET parent_id=%s WHERE tenant_id=%s AND hierarchy_id=%s",
            (body.targetParentId, tid, node_id))
        # 트리아지 #24 — 주소 참조 자산 연쇄 갱신 (이동 후에도 연결 유지)
        relinked = 0
        for tbl, _code_col, _label in _H_REF_TABLES:
            cur.execute(
                f"""UPDATE {tbl}
                    SET hierarchy_address = %s || substring(hierarchy_address from %s)
                    WHERE tenant_id=%s AND (hierarchy_address=%s
                          OR hierarchy_address LIKE %s OR hierarchy_address LIKE %s)""",
                (new_addr, len(old_addr) + 1, tid, old_addr, old_addr + "/%", old_addr + ".%"))
            relinked += cur.rowcount
        _audit(cur, tid, "sys_hierarchy", node_id, "NODE_MOVE", request.state.user_id,
               after={"from": old_addr, "to": new_addr, "moved": moved, "relinked": relinked})
    return {"hierarchyId": node_id, "newAddress": new_addr, "moved": moved, "relinked": relinked}


@router.delete("/hierarchy/nodes/{node_id}", dependencies=[SETUP])
def hierarchy_node_delete(node_id: int, request: Request) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT is_locked FROM sys_hierarchy WHERE tenant_id=%s AND hierarchy_id=%s",
                    (tid, node_id))
        lk = cur.fetchone()
        if lk and lk[0]:
            raise HTTPException(409, detail="잠금 노드는 삭제 불가 — 잠금 해제 후 진행 (🔒)")
        cur.execute("SELECT 1 FROM sys_hierarchy WHERE parent_id=%s LIMIT 1", (node_id,))
        if cur.fetchone():
            raise HTTPException(409, detail="하위 노드가 있는 노드는 삭제 불가")
        # 트리아지 #25 — 주소 참조 자산이 있으면 삭제 차단 (영향 분석 확인 유도)
        cur.execute("SELECT address FROM sys_hierarchy WHERE tenant_id=%s AND hierarchy_id=%s",
                    (tid, node_id))
        arow = cur.fetchone()
        if not arow:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
        refs = _hierarchy_refs(cur, tid, arow[0])
        if refs:
            summary = " · ".join(f"{x['label']} {x['count']}건" for x in refs)
            raise HTTPException(409, detail=f"참조 자산이 있는 노드는 삭제 불가 — {summary} (영향 분석으로 확인)")
        cur.execute(
            "DELETE FROM sys_hierarchy WHERE tenant_id=%s AND hierarchy_id=%s RETURNING address",
            (tid, node_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"노드 없음: #{node_id}")
        _audit(cur, tid, "sys_hierarchy", node_id, "DELETE", request.state.user_id,
               {"address": row[0]})
    return {"deleted": node_id}


class AllocateCode(BaseModel):
    docType: str = "DOC"
    dept: str = ""


_DOC_NO_DEFAULT_TEMPLATE = "{TYPE}-{YYYY}-{SEQ:4}"


def _doc_numbering_rule(cur) -> dict[str, str]:
    """U23 — 테넌트 채번 규칙 (sys_tenant.settings.docNumbering, 부재 시 기본)."""
    tid = _tenant_id(cur)
    cur.execute("SELECT COALESCE(settings,'{}'::jsonb) FROM sys_tenant WHERE tenant_id=%s", (tid,))
    st = cur.fetchone()[0] or {}
    rule = st.get("docNumbering") if isinstance(st, dict) else None
    rule = rule if isinstance(rule, dict) else {}
    return {"template": str(rule.get("template") or _DOC_NO_DEFAULT_TEMPLATE),
            "dept": str(rule.get("dept") or "HQ")}


def _render_doc_no(template: str, *, dept: str, doc_type: str, seq: int) -> str:
    m = re.search(r"\{SEQ:?(\d*)\}", template)
    width = int(m.group(1) or 4) if m and m.group(1) else 4
    yyyy = date.today().strftime("%Y")
    out = (template.replace("{DEPT}", dept).replace("{TYPE}", doc_type)
           .replace("{YYYY}", yyyy).replace("{YY}", yyyy[2:]))
    return re.sub(r"\{SEQ:?\d*\}", str(seq).zfill(width), out)


@router.post("/documents/allocate-code", dependencies=[SETUP])
def document_allocate_code(body: AllocateCode) -> dict[str, Any]:
    """U23 — 규칙 기반 문서 채번 (슬라이드 53 Document Code): settings.docNumbering.template,
    토큰 {DEPT}·{TYPE}·{YYYY}·{YY}·{SEQ:n}. 유형별 순번, 중복 회피."""
    dt = body.docType.strip().upper()[:10] or "DOC"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        rule = _doc_numbering_rule(cur)
        dept = (body.dept.strip() or rule["dept"]).upper()[:10]
        cur.execute("SELECT count(*) FROM doc_control WHERE tenant_id=%s AND doc_type=%s", (tid, dt))
        seq = cur.fetchone()[0] + 1
        while True:
            candidate = _render_doc_no(rule["template"], dept=dept, doc_type=dt, seq=seq)
            cur.execute("SELECT 1 FROM doc_control WHERE tenant_id=%s AND doc_no=%s", (tid, candidate))
            if not cur.fetchone():
                break
            seq += 1
    return {"docNo": candidate, "docType": dt, "template": rule["template"], "dept": dept}


@router.get("/documents/numbering-rule")
def document_numbering_rule() -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        rule = _doc_numbering_rule(cur)
    return {**rule, "sample": _render_doc_no(rule["template"], dept=rule["dept"], doc_type="DOC", seq=1)}


class NumberingRulePut(BaseModel):
    template: str
    dept: str = "HQ"


@router.put("/documents/numbering-rule", dependencies=[SETUP])
def document_numbering_rule_put(request: Request, body: NumberingRulePut) -> dict[str, Any]:
    tpl = body.template.strip()
    if not re.search(r"\{SEQ:?\d*\}", tpl):
        raise HTTPException(422, detail="템플릿에 {SEQ} 토큰이 필요합니다 (예: {DEPT}-{TYPE}-{YYYY}-{SEQ:4})")
    bad = re.findall(r"\{(?!DEPT\}|TYPE\}|YYYY\}|YY\}|SEQ)[^}]*\}", tpl)
    if bad:
        raise HTTPException(422, detail=f"미지원 토큰: {', '.join(bad[:3])} (지원: DEPT/TYPE/YYYY/YY/SEQ:n)")
    dept = body.dept.strip().upper()[:10] or "HQ"
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE sys_tenant
               SET settings = COALESCE(settings,'{}'::jsonb)
                   || jsonb_build_object('docNumbering', jsonb_build_object('template', %s::text, 'dept', %s::text))
               WHERE tenant_id=%s""", (tpl[:80], dept, tid))
        _audit(cur, tid, "sys_tenant", tid, "DOC_NUMBERING_SET", request.state.user_id,
               {"template": tpl[:80], "dept": dept})
    return {"template": tpl[:80], "dept": dept,
            "sample": _render_doc_no(tpl, dept=dept, doc_type="DOC", seq=1)}


DOC_TRANSITIONS = {"SET_UP": ("CHECK",), "CHECK": ("APPROVE", "SET_UP"),
                   "APPROVE": ("ACCEPTED", "SET_UP"), "ACCEPTED": ()}


class DocStatusPatch(BaseModel):
    status: str


@router.patch("/documents/{doc_no}/status", dependencies=[SETUP])
def document_status(doc_no: str, request: Request, body: DocStatusPatch) -> dict[str, Any]:
    """문서 상태 전이 — SET_UP→CHECK→APPROVE→ACCEPTED (반려=SET_UP 복귀, 유효 전이만)."""
    new = body.status.strip().upper()
    if new not in DOC_TRANSITIONS:
        raise HTTPException(422, detail=f"상태 오류: {body.status} ({'/'.join(DOC_TRANSITIONS)})")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT doc_control_id, released_status FROM doc_control
               WHERE tenant_id=%s AND doc_no=%s ORDER BY doc_control_id DESC LIMIT 1""",
            (tid, doc_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"문서 없음: {doc_no}")
        if new not in DOC_TRANSITIONS[row[1]]:
            raise HTTPException(409, detail=f"전이 불가: {row[1]} → {new} (허용: {', '.join(DOC_TRANSITIONS[row[1]]) or '없음'})")
        cur.execute(
            """UPDATE doc_control SET released_status=%s,
               approval_date=CASE WHEN %s='ACCEPTED' THEN CURRENT_DATE ELSE approval_date END,
               approver_id=CASE WHEN %s='ACCEPTED' THEN %s ELSE approver_id END,
               updated_at=now() WHERE doc_control_id=%s""",
            (new, new, new, request.state.user_id, row[0]))
        _audit(cur, tid, "doc_control", row[0], "STATUS", request.state.user_id,
               {"docNo": doc_no, "from": row[1], "to": new})
    return {"docNo": doc_no, "status": new}


class RelationshipAdd(BaseModel):
    mother: str
    child: str
    qty: float = 1


@router.post("/codes/relationships", status_code=201, dependencies=[SETUP])
def relationship_add(request: Request, body: RelationshipAdd) -> dict[str, Any]:
    """Child 추가 (S-1-4) — DRAFT 등록, Running Test 통과 후 승인 (CODE-009)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        ids = {}
        for label, code in (("mother", body.mother.strip()), ("child", body.child.strip())):
            cur.execute(
                "SELECT min(product_code_id) FROM product_code WHERE tenant_id=%s AND main_code=%s",
                (tid, code))
            r = cur.fetchone()
            if not r or not r[0]:
                raise HTTPException(422, detail=f"코드 없음: {code}")
            ids[label] = r[0]
        if ids["mother"] == ids["child"]:
            raise HTTPException(422, detail="Mother 와 Child 가 같습니다")
        cur.execute(
            """SELECT 1 FROM code_relationship
               WHERE tenant_id=%s AND mother_code_id=%s AND child_code_id=%s""",
            (tid, ids["mother"], ids["child"]))
        if cur.fetchone():
            raise HTTPException(409, detail=f"이미 등록된 관계: {body.mother} → {body.child}")
        cur.execute(
            """INSERT INTO code_relationship (tenant_id, mother_code_id, child_code_id, quantity,
               approval_status, created_by) VALUES (%s,%s,%s,%s,'DRAFT',%s) RETURNING rel_id""",
            (tid, ids["mother"], ids["child"], body.qty, request.state.login))
        rel_id = cur.fetchone()[0]
        _audit(cur, tid, "code_relationship", rel_id, "CREATE", request.state.user_id,
               {"mother": body.mother, "child": body.child, "qty": body.qty})
    return {"relId": rel_id, "status": "DRAFT"}


@router.delete("/codes/relationships/{rel_id}", dependencies=[SETUP])
def relationship_delete(rel_id: int, request: Request) -> dict[str, Any]:
    """관계 삭제 — DRAFT 한정 (승인 관계 보호)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """DELETE FROM code_relationship
               WHERE tenant_id=%s AND rel_id=%s AND approval_status='DRAFT' RETURNING rel_id""",
            (tid, rel_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"DRAFT 관계 없음: #{rel_id}")
    return {"deleted": rel_id}


@router.get("/erp/projects/check-duplicate")
def project_check_duplicate(name: str = "", no: str = "") -> dict[str, Any]:
    """프로젝트 중복검토 (S-3-5) — 이름/번호 ILIKE 실질의."""
    if not name.strip() and not no.strip():
        raise HTTPException(422, detail="검토할 이름 또는 번호를 입력하십시오")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT project_no, project_name FROM prj_project
               WHERE tenant_id=%s AND (project_name ILIKE %s OR project_no ILIKE %s)
               LIMIT 10""",
            (tid, f"%{name.strip() or chr(1)}%", f"%{no.strip() or chr(1)}%"))
        matches = [{"no": r[0], "name": r[1]} for r in cur.fetchall()]
    return {"duplicate": bool(matches), "matches": matches}


# ── B19 창고·저장위치 계층 — erp_warehouse (ERP-020/021) ──

WAREHOUSE_TYPES = ("REGION", "PLANT", "WAREHOUSE", "STORAGE", "SECTOR")


@router.get("/erp/warehouses")
def warehouse_tree() -> list[dict[str, Any]]:
    """창고/저장위치 계층 — 재귀 CTE 로 경로 정렬 (depth 포함)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """WITH RECURSIVE t AS (
                 SELECT warehouse_id, parent_id, location_type, location_code, location_name,
                        hazard_allowed, inspection_cycle, remarks, 0 AS depth,
                        location_code::text AS path
                 FROM erp_warehouse WHERE tenant_id=%s AND parent_id IS NULL
                 UNION ALL
                 SELECT w.warehouse_id, w.parent_id, w.location_type, w.location_code,
                        w.location_name, w.hazard_allowed, w.inspection_cycle, w.remarks,
                        t.depth+1, t.path || '/' || w.location_code
                 FROM erp_warehouse w JOIN t ON w.parent_id=t.warehouse_id)
               SELECT warehouse_id, parent_id, location_type, location_code, location_name,
                      COALESCE(hazard_allowed,''), COALESCE(inspection_cycle,''),
                      COALESCE(remarks,''), depth, path
               FROM t ORDER BY path""", (tid,))
        return [
            {"warehouseId": r[0], "parentId": r[1], "type": r[2], "code": r[3], "name": r[4],
             "hazard": r[5], "inspection": r[6], "remarks": r[7], "depth": r[8], "path": r[9]}
            for r in cur.fetchall()
        ]


class WarehouseCreate(BaseModel):
    parentCode: str = ""         # 빈 값 = 최상위 (REGION)
    locationType: str
    code: str
    name: str
    hazard: str = ""
    inspection: str = ""
    remarks: str = ""


@router.post("/erp/warehouses", status_code=201, dependencies=[SETUP])
def warehouse_create(request: Request, body: WarehouseCreate) -> dict[str, Any]:
    lt = body.locationType.strip().upper()
    if lt not in WAREHOUSE_TYPES:
        raise HTTPException(422, detail=f"위치 유형 오류: {body.locationType} ({'|'.join(WAREHOUSE_TYPES)})")
    code, name = body.code.strip(), body.name.strip()
    if not code or not name:
        raise HTTPException(422, detail="위치 코드·이름은 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM erp_warehouse WHERE tenant_id=%s AND location_code=%s", (tid, code))
        if cur.fetchone():
            raise HTTPException(409, detail=f"위치 코드 중복: {code}")
        parent_id = None
        if body.parentCode.strip():
            cur.execute(
                """SELECT warehouse_id, location_type FROM erp_warehouse
                   WHERE tenant_id=%s AND location_code=%s""", (tid, body.parentCode.strip()))
            pr = cur.fetchone()
            if not pr:
                raise HTTPException(422, detail=f"상위 위치 없음: {body.parentCode}")
            # 계층 순서 강제 — 자식 유형은 부모 유형보다 하위여야 함
            if WAREHOUSE_TYPES.index(lt) <= WAREHOUSE_TYPES.index(pr[1]):
                raise HTTPException(422, detail=f"계층 오류: {pr[1]} 아래에 {lt} 불가 (REGION→PLANT→WAREHOUSE→STORAGE→SECTOR)")
            parent_id = pr[0]
        elif lt != "REGION":
            raise HTTPException(422, detail="최상위는 REGION 만 가능")
        cur.execute(
            """INSERT INTO erp_warehouse (tenant_id, parent_id, location_type, location_code,
               location_name, hazard_allowed, inspection_cycle, remarks, created_by)
               VALUES (%s,%s,%s,%s,%s,NULLIF(%s,''),NULLIF(%s,''),NULLIF(%s,''),%s)
               RETURNING warehouse_id""",
            (tid, parent_id, lt, code[:30], name[:100], body.hazard.strip()[:100],
             body.inspection.strip()[:30], body.remarks.strip()[:300], request.state.login))
        wid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'erp_warehouse',%s,'CREATE',%s,%s)""",
            (tid, wid, request.state.user_id, json.dumps({"code": code, "type": lt})))
    return {"warehouseId": wid, "code": code}


@router.delete("/erp/warehouses/{code}", dependencies=[SETUP])
def warehouse_delete(code: str, request: Request) -> dict[str, Any]:
    """위치 삭제 — 하위 위치 존재 시 409 보호."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT warehouse_id FROM erp_warehouse WHERE tenant_id=%s AND location_code=%s",
                    (tid, code))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"위치 없음: {code}")
        cur.execute("SELECT 1 FROM erp_warehouse WHERE parent_id=%s LIMIT 1", (row[0],))
        if cur.fetchone():
            raise HTTPException(409, detail=f"하위 위치가 있는 노드는 삭제 불가: {code}")
        cur.execute("DELETE FROM erp_warehouse WHERE warehouse_id=%s", (row[0],))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'erp_warehouse',%s,'DELETE',%s,%s)""",
            (tid, row[0], request.state.user_id, json.dumps({"code": code})))
    return {"deleted": code}


# ── B19 QCR 발행·PO 문서 등록 — 구매 상세 (ERP-017) ──

class QcrIssue(BaseModel):
    codes: list[str]
    note: str = ""


@router.post("/erp/qcr", status_code=201, dependencies=[SETUP])
def qcr_issue(request: Request, body: QcrIssue) -> dict[str, Any]:
    """견적 요청(QCR) 발행 — 감사 기록 + 구매 담당(SETUP+) 알림 (공급자 회신 대기)."""
    codes = [c.strip() for c in body.codes if c.strip()]
    if not codes:
        raise HTTPException(422, detail="발행 대상 품목이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*)+1 FROM sys_history WHERE tenant_id=%s AND action='QCR_ISSUE'",
                    (tid,))
        qcr_no = f"QCR-{cur.fetchone()[0]:04d}"
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'erp_qcr',0,'QCR_ISSUE',%s,%s) RETURNING history_id""",
            (tid, request.state.user_id,
             json.dumps({"qcrNo": qcr_no, "codes": codes, "note": body.note.strip()[:300]})))
        cur.execute(
            """SELECT user_id FROM sys_user
               WHERE tenant_id=%s AND user_level IN ('SETUP','ADMIN') AND user_id<>%s""",
            (tid, request.state.user_id))
        for (uid,) in cur.fetchall():
            _notify(cur, tid, uid, "QCR_ISSUE",
                    f"견적 요청 발행 — {qcr_no} ({len(codes)}품목, 공급자 회신 대기)", "/erp")
    return {"qcrNo": qcr_no, "codes": len(codes)}


class PoCreate(BaseModel):
    codes: list[str]
    totalK: float = 0
    deliveryTerms: str = "EXW 창원공장"      # ERP-017 납품조건
    transport: str = "육로 (트럭)"           # 운송수단
    minOrderQty: int = 1                     # 최소구매수량
    certRequired: bool = False               # 인증서 요구


@router.post("/erp/po", status_code=201, dependencies=[SETUP])
def po_create(request: Request, body: PoCreate) -> dict[str, Any]:
    """발주 생성 — PO 를 doc_control 문서로 영속 (구매 조건 remarks·공급자 코드 라인 포함)."""
    codes = [c.strip() for c in body.codes if c.strip()]
    if not codes:
        raise HTTPException(422, detail="발주 품목이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT count(*)+1 FROM doc_control WHERE tenant_id=%s AND doc_type='PO'", (tid,))
        seq = cur.fetchone()[0]
        po_no = f"PO-61313-{seq}"
        # 공급자 코드 매핑 병기 (ERP-018 — 발주 문서에 공급자 코드 표시)
        sup_lines: list[str] = []
        for c in codes:
            cur.execute(
                """SELECT m.supplier_code, co.company_name
                   FROM prt_supplier_code_map m
                   JOIN com_company co ON co.company_id=m.supplier_id
                   LEFT JOIN prt_part p ON p.part_id=m.part_id
                   LEFT JOIN product_code pc
                     ON pc.product_code_id=COALESCE(m.product_code_id, p.product_code_id)
                   WHERE m.tenant_id=%s AND pc.main_code=%s LIMIT 1""", (tid, c))
            m = cur.fetchone()
            sup_lines.append(f"{c}↔{m[0]}({m[1]})" if m else c)
        terms = (f"납품:{body.deliveryTerms.strip()[:40]} · 운송:{body.transport.strip()[:30]}"
                 f" · 최소수량:{body.minOrderQty} · 인증서:{'요구' if body.certRequired else '불요'}"
                 f" · 품목:{', '.join(sup_lines)}")
        cur.execute(
            """INSERT INTO doc_control (tenant_id, doc_no, title, doc_type, released_status,
               version, person, management_grade, remarks, created_by)
               VALUES (%s,%s,%s,'PO','SET_UP','v1.0',%s,'S-3',%s,%s) RETURNING doc_control_id""",
            (tid, po_no, f"발주서 {po_no} — {len(codes)}품목 {body.totalK:,.0f}K",
             request.state.login, terms[:500], request.state.login))
        doc_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'doc_control',%s,'PO_CREATE',%s,%s)""",
            (tid, doc_id, request.state.user_id,
             json.dumps({"poNo": po_no, "codes": codes, "terms": terms[:300]})))
    return {"poNo": po_no, "docNo": po_no, "terms": terms}


# ── G3 발주 라이프사이클 (erp_po) — 헤더+라인·승인·입고(GR)·3-way 수량 match ──
PO_STATUS_LABEL = {"DRAFT": "작성", "APPROVED": "승인", "RECEIVING": "입고중",
                   "CLOSED": "완료", "CANCELLED": "취소"}


class PoLineIn(BaseModel):
    itemCode: str = ""
    itemName: str
    qty: float = 1
    unitPrice: float = 0


class PoLifecycleCreate(BaseModel):
    supplier: str = ""
    expectedDate: str = ""
    note: str = ""
    items: list[PoLineIn] = []


@router.post("/erp/pos", status_code=201, dependencies=[SETUP])
def po_lc_create(request: Request, body: PoLifecycleCreate) -> dict[str, Any]:
    """구조화 발주 생성 (G3) — 헤더 + 라인아이템, DRAFT."""
    items = [i for i in body.items if i.itemName.strip() and i.qty > 0]
    if not items:
        raise HTTPException(422, detail="발주 라인아이템이 없습니다 (품명·수량>0)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM erp_po WHERE tenant_id=%s", (tid,))
        seq = cur.fetchone()[0] + 1
        while True:
            po_no = f"PO-{seq:05d}"
            cur.execute("SELECT 1 FROM erp_po WHERE tenant_id=%s AND po_no=%s", (tid, po_no))
            if not cur.fetchone():
                break
            seq += 1
        cur.execute(
            """INSERT INTO erp_po (tenant_id, po_no, supplier, expected_date, note, created_by)
               VALUES (%s,%s,%s,%s,%s,%s) RETURNING po_id""",
            (tid, po_no, body.supplier.strip()[:200] or None,
             body.expectedDate.strip() or None, body.note.strip()[:500] or None,
             str(request.state.user_id)))
        po_id = cur.fetchone()[0]
        for n, it in enumerate(items):
            cur.execute(
                """INSERT INTO erp_po_item (po_id, item_code, item_name, order_qty, unit_price, sort_order)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (po_id, it.itemCode.strip()[:50] or None, it.itemName.strip()[:200],
                 it.qty, it.unitPrice, n))
        _audit(cur, tid, "erp_po", po_id, "PO_LC_CREATE", request.state.user_id,
               {"poNo": po_no, "lines": len(items)})
    return {"poNo": po_no, "status": "DRAFT", "lines": len(items)}


@router.get("/erp/pos")
def po_lc_list(status: str = "") -> list[dict[str, Any]]:
    """발주 목록 (G3) — 총액·입고 진척·3-way match 상태."""
    clause, params = "", []
    st = status.strip().upper()
    if st in PO_STATUS_LABEL:
        clause = " AND p.status=%s"
        params.append(st)
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT p.po_no, COALESCE(p.supplier,''), p.status,
                       to_char(p.order_date,'YYYY-MM-DD'), to_char(p.expected_date,'YYYY-MM-DD'),
                       count(i.po_item_id), COALESCE(sum(i.order_qty*i.unit_price),0),
                       COALESCE(sum(i.order_qty),0), COALESCE(sum(i.received_qty),0)
                FROM erp_po p LEFT JOIN erp_po_item i ON i.po_id=p.po_id
                WHERE p.tenant_id=%s{clause}
                GROUP BY p.po_id ORDER BY p.po_id DESC""", (tid, *params))
        rows = []
        for r in cur.fetchall():
            oq, rq = float(r[7]), float(r[8])
            rows.append({"poNo": r[0], "supplier": r[1], "status": r[2],
                         "statusLabel": PO_STATUS_LABEL.get(r[2], r[2]),
                         "orderDate": r[3], "expectedDate": r[4], "lines": r[5],
                         "amount": float(r[6]), "orderQty": oq, "receivedQty": rq,
                         "progress": round(rq / oq * 100) if oq else 0,
                         "matched": oq > 0 and rq == oq})
        return rows


@router.get("/erp/pos/{po_no}")
def po_lc_detail(po_no: str) -> dict[str, Any]:
    """발주 상세 + 라인 (G3)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT po_id, po_no, COALESCE(supplier,''), status, to_char(order_date,'YYYY-MM-DD'),
                      to_char(expected_date,'YYYY-MM-DD'), COALESCE(note,'')
               FROM erp_po WHERE tenant_id=%s AND po_no=%s""", (tid, po_no))
        h = cur.fetchone()
        if not h:
            raise HTTPException(404, detail=f"발주 없음: {po_no}")
        cur.execute(
            """SELECT po_item_id, COALESCE(item_code,''), item_name, order_qty, unit_price, received_qty
               FROM erp_po_item WHERE po_id=%s ORDER BY sort_order, po_item_id""", (h[0],))
        items = [{"poItemId": r[0], "itemCode": r[1], "itemName": r[2], "orderQty": float(r[3]),
                  "unitPrice": float(r[4]), "receivedQty": float(r[5]),
                  "remaining": float(r[3]) - float(r[5])} for r in cur.fetchall()]
        return {"poNo": h[1], "supplier": h[2], "status": h[3], "statusLabel": PO_STATUS_LABEL.get(h[3], h[3]),
                "orderDate": h[4], "expectedDate": h[5], "note": h[6], "items": items}


@router.patch("/erp/pos/{po_no}/approve", dependencies=[SETUP])
def po_lc_approve(po_no: str, request: Request) -> dict[str, Any]:
    """발주 승인 (G3) — DRAFT→APPROVED (입고 가능)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT po_id, status FROM erp_po WHERE tenant_id=%s AND po_no=%s", (tid, po_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"발주 없음: {po_no}")
        if row[1] != "DRAFT":
            raise HTTPException(409, detail=f"승인 불가 — 현재 {PO_STATUS_LABEL.get(row[1], row[1])}")
        cur.execute("UPDATE erp_po SET status='APPROVED', approved_at=now() WHERE po_id=%s", (row[0],))
        _audit(cur, tid, "erp_po", row[0], "PO_APPROVE", request.state.user_id, {"poNo": po_no})
    return {"poNo": po_no, "status": "APPROVED"}


class PoReceiveLine(BaseModel):
    poItemId: int
    qty: float


class PoReceive(BaseModel):
    items: list[PoReceiveLine] = []


@router.post("/erp/pos/{po_no}/receive", dependencies=[SETUP])
def po_lc_receive(po_no: str, request: Request, body: PoReceive) -> dict[str, Any]:
    """입고(GR) 처리 (G3) — 라인별 received_qty += qty, 발주 초과 409(3-way 수량 match).
    승인/입고중 상태만. 전 라인 입고 완료 시 CLOSED. inv_movement(IN) 기록."""
    if not body.items:
        raise HTTPException(422, detail="입고 라인이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT po_id, status FROM erp_po WHERE tenant_id=%s AND po_no=%s", (tid, po_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"발주 없음: {po_no}")
        po_id, status = row
        if status not in ("APPROVED", "RECEIVING"):
            raise HTTPException(409, detail=f"입고 불가 — 승인 후 가능 (현재 {PO_STATUS_LABEL.get(status, status)})")
        received = 0
        for line in body.items:
            if line.qty <= 0:
                continue
            cur.execute(
                "SELECT item_code, item_name, order_qty, received_qty FROM erp_po_item "
                "WHERE po_item_id=%s AND po_id=%s", (line.poItemId, po_id))
            it = cur.fetchone()
            if not it:
                raise HTTPException(404, detail=f"발주 라인 없음: {line.poItemId}")
            code, name, oq, rq = it[0], it[1], float(it[2]), float(it[3])
            if rq + line.qty > oq + 1e-6:
                raise HTTPException(409, detail=f"입고 초과 — {name}: 발주 {oq}, 기입고 {rq}, 요청 {line.qty} (3-way 불일치)")
            cur.execute("UPDATE erp_po_item SET received_qty=received_qty+%s WHERE po_item_id=%s",
                        (line.qty, line.poItemId))
            cur.execute(
                """INSERT INTO inv_movement (tenant_id, item_code, location_code, movement_type,
                   quantity, ref_type, ref_no)
                   VALUES (%s,%s,%s,'IN',%s,'PO',%s)""",
                (tid, code or name[:50], 'WH-RECV', line.qty, po_no))
            received += 1
        # 상태 전이 — 전 라인 입고 완료면 CLOSED, 아니면 RECEIVING
        cur.execute(
            "SELECT COALESCE(sum(order_qty),0), COALESCE(sum(received_qty),0) FROM erp_po_item WHERE po_id=%s",
            (po_id,))
        oq, rq = cur.fetchone()
        new_status = "CLOSED" if float(rq) >= float(oq) and float(oq) > 0 else "RECEIVING"
        cur.execute("UPDATE erp_po SET status=%s WHERE po_id=%s", (new_status, po_id))
        _audit(cur, tid, "erp_po", po_id, "PO_RECEIVE", request.state.user_id,
               {"poNo": po_no, "lines": received, "status": new_status})
    return {"poNo": po_no, "received": received, "status": new_status}


# ── B18 수익성(PCR)·견적 lifecycle — cst_pcr·cst_quotation ──

PCR_BUSINESS_TYPES = ("PRE_SALES", "MAIN")


def _latest_cost_base(cur, tid: int, project_no: str = "") -> tuple[int, dict[str, float]]:
    """최근 SUCCESS Run 의 cst_calc 합계 → (run_id, {MATERIAL, MANUFACTURING, DIRECT}).

    project_no 지정 시 해당 프로젝트(run→selection→project) 스코프의 최근 Run 만."""
    if project_no.strip():
        cur.execute(
            """SELECT c.run_id, c.calc_type, c.total_amount
               FROM cst_calc c
               JOIN cpq_run r ON r.run_id=c.run_id AND r.status='SUCCESS'
               JOIN cpq_selection s ON s.selection_id=r.selection_id
               JOIN prj_project p ON p.project_id=s.project_id
               WHERE c.tenant_id=%s AND p.project_no=%s AND c.run_id=(
                 SELECT max(c2.run_id) FROM cst_calc c2
                 JOIN cpq_run r2 ON r2.run_id=c2.run_id AND r2.status='SUCCESS'
                 JOIN cpq_selection s2 ON s2.selection_id=r2.selection_id
                 JOIN prj_project p2 ON p2.project_id=s2.project_id
                 WHERE c2.tenant_id=%s AND p2.project_no=%s)""",
            (tid, project_no.strip(), tid, project_no.strip()))
    else:
        cur.execute(
            """SELECT c.run_id, c.calc_type, c.total_amount
               FROM cst_calc c JOIN cpq_run r ON r.run_id=c.run_id AND r.status='SUCCESS'
               WHERE c.tenant_id=%s AND c.run_id=(
                 SELECT max(c2.run_id) FROM cst_calc c2
                 JOIN cpq_run r2 ON r2.run_id=c2.run_id AND r2.status='SUCCESS'
                 WHERE c2.tenant_id=%s)""", (tid, tid))
    rows = cur.fetchall()
    if not rows:
        detail = (f"프로젝트 {project_no} 의 원가 상세 SUCCESS Run 이 없습니다"
                  if project_no.strip() else "원가 상세가 있는 SUCCESS Run 이 없습니다 — EDIM Run 먼저 실행")
        raise HTTPException(409, detail=detail)
    totals = {r[1]: float(r[2]) for r in rows}
    return rows[0][0], totals


class PcrCreate(BaseModel):
    businessType: str = "PRE_SALES"
    marginRate: float = 0.35     # 목표 마진율 — 매출 = 직접비 × (1+rate)


@router.post("/cost/pcr", dependencies=[SETUP])
def pcr_upsert(request: Request, body: PcrCreate) -> dict[str, Any]:
    """PCR 수익성 보고서 — 최근 Run 원가 기반 산출 (UNIQUE selection×business_type upsert)."""
    bt = body.businessType.strip().upper()
    if bt not in PCR_BUSINESS_TYPES:
        raise HTTPException(422, detail=f"사업유형 오류: {body.businessType} (PRE_SALES|MAIN)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT min(selection_id) FROM cpq_selection WHERE tenant_id=%s", (tid,))
        sel = cur.fetchone()
        if not sel or not sel[0]:
            raise HTTPException(409, detail="cpq_selection 없음")
        run_id, totals = _latest_cost_base(cur, tid)
        direct_total = sum(totals.values())
        revenue = round(direct_total * (1 + body.marginRate))
        sga = round(revenue * 0.08)
        margin = revenue - direct_total
        ebit = margin - sga
        sections = {
            "runId": run_id, "revenue": revenue,
            "material": totals.get("MATERIAL", 0),
            "manufacturing": totals.get("MANUFACTURING", 0),
            "direct": totals.get("DIRECT", 0),
            "sga": sga, "marginRate": body.marginRate,
        }
        cur.execute(
            """UPDATE cst_pcr SET sections=%s, direct_cost_total=%s, contribution_margin=%s,
               ebit=%s, status='DRAFT', updated_by=%s, updated_at=now()
               WHERE tenant_id=%s AND selection_id=%s AND business_type=%s RETURNING pcr_id""",
            (json.dumps(sections), direct_total, margin, ebit, request.state.login,
             tid, sel[0], bt))
        row = cur.fetchone()
        if not row:
            cur.execute(
                """INSERT INTO cst_pcr (tenant_id, selection_id, business_type, sections,
                   direct_cost_total, contribution_margin, ebit, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING pcr_id""",
                (tid, sel[0], bt, json.dumps(sections), direct_total, margin, ebit,
                 request.state.login))
            row = cur.fetchone()
    return {"pcrId": row[0], "businessType": bt, "revenue": revenue,
            "directCostTotal": direct_total, "contributionMargin": margin, "ebit": ebit}


# ── U19 PCR 세부 비용 체계 (슬라이드 74) — 비용 트리(조달/부제조/직접/판관 분해) ──

@router.get("/cost/pcr/{pcr_id}/breakdown")
def pcr_breakdown(pcr_id: int) -> dict[str, Any]:
    """PCR 비용 트리 (U19) — Run 원가 라인(재료/제조/직접) + 판관 분해(율 기반 근사, 합계=SGA 8%)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT business_type, sections, direct_cost_total, contribution_margin, ebit
               FROM cst_pcr WHERE tenant_id=%s AND pcr_id=%s""", (tid, pcr_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"PCR 없음: #{pcr_id}")
        sec = row[1] or {}
        run_id = sec.get("runId")
        lines_by_type: dict[str, list[dict[str, Any]]] = {}
        if run_id:
            cur.execute(
                "SELECT calc_type, detail FROM cst_calc WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
            for ct, det in cur.fetchall():
                lines_by_type[ct] = (det or {}).get("lines", [])
    revenue = float(sec.get("revenue", 0))
    sga_total = float(sec.get("sga", 0))
    # 판관 분해 — 슬라이드 74 세부 항목을 율 기반 근사 (합계 = SGA 8%, 근거 명시)
    sga_rows = [
        {"name": "Sales overhead (3%)", "amount": round(revenue * 0.03)},
        {"name": "Adm overhead (2%)", "amount": round(revenue * 0.02)},
        {"name": "Tech. R&D Cost (2%)", "amount": round(revenue * 0.02)},
    ]
    sga_rows.append({"name": "기타 (Travel·Maintenance 등)", "amount": round(sga_total - sum(r["amount"] for r in sga_rows))})
    def sect(title: str, ct: str, fallback: float) -> dict[str, Any]:
        rows = [{"name": ln.get("name") or ln.get("code", "?"), "amount": float(ln.get("amount", 0))}
                for ln in lines_by_type.get(ct, [])]
        return {"title": title, "rows": rows[:20],
                "subtotal": float(sec.get(fallback, 0)) if isinstance(fallback, str) else fallback}
    sections = [
        {**sect("Procurement cost (재료비)", "MATERIAL", "material")},
        {**sect("Sub-manufacturing cost (제조비)", "MANUFACTURING", "manufacturing")},
        {**sect("Other direct cost (직접경비)", "DIRECT", "direct")},
    ]
    return {
        "pcrId": pcr_id, "businessType": row[0], "revenue": revenue,
        "sections": sections,
        "directCostTotal": float(row[2]), "contributionMargin": float(row[3]),
        "sga": {"rows": sga_rows, "subtotal": sga_total, "basis": "율 기반 근사 — 합계 = SGA 8% (CST-005 상세율은 PCR 기준 관리 후속)"},
        "fullCosts": float(row[2]) + sga_total, "ebit": float(row[4]),
    }


@router.get("/cost/pcr")
def pcr_list(request: Request) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.pcr_id, p.business_type, p.sections, p.direct_cost_total,
                      p.contribution_margin, p.ebit, p.status, s.finished_goods_code
               FROM cst_pcr p JOIN cpq_selection s ON s.selection_id=p.selection_id
               WHERE p.tenant_id=%s ORDER BY p.pcr_id""", (tid,))
        rows = cur.fetchall()
        # 1.5 — 원가 열람 모드: summary 는 상세 sections 제거(요약만), masked/hidden 은 금액 마스킹
        cm = _info_mode(cur, tid, request, "cost")
        if cm != "full":
            _audit(cur, tid, "cst_pcr", 0, "MASKED_READ", request.state.user_id, {"cost": cm})
        return [
            {"pcrId": r[0], "businessType": r[1],
             "sections": {} if cm in ("summary", "hidden", "masked") else r[2],
             "directCostTotal": _mask_num(float(r[3]), cm),
             "contributionMargin": _mask_num(float(r[4]) if r[4] is not None else None, cm),
             "ebit": _mask_num(float(r[5]) if r[5] is not None else None, cm),
             "status": r[6], "code": r[7], "maskMode": cm}
            for r in rows
        ]


@router.get("/cost/pcr/{pcr_id}/actual")
def pcr_actual(pcr_id: int) -> dict[str, Any]:
    """실적 반영 PCR 재계산 (D6) — 매출 고정, 직접비를 실적(cst_actual, 프로젝트 귀속)으로 치환 →
    기여마진·EBIT 재산출 + 추정 대비 차이. 실적 없으면 actualAvailable=false."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.sections, p.direct_cost_total, p.contribution_margin, p.ebit, p.selection_id
               FROM cst_pcr p WHERE p.tenant_id=%s AND p.pcr_id=%s""", (tid, pcr_id))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"PCR 없음: #{pcr_id}")
        sections, est_direct, est_margin, est_ebit, sel_id = r
        revenue = float(sections.get("revenue", 0))
        sga = float(sections.get("sga", 0))
        # PCR 프로젝트 (selection→project)
        cur.execute(
            """SELECT pj.project_no FROM cpq_selection s
               JOIN prj_project pj ON pj.project_id=s.project_id
               WHERE s.selection_id=%s""", (sel_id,))
        prow = cur.fetchone()
        project_no = prow[0] if prow else ""
        # 실적 직접비 — 프로젝트 귀속분 합계
        if project_no:
            cur.execute("SELECT COALESCE(sum(amount),0), count(*) FROM cst_actual "
                        "WHERE tenant_id=%s AND project_no=%s", (tid, project_no))
        else:
            cur.execute("SELECT COALESCE(sum(amount),0), count(*) FROM cst_actual WHERE tenant_id=%s", (tid,))
        act_row = cur.fetchone()
        act_direct, act_count = float(act_row[0]), act_row[1]
    est_direct = float(est_direct)
    est_margin = float(est_margin) if est_margin is not None else revenue - est_direct
    est_ebit = float(est_ebit) if est_ebit is not None else est_margin - sga
    act_margin = round(revenue - act_direct, 2)
    act_ebit = round(act_margin - sga, 2)
    return {
        "pcrId": pcr_id, "projectNo": project_no, "revenue": revenue, "sga": sga,
        "actualAvailable": act_count > 0, "actualCount": act_count,
        "estimate": {"directCost": est_direct, "margin": est_margin, "ebit": est_ebit,
                     "marginPct": round(est_margin / revenue, 4) if revenue else 0},
        "actual": {"directCost": act_direct, "margin": act_margin, "ebit": act_ebit,
                   "marginPct": round(act_margin / revenue, 4) if revenue else 0},
        "variance": {"directCost": round(act_direct - est_direct, 2),
                     "margin": round(act_margin - est_margin, 2),
                     "ebit": round(act_ebit - est_ebit, 2),
                     "marginPctDelta": round((act_margin - est_margin) / revenue, 4) if revenue else 0},
    }


@router.get("/reports/catalog")
def reports_catalog() -> list[dict[str, Any]]:
    """리포트 센터 카탈로그 — 산발 리포트 생성기 목록 + 건수(데이터 기반)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT count(*) FROM cst_pcr WHERE tenant_id=%s", (tid,))
        pcr = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM cst_quotation WHERE tenant_id=%s", (tid,))
        quot = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM doc_control WHERE tenant_id=%s", (tid,))
        docs = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM dwg_file WHERE tenant_id=%s", (tid,))
        dwg = cur.fetchone()[0]
    return [
        {"id": "pcr", "name": "PCR 수익성 보고서", "category": "원가", "kind": "PDF", "count": pcr,
         "screen": "cpq-run", "desc": "기여마진·EBIT 수익성 분석 (RPT-07)"},
        {"id": "quotation", "name": "견적서", "category": "영업", "kind": "PDF", "count": quot,
         "screen": "cpq-run", "desc": "PCR 기반 견적서 렌더"},
        {"id": "document", "name": "문서(Grade 워터마크)", "category": "문서", "kind": "PDF", "count": docs,
         "screen": "cpq-docmgmt", "desc": "doc_control 문서 렌더 (S-1/S-2 CONFIDENTIAL)"},
        {"id": "cad-plot", "name": "CAD 축척 도면", "category": "설계", "kind": "PDF", "count": dwg,
         "screen": "com-folder", "desc": "DXF 1:scale 벡터 출력"},
        {"id": "audit", "name": "감사 로그", "category": "보안", "kind": "XLSX", "count": None,
         "screen": "erp-audit", "desc": "sys_history 기간/사용자/작업 필터 export"},
    ]


@router.get("/reports/pcr/{pcr_id}.pdf")
def pcr_report_pdf(pcr_id: int) -> StreamingResponse:
    """PCR 수익성 보고서 PDF (RPT-07) — 기여마진·EBIT·원가 구성."""
    import io as _io

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfgen import canvas

    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.business_type, p.sections, p.direct_cost_total, p.contribution_margin,
                      p.ebit, p.status, s.finished_goods_code, COALESCE(pr.project_no,''),
                      to_char(p.created_at,'YYYY-MM-DD')
               FROM cst_pcr p JOIN cpq_selection s ON s.selection_id=p.selection_id
               LEFT JOIN prj_project pr ON pr.project_id=s.project_id
               WHERE p.tenant_id=%s AND p.pcr_id=%s""", (tid, pcr_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"PCR 없음: {pcr_id}")
    biz, sections, direct, margin, ebit, status, code, proj, created = row

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
        font = "HYSMyeongJo-Medium"
    except Exception:
        font = "Helvetica"
    buf = _io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    y = H - 60
    c.setFont(font, 16); c.setFillColorRGB(0.09, 0.16, 0.31)
    c.drawString(50, y, "PCR 수익성 보고서 (RPT-07)"); y -= 8
    c.setStrokeColorRGB(0.09, 0.16, 0.31); c.setLineWidth(1.2); c.line(50, y, W - 50, y); y -= 24

    def line(label: str, val: str, bold: bool = False) -> None:
        nonlocal y
        c.setFont(font, 11 if not bold else 12)
        c.setFillColorRGB(0.35, 0.4, 0.5); c.drawString(54, y, label)
        c.setFillColorRGB(0.1, 0.12, 0.18); c.drawRightString(W - 54, y, val)
        y -= 20

    def won(v: Any) -> str:
        return f"{float(v):,.0f} 원" if v is not None else "—"

    line("제품 코드", str(code)); line("프로젝트", proj or "—")
    line("사업 유형", str(biz)); line("상태", str(status)); line("작성일", str(created))
    y -= 6; c.setStrokeColorRGB(0.8, 0.84, 0.9); c.line(50, y, W - 50, y); y -= 22
    line("직접비 합계", won(direct))
    if isinstance(sections, (list, tuple)):
        for sec in sections:
            if isinstance(sec, dict):
                nm = sec.get("name") or sec.get("label") or sec.get("key") or "-"
                amt = sec.get("amount", sec.get("value", sec.get("total")))
                line(f"  · {nm}", won(amt) if isinstance(amt, (int, float)) else str(amt))
    elif isinstance(sections, dict):
        for k, v in sections.items():
            line(f"  · {k}", won(v) if isinstance(v, (int, float)) else str(v))
    y -= 6; c.setStrokeColorRGB(0.8, 0.84, 0.9); c.line(50, y, W - 50, y); y -= 22
    line("기여마진 (Contribution Margin)", won(margin), True)
    line("EBIT", won(ebit), True)

    c.setFont(font, 8); c.setFillColorRGB(0.5, 0.55, 0.62)
    c.drawString(50, 40, f"EDIM · PCR #{pcr_id} · cst_pcr")
    c.showPage(); c.save()
    return StreamingResponse(
        _io.BytesIO(buf.getvalue()), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="PCR_{pcr_id}.pdf"'})


class QuotationCreate(BaseModel):
    businessType: str = "PRE_SALES"
    validityPeriod: str = "견적일로부터 30일"
    deliveryTerms: str = "FOB 부산"
    paymentTerms: str = "T/T 30일"
    currency: str = "KRW"
    taxCode: str = ""


@router.post("/cost/quotations", status_code=201, dependencies=[SETUP])
def quotation_create(request: Request, body: QuotationCreate) -> dict[str, Any]:
    """견적 확정 — PCR 기반 cst_quotation 행 생성 (line_items = 최근 Run 원가 라인)."""
    bt = body.businessType.strip().upper()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT p.pcr_id, p.sections, p.selection_id, s.project_id
               FROM cst_pcr p JOIN cpq_selection s ON s.selection_id=p.selection_id
               WHERE p.tenant_id=%s AND p.business_type=%s
               ORDER BY p.pcr_id DESC LIMIT 1""", (tid, bt))
        pcr = cur.fetchone()
        if not pcr:
            raise HTTPException(409, detail=f"PCR 없음 ({bt}) — 수익성 보고서 먼저 생성")
        cur.execute(
            "SELECT company_id FROM com_company WHERE tenant_id=%s ORDER BY company_id LIMIT 1", (tid,))
        customer = cur.fetchone()[0]
        run_id = pcr[1].get("runId")
        cur.execute(
            """SELECT detail FROM cst_calc WHERE tenant_id=%s AND run_id=%s AND calc_type='MATERIAL'""",
            (tid, run_id))
        mat = cur.fetchone()
        line_items = (mat[0].get("lines", []) if mat else [])
        revenue_krw = float(pcr[1].get("revenue", 0))   # PCR 매출(기준통화 KRW)
        # 다통화/세금엔진 — 통화 환산 + 세액 적재
        cur_c = body.currency.strip().upper()[:3] or "KRW"
        rate = _fx_rate(cur, tid, cur_c)
        if rate is None:
            raise HTTPException(422, detail=f"환율 미등록 통화: {cur_c} (다통화 마스터 등록 필요)")
        pct = 0.0
        if body.taxCode.strip():
            cur.execute("SELECT rate_pct FROM tax_code WHERE tenant_id=%s AND code=%s",
                        (tid, body.taxCode.strip().upper()))
            tr = cur.fetchone()
            if not tr:
                raise HTTPException(422, detail=f"세금코드 없음: {body.taxCode}")
            pct = float(tr[0])
        subtotal = round(revenue_krw / rate, 2)      # 견적 통화 공급가액
        tax = round(subtotal * pct / 100, 2)
        total = round(subtotal + tax, 2)             # 세액 포함 합계(견적 통화)
        vat_mode = body.taxCode.strip().upper() or "별도"
        cur.execute("SELECT count(*)+1 FROM cst_quotation WHERE tenant_id=%s", (tid,))
        seq = cur.fetchone()[0]
        qno = f"QT-{run_id}-{seq:03d}"
        cur.execute(
            """INSERT INTO cst_quotation (tenant_id, quotation_no, pcr_id, project_id, customer_id,
               total_amount, currency, vat_mode, validity_period, delivery_terms, payment_terms,
               line_items, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING quotation_id""",
            (tid, qno, pcr[0], pcr[3], customer, total, cur_c, vat_mode,
             body.validityPeriod.strip()[:50], body.deliveryTerms.strip()[:200],
             body.paymentTerms.strip()[:200], json.dumps(line_items), request.state.login))
        qid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'cst_quotation',%s,'CREATE',%s,%s)""",
            (tid, qid, request.state.user_id,
             json.dumps({"quotationNo": qno, "currency": cur_c, "subtotal": subtotal, "tax": tax, "total": total})))
    return {"quotationId": qid, "quotationNo": qno, "currency": cur_c, "rate": rate,
            "taxPct": pct, "subtotal": subtotal, "tax": tax, "total": total}


@router.get("/cost/quotations")
def quotation_list() -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT code, rate_pct FROM tax_code WHERE tenant_id=%s", (tid,))
        tax_rates = {c: float(p) for c, p in cur.fetchall()}
        cur.execute(
            """SELECT q.quotation_id, q.quotation_no, q.total_amount, q.currency, q.status,
                      to_char(q.created_at,'YYYY-MM-DD'), p.project_no, c.company_name,
                      q.validity_period, q.delivery_terms, q.payment_terms, q.vat_mode
               FROM cst_quotation q
               JOIN prj_project p ON p.project_id=q.project_id
               JOIN com_company c ON c.company_id=q.customer_id
               WHERE q.tenant_id=%s ORDER BY q.quotation_id DESC""", (tid,))
        out = []
        for r in cur.fetchall():
            total = float(r[2])
            pct = tax_rates.get(r[11], 0.0)   # vat_mode = 세금코드
            subtotal = round(total / (1 + pct / 100), 2) if pct else total
            out.append({"quotationId": r[0], "quotationNo": r[1], "total": total, "currency": r[3],
                        "status": r[4], "date": r[5], "project": r[6], "customer": r[7],
                        "validity": r[8], "delivery": r[9], "payment": r[10],
                        "taxCode": r[11], "taxPct": pct, "subtotal": subtotal,
                        "tax": round(total - subtotal, 2)})
        return out


@router.delete("/cost/quotations/{quotation_id}", dependencies=[SETUP])
def quotation_delete(quotation_id: int, request: Request) -> dict[str, Any]:
    """견적 삭제 — DRAFT 한정 (발행/승인 견적 보호)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT status, quotation_no FROM cst_quotation WHERE tenant_id=%s AND quotation_id=%s",
            (tid, quotation_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"견적 없음: #{quotation_id}")
        if row[0] != "DRAFT":
            raise HTTPException(409, detail=f"DRAFT 견적만 삭제 가능 (현재 {row[0]})")
        cur.execute("DELETE FROM cst_quotation WHERE quotation_id=%s", (quotation_id,))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'cst_quotation',%s,'DELETE',%s,%s)""",
            (tid, quotation_id, request.state.user_id, json.dumps({"quotationNo": row[1]})))
    return {"deleted": quotation_id}


QUOTE_TRANSITIONS = {"DRAFT": ("SENT",), "SENT": ("ORDERED", "LOST"), "ORDERED": (), "LOST": ()}


class QuoteStatusPatch(BaseModel):
    status: str
    contractAmount: float | None = None
    expectedDelivery: str | None = None   # YYYY-MM-DD (ORDERED 시)


@router.patch("/cost/quotations/{quotation_id}/status", dependencies=[SETUP])
def quotation_status(quotation_id: int, request: Request, body: QuoteStatusPatch) -> dict[str, Any]:
    """견적 lifecycle 전이 (D1) — DRAFT→SENT→ORDERED/LOST. ORDERED = 수주(계약금액·납기·영업단계 CONTRACT)."""
    new = body.status.strip().upper()
    if new not in ("SENT", "ORDERED", "LOST"):
        raise HTTPException(422, detail="상태 오류 (SENT/ORDERED/LOST)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT status, project_id, total_amount, quotation_no
               FROM cst_quotation WHERE tenant_id=%s AND quotation_id=%s""", (tid, quotation_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"견적 없음: #{quotation_id}")
        cur_status, project_id, total, qno = row
        if new not in QUOTE_TRANSITIONS.get(cur_status, ()):
            raise HTTPException(409, detail=f"전이 불가: {cur_status} → {new} "
                                f"(허용: {', '.join(QUOTE_TRANSITIONS.get(cur_status, ())) or '없음'})")
        if new == "ORDERED":
            amt = body.contractAmount if body.contractAmount is not None else float(total)
            cur.execute(
                """UPDATE cst_quotation SET status='ORDERED', contract_amount=%s,
                   order_date=CURRENT_DATE, expected_delivery=%s, updated_at=now()
                   WHERE quotation_id=%s""", (amt, body.expectedDelivery, quotation_id))
            # 프로젝트 영업단계 자동 전이 → CONTRACT (D1 고리)
            cur.execute(
                """UPDATE prj_project SET sales_stage='CONTRACT', updated_at=now()
                   WHERE project_id=%s AND sales_stage NOT IN ('CONTRACT','CLOSED')""", (project_id,))
            # 마일스톤 자동 시딩 (D1→D7 고리) — 표준 5단계, 기존 단계는 보존(ON CONFLICT)
            cur.execute("SELECT project_no FROM prj_project WHERE project_id=%s", (project_id,))
            pno = cur.fetchone()
            offs = {"ORDER": 0, "DESIGN": 14, "PURCHASE": 30, "PRODUCTION": 50, "SHIPMENT": 70}
            if pno:
                for stg in MILESTONE_STAGES:
                    if stg == "SHIPMENT" and body.expectedDelivery:
                        cur.execute(
                            """INSERT INTO prj_milestone (tenant_id, project_no, stage, planned_date, note, created_by)
                               VALUES (%s,%s,%s,%s::date,'수주 자동 시딩',%s)
                               ON CONFLICT (tenant_id, project_no, stage) DO NOTHING""",
                            (tid, pno[0], stg, body.expectedDelivery, str(request.state.user_id)))
                    else:
                        cur.execute(
                            """INSERT INTO prj_milestone (tenant_id, project_no, stage, planned_date, note, created_by)
                               VALUES (%s,%s,%s,CURRENT_DATE + (%s||' days')::interval,'수주 자동 시딩',%s)
                               ON CONFLICT (tenant_id, project_no, stage) DO NOTHING""",
                            (tid, pno[0], stg, offs[stg], str(request.state.user_id)))
            # 후속 이벤트 자동 생성 (D1→업무함 고리) — OR(수주)의 후행 프로세스를 TODO 로 시딩(설계 착수 TODO)
            followups = _seed_order_followups(cur, tid, project_id, quotation_id, qno, request.state.user_id)
        else:
            cur.execute("UPDATE cst_quotation SET status=%s, updated_at=now() WHERE quotation_id=%s",
                        (new, quotation_id))
        _audit(cur, tid, "cst_quotation", quotation_id, "STATUS", request.state.user_id,
               {"quotationNo": qno, "from": cur_status, "to": new,
                "followups": [f["code"] for f in followups] if new == "ORDERED" else None})
    return {"quotationId": quotation_id, "status": new,
            "followupEvents": followups if new == "ORDERED" else []}


def _seed_order_followups(cur: Any, tid: int, project_id: int, quotation_id: int,
                          qno: str, user_id: int) -> list[dict[str, Any]]:
    """수주 확정 시 OR(수주)의 후행 프로세스를 TODO 이벤트로 생성 (업무함 착수 TODO).

    erp_process_edge 로 OR 후행(AP 승인도서·PL Part List)을 조회 → 프로젝트별 미존재 시 생성·담당자 알림.
    담당자: 프로젝트 관리자 우선, 없으면 수행자. 중복 방지(같은 proc×project 있으면 skip)."""
    cur.execute("SELECT proc_def_id FROM erp_process_def WHERE tenant_id=%s AND proc_code='OR'", (tid,))
    or_def = cur.fetchone()
    if not or_def:
        return []
    cur.execute(
        """SELECT d.proc_def_id, d.proc_code, d.proc_name
           FROM erp_process_edge e JOIN erp_process_def d ON d.proc_def_id=e.to_def_id
           WHERE e.tenant_id=%s AND e.from_def_id=%s ORDER BY d.proc_code""", (tid, or_def[0]))
    successors = cur.fetchall()
    cur.execute("SELECT manager_id FROM prj_project WHERE project_id=%s", (project_id,))
    mgr = cur.fetchone()
    assignee = mgr[0] if mgr and mgr[0] else user_id
    lead = {"AP": 7, "PL": 10}   # 착수 리드타임(영업일 아닌 캘린더 근사)
    seeded: list[dict[str, Any]] = []
    for pdid, pcode, pname in successors:
        cur.execute(
            "SELECT 1 FROM erp_process_event WHERE tenant_id=%s AND proc_def_id=%s AND project_id=%s LIMIT 1",
            (tid, pdid, project_id))
        if cur.fetchone():
            continue
        cur.execute(
            """INSERT INTO erp_process_event (tenant_id, proc_def_id, project_id, ref_type, ref_id,
               status, assignee_id, due_date, created_by)
               VALUES (%s,%s,%s,'QUOTATION',%s,'TODO',%s,
                       CURRENT_DATE + (%s||' days')::interval, %s) RETURNING event_id""",
            (tid, pdid, project_id, quotation_id, assignee, lead.get(pcode, 7), str(user_id)))
        eid = cur.fetchone()[0]
        seeded.append({"eventId": eid, "code": pcode, "name": pname})
        if assignee:
            _notify(cur, tid, int(assignee), "ORDER_FOLLOWUP",
                    f"수주 후속 착수 — {pname} ({qno})", "/erp")
    return seeded


# ── U17 설계우선순위 테이블 (슬라이드 44, S-4-1-2) ──

class DesignParamItem(BaseModel):
    no: str
    designPriority: int | None = None
    dataPriority: int | None = None
    basePoint: str = ""
    errorCheck: str = ""
    remarks: str = ""


class DesignParamsSave(BaseModel):
    drawing: str
    items: list[DesignParamItem] = []


@router.get("/drawings/dimensions/design-params")
def design_params(drawing: str = "KDCR 3-13") -> list[dict[str, Any]]:
    """치수별 설계 파라미터 (U17) — 설계/자료 우선순위·기준점·오류체크."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT d.dim_label, d.dim_type, d.design_priority, d.data_priority,
                      COALESCE(d.base_point,''), COALESCE(d.error_check,''), COALESCE(d.remarks,'')
               FROM dwg_dimension d JOIN dwg_drawing w ON w.drawing_id=d.drawing_id
               WHERE d.tenant_id=%s AND w.drawing_no=%s
               ORDER BY d.design_priority NULLS LAST, d.dim_label""", (tid, drawing))
        return [{"no": r[0], "kind": r[1], "designPriority": r[2], "dataPriority": r[3],
                 "basePoint": r[4], "errorCheck": r[5], "remarks": r[6]} for r in cur.fetchall()]


@router.put("/drawings/dimensions/design-params", dependencies=[SETUP])
def design_params_save(request: Request, body: DesignParamsSave) -> dict[str, Any]:
    """설계 파라미터 일괄 저장 (U17) — dim_label 기준 갱신."""
    n = 0
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT drawing_id FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s",
            (tid, body.drawing.strip()))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"도면 없음: {body.drawing}")
        for it in body.items:
            cur.execute(
                """UPDATE dwg_dimension SET design_priority=%s, data_priority=%s,
                       base_point=NULLIF(%s,''), error_check=NULLIF(%s,''), remarks=NULLIF(%s,'')
                   WHERE tenant_id=%s AND drawing_id=%s AND dim_label=%s""",
                (it.designPriority, it.dataPriority, it.basePoint.strip()[:100],
                 it.errorCheck.strip()[:100], it.remarks.strip()[:300],
                 tid, row[0], it.no.strip()[:10]))
            n += cur.rowcount
        _audit(cur, tid, "dwg_dimension", row[0], "DESIGN_PARAMS_SAVE", request.state.user_id,
               after={"drawing": body.drawing, "rows": n})
    return {"updated": n}


@router.get("/erp/production/schedule")
def production_schedule() -> dict[str, Any]:
    """생산 스케줄·Capacity 1차 (U4·ERP-023, 슬라이드 46) — 미완료 작업지시 × Work Process 공수.

    소요 공수 = 해당 도면 코드의 erp_work_process MAKE 행 work_time 합(분, U3 파라미터).
    작업장 Capacity = 작업장별 인원 최대치 × 480분/일 근사 (상세 캘린더는 후속).
    """
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT wo.wo_no, wo.title, wo.drawing_no, wo.project_no, wo.status,
                      COALESCE(wo.assignee,''), to_char(wo.issued_at,'YYYY-MM-DD'),
                      GREATEST(0, (CURRENT_DATE - wo.issued_at::date)) AS age_days,
                      COALESCE((SELECT SUM(wp.work_time) FROM erp_work_process wp
                                JOIN product_code pc ON pc.product_code_id = wp.product_code_id
                                WHERE wp.tenant_id = wo.tenant_id AND pc.main_code = wo.drawing_no
                                  AND wp.make_or_buy = 'MAKE'), 0) AS work_min
               FROM erp_work_order wo
               WHERE wo.tenant_id=%s AND wo.status IN ('ISSUED','STARTED')
               ORDER BY wo.issued_at""", (tid,))
        orders = [{"woNo": r[0], "title": r[1], "drawingNo": r[2], "projectNo": r[3],
                   "status": r[4], "assignee": r[5], "issuedAt": r[6],
                   "ageDays": int(r[7]), "workMin": float(r[8])} for r in cur.fetchall()]
        # 작업장 부하 — 미완료 WO 도면들의 MAKE 공수 × 작업장 (U3 workshop 파라미터)
        cur.execute(
            """SELECT COALESCE(wp.workshop,'(미지정)'),
                      SUM(wp.work_time) AS load_min,
                      MAX(COALESCE(wp.person_count,1)) AS persons
               FROM erp_work_order wo
               JOIN product_code pc ON pc.main_code = wo.drawing_no
               JOIN erp_work_process wp ON wp.tenant_id = wo.tenant_id
                    AND wp.product_code_id = pc.product_code_id AND wp.make_or_buy='MAKE'
               WHERE wo.tenant_id=%s AND wo.status IN ('ISSUED','STARTED')
                 AND wp.work_time IS NOT NULL
               GROUP BY COALESCE(wp.workshop,'(미지정)') ORDER BY 2 DESC""", (tid,))
        workshops = []
        for r in cur.fetchall():
            load_min = float(r[1] or 0)
            persons = int(r[2] or 1)
            cap_min = persons * 480.0
            workshops.append({"workshop": r[0], "loadMin": load_min, "persons": persons,
                              "capMinPerDay": cap_min,
                              "loadPct": round(load_min / cap_min * 100, 1) if cap_min else 0,
                              "daysNeeded": round(load_min / cap_min, 2) if cap_min else 0})
    return {"orders": orders, "workshops": workshops,
            "openCount": len(orders), "totalWorkMin": sum(o["workMin"] for o in orders)}


@router.get("/erp/mrp")
def mrp_plan(leadDays: int = 14) -> dict[str, Any]:
    """MRP 자재 소요 계획 1차 (U4·M-8-5, ERP-022) — 수주(ORDERED 견적) 자재 라인 × 현재고 대비.

    소요 = ORDERED 견적 line_items(재료비 라인: code·qty) 집계, 보유 = erp_stock 합계.
    부족분 발주 권장일 = 최단 납기 - leadDays (리드타임 기본 14일 — 공급처별 리드타임은 후속).
    """
    lead = max(0, min(90, leadDays))
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT q.quotation_no, to_char(q.expected_delivery,'YYYY-MM-DD'),
                      p.project_no, q.line_items
               FROM cst_quotation q JOIN prj_project p ON p.project_id=q.project_id
               WHERE q.tenant_id=%s AND q.status='ORDERED'
               ORDER BY q.expected_delivery NULLS LAST""", (tid,))
        req: dict[str, dict[str, Any]] = {}
        order_count = 0
        for qno, due, prj, items in cur.fetchall():
            order_count += 1
            for it in (items or []):
                code = str(it.get("code", "")).strip()
                if not code:
                    continue
                qty = float(it.get("qty", 0) or 0)
                r = req.setdefault(code, {"code": code, "name": it.get("name", ""),
                                          "required": 0.0, "dueDate": due, "orders": []})
                r["required"] += qty
                if due and (not r["dueDate"] or due < r["dueDate"]):
                    r["dueDate"] = due
                if qno not in r["orders"]:
                    r["orders"].append(qno)
        cur.execute(
            "SELECT item_code, COALESCE(SUM(quantity),0) FROM inv_stock WHERE tenant_id=%s GROUP BY item_code",
            (tid,))
        on_hand = {r[0]: float(r[1]) for r in cur.fetchall()}
    rows = []
    for r in sorted(req.values(), key=lambda x: (x["dueDate"] or "9999", x["code"])):
        oh = on_hand.get(r["code"], 0.0)
        shortage = max(0.0, r["required"] - oh)
        order_by = ""
        if r["dueDate"]:
            from datetime import datetime, timedelta
            try:
                order_by = (datetime.strptime(r["dueDate"], "%Y-%m-%d") - timedelta(days=lead)).strftime("%Y-%m-%d")
            except ValueError:
                order_by = ""
        rows.append({**r, "onHand": oh, "shortage": shortage, "orderBy": order_by,
                     "status": "SHORT" if shortage > 0 else "OK"})
    return {"rows": rows, "orderCount": order_count,
            "shortCount": sum(1 for x in rows if x["status"] == "SHORT"), "leadDays": lead}


@router.get("/cost/orders")
def sales_orders() -> dict[str, Any]:
    """수주 잔고 (D1) — ORDERED 견적(프로젝트별 수주액·납기·단계) + 수주율 지표."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT q.quotation_no, q.contract_amount, q.total_amount,
                      to_char(q.order_date,'YYYY-MM-DD'), to_char(q.expected_delivery,'YYYY-MM-DD'),
                      p.project_no, p.project_name, p.sales_stage, c.company_name
               FROM cst_quotation q
               JOIN prj_project p ON p.project_id=q.project_id
               JOIN com_company c ON c.company_id=q.customer_id
               WHERE q.tenant_id=%s AND q.status='ORDERED'
               ORDER BY q.order_date DESC NULLS LAST, q.quotation_id DESC""", (tid,))
        orders = [
            {"quotationNo": r[0], "contractAmount": float(r[1]) if r[1] is not None else float(r[2]),
             "quoteAmount": float(r[2]), "orderDate": r[3], "expectedDelivery": r[4],
             "project": r[5], "projectName": r[6], "stage": r[7], "customer": r[8]}
            for r in cur.fetchall()
        ]
        cur.execute(
            """SELECT count(*) FILTER (WHERE status='ORDERED'),
                      count(*) FILTER (WHERE status IN ('SENT','ORDERED','LOST'))
               FROM cst_quotation WHERE tenant_id=%s""", (tid,))
        won, decided = cur.fetchone()
    total_amt = sum(o["contractAmount"] for o in orders)
    return {"orders": orders, "orderCount": len(orders),
            "orderRate": round(won / decided, 3) if decided else 0.0,
            "totalAmount": total_amt}


@router.get("/cost/quotations/{quotation_id}/render.pdf")
def quotation_render(quotation_id: int) -> Any:
    """견적서 PDF 렌더 (U19) — CLT 공식 양식 (슬라이드 74: 헤더 메타·품목표·합계·조건)."""
    from fastapi.responses import Response
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT q.quotation_no, q.line_items, q.total_amount, p.project_no, q.currency, q.vat_mode,
                      p.project_name, c.company_name, COALESCE(q.delivery_terms,''),
                      COALESCE(q.payment_terms,''), COALESCE(q.validity_period,''),
                      to_char(q.created_at,'YYYY-MM-DD')
               FROM cst_quotation q
               JOIN prj_project p ON p.project_id=q.project_id
               JOIN com_company c ON c.company_id=q.customer_id
               WHERE q.tenant_id=%s AND q.quotation_id=%s""", (tid, quotation_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"견적 없음: #{quotation_id}")
        total = float(row[2])
        pct = 0.0
        if row[5] and row[5] != "별도":
            cur.execute("SELECT rate_pct FROM tax_code WHERE tenant_id=%s AND code=%s", (tid, row[5]))
            tr = cur.fetchone()
            if tr:
                pct = float(tr[0])
    subtotal = round(total / (1 + pct / 100), 2) if pct else total
    pdf = rp.build_clt_quotation_pdf(
        quotation_no=row[0], project_name=f"{row[6]} ({row[3]})", customer=row[7],
        items=[{"code": ln.get("code", "?"), "name": ln.get("name", ""),
                "qty": ln.get("qty", 1), "priceK": ln.get("priceK")} for ln in row[1]],
        currency=row[4], subtotal=subtotal, tax=round(total - subtotal, 2), total=total,
        delivery_terms=row[8], payment_terms=row[9], validity=row[10], quote_date=row[11])
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename=\"{row[0]}.pdf\""})


class SelectionCreate(BaseModel):
    projectNo: str
    rootCode: str = "KDCR 3-13"
    finishedGoodsCode: str
    slotValues: dict[str, str] = {}
    specInput: dict[str, Any] | None = None


@router.get("/cpq/selections")
def list_selections(projectNo: str = "") -> list[dict[str, Any]]:
    """견적안 목록 (C1 — 프로젝트별 저장된 구성/불러오기 대상)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if projectNo.strip():
            cur.execute(
                """SELECT s.selection_id, s.finished_goods_code, s.slot_values, s.status,
                          to_char(s.created_at,'MM-DD HH24:MI'), s.x_code_status, s.is_standard,
                          (SELECT count(*) FROM cpq_run r WHERE r.selection_id=s.selection_id)
                   FROM cpq_selection s JOIN prj_project p ON p.project_id=s.project_id
                   WHERE s.tenant_id=%s AND p.project_no=%s ORDER BY s.selection_id DESC""",
                (tid, projectNo.strip()))
        else:
            cur.execute(
                """SELECT s.selection_id, s.finished_goods_code, s.slot_values, s.status,
                          to_char(s.created_at,'MM-DD HH24:MI'), s.x_code_status, s.is_standard,
                          (SELECT count(*) FROM cpq_run r WHERE r.selection_id=s.selection_id)
                   FROM cpq_selection s WHERE s.tenant_id=%s ORDER BY s.selection_id DESC LIMIT 20""",
                (tid,))
        return [{"selectionId": r[0], "finishedGoodsCode": r[1], "slotValues": r[2],
                 "status": r[3], "createdAt": r[4], "xCodeStatus": r[5],
                 "isStandard": r[6], "runCount": r[7]} for r in cur.fetchall()]


@router.post("/cpq/selections", status_code=201, dependencies=[SETUP])
def create_selection(request: Request, body: SelectionCreate) -> dict[str, Any]:
    """견적안 저장 (C1) — 현재 C-1 구성(슬롯·완성 코드·사양)을 cpq_selection 으로 영속."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT project_id FROM prj_project WHERE tenant_id=%s AND project_no=%s",
                    (tid, body.projectNo.strip()))
        prj = cur.fetchone()
        if not prj:
            raise HTTPException(404, detail=f"프로젝트 없음: {body.projectNo}")
        cur.execute("SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
                    (tid, body.rootCode.strip()))
        pc = cur.fetchone()
        if not pc:
            raise HTTPException(422, detail=f"루트 코드 없음: {body.rootCode}")
        is_std = not body.finishedGoodsCode.upper().startswith("X")
        cur.execute(
            """INSERT INTO cpq_selection (tenant_id, project_id, finished_goods_code,
               product_code_id, slot_values, spec_input, is_standard, status,
               x_code_status, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,'DRAFT',%s,%s) RETURNING selection_id""",
            (tid, prj[0], body.finishedGoodsCode.strip()[:100], pc[0],
             json.dumps(body.slotValues), json.dumps(body.specInput) if body.specInput else None,
             is_std, None if is_std else "PENDING", str(request.state.user_id)))
        sid = cur.fetchone()[0]
        _audit(cur, tid, "cpq_selection", sid, "CREATE", request.state.user_id,
               {"finishedGoodsCode": body.finishedGoodsCode, "project": body.projectNo})
    return {"selectionId": sid, "status": "DRAFT"}


@router.get("/cpq/x-review")
def x_review_list() -> list[dict[str, Any]]:
    """X-code 검토 대기열 — 비표준(X) 코드 견적안 중 x_code_status='PENDING'."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT s.selection_id, s.finished_goods_code, s.slot_values, p.project_no,
                      COALESCE(p.project_name,''), to_char(s.created_at,'YYYY-MM-DD HH24:MI'), s.created_by
               FROM cpq_selection s JOIN prj_project p ON p.project_id=s.project_id
               WHERE s.tenant_id=%s AND s.x_code_status='PENDING'
               ORDER BY s.selection_id DESC""", (tid,))
        return [{"selectionId": r[0], "finishedGoodsCode": r[1], "slotValues": r[2],
                 "projectNo": r[3], "projectName": r[4], "createdAt": r[5], "createdBy": r[6]}
                for r in cur.fetchall()]


class XReviewRequest(BaseModel):
    decision: str          # APPROVE | REJECT
    comment: str = ""


@router.post("/cpq/selections/{selection_id}/x-review", dependencies=[SETUP])
def x_review(selection_id: int, request: Request, body: XReviewRequest) -> dict[str, Any]:
    """X-code 검토 결정 — PENDING → APPROVED/REJECTED, 요청자 알림·감사."""
    new_status = {"APPROVE": "APPROVED", "REJECT": "REJECTED"}.get(body.decision.strip().upper())
    if not new_status:
        raise HTTPException(422, detail="decision 은 APPROVE 또는 REJECT")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """UPDATE cpq_selection SET x_code_status=%s
               WHERE tenant_id=%s AND selection_id=%s AND x_code_status='PENDING'
               RETURNING finished_goods_code, created_by""",
            (new_status, tid, selection_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(409, detail="검토 불가 — PENDING X-code 견적안 아님")
        code, created_by = row
        _audit(cur, tid, "cpq_selection", selection_id, "X_CODE_REVIEW", request.state.user_id,
               {"decision": new_status, "code": code, "comment": body.comment})
        if created_by and str(created_by).isdigit():
            _notify(cur, tid, int(created_by), "X_CODE_RESULT",
                    f"X-code {code} {'승인' if new_status == 'APPROVED' else '반려'}", "/cpq")
    return {"selectionId": selection_id, "finishedGoodsCode": code, "xCodeStatus": new_status}


@router.delete("/cpq/selections/{selection_id}", dependencies=[SETUP])
def delete_selection(selection_id: int, request: Request) -> dict[str, Any]:
    """견적안 삭제 (C1) — Run 참조 시 409 보호."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM cpq_run WHERE selection_id=%s LIMIT 1", (selection_id,))
        if cur.fetchone():
            raise HTTPException(409, detail="Run 이력이 있는 견적안은 삭제 불가 (참조 보호)")
        cur.execute(
            "DELETE FROM cpq_selection WHERE tenant_id=%s AND selection_id=%s RETURNING finished_goods_code",
            (tid, selection_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"견적안 없음: #{selection_id}")
        _audit(cur, tid, "cpq_selection", selection_id, "DELETE", request.state.user_id,
               {"finishedGoodsCode": row[0]})
    return {"deleted": selection_id}


@router.post("/cpq/runs", status_code=202)
async def start_run(body: RunRequest) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        if body.selectionId:   # C1 — 지정 견적안
            cur.execute(
                """SELECT s.selection_id, s.slot_values, p.project_no
                   FROM cpq_selection s JOIN prj_project p ON p.project_id=s.project_id
                   WHERE s.tenant_id=%s AND s.selection_id=%s""", (tid, body.selectionId))
        else:                  # 미지정 — 최신 견적안
            cur.execute(
                """SELECT s.selection_id, s.slot_values, p.project_no
                   FROM cpq_selection s JOIN prj_project p ON p.project_id=s.project_id
                   WHERE s.tenant_id=%s ORDER BY s.selection_id DESC LIMIT 1""", (tid,))
        sel = cur.fetchone()
        if not sel:
            raise HTTPException(404 if body.selectionId else 503,
                                detail="견적안 없음" if body.selectionId else "seed selection missing")
        cur.execute(
            """INSERT INTO cpq_run (tenant_id, selection_id, run_type, status, is_test)
               VALUES (%s,%s,%s,'RUNNING',%s) RETURNING run_id""",
            (tid, sel[0], body.runType, body.isTest))
        run_id = cur.fetchone()[0]
    _runs[run_id] = {
        "status": "RUNNING", "current": -1, "outputs": [], "logs": [],
        "steps": [
            {"no": i + 1, "task": t, "measured": "—", "elapsed": "", "status": "PENDING"}
            for i, t in enumerate(RUN_TASKS)
        ],
    }
    asyncio.get_running_loop().create_task(
        _advance(run_id, tid, sel[0], sel[1] or {}, sel[2]))
    return {"runId": run_id, "status": "RUNNING", "statusUrl": f"/api/v1/cpq/runs/{run_id}"}


@router.get("/cpq/runs/{run_id}")
def run_status(run_id: int) -> dict[str, Any]:
    state = _runs.get(run_id)
    if state is None:
        # 백엔드 재기동 후 — DB 에서 산출물 복원
        with _conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT status FROM cpq_run WHERE run_id=%s", (run_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, detail=f"run not found: {run_id}")
            cur.execute(
                """SELECT o.output_type, o.data->>'file', o.data->>'fileType', o.file_id
                   FROM cpq_output o WHERE o.run_id=%s ORDER BY o.output_id""", (run_id,))
            outputs = [_out_row(r[0], r[1] or "output", r[2] or "PDF", r[3])
                       for r in cur.fetchall()]
        return {
            "runId": run_id, "status": row[0], "progress": 1.0,
            "steps": [
                {"no": i + 1, "task": t, "measured": "(재기동 — 이력)",
                 "elapsed": "", "status": "DONE"}
                for i, t in enumerate(RUN_TASKS)
            ],
            "outputs": outputs, "logs": [],
        }
    done = state["status"] != "RUNNING"
    cur_i = state["current"]
    progress = 1.0 if done else max(0.0, (cur_i + 0.5) / len(RUN_TASKS))
    return {
        "runId": run_id, "status": state["status"], "progress": progress,
        "steps": state["steps"], "outputs": state["outputs"], "logs": state["logs"],
    }


# ── E3 Run 산출물 누적 관리 (cpq_run 이력·정리) ──
def _run_refs(cur, tid: int) -> tuple[set[int], int | None]:
    """(견적/PCR 이 참조하는 run_id 집합, 최신 SUCCESS run_id) — 삭제 보호 판단."""
    cur.execute(
        "SELECT DISTINCT (sections->>'runId')::bigint FROM cst_pcr "
        "WHERE tenant_id=%s AND sections->>'runId' IS NOT NULL", (tid,))
    refs = {r[0] for r in cur.fetchall() if r[0] is not None}
    cur.execute(
        "SELECT max(run_id) FROM cpq_run WHERE tenant_id=%s AND status='SUCCESS'", (tid,))
    latest = cur.fetchone()[0]
    return refs, latest


@router.get("/cpq/runs")
def run_list() -> list[dict[str, Any]]:
    """Run 이력 (E3) — 전체 Run 목록 + 산출물 수 + 참조/최신 보호 플래그."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        refs, latest = _run_refs(cur, tid)
        cur.execute(
            """SELECT r.run_id, r.status, r.run_type, to_char(r.started_at,'MM-DD HH24:MI'),
                      EXTRACT(EPOCH FROM (r.finished_at - r.started_at)),
                      (SELECT count(*) FROM cpq_output o WHERE o.run_id=r.run_id),
                      COALESCE(r.created_by,'system'), r.is_test
               FROM cpq_run r WHERE r.tenant_id=%s ORDER BY r.run_id DESC""", (tid,))
        return [{"runId": x[0], "status": x[1], "runType": x[2], "startedAt": x[3],
                 "durationSec": round(float(x[4]), 1) if x[4] is not None else None,
                 "outputCount": x[5], "createdBy": x[6], "isTest": x[7],
                 "latest": x[0] == latest, "referenced": x[0] in refs,
                 "protected": x[0] == latest or x[0] in refs} for x in cur.fetchall()]


@router.get("/cpq/runs/{run_id}/bom-snapshot")
def run_bom_snapshot(run_id: int) -> dict[str, Any]:
    """BOM Snapshot 조회 (트리아지 #41) — Run 완료 시 고정된 전개 결과."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT bom_snapshot FROM cpq_run WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"run not found: {run_id}")
    rows = row[0] or []
    return {"runId": run_id, "count": len(rows), "rows": rows}


@router.get("/cpq/runs/{run_id}/bom-basis")
def run_bom_basis(run_id: int) -> dict[str, Any]:
    """BOM 전개 근거 대조 (#40) — Run 이 고정한 관계 Revision 집합 vs 지금 다시 폈을 때의 집합.

    Snapshot(1.7)이 *결과*를 불변 보존한다면, 여기는 *근거*를 대조한다.
    stable=true 면 지금 재실행해도 같은 BOM 이 나온다는 뜻이고, false 면 무엇이 달라졌는지
    관계 단위(added/removed/revised)로 짚어준다. 근거가 바뀌어도 과거 BOM 은 바뀌지 않는다."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.rel_basis, s.slot_values, r.status
               FROM cpq_run r LEFT JOIN cpq_selection s ON s.selection_id=r.selection_id
               WHERE r.tenant_id=%s AND r.run_id=%s""", (tid, run_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"run not found: {run_id}")
        pinned = row[0]
        if not pinned:
            # #40 이전 Run — 근거가 기록되지 않았다. 없는 것을 있는 척하지 않는다.
            return {"runId": run_id, "pinned": None, "stable": None,
                    "reason": "이 Run 은 전개 근거 도입(#40) 이전에 실행되어 근거가 없습니다"}
        cur_rows = _expand_rows(cur, tid, "KDCR 3-13", row[1] or {})
        current = _rel_basis(cur_rows)
        pin_map = {e["relId"]: e["revisionNo"] for e in pinned.get("edges", [])}
        cur_map = {e["relId"]: e["revisionNo"] for e in current["edges"]}
        # 관계 표시용 이름 (rel_id → 'Mother > Child')
        ids = sorted(set(pin_map) | set(cur_map))
        names: dict[int, str] = {}
        if ids:
            cur.execute(
                """SELECT cr.rel_id, m.main_code || ' > ' || c.main_code
                   FROM code_relationship cr
                   JOIN product_code m ON m.product_code_id=cr.mother_code_id
                   JOIN product_code c ON c.product_code_id=cr.child_code_id
                   WHERE cr.tenant_id=%s AND cr.rel_id = ANY(%s)""", (tid, ids))
            names = {r[0]: r[1] for r in cur.fetchall()}
        diff = []
        for rid in ids:
            label = names.get(rid, f"#{rid}")
            if rid not in cur_map:
                diff.append({"relId": rid, "label": label, "change": "removed",
                             "pinnedRevision": pin_map[rid], "currentRevision": None})
            elif rid not in pin_map:
                diff.append({"relId": rid, "label": label, "change": "added",
                             "pinnedRevision": None, "currentRevision": cur_map[rid]})
            elif pin_map[rid] != cur_map[rid]:
                diff.append({"relId": rid, "label": label, "change": "revised",
                             "pinnedRevision": pin_map[rid], "currentRevision": cur_map[rid]})
        return {"runId": run_id, "pinned": pinned, "current": current,
                "stable": pinned.get("checksum") == current["checksum"],
                "edgeCount": len(pin_map), "diff": diff}


# ── 1.7 Snapshot 레지스트리 (요구 #9) — 실행 결과 고정·재현·무결성 ──

class SnapshotCreate(BaseModel):
    runId: int
    note: str = ""


def _run_state(cur, tid: int, run_id: int) -> dict[str, Any]:
    """Run 시점 상태 수집 — Snapshot payload 이자 재현(drift) 대조의 기준."""
    cur.execute(
        """SELECT r.status, r.run_type, COALESCE(r.is_test,false), r.bom_snapshot,
                  s.selection_id, s.finished_goods_code, r.rel_basis,
                  jsonb_build_object('slotValues', s.slot_values, 'specInput', s.spec_input,
                                     'arrangementId', s.arrangement_id,
                                     'isStandard', s.is_standard, 'status', s.status),
                  p.project_no
           FROM cpq_run r
           LEFT JOIN cpq_selection s ON s.selection_id=r.selection_id
           LEFT JOIN prj_project p ON p.project_id=s.project_id
           WHERE r.tenant_id=%s AND r.run_id=%s""", (tid, run_id))
    r = cur.fetchone()
    if not r:
        raise HTTPException(404, detail=f"run not found: {run_id}")
    cur.execute(
        """SELECT calc_type, total_amount FROM cst_calc
           WHERE tenant_id=%s AND run_id=%s ORDER BY calc_type""", (tid, run_id))
    costs = [{"calcType": c[0], "total": float(c[1])} for c in cur.fetchall()]
    cur.execute(
        """SELECT output_type, data->>'file', data->>'fileType' FROM cpq_output
           WHERE run_id=%s ORDER BY output_type, output_id""", (run_id,))
    outputs = [{"type": o[0], "file": o[1], "fileType": o[2]} for o in cur.fetchall()]
    bom = r[3] or []
    return {
        "runId": run_id, "status": r[0], "runType": r[1], "isTest": r[2],
        "selectionId": r[4], "finishedGoodsCode": r[5],
        "selections": r[7], "projectNo": r[8],
        "bomRows": len(bom), "bom": bom, "costs": costs, "outputs": outputs,
        "relBasis": r[6],   # #40 — 전개 근거도 Snapshot payload 에 동결
    }


def _snapshot_checksum(payload: dict[str, Any]) -> str:
    """정규화 직렬화 기준 SHA-256 — 키 순서·공백에 무관한 안정 해시."""
    canon = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(canon.encode("utf-8")).hexdigest()


def _freeze_snapshot(cur, tid: int, run_id: int, actor_id: int, note: str = "") -> dict[str, Any]:
    """Run 상태를 Snapshot 으로 동결 — 같은 Run 재고정은 version+1 (기존 건 불변 보존)."""
    payload = _run_state(cur, tid, run_id)
    checksum = _snapshot_checksum(payload)
    cur.execute(
        """SELECT COALESCE(max(version_no),0) FROM sys_snapshot
           WHERE tenant_id=%s AND snapshot_type='CPQ_RUN' AND source_id=%s""", (tid, run_id))
    ver = int(cur.fetchone()[0]) + 1
    code = f"SNAP-R{run_id}-v{ver}"
    cur.execute(
        """INSERT INTO sys_snapshot (tenant_id, snapshot_code, snapshot_type, source_id,
                                     version_no, payload, checksum, note, created_by)
           VALUES (%s,%s,'CPQ_RUN',%s,%s,%s,%s,%s,%s)
           RETURNING snapshot_id, to_char(created_at,'YYYY-MM-DD HH24:MI')""",
        (tid, code, run_id, ver, json.dumps(payload), checksum, note[:300] or None, actor_id))
    row = cur.fetchone()
    _audit(cur, tid, "sys_snapshot", row[0], "SNAPSHOT_FREEZE", actor_id,
           {"runId": run_id, "code": code, "checksum": checksum[:12]})
    return {"snapshotId": row[0], "snapshotCode": code, "version": ver,
            "checksum": checksum, "createdAt": row[1], "bomRows": payload["bomRows"]}


@router.post("/snapshots", status_code=201, dependencies=[SETUP])
def snapshot_create(request: Request, body: SnapshotCreate) -> dict[str, Any]:
    """Run 결과 Snapshot 고정 (요구 #9) — 이후 원본이 바뀌어도 이 ID 로 근거가 재현된다."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        return _freeze_snapshot(cur, tid, body.runId, request.state.user_id, body.note)


@router.get("/snapshots")
def snapshot_list(sourceId: int = 0) -> list[dict[str, Any]]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        sql = ("""SELECT s.snapshot_id, s.snapshot_code, s.snapshot_type, s.source_id,
                         s.version_no, s.checksum, COALESCE(s.note,''),
                         to_char(s.created_at,'MM-DD HH24:MI'),
                         COALESCE((s.payload->>'bomRows')::int,0),
                         COALESCE(s.payload->>'finishedGoodsCode',''),
                         COALESCE(s.payload->>'projectNo',''),
                         EXISTS (SELECT 1 FROM erp_handoff h WHERE h.snapshot_id=s.snapshot_id)
                  FROM sys_snapshot s WHERE s.tenant_id=%s""")
        params: list[Any] = [tid]
        if sourceId:
            sql += " AND s.source_id=%s"
            params.append(sourceId)
        cur.execute(sql + " ORDER BY s.snapshot_id DESC LIMIT 100", tuple(params))
        return [{"snapshotId": r[0], "snapshotCode": r[1], "snapshotType": r[2],
                 "sourceId": r[3], "version": r[4], "checksum": r[5][:12],
                 "note": r[6], "createdAt": r[7], "bomRows": r[8],
                 "finishedGoodsCode": r[9], "projectNo": r[10], "handedOff": r[11]}
                for r in cur.fetchall()]


@router.get("/snapshots/{snapshot_id}")
def snapshot_get(snapshot_id: int) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT snapshot_code, snapshot_type, source_id, version_no, payload, checksum,
                      COALESCE(note,''), to_char(created_at,'YYYY-MM-DD HH24:MI')
               FROM sys_snapshot WHERE tenant_id=%s AND snapshot_id=%s""", (tid, snapshot_id))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"snapshot not found: {snapshot_id}")
    return {"snapshotId": snapshot_id, "snapshotCode": r[0], "snapshotType": r[1],
            "sourceId": r[2], "version": r[3], "payload": r[4], "checksum": r[5],
            "note": r[6], "createdAt": r[7]}


@router.get("/snapshots/{snapshot_id}/verify")
def snapshot_verify(snapshot_id: int) -> dict[str, Any]:
    """재현 검증 (요구 #9) — ① 저장 payload 무결성(checksum) ② 현재 원본과의 drift 비교.

    drift 는 '원본이 바뀌었다'는 사실 보고이지 Snapshot 훼손이 아니다 (Snapshot 은 불변).
    """
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT snapshot_code, source_id, payload, checksum
               FROM sys_snapshot WHERE tenant_id=%s AND snapshot_id=%s""", (tid, snapshot_id))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, detail=f"snapshot not found: {snapshot_id}")
        code, run_id, payload, checksum = r
        intact = _snapshot_checksum(payload) == checksum
        try:
            current = _run_state(cur, tid, run_id)
            source_exists = True
        except HTTPException:
            current, source_exists = {}, False

    drift: list[dict[str, Any]] = []
    if source_exists:
        for key in ("status", "finishedGoodsCode", "projectNo", "bomRows"):
            if payload.get(key) != current.get(key):
                drift.append({"field": key, "snapshot": payload.get(key), "current": current.get(key)})
        # #40 — 행 수만 보던 BOM 대조를 내용 기준으로 강화 (행 수가 같아도 수량·코드가 바뀔 수 있다)
        if _snapshot_checksum({"b": payload.get("bom", [])}) != _snapshot_checksum({"b": current.get("bom", [])}):
            drift.append({"field": "bom", "snapshot": f"{len(payload.get('bom', []))}행",
                          "current": f"{len(current.get('bom', []))}행 (내용 상이)"})
        # #40 — 전개 근거(관계 Revision) 이동 여부
        if (payload.get("relBasis") or {}).get("checksum") != (current.get("relBasis") or {}).get("checksum"):
            drift.append({"field": "relBasis",
                          "snapshot": (payload.get("relBasis") or {}).get("checksum"),
                          "current": (current.get("relBasis") or {}).get("checksum")})
        if [c["total"] for c in payload.get("costs", [])] != [c["total"] for c in current.get("costs", [])]:
            drift.append({"field": "costs", "snapshot": payload.get("costs"), "current": current.get("costs")})
        if len(payload.get("outputs", [])) != len(current.get("outputs", [])):
            drift.append({"field": "outputs",
                          "snapshot": len(payload.get("outputs", [])),
                          "current": len(current.get("outputs", []))})
    return {"snapshotId": snapshot_id, "snapshotCode": code, "intact": intact,
            "sourceExists": source_exists, "reproducible": intact,
            "drift": drift, "driftCount": len(drift)}


class HandoffCreate(BaseModel):
    runId: int


@router.post("/erp/handoffs", status_code=201, dependencies=[SETUP])
def handoff_create(request: Request, body: HandoffCreate) -> dict[str, Any]:
    """ERP Handoff 생성 (트리아지 #44~47) — Validation(pass/warning/fail) 후 승인 요청.

    ERP 는 승인된 Handoff 만 수신한다. 같은 프로젝트 재생성 = 새 Version, 이전 미수신 건은 superseded."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.status, r.bom_snapshot, p.project_no
               FROM cpq_run r
               JOIN cpq_selection s ON s.selection_id=r.selection_id
               JOIN prj_project p ON p.project_id=s.project_id
               WHERE r.tenant_id=%s AND r.run_id=%s""", (tid, body.runId))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"run not found: {body.runId}")
        status, bom, project_no = row
        # Guard — 임시/실패 Run·빈 BOM 차단 (fail 등급)
        checks = []
        if status != "SUCCESS":
            checks.append({"check": "run", "grade": "fail", "detail": f"Run 상태 {status} — SUCCESS 만 Handoff 가능"})
        bom_rows = len(bom or [])
        checks.append({"check": "bom", "grade": "pass" if bom_rows > 0 else "fail",
                       "detail": f"BOM Snapshot {bom_rows}행"})
        cur.execute("SELECT count(*) FROM cst_calc WHERE tenant_id=%s AND run_id=%s", (tid, body.runId))
        cost_n = cur.fetchone()[0]
        checks.append({"check": "cost", "grade": "pass" if cost_n else "warning",
                       "detail": f"원가 상세 {cost_n}건" if cost_n else "원가 상세 없음"})
        cur.execute("SELECT count(*) FROM cpq_output WHERE run_id=%s AND output_type='QUOTATION'", (body.runId,))
        q_n = cur.fetchone()[0]
        checks.append({"check": "quotation", "grade": "pass" if q_n else "warning",
                       "detail": "견적서 산출물" if q_n else "견적서 산출물 없음"})
        if any(c["grade"] == "fail" for c in checks):
            raise HTTPException(422, detail="Handoff Validation 실패 — " +
                                "; ".join(c["detail"] for c in checks if c["grade"] == "fail"))
        grade = "warning" if any(c["grade"] == "warning" for c in checks) else "pass"
        # 새 Version — 이전 미수신(validated/approval_requested/approved) 건 supersede
        cur.execute(
            "SELECT COALESCE(max(version),0)+1 FROM erp_handoff WHERE tenant_id=%s AND project_no=%s",
            (tid, project_no))
        version = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO erp_handoff (tenant_id, project_no, run_id, version, status, validation, created_by)
               VALUES (%s,%s,%s,%s,'validated',%s,%s) RETURNING handoff_id""",
            (tid, project_no, body.runId, version,
             json.dumps({"grade": grade, "checks": checks}), request.state.login))
        hid = cur.fetchone()[0]
        # 1.7 — Handoff 는 Snapshot 을 근거로 넘긴다 (요구 #9): 생성 시점 상태를 자동 고정·연결
        snap = _freeze_snapshot(cur, tid, body.runId, request.state.user_id,
                                f"ERP Handoff {project_no} v{version} 자동 고정")
        cur.execute("UPDATE erp_handoff SET snapshot_id=%s WHERE handoff_id=%s", (snap["snapshotId"], hid))
        cur.execute(
            """UPDATE erp_handoff SET status='superseded', superseded_by=%s
               WHERE tenant_id=%s AND project_no=%s AND handoff_id<>%s
                 AND status IN ('validated','approval_requested','approved')""",
            (hid, tid, project_no, hid))
        # 승인 요청 자동 생성 → 승인함 (승인 후에만 ERP 수신 가능)
        cur.execute(
            """INSERT INTO sys_approval_request (tenant_id, target_table, target_id,
               request_type, step, requester_id, comment)
               VALUES (%s,'erp_handoff',%s,'CREATE','승인',%s,%s)""",
            (tid, hid, request.state.user_id, f"ERP Handoff — {project_no} v{version} (run #{body.runId}, {grade})"))
        cur.execute("UPDATE erp_handoff SET status='approval_requested' WHERE handoff_id=%s", (hid,))
        _audit(cur, tid, "erp_handoff", hid, "HANDOFF_CREATE", request.state.user_id,
               {"projectNo": project_no, "runId": body.runId, "version": version, "grade": grade})
    return {"handoffId": hid, "projectNo": project_no, "version": version,
            "status": "approval_requested", "grade": grade, "checks": checks,
            "snapshotId": snap["snapshotId"], "snapshotCode": snap["snapshotCode"]}


@router.get("/erp/handoffs")
def handoff_list(project: str = "") -> list[dict[str, Any]]:
    """Handoff 목록 (트리아지 #49) — 수신 상태 + FG Code·Snapshot ID 병기 (#38 표시/추적 분리)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        params: list[Any] = [tid]
        clause = ""
        if project.strip():
            clause = " AND h.project_no=%s"
            params.append(project.strip())
        cur.execute(
            f"""SELECT h.handoff_id, h.project_no, h.run_id, h.version, h.status, h.validation,
                       to_char(h.created_at,'MM-DD HH24:MI'), COALESCE(h.created_by,''),
                       to_char(h.accepted_at,'MM-DD HH24:MI'),
                       s.finished_goods_code, s.selection_id,
                       h.snapshot_id, sn.snapshot_code
               FROM erp_handoff h
               LEFT JOIN sys_snapshot sn ON sn.snapshot_id=h.snapshot_id
               LEFT JOIN cpq_run r ON r.run_id=h.run_id
               LEFT JOIN cpq_selection s ON s.selection_id=r.selection_id
               WHERE h.tenant_id=%s{clause}
               ORDER BY h.handoff_id DESC LIMIT 50""", tuple(params))
        return [{"handoffId": r[0], "projectNo": r[1], "runId": r[2], "version": r[3],
                 "status": r[4], "grade": (r[5] or {}).get("grade"),
                 "checks": (r[5] or {}).get("checks", []),
                 "createdAt": r[6], "createdBy": r[7], "acceptedAt": r[8],
                 "finishedGoodsCode": r[9] or "", "configSnapshotId": r[10],
                 "snapshotId": r[11], "snapshotCode": r[12] or ""}
                for r in cur.fetchall()]


@router.get("/tenant/export.zip", dependencies=[ADMIN])
def tenant_export(request: Request) -> StreamingResponse:
    """테넌트 오프보딩/백업 export (트리아지 #13) — 코어 테이블 JSON 덤프 ZIP (ADMIN, 감사 기록).

    파일 원본(MinIO)은 기존 /files/zip 채널 — 본 export 는 구조화 데이터 반출."""
    import zipfile
    TABLES = [
        ("projects", "SELECT * FROM prj_project WHERE tenant_id=%s"),
        ("companies", "SELECT * FROM com_company WHERE tenant_id=%s"),
        ("product_codes", "SELECT * FROM product_code WHERE tenant_id=%s"),
        ("code_relationships", "SELECT * FROM code_relationship WHERE tenant_id=%s"),
        ("drawings", "SELECT * FROM dwg_drawing WHERE tenant_id=%s"),
        ("parts", "SELECT * FROM prt_part WHERE tenant_id=%s"),
        ("warehouses", "SELECT * FROM erp_warehouse WHERE tenant_id=%s"),
        ("quotations", "SELECT * FROM cst_quotation WHERE tenant_id=%s"),
        ("documents", "SELECT * FROM doc_control WHERE tenant_id=%s"),
        ("selections", "SELECT * FROM cpq_selection WHERE tenant_id=%s"),
        ("runs", "SELECT run_id, selection_id, run_type, status, started_at, finished_at, is_test "
                 "FROM cpq_run WHERE tenant_id=%s"),
        ("handoffs", "SELECT * FROM erp_handoff WHERE tenant_id=%s"),
        ("users", "SELECT user_id, login_id, user_name, department, email, user_level, status, created_at "
                  "FROM sys_user WHERE tenant_id=%s"),   # 비밀번호 해시 등 민감 필드 제외
    ]
    buf = io.BytesIO()
    manifest = ["EDIM 테넌트 데이터 export (오프보딩/백업)", "-" * 40]
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, sql in TABLES:
                cur.execute(sql, (tid,))
                cols = [d[0] for d in cur.description]
                rows = [dict(zip(cols, r)) for r in cur.fetchall()]
                zf.writestr(f"{name}.json", json.dumps(rows, ensure_ascii=False, default=str, indent=1))
                manifest.append(f"{name}: {len(rows)}건")
            zf.writestr("manifest.txt", "\n".join(manifest))
        _audit(cur, tid, "sys_tenant", tid, "TENANT_EXPORT", request.state.user_id,
               {"tables": len(TABLES)})
    from urllib.parse import quote as _q
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_q('tenant_export.zip')}",
                                      "X-Table-Count": str(len(TABLES))})


@router.get("/projects/{project_no}/output-packages")
def project_output_packages(project_no: str) -> list[dict[str, Any]]:
    """Project Output Package 조회 (트리아지 #42) — SUCCESS Run 단위 산출물 묶음 뷰.

    Package = FG Code(표시) + Config/BOM Snapshot ID(추적) + 산출물 수 + Handoff 상태."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT r.run_id, to_char(r.started_at,'MM-DD HH24:MI'),
                      (SELECT count(*) FROM cpq_output o WHERE o.run_id=r.run_id),
                      COALESCE(jsonb_array_length(r.bom_snapshot),0),
                      s.finished_goods_code, s.selection_id,
                      (SELECT h.status FROM erp_handoff h
                       WHERE h.tenant_id=r.tenant_id AND h.run_id=r.run_id
                       ORDER BY h.handoff_id DESC LIMIT 1)
               FROM cpq_run r
               JOIN cpq_selection s ON s.selection_id=r.selection_id
               JOIN prj_project p ON p.project_id=s.project_id
               WHERE r.tenant_id=%s AND p.project_no=%s AND r.status='SUCCESS' AND NOT r.is_test
               ORDER BY r.run_id DESC LIMIT 20""", (tid, project_no))
        return [{"packageId": r[0], "at": r[1], "outputCount": r[2], "bomRows": r[3],
                 "finishedGoodsCode": r[4] or "", "configSnapshotId": r[5],
                 "handoffStatus": r[6]} for r in cur.fetchall()]


@router.post("/erp/handoffs/{handoff_id}/accept", dependencies=[SETUP])
def handoff_accept(handoff_id: int, request: Request) -> dict[str, Any]:
    """ERP 수신 (트리아지 #46) — 승인된 Handoff 만 수신 가능."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT status, project_no FROM erp_handoff WHERE tenant_id=%s AND handoff_id=%s",
                    (tid, handoff_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"handoff not found: {handoff_id}")
        if row[0] != "approved":
            raise HTTPException(409, detail=f"승인된 Handoff 만 수신 가능 (현재 {row[0]})")
        cur.execute("UPDATE erp_handoff SET status='accepted', accepted_at=now() WHERE handoff_id=%s",
                    (handoff_id,))
        _audit(cur, tid, "erp_handoff", handoff_id, "HANDOFF_ACCEPT", request.state.user_id,
               {"projectNo": row[1]})
    return {"handoffId": handoff_id, "status": "accepted"}


@router.get("/cpq/runs/{run_id}/outputs")
def run_outputs(run_id: int) -> list[dict[str, Any]]:
    """Run 산출물 드릴다운 (E3)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM cpq_run WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"run not found: {run_id}")
        cur.execute(
            """SELECT output_type, data->>'file', data->>'fileType', to_char(created_at,'MM-DD HH24:MI')
               FROM cpq_output WHERE run_id=%s ORDER BY output_id""", (run_id,))
        return [{"outputType": r[0], "file": r[1] or "output", "fileType": r[2] or "PDF",
                 "createdAt": r[3]} for r in cur.fetchall()]


def _delete_run(cur, tid: int, run_id: int) -> None:
    """Run 물리 삭제 — 자식(cst_calc·cpq_output) 먼저 정리 (FK). dwg_file/MinIO 는 보존."""
    cur.execute("DELETE FROM cst_calc WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
    cur.execute("DELETE FROM cpq_output WHERE run_id=%s", (run_id,))
    cur.execute("DELETE FROM cpq_run WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
    _runs.pop(run_id, None)


@router.delete("/cpq/runs/{run_id}", dependencies=[SETUP])
def run_delete(run_id: int, request: Request) -> dict[str, Any]:
    """Run 정리 (E3) — 견적(PCR) 참조·최신 SUCCESS Run 은 409 보호."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM cpq_run WHERE tenant_id=%s AND run_id=%s", (tid, run_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"run not found: {run_id}")
        refs, latest = _run_refs(cur, tid)
        if run_id == latest:
            raise HTTPException(409, detail="최신 SUCCESS Run 은 정리 불가 — 현재 원가/견적 기준")
        if run_id in refs:
            raise HTTPException(409, detail="견적(PCR)이 참조 중인 Run — 정리 불가")
        _delete_run(cur, tid, run_id)
        _audit(cur, tid, "cpq_run", run_id, "RUN_DELETE", request.state.user_id, {"runId": run_id})
    return {"runId": run_id, "deleted": True}


class RunCleanup(BaseModel):
    keepLatest: int = 5   # 최근 N건 유지


@router.post("/cpq/runs/cleanup", dependencies=[SETUP])
def run_cleanup(request: Request, body: RunCleanup) -> dict[str, Any]:
    """보관 정책 정리 (E3) — 최근 N건 유지, 그 외 미참조 Run 일괄 정리 (참조·최신 보호)."""
    keep = max(1, min(body.keepLatest, 100))
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        refs, latest = _run_refs(cur, tid)
        cur.execute(
            "SELECT run_id FROM cpq_run WHERE tenant_id=%s ORDER BY run_id DESC", (tid,))
        all_ids = [r[0] for r in cur.fetchall()]
        keep_ids = set(all_ids[:keep])
        if latest is not None:
            keep_ids.add(latest)
        deleted, skipped = [], []
        for rid in all_ids[keep:]:
            if rid in keep_ids or rid in refs:
                skipped.append(rid)
                continue
            _delete_run(cur, tid, rid)
            deleted.append(rid)
        if deleted:
            _audit(cur, tid, "cpq_run", 0, "RUN_CLEANUP", request.state.user_id,
                   {"deleted": deleted, "keepLatest": keep})
    return {"deleted": len(deleted), "skipped": len(skipped),
            "keptLatest": keep, "deletedIds": deleted}


# ── B7 — PLM 도면 대장 (dwg_drawing·dwg_revision·dwg_supersedure 개방) ──

def _rev_next(rev: str) -> str:
    """Rev 문자 증가 — A→B…Z→AA (엑셀 컬럼 방식)."""
    letters = list(rev.strip().upper() or "@")   # '@'+1 == 'A'
    i = len(letters) - 1
    while i >= 0:
        if letters[i] != "Z":
            letters[i] = chr(ord(letters[i]) + 1)
            return "".join(letters)
        letters[i] = "A"
        i -= 1
    return "A" + "".join(letters)


def _drawing_id(cur, tid: int, drawing_no: str) -> int:
    cur.execute(
        "SELECT drawing_id FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s",
        (tid, drawing_no.strip()))
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, detail=f"도면 없음: {drawing_no}")
    return row[0]


@router.get("/drawings")
def drawings_list(code: str = "") -> list[dict[str, Any]]:
    """도면 대장 — dwg_drawing + Rev 수·최신 DXF 연결. code 지정 시 해당 도면번호만."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        where, params = "d.tenant_id=%s", [tid]
        if code.strip():
            where += " AND d.drawing_no=%s"
            params.append(code.strip())
        cur.execute(
            f"""SELECT d.drawing_no, d.drawing_name, d.drawing_type, d.dwg_kind,
                       d.current_rev, d.status,
                       (SELECT COUNT(*) FROM dwg_revision r WHERE r.drawing_id=d.drawing_id),
                       f.file_id, f.file_name,
                       EXISTS (SELECT 1 FROM dwg_supersedure s WHERE s.old_drawing_id=d.drawing_id)
                FROM dwg_drawing d
                LEFT JOIN LATERAL (
                    SELECT file_id, file_name FROM dwg_file
                    WHERE drawing_id=d.drawing_id AND file_type='DXF'
                    ORDER BY file_id DESC LIMIT 1) f ON TRUE
                WHERE {where} ORDER BY d.drawing_no""", params)
        return [
            {"drawingNo": r[0], "name": r[1], "type": r[2], "kind": r[3],
             "rev": r[4], "status": r[5], "revCount": r[6],
             "fileId": r[7], "fileName": r[8], "superseded": bool(r[9])}
            for r in cur.fetchall()
        ]


class DrawingCreate(BaseModel):
    drawingNo: str
    name: str
    drawingType: str = "PART"
    kind: str = "STANDARD"


@router.post("/drawings", status_code=201, dependencies=[SETUP])
def create_drawing(request: Request, body: DrawingCreate) -> dict[str, Any]:
    """도면 등록 — dwg_drawing insert + Rev.A 이력 (중복 409)."""
    no, name = body.drawingNo.strip(), body.name.strip()
    if not no or not name:
        raise HTTPException(422, detail="도면번호·도면명은 필수입니다")
    if body.drawingType not in ("ASSEMBLY", "PART", "LAYOUT"):
        raise HTTPException(422, detail=f"유형 오류: {body.drawingType} (ASSEMBLY|PART|LAYOUT)")
    if body.kind not in ("APPROVAL", "MANUFACTURING", "STANDARD"):
        raise HTTPException(422, detail=f"Kind 오류: {body.kind}")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT 1 FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s", (tid, no))
        if cur.fetchone():
            raise HTTPException(409, detail=f"도면번호 중복: {no}")
        cur.execute(
            """INSERT INTO dwg_drawing (tenant_id, drawing_no, drawing_name, drawing_type,
               dwg_kind, current_rev, status, created_by)
               VALUES (%s,%s,%s,%s,%s,'A','DRAFT',%s) RETURNING drawing_id""",
            (tid, no, name[:200], body.drawingType, body.kind, request.state.login))
        drawing_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO dwg_revision (drawing_id, rev_no, rev_date, rev_reason, revised_by)
               VALUES (%s,'A',CURRENT_DATE,'최초 발행',%s)""",
            (drawing_id, request.state.login))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_drawing',%s,'CREATE',%s,%s)""",
            (tid, drawing_id, request.state.user_id,
             json.dumps({"drawingNo": no, "name": name})))
    return {"drawingId": drawing_id, "rev": "A", "status": "DRAFT"}


@router.delete("/drawings/{drawing_no}", dependencies=[SETUP])
def delete_drawing(drawing_no: str, request: Request) -> dict[str, Any]:
    """도면 삭제 — RELEASED 보호 (발행 도면 불가). Rev·Supersedure·승인·블록·관계 연쇄 정리,
    연결 파일은 보존(도면 링크만 해제)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT drawing_id, status FROM dwg_drawing WHERE tenant_id=%s AND drawing_no=%s",
            (tid, drawing_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"도면 없음: {drawing_no}")
        did, status = row
        if status == "RELEASED":
            raise HTTPException(409, detail=f"RELEASED 도면은 삭제 불가 (현재 {status})")
        cur.execute(
            "DELETE FROM dwg_supersedure WHERE old_drawing_id=%s OR new_drawing_id=%s",
            (did, did))
        cur.execute("DELETE FROM dwg_approval WHERE drawing_id=%s", (did,))
        cur.execute("DELETE FROM dwg_document WHERE drawing_id=%s", (did,))
        cur.execute("DELETE FROM dwg_part_relation WHERE drawing_id=%s", (did,))
        cur.execute("UPDATE dwg_file SET drawing_id=NULL WHERE drawing_id=%s", (did,))
        cur.execute("DELETE FROM dwg_revision WHERE drawing_id=%s", (did,))
        cur.execute("DELETE FROM dwg_drawing WHERE drawing_id=%s", (did,))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_drawing',%s,'DELETE',%s,%s)""",
            (tid, did, request.state.user_id, json.dumps({"drawingNo": drawing_no})))
    return {"deleted": drawing_no}


@router.get("/drawings/{drawing_no}/revisions")
def drawing_revisions(drawing_no: str) -> list[dict[str, Any]]:
    """Rev 이력 — dwg_revision (최신 우선)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT rev_no, to_char(rev_date,'YYYY-MM-DD'), COALESCE(rev_reason,''), revised_by
               FROM dwg_revision WHERE drawing_id=%s ORDER BY revision_id DESC""", (did,))
        return [{"rev": r[0], "date": r[1], "reason": r[2], "by": r[3]} for r in cur.fetchall()]


class RevUpRequest(BaseModel):
    reason: str = ""


@router.post("/drawings/{drawing_no}/revisions", status_code=201, dependencies=[SETUP])
def rev_up(drawing_no: str, request: Request, body: RevUpRequest) -> dict[str, Any]:
    """Rev 올리기 — current_rev 증가 + dwg_revision 행 + 이력 (B→C)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT drawing_id, current_rev FROM dwg_drawing
               WHERE tenant_id=%s AND drawing_no=%s""", (tid, drawing_no.strip()))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"도면 없음: {drawing_no}")
        did, cur_rev = row
        new_rev = _rev_next(cur_rev)
        cur.execute(
            """INSERT INTO dwg_revision (drawing_id, rev_no, rev_date, rev_reason, revised_by)
               VALUES (%s,%s,CURRENT_DATE,%s,%s)""",
            (did, new_rev, body.reason.strip()[:500] or f"Rev {cur_rev}→{new_rev}",
             request.state.login))
        cur.execute(
            """UPDATE dwg_drawing SET current_rev=%s, updated_by=%s, updated_at=now()
               WHERE drawing_id=%s""", (new_rev, request.state.login, did))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_drawing',%s,'REV_UP',%s,%s)""",
            (tid, did, request.state.user_id,
             json.dumps({"from": cur_rev, "to": new_rev, "reason": body.reason})))
    return {"drawingNo": drawing_no, "rev": new_rev, "prevRev": cur_rev}


@router.get("/drawings/supersedures")
def supersedures_list() -> list[dict[str, Any]]:
    """Supersedure — Rev 대체 이력 (구도면 → 신도면)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT o.drawing_no, o.drawing_name, n.drawing_no, n.current_rev,
                      COALESCE(s.reason,''), to_char(s.superseded_date,'YYYY-MM-DD')
               FROM dwg_supersedure s
               JOIN dwg_drawing o ON o.drawing_id=s.old_drawing_id
               JOIN dwg_drawing n ON n.drawing_id=s.new_drawing_id
               WHERE s.tenant_id=%s ORDER BY s.supersedure_id DESC""", (tid,))
        return [
            {"oldNo": r[0], "oldName": r[1], "newNo": r[2], "newRev": r[3],
             "reason": r[4], "date": r[5]}
            for r in cur.fetchall()
        ]


class SupersedeRequest(BaseModel):
    oldNo: str
    newNo: str
    reason: str = ""


@router.post("/drawings/supersedures", status_code=201, dependencies=[SETUP])
def create_supersedure(request: Request, body: SupersedeRequest) -> dict[str, Any]:
    """대체 등록 — old 도면을 new 도면으로 대체 (old 는 1회만, 409)."""
    if body.oldNo.strip() == body.newNo.strip():
        raise HTTPException(422, detail="구도면과 신도면이 같습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        old_id = _drawing_id(cur, tid, body.oldNo)
        new_id = _drawing_id(cur, tid, body.newNo)
        cur.execute("SELECT 1 FROM dwg_supersedure WHERE old_drawing_id=%s", (old_id,))
        if cur.fetchone():
            raise HTTPException(409, detail=f"이미 대체된 도면: {body.oldNo}")
        cur.execute(
            """INSERT INTO dwg_supersedure (tenant_id, old_drawing_id, new_drawing_id,
               reason, superseded_date, created_by)
               VALUES (%s,%s,%s,%s,CURRENT_DATE,%s) RETURNING supersedure_id""",
            (tid, old_id, new_id, body.reason.strip()[:500] or None, request.state.login))
        sid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_supersedure',%s,'CREATE',%s,%s)""",
            (tid, sid, request.state.user_id,
             json.dumps({"old": body.oldNo, "new": body.newNo, "reason": body.reason})))
    return {"supersedureId": sid}


# ── B16 도면 상세 — Variants·첨부·단계별 승인·블록·부품 관계 (dwg_* 도메인 완결) ──

@router.get("/drawings/{drawing_no}/variants")
def drawing_variants(drawing_no: str) -> list[dict[str, Any]]:
    """Variants — 동일 패밀리 도면 (도면번호 마지막 '-' 토큰 앞 접두 일치, 자신 제외)."""
    no = drawing_no.strip()
    prefix = no.rsplit("-", 1)[0] + "-" if "-" in no else no
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT d.drawing_no, d.drawing_name, d.current_rev, d.status,
                      EXISTS(SELECT 1 FROM dwg_supersedure s WHERE s.old_drawing_id=d.drawing_id)
               FROM dwg_drawing d
               WHERE d.tenant_id=%s AND d.drawing_no LIKE %s AND d.drawing_no<>%s
               ORDER BY d.drawing_no""", (tid, prefix + "%", no))
        return [{"drawingNo": r[0], "name": r[1], "rev": r[2], "status": r[3],
                 "superseded": r[4]} for r in cur.fetchall()]


@router.get("/drawings/{drawing_no}/files")
def drawing_files(drawing_no: str) -> list[dict[str, Any]]:
    """첨부 — 이 도면에 연결된 dwg_file (Run DXF·업로드본)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT file_id, file_name, file_type, COALESCE(file_size,0),
                      to_char(uploaded_date,'YYYY-MM-DD HH24:MI')
               FROM dwg_file WHERE tenant_id=%s AND drawing_id=%s
               ORDER BY file_id DESC""", (tid, did))
        return [{"fileId": r[0], "fileName": r[1], "fileType": r[2], "size": r[3],
                 "date": r[4]} for r in cur.fetchall()]


DWG_APPROVAL_STEPS = ("WRITE", "REVIEW", "APPROVE")


@router.get("/drawings/{drawing_no}/approvals")
def drawing_approvals(drawing_no: str) -> list[dict[str, Any]]:
    """단계별 승인 이력 — dwg_approval (WRITE→REVIEW→APPROVE)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT a.approval_id, a.step, a.result, COALESCE(a.comment,''),
                      to_char(a.approval_date,'YYYY-MM-DD'), u.user_name
               FROM dwg_approval a JOIN sys_user u ON u.user_id=a.approver_id
               WHERE a.drawing_id=%s ORDER BY a.approval_id""", (did,))
        return [{"approvalId": r[0], "step": r[1], "result": r[2], "comment": r[3],
                 "date": r[4], "by": r[5]} for r in cur.fetchall()]


class DwgApprovalCreate(BaseModel):
    step: str                 # WRITE | REVIEW | APPROVE
    approve: bool = True
    comment: str = ""


@router.post("/drawings/{drawing_no}/approvals", status_code=201, dependencies=[SETUP])
def drawing_approval_decide(drawing_no: str, request: Request, body: DwgApprovalCreate) -> dict[str, Any]:
    """단계 결정 — 순서 강제 (이전 단계 APPROVED 필요), 반려 시 도면 DRAFT 복귀.

    상태 전이: WRITE 승인→REVIEW · APPROVE 승인→APPROVED · 반려→DRAFT (이후 단계 재진행)."""
    step = body.step.strip().upper()
    if step not in DWG_APPROVAL_STEPS:
        raise HTTPException(422, detail=f"step 오류: {body.step} (WRITE|REVIEW|APPROVE)")
    if not body.approve and not body.comment.strip():
        raise HTTPException(422, detail="반려는 코멘트 필수")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT step FROM dwg_approval
               WHERE drawing_id=%s AND result='APPROVED'""", (did,))
        approved = {r[0] for r in cur.fetchall()}
        idx = DWG_APPROVAL_STEPS.index(step)
        if step in approved:
            raise HTTPException(409, detail=f"이미 승인된 단계: {step}")
        if idx > 0 and DWG_APPROVAL_STEPS[idx - 1] not in approved:
            raise HTTPException(409, detail=f"이전 단계({DWG_APPROVAL_STEPS[idx - 1]}) 승인 후 진행 가능")
        result = "APPROVED" if body.approve else "REJECTED"
        cur.execute(
            """INSERT INTO dwg_approval (drawing_id, step, approver_id, approval_date, result, comment)
               VALUES (%s,%s,%s,CURRENT_DATE,%s,NULLIF(%s,'')) RETURNING approval_id""",
            (did, step, request.state.user_id, result, body.comment.strip()[:1000]))
        approval_id = cur.fetchone()[0]
        new_status = None
        if not body.approve:
            new_status = "DRAFT"
            cur.execute("DELETE FROM dwg_approval WHERE drawing_id=%s AND result='APPROVED'", (did,))
        elif step == "WRITE":
            new_status = "REVIEW"
        elif step == "APPROVE":
            new_status = "APPROVED"
        if new_status:
            cur.execute(
                """UPDATE dwg_drawing SET status=%s, updated_by=%s, updated_at=now()
                   WHERE drawing_id=%s AND status<>'RELEASED'""",
                (new_status, request.state.login, did))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_approval',%s,'DECIDE',%s,%s)""",
            (tid, approval_id, request.state.user_id,
             json.dumps({"drawingNo": drawing_no, "step": step, "result": result})))
    return {"approvalId": approval_id, "step": step, "result": result,
            "drawingStatus": new_status}


@router.get("/drawings/{drawing_no}/blocks")
def drawing_blocks(drawing_no: str) -> list[dict[str, Any]]:
    """도면 블록 — dwg_document (Design Editor Block 패널 원천, content=좌표/치수 JSONB)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT document_id, COALESCE(block_name,''), content, origin_x, origin_y
               FROM dwg_document WHERE tenant_id=%s AND drawing_id=%s
               ORDER BY document_id""", (tid, did))
        return [
            {"documentId": r[0], "blockName": r[1], "content": r[2],
             "originX": float(r[3]) if r[3] is not None else None,
             "originY": float(r[4]) if r[4] is not None else None}
            for r in cur.fetchall()
        ]


@router.get("/drawings/{drawing_no}/relations")
def drawing_relations(drawing_no: str) -> list[dict[str, Any]]:
    """부품 관계 — dwg_part_relation (정렬·접촉 조건 + Macro 규칙, 부품 상세 실데이터)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT r.relation_id, r.block_a, r.block_b, COALESCE(r.condition_align,''),
                      COALESCE(r.condition_contact,''), m.macro_name, r.priority, r.approval_status
               FROM dwg_part_relation r
               LEFT JOIN tbx_macro m ON m.macro_id=r.value_macro_id
               WHERE r.tenant_id=%s AND r.drawing_id=%s ORDER BY r.priority, r.relation_id""",
            (tid, did))
        return [{"relationId": r[0], "blockA": r[1], "blockB": r[2], "align": r[3],
                 "contact": r[4], "macro": r[5], "priority": r[6], "status": r[7]}
                for r in cur.fetchall()]


# ── B17 부품 마스터 — prt_part·dwg_bom·prt_supplier_code_map·product_code_item ──

@router.get("/parts")
def parts_list() -> list[dict[str, Any]]:
    """부품 대장 — 재질·공급처·제품코드 조인 (M-4-7)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT p.part_id, p.part_no, p.part_name, COALESCE(p.specification,''),
                      m.material_code, c.company_name, pc.main_code, p.unit,
                      p.weight, p.is_standard,
                      (SELECT count(*) FROM dwg_bom b WHERE b.part_id=p.part_id),
                      to_char(COALESCE(p.updated_at,p.created_at),'{_VER_FMT}')
               FROM prt_part p
               LEFT JOIN mat_material m ON m.material_id=p.material_id
               LEFT JOIN com_company c ON c.company_id=p.supplier_id
               LEFT JOIN product_code pc ON pc.product_code_id=p.product_code_id
               WHERE p.tenant_id=%s ORDER BY p.part_no""", (tid,))
        return [
            {"partId": r[0], "partNo": r[1], "name": r[2], "spec": r[3],
             "material": r[4], "supplier": r[5], "productCode": r[6], "unit": r[7],
             "weight": float(r[8]) if r[8] is not None else None,
             "isStandard": r[9], "bomCount": r[10], "updatedAt": r[11]}   # D9 잠금 토큰
            for r in cur.fetchall()
        ]


@router.get("/parts/detail")
def part_detail(drawing: str = "KDCR 3-13", block: str = "") -> dict[str, Any]:
    """부품 상세 집계 (G3-b) — 도면 BOM 에서 블록명 매칭 부품 + 실 치수 + 공정.

    블록↔부품은 도면 BOM 내 이름 스코프 매칭(B17 패턴, 전역 아님). 미매칭 시 part=null(정직).
    치수 = dwg_dimension(도면 실 A/B/C…), 공정 = 부품 product_code 의 erp_work_process.
    """
    want = block.strip().lower()
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing)
        cur.execute(
            """SELECT b.item_no, b.quantity, b.assembly_seq, COALESCE(b.assembly_note,''),
                      p.part_no, p.part_name, COALESCE(p.specification,''),
                      m.material_code, c.company_name, p.weight, p.is_standard, p.product_code_id
               FROM dwg_bom b JOIN prt_part p ON p.part_id=b.part_id
               LEFT JOIN mat_material m ON m.material_id=p.material_id
               LEFT JOIN com_company c ON c.company_id=p.supplier_id
               WHERE b.drawing_id=%s ORDER BY b.item_no""", (did,))
        rows = cur.fetchall()
        part = None
        for r in rows:
            pn = (r[5] or "").lower()
            first = pn.split(" ")[0] if pn else ""
            if want and (pn.startswith(want.split(" ")[0]) or want in pn or (first and first in want)):
                part = r
                break
        if part is None and not want and rows:
            part = rows[0]

        process = None
        if part and part[11]:
            cur.execute(
                """SELECT process_type, workshop, person_count, skill_grade, work_time, make_or_buy
                   FROM erp_work_process WHERE tenant_id=%s AND product_code_id=%s
                   ORDER BY seq_no LIMIT 1""", (tid, part[11]))
            pr = cur.fetchone()
            if pr:
                process = {"process": pr[0], "workplace": pr[1] or "-", "person": pr[2] or 0,
                           "skill": pr[3] or "-", "wtimeHr": float(pr[4]) if pr[4] is not None else 0,
                           "makeBuy": pr[5] or "-"}

        cur.execute(
            """SELECT d.dim_label, d.dim_type, d.variant_value, mm.macro_expr
               FROM dwg_dimension d LEFT JOIN tbx_macro mm ON mm.macro_id=d.macro_id
               WHERE d.tenant_id=%s AND d.drawing_id=%s ORDER BY d.dim_label""", (tid, did))
        dims = [{"no": r[0], "value": str(r[3]) if r[3] else (str(r[2]) if r[2] is not None else ""),
                 "binding": "MACRO" if r[3] else "VARIANT", "kind": r[1]} for r in cur.fetchall()]

    part_obj = None
    if part:
        part_obj = {
            "partNo": part[4], "name": part[5], "spec": part[6],
            "material": part[7] or "-", "supplier": part[8] or "-",
            "weight": float(part[9]) if part[9] is not None else None, "isStandard": part[10],
            "makeBuy": process["makeBuy"] if process else "-",
            "assemblySeq": part[2], "assemblyNote": part[3], "qty": float(part[1]), "itemNo": part[0],
        }
    return {"drawing": drawing, "block": block, "part": part_obj, "dims": dims, "process": process}


class PartCreate(BaseModel):
    partNo: str
    name: str
    spec: str = ""
    materialCode: str = ""       # mat_material.material_code (없으면 미연결)
    supplier: str = ""           # com_company 이름 — 미존재 시 자동 생성 (단가 등록과 동일)
    productCode: str = ""        # product_code.main_code
    unit: str = "EA"
    weight: float | None = None
    isStandard: bool = False


@router.post("/parts", status_code=201, dependencies=[SETUP])
def part_create(request: Request, body: PartCreate) -> dict[str, Any]:
    no, name = body.partNo.strip(), body.name.strip()
    if not no or not name:
        raise HTTPException(422, detail="부품번호·부품명은 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT 1 FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, no))
        if cur.fetchone():
            raise HTTPException(409, detail=f"부품번호 중복: {no}")
        mid = None
        if body.materialCode.strip():
            cur.execute("SELECT material_id FROM mat_material WHERE tenant_id=%s AND material_code=%s",
                        (tid, body.materialCode.strip()))
            m = cur.fetchone()
            if not m:
                raise HTTPException(422, detail=f"재질 코드 없음: {body.materialCode} (M-3-2 등록 필요)")
            mid = m[0]
        sid = None
        if body.supplier.strip():
            cur.execute("SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s LIMIT 1",
                        (tid, body.supplier.strip()))
            s = cur.fetchone()
            if s:
                sid = s[0]
            else:
                cur.execute(
                    """INSERT INTO com_company (tenant_id, company_name, company_type)
                       VALUES (%s,%s,'SUPPLIER') RETURNING company_id""", (tid, body.supplier.strip()))
                sid = cur.fetchone()[0]
        pcid = None
        if body.productCode.strip():
            cur.execute("SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s LIMIT 1",
                        (tid, body.productCode.strip()))
            pc = cur.fetchone()
            if not pc:
                raise HTTPException(422, detail=f"제품코드 없음: {body.productCode}")
            pcid = pc[0]
        cur.execute(
            """INSERT INTO prt_part (tenant_id, part_no, part_name, specification, material_id,
               supplier_id, product_code_id, unit, weight, is_standard, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING part_id""",
            (tid, no, name[:200], body.spec.strip()[:300] or None, mid, sid, pcid,
             body.unit.strip()[:10] or "EA", body.weight, body.isStandard, request.state.login))
        part_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'prt_part',%s,'CREATE',%s,%s)""",
            (tid, part_id, request.state.user_id, json.dumps({"partNo": no, "name": name})))
    return {"partId": part_id, "partNo": no}


@router.post("/parts/import-excel", dependencies=[SETUP])
async def import_parts_excel(request: Request, uploadedFile: UploadFile = File(...)) -> dict[str, Any]:
    """부품 대량 등록 — 헤더: 부품번호·부품명·사양·단위·중량·공급처 (중복 부품번호 거부, 공급처 자동생성)."""
    ws, idx = _load_ws(await uploadedFile.read(), ["부품번호", "부품명"])
    inserted, rejected = 0, []
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            def cell(col: str) -> str:
                i = idx.get(col)
                v = row[i].value if i is not None else None
                return str(v).strip() if v is not None else ""
            no, name = cell("부품번호"), cell("부품명")
            if not no or not name:
                if no or name:
                    rejected.append(f"{r_i}행: 부품번호·부품명 필수")
                continue
            cur.execute("SELECT 1 FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, no))
            if cur.fetchone():
                rejected.append(f"{r_i}행 {no}: 중복"); continue
            weight = None
            wtxt = cell("중량")
            if wtxt:
                try:
                    weight = float(wtxt)
                except ValueError:
                    weight = None
            sid = None
            sup = cell("공급처")
            if sup:
                cur.execute("SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s LIMIT 1", (tid, sup))
                s = cur.fetchone()
                if s:
                    sid = s[0]
                else:
                    cur.execute("INSERT INTO com_company (tenant_id, company_name, company_type) "
                                "VALUES (%s,%s,'SUPPLIER') RETURNING company_id", (tid, sup[:200]))
                    sid = cur.fetchone()[0]
            cur.execute(
                """INSERT INTO prt_part (tenant_id, part_no, part_name, specification, supplier_id,
                   unit, weight, is_standard, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,TRUE,%s)""",
                (tid, no[:50], name[:200], cell("사양")[:300] or None, sid,
                 cell("단위")[:10] or "EA", weight, request.state.login))
            inserted += 1
    return {"inserted": inserted, "rejected": rejected, "rejectedCount": len(rejected)}


@router.delete("/parts/{part_no}", dependencies=[SETUP])
def part_delete(part_no: str, request: Request) -> dict[str, Any]:
    """부품 삭제 — BOM 참조 중이면 409 보호, 공급자 코드 매핑은 연쇄 정리."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT part_id FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, part_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"부품 없음: {part_no}")
        pid = row[0]
        cur.execute("SELECT 1 FROM dwg_bom WHERE part_id=%s LIMIT 1", (pid,))
        if cur.fetchone():
            raise HTTPException(409, detail=f"BOM 이 참조하는 부품은 삭제 불가: {part_no}")
        cur.execute("DELETE FROM prt_supplier_code_map WHERE part_id=%s", (pid,))
        cur.execute("DELETE FROM prt_part WHERE part_id=%s", (pid,))
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'prt_part',%s,'DELETE',%s,%s)""",
            (tid, pid, request.state.user_id, json.dumps({"partNo": part_no})))
    return {"deleted": part_no}


@router.get("/drawings/{drawing_no}/bom")
def drawing_bom(drawing_no: str) -> list[dict[str, Any]]:
    """도면 BOM — dwg_bom×prt_part (조립순서 assembly_seq 정렬, 부품 상세 ◆ 원천)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute(
            """SELECT b.bom_id, b.item_no, p.part_no, p.part_name, b.quantity,
                      b.assembly_seq, COALESCE(b.assembly_note,''), p.unit, p.is_standard
               FROM dwg_bom b JOIN prt_part p ON p.part_id=b.part_id
               WHERE b.drawing_id=%s
               ORDER BY COALESCE(b.assembly_seq, 999), b.item_no""", (did,))
        return [
            {"bomId": r[0], "itemNo": r[1], "partNo": r[2], "partName": r[3],
             "qty": float(r[4]), "assemblySeq": r[5], "assemblyNote": r[6],
             "unit": r[7], "isStandard": r[8]}
            for r in cur.fetchall()
        ]


class BomAdd(BaseModel):
    partNo: str
    qty: float = 1
    assemblySeq: int | None = None
    assemblyNote: str = ""


@router.post("/drawings/{drawing_no}/bom", status_code=201, dependencies=[SETUP])
def drawing_bom_add(drawing_no: str, request: Request, body: BomAdd) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute("SELECT part_id FROM prt_part WHERE tenant_id=%s AND part_no=%s",
                    (tid, body.partNo.strip()))
        p = cur.fetchone()
        if not p:
            raise HTTPException(422, detail=f"부품 없음: {body.partNo} (부품 대장 등록 필요)")
        cur.execute("SELECT 1 FROM dwg_bom WHERE drawing_id=%s AND part_id=%s", (did, p[0]))
        if cur.fetchone():
            raise HTTPException(409, detail=f"이미 BOM 에 있는 부품: {body.partNo}")
        cur.execute("SELECT COALESCE(max(item_no),0)+1 FROM dwg_bom WHERE drawing_id=%s", (did,))
        item_no = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO dwg_bom (drawing_id, part_id, item_no, quantity, assembly_seq, assembly_note)
               VALUES (%s,%s,%s,%s,%s,NULLIF(%s,'')) RETURNING bom_id""",
            (did, p[0], item_no, body.qty, body.assemblySeq, body.assemblyNote.strip()[:500]))
        bom_id = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO sys_history (tenant_id, target_table, target_id, action, actor_id, after_data)
               VALUES (%s,'dwg_bom',%s,'CREATE',%s,%s)""",
            (tid, bom_id, request.state.user_id,
             json.dumps({"drawingNo": drawing_no, "partNo": body.partNo, "qty": body.qty})))
    return {"bomId": bom_id, "itemNo": item_no}


@router.delete("/drawings/{drawing_no}/bom/{bom_id}", dependencies=[SETUP])
def drawing_bom_delete(drawing_no: str, bom_id: int) -> dict[str, Any]:
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        did = _drawing_id(cur, tid, drawing_no)
        cur.execute("DELETE FROM dwg_bom WHERE drawing_id=%s AND bom_id=%s RETURNING bom_id",
                    (did, bom_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"BOM 행 없음: #{bom_id}")
    return {"deleted": bom_id}


@router.get("/parts/{part_no}/supplier-codes")
def part_supplier_codes(part_no: str) -> list[dict[str, Any]]:
    """공급자 코드 매핑 (ERP-018) — 부품 기준."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT m.map_id, c.company_name, m.supplier_code, COALESCE(m.supplier_name,'')
               FROM prt_supplier_code_map m
               JOIN prt_part p ON p.part_id=m.part_id
               JOIN com_company c ON c.company_id=m.supplier_id
               WHERE m.tenant_id=%s AND p.part_no=%s ORDER BY m.map_id""", (tid, part_no))
        return [{"mapId": r[0], "supplier": r[1], "supplierCode": r[2], "supplierName": r[3]}
                for r in cur.fetchall()]


class SupplierCodeAdd(BaseModel):
    supplier: str
    supplierCode: str
    supplierName: str = ""


@router.post("/parts/{part_no}/supplier-codes", status_code=201, dependencies=[SETUP])
def part_supplier_code_add(part_no: str, request: Request, body: SupplierCodeAdd) -> dict[str, Any]:
    if not body.supplier.strip() or not body.supplierCode.strip():
        raise HTTPException(422, detail="공급처·공급자 코드는 필수입니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT part_id FROM prt_part WHERE tenant_id=%s AND part_no=%s", (tid, part_no))
        p = cur.fetchone()
        if not p:
            raise HTTPException(404, detail=f"부품 없음: {part_no}")
        cur.execute("SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s LIMIT 1",
                    (tid, body.supplier.strip()))
        s = cur.fetchone()
        if s:
            sid = s[0]
        else:
            cur.execute("""INSERT INTO com_company (tenant_id, company_name, company_type)
                           VALUES (%s,%s,'SUPPLIER') RETURNING company_id""", (tid, body.supplier.strip()))
            sid = cur.fetchone()[0]
        cur.execute(
            """INSERT INTO prt_supplier_code_map (tenant_id, part_id, supplier_id, supplier_code,
               supplier_name, created_by) VALUES (%s,%s,%s,%s,NULLIF(%s,''),%s) RETURNING map_id""",
            (tid, p[0], sid, body.supplierCode.strip()[:50], body.supplierName.strip()[:200],
             request.state.login))
        return {"mapId": cur.fetchone()[0]}


@router.get("/codes/{code}/supplier-codes")
def code_supplier_codes(code: str) -> list[dict[str, Any]]:
    """공급자 코드 매핑 — 제품코드 기준 (발주 문서 표시용, 부품 경유 포함)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT m.map_id, c.company_name, m.supplier_code, COALESCE(m.supplier_name,'')
               FROM prt_supplier_code_map m
               JOIN com_company c ON c.company_id=m.supplier_id
               LEFT JOIN prt_part p ON p.part_id=m.part_id
               LEFT JOIN product_code pc
                 ON pc.product_code_id=COALESCE(m.product_code_id, p.product_code_id)
               WHERE m.tenant_id=%s AND pc.main_code=%s ORDER BY m.map_id""", (tid, code))
        return [{"mapId": r[0], "supplier": r[1], "supplierCode": r[2], "supplierName": r[3]}
                for r in cur.fetchall()]


@router.get("/codes/{code}/slot-items")
def code_slot_items(code: str) -> list[dict[str, Any]]:
    """제품코드 필수 슬롯 정의 — product_code_item×code_item (C-1 슬롯 구성의 DB 원천)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT i.pc_item_id, i.item_slot, ci.item_name, i.is_required, i.sort_order
               FROM product_code_item i
               JOIN product_code pc ON pc.product_code_id=i.product_code_id AND pc.tenant_id=%s
               JOIN code_item ci ON ci.item_id=i.source_item_id
               WHERE pc.main_code=%s ORDER BY i.sort_order, i.item_slot""", (tid, code))
        return [{"pcItemId": r[0], "slot": r[1], "itemName": r[2], "required": r[3],
                 "sortOrder": r[4]} for r in cur.fetchall()]


@router.get("/codes/{code}/approval-history")
def code_approval_history(code: str) -> list[dict[str, Any]]:
    """코드 승인 이력 — sys_approval_request 실조회 (코드 상세 CODE_APPROVAL_HIST mock 대체).
    product_code 직접 target + comment 라벨 매칭 (범용 승인 요청의 코드 연결)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
            (tid, code))
        pc = cur.fetchone()
        cur.execute(
            """SELECT a.request_type, a.result, COALESCE(a.comment,''),
                      to_char(a.requested_at,'YYYY-MM-DD'), to_char(a.decided_at,'YYYY-MM-DD'),
                      ru.user_name, au.user_name
               FROM sys_approval_request a
               JOIN sys_user ru ON ru.user_id=a.requester_id
               LEFT JOIN sys_user au ON au.user_id=a.approver_id
               WHERE a.tenant_id=%s AND ((a.target_table='product_code' AND a.target_id=%s)
                     OR a.comment ILIKE %s)
               ORDER BY a.approval_id""",
            (tid, pc[0] if pc else -1, f"%{code}%"))
        rows: list[dict[str, Any]] = []
        for rt, result, comment, req_d, dec_d, req_by, app_by in cur.fetchall():
            rows.append({"date": req_d, "action": f"승인 요청 ({rt})", "by": req_by, "note": comment})
            if result:
                label = "승인 (APPROVED)" if result == "APPROVED" else f"반려 ({result})"
                rows.append({"date": dec_d or req_d, "action": label, "by": app_by or "-", "note": ""})
        return rows


# ══ F5 — 마스터 데이터 수정·정정 전면 (등록만 있고 Update 가 없던 도메인) ══


class CompanyPatch(BaseModel):
    name: str | None = None
    companyType: str | None = None
    nation: str | None = None
    grade: str | None = None
    terms: str | None = None
    remarks: str | None = None
    active: bool | None = None      # 거래처 비활성/재활성 (소프트 — 선택 리스트 제외)


@router.put("/companies/{company_id}", dependencies=[SETUP])
def update_company(company_id: int, body: CompanyPatch, request: Request) -> dict[str, Any]:
    """공급처·거래처 수정 (F5) — M-14-2 행 더블클릭."""
    if body.companyType and body.companyType not in ("CUSTOMER", "SUPPLIER", "PARTNER", "BANK"):
        raise HTTPException(422, detail=f"유형 오류: {body.companyType}")
    if body.name is not None and not body.name.strip():
        raise HTTPException(422, detail="업체명은 비울 수 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT company_name FROM com_company WHERE tenant_id=%s AND company_id=%s",
                    (tid, company_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"company not found: {company_id}")
        if body.name and body.name.strip() != row[0]:
            cur.execute(
                """SELECT 1 FROM com_company WHERE tenant_id=%s AND company_name=%s
                   AND company_id<>%s""", (tid, body.name.strip(), company_id))
            if cur.fetchone():
                raise HTTPException(409, detail=f"업체명 중복: {body.name.strip()}")
        fields = {
            "company_name": body.name.strip() if body.name is not None else None,
            "company_type": body.companyType,
            "nation": body.nation,
            "evaluation_grade": body.grade,
            "payment_terms": body.terms,
            "remarks": body.remarks,
        }
        sets = {k: v for k, v in fields.items() if v is not None}
        if body.active is not None:            # bool 은 None-필터에서 걸러지므로 별도 처리
            sets["is_active"] = body.active
        if not sets:
            raise HTTPException(422, detail="수정할 필드가 없습니다")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE com_company SET {assign}, updated_by=%s, updated_at=now() WHERE company_id=%s",
            (*sets.values(), request.state.login, company_id))
        _audit(cur, tid, "com_company", company_id, "COMPANY_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"name": row[0]})
    return {"companyId": company_id, "updated": sorted(sets)}


class CompanyBatch(BaseModel):
    ids: list[int]
    active: bool           # True=재활성 · False=비활성


@router.post("/companies/batch", dependencies=[SETUP])
def batch_company(request: Request, body: CompanyBatch) -> dict[str, Any]:
    """거래처 일괄 활성/비활성 (그리드 다중 선택) — 소프트, 최대 200건."""
    ids = list(dict.fromkeys(body.ids))[:200]
    if not ids:
        raise HTTPException(422, detail="대상이 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "UPDATE com_company SET is_active=%s, updated_by=%s, updated_at=now() "
            "WHERE tenant_id=%s AND company_id = ANY(%s) RETURNING company_id",
            (body.active, request.state.login, tid, ids))
        changed = [r[0] for r in cur.fetchall()]
        _audit(cur, tid, "com_company", 0, "COMPANY_BATCH", request.state.user_id,
               {"active": body.active, "count": len(changed), "ids": changed[:50]})
    return {"active": body.active, "done": len(changed), "requested": len(ids)}


class PartPatch(BaseModel):
    name: str | None = None
    specification: str | None = None
    materialCode: str | None = None      # 빈 문자열 = 연결 해제
    supplier: str | None = None          # 빈 문자열 = 연결 해제, 미존재 시 자동 생성
    code: str | None = None              # 제품코드 연결, 빈 문자열 = 해제
    weight: float | None = None
    isStandard: bool | None = None
    baseUpdatedAt: str = ""              # D9 — 낙관적 잠금 (불일치 409)


@router.put("/parts/{part_no}", dependencies=[SETUP])
def update_part(part_no: str, body: PartPatch, request: Request) -> dict[str, Any]:
    """부품 속성 수정 (F5) — M-4-7 (등록·삭제만 있던 도메인). D9 낙관적 잠금."""
    if body.name is not None and not body.name.strip():
        raise HTTPException(422, detail="부품명은 비울 수 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            f"""SELECT part_id, part_name, to_char(COALESCE(updated_at,created_at),'{_VER_FMT}')
               FROM prt_part WHERE tenant_id=%s AND part_no=%s""", (tid, part_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"part not found: {part_no}")
        if body.baseUpdatedAt.strip() and body.baseUpdatedAt.strip() != row[2]:
            raise HTTPException(
                409, detail="다른 사용자가 먼저 수정했습니다 — 재조회 후 다시 시도하십시오 (동시 편집 충돌)")
        sets: dict[str, Any] = {}
        if body.name is not None:
            sets["part_name"] = body.name.strip()
        if body.specification is not None:
            sets["specification"] = body.specification.strip() or None
        if body.weight is not None:
            sets["weight"] = body.weight
        if body.isStandard is not None:
            sets["is_standard"] = body.isStandard
        if body.materialCode is not None:
            if body.materialCode.strip():
                cur.execute(
                    "SELECT material_id FROM mat_material WHERE tenant_id=%s AND material_code=%s",
                    (tid, body.materialCode.strip()))
                m = cur.fetchone()
                if not m:
                    raise HTTPException(422, detail=f"재질 없음: {body.materialCode} (M-3-2 등록 필요)")
                sets["material_id"] = m[0]
            else:
                sets["material_id"] = None
        if body.supplier is not None:
            if body.supplier.strip():
                cur.execute(
                    "SELECT company_id FROM com_company WHERE tenant_id=%s AND company_name=%s",
                    (tid, body.supplier.strip()))
                c = cur.fetchone()
                if c:
                    sets["supplier_id"] = c[0]
                else:
                    cur.execute(
                        """INSERT INTO com_company (tenant_id, company_type, company_name)
                           VALUES (%s,'SUPPLIER',%s) RETURNING company_id""",
                        (tid, body.supplier.strip()))
                    sets["supplier_id"] = cur.fetchone()[0]
            else:
                sets["supplier_id"] = None
        if body.code is not None:
            if body.code.strip():
                cur.execute(
                    "SELECT product_code_id FROM product_code WHERE tenant_id=%s AND main_code=%s",
                    (tid, body.code.strip()))
                pc = cur.fetchone()
                if not pc:
                    raise HTTPException(422, detail=f"제품코드 없음: {body.code}")
                sets["product_code_id"] = pc[0]
            else:
                sets["product_code_id"] = None
        if not sets:
            raise HTTPException(422, detail="수정할 필드가 없습니다")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE prt_part SET {assign}, updated_by=%s, updated_at=now() WHERE part_id=%s",
            (*sets.values(), request.state.login, row[0]))
        _audit(cur, tid, "prt_part", row[0], "PART_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"name": row[1]})
    return {"partNo": part_no, "updated": sorted(sets)}


class MaterialPatch(BaseModel):
    name: str | None = None
    materialType: str | None = None
    density: float | None = None
    standard: str | None = None
    hazard: str | None = None


@router.put("/materials/{code}", dependencies=[SETUP])
def update_material(code: str, body: MaterialPatch, request: Request) -> dict[str, Any]:
    """재질 수정 (F5) — M-3-2."""
    if body.name is not None and not body.name.strip():
        raise HTTPException(422, detail="재질명은 비울 수 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT material_id, material_name FROM mat_material WHERE tenant_id=%s AND material_code=%s",
            (tid, code))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"material not found: {code}")
        fields = {
            "material_name": body.name.strip() if body.name is not None else None,
            "material_type": body.materialType,
            "density": body.density,
            "standard": body.standard,
            "hazard_class": body.hazard,
        }
        sets = {k: v for k, v in fields.items() if v is not None}
        if not sets:
            raise HTTPException(422, detail="수정할 필드가 없습니다")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE mat_material SET {assign}, updated_by=%s, updated_at=now() WHERE material_id=%s",
            (*sets.values(), request.state.login, row[0]))
        _audit(cur, tid, "mat_material", row[0], "MATERIAL_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"name": row[1]})
    return {"code": code, "updated": sorted(sets)}


class VerificationPatch(BaseModel):
    ruleName: str | None = None
    warningMessage: str | None = None
    isActive: bool | None = None


@router.put("/verifications/{verification_id}", dependencies=[SETUP])
def update_verification(verification_id: int, body: VerificationPatch,
                        request: Request) -> dict[str, Any]:
    """검증 규칙 수정·비활성 (F5) — M-4-5."""
    if body.ruleName is not None and not body.ruleName.strip():
        raise HTTPException(422, detail="규칙명은 비울 수 없습니다")
    fields = {
        "rule_name": body.ruleName.strip() if body.ruleName is not None else None,
        "warning_message": body.warningMessage,
        "is_active": body.isActive,
    }
    sets = {k: v for k, v in fields.items() if v is not None}
    if not sets:
        raise HTTPException(422, detail="수정할 필드가 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT rule_name FROM dwg_verification WHERE tenant_id=%s AND verification_id=%s",
            (tid, verification_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"verification not found: {verification_id}")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE dwg_verification SET {assign}, updated_by=%s, updated_at=now() "
            "WHERE verification_id=%s",
            (*sets.values(), request.state.login, verification_id))
        _audit(cur, tid, "dwg_verification", verification_id, "VERIFY_UPDATE",
               request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"name": row[0]})
    return {"verificationId": verification_id, "updated": sorted(sets)}


class ValuePatch(BaseModel):
    valueName: str | None = None
    description: str | None = None
    deprecate: bool = False
    approve: bool = False   # #28 — 미승인 값 승인 (조합 대상 편입)


@router.patch("/codes/values/{value_id}", dependencies=[SETUP])
def patch_code_value(value_id: int, body: ValuePatch, request: Request) -> dict[str, Any]:
    """Variant 값 수정·폐기 (F5) — S-1-2. 폐기 = approval_status DEPRECATED."""
    sets: dict[str, Any] = {}
    if body.valueName is not None:
        sets["value_name"] = body.valueName.strip()
    if body.description is not None:
        sets["description"] = body.description.strip() or None
    if body.deprecate:
        sets["approval_status"] = "DEPRECATED"
    if body.approve:
        if body.deprecate:
            raise HTTPException(422, detail="승인과 폐기를 동시에 지정할 수 없습니다")
        sets["approval_status"] = "APPROVED"
    if not sets:
        raise HTTPException(422, detail="수정할 필드가 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT value_code, approval_status FROM code_item_value WHERE tenant_id=%s AND value_id=%s",
            (tid, value_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"value not found: {value_id}")
        assign = ", ".join(f"{k}=%s" for k in sets)
        # #28 — 내용이 바뀌면 Sub Code Revision 을 올린다. 이미 고정된 제품 코드 조합은
        # 생성 시점 Revision 을 보존하므로, 조합 상세에서 revDrift 로 드러난다(조합 자체는 불변).
        bump = ", revision_no=revision_no+1" if ("value_name" in sets or "description" in sets) else ""
        cur.execute(
            f"UPDATE code_item_value SET {assign}{bump}, updated_by=%s, updated_at=now() WHERE value_id=%s",
            (*sets.values(), request.state.login, value_id))
        _audit(cur, tid, "code_item_value", value_id, "VALUE_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()},
               before={"code": row[0], "status": row[1]})
    return {"valueId": value_id, "updated": sorted(sets)}


class WarehousePatch(BaseModel):
    name: str | None = None
    hazard: str | None = None
    inspection: str | None = None
    remarks: str | None = None


@router.patch("/erp/warehouses/{code}", dependencies=[SETUP])
def patch_warehouse(code: str, body: WarehousePatch, request: Request) -> dict[str, Any]:
    """창고 노드 개명·속성 수정 (F5) — M-8-4 (code 는 불변)."""
    if body.name is not None and not body.name.strip():
        raise HTTPException(422, detail="위치명은 비울 수 없습니다")
    fields = {
        "location_name": body.name.strip() if body.name is not None else None,
        "hazard_allowed": body.hazard,
        "inspection_cycle": body.inspection,
        "remarks": body.remarks,
    }
    sets = {k: v for k, v in fields.items() if v is not None}
    if not sets:
        raise HTTPException(422, detail="수정할 필드가 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT warehouse_id, location_name FROM erp_warehouse WHERE tenant_id=%s AND location_code=%s",
            (tid, code))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"warehouse not found: {code}")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE erp_warehouse SET {assign}, updated_by=%s, updated_at=now() WHERE warehouse_id=%s",
            (*sets.values(), request.state.login, row[0]))
        _audit(cur, tid, "erp_warehouse", row[0], "WH_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"name": row[1]})
    return {"code": code, "updated": sorted(sets)}


class PriceClose(BaseModel):
    validTo: str          # YYYY-MM-DD — 적용 종료일 마감(정정)


@router.patch("/prices/{price_id}", dependencies=[SETUP])
def close_price(price_id: int, body: PriceClose, request: Request) -> dict[str, Any]:
    """단가 적용 종료일 마감 (F5) — M-12-5 (잘못 등록한 단가 정정 동선)."""
    try:
        to = date.fromisoformat(body.validTo.strip())
    except ValueError:
        raise HTTPException(422, detail=f"날짜 형식 오류: {body.validTo} (YYYY-MM-DD)")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT valid_from, valid_to FROM cst_price WHERE tenant_id=%s AND price_id=%s",
            (tid, price_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"price not found: {price_id}")
        if to < row[0]:
            raise HTTPException(422, detail=f"종료일이 적용 시작({row[0]}) 이전입니다")
        cur.execute(
            "UPDATE cst_price SET valid_to=%s, updated_by=%s, updated_at=now() WHERE price_id=%s",
            (to, request.state.login, price_id))
        _audit(cur, tid, "cst_price", price_id, "PRICE_CLOSE", request.state.user_id,
               after={"validTo": body.validTo},
               before={"validTo": row[1].isoformat() if row[1] else None})
    return {"priceId": price_id, "validTo": body.validTo}


class DocMetaPatch(BaseModel):
    title: str | None = None
    docType: str | None = None
    grade: str | None = None


@router.patch("/documents/{doc_no}/meta", dependencies=[SETUP])
def patch_document_meta(doc_no: str, body: DocMetaPatch, request: Request) -> dict[str, Any]:
    """문서 메타(제목·유형·Grade) 수정 (F5) — ACCEPTED(승인 완료) 문서는 409 통제."""
    if body.title is not None and not body.title.strip():
        raise HTTPException(422, detail="제목은 비울 수 없습니다")
    if body.grade and body.grade not in ("S-1", "S-2", "S-3", "S-4"):
        raise HTTPException(422, detail=f"Grade 오류: {body.grade}")
    fields = {
        "title": body.title.strip() if body.title is not None else None,
        "doc_type": body.docType,
        "management_grade": body.grade,
    }
    sets = {k: v for k, v in fields.items() if v is not None}
    if not sets:
        raise HTTPException(422, detail="수정할 필드가 없습니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT doc_control_id, released_status, title FROM doc_control
               WHERE tenant_id=%s AND doc_no=%s
               ORDER BY doc_control_id DESC LIMIT 1""", (tid, doc_no))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"document not found: {doc_no}")
        if row[1] == "ACCEPTED":
            raise HTTPException(409, detail="승인 완료(ACCEPTED) 문서는 메타 수정 불가 — 개정 절차 사용")
        assign = ", ".join(f"{k}=%s" for k in sets)
        cur.execute(
            f"UPDATE doc_control SET {assign}, updated_by=%s, updated_at=now() WHERE doc_control_id=%s",
            (*sets.values(), request.state.login, row[0]))
        _audit(cur, tid, "doc_control", row[0], "DOC_META_UPDATE", request.state.user_id,
               after={k: str(v) for k, v in sets.items()}, before={"title": row[2]})
    return {"docNo": doc_no, "updated": sorted(sets)}


@router.delete("/templets/{name}", dependencies=[SETUP])
def delete_templet(name: str, request: Request) -> dict[str, Any]:
    """Templet 삭제 (F5) — 시스템·게시(RELEASED) 보호 409."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            "SELECT templet_id, approval_status, is_system FROM tbx_templet "
            "WHERE tenant_id=%s AND templet_name=%s", (tid, name))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"templet not found: {name}")
        if row[2]:
            raise HTTPException(409, detail="시스템 Templet 은 삭제 불가")
        if row[1] == "RELEASED":
            raise HTTPException(409, detail="게시(RELEASED) Templet 은 삭제 불가 — 개정 절차 사용")
        cur.execute("DELETE FROM tbx_templet WHERE templet_id=%s", (row[0],))
        _audit(cur, tid, "tbx_templet", row[0], "TEMPLET_DELETE", request.state.user_id,
               after={"name": name})
    return {"deleted": name}


@router.delete("/macros/{name}", dependencies=[SETUP])
def delete_macro(name: str, request: Request) -> dict[str, Any]:
    """Macro 삭제 (F5) — 치수식·검증 규칙·구성 join 참조 보호 409 (Studio F3 실배선)."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute("SELECT macro_id FROM tbx_macro WHERE tenant_id=%s AND macro_name=%s",
                    (tid, name))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"macro not found: {name}")
        mid = row[0]
        for sql, label in (
            ("SELECT count(*) FROM dwg_dimension WHERE macro_id=%s", "치수식"),
            ("SELECT count(*) FROM dwg_verification WHERE macro_id=%s", "검증 규칙"),
            ("SELECT count(*) FROM arrangement_component WHERE join_macro_id=%s", "구성 join"),
        ):
            cur.execute(sql, (mid,))
            c = cur.fetchone()[0]
            if c:
                raise HTTPException(409, detail=f"참조 존재 — {label} {c}건 (삭제 불가)")
        cur.execute("DELETE FROM tbx_macro_ref WHERE macro_id=%s", (mid,))
        cur.execute("DELETE FROM tbx_macro WHERE macro_id=%s", (mid,))
        _audit(cur, tid, "tbx_macro", mid, "MACRO_DELETE", request.state.user_id,
               after={"name": name})
    return {"deleted": name}


class ComponentPatch(BaseModel):
    quantity: float


@router.patch("/arrangements/{code}/components/{component_id}", dependencies=[SETUP])
def patch_arrangement_component(code: str, component_id: int, body: ComponentPatch,
                                request: Request) -> dict[str, Any]:
    """Arrangement 구성품 수량 수정 (F5) — M-4-2."""
    if body.quantity <= 0:
        raise HTTPException(422, detail="수량은 0보다 커야 합니다")
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """SELECT c.component_id, c.quantity FROM arrangement_component c
               JOIN arrangement_code a ON a.arrangement_id=c.arrangement_id
               WHERE a.tenant_id=%s AND a.arrangement_code=%s AND c.component_id=%s""",
            (tid, code, component_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, detail=f"component not found: {component_id}")
        cur.execute("UPDATE arrangement_component SET quantity=%s WHERE component_id=%s",
                    (body.quantity, component_id))
        _audit(cur, tid, "arrangement_component", component_id, "ARR_COMP_UPDATE",
               request.state.user_id,
               after={"quantity": body.quantity}, before={"quantity": float(row[1])})
    return {"componentId": component_id, "quantity": body.quantity}


@router.delete("/arrangements/{code}/components/{component_id}", dependencies=[SETUP])
def delete_arrangement_component(code: str, component_id: int, request: Request) -> dict[str, Any]:
    """Arrangement 구성품 삭제 (F5) — M-4-2."""
    with _conn() as conn, conn.cursor() as cur:
        tid = _tenant_id(cur)
        cur.execute(
            """DELETE FROM arrangement_component c
               USING arrangement_code a
               WHERE a.arrangement_id=c.arrangement_id AND a.tenant_id=%s
                 AND a.arrangement_code=%s AND c.component_id=%s
               RETURNING c.component_id""", (tid, code, component_id))
        if not cur.fetchone():
            raise HTTPException(404, detail=f"component not found: {component_id}")
        _audit(cur, tid, "arrangement_component", component_id, "ARR_COMP_DELETE",
               request.state.user_id, after={"arrangement": code})
    return {"deleted": component_id}
