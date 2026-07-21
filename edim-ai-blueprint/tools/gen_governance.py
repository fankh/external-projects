# -*- coding: utf-8 -*-
"""거버넌스 정의서 생성기 (요구 #71) — 코드에서 뽑아 쓴다.

권한 Matrix·상태 흐름·경계 정의는 손으로 쓰면 쓰는 순간부터 낡는다. 여기서는 **실제로
강제되는 것**만 뽑는다.
  · 권한 Matrix : 라우터 데코레이터의 dependencies=[SETUP|ADMIN|PLATFORM]
  · 상태 흐름   : alembic 의 CHECK (... IN (...)) 로 선언된 상태 집합 +
                  라우터의 SET status='X' 로 실제 일어나는 전이
  · 경계 정의표 : 기준/Runtime Snapshot/실행/분석 계층 분류 (ERP 기준 문서 §4.2)
  · 용어사전    : 위에서 등장한 테이블·상태의 정의

사용:
  py tools/gen_governance.py            # docs/EDIM_거버넌스정의서.xlsx 재생성
  py tools/gen_governance.py --print    # 추출 결과만 출력(비교·점검용)
"""
from __future__ import annotations

import pathlib
import re
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
ROUTER = ROOT / "backend" / "app" / "routers" / "edim.py"
ALEMBIC = ROOT / "backend" / "alembic" / "versions"
OUT = ROOT / "docs" / "EDIM_거버넌스정의서.xlsx"

LEVELS = ("PLATFORM", "ADMIN", "SETUP")

# ERP 기준 문서 §4.2 — 기준 Data / Runtime Snapshot / 실행 Data / 분석 Data 는 섞지 않는다(§10.9).
LAYER_RULES = [
    ("기준 Data", "승인되어 재사용되는 제품·업무 기준. 직접 변경하지 않고 승인·개정으로 다룬다.",
     ("code_group", "code_item", "code_item_value", "product_code", "code_relationship",
      "arrangement", "tbx_", "erp_domain_catalog", "erp_process_catalog",
      "erp_workflow_", "sys_head", "sys_hierarchy", "sys_process_node", "customer_company")),
    ("Runtime Snapshot", "특정 Project 에서 사용된 고정 결과. 재실행해도 같은 값이어야 한다.",
     ("sys_snapshot", "cpq_run", "dwg_run_job", "sys_setup_version")),
    ("실행 Data", "ERP/MES/WMS/QMS 에서 실제 진행된 업무 결과.",
     ("erp_", "inv_", "qc_", "prj_", "eco_", "prt_", "com_", "doc_control", "work_order")),
    ("분석 Data", "실행 Data 를 집계한 상태. 기준 Data 를 바꾸지 않는다.",
     ("cst_", "dwg_text_index")),
    ("기준 참조", "달력·환율·세율 등 업무 계산의 공통 참조값.",
     ("cal_", "fx_", "tax_")),
    ("통제·감사", "권한·승인·점유·감사 등 시스템 통제 자체의 기록.",
     ("sys_", "customer_logo_asset")),
    ("개발 관리", "제품 자체의 개발·요구 관리 기록(운영 업무 데이터가 아님).",
     ("dev_",)),
]


def _layer(table: str) -> tuple[str, str]:
    for name, desc, prefixes in LAYER_RULES:
        if any(table.startswith(p) or table == p for p in prefixes):
            return name, desc
    return "미분류", "계층이 지정되지 않음 — 분류 규칙 보완 필요"


def extract_routes() -> list[tuple[str, str, str, str]]:
    """(method, path, required_level, handler) — 데코레이터에서 그대로 읽는다."""
    src = ROUTER.read_text(encoding="utf-8")
    rows: list[tuple[str, str, str, str]] = []
    pat = re.compile(
        r"@router\.(get|post|put|patch|delete)\(\s*[\"']([^\"']+)[\"']([^)]*)\)\s*\n"
        r"(?:async\s+)?def\s+(\w+)", re.S)
    for m in pat.finditer(src):
        method, path, extra, fn = m.group(1).upper(), m.group(2), m.group(3), m.group(4)
        level = "GENERAL"
        for lv in LEVELS:
            if re.search(r"dependencies=\[[^\]]*\b" + lv + r"\b", extra):
                level = lv
                break
        rows.append((method, path, level, fn))
    return sorted(rows, key=lambda r: (r[1], r[0]))


def extract_states() -> list[tuple[str, str, str, str, str, str]]:
    """상태 흐름 — CHECK 로 선언된 집합, DEFAULT 초기값, UPDATE 로 일어나는 전이를 나눠 싣는다.

    주의: 여기서 읽는 전이는 **라우터의 UPDATE ... SET 리터럴**뿐이다. INSERT 로만 놓이는
    상태나 변수로 넘어가는 상태는 잡히지 않으므로, 'UPDATE 경로 없음' 을 곧 '도달 불가'로
    읽어서는 안 된다. 그렇게 단정하면 초기 상태(DEFAULT)까지 미사용으로 잘못 몰린다."""
    states: dict[tuple[str, str], set[str]] = {}
    defaults: dict[tuple[str, str], str] = {}
    for f in sorted(ALEMBIC.glob("*.py")):
        txt = f.read_text(encoding="utf-8")
        for tm in re.finditer(r"CREATE TABLE IF NOT EXISTS (\w+)(.*?)\n\s*\)\"\"\"", txt, re.S):
            table, body = tm.group(1), tm.group(2)
            for cm in re.finditer(r"CHECK\s*\(\s*(\w+)\s+IN\s*\(([^)]*)\)", body):
                col = cm.group(1)
                vals = {v.strip().strip("'") for v in cm.group(2).split(",") if v.strip()}
                states.setdefault((table, col), set()).update(vals)
            for dm in re.finditer(r"^\s*(\w+)\s+VARCHAR\([^)]*\)[^,\n]*DEFAULT\s+'(\w+)'",
                                  body, re.M):
                defaults[(table, dm.group(1))] = dm.group(2)

    src = ROUTER.read_text(encoding="utf-8")
    transitions: dict[str, set[str]] = {}
    for um in re.finditer(r"UPDATE\s+(\w+)\s+SET\s+([^\n]*)", src):
        table, tail = um.group(1), um.group(2)
        for sm in re.finditer(r"(\w*status|result)\s*=\s*'(\w+)'", tail):
            transitions.setdefault(f"{table}.{sm.group(1)}", set()).add(sm.group(2))

    rows = []
    for (table, col), vals in sorted(states.items()):
        applied = sorted(transitions.get(f"{table}.{col}", set()))
        init = defaults.get((table, col), "")
        rest = sorted(vals - set(applied) - ({init} if init else set()))
        rows.append((
            table, col,
            # 허용 상태는 **집합**이다. '→' 로 이으면 있지도 않은 순서를 읽게 되므로 쓰지 않는다.
            " / ".join(sorted(vals)),
            init or "(DEFAULT 없음)",
            ", ".join(applied) or "(없음)",
            ", ".join(rest) or "-",
        ))
    return rows


def extract_tables() -> list[str]:
    seen: list[str] = []
    for f in sorted(ALEMBIC.glob("*.py")):
        for tm in re.finditer(r"CREATE TABLE IF NOT EXISTS (\w+)",
                              f.read_text(encoding="utf-8")):
            if tm.group(1) not in seen:
                seen.append(tm.group(1))
    return seen


def build() -> dict[str, list[list[str]]]:
    routes = extract_routes()
    states = extract_states()
    tables = extract_tables()

    perm = [["Method", "경로", "필요 등급", "핸들러"]]
    perm += [[m, p, lv, fn] for m, p, lv, fn in routes]

    flow = [["테이블", "컬럼", "허용 상태(CHECK·집합)", "초기값(DEFAULT)",
             "UPDATE 로 일어나는 전이", "UPDATE 경로가 확인되지 않은 상태"]]
    flow += [list(r) for r in states]

    bound = [["테이블", "계층", "계층의 의미"]]
    for t in tables:
        layer, desc = _layer(t)
        bound.append([t, layer, desc])

    glossary = [["용어", "구분", "정의"]]
    glossary += [
        ["Tenant", "구조", "회사 단위 데이터 경계. 모든 업무 객체는 tenant_id 를 가진다."],
        ["GENERAL / SETUP / ADMIN / PLATFORM", "권한",
         "사용 등급. 상위 등급은 하위 등급의 권한을 포함한다(LEVEL_RANK)."],
        ["Snapshot", "구조",
         "특정 시점의 결과를 고정한 기록. 같은 Snapshot 이면 재실행 결과가 같아야 한다."],
        ["Set-up Version", "구조",
         "게시 시점 Set-up 내용의 지문. 이후 변경은 drift 로 드러난다."],
        ["Work Lock", "통제",
         "자원 단위 편집 점유. 다른 사용자가 점유 중이면 보유자·만료를 밝히고 거부한다."],
        ["PENDING / APPROVED / REJECTED / SUPERSEDED", "상태",
         "승인 대상의 공통 상태. 표시·사용은 APPROVED 만 허용한다."],
        ["DRAFT / PUBLISHED", "상태",
         "편집 중 / 게시됨. 게시 시점에 성립 조건을 강제한다."],
        ["미설정 = 허용", "원칙",
         "선언되지 않은 대상은 막지 않는다. 통제를 새로 넣어도 기존 운영이 멈추지 않게 한다."],
        ["선언 후 게시 강제", "원칙",
         "등록·편집 단계는 불완전해도 허용하고, 게시 시점에 성립 여부를 따진다."],
    ]
    method = [["항목", "내용"]]
    method += [
        ["생성 방법", "이 문서는 손으로 쓰지 않고 tools/gen_governance.py 가 코드에서 뽑는다. "
                  "tests/check_governance.py 가 코드와 어긋나면 실패시킨다."],
        ["권한Matrix 근거", "backend/app/routers/edim.py 의 @router 데코레이터와 "
                        "dependencies=[SETUP|ADMIN|PLATFORM]. 등급 표기가 없으면 GENERAL(로그인만)."],
        ["상태흐름 근거", "alembic 의 CHECK (... IN (...)) 와 컬럼 DEFAULT, "
                     "그리고 라우터의 UPDATE ... SET <컬럼>='값' 리터럴."],
        ["상태흐름 한계", "INSERT 로만 놓이거나 변수로 전달되는 상태는 잡히지 않는다. "
                     "'UPDATE 경로가 확인되지 않은 상태' 를 '도달 불가' 로 읽으면 안 된다."],
        ["경계정의 근거", "ERP 기준 문서 §4.2 의 계층 구분을 테이블 접두사 규칙으로 적용한 것. "
                     "규칙은 tools/gen_governance.py 의 LAYER_RULES 에 있다."],
        ["용어사전 근거", "코드에서 반복적으로 강제되는 개념만 싣는다(장식적 정의는 넣지 않는다)."],
    ]
    return {"생성기준": method, "권한Matrix": perm, "상태흐름": flow,
            "경계정의": bound, "용어사전": glossary}


def write_xlsx(data: dict[str, list[list[str]]]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    for sheet, rows in data.items():
        ws = wb.create_sheet(sheet)
        for r in rows:
            ws.append(r)
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for i, _ in enumerate(rows[0], start=1):
            width = max((len(str(r[i - 1])) for r in rows if len(r) >= i), default=10)
            ws.column_dimensions[chr(64 + i)].width = min(max(width + 2, 12), 70)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.font = Font(size=10)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    d = build()
    if "--print" in sys.argv:
        for k, rows in d.items():
            print(f"== {k}: {len(rows) - 1}행")
        sys.exit(0)
    write_xlsx(d)
    print(f"생성 — {OUT.relative_to(ROOT)}")
    for k, rows in d.items():
        print(f"  {k}: {len(rows) - 1}행")
