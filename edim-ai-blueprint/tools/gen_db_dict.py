# -*- coding: utf-8 -*-
"""DB 정의서 생성기 (18.2) — 실제 스키마에서 뽑고, 사람 설명은 병합한다.

배경: 종전 `docs/EDIM_DB정의서.xlsx` 는 설계 초안 MD(v0.1, 2026-07-07)에서 만든 것이라
**54 테이블·516 컬럼**만 담고 있었다. 실제 스키마는 그 사이 alembic 59 리비전을 거쳐
**107 테이블·1,181 컬럼**이 되었고, 문서에는 그 절반이 아예 **행조차 없었다**(빈 칸이 아니라
행이 없으므로 읽는 사람은 누락을 알 수 없다 — 17.2 에서 거버넌스 정의서가 같은 상태였다).
'제약' 열도 전부 비어 있어, 17.1 에서 새로 건 상태 CHECK 가 문서에 반영될 자리가 없었다.

그래서 **스키마는 실 DB 에서** 읽고(information_schema + pg_constraint), **설명은 MD 에서**
병합한다. 설명이 없는 항목은 비워 두지 않고 '미기재' 시트에 모아 몇 건인지 드러낸다 —
문서가 조용히 절반만 설명하는 상태로 돌아가지 않게 하려는 것이다.

사용:
  py tools/gen_db_dict.py --dump      # 서버에서 스키마 스냅샷 갱신 (ssh 필요)
  py tools/gen_db_dict.py             # 스냅샷 + MD → docs/EDIM_DB정의서.xlsx 재생성
  py tools/gen_db_dict.py --print     # 수치만 출력(점검용)

스냅샷(`docs/ddl/schema_snapshot.json`)을 커밋해 두므로 생성·검사는 오프라인에서도 된다.
서버가 스냅샷과 어긋나면 --dump 로 갱신하는 것이 정답이며, 그 차이 자체가 신호다.
"""
from __future__ import annotations

import json
import pathlib
import re
import subprocess
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
MD = ROOT / "docs" / "EDIM_DB_정의서.md"
SNAP = ROOT / "docs" / "ddl" / "schema_snapshot.json"
# 사람이 쓴 설명 사전. 설계 MD 만 긁으면 종전 문서에 있던 설명 60건이 사라졌다 —
# **재생성이 정보를 잃으면 안 된다**. 사전을 커밋해 두고 MD 보다 우선해 병합한다.
# 새 테이블·컬럼 설명은 여기에 채워 넣는다(설명을 늘리는 유일한 자리).
DESC = ROOT / "docs" / "ddl" / "column_descriptions.json"
OUT = ROOT / "docs" / "EDIM_DB정의서.xlsx"

# 도메인 접두어 — MD §1.2 의 규칙을 그대로 옮긴다(문서와 코드가 같은 표를 보게 한다)
DOMAINS = [
    ("sys_", "시스템 공통 — 사용자·권한·Hierarchy·승인·이력"),
    ("code_", "RCCS 코드 — Sub/Product/Relationship/Arrangement"),
    ("prj_", "프로젝트"),
    ("dwg_", "도면/PLM — 도면·개정·치수·부품관계·검증"),
    ("prt_", "부품"),
    ("mat_", "재질"),
    ("cpq_", "CPQ — Selection·Run·산출물"),
    ("tbl_", "데이터 Table — Variant/Tech/Material Table"),
    ("tbx_", "Toolbox — UI Form·Macro·Templet·Print Form"),
    ("cst_", "원가/견적 — 단가·원가계산·PCR·견적"),
    ("erp_", "ERP 프로세스"),
    ("com_", "회사 마스터"),
    ("inv_", "재고"),
    ("qc_", "품질"),
    ("eco_", "설계변경"),
    ("doc_", "문서 통제"),
    ("cal_", "달력·휴일"),
    ("fx_", "환율"),
    ("tax_", "세율"),
    ("ai_", "AI 데이터 준비(제품 개발용)"),
    ("dev_", "개발 관리(제품 자체)"),
    ("customer_", "고객사 자산(로고 등)"),
    ("arrangement", "Arrangement"),
    ("work_order", "작업지시"),
    ("product_code", "RCCS 코드 — Product Code"),
]

_DUMP_SQL = """COPY (SELECT replace(json_build_object(
 'columns', (SELECT json_agg(to_jsonb(q)) FROM (
    SELECT table_name AS t, ordinal_position AS pos, column_name AS c,
           data_type AS ty, character_maximum_length AS len,
           numeric_precision AS np, numeric_scale AS ns,
           is_nullable AS nullable, column_default AS def
      FROM information_schema.columns
     WHERE table_schema='public' AND table_name<>'alembic_version'
     ORDER BY table_name, ordinal_position) q),
 'constraints', (SELECT json_agg(to_jsonb(q)) FROM (
    SELECT co.conname AS name, co.contype AS ty, cl.relname AS t,
           (SELECT string_agg(a.attname, ',' ORDER BY x.ord)
              FROM unnest(co.conkey) WITH ORDINALITY AS x(attnum, ord)
              JOIN pg_attribute a ON a.attrelid=cl.oid AND a.attnum=x.attnum) AS cols,
           (SELECT rl.relname FROM pg_class rl WHERE rl.oid=co.confrelid) AS ref,
           pg_get_constraintdef(co.oid) AS def
      FROM pg_constraint co JOIN pg_class cl ON cl.oid=co.conrelid
      JOIN pg_namespace n ON n.oid=cl.relnamespace
     WHERE n.nspname='public' ORDER BY cl.relname, co.contype, co.conname) q)
)::text, chr(10), ' ')) TO STDOUT"""


def dump() -> None:
    """서버 스키마 → 스냅샷 파일. psql 의 json_agg 는 원소 사이에 개행을 넣으므로 서버에서
    미리 없앤다(COPY 텍스트 형식이 개행을 \\n 두 글자로 escape 해 파싱이 깨진다)."""
    cmd = ["ssh", "edim-server",
           f'sudo docker exec edim-postgres psql -U edim -d edim -Atc "{_DUMP_SQL}"']
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=180,
                       encoding="utf-8", errors="replace")
    out = r.stdout or ""
    if "{" not in out:
        print(f"FAIL — 스키마 덤프 실패: {(r.stderr or out)[:200]}")
        raise SystemExit(1)
    data = json.loads(out[out.index("{"):])
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    SNAP.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"스냅샷 저장 — 테이블 {len({c['t'] for c in data['columns']})} · "
          f"컬럼 {len(data['columns'])} · 제약 {len(data['constraints'])}")


def md_descriptions() -> tuple[dict[str, str], dict[tuple[str, str], str]]:
    """설계 MD 의 표에서 테이블·컬럼 설명을 긁는다.

    MD 는 사람이 쓴 설계 문서이고 형식이 완전히 규칙적이지 않다. 그래서 **긁히는 것만**
    쓰고, 못 긁은 것은 '미기재' 로 남긴다 — 추측으로 채우면 문서가 그럴듯해지면서 틀린다.
    """
    if not MD.exists():
        return {}, {}
    tbl: dict[str, str] = {}
    col: dict[tuple[str, str], str] = {}
    cur_tbl = ""
    for line in MD.read_text(encoding="utf-8").splitlines():
        # "### 2.1 `code_group` — 제품군" 같은 제목에서 테이블과 설명을 얻는다
        m = re.match(r"^#{2,4}\s.*?`(\w+)`\s*[—\-–:]\s*(.+)$", line)
        if m:
            cur_tbl = m.group(1)
            tbl.setdefault(cur_tbl, m.group(2).strip())
            continue
        m = re.match(r"^#{2,4}\s.*?`(\w+)`\s*$", line)
        if m:
            cur_tbl = m.group(1)
            continue
        # 컬럼 표: | `col` | TYPE | ... | 설명 |
        if cur_tbl and line.startswith("|"):
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 3:
                continue
            cm = re.match(r"^`?(\w+)`?$", cells[0])
            if not cm or cells[0].lower() in ("컬럼", "항목", "대상"):
                continue
            desc = cells[-1].strip()
            if desc and not re.match(r"^[-–—]*$", desc) and desc != cells[0]:
                col.setdefault((cur_tbl, cm.group(1)), desc[:200])
    # 사전이 MD 보다 우선한다 — 사전은 검토를 거친 문장이고, MD 긁기는 형식 추정이다.
    if DESC.exists():
        d = json.loads(DESC.read_text(encoding="utf-8"))
        tbl.update({k: v for k, v in d.get("tables", {}).items() if v})
        for key, v in d.get("columns", {}).items():
            if "." in key and v:
                t, _, c = key.partition(".")
                col[(t, c)] = v
    return tbl, col


def _type_text(c: dict) -> str:
    ty, ln, np_, ns = c["ty"], c["len"], c["np"], c["ns"]
    if ty in ("character varying", "character"):
        return f"VARCHAR({ln})" if ln else "VARCHAR"
    if ty == "numeric":
        return f"NUMERIC({np_},{ns})" if np_ is not None else "NUMERIC"
    return {"timestamp with time zone": "TIMESTAMPTZ", "bigint": "BIGINT",
            "integer": "INTEGER", "smallint": "SMALLINT", "boolean": "BOOLEAN",
            "text": "TEXT", "date": "DATE", "jsonb": "JSONB", "bytea": "BYTEA",
            "double precision": "DOUBLE PRECISION"}.get(ty, ty.upper())


def _domain(table: str) -> str:
    for pref, desc in DOMAINS:
        if table.startswith(pref):
            return desc
    return "(분류 없음)"


def build() -> dict[str, list[list]]:
    if not SNAP.exists():
        print(f"FAIL — 스냅샷 없음: {SNAP.relative_to(ROOT)} → py tools/gen_db_dict.py --dump")
        raise SystemExit(1)
    snap = json.loads(SNAP.read_text(encoding="utf-8"))
    cols = snap["columns"]
    cons = snap["constraints"]
    tbl_desc, col_desc = md_descriptions()

    # 컬럼별 제약 표기 — PK/FK/UQ/CHECK 을 한 칸에 모은다
    marks: dict[tuple[str, str], list[str]] = {}
    checks: list[list] = []
    for k in cons:
        cl = (k["cols"] or "").split(",") if k["cols"] else []
        for c in cl:
            key = (k["t"], c)
            if k["ty"] == "p":
                marks.setdefault(key, []).append("PK")
            elif k["ty"] == "f":
                marks.setdefault(key, []).append(f"FK→{k['ref']}")
            elif k["ty"] == "u":
                marks.setdefault(key, []).append("UQ")
            elif k["ty"] == "x":
                marks.setdefault(key, []).append("EXCLUDE")
        if k["ty"] == "c":
            # pg_get_constraintdef 은 괄호를 겹쳐 낸다: CHECK (((col)::text = ANY (...)))
            # `\(\(?` 로는 3중 괄호를 못 넘어 대상 컬럼을 통째로 놓쳤다.
            m = re.search(r"CHECK \(+(\w+)", k["def"] or "")
            target = m.group(1) if m else ""
            if target:
                marks.setdefault((k["t"], target), []).append("CHECK")
            vals = re.findall(r"'([^']+)'::", k["def"] or "")
            checks.append([k["t"], target or "(복합)", k["name"],
                           " / ".join(dict.fromkeys(vals)) or "(집합 아님)",
                           (k["def"] or "")[:160]])

    by_tbl: dict[str, list[dict]] = {}
    for c in cols:
        by_tbl.setdefault(c["t"], []).append(c)

    info = [["EDIM DB정의서", ""],
            ["생성 방식", "실 스키마(information_schema·pg_constraint) 스냅샷 + 설계 MD 설명 병합 "
                       "— tools/gen_db_dict.py"],
            ["스냅샷", str(SNAP.relative_to(ROOT)).replace("\\", "/")],
            ["대상 DBMS", "PostgreSQL 16 (운영 실측 — 종전 문서의 '가정 미확정' 표기 정정)"],
            ["테이블 수", f"{len(by_tbl)}개 (설계 MD 기재 {len(tbl_desc)}개)"],
            ["컬럼 수", f"{len(cols)}개"],
            ["제약 수", f"{len(cons)}개 (CHECK {sum(1 for k in cons if k['ty']=='c')} · "
                      f"FK {sum(1 for k in cons if k['ty']=='f')} · "
                      f"UNIQUE {sum(1 for k in cons if k['ty']=='u')})"],
            ["문자 인코딩", "UTF-8"],
            ["시간대", "저장은 timestamptz(UTC), 업무 판정은 EDIM_TZ(기본 Asia/Seoul)"],
            ["설명 출처", "설계 MD 에서 긁힌 것만 싣는다. 못 긁은 것은 추측으로 채우지 않고 "
                       "'미기재' 시트에 모은다 — 그럴듯한 오답보다 공백이 낫다."]]

    tables = [["No", "도메인", "테이블명", "테이블 설명", "컬럼 수", "제약 수", "설명 출처"]]
    for i, t in enumerate(sorted(by_tbl), start=1):
        ncon = sum(1 for k in cons if k["t"] == t)
        tables.append([i, _domain(t), t, tbl_desc.get(t, ""), len(by_tbl[t]), ncon,
                       "설계 MD" if t in tbl_desc else "미기재"])

    coldefs = [["테이블명", "No", "컬럼명", "타입", "Null", "키/제약", "기본값", "설명"]]
    missing: list[list] = [["구분", "테이블명", "컬럼명", "비고"]]
    for t in sorted(by_tbl):
        if t not in tbl_desc:
            missing.append(["테이블", t, "", "설계 MD 에 설명 없음(alembic 로 추가된 테이블 등)"])
        for c in by_tbl[t]:
            key = (t, c["c"])
            d = col_desc.get(key, "")
            coldefs.append([t, c["pos"], c["c"], _type_text(c),
                            "NULL" if c["nullable"] == "YES" else "NOT NULL",
                            " · ".join(dict.fromkeys(marks.get(key, []))),
                            (c["def"] or "")[:60], d])
            if not d:
                missing.append(["컬럼", t, c["c"], "설계 MD 에 설명 없음"])

    checks.sort(key=lambda r: (r[0], r[1]))
    return {"문서정보": info, "테이블목록": tables, "컬럼정의": coldefs,
            "제약집합": [["테이블명", "컬럼", "제약명", "허용 값(집합)", "정의"]] + checks,
            "미기재": missing}


def write_xlsx(data: dict[str, list[list]]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F3864")
    hf = Font(color="FFFFFF", bold=True, size=10)
    for sheet, rows in data.items():
        ws = wb.create_sheet(sheet)
        for r in rows:
            ws.append(r)
        for c in ws[1]:
            c.fill = fill
            c.font = hf
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for i in range(1, len(rows[0]) + 1):
            w = max((len(str(r[i - 1])) for r in rows if len(r) >= i), default=10)
            ws.column_dimensions[chr(64 + i)].width = min(max(w + 2, 10), 60)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.font = Font(size=10)
    wb.save(OUT)


if __name__ == "__main__":
    if "--dump" in sys.argv:
        dump()
        raise SystemExit(0)
    d = build()
    if "--print" in sys.argv:
        for k, rows in d.items():
            print(f"== {k}: {len(rows) - 1}행")
        raise SystemExit(0)
    write_xlsx(d)
    print(f"생성 — {OUT.relative_to(ROOT)}")
    for k, rows in d.items():
        print(f"  {k}: {len(rows) - 1}행")
