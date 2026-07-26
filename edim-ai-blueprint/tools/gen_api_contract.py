# -*- coding: utf-8 -*-
"""API 계약 문서 생성기 (18.6) — 실제로 배포된 스펙에서 뽑는다.

배경: `docs/EDIM_컴포넌트정의서.xlsx` 의 API목록(108건)과 그것에서 만든
`docs/api/edim-openapi.yaml`(95경로)은 **2026-07-07 설계 초안**이고, 구현은 그 뒤로 갈라졌다.
대조해 보니 **108건 중 라우터에 그대로 있는 것은 42건**뿐이었다:

  · 일치 42 · 메서드만 다름 9 · 경로 상이 30 · 없는 것 27

문서는 "대표 엔드포인트" 라고 범위를 밝혔지만, 적힌 것이 **호출되지 않는다**는 사실은 밝히지
않았다. 누락보다 나쁘다 — 이 문서를 보고 연동하는 사람은 404 를 만난다. 실제 스펙은
358 경로·446 오퍼레이션이다.

그래서 **as-built(실제 배포된 스펙)를 문서의 기준으로 삼고**, 설계 의도는 별도 시트에
대조 상태와 함께 남긴다(의도를 지우지 않되, 구현으로 착각하게 두지도 않는다).

사용:
  py tools/gen_api_contract.py --dump    # 배포 컨테이너의 app.openapi() → 스냅샷 (ssh 필요)
  py tools/gen_api_contract.py           # 스냅샷 → 컴포넌트정의서 API 시트 재생성
  py tools/gen_api_contract.py --print   # 수치만
"""
from __future__ import annotations

import json
import pathlib
import re
import subprocess
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
SNAP = ROOT / "docs" / "api" / "openapi_asbuilt.json"
DESIGN = ROOT / "docs" / "api" / "design_api_list.json"
OUT = ROOT / "docs" / "EDIM_컴포넌트정의서.xlsx"
METHODS = ("get", "post", "put", "patch", "delete")

# 경로 접두어 → 서비스 구분. 설계 문서의 SVC 번호 체계를 유지해 두 시트를 나란히 읽게 한다.
SVC = [
    ("/auth", "SVC-01 Auth"), ("/users", "SVC-01 Auth"), ("/platform", "SVC-01 Auth"),
    ("/i18n", "SVC-01 Auth"), ("/access", "SVC-01 Auth"), ("/roles", "SVC-01 Auth"),
    ("/heads", "SVC-02 Hierarchy"), ("/hierarchy", "SVC-02 Hierarchy"),
    ("/codes", "SVC-03 Code"), ("/arrangements", "SVC-03 Code"),
    ("/drawings", "SVC-04 Drawing"), ("/cad", "SVC-04 Drawing"),
    ("/tables", "SVC-05 Table"),
    ("/toolbox", "SVC-06 Toolbox"), ("/macros", "SVC-06 Toolbox"),
    ("/cpq", "SVC-07 CPQ"),
    ("/cost", "SVC-08 Cost"), ("/prices", "SVC-08 Cost"), ("/parts", "SVC-08 Cost"),
    ("/erp", "SVC-09 ERP"), ("/companies", "SVC-09 ERP"), ("/inventory", "SVC-09 ERP"),
    ("/qc", "SVC-09 ERP"), ("/projects", "SVC-09 ERP"),
    ("/approvals", "SVC-10 Approval"),
    ("/documents", "SVC-11 Print"), ("/render", "SVC-11 Print"), ("/reports", "SVC-11 Print"),
    ("/files", "SVC-12 File"),
    ("/notifications", "SVC-13 Notify"), ("/prefs", "SVC-13 Notify"),
    ("/ai", "AI"), ("/assistant", "AI"),
    ("/search", "INT 통합검색"), ("/history", "INT 감사"), ("/audit", "INT 감사"),
    ("/system", "INT 운영"), ("/health", "INT 운영"), ("/dev", "개발 관리"),
    ("/setup", "SVC-06 Toolbox"), ("/tenant", "SVC-01 Auth"),
    # 아래는 as-built 에만 있는 영역 — 설계 초안의 SVC 체계에 자리가 없어 새로 분류했다.
    ("/anomalies", "INT 이상감지"), ("/locks", "INT 점유·잠금"),
    ("/snapshots", "SVC-07 CPQ"), ("/verifications", "SVC-04 Drawing"),
    ("/eco", "SVC-04 Drawing"), ("/templets", "SVC-06 Toolbox"),
    ("/materials", "SVC-08 Cost"), ("/finance", "SVC-08 Cost"),
    ("/calendar", "기준 참조"), ("/process", "SVC-09 ERP"),
    ("/customers", "SVC-09 ERP"), ("/support", "INT 운영"),
    ("/config", "INT 운영"), ("/settings", "INT 운영"), ("/metrics", "INT 운영"),
    ("/menu", "SVC-01 Auth"),
]


def _svc(path: str) -> str:
    # /api/v1 을 붙이지 않은 구형 경로가 아직 마운트돼 있다. 분류로 덮어 감추지 않고
    # 별도 구분으로 드러낸다 — 규약(§3 REST /api/v1)에서 벗어난 표면이기 때문이다(18.6).
    if path.startswith("/api/") and not path.startswith("/api/v1"):
        return "구형 /api (v1 규약 밖)"
    p = path.replace("/api/v1", "")
    for pref, name in SVC:
        if p.startswith(pref):
            return name
    return "(미분류)"


def dump() -> None:
    """배포된 백엔드의 app.openapi() 를 그대로 받는다 — 소스 추정이 아니라 실물이다."""
    cmd = ["ssh", "edim-server",
           'sudo docker exec edim-backend python -c "'
           "import json;from app.main import app;print(json.dumps(app.openapi(),ensure_ascii=False))"
           '"']
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=180,
                       encoding="utf-8", errors="replace")
    out = r.stdout or ""
    if "{" not in out:
        print(f"FAIL — 스펙 덤프 실패: {(r.stderr or out)[:200]}")
        raise SystemExit(1)
    spec = json.loads(out[out.index("{"):])
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    SNAP.write_text(json.dumps(spec, ensure_ascii=False, indent=1), encoding="utf-8")
    ops = sum(len([m for m in v if m in METHODS]) for v in spec["paths"].values())
    print(f"as-built 스펙 저장 — 경로 {len(spec['paths'])} · 오퍼레이션 {ops} · "
          f"스키마 {len(spec.get('components', {}).get('schemas', {}))}")


def _norm(p: str) -> str:
    return re.sub(r"\{[^}]+\}", "{}", p.replace("/api/v1", "").rstrip("/")) or "/"


def build() -> dict[str, list[list]]:
    if not SNAP.exists():
        print(f"FAIL — 스냅샷 없음: {SNAP.relative_to(ROOT)} → --dump")
        raise SystemExit(1)
    spec = json.loads(SNAP.read_text(encoding="utf-8"))
    live: list[tuple[str, str, str]] = []      # (method, path, summary)
    for path, ops in sorted(spec["paths"].items()):
        for m in METHODS:
            if m in ops:
                op = ops[m]
                live.append((m.upper(), path,
                             (op.get("summary") or op.get("description") or "").split("\n")[0]))
    liveset = {(m, _norm(p)) for m, p, _ in live}
    livepaths = {_norm(p) for _, p, _ in live}

    api = [["No", "서비스", "Method", "경로", "요약"]]
    for i, (m, p, s) in enumerate(
            sorted(live, key=lambda x: (_svc(x[1]), x[1], x[0])), start=1):
        api.append([i, _svc(p), m, p, s[:160]])

    # 설계 초안 대조 — 의도를 지우지 않되 구현으로 착각하게 두지 않는다
    cmp_rows = [["No", "서비스(설계)", "Method", "설계 경로", "구현 상태", "비고"]]
    design = json.loads(DESIGN.read_text(encoding="utf-8"))["apis"] if DESIGN.exists() else []
    tails: dict[str, list[str]] = {}
    for m, p, _ in live:
        tails.setdefault(_norm(p).rsplit("/", 1)[-1], []).append(f"{m} {_norm(p)}")
    for i, d in enumerate(design, start=1):
        m, p = d["method"], _norm(d["path"])
        if (m, p) in liveset:
            state, note = "구현됨", ""
        elif p in livepaths:
            state = "메서드 상이"
            note = "같은 경로가 다른 메서드로 존재 — 연동 시 메서드를 확인해야 한다"
        else:
            cand = tails.get(p.rsplit("/", 1)[-1]) or []
            if cand:
                state, note = "경로 상이", f"유사 후보: {', '.join(cand[:2])}"
            else:
                state, note = "미구현", "이 경로로 호출하면 404 — 설계 의도만 남아 있다"
        cmp_rows.append([i, d["svc"], m, d["path"], state, note])
    return {"API목록": api, "설계-구현 대조": cmp_rows}


def write_sheets(data: dict[str, list[list]]) -> None:
    """컴포넌트정의서의 API 관련 시트만 갈아끼운다 — 다른 시트(문서정보·컴포넌트목록)는
    사람이 쓴 것이므로 건드리지 않는다."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.load_workbook(OUT)
    fill = PatternFill("solid", fgColor="1F3864")
    hf = Font(color="FFFFFF", bold=True, size=10)
    for sheet, rows in data.items():
        if sheet in wb.sheetnames:
            idx = wb.sheetnames.index(sheet)
            wb.remove(wb[sheet])
            ws = wb.create_sheet(sheet, idx)
        else:
            ws = wb.create_sheet(sheet)
        for r in rows:
            ws.append(r)
        for c in ws[1]:
            c.fill = fill
            c.font = hf
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for i in range(1, len(rows[0]) + 1):
            w = max((len(str(r[i - 1])) for r in rows if len(r) >= i), default=10)
            ws.column_dimensions[chr(64 + i)].width = min(max(w + 2, 10), 62)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.font = Font(size=10)
    wb.save(OUT)


if __name__ == "__main__":
    if "--dump" in sys.argv:
        dump()
        raise SystemExit(0)
    d = build()
    if "--print" in sys.argv:
        for k, rows in d.items():
            print(f"== {k}: {len(rows) - 1}행")
        raise SystemExit(0)
    write_sheets(d)
    print(f"생성 — {OUT.relative_to(ROOT)}")
    for k, rows in d.items():
        print(f"  {k}: {len(rows) - 1}행")
    from collections import Counter
    c = Counter(r[4] for r in d["설계-구현 대조"][1:])
    print("  설계 대조:", dict(c))
