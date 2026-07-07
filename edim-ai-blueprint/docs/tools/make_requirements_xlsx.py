# -*- coding: utf-8 -*-
"""EDIM 요구사항정의서.xlsx 생성.

기준: EDIM_기능정의서 v0.2 (178기능) · 개요서 · 컴포넌트정의서 — 요구사항 수준으로 상향 집계
실행: py docs/tools/make_requirements_xlsx.py  (저장: docs/02_요구사항/EDIM_요구사항정의서.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "02_요구사항", "EDIM_요구사항정의서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
GRP_FILLS = ["FFF8E7", "E8F4FD", "F0FFF0", "FFF0F5", "F5F0FF", "FFEFC2", "E0F7FA",
             "FDF2E9", "EAF2F8", "F9EBEA", "F4ECF7", "E8F8F5", "FEF9E7", "EBEDEF"]

# ---------------------------------------------------------------- 기능 요구사항
# (REQ ID, 구분, 요구사항명, 상세 설명, 중요도, Phase, 관련 기능 ID, 비고)
FUNC = [
 ("시스템 공통", [
  ("REQ-F-001", "사용자 인증·세션", "ID/PW 로그인, 토큰 갱신, 실패 잠금, 세션 만료 관리", "필수", "P1", "SYS-001/002", ""),
  ("REQ-F-002", "역할 기반 권한", "PLATFORM/ADMIN/SETUP/GENERAL 4단계 + Head Tab·Hierarchy·기능·Table 단위 권한, 권한 없는 항목 미표시", "필수", "P1", "SYS-003~005", ""),
  ("REQ-F-003", "Hierarchy·DB Address", "제품/일반/설정 Tree 관리, Materialized Address 발급, 이동·개명 시 데이터 연결 추적, 심볼·검색", "필수", "P1", "SYS-006~008", "개요 §4"),
  ("REQ-F-004", "공통 승인 워크플로우", "모든 Set-up 자산의 DRAFT→PENDING→APPROVED 전이, 승인함, 승인 이력", "필수", "P1", "SYS-009~011", ""),
  ("REQ-F-005", "변경 이력·감사", "생성·편집·삭제·승인의 작업자/일시/전후 데이터 기록·조회", "필수", "P2", "SYS-012", ""),
  ("REQ-F-006", "Main Work Frame", "Head Tab + 3판넬(좌 Tree·중앙 작업·우 Templet) 공통 화면, 일정·To-do·알림·업무 SNS·개인화", "필수", "P1~", "SYS-013~020", ""),
 ]),
 ("RCCS 코드", [
  ("REQ-F-007", "Sub Code 등록", "코드 그룹·자릿수(Slot)·값 등록, 중복검토, Table 참조 값, Excel Import/Export, 코드별 DWG/자료 연결", "필수", "P1", "CODE-001~005/013~016", ""),
  ("REQ-F-008", "Product Code 관리", "Main Code 등록(중복검토)·자릿수 구성(Sub Code 연결)", "필수", "P1", "CODE-006/007", ""),
  ("REQ-F-009", "Code Relationship", "Mother-Child 관계 + 수량 + Slot 매핑(match_condition), Part List Running Test 검증", "필수", "P1", "CODE-008/009", "관계형 BOM 핵심"),
  ("REQ-F-010", "BOM 자동 전개", "완성품 Code → 재귀 전개 → 전체 Part List (깊이 제한·순환 차단)", "필수", "P1", "CODE-010", ""),
  ("REQ-F-011", "Arrangement", "표준 구성 코드 등록, 구성품·결합조건(Macro)·Fan Direction 설정", "필수", "P2", "CODE-011/012", ""),
 ]),
 ("도면·PLM", [
  ("REQ-F-012", "도면 관리", "도면 등록(도번·유형·축척), 개정(Rev), 승인(작성→검토→승인→발행)", "필수", "P2", "DWG-001/013/014", ""),
  ("REQ-F-013", "Drawing Editor", "CAD형 캔버스 — 부품 Drag 배치, CAD 명령(복사·회전·Trim·Block 등)·스냅·측정, 도면 Templet, Block 단위 저장", "필수", "P2", "DWG-002~005/012/022/025", "슬라이드 61 명령표 = 수용 기준"),
  ("REQ-F-014", "파라메트릭 치수", "치수별 Macro/Variant 바인딩, 지시선 자동번호(A~N), KEY/DETAIL 구분, 치수↔부품 동기화, 설계·자료 우선순위(순환 점검)", "필수", "P2", "DWG-006~008", ""),
  ("REQ-F-015", "부품 관계·검증", "Block 간 결합 조건(수직·수평·접촉 등)+관계값 Macro, 설계 검증 규칙(경고), 조립 순서", "필수", "P2", "DWG-009~011/024", ""),
  ("REQ-F-016", "도면 Import/Export", "DXF/DWG(ODA)/PDF/STEP Import → DrawingDocument, DXF(R2010)/PDF Export, CAD Mapping", "필수", "P2", "DWG-015/016/023", ""),
  ("REQ-F-017", "부품·재질·PLM 조회", "부품/재질 마스터, Referencers(역참조)·Variants·Supersedure(대체) 관리", "필수", "P2~P3", "DWG-017~020/026/027", ""),
 ]),
 ("데이터 Table", [
  ("REQ-F-018", "데이터 Table 관리", "Variant/Tech/Material Table 정의·행 관리·Excel Import·Macro 참조 범위 조회·승인 (Constant·단위계 포함)", "필수", "P1", "TBL-001~006", "Constant/Metric 기능 v0.3 보강"),
 ]),
 ("Toolbox", [
  ("REQ-F-019", "UI Designer", "위젯 팔레트 Drag&Drop 폼 빌더, 동작 Templet 바인딩, 버전·승인·게시, 동적 렌더링", "필수", "P4", "TBX-001~004/013", ""),
  ("REQ-F-020", "Macro 엔진", "Excel 호환 문법 파서·평가(Table/Var/PreC 참조·공학 함수), Test Run, 버전·승인", "필수", "P1~P2", "TBX-006/011/012", "문법 특허 검토 연동"),
  ("REQ-F-021", "4-Way Sync", "Prompt↔Macro↔Flowchart↔Coding 실시간 상호 변환 (Description 동기)", "필수", "P4", "TBX-005/007/008", "핵심 차별점"),
  ("REQ-F-022", "지원 도구", "함수·그래프 마법사, Data Information Call, 기능 찾기(자연어), 인터넷 검색·AI 질의", "선택", "P4~P5", "TBX-009/010/014/015", ""),
 ]),
 ("CPQ", [
  ("REQ-F-023", "제품 선정", "Product Tree 선택, Arrangement 모듈 Drag 조합(분할·통합·3각법 View), 사양 입력, DWG View", "필수", "P2", "CPQ-001~003/010", ""),
  ("REQ-F-024", "사양 Excel Import", "정형 Excel 양식 Import → Selection 입력 자동 적용", "필수", "P3", "CPQ-004", ""),
  ("REQ-F-025", "기술 데이터·Document", "선정 상세 옵션(Impeller·Motor 등)·성능 곡선, 입력→계산→출력 Document Template", "필수", "P3", "CPQ-005/006", ""),
  ("REQ-F-026", "완성품 Code·X-Code", "Slot 조합 확정 → 완성품 Code 생성, 비규격은 X-Code 분기(R&D 검토→신규 코드→복귀)", "필수", "P2~P3", "CPQ-007~009", ""),
  ("REQ-F-027", "CPQ Set-up", "Selection/Technical/Document/Print Form 구성, Toolbar·Module 이미지 제작", "필수", "P3", "CPQ-011~015", ""),
 ]),
 ("EDIM Run", [
  ("REQ-F-028", "일괄 자동 생성", "BOM→치수 Macro 실행(계산값 Table)→승인·제작도면→원가→기술자료→서류 비동기 생성", "필수", "P2~P3", "RUN-001~007", "개요 §6"),
  ("REQ-F-029", "Run 모니터링·산출물", "진행률·오류 목록·재시도, Project Folder(DWG/Price/Data/BOM) 저장·조회", "필수", "P2", "RUN-008/009", ""),
  ("REQ-F-030", "부분 Run·Export", "EBOM Run(BOM만)/전체 Run 구분, BOM Excel Export", "필수", "P2", "RUN-010", ""),
 ]),
 ("원가·견적", [
  ("REQ-F-031", "단가 관리", "견적/구매이력/재고단가/견적적용 4종 Table, 적용 우선순위, 재고단가 4값(최고·최저·평균·최근) 자동 산출", "필수", "P3", "CST-001, ERP-021", ""),
  ("REQ-F-032", "원가 계산", "자재비(면적·무게·부피×단가)·제조비(시간·임율·장비)·직접비 집계", "필수", "P3", "CST-002~004, ERP-028", ""),
  ("REQ-F-033", "PCR·견적", "Business Type별 원가 전개(EBIT까지), 견적서 생성·PDF 출력·이력", "필수", "P3", "CST-005~007", ""),
 ]),
 ("문서 관리", [
  ("REQ-F-034", "문서 통제", "Released Status(Set-up→Check→Approve→Accepted)·Version·DOC No. 관리, Management Grade별 열람·출력 통제, 보안 솔루션 연계", "필수", "P3~P5", "DOC-001~004", "보안 솔루션 범위 협의"),
 ]),
 ("ERP", [
  ("REQ-F-035", "프로젝트 관리", "Project 등록(PS)·영업 단계(기술제안~종료)·접수 자료", "필수", "P2", "ERP-001", ""),
  ("REQ-F-036", "프로세스 상태기계", "프로세스 정의(40종 코드) Set-up(선행·후행·기한·자동), 이벤트 상태 전이, 부서별 업무함", "필수", "P5", "ERP-002/003/016", "슬라이드 17 5-1~5-3"),
  ("REQ-F-037", "영업 프로세스", "견적검토(PCR)→견적(QCR)→수주(OR)→승인도서(AP/APP)→출고·납품(FDR/SFO/DF)→기성(IR)", "필수", "P5", "ERP-004/005/009/010", ""),
  ("REQ-F-038", "구매·자재", "구매 조건·공급자 코드 매핑, 발주(PR/PO)·입출고(MI/MO/RM), 재고·창고(위험물성·위치계층), MRP, 일반발주", "필수", "P5", "ERP-007/017~022/026/029", "슬라이드 46"),
  ("REQ-F-039", "생산", "생산계획(MP)·Scheduling·Capacity, 작업지시(WR), 공정 관리(실시간 수집), 반·완성품, 외주·재작업", "필수", "P5", "ERP-006/008/023~025/030, DWG-021", ""),
  ("REQ-F-040", "품질", "자재·반완성품 검사(ER/EF), 부적합(NCF), 검수, 인증서 발급", "필수", "P5", "ERP-012/027", ""),
  ("REQ-F-041", "재무", "세금계산 매입/매출(II/IO), 출금/입금(PA/PI), 정산(RR/RA), Advance(AR), 재무관리(FM)", "필수", "P5", "ERP-011", "AR 기능 v0.3 보강"),
  ("REQ-F-042", "Dashboard·마스터", "전후 공정·일정·이상 경고(시간/자금)·손익·자금 흐름 집계, 회사 마스터, 업체/Project 평가", "필수", "P2~P5", "ERP-013~015", ""),
 ]),
 ("AI", [
  ("REQ-F-043", "AI 학습 파이프라인", "CAD 변환→메타데이터/2D 엔티티/OCR 추출→RAG 인덱싱 (3D Feature는 후순위), Platform 전용", "선택", "P5", "AI-001~003", "개요 §9"),
  ("REQ-F-044", "AI 생성", "Prompt→Macro/Flowchart/Coding 생성·상호 변환, Application 설명→UI 초안 제안", "필수", "P4", "AI-004~006", ""),
  ("REQ-F-045", "사내 자료 챗봇", "도면·기술자료·서류 RAG Q&A, ERP 연동 서류 생성 지원", "선택", "P5", "AI-007", ""),
  ("REQ-F-046", "AI 승인 게이트", "AI 생성물은 검증(Test)·승인 통과 후에만 사용", "필수", "P4", "AI-008", ""),
 ]),
 ("모바일", [
  ("REQ-F-047", "모바일 기본", "QR 스캔 정보 열람(도면·서류·Project·업무), 모바일 승인, Project 소통(History), 공지·푸시", "필수", "P5", "APP-001~003/007", ""),
  ("REQ-F-048", "현장 업무", "자재 입출고, 검수(사진 첨부), 유지보수 접수, AR 뷰(후순위), 오프라인 캐시", "필수", "P5", "APP-004~006/008", ""),
 ]),
 ("관리", [
  ("REQ-F-049", "관리 콘솔", "테넌트·사용자·권한·승인 일괄 관리, AI 학습 관리(Platform), 감사·사용량", "필수", "P1~P5", "ADM-001~005", ""),
 ]),
 ("건축 설비", [
  ("REQ-F-050", "Duct 자동 설계", "건축도 AI 판독(설치불가 구분), 설계 조건(풍량·유속), 최단경로 자동 배치·Diffuser, 기술자료 계산, EDIM 연동", "선택", "P5", "DUCT-001~007", "사업 범위 확정 대상"),
 ]),
]

# ---------------------------------------------------------------- 비기능 요구사항
# (REQ ID, 분류, 요구사항명, 상세, 측정 기준/목표, 비고)
NONFUNC = [
 ("REQ-N-001", "성능", "산출물 생성 시간", "CPQ 선정 후 EDIM Run 전체 산출물(BOM·도면·견적·기술자료) 생성 완료", "1시간 이내 (표준 제품 목표 10분)", "발표자료 핵심 지표"),
 ("REQ-N-002", "성능", "화면 응답", "일반 조회/저장 트랜잭션 응답", "조회 2초·저장 3초 이내 (95%ile)", ""),
 ("REQ-N-003", "성능", "BOM 전개", "완성품 Code 재귀 전개", "5,000 파트 30초 이내", "가정 — 협의"),
 ("REQ-N-004", "성능", "도면 캔버스", "SVG 렌더 팬줌 반응성", "5,000 엔티티 원활, 1만 초과 시 Canvas/WebGL 전환", "개요 설계노트"),
 ("REQ-N-005", "성능", "동시 사용자", "테넌트당 동시 사용자", "100명 (가정 — 협의)", ""),
 ("REQ-N-006", "보안", "전송·저장 암호화", "전 구간 TLS, 비밀번호 해시(bcrypt/argon2), 비밀키 암호화 보관", "필수", ""),
 ("REQ-N-007", "보안", "접근 통제", "RBAC(4단계 × 리소스 단위) + 테넌트 데이터 격리(RLS)", "필수", ""),
 ("REQ-N-008", "보안", "문서 보안", "Management Grade별 열람·출력 통제, 출력물 워터마크", "필수", "보안 솔루션 범위 협의"),
 ("REQ-N-009", "보안", "감사 로그", "전 변경의 작업자·일시·전후 데이터 기록", "보관 5년 (가정)", ""),
 ("REQ-N-010", "보안", "계정 정책", "비밀번호 정책·실패 잠금·세션 타임아웃", "필수", ""),
 ("REQ-N-011", "가용성", "서비스 가용성", "SaaS 운영 기준 가용률", "99.5% (협의)", ""),
 ("REQ-N-012", "가용성", "백업·복구", "DB 일 1회 전체 + PITR, 파일 스토리지 버전닝, 복구 목표 RPO 24h/RTO 4h", "가정 — 협의", ""),
 ("REQ-N-013", "데이터", "승인 게이트", "미승인(DRAFT/PENDING) 자산은 CPQ·Run에서 사용 불가", "필수", "개요 §1.4"),
 ("REQ-N-014", "데이터", "무결성 검증", "Code 관계·Macro 참조·설계 우선순위의 순환 참조 자동 검증", "저장 시 차단", ""),
 ("REQ-N-015", "사용성", "다국어", "한국어 기본, 다국어 확장 구조(i18n)", "영어 1차 (협의)", ""),
 ("REQ-N-016", "사용성", "개인화", "Frame·Canvas 색상, 언어, 판넬 크기(아코디언·펼침), 회사 로고", "제공", "슬라이드 57"),
 ("REQ-N-017", "사용성", "지원 환경", "브라우저 Chrome/Edge 최신 2개 버전, 모바일 Android/iOS", "협의", ""),
 ("REQ-N-018", "운영", "배포 형태", "SaaS 멀티테넌트 + Self-managed Server 패키지 겸용", "필수", "슬라이드 20"),
 ("REQ-N-019", "운영", "모니터링", "메트릭·로그·경고(시간/자금 이상 경고 포함) 대시보드", "필수", ""),
 ("REQ-N-020", "운영", "확장성", "Run·AI·변환 워커 수평 확장(큐 기반), 코어 모듈 경계 유지로 MSA 전환 가능", "필수", "컴포넌트 §1.2"),
 ("REQ-N-021", "제약", "Macro 문법", "Excel 호환 문법은 특허 검토 결과에 따라 조정될 수 있음", "법무 검토 선행", "슬라이드 27 노트"),
 ("REQ-N-022", "제약", "Platform 권한", "AI 학습은 Platform 제공자 전용, System DB 영향 변경은 Platform 승인 필수", "필수", "슬라이드 18"),
]

# ---------------------------------------------------------------- 인터페이스 요구사항
# (REQ ID, 대상, 방향, 내용, 방식, 비고)
INTERFACE = [
 ("REQ-I-001", "외부 ERP", "양방향", "BOM·발주·회계 데이터 교환", "REST/파일 어댑터 (플러그블)", "대상 ERP 미확정"),
 ("REQ-I-002", "Digital Twin (DTDesigner)", "송신", "치수 Table·3D 모델·실시간 Code 연결, AR/XR 데이터", "API (협의)", "슬라이드 77"),
 ("REQ-I-003", "Excel", "양방향", "고객 사양 Import, Table 행 Import, BOM·견적 Export", "정형 양식 파일", ""),
 ("REQ-I-004", "CAD", "양방향", "DWG↔DXF 변환(ODA), PDF→DXF, 2D→3D(STEP)", "변환 워커 (바이너리 격리)", ""),
 ("REQ-I-005", "LLM API", "송신", "Macro/UI 생성·챗봇·학습 파이프라인용 LLM 호출", "Claude API (온프레미스 옵션 협의)", ""),
 ("REQ-I-006", "QR", "양방향", "도면·서류·Project·자산 QR 발급/해석", "QR 코드", "모바일 진입점"),
 ("REQ-I-007", "모바일 푸시", "송신", "승인·공지·경고 푸시 알림", "FCM/APNs", ""),
 ("REQ-I-008", "Object Storage", "양방향", "CAD·PDF·이미지 파일 저장, Project Folder 규약", "S3 API (MinIO — 개발 구축완료)", ""),
]

# ---------------------------------------------------------------- 용어
GLOSSARY = [
 ("EDIM", "Enterprise Digital Integration Management — CPQ+PLM+ERP+Digital Twin 통합 플랫폼"),
 ("CTO / ETO", "Configure-to-Order / Engineer-to-Order — 주문 구성/설계 생산 방식"),
 ("RCCS™", "Relational Code Configuration System — 관계형 코드 체계 (Sub/Product/Arrangement Code)"),
 ("CPQ", "Configure-Price-Quote — 제품 선정·견적"),
 ("TLM / PLM", "코드·제품 Set-up 영역 명칭 (발표자료 혼용) / Product Lifecycle Management"),
 ("Arrangement", "표준 구성 조합 (모듈 배치·결합 조건)"),
 ("완성품 Code", "CPQ Slot 선택 조합으로 생성되는 Finished Goods Code"),
 ("X-Code", "비규격 Option 표시 — R&D 검토·신규 코드 등록 분기"),
 ("Macro", "Excel 호환 문법 계산식 — 치수·검증·원가·기술 계산에 바인딩"),
 ("4-Way Sync", "Prompt↔Macro↔Flowchart↔Coding 상호 변환"),
 ("EDIM Run", "BOM·도면·원가·기술자료·서류 일괄 자동 생성 실행"),
 ("Hierarchy Address", "Tree 노드 기반 데이터 주소 체계 — 이동·개명에도 연결 유지"),
 ("PCR", "Pre-Calculation Report — Business Type별 원가 전개 (문맥에 따라 견적검토 프로세스 코드)"),
 ("DrawingDocument", "도면 기하 표준 JSON (line/polyline/circle/arc/text + layers)"),
 ("Management Grade", "문서 보안 등급 (S-1~S-n) — 열람·출력 통제"),
 ("Running Test", "Mother-Child 관계의 Part List 전개 검증"),
]


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def fill_rows(ws, start_col_center=(1,)):
    pass


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()

    # ---- 문서정보 ----
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 요구사항정의서"
    ws["B2"].font = TITLE_FONT
    total_f = sum(len(rows) for _, rows in FUNC)
    info = [
        ("문서 버전", "v0.1"),
        ("작성일", "2026-07-07"),
        ("근거", "EDIM Tool System EP2.pptx (78슬라이드) 분석 → 기능정의서 v0.2(178기능)·메뉴정의서 v0.3(97메뉴)·컴포넌트정의서에서 도출"),
        ("요구사항 수", f"기능 {total_f} · 비기능 {len(NONFUNC)} · 인터페이스 {len(INTERFACE)}"),
        ("ID 체계", "REQ-F-xxx(기능) / REQ-N-xxx(비기능) / REQ-I-xxx(인터페이스)"),
        ("중요도", "필수 = v1 범위 / 선택 = 협의·후순위"),
        ("Phase", "P1 RCCS코어 / P2 설계·Run / P3 원가·문서 / P4 Toolbox·AI / P5 ERP·모바일"),
        ("추적", "관련 기능 ID → EDIM_기능정의서.xlsx, 화면 → EDIM_화면설계서.html, DB → EDIM_DB정의서.xlsx"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 100

    # ---- 요구사항총괄 ----
    ws = wb.create_sheet("요구사항총괄")
    ws.append(["구분", "그룹", "건수", "필수", "선택"])
    style_header(ws, 5)
    for gi, (grp, rows) in enumerate(FUNC):
        must = sum(1 for x in rows if x[4] == "필수")
        ws.append(["기능", grp, len(rows), must, len(rows) - must])
        for c in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = CENTER
        ws.cell(row=ws.max_row, column=2).fill = PatternFill("solid", fgColor=GRP_FILLS[gi % len(GRP_FILLS)])
    for name, data in (("비기능", NONFUNC), ("인터페이스", INTERFACE)):
        ws.append([name, "-", len(data), len(data), 0])
        for c in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = CENTER
    ws.append(["합계", "", total_f + len(NONFUNC) + len(INTERFACE), "", ""])
    for c in range(1, 6):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = BOLD_FONT; cell.border = BORDER; cell.alignment = CENTER
    for i, w in enumerate([10, 18, 8, 8, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 기능 요구사항 ----
    ws = wb.create_sheet("기능요구사항")
    headers = ["No", "요구사항 ID", "구분", "요구사항명", "상세 설명", "중요도", "Phase", "관련 기능 ID", "비고"]
    ws.append(headers)
    style_header(ws, len(headers))
    n = 0
    for gi, (grp, rows) in enumerate(FUNC):
        fill = PatternFill("solid", fgColor=GRP_FILLS[gi % len(GRP_FILLS)])
        for rid, name, desc, imp, phase, fids, note in rows:
            n += 1
            ws.append([n, rid, grp, name, desc, imp, phase, fids, note])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=c)
                cell.font = BODY_FONT; cell.border = BORDER
                cell.alignment = CENTER if c in (1, 2, 6, 7) else WRAP
            ws.cell(row=ws.max_row, column=3).fill = fill
    for i, w in enumerate([5, 12, 12, 20, 62, 8, 9, 22, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---- 비기능 요구사항 ----
    ws = wb.create_sheet("비기능요구사항")
    headers = ["No", "요구사항 ID", "분류", "요구사항명", "상세 설명", "측정 기준·목표", "비고"]
    ws.append(headers)
    style_header(ws, len(headers))
    for n, row in enumerate(NONFUNC, 1):
        ws.append([n] + list(row))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 3) else WRAP
    for i, w in enumerate([5, 12, 9, 16, 54, 30, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    # ---- 인터페이스 요구사항 ----
    ws = wb.create_sheet("인터페이스요구사항")
    headers = ["No", "요구사항 ID", "대상 시스템", "방향", "내용", "방식", "비고"]
    ws.append(headers)
    style_header(ws, len(headers))
    for n, row in enumerate(INTERFACE, 1):
        ws.append([n] + list(row))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 4) else WRAP
    for i, w in enumerate([5, 12, 22, 8, 46, 30, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- 용어정의 ----
    ws = wb.create_sheet("용어정의")
    ws.append(["No", "용어", "정의"])
    style_header(ws, 3)
    for n, (term, desc) in enumerate(GLOSSARY, 1):
        ws.append([n, term, desc])
        for c in range(1, 4):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c == 1 else WRAP
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 96
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (기능 {total_f} · 비기능 {len(NONFUNC)} · 인터페이스 {len(INTERFACE)} · 용어 {len(GLOSSARY)})")


if __name__ == "__main__":
    main()
