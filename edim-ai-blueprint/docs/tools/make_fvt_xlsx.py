# -*- coding: utf-8 -*-
"""EDIM 기능확인서(FVT).xlsx 생성 — 기능정의서 178건 + 비기능(REQ-N) 자동 생성.

실행: py docs/tools/make_fvt_xlsx.py  (저장: docs/03_기능확인서_FVT/EDIM_기능확인서.xlsx)
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..")
F_FEAT = os.path.join(BASE, "EDIM_기능정의서.xlsx")
F_REQ = os.path.join(BASE, "02_요구사항", "EDIM_요구사항정의서.xlsx")
OUT = os.path.join(BASE, "03_기능확인서_FVT", "EDIM_기능확인서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
PALETTE = ["FFF8E7", "E8F4FD", "F0FFF0", "FFF0F5", "F5F0FF", "FFEFC2",
           "E0F7FA", "FDF2E9", "EAF2F8", "F9EBEA", "F4ECF7", "E8F8F5", "FEF9E7", "EBEDEF"]


def style_header(ws, cols, row=1):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def body_cell(cell, center=False):
    cell.font = BODY_FONT
    cell.border = BORDER
    cell.alignment = CENTER if center else WRAP


def _read_frozen(sheet, id_col, verdict_cols, verdict_at):
    """기존 FVT xlsx 에서 판정 기입 행을 보존 (C13 freeze — 재생성 덮어쓰기 방지).

    반환 {기능/REQ ID: {열idx: 값}} — 판정(verdict_at) 이 채워진 행만.
    """
    if not os.path.exists(OUT):
        return {}
    try:
        wb = openpyxl.load_workbook(OUT, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return {}
        frozen = {}
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if len(row) <= verdict_at or not row[id_col]:
                continue
            verdict = row[verdict_at]
            if verdict not in (None, ""):
                frozen[row[id_col]] = {c: row[c] for c in verdict_cols
                                       if len(row) > c and row[c] not in (None, "")}
        wb.close()
        return frozen
    except Exception:  # noqa: BLE001
        return {}


def main():
    # C13 — 기존 판정 보존 (기능: id=col1, 판정~비고=col6~10 / 비기능: id=col1, 판정~비고=col5~9)
    frozen_feat = _read_frozen("기능확인항목", 1, [6, 7, 8, 9, 10], 6)
    frozen_nfr = _read_frozen("비기능확인항목", 1, [5, 6, 7, 8, 9], 5)
    if frozen_feat or frozen_nfr:
        print(f"freeze — 보존 판정: 기능 {len(frozen_feat)}·비기능 {len(frozen_nfr)}")
    # 기능정의서 로드
    feats = []
    wb = openpyxl.load_workbook(F_FEAT, read_only=True)
    for row in wb["기능목록"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            feats.append({"id": row[1], "module": row[2], "name": row[4],
                          "desc": row[5], "phase": row[10]})
    wb.close()
    # 비기능 로드
    nfrs = []
    wb = openpyxl.load_workbook(F_REQ, read_only=True)
    for row in wb["비기능요구사항"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            nfrs.append({"id": row[1], "cat": row[2], "name": row[3], "crit": row[5]})
    wb.close()

    out = Workbook()

    # ---- 문서정보 ----
    ws = out.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 기능확인서 (FVT)"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1 (자동 생성 — 기능정의서 v0.2 기준)"),
        ("작성일", "2026-07-07"),
        ("목적", "구축된 EDIM 시스템이 기능정의서·요구사항정의서에 맞게 구현되었는지 항목별 확인 (검수 근거)"),
        ("확인 항목", f"기능 {len(feats)}건 · 비기능 {len(nfrs)}건"),
        ("판정 기준", "P(Pass) / F(Fail — 결함번호 기재) / C(조건부 — 비고 기재) / N/A(범위 제외)"),
        ("합격 기준", "필수 기능 P 100%, F 0건 (C는 조치 계획 합의 시 허용) — 검수 협의로 확정"),
        ("수행 시점", "WBS 5.4(W37~38) 준비 → 8.1(W43) 검수 ◆M4"),
        ("재생성", "기능정의서 변경 시 make_fvt_xlsx.py 재실행 (판정 기입 전까지만 — 기입 후에는 버전 고정)"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 100

    # ---- 시스템환경 ----
    ws = out.create_sheet("시스템환경")
    ws.append(["구분", "항목", "내용"])
    style_header(ws, 3)
    envs = [
        ("검수 환경", "URL", "(운영/검수 환경 확정 후 기입 — 개발: https://edim.seekerslab.com)"),
        ("검수 환경", "서버", "(기입) — 개발 서버: Ubuntu 24.04, Docker, PostgreSQL 16, MinIO, nginx/HTTPS"),
        ("검수 환경", "DB", "PostgreSQL 16 — 스키마 53테이블 (ddl/edim_schema.sql)"),
        ("클라이언트", "브라우저", "Chrome/Edge 최신 (REQ-N-017)"),
        ("클라이언트", "모바일", "Android/iOS (P5 범위)"),
        ("계정", "확인용 계정", "PLATFORM/ADMIN/SETUP/GENERAL 각 1 (권한승인정의서 매트릭스 기준)"),
        ("데이터", "확인 데이터", "이행 데이터 또는 표준 샘플 (KDCR 3-13 코드 체계 — verify_runtime.sql 준용)"),
    ]
    for row in envs:
        ws.append(list(row))
        for c in range(1, 4):
            body_cell(ws.cell(row=ws.max_row, column=c), center=(c == 1))
    for i, w in enumerate([12, 14, 90], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 기능 확인 항목 ----
    ws = out.create_sheet("기능확인항목")
    headers = ["No", "기능 ID", "모듈", "기능명", "확인 내용", "Phase",
               "판정", "결함번호", "확인자", "확인일", "비고"]
    ws.append(headers)
    style_header(ws, len(headers))
    module_fill = {}
    mi = 0
    for n, f in enumerate(feats, 1):
        if f["module"] not in module_fill:
            module_fill[f["module"]] = PatternFill("solid", fgColor=PALETTE[mi % len(PALETTE)])
            mi += 1
        fr = frozen_feat.get(f["id"], {})
        ws.append([n, f["id"], f["module"], f["name"], f["desc"], f["phase"],
                   fr.get(6, ""), fr.get(7, ""), fr.get(8, ""), fr.get(9, ""), fr.get(10, "")])
        for c in range(1, len(headers) + 1):
            body_cell(ws.cell(row=ws.max_row, column=c), center=c in (1, 2, 6, 7, 8, 10))
        ws.cell(row=ws.max_row, column=3).fill = module_fill[f["module"]]
    for i, w in enumerate([5, 10, 13, 22, 56, 7, 6, 9, 8, 9, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---- 비기능 확인 항목 ----
    ws = out.create_sheet("비기능확인항목")
    headers = ["No", "REQ ID", "분류", "요구사항명", "측정 기준·목표", "판정", "측정값", "확인자", "확인일", "비고"]
    ws.append(headers)
    style_header(ws, len(headers))
    for n, f in enumerate(nfrs, 1):
        fr = frozen_nfr.get(f["id"], {})
        ws.append([n, f["id"], f["cat"], f["name"], f["crit"],
                   fr.get(5, ""), fr.get(6, ""), fr.get(7, ""), fr.get(8, ""), fr.get(9, "")])
        for c in range(1, len(headers) + 1):
            body_cell(ws.cell(row=ws.max_row, column=c), center=c in (1, 2, 3, 6, 9))
    for i, w in enumerate([5, 11, 9, 18, 36, 6, 14, 8, 9, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---- 결함 목록 ----
    ws = out.create_sheet("결함목록")
    ws.append(["결함번호", "관련 항목(기능 ID)", "결함 내용", "심각도", "발견일", "상태", "조치 내용", "조치일", "재확인"])
    style_header(ws, 9)
    ws.append(["D-001", "", "", "치명/중/경", "", "OPEN/FIXED/CLOSED", "", "", "P/F"])
    for c in range(1, 10):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(name="맑은 고딕", size=9, color="999999")
        cell.border = BORDER
        cell.alignment = WRAP
    for i, w in enumerate([9, 15, 42, 10, 9, 16, 34, 9, 7], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 최종승인 ----
    ws = out.create_sheet("최종승인")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "최종 확인 및 승인"
    ws["B2"].font = TITLE_FONT
    ws["B4"] = "확인 결과 요약"
    ws["B4"].font = BOLD_FONT
    ws.append([])
    rows = [
        ("", "기능 확인", f"총 {len(feats)}건 — P (  ) · F (  ) · C (  ) · N/A (  )"),
        ("", "비기능 확인", f"총 {len(nfrs)}건 — P (  ) · F (  ) · C (  ) · N/A (  )"),
        ("", "미결 결함", "치명 (  ) · 중 (  ) · 경 (  )"),
    ]
    r = 5
    for _, k, v in rows:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    r += 2
    ws.cell(row=r, column=2, value="공급자 (NOVA Solution)").font = BOLD_FONT
    ws.cell(row=r + 1, column=2, value="확인자:                    (서명)      일자:")
    ws.cell(row=r + 3, column=2, value="고객사").font = BOLD_FONT
    ws.cell(row=r + 4, column=2, value="검수자:                    (서명)      일자:")
    ws.cell(row=r + 6, column=2, value="※ 본 확인서 승인으로 검수(M4)가 완료되며, 잔여 C 항목은 합의된 조치 계획에 따른다.").font = Font(name="맑은 고딕", size=9, color="777777")
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 60

    out.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (기능 {len(feats)}·비기능 {len(nfrs)})")


if __name__ == "__main__":
    main()
