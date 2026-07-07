# -*- coding: utf-8 -*-
"""EDIM WBS.xlsx 생성 — SI 8단계 × 개발 Phase(P1~P5) 내용화, 주차 간트 포함.

시작일은 가정(2026-08-03) — 계약 확정 시 START만 바꿔 재생성.
실행: py docs/tools/make_wbs_xlsx.py  (저장: docs/04_WBS/EDIM_WBS.xlsx)
"""
import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "04_WBS", "EDIM_WBS.xlsx")
START = date(2026, 8, 3)   # 가정 — 월요일
TOTAL_WEEKS = 44

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
SMALL_FONT = Font(name="맑은 고딕", size=8)
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

PHASE_COLOR = {
    "착수": "8E7CC3", "분석": "6FA8DC", "설계": "76A5AF", "구현": "93C47D",
    "테스트": "F6B26B", "이행": "E06666", "안정화": "C27BA0", "종료": "999999",
}
PHASE_BG = {
    "착수": "F5F0FF", "분석": "E8F4FD", "설계": "E0F7FA", "구현": "F0FFF0",
    "테스트": "FFF8E7", "이행": "FFF0F5", "안정화": "FDF2E9", "종료": "F0F1F3",
}

# (WBS, 단계, Task, 시작주, 기간(주), 담당 역할, 산출물, Phase/비고)
TASKS = [
 ("1",   "착수", "착수 단계", 1, 2, "", "", ""),
 ("1.1", "착수", "킥오프·사업수행계획 수립", 1, 1, "PM", "사업수행계획서", ""),
 ("1.2", "착수", "개발 환경·CI 파이프라인 구성", 1, 2, "인프라", "Jenkins 파이프라인", "서버·Docker·PG·MinIO 선구축 완료"),
 ("1.3", "착수", "개발 표준 확정 (v1.0 승격)", 2, 1, "아키텍트", "개발표준정의서 v1.0", "백엔드 언어 확정 포함"),
 ("2",   "분석", "분석 단계", 3, 4, "", "", ""),
 ("2.1", "분석", "AS-IS 업무·시스템·자료 조사", 3, 2, "컨설턴트", "현행분석서", "고객사 인터뷰"),
 ("2.2", "분석", "요구사항 확정 — 협의 4건 결정", 3, 3, "PM·전 파트", "요구사항정의서 v1.0", "보안솔루션·DUCT·ERP경계·CAD명령표"),
 ("2.3", "분석", "데이터 이행 대상 조사 (도면·코드·단가)", 5, 2, "DBA·AI", "데이터 이행 계획서 초안", "AI 학습(RAW DB) 연계"),
 ("2.4", "분석", "요구사항 검토회 ◆M1", 6, 1, "전체", "검토회의록", "M1: 요구 확정"),
 ("3",   "설계", "설계 단계", 7, 6, "", "", ""),
 ("3.1", "설계", "화면 상세 설계 (와이어프레임 16 → 전 화면)", 7, 4, "UI/UX", "화면설계서 v1.0", ""),
 ("3.2", "설계", "DB 물리 설계·마이그레이션 체계", 7, 2, "DBA", "DB정의서 v1.0 + 마이그레이션", "DDL v0.4.1 기반"),
 ("3.3", "설계", "인터페이스 정의서 (OpenAPI 106+)", 8, 3, "백엔드", "인터페이스정의서", ""),
 ("3.4", "설계", "클래스·모듈 설계", 9, 3, "아키텍트", "클래스정의서", ""),
 ("3.5", "설계", "설계 검토회 ◆M2", 12, 1, "전체", "검토회의록", "M2: 설계 완료"),
 ("4",   "구현", "구현 단계", 9, 28, "", "", "Phase 중첩 진행"),
 ("4.1", "구현", "P1 RCCS 코어 — 코드·Hierarchy·Table·Macro엔진·승인", 9, 6, "백엔드 2·FE 1", "P1 모듈+단위테스트", "BOM 전개·Running Test"),
 ("4.2", "구현", "P2 설계·Run — Drawing Editor·EDIM Run·CPQ", 13, 8, "백엔드 2·FE 2", "P2 모듈", "◆M3(W20): 코드→Run→견적 E2E 데모"),
 ("4.3", "구현", "P3 원가·문서 — 단가·PCR·견적·Print·문서관리", 19, 6, "백엔드 2·FE 1", "P3 모듈", ""),
 ("4.4", "구현", "P4 Toolbox·AI — UI Designer·4-Way Sync·AI 생성", 23, 8, "백엔드 1·FE 2·AI 1", "P4 모듈", "AI 승인 게이트 포함"),
 ("4.5", "구현", "P5 ERP·모바일 — 프로세스·부서업무·Dashboard·App", 27, 10, "백엔드 2·FE 1·모바일 1", "P5 모듈", ""),
 ("4.6", "구현", "단위테스트 (병행, 커버리지 80%)", 9, 28, "전 개발", "단위테스트 결과", "개발표준 §7"),
 ("5",   "테스트", "테스트 단계", 31, 8, "", "", ""),
 ("5.1", "테스트", "통합 테스트 — E2E 시나리오", 31, 6, "QA·전 파트", "통합테스트 결과", "개요 §7 워크플로우 기반"),
 ("5.2", "테스트", "성능 시험 — REQ-N-001~005", 35, 3, "QA·인프라", "성능시험 결과", "산출물 1시간·BOM 30초·응답 2/3초"),
 ("5.3", "테스트", "결함 조치·회귀", 33, 6, "전 개발", "결함관리대장", ""),
 ("5.4", "테스트", "FVT 확인 항목 준비 (기능 178건)", 37, 2, "QA·PM", "기능확인서 초안", ""),
 ("6",   "이행", "이행 단계", 35, 6, "", "", ""),
 ("6.1", "이행", "데이터 이행 — 코드·도면·단가·BOM", 35, 4, "DBA·컨설턴트", "이행 결과서", "★검증 필수"),
 ("6.2", "이행", "AI 학습 — 사내 도면·서류 (Platform)", 37, 4, "AI", "학습 DB·품질 리포트", "슬라이드 25 파이프라인"),
 ("6.3", "이행", "운영 환경 구축·배포", 38, 3, "인프라", "설치·배포 매뉴얼", "Self-managed 대응"),
 ("7",   "안정화", "안정화 단계", 39, 4, "", "", ""),
 ("7.1", "안정화", "파일럿 운영·모니터링·튜닝", 39, 4, "전체", "운영 리포트", ""),
 ("7.2", "안정화", "사용자 교육 — Set-up(코드·Macro) 중심", 40, 2, "PM·컨설턴트", "교육 자료·결과", ""),
 ("7.3", "안정화", "사용자·운영자 매뉴얼", 39, 3, "테크니컬라이터", "매뉴얼 2종", ""),
 ("8",   "종료", "종료 단계", 43, 2, "", "", ""),
 ("8.1", "종료", "FVT·검수 ◆M4", 43, 1, "고객·PM", "기능확인서 승인", "M4: 검수"),
 ("8.2", "종료", "인수인계·최종 보고", 44, 1, "PM", "완료보고서", ""),
]

MILESTONES = [
 ("M1", "요구사항 확정", 6, "요구사항정의서 v1.0 승인 — 협의 4건 결정 완료"),
 ("M2", "설계 완료", 12, "화면·DB·인터페이스·클래스 정의서 v1.0"),
 ("M3", "핵심 E2E 데모", 20, "코드 등록→CPQ→EDIM Run→도면·견적 자동 생성 시연 (P2)"),
 ("M4", "검수", 43, "FVT 승인·인수"),
]


def week_date(w):
    return START + timedelta(weeks=w - 1)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM WBS·일정표"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1"),
        ("작성일", "2026-07-07"),
        ("기간", f"총 {TOTAL_WEEKS}주 (약 10개월) — W1 = {START.isoformat()} 시작 **가정**, 계약 확정 시 재산정"),
        ("구조", "SI 8단계(착수~종료) × 개발 Phase(P1~P5, 중첩 진행)"),
        ("근거", "기능정의서 Phase 배분(P1:30 P2:52 P3:27 P4:18 P5:51), 컴포넌트정의서 §10"),
        ("마일스톤", "M1 요구확정(W6) · M2 설계완료(W12) · M3 핵심 E2E 데모(W20) · M4 검수(W43)"),
        ("투입(안)", "PM 1 · 아키텍트 1 · 백엔드 2~3 · FE 2 · 모바일 1 · AI 1 · DBA 1 · QA 1 · 컨설턴트 1 — 협의 대상"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 100

    # ---- WBS + 간트 ----
    ws = wb.create_sheet("WBS")
    fixed = ["WBS", "단계", "Task", "시작", "종료", "주", "담당", "산출물", "비고"]
    headers = fixed + [f"W{w}" for w in range(1, TOTAL_WEEKS + 1)]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT if c <= len(fixed) else Font(name="맑은 고딕", size=7, bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BORDER
    for t in TASKS:
        wbs, phase, name, sw, dur, owner, deliv, note = t
        is_group = "." not in wbs
        s, e = week_date(sw), week_date(sw + dur - 1) + timedelta(days=4)
        ws.append([wbs, phase, name, s.strftime("%m-%d"), e.strftime("%m-%d"), dur, owner, deliv, note] + [""] * TOTAL_WEEKS)
        row = ws.max_row
        for c in range(1, len(fixed) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BOLD_FONT if is_group else BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 4, 5, 6) else WRAP
            if is_group:
                cell.fill = PatternFill("solid", fgColor=PHASE_BG[phase])
        bar = PatternFill("solid", fgColor=PHASE_COLOR[phase])
        for w in range(sw, sw + dur):
            cell = ws.cell(row=row, column=len(fixed) + w)
            cell.fill = bar
            cell.border = BORDER
        for w in range(1, TOTAL_WEEKS + 1):
            ws.cell(row=row, column=len(fixed) + w).border = BORDER
    widths = [6, 7, 42, 7, 7, 4, 15, 22, 26] + [2.4] * TOTAL_WEEKS
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "J2"

    # ---- 마일스톤 ----
    ws = wb.create_sheet("마일스톤")
    ws.append(["ID", "마일스톤", "주차", "예정일(가정)", "기준"])
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for mid, name, w, crit in MILESTONES:
        ws.append([mid, name, f"W{w}", week_date(w).isoformat(), crit])
        for c in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 3, 4) else WRAP
    for i, w in enumerate([6, 20, 6, 13, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (Task {len(TASKS)}·마일스톤 {len(MILESTONES)}·{TOTAL_WEEKS}주 간트)")


if __name__ == "__main__":
    main()
