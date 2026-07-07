# -*- coding: utf-8 -*-
"""EDIM 권한승인정의서.xlsx 생성.

권한 매트릭스는 EDIM_메뉴정의서.xlsx(97메뉴)의 '사용 권한' 열에서 자동 도출 — 메뉴 변경 시 재생성.
근거: 개요 §2(3단계 권한+Platform)·§4(권한 필터), DB정의서 sys_role_permission,
      슬라이드 18/46/57(※ System DB 영향 = Platform 승인), 58(Management Grade)
실행: py docs/tools/make_authz_xlsx.py  (저장: docs/EDIM_권한승인정의서.xlsx)
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..")
F_MENU = os.path.join(BASE, "EDIM_메뉴정의서.xlsx")
OUT = os.path.join(BASE, "EDIM_권한승인정의서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
FILL_FULL = PatternFill("solid", fgColor="C8E6C9")   # ●
FILL_EDIT = PatternFill("solid", fgColor="FFF3C4")   # C
FILL_VIEW = PatternFill("solid", fgColor="E3ECF7")   # R
FILL_NO = PatternFill("solid", fgColor="F0F1F3")     # —

# ---------------------------------------------------------------- 역할 정의
ROLES = [
 ("PLATFORM", "플랫폼 운영자 (NOVA)", "EDIM 제공자 — System DB·AI 학습·is_system 자산 관리, 테넌트 운영",
  "AI 학습은 Platform 전용(슬라이드 18) · System DB 영향 작업 최종 승인(슬라이드 46/57)"),
 ("ADMIN", "기업 관리자", "테넌트 내 사용자·권한·Head 편집(ERP 하위만)·승인 결정·마스터 관리",
  "Head Item 편집은 ERP 하부 Item만 가능(슬라이드 57)"),
 ("SETUP", "Set-up 사용자", "코드·도면·Macro·Table·UI Form 등 Set-up 자산 작성 (승인 요청까지)",
  "작성 자산은 DRAFT — 승인 전 CPQ/Run 사용 불가"),
 ("GENERAL", "일반 사용자", "CPQ 선정·ERP 부서 업무·조회 — 부서(영업/기술/자재/생산/QC/CS/재무/HR) 소속에 따라 처리 범위 제한",
  "권한 없는 메뉴·Tree 항목은 미표시(개요 §4)"),
]

ACTIONS = [
 ("VIEW", "조회 — 메뉴 진입·데이터 열람"),
 ("EDIT", "편집 — 생성·수정·삭제 (승인 대상 자산은 DRAFT 저장)"),
 ("APPROVE", "승인 — PENDING 자산의 승인/반려 결정"),
 ("SETUP", "설정 — 시스템 구조(Head·Templet·프로세스 정의) 변경"),
]

# 메뉴 '사용 권한' 문자열 → (PLATFORM, ADMIN, SETUP, GENERAL) 셀 코드
# 코드: ● 전체 / C 편집 / A 승인 / R 조회 / — 불가
def derive(perm):
    p = (perm or "").strip()
    if p == "PLATFORM":
        return ("●", "R", "—", "—", "")
    if p in ("ADMIN", "ADMIN↑"):
        return ("●", "●", "—", "—", "")
    if p in ("ADMIN·경영",):
        return ("●", "●", "—", "R", "경영진 조회 포함")
    if p in ("ADMIN·재무",):
        return ("●", "●", "—", "C", "재무 부서")
    if p in ("SETUP", "SETUP↑"):
        return ("●", "A", "C", "R", "")
    if p in ("SETUP·GENERAL",):
        return ("●", "A", "C", "C", "")
    if p == "GENERAL":
        return ("●", "A", "C", "C", "")
    if p == "전체":
        return ("●", "●", "●", "●", "")
    if p == "시스템":
        return ("—", "—", "—", "—", "시스템 내부 실행 (메뉴 아님)")
    if p == "현장":
        return ("●", "A", "—", "C", "현장 모바일 사용자")
    # 부서 제한 (영업/기술/자재/생산/QC/CS/재무/HR 및 복합)
    return ("●", "A", "C", "C", f"부서 제한: {p}")


# ---------------------------------------------------------------- 승인 상태기계
# (자산, 상태 흐름, 요청(작성) 권한, 검토·승인 권한, 선행 조건·특이 규칙, 관련)
FLOWS = [
 ("코드 그룹·자릿수·값", "DRAFT → PENDING → APPROVED / REJECTED", "SETUP", "ADMIN",
  "중복검토 통과 후 승인 요청 가능. 미승인 값은 CPQ 콤보에 미표시", "CODE-003/013, W-04"),
 ("Product Code", "DRAFT → PENDING → APPROVED / REJECTED", "SETUP", "ADMIN",
  "Main Code 중복검토 필수", "CODE-006"),
 ("Code Relationship", "DRAFT → PENDING → APPROVED / REJECTED", "SETUP", "ADMIN",
  "**Part List Running Test 통과 후에만** 승인 요청 가능, 순환(DAG) 검증", "CODE-009, W-05"),
 ("Arrangement Code·구성", "DRAFT → PENDING → APPROVED / REJECTED", "SETUP", "ADMIN",
  "결합 조건 Macro는 별도 승인된 것만 참조", "CODE-011/012"),
 ("도면 (Drawing)", "작성중 → 검토 → 승인 → 발행", "SETUP", "검토: 선임 SETUP / 승인: ADMIN",
  "Breadcrumb 상태 표시(Design>Check>Approve>Accepted). 개정(Rev)마다 재승인", "DWG-014, 슬라이드 26-7"),
 ("도면 Block", "임시 저장 → 승인 → 사용", "SETUP", "ADMIN",
  "저장 장소 확인 절차, 상위 호출 시 Block 상태 (개요 §7.1)", "DWG-012"),
 ("Macro (4-Way)", "DRAFT → TESTED → PENDING → APPROVED", "SETUP", "ADMIN",
  "**Test Run 통과(TESTED) 필수**. AI 생성물도 동일 게이트. 버전별 승인", "TBX-011/012, AI-008, W-12"),
 ("UI Form·Templet", "DRAFT → PENDING → APPROVED(게시)", "SETUP", "ADMIN",
  "게시된 버전만 동적 렌더러가 실행", "TBX-003"),
 ("데이터 Table", "DRAFT → PENDING → APPROVED", "SETUP", "ADMIN",
  "Macro 참조 대상은 승인본만. 대량 Import는 형식 검증 선행", "TBL-006"),
 ("Hierarchy Tree 편집", "편집 → 점검 → 승인 → 반영", "SETUP(사용자 Tree)", "ADMIN",
  "DB 상호관계 이상 유무 점검 후 저장. is_system Tree는 Platform만", "SYS-007, 슬라이드 57"),
 ("문서 (Document)", "Set-up → Check → Approve → Accepted", "GENERAL/SETUP", "ADMIN(Grade 권한자)",
  "Released Status·Version·DOC No. 기록. Grade별 열람·출력 통제", "DOC-001/002, W-11"),
 ("ERP 프로세스 이벤트", "TODO → IN_PROGRESS → DONE (+ALERT)", "담당 부서 GENERAL", "프로세스 정의의 승인 단계 담당",
  "선행 프로세스 DONE 조건 검사, 기한 초과 시 이상 경고", "ERP-002, W-14"),
 ("X-Code (비규격)", "X-Code 표시 → 검토 요청 → R&D 설계 → 신규 코드 등록 → 승인 → 복귀", "GENERAL(영업)", "R&D·ADMIN",
  "신규 코드는 표준 코드 승인 절차로 합류", "CPQ-009, 개요 §6"),
]

# ---------------------------------------------------------------- Platform 승인 범위
PLATFORM_SCOPE = [
 ("Head 구조 변경", "Head Tab 추가·삭제·구조 변경 (기업 ADMIN은 ERP 하위 Item만 편집 가능)", "슬라이드 57"),
 ("is_system Hierarchy Tree", "EDIM 제공 Tree의 구조·항목 변경", "개요 §4"),
 ("System Templet", "EDIM 제공 화면·동작 Templet(E-1~E-4) 변경", "슬라이드 17/64"),
 ("System DB 영향 작업", "공통 스키마·System DB에 영향을 주는 모든 Set-up (Form 생성·DB 구성·Macro 작성 중 해당분)", "슬라이드 18/46 ※ 표기"),
 ("AI 학습", "사내 자료 학습 DB 구축·파이프라인 실행 — Platform 제공자만 작업 가능", "슬라이드 18, 개요 §9"),
 ("프로세스 코드 신설", "ERP 프로세스 정의(코드 40종) 추가·삭제 — 테넌트 커스터마이징(선행·기한 등)은 ADMIN", "슬라이드 17 5-1"),
 ("테넌트·요금", "테넌트 생성·정지, 사용량·과금", "ADM-001"),
 ("Special Customized Program", "사용자가 만들 수 없는 프로그램의 개발 지원(EDIM Developer)", "슬라이드 13/19"),
]

# ---------------------------------------------------------------- 문서 보안 등급 (안)
GRADES = [
 ("S-1", "대외비 (최고)", "임원·지정 승인자", "차단", "차단", "워터마크+열람 로그, 승인자만 Grade 변경"),
 ("S-2", "사내 한정", "부서장·담당 부서", "워터마크 출력", "차단", "슬라이드 58 예시 등급"),
 ("S-3", "일반 업무", "전 사용자(테넌트 내)", "허용", "허용(로그)", "기본값"),
]


def style_header(ws, cols, row=1):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def body(ws, center_cols=()):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if cell.column in center_cols else WRAP


def main():
    # 메뉴 로드
    menus = []
    wb = openpyxl.load_workbook(F_MENU, read_only=True)
    for row in wb["메뉴목록"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            menus.append({"id": row[1], "lv1": row[2], "lv2": row[3], "lv3": row[4], "perm": row[8]})
    wb.close()

    out = Workbook()

    # ---- 문서정보 ----
    ws = out.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 권한·승인 정의서"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1"),
        ("작성일", "2026-07-07"),
        ("근거", "개요 §2/§4 · DB정의서 sys_role_permission · 슬라이드 18/46/57(Platform 승인)·58(Grade) · 메뉴정의서 v0.3"),
        ("구성", "역할정의 / 권한매트릭스(메뉴 97 자동 도출) / 승인상태기계(자산 13종) / Platform승인범위 / 문서보안등급"),
        ("셀 코드", "● 전체(승인 포함) / C 편집(DRAFT 작성) / A 승인 / R 조회 / — 접근 불가"),
        ("자동화", "권한매트릭스는 메뉴정의서에서 도출 — 메뉴 변경 시 make_authz_xlsx.py 재실행"),
        ("원칙", "① 권한 없는 항목 미표시 ② 미승인 자산 사용 불가 ③ System DB 영향 = Platform 승인"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 104

    # ---- 역할정의 ----
    ws = out.create_sheet("역할정의")
    ws.append(["역할(user_level)", "명칭", "책임 범위", "특이 규칙"])
    style_header(ws, 4)
    for row in ROLES:
        ws.append(list(row))
    body(ws, center_cols=(1,))
    ws.append([])
    ws.append(["액션", "정의", "", ""])
    style_header(ws, 2, row=ws.max_row)
    for a, d in ACTIONS:
        ws.append([a, d])
        for cell in ws[ws.max_row][:2]:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP
    ws.append([])
    ws.append(["리소스 단위", "HEAD_TAB / HIERARCHY(노드) / FEATURE(기능 코드) / TABLE — DB정의서 sys_role_permission", "", ""])
    ws[ws.max_row][0].font = BOLD_FONT
    ws[ws.max_row][1].font = BODY_FONT
    for i, w in enumerate([16, 18, 62, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 권한매트릭스 ----
    ws = out.create_sheet("권한매트릭스")
    headers = ["No", "메뉴 ID", "Lv1", "Lv2", "메뉴", "PLATFORM", "ADMIN", "SETUP", "GENERAL", "비고(부서 제한 등)"]
    ws.append(headers)
    style_header(ws, len(headers))
    fills = {"●": FILL_FULL, "C": FILL_EDIT, "A": FILL_EDIT, "R": FILL_VIEW, "—": FILL_NO}
    for n, m in enumerate(menus, 1):
        plat, adm, setup, gen, note = derive(m["perm"])
        ws.append([n, m["id"], m["lv1"], m["lv2"], m["lv3"], plat, adm, setup, gen, note])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 6, 7, 8, 9) else WRAP
        for ci, code in ((6, plat), (7, adm), (8, setup), (9, gen)):
            f = fills.get(code)
            if f:
                ws.cell(row=ws.max_row, column=ci).fill = f
    for i, w in enumerate([5, 9, 15, 18, 34, 10, 10, 10, 10, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---- 승인상태기계 ----
    ws = out.create_sheet("승인상태기계")
    headers = ["No", "대상 자산", "상태 흐름", "작성(요청) 권한", "검토·승인 권한", "선행 조건·특이 규칙", "관련"]
    ws.append(headers)
    style_header(ws, len(headers))
    for n, row in enumerate(FLOWS, 1):
        ws.append([n] + [x.replace("**", "") for x in row])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1,) else WRAP
    for i, w in enumerate([5, 20, 34, 16, 20, 52, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Platform승인범위 ----
    ws = out.create_sheet("Platform승인범위")
    ws.append(["No", "작업", "내용", "근거"])
    style_header(ws, 4)
    for n, row in enumerate(PLATFORM_SCOPE, 1):
        ws.append([n] + list(row))
    body(ws, center_cols=(1,))
    for i, w in enumerate([5, 26, 74, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- 문서보안등급 ----
    ws = out.create_sheet("문서보안등급")
    ws.append(["Grade", "명칭", "열람", "출력", "다운로드", "비고"])
    style_header(ws, 6)
    for row in GRADES:
        ws.append(list(row))
    body(ws, center_cols=(1,))
    ws.append([])
    ws.append(["※ 등급 수·명칭·통제 수준은 고객 협의 대상 (DOC-002/004, 보완노트 §3.3). 상위 Grade 문서는 목록에서 마스킹."])
    ws[ws.max_row][0].font = Font(name="맑은 고딕", size=10, color="C0392B")
    for i, w in enumerate([8, 16, 22, 16, 14, 44], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (매트릭스 {len(menus)}메뉴 · 승인 흐름 {len(FLOWS)} · Platform 범위 {len(PLATFORM_SCOPE)} · Grade {len(GRADES)})")


if __name__ == "__main__":
    main()
