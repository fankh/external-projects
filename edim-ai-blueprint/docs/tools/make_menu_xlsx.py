# -*- coding: utf-8 -*-
"""EDIM 메뉴정의서.xlsx 생성.

기준: 슬라이드 18 (H-1 Head System Set-Up), 56/59 (EDIM Toolbar), 기능정의서 v0.2
실행: py docs/tools/make_menu_xlsx.py  (저장: docs/EDIM_메뉴정의서.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "EDIM_메뉴정의서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
LV1_FILLS = {
    "EDIM Set-Up": "F5F0FF", "EDIM Toolbox": "F0FFF0", "CPQ Set-up": "E8F4FD",
    "Code(TLM) Set-up": "FFF8E7", "PLM Set-up": "FFEFC2", "CPQ": "E0F7FA",
    "ERP": "FFF0F5", "공통": "FDF2E9", "Mobile App": "EAF2F8",
}

# (메뉴ID, Lv1, Lv2, Lv3, 화면ID, 기능코드, 관련 기능ID, 권한, Phase, 비고)
MENUS = [
 # ---- EDIM Set-Up (Platform 전용) — 슬라이드 18 좌측 ----
 ("M-0-1", "EDIM Set-Up", "EDIM AI", "AI Study (DWG·Document)", "-", "-", "AI-001~003", "PLATFORM", "P5", "AI 학습은 Platform 제공자만"),
 ("M-0-2", "EDIM Set-Up", "EDIM AI", "AI Programming", "-", "-", "AI-004~006", "PLATFORM", "P4", ""),
 ("M-0-3", "EDIM Set-Up", "System Structure", "Main", "-", "E-1", "SYS-015/016", "PLATFORM", "P1", ""),
 ("M-0-4", "EDIM Set-Up", "System Structure", "Main Work Place", "-", "E-2", "DWG-025, CPQ-014", "PLATFORM", "P2", "Toolbar Module"),
 ("M-0-5", "EDIM Set-Up", "System Structure", "Key Work Place Hierarchy", "-", "E-3", "SYS-006~008", "PLATFORM", "P1", ""),
 ("M-0-6", "EDIM Set-Up", "System Structure", "Sub Work Place Templet", "-", "E-4", "TBX-013", "PLATFORM", "P2", ""),

 # ---- EDIM Toolbox — S-2-x ----
 ("M-1-1", "EDIM Toolbox", "UI Design", "CPQ", "W-08", "S-2-1", "TBX-001~004", "SETUP", "P4", "UI Designer 진입 컨텍스트별"),
 ("M-1-2", "EDIM Toolbox", "UI Design", "TLM", "W-08", "S-2-1", "TBX-001~004", "SETUP", "P4", ""),
 ("M-1-3", "EDIM Toolbox", "UI Design", "Print Form", "W-08", "S-2-1", "CPQ-013", "SETUP", "P3", ""),
 ("M-1-4", "EDIM Toolbox", "UI Design", "User ERP", "W-08", "S-2-1", "ERP-016", "SETUP", "P5", ""),
 ("M-1-5", "EDIM Toolbox", "UI Design", "Tool Bar", "W-08", "S-2-1", "CPQ-014", "SETUP", "P3", ""),
 ("M-1-6", "EDIM Toolbox", "Program Macro", "Templet", "W-07", "S-2-2", "TBX-013", "SETUP", "P4", ""),
 ("M-1-7", "EDIM Toolbox", "Program Macro", "AI (Macro·Flowchart·Coding)", "W-07", "S-2-2", "TBX-005~012", "SETUP", "P2~P4", "4-Way Sync"),

 # ---- CPQ Set-up — S-3-x ----
 ("M-2-1", "CPQ Set-up", "CPQ", "Selection", "-", "S-3-1", "CPQ-011", "SETUP", "P3", ""),
 ("M-2-2", "CPQ Set-up", "CPQ", "Technical", "-", "S-3-2", "CPQ-012", "SETUP", "P3", ""),
 ("M-2-3", "CPQ Set-up", "CPQ", "Document", "-", "S-3-3", "CPQ-012", "SETUP", "P3", ""),
 ("M-2-4", "CPQ Set-up", "CPQ", "Print Set-up", "-", "S-3-4", "CPQ-013", "SETUP", "P3", ""),
 ("M-2-5", "CPQ Set-up", "User Management", "User Customizing ERP", "-", "S-3-5", "ERP-001/016", "SETUP", "P5", ""),

 # ---- Code(TLM) Set-up — S-1-x ----
 ("M-3-1", "Code(TLM) Set-up", "Code", "Sub Code Spec.", "W-04", "S-1-1", "CODE-001~004/015", "SETUP", "P1", ""),
 ("M-3-2", "Code(TLM) Set-up", "Code", "Sub Code Raw Material·GPI", "W-04", "S-1-2", "CODE-005", "SETUP", "P1", "슬라이드 34 breadcrumb: Raw material Or GPI"),
 ("M-3-3", "Code(TLM) Set-up", "Code", "Product Code", "W-04", "S-1-3", "CODE-006/007", "SETUP", "P1", ""),
 ("M-3-4", "Code(TLM) Set-up", "Code", "Code Relationship", "W-05", "S-1-4", "CODE-008~010", "SETUP", "P1", "관계형 BOM"),
 ("M-3-5", "Code(TLM) Set-up", "Code", "Arrangement Code", "W-04", "S-1-5", "CODE-011", "SETUP", "P2", ""),
 ("M-3-6", "Code(TLM) Set-up", "Code", "Arrangement Set-up", "W-06", "S-1-6", "CODE-012", "SETUP", "P2", ""),
 ("M-3-7", "Code(TLM) Set-up", "Data", "Table·Variant·Constant·단위계(Metric)", "-", "E-4", "TBL-001~006", "SETUP", "P1", "슬라이드 46 — Constant List·Metric System은 기능정의서 v0.3 보강 필요"),

 # ---- PLM Set-up — S-4-x ----
 ("M-4-1", "PLM Set-up", "Design (DWG)", "Design", "W-06", "S-4-1-1", "DWG-001~016/022~027", "SETUP", "P2", "Drawing Editor — 표시 경로: Drawing Info. > Production Design Management > Design > Check > Approve > Accepted"),
 ("M-4-2", "PLM Set-up", "Work Process", "All Department", "-", "S-4-1-2", "DWG-021, ERP-025", "SETUP", "P3", "하위 실체: Design·Manufacturing·Material·Quality (제품별 Work Process Tree)"),
 ("M-4-3", "PLM Set-up", "Design (DWG)", "건축 설비 Design", "-", "E-4-2", "DUCT-001~007", "SETUP·GENERAL", "P5", "슬라이드 68 — 사업 범위 확정 대상"),

 # ---- CPQ 사용 — C-x (Enterprise User) ----
 ("M-5-1", "CPQ", "Selection", "Product", "W-02", "C-1", "CPQ-001~004/007~010", "GENERAL", "P2", "제품 선정"),
 ("M-5-2", "CPQ", "Technical", "Product·Item", "W-03", "C-2", "CPQ-005", "GENERAL", "P3", ""),
 ("M-5-3", "CPQ", "Document", "Product·Item", "-", "C-3", "CPQ-006", "GENERAL", "P3", ""),
 ("M-5-4", "CPQ", "Document Management", "문서함 (Set-up→Check→Approve→Accepted)", "-", "-", "DOC-001~004", "GENERAL", "P3", "슬라이드 20/58 breadcrumb — Released Status·Grade 관리"),

 # ---- ERP — 부서 Tab (§11 프로세스 코드), Toolbar 사용자 편집 가능 ----
 ("M-6-1", "ERP", "Sales", "고객 관리", "-", "§11", "ERP-015", "영업", "P5", "ERP Toolbar 사용자 편집 가능 (슬라이드 59)"),
 ("M-6-2", "ERP", "Sales", "Project 등록 (PS)", "W-09", "§11", "ERP-001", "영업", "P2", ""),
 ("M-6-3", "ERP", "Sales", "Project 관리", "W-09", "§11", "ERP-001/002", "영업", "P2", ""),
 ("M-6-4", "ERP", "Sales", "견적 검토 (PCR)", "-", "§11", "CST-005, ERP-004", "영업", "P3", ""),
 ("M-6-5", "ERP", "Sales", "견적 (QCR)", "-", "§11", "CST-006, ERP-004", "영업", "P3", ""),
 ("M-6-6", "ERP", "Sales", "수주 (OR)", "-", "§11", "ERP-004", "영업", "P5", ""),
 ("M-6-7", "ERP", "Sales", "승인 도서 (AP)·고객 승인 (APP)", "-", "§11", "ERP-005", "영업", "P5", ""),
 ("M-6-8", "ERP", "Sales", "출고 요청 (FDR)·고객 출고 (SFO)·납품 (DF)", "-", "§11", "ERP-009", "영업", "P5", ""),
 ("M-6-9", "ERP", "Sales", "기성 청구 (IR)", "-", "§11", "ERP-009", "영업", "P5", ""),
 ("M-6-10", "ERP", "Sales", "시운전 요청 (CR)", "-", "§11", "ERP-010", "영업", "P5", ""),
 ("M-7-1", "ERP", "Tech.", "제작 의뢰 (MR)·제작 점검 (MRR)", "-", "§11", "ERP-006", "기술", "P5", ""),
 ("M-7-2", "ERP", "Tech.", "Part List (PL)·BOM 조회", "-", "§11", "RUN-002/010", "기술", "P2", ""),
 ("M-7-3", "ERP", "Tech.", "제품 Set-up (PSU)·자동 제작 설계 (PLM)", "-", "§11", "RUN-003~005", "기술", "P2", ""),
 ("M-8-1", "ERP", "Purchasing", "구매 관리", "-", "§11", "ERP-017/018", "자재", "P5", ""),
 ("M-8-2", "ERP", "Purchasing", "자재발주요청 (PR)·발주 (PO)", "-", "§11", "ERP-007", "자재", "P5", ""),
 ("M-8-3", "ERP", "Material", "자재 입고 (MI)·출고 (MO)·수령 (RM)", "-", "§11", "ERP-007", "자재", "P5", ""),
 ("M-8-4", "ERP", "Material", "재고·창고 관리", "-", "§11", "ERP-019~021", "자재", "P5", ""),
 ("M-8-5", "ERP", "Material", "MRP 자재 소요 계획", "-", "§11", "ERP-022", "자재", "P5", ""),
 ("M-8-6", "ERP", "Purchasing", "일반 발주 (PR)·구매 승인 (RA)", "-", "§11", "ERP-017", "자재", "P5", "슬라이드 10 — 비프로젝트(재작업·공구 등) 구매"),
 ("M-9-1", "ERP", "Product(생산)", "생산계획 (MP)·Scheduling", "-", "§11", "ERP-023", "생산", "P5", ""),
 ("M-9-2", "ERP", "Product(생산)", "작업 지시 (WR)", "-", "§11", "ERP-024", "생산", "P5", ""),
 ("M-9-3", "ERP", "Product(생산)", "공정 관리", "-", "§11", "ERP-025", "생산", "P5", ""),
 ("M-9-4", "ERP", "Product(생산)", "반·완성품 (SFI·EF·FF)·Stock 생산 (SPP)", "-", "§11", "ERP-008", "생산", "P5", ""),
 ("M-9-5", "ERP", "Product(생산)", "외주 제조 관리", "-", "§11", "ERP-030", "생산", "P5", ""),
 ("M-9-6", "ERP", "Product(생산)", "재작업 요청 (WR)", "-", "§11", "ERP-029/030", "생산", "P5", "슬라이드 10"),
 ("M-10-1", "ERP", "QC", "자재 품질 검사 (ER)", "-", "§11", "ERP-027", "QC", "P5", ""),
 ("M-10-2", "ERP", "QC", "반·완성품 검사 (EF)", "-", "§11", "ERP-027", "QC", "P5", ""),
 ("M-10-3", "ERP", "QC", "부적합 처리 (NCF)", "-", "§11", "ERP-012", "QC", "P5", ""),
 ("M-10-4", "ERP", "QC", "인증서 발급 (CR)", "-", "§11", "ERP-030", "QC", "P5", "슬라이드 10"),
 ("M-11-1", "ERP", "A/S", "시운전 (CF)", "-", "§11", "ERP-010", "CS", "P5", ""),
 ("M-11-2", "ERP", "A/S", "유지보수", "-", "§11", "APP-006", "CS", "P5", ""),
 ("M-12-1", "ERP", "Finance", "세금계산 매입 (II)·매출 (IO)", "-", "§11", "ERP-011", "재무", "P5", ""),
 ("M-12-2", "ERP", "Finance", "매입 출금 (PA)·매출 입금 (PI)", "-", "§11", "ERP-011", "재무", "P5", ""),
 ("M-12-3", "ERP", "Finance", "정산 (RR)·정산승인 (RA)·재무관리 (FM)", "-", "§11", "ERP-011", "재무", "P5", ""),
 ("M-12-4", "ERP", "Finance", "Advance 선지급 (AR)", "-", "§11", "ERP-011", "재무", "P5", "슬라이드 10 — 기능정의서 v0.3 보강 필요"),
 ("M-13-1", "ERP", "HR", "임직원 관리·임직원 출금 (CO)", "-", "§11", "-", "HR", "P5", "상세 v0.3"),
 ("M-14-1", "ERP", "Company Info.", "ERP System·Department", "-", "§11", "SYS-019/020", "ADMIN", "P5", ""),
 ("M-14-2", "ERP", "Company Info.", "Company DB (고객·공급·파트너·은행)", "-", "§11", "ERP-015", "ADMIN", "P2", ""),
 ("M-14-3", "ERP", "Company Info.", "업체 평가 (CE)·Project 평가 (PE)", "-", "§11", "ERP-013", "ADMIN", "P5", ""),
 ("M-14-4", "ERP", "Company Info.", "Dashboard", "W-10", "§11", "ERP-014", "ADMIN·경영", "P5", ""),
 ("M-14-5", "ERP", "Company Info.", "PCR 기준 관리", "-", "§11", "CST-005", "ADMIN·재무", "P3", "슬라이드 18 — Company info. 하위 3. PCR"),
 ("M-14-6", "ERP", "Company Info.", "Access Control·Version Management", "-", "D-1", "SYS-003~005/012", "ADMIN", "P1", "슬라이드 72 — 권한 설정·버전 관리"),

 # ---- 공통 (모든 Head에서 접근) ----
 ("M-15-1", "공통", "로그인", "-", "W-01", "-", "SYS-001/002", "전체", "P1", "슬라이드 3"),
 ("M-15-2", "공통", "승인함", "승인 요청·처리·이력", "-", "-", "SYS-009~011", "전체", "P1", "좌측 하단 위젯 + 전용 화면"),
 ("M-15-3", "공통", "일정·To-do", "Schedule·To-do·Done", "-", "E-3", "SYS-017", "전체", "P2", ""),
 ("M-15-4", "공통", "알림", "알림 목록·공지", "-", "-", "SYS-013/014", "전체", "P2", ""),
 ("M-15-5", "공통", "업무 소통", "Project 중심 대화 (SNS)", "-", "-", "SYS-018", "전체", "P5", "슬라이드 57"),
 ("M-15-6", "공통", "검색", "Hierarchy·자료 통합 검색", "-", "E-3", "SYS-006", "전체", "P2", ""),
 ("M-15-7", "공통", "개인 설정", "색상·언어·판넬 화면 조정", "-", "E-1", "SYS-016", "전체", "P4", "슬라이드 57 — Frame/Canvas 색상, 아코디언·펼침"),
 ("M-15-8", "공통", "Project Folder", "산출물 탐색 (DWG·Price·Data·BOM)", "-", "E-3", "RUN-009", "전체", "P2", "슬라이드 64 Hierarchy List 4. Project"),

 # ---- Mobile App — 슬라이드 77 ----
 ("M-16-1", "Mobile App", "QR", "QR 정보 열람 (도면·서류·Project·업무)", "-", "§12", "APP-001", "전체", "P5", ""),
 ("M-16-2", "Mobile App", "승인", "업무 승인", "-", "§12", "APP-002", "ADMIN↑", "P5", ""),
 ("M-16-3", "Mobile App", "소통", "Project 중심 대화·History", "-", "§12", "APP-003", "전체", "P5", ""),
 ("M-16-4", "Mobile App", "자재", "자재 입출고", "-", "§12", "APP-004", "자재", "P5", ""),
 ("M-16-5", "Mobile App", "검수", "자재·완성품·설치완료 검수", "-", "§12", "APP-005", "QC", "P5", ""),
 ("M-16-6", "Mobile App", "유지보수", "Client 유지보수 접수·처리", "-", "§12", "APP-006", "CS", "P5", ""),
 ("M-16-7", "Mobile App", "공지", "공지·알림 수신", "-", "§12", "APP-007", "전체", "P5", ""),
 ("M-16-8", "Mobile App", "AR", "증강현실 View", "-", "§12", "APP-008", "현장", "P5", "후순위"),
]

HEADERS = ["No", "메뉴 ID", "Lv1 (Head)", "Lv2", "Lv3", "화면 ID", "기능코드", "관련 기능 ID", "사용 권한", "Phase", "비고"]
WIDTHS = [5, 9, 16, 20, 34, 8, 9, 22, 11, 8, 30]


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 메뉴정의서"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.2 — PPT 전수 대조 검증 반영 (누락 12개 영역 20메뉴 추가: Document Mgmt·건축설비·모바일·AR 선지급 등)"),
        ("작성일", "2026-07-07"),
        ("근거", "슬라이드 10/18/20/46/56~59/64/68/72/77 전수 대조, 기능정의서 v0.2"),
        ("메뉴 수", f"{len(MENUS)}개"),
        ("구조", "Lv1 = Head Tab / Lv2 = 작업 영역 / Lv3 = 메뉴 항목"),
        ("표시 규칙", "권한 없는 메뉴 미표시 (SYS-005), ERP Toolbar는 사용자 편집 가능 (ERP-016)"),
        ("화면 ID", "EDIM_화면설계서.html 의 W-01~W-10 와이어프레임과 연결"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 90

    ws = wb.create_sheet("메뉴목록")
    ws.append(HEADERS)
    style_header(ws, len(HEADERS))
    for n, row in enumerate(MENUS, 1):
        mid, lv1, lv2, lv3, wid, fcode, fids, perm, phase, note = row
        ws.append([n, mid, lv1, lv2, lv3, wid, fcode, fids, perm, phase, note])
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 6, 7, 9, 10) else WRAP
        ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=LV1_FILLS.get(lv1, "FFFFFF"))
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (메뉴 {len(MENUS)}개)")


if __name__ == "__main__":
    main()
