# -*- coding: utf-8 -*-
"""EDIM 컴포넌트정의서.xlsx 생성.

기준: docs/EDIM_컴포넌트_정의서.md
실행: py docs/tools/make_component_xlsx.py  (저장: docs/EDIM_컴포넌트정의서.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "EDIM_컴포넌트정의서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
PALETTE = {"클라이언트": "FFF8E7", "게이트웨이": "E8F4FD", "도메인 서비스": "F0FFF0",
           "엔진": "FFF0F5", "AI": "F5F0FF", "연계": "FFEFC2", "인프라": "E0F7FA"}

# (ID, 계층, 이름, 책임 요약, 기능코드, 주 DB, 의존, 기술 후보, Phase)
COMPONENTS = [
 ("FE-01", "클라이언트", "EDIM Web Console", "Enterprise User 메인 웹앱 — Main Work Frame(3판넬), CPQ 사용 화면, ERP 업무, Dashboard, Drawing Viewer(SVG), 동적 Form 렌더러", "E-1~E-4, C-1~C-3", "-", "GW-01, SVC-02/03/04/05/07/08/09/10, ENG-02", "React+TypeScript+Vite", "P1~"),
 ("FE-02", "클라이언트", "Set-up Studio", "시스템 구축 도구 — 코드 등록(중복검토·Running Test), Drawing Editor(CAD형 캔버스·치수·Part Relation·검증), Work Process, CPQ Set-up", "S-1-1~6, S-3-1~4, S-4-1-1/2", "-", "SVC-02/03/04/05/06/10, ENG-01", "FE-01 스택 + Canvas 에디터 모듈", "P1~"),
 ("FE-03", "클라이언트", "Toolbox Designer", "UI Designer(위젯 팔레트 Drag&Drop, AI 제안) + Macro Studio(4-Way Sync, 함수/그래프 마법사, Test Run)", "S-2-1, S-2-2", "-", "SVC-05/06/10, ENG-01, AI-03/04", "FE-01 스택", "P4"),
 ("FE-04", "클라이언트", "Admin Console", "테넌트·사용자·권한·승인 관리, AI 학습 관리(Platform 전용), 감사·사용량", "-", "-", "SVC-01/02/10, AI-02, INF-07", "FE-01 스택", "P1~"),
 ("FE-05", "클라이언트", "EDIM App (Mobile)", "QR 열람·모바일 승인·Project 소통·자재 입출고·검수·유지보수·AR", "§12", "-", "GW-01, SVC-04/09/10/13, INT-02/05", "React Native 또는 Flutter", "P5"),
 ("GW-01", "게이트웨이", "API Gateway", "단일 진입점 — TLS·JWT 검증·테넌트 해석·라우팅·Rate Limit", "-", "-", "SVC-01", "Kong/Traefik/SCG (미결정)", "P1"),
 ("SVC-01", "게이트웨이", "Auth Service", "인증(OIDC)·토큰·RBAC 평가(HEAD_TAB/HIERARCHY/FEATURE/TABLE × VIEW/EDIT/APPROVE/SETUP)", "-", "sys_user/role/permission", "INF-01/04", "코어 모듈", "P1"),
 ("SVC-02", "도메인 서비스", "Hierarchy Service", "Tree CRUD·이동·Address 발급/추적·심볼·검색·관계 점검", "E-3, H-1", "sys_hierarchy, sys_history", "SVC-10", "코어 모듈", "P1"),
 ("SVC-03", "도메인 서비스", "Code Service (RCCS)", "코드 그룹/자릿수/값·Product Code·Mother-Child 관계·BOM 전개(재귀)·Running Test·Arrangement", "S-1-1~6, D-1", "code_* 8테이블", "SVC-02/05/10", "코어 모듈", "P1"),
 ("SVC-04", "도메인 서비스", "Drawing Service (PLM)", "도면·개정·승인, DrawingDocument(Block) 저장/호출, 치수·부품관계·검증 규칙, Import/Export", "S-4-1-1, E-4", "dwg_* 12, prt_, mat_", "INT-04, ENG-03, SVC-10/12", "코어 모듈", "P2"),
 ("SVC-05", "도메인 서비스", "Table Service", "데이터 Table 정의·행 관리·Excel Import·Macro 참조 범위 조회", "E-4", "tbl_data_table/row", "INT-03, INF-04", "코어 모듈", "P1"),
 ("SVC-06", "도메인 서비스", "Toolbox Service", "UI Form·Templet 저장/버전/승인, Form 렌더 정의 서빙", "S-2-1", "tbx_ui_form/templet", "SVC-10", "코어 모듈", "P4"),
 ("SVC-07", "도메인 서비스", "CPQ Service", "선정 세션(slot·사양·Excel Import), Arrangement 구성, 완성품 Code, X-Code 분기, Run 요청", "C-1~C-3, S-3-1~3", "cpq_* 4테이블", "SVC-03, ENG-02", "코어 모듈", "P2"),
 ("SVC-08", "도메인 서비스", "Cost Service", "4종 단가·자재비/제조비/직접비·PCR(EBIT)·견적서", "D-3, D-4", "cst_* 4테이블", "SVC-11, ENG-02", "코어 모듈", "P3"),
 ("SVC-09", "도메인 서비스", "ERP Process Service", "프로세스 정의(40종)·이벤트 상태기계·To-do·Work Process·Dashboard 집계", "S-3-5, S-4-1-2, §11", "erp_*, com_company", "SVC-13", "코어 모듈", "P5"),
 ("SVC-10", "도메인 서비스", "Approval Service", "공통 승인 워크플로우 — 요청→검토→승인/반려, approval_status 전이", "전 영역", "sys_approval_request", "SVC-13", "코어 모듈", "P1"),
 ("SVC-11", "도메인 서비스", "Document/Print Service", "Print Form 렌더(PDF), 워터마크·머리글/바닥글, Office 내보내기", "S-3-4", "tbx_ui_form(PRINT), dwg_file", "SVC-12", "Headless Chromium/리포트엔진", "P3"),
 ("SVC-12", "도메인 서비스", "File Service", "Object Storage 추상화, Project Folder 규약, 서명 URL", "-", "dwg_file", "INF-02", "코어 모듈", "P2"),
 ("SVC-13", "도메인 서비스", "Notification Service", "알림 생성/조회, SSE·WebSocket·모바일 푸시, 공지", "-", "sys_notification", "INF-03", "코어 모듈", "P4"),
 ("ENG-01", "엔진", "Macro Engine", "Excel 문법 파서(AST)·참조 해석(Table/Var/PreC)·평가기(공학 함수)·DAG 순환 검증·Test Runner", "S-2-2", "tbx_macro/_ref", "SVC-05", "코어 내장 라이브러리+샌드박스", "P1"),
 ("ENG-02", "엔진", "EDIM Run Orchestrator", "비동기 잡 워커 — BOM→치수 Macro→도면→원가→기술자료→서류, 산출물 저장, 부분 실패 처리", "D-2", "cpq_run/output", "SVC-03/08/12, ENG-01/03, INF-03", "별도 워커 프로세스", "P2"),
 ("ENG-03", "엔진", "Drawing Generation Engine", "Block+치수값→완성 도면 합성(Parametric), DXF(R2010)/PDF 출력, 2D→3D 연계", "D-2, §7.1", "dwg_document", "SVC-04", "ezdxf 등 — 프로토타입 승계", "P2"),
 ("AI-01", "AI", "AI Gateway", "LLM 호출 단일화(Claude API), 프롬프트 템플릿·버전, 토큰/비용 계측, PII 필터", "-", "-", "외부 LLM API", "AI 워커", "P4"),
 ("AI-02", "AI", "Drawing Learning Pipeline", "CAD 변환→메타데이터/2D 엔티티/OCR/3D Feature 추출→RAG 인덱싱 (Platform 전용)", "§9", "(RAG 인덱스)", "INT-04, INF-05, AI-01", "AI 워커", "P5"),
 ("AI-03", "AI", "Macro Generation", "Prompt→Macro/Flowchart/Description/Coding 생성, 4-Way 상호 변환, 기능 찾기", "S-2-2", "tbx_macro", "ENG-01, AI-01", "AI 워커", "P4"),
 ("AI-04", "AI", "UI Generation", "Application 설명→용도/항목/필요 DB Table→UI Form 초안", "S-2-1", "tbx_ui_form", "SVC-06, AI-01", "AI 워커", "P4"),
 ("AI-05", "AI", "Knowledge Chatbot", "사내 자료 RAG Q&A, ERP 연동 서류 생성 지원", "§9", "(RAG 인덱스)", "INF-05, AI-01", "AI 워커", "P5"),
 ("INT-01", "연계", "External ERP Connector", "고객사 기존 ERP와 BOM·발주·회계 교환 (어댑터 플러그블)", "-", "-", "SVC-09", "대상 ERP 미확정", "P5"),
 ("INT-02", "연계", "Digital Twin Connector", "DTDesigner 연계 — 치수 Table·3D·실시간 Code, AR/XR 데이터", "§12", "-", "ENG-03", "-", "P5"),
 ("INT-03", "연계", "Excel Import/Export", "고객 사양·Table 행 Import, 견적·BOM Export", "-", "-", "SVC-05/07", "openpyxl/SheetJS", "P1"),
 ("INT-04", "연계", "CAD Converter", "DWG↔DXF(ODA), PDF→DXF, 2D→3D(STEP) 변환 워커", "-", "-", "SVC-04, AI-02", "ODA File Converter 격리 실행", "P2"),
 ("INT-05", "연계", "QR Service", "도면·서류·Project·자산 QR 발급/해석", "§12", "-", "FE-05", "-", "P5"),
 ("INF-01", "인프라", "PostgreSQL 16 (HA)", "주 DB — 46테이블", "-", "-", "-", "-", "P1"),
 ("INF-02", "인프라", "Object Storage", "S3 호환 — CAD·PDF·이미지, Project Folder", "-", "-", "-", "-", "P2"),
 ("INF-03", "인프라", "Message Queue", "Run·AI·변환 잡, 이벤트 (Redis Streams/RabbitMQ)", "-", "-", "-", "-", "P2"),
 ("INF-04", "인프라", "Redis Cache", "세션·권한·Table 조회 캐시", "-", "-", "-", "-", "P1"),
 ("INF-05", "인프라", "Vector Store", "AI RAG 인덱스 (pgvector/OpenSearch)", "-", "-", "-", "-", "P4"),
 ("INF-06", "인프라", "Kubernetes + CI/CD", "SaaS 멀티테넌트 / Self-managed 패키지 겸용 배포", "-", "-", "-", "-", "P1"),
 ("INF-07", "인프라", "모니터링", "Prometheus·Grafana·Loki — 메트릭·로그·알림", "-", "-", "-", "-", "P1"),
]
# 구축 상태 (2026-07-09, 개발 서버 edim.seekerslab.com 기준)
STATUS = {
 "FE-01":  ("구축중", "edim-web(React19+Vite+TS, dense B안) — 6모듈 셸(MDI·F-key), 와이어프레임 24화면 전량 + 상세 드릴다운 4종, mock API(OpenAPI 대응). 루트(/)는 Drawing Viewer 프로토타입 유지"),
 "FE-02":  ("구축중", "edim-web /code·/plm — Sub Code·Relationship(Running Test)·데이터 Table·Design Editor·Work Process·Duct — mock API"),
 "FE-03":  ("구축중", "edim-web /toolbox — Macro Studio(4-Way Sync·TESTED 게이트)·UI Designer(팔레트·Inspector·AI 초안) — mock"),
 "GW-01":  ("개발 대체 구축", "nginx 1.24 — TLS(Let's Encrypt)·라우팅(/api·/jenkins·/minio/ui)·SPA 서빙"),
 "SVC-01": ("구축중", "/api/v1/auth/login — sys_user 실검증(sha256)·HMAC 토큰"),
 "SVC-03": ("구축중", "/api/v1 codes — group slots·BOM expand(재귀 CTE+slot_map, 실 PG)"),
 "SVC-05": ("구축중", "/api/v1/tables/tech-data — tbl_data_row(row_key_num) 조회"),
 "SVC-07": ("구축중", "/api/v1/cpq/runs — 202+폴링, cpq_run/cpq_output 영속 (파이프라인 단계는 mock)"),
 "SVC-08": ("구축중", "/api/v1/prices/resolve — cst_price 우선순위 resolve(APPLIED→…→QUOTE)"),
 "SVC-04": ("프로토타입 일부", "FastAPI backend — DXF/IFC Import·DXF Export·DrawingDocument JSON"),
 "ENG-03": ("프로토타입 일부", "ezdxf 기반 DXF(R2010) Export 경로 검증됨"),
 "AI-01":  ("프로토타입 일부", "Claude API 연동(ai_generator) — ANTHROPIC_API_KEY 미설정 시 샘플 모드"),
 "INT-04": ("프로토타입", "DwgToDxfConverter 플러그블 인터페이스 — ODA 바이너리 미설치(501)"),
 "INF-01": ("구축완료", "PostgreSQL 16 컨테이너(edim-postgres, 127.0.0.1:5432) — EDIM 스키마 53테이블 적용·검증됨"),
 "INF-02": ("구축완료", "MinIO — S3 API 127.0.0.1:9000(내부), 콘솔 /minio/ui, 버킷 edim, 볼륨 minio_data"),
 "INF-06": ("부분 구축", "Docker 29.6.1 + Compose v5.3, Jenkins LTS(/jenkins) — k8s는 운영 전환 시 도입"),
}
DEFAULT_STATUS = ("미착수", "")

# (서비스, Method, Path, 설명)
APIS = [
 ("SVC-01 Auth", "POST", "/api/v1/auth/login", "로그인 — 토큰 발급"),
 ("SVC-01 Auth", "POST", "/api/v1/auth/refresh", "토큰 갱신"),
 ("SVC-01 Auth", "GET", "/api/v1/auth/me", "내 정보·권한 요약"),
 ("SVC-01 Auth", "PATCH", "/api/v1/auth/password", "비밀번호 변경 (정책 검증)"),
 ("SVC-01 Auth", "GET", "/api/v1/i18n/{locale}", "데이터 라벨 번역 번들 (ko/en/ja/zh, KO 폴백) — SYS-021"),
 ("SVC-01 Auth", "PUT", "/api/v1/i18n/translations", "번역 등록·수정 (ADMIN, 일괄 Import 지원)"),
 ("SVC-01 Auth", "GET", "/api/v1/auth/permissions", "리소스별 권한 조회"),
 ("SVC-01 Auth", "GET", "/api/v1/users", "사용자 목록 (부서·상태 필터, ADMIN)"),
 ("SVC-01 Auth", "POST", "/api/v1/users", "사용자 등록 (SYS-003)"),
 ("SVC-01 Auth", "PATCH", "/api/v1/users/{id}", "사용자 수정·상태 변경 (ACTIVE/LOCKED/RETIRED)"),
 ("SVC-01 Auth", "POST", "/api/v1/roles", "역할 생성·권한 부여 (SYS-004/005)"),
 ("SVC-01 Auth", "GET", "/api/v1/tenants", "테넌트 목록·상태 (Platform 전용)"),
 ("SVC-01 Auth", "POST", "/api/v1/tenants", "테넌트 생성 — sys_tenant (ADM-001, Platform 전용)"),
 ("SVC-02 Hierarchy", "GET", "/api/v1/hierarchy/trees", "Tree 목록 (type 필터)"),
 ("SVC-02 Hierarchy", "GET", "/api/v1/hierarchy/nodes/{id}/children", "하위 노드 조회"),
 ("SVC-02 Hierarchy", "POST", "/api/v1/hierarchy/nodes", "노드 생성 (승인 대상)"),
 ("SVC-02 Hierarchy", "PATCH", "/api/v1/hierarchy/nodes/{id}/move", "노드 이동 — Address 추적"),
 ("SVC-02 Hierarchy", "GET", "/api/v1/hierarchy/resolve", "address→노드 해석"),
 ("SVC-02 Hierarchy", "GET", "/api/v1/hierarchy/search", "이름·비고·심볼 통합 검색 (M-15-6)"),
 ("SVC-02 Hierarchy", "PATCH", "/api/v1/hierarchy/nodes/{id}", "노드 개명·속성 수정 (승인 대상)"),
 ("SVC-02 Hierarchy", "DELETE", "/api/v1/hierarchy/nodes/{id}", "노드 삭제 — DB 상호관계 점검 후"),
 ("SVC-03 Code", "POST", "/api/v1/codes/groups", "코드 그룹 등록"),
 ("SVC-03 Code", "POST", "/api/v1/codes/products", "Product Code 등록"),
 ("SVC-03 Code", "GET", "/api/v1/codes/products/{id}/check-duplicate", "Main Code 중복검토"),
 ("SVC-03 Code", "POST", "/api/v1/codes/relationships", "Mother-Child 관계 등록"),
 ("SVC-03 Code", "POST", "/api/v1/codes/products/{id}/expand", "BOM 전개"),
 ("SVC-03 Code", "POST", "/api/v1/codes/relationships/running-test", "Part List Running Test"),
 ("SVC-03 Code", "POST", "/api/v1/codes/arrangements", "Arrangement 등록"),
 ("SVC-03 Code", "POST", "/api/v1/codes/values/import-excel", "코드 값 Excel Import (CODE-016)"),
 ("SVC-03 Code", "GET", "/api/v1/codes/groups/{id}/export-excel", "Registered Code Table Export (CODE-016)"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings", "도면 등록"),
 ("SVC-04 Drawing", "GET", "/api/v1/drawings/{id}/document", "DrawingDocument 조회 (block 파라미터)"),
 ("SVC-04 Drawing", "PUT", "/api/v1/drawings/{id}/document", "DrawingDocument 저장"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/{id}/revisions", "개정 생성"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/{id}/dimensions", "치수 등록 (Macro 바인딩)"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/{id}/relations", "부품 관계 등록"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/import", "DXF/DWG/PDF/STEP Import"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/{id}/export", "DXF/PDF Export"),
 ("SVC-04 Drawing", "GET", "/api/v1/drawings/{id}/referencers", "Where-used 역참조 조회 (DWG-026)"),
 ("SVC-04 Drawing", "GET", "/api/v1/drawings/{id}/variants", "Variants 목록 (DWG-027)"),
 ("SVC-04 Drawing", "POST", "/api/v1/drawings/{id}/supersede", "대체 관계 등록 — old→new (DWG-027)"),
 ("SVC-04 Drawing", "PATCH", "/api/v1/drawings/{id}/dimensions/{dimId}", "치수 Macro/Variant 바인딩 수정 (XOR 검증)"),
 ("SVC-04 Drawing", "GET", "/api/v1/parts", "부품 마스터 목록·검색 (DWG-017)"),
 ("SVC-04 Drawing", "POST", "/api/v1/parts", "부품 등록"),
 ("SVC-04 Drawing", "POST", "/api/v1/materials", "재질 등록 — hazard_class 포함 (DWG-018)"),
 ("SVC-05 Table", "POST", "/api/v1/tables", "Table 정의 생성"),
 ("SVC-05 Table", "POST", "/api/v1/tables/{id}/rows:bulk", "행 일괄 입력"),
 ("SVC-05 Table", "GET", "/api/v1/tables/{id}/query", "범위 조회 (Macro 참조용)"),
 ("SVC-05 Table", "POST", "/api/v1/tables/import-excel", "Excel Import"),
 ("SVC-06 Toolbox", "POST", "/api/v1/toolbox/forms", "UI Form 저장"),
 ("SVC-06 Toolbox", "GET", "/api/v1/toolbox/forms/{id}", "Form 렌더 정의 조회"),
 ("SVC-06 Toolbox", "POST", "/api/v1/toolbox/forms/{id}/publish", "Form 게시 (승인 후)"),
 ("SVC-06 Toolbox", "GET", "/api/v1/toolbox/templets", "Templet 목록"),
 ("SVC-06 Toolbox", "DELETE", "/api/v1/toolbox/forms/{id}", "DRAFT Form 삭제 (게시본 불가)"),
 ("SVC-07 CPQ", "GET", "/api/v1/cpq/selections", "선정 세션 목록 (project·상태 필터)"),
 ("SVC-07 CPQ", "POST", "/api/v1/cpq/selections", "선정 세션 생성"),
 ("SVC-07 CPQ", "PATCH", "/api/v1/cpq/selections/{id}/slots", "Slot 값 선택"),
 ("SVC-07 CPQ", "POST", "/api/v1/cpq/selections/{id}/finalize", "완성품 Code 확정"),
 ("SVC-07 CPQ", "POST", "/api/v1/cpq/selections/{id}/x-code-review", "비규격 검토 요청"),
 ("SVC-07 CPQ", "POST", "/api/v1/cpq/selections/{id}/runs", "EDIM Run 실행 요청"),
 ("SVC-07 CPQ", "GET", "/api/v1/cpq/runs/{runId}", "Run 상태 조회"),
 ("SVC-07 CPQ", "GET", "/api/v1/cpq/runs/{runId}/outputs", "산출물 목록"),
 ("SVC-07 CPQ", "GET", "/api/v1/cpq/runs/{runId}/bom/export", "BOM Excel Export (RUN-010)"),
 ("SVC-08 Cost", "POST", "/api/v1/cost/prices", "단가 등록"),
 ("SVC-08 Cost", "GET", "/api/v1/cost/prices/resolve", "적용 단가 해석 (source 우선순위)"),
 ("SVC-08 Cost", "POST", "/api/v1/cost/pcr", "PCR 생성"),
 ("SVC-08 Cost", "POST", "/api/v1/cost/quotations", "견적서 생성"),
 ("SVC-08 Cost", "POST", "/api/v1/cost/quotations/{id}/render", "견적서 PDF 출력"),
 ("SVC-09 ERP", "POST", "/api/v1/projects", "Project 등록 PS — 채번 (ERP-001)"),
 ("SVC-09 ERP", "GET", "/api/v1/projects", "Project 목록 — 영업단계 필터"),
 ("SVC-09 ERP", "GET", "/api/v1/erp/processes", "프로세스 정의 목록"),
 ("SVC-09 ERP", "POST", "/api/v1/erp/process-defs", "프로세스 정의 생성 (W-14, ADMIN — 코드 신설은 Platform)"),
 ("SVC-09 ERP", "PATCH", "/api/v1/erp/process-defs/{id}", "정의 수정 — 선행/후행(edge)·기한·Form"),
 ("SVC-09 ERP", "POST", "/api/v1/erp/events", "프로세스 이벤트 생성"),
 ("SVC-09 ERP", "PATCH", "/api/v1/erp/events/{id}/status", "이벤트 상태 전이"),
 ("SVC-09 ERP", "GET", "/api/v1/erp/events", "업무함 (assignee/status 필터)"),
 ("SVC-09 ERP", "GET", "/api/v1/erp/dashboard", "전사 Dashboard — 부서별 Event·이상 경고 (W-10)"),
 ("SVC-09 ERP", "GET", "/api/v1/erp/dashboard/project/{projectId}", "프로젝트 Dashboard 집계"),
 ("SVC-09 ERP", "POST", "/api/v1/erp/work-processes", "Work Process 공정 데이터 등록"),
 ("SVC-09 ERP", "POST", "/api/v1/erp/supplier-code-maps", "사용자↔공급자 코드 매핑 (ERP-018)"),
 ("SVC-10 Approval", "POST", "/api/v1/approvals", "승인 요청 생성"),
 ("SVC-10 Approval", "GET", "/api/v1/approvals/inbox", "승인함"),
 ("SVC-10 Approval", "POST", "/api/v1/approvals/{id}/decide", "승인/반려 결정"),
 ("SVC-10 Approval", "GET", "/api/v1/approvals/history", "승인 이력"),
 ("SVC-11 Print", "POST", "/api/v1/documents/render", "Print Form + 데이터 → PDF"),
 ("SVC-11 Print", "POST", "/api/v1/documents/export-office", "xlsx/docx 내보내기"),
 ("SVC-11 Print", "GET", "/api/v1/doc-controls", "문서함 목록 — Released Status·Grade 필터 (DOC-001)"),
 ("SVC-11 Print", "PATCH", "/api/v1/doc-controls/{id}/status", "Released Status 전이 — Set-up→Check→Approve→Accepted"),
 ("SVC-11 Print", "POST", "/api/v1/doc-controls/allocate-code", "Document Code 채번 (DOC-003)"),
 ("SVC-12 File", "POST", "/api/v1/files/upload-url", "업로드 서명 URL"),
 ("SVC-12 File", "GET", "/api/v1/files/{id}/download-url", "다운로드 서명 URL"),
 ("SVC-12 File", "GET", "/api/v1/files", "Project Folder 파일 목록"),
 ("SVC-13 Notify", "GET", "/api/v1/notifications", "알림 목록"),
 ("SVC-13 Notify", "POST", "/api/v1/notifications/read", "읽음 처리"),
 ("SVC-13 Notify", "WS", "/ws/notifications", "실시간 알림 채널"),
 ("ENG-01 Macro", "GET", "/api/v1/macros", "Macro 목록·검색 (v0.4 — CRUD 부재 결함 수정)"),
 ("ENG-01 Macro", "POST", "/api/v1/macros", "Macro 생성 — 4표현 저장, 참조 추출·DAG 검증"),
 ("ENG-01 Macro", "GET", "/api/v1/macros/{id}", "Macro 상세 (버전 포함)"),
 ("ENG-01 Macro", "PATCH", "/api/v1/macros/{id}", "Macro 수정 — 새 버전 생성"),
 ("ENG-01 Macro", "DELETE", "/api/v1/macros/{id}", "DRAFT Macro 삭제 (승인본 삭제 불가)"),
 ("ENG-01 Macro", "POST", "/api/v1/macros/{id}/test-run", "Macro Test Run (조건값→결과)"),
 ("AI", "POST", "/api/v1/ai/macro/generate", "Prompt→Macro 생성"),
 ("AI", "POST", "/api/v1/ai/macro/convert", "표현 간 변환 (from/to)"),
 ("AI", "POST", "/api/v1/ai/ui/suggest", "UI Form 초안 제안"),
 ("AI", "POST", "/api/v1/ai/chat", "사내 자료 챗봇"),
 ("AI", "POST", "/api/v1/ai/learning/jobs", "학습 잡 실행 (Platform 전용)"),
 ("INT-05 QR", "POST", "/api/v1/qr/issue", "도면·서류·Project·자산 QR 발급 (APP-001)"),
 ("INT-05 QR", "GET", "/api/v1/qr/resolve", "QR 해석 → 대상 링크·권한 검사"),
]


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 컴포넌트정의서"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1 (초안)"), ("작성일", "2026-07-07"),
        ("기준 문서", "EDIM_컴포넌트_정의서.md (아키텍처 다이어그램은 MD 참조)"),
        ("아키텍처 방침", "모듈러 모놀리스 코어 + 분리형 워커(Run·AI·변환)"),
        ("컴포넌트 수", f"{len(COMPONENTS)}개"),
        ("API 수", f"{len(APIS)}개 (대표 엔드포인트 — 상세 스펙은 OpenAPI 원천, 개발표준 §3)"),
        ("API 규약", "REST /api/v1, 커서 페이지네이션, RFC 9457 오류, tenant 필수"),
        ("Phase 정의", "P1 RCCS코어 / P2 설계·Run / P3 원가·문서 / P4 Toolbox·AI / P5 ERP·모바일"),
        ("구축 현황 기준", "2026-07-07 — 개발 서버 edim.seekerslab.com (Ubuntu 24.04, Docker·nginx·HTTPS·ufw, MinIO·Jenkins 가동)"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 80

    ws = wb.create_sheet("컴포넌트목록")
    headers = ["No", "ID", "계층", "이름", "책임", "기능코드", "주 DB", "의존", "기술 후보", "Phase", "구축 상태", "구축 내역 (2026-07-07)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for n, row in enumerate(COMPONENTS, 1):
        cid, layer, name, resp, code, db, dep, tech, phase = row
        st, detail = STATUS.get(cid, DEFAULT_STATUS)
        ws.append([n, cid, layer, name, resp, code, db, dep, tech, phase, st, detail])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 10, 11) else WRAP
        ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=PALETTE.get(layer, "FFFFFF"))
        if st == "구축완료":
            ws.cell(row=ws.max_row, column=11).fill = PatternFill("solid", fgColor="C8E6C9")
        elif st in ("부분 구축", "프로토타입 배포", "개발 대체 구축"):
            ws.cell(row=ws.max_row, column=11).fill = PatternFill("solid", fgColor="FFF3C4")
        elif st.startswith("프로토타입"):
            ws.cell(row=ws.max_row, column=11).fill = PatternFill("solid", fgColor="E3ECF7")
    for i, w in enumerate([5, 9, 13, 24, 60, 16, 22, 26, 24, 7, 12, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    ws = wb.create_sheet("API목록")
    headers = ["No", "서비스", "Method", "Path", "설명"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for n, row in enumerate(APIS, 1):
        ws.append([n] + list(row))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 3) else WRAP
    for i, w in enumerate([5, 18, 9, 50, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (컴포넌트 {len(COMPONENTS)}개, API {len(APIS)}개)")


if __name__ == "__main__":
    main()
