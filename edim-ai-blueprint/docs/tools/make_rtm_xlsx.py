# -*- coding: utf-8 -*-
"""EDIM 요구사항추적표(RTM).xlsx 생성 — 기존 정의서 xlsx를 조인하여 자동 생성.

체인: 요구사항(REQ) → 기능 ID → [기능정의서: 기능명·모듈·컴포넌트·DB·Phase]
                              → [메뉴정의서: 메뉴 ID·화면 ID]
커버리지 검증: REQ 미연결 기능 / 해석 불가 참조 목록 포함.
실행: py docs/tools/make_rtm_xlsx.py  (저장: docs/EDIM_요구사항추적표.xlsx)
"""
import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..")
F_REQ = os.path.join(BASE, "02_요구사항", "EDIM_요구사항정의서.xlsx")
F_FEAT = os.path.join(BASE, "EDIM_기능정의서.xlsx")
F_MENU = os.path.join(BASE, "EDIM_메뉴정의서.xlsx")
OUT = os.path.join(BASE, "EDIM_요구사항추적표.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
WARN_FILL = PatternFill("solid", fgColor="FDEAEA")

ID_RE = re.compile(r"^([A-Z]+)-(.+)$")


def expand_ids(expr):
    """'CODE-001~005/015, ERP-021' → {CODE-001..005, CODE-015, ERP-021}"""
    out = set()
    if not expr:
        return out
    for token in str(expr).split(","):
        token = token.strip()
        m = ID_RE.match(token)
        if not m:
            continue
        prefix, rest = m.group(1), m.group(2)
        for seg in rest.split("/"):
            seg = seg.strip()
            if "~" in seg:
                a, b = seg.split("~")
                a, b = a.strip(), b.strip()
                if a.isdigit() and b.isdigit():
                    width = len(a)
                    for n in range(int(a), int(b) + 1):
                        out.add(f"{prefix}-{str(n).zfill(width)}")
            elif seg.isdigit():
                out.add(f"{prefix}-{seg}")
    return out


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    # ---- 기능정의서 로드 ----
    feat = {}  # ID -> dict
    wb = openpyxl.load_workbook(F_FEAT, read_only=True)
    for row in wb["기능목록"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            feat[row[1]] = {"module": row[2], "name": row[4], "comp": row[8],
                            "db": row[9], "phase": row[10], "assignee": row[11] if len(row) > 11 else "미정"}
    wb.close()

    # ---- 메뉴정의서 로드: 기능ID -> [(메뉴ID, 화면ID)] ----
    f2menu = {}
    wb = openpyxl.load_workbook(F_MENU, read_only=True)
    for row in wb["메뉴목록"].iter_rows(min_row=2, values_only=True):
        mid, screen, fids = row[1], row[5], row[7]
        for fid in expand_ids(fids):
            f2menu.setdefault(fid, []).append((mid, screen if screen and screen != "-" else None))
    wb.close()

    # ---- 요구사항정의서 로드 ----
    reqs = []  # (REQ ID, 구분, 명, 중요도, Phase, expr)
    wb = openpyxl.load_workbook(F_REQ, read_only=True)
    for row in wb["기능요구사항"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            reqs.append({"id": row[1], "grp": row[2], "name": row[3],
                         "imp": row[5], "phase": row[6], "expr": row[7]})
    wb.close()

    # ---- 조인 ----
    rtm_rows = []
    covered = set()
    unresolved = []
    for r in reqs:
        fids = sorted(expand_ids(r["expr"]))
        if not fids:
            unresolved.append((r["id"], r["expr"], "기능 ID 해석 불가"))
            continue
        for fid in fids:
            f = feat.get(fid)
            if not f:
                unresolved.append((r["id"], fid, "기능정의서에 없음"))
                continue
            covered.add(fid)
            menus = f2menu.get(fid, [])
            menu_ids = " · ".join(sorted({m for m, _ in menus})) or "-"
            screens = " · ".join(sorted({s for _, s in menus if s})) or "-"
            rtm_rows.append([r["id"], r["name"], fid, f["name"], f["module"],
                             menu_ids, screens, f["comp"], f["db"], f["phase"], f["assignee"]])
    orphan_feats = sorted(set(feat) - covered)

    # ---- 출력 ----
    out = Workbook()
    ws = out.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 요구사항추적표 (RTM)"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1 (자동 생성)"),
        ("작성일", "2026-07-07"),
        ("생성 방법", "make_rtm_xlsx.py — 요구사항정의서·기능정의서·메뉴정의서 xlsx 조인 (수정 시 재생성)"),
        ("추적 체인", "요구사항 → 기능 → 메뉴 → 화면 → 컴포넌트 → DB → Phase"),
        ("추적 행", f"{len(rtm_rows)}건 (요구사항 {len(reqs)} × 기능 매핑)"),
        ("커버리지", f"기능 {len(covered)}/{len(feat)} 연결 · 미연결 {len(orphan_feats)}건 · 해석불가 {len(unresolved)}건 (커버리지 시트 참조)"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 96

    ws = out.create_sheet("추적표")
    headers = ["No", "요구사항 ID", "요구사항명", "기능 ID", "기능명", "모듈",
               "메뉴 ID", "화면 ID", "컴포넌트", "주 DB", "Phase", "작업자"]
    ws.append(headers)
    style_header(ws, len(headers))
    prev_req = None
    for n, row in enumerate(rtm_rows, 1):
        ws.append([n] + row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 4, 8, 11, 12) else WRAP
        if row[0] != prev_req:
            ws.cell(row=ws.max_row, column=2).font = BOLD_FONT
            prev_req = row[0]
    for i, w in enumerate([5, 12, 22, 10, 24, 12, 20, 12, 20, 24, 7, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    ws = out.create_sheet("커버리지")
    ws.append(["구분", "대상", "내용"])
    style_header(ws, 3)
    for fid in orphan_feats:
        f = feat[fid]
        ws.append(["REQ 미연결 기능", fid, f"{f['name']} ({f['module']}, {f['phase']})"])
    for rid, target, why in unresolved:
        ws.append(["해석 불가 참조", f"{rid} → {target}", why])
    if not orphan_feats and not unresolved:
        ws.append(["-", "-", "전 기능이 요구사항에 연결됨 (100%)"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP
            if row[0].value == "해석 불가 참조":
                cell.fill = WARN_FILL
    for i, w in enumerate([18, 22, 70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}")
    print(f"  추적 {len(rtm_rows)}행 · 기능 커버 {len(covered)}/{len(feat)} · 미연결 {len(orphan_feats)} · 해석불가 {len(unresolved)}")
    if orphan_feats:
        print("  미연결 기능:", ", ".join(orphan_feats[:20]), "..." if len(orphan_feats) > 20 else "")


if __name__ == "__main__":
    main()
