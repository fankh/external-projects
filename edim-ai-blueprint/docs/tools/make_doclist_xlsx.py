# -*- coding: utf-8 -*-
"""EDIM 산출물목록.xlsx 생성 — SI 표준 산출물 체계 대비 현황·계획.

실행: py docs/tools/make_doclist_xlsx.py  (저장: docs/EDIM_산출물목록.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "EDIM_산출물목록.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
STATUS_FILL = {
    "완료": "C8E6C9", "진행": "FFF3C4", "템플릿": "E3ECF7", "예정": "F0F1F3",
}
PHASE_FILL = {"착수": "FFF8E7", "분석": "E8F4FD", "설계": "F0FFF0", "구현·시험": "FFF0F5",
              "이행·운영": "F5F0FF", "영업·계약": "FFEFC2"}

# 작업자 배정: 산출물명 → 실명 (미기재 = 미정)
ASSIGNEES = {
    # 예: "요구사항정의서": "홍길동",
}
# (단계, 산출물, 형식, 상태, 버전, 파일/위치, 비고·선행조건)
DOCS = [
 # ---- 영업·계약 ----
 ("영업·계약", "제안서", "HTML→PDF", "예정", "-", "_template/01_제안서 (필요 시 복사)", "사업 조건 확정 후 작성 — 중복 템플릿은 정리(2026-07-07)"),
 ("영업·계약", "견적·계약 문서", "-", "예정", "-", "-", "제안 확정 후"),

 # ---- 착수 ----
 ("착수", "사업수행계획서", "MD", "완료", "v0.1", "docs/EDIM_사업수행계획서.md", "범위·조직·일정·관리 체계 — 계약 확정(시작일·인력·협의 4건) 시 v1.0 승격"),
 ("착수", "WBS·일정표", "Excel", "완료", "v0.1", "docs/04_WBS/EDIM_WBS.xlsx", "38 Task·44주 간트·M1~M4 — 시작일 가정, 계약 시 재산정"),
 ("착수", "위험관리대장", "Excel", "완료", "v0.1", "docs/EDIM_위험관리대장.xlsx", "초기 위험 14건(협의 4건·미결정·이행·일정) — make_risk_xlsx.py 재생성, 주간 갱신"),
 ("착수", "보안관리계획서", "MD", "완료", "v0.1", "docs/EDIM_보안관리계획서.md", "REQ-N-006~010·022 기준선 + Grade·개인정보·서버 보안 — 보안 솔루션 범위 협의 후 v0.2"),

 # ---- 분석 ----
 ("분석", "시스템 개요서", "MD", "완료", "v0.1", "docs/EDIM_개요.md", "발표자료 78슬라이드 분석 — 기준 문서"),
 ("분석", "요구사항정의서", "Excel", "완료", "v0.2", "docs/02_요구사항/EDIM_요구사항정의서.xlsx", "기능 50·비기능 22·인터페이스 8"),
 ("분석", "요구사항추적표 (RTM)", "Excel(자동)", "완료", "v0.1", "docs/EDIM_요구사항추적표.xlsx", "REQ→기능→메뉴→화면→컴포넌트→DB, 커버리지 179/179"),
 ("분석", "기능정의서", "Excel", "완료", "v0.2", "docs/EDIM_기능정의서.xlsx", "14모듈 179기능(SYS-021 i18n) — v0.3"),
 ("분석", "메뉴정의서", "Excel", "완료", "v0.3", "docs/EDIM_메뉴정의서.xlsx", "98메뉴 — PPT 3차 전수 대조+다국어 관리"),
 ("분석", "요구사항 보완노트", "MD", "완료", "-", "docs/EDIM_요구사항_보완노트.md", "재검토 근거·협의 필요 4건"),
 ("분석", "현행(AS-IS) 분석서", "문서", "예정", "-", "-", "고객사 확정 후 — 기존 업무·시스템·데이터 조사"),

 # ---- 설계 ----
 ("설계", "컴포넌트(아키텍처) 정의서", "MD+Excel", "완료", "v0.2", "docs/EDIM_컴포넌트_정의서.md / EDIM_컴포넌트정의서.xlsx", "39컴포넌트·API 108·구축 상태 반영"),
 ("설계", "DB 정의서", "MD+Excel", "완료", "v0.2", "docs/EDIM_DB_정의서.md / EDIM_DB정의서.xlsx", "54테이블 462컬럼 — v0.5 i18n(sys_translation, 미결정 #7 해소)"),
 ("설계", "화면설계서 (와이어프레임)", "HTML", "완료", "v0.3", "docs/EDIM_화면설계서.html · https://edim.seekerslab.com/design/", "24화면 — 메뉴 미배정 주요 화면 보완(v0.3)"),
 ("설계", "DB DDL·검증 스크립트", "SQL", "완료", "v0.4.1", "docs/ddl/edim_schema.sql · verify_runtime.sql", "실 PG16 적용·BOM 재귀/제약 검증 통과 — 마이그레이션 도구 도입 시 초기본"),
 ("설계", "디자인 시안 A (Modern)", "HTML", "완료", "v0.1", "docs/EDIM_디자인시안.html · /design/hifi/", "브랜드 접점용(로그인·Dashboard·모바일) — B-06 권고 참조"),
 ("설계", "디자인 시안 B (Dense) ★권고", "HTML", "완료", "v0.1", "docs/EDIM_디자인시안_B_dense.html · /design/dense/", "레거시 패턴 조사(SAP·더존·CAD·Excel)+22px 컨트롤+MDI·F-key — 업무 화면 표준"),
 ("설계", "클래스 정의서", "MD", "완료", "v0.1", "docs/EDIM_클래스정의서.md", "언어 중립 — 11도메인·Aggregate 22·인바리언트↔DB제약 대응표"),
 ("설계", "인터페이스 정의서", "MD+YAML", "완료", "v0.1", "docs/EDIM_인터페이스정의서.md · docs/api/edim-openapi.yaml", "OpenAPI 3.1 자동 생성(107 op)+Accept-Language·외부 연계 8종"),
 ("설계", "권한·승인 정의서", "Excel", "완료", "v0.1", "docs/EDIM_권한승인정의서.xlsx", "매트릭스 97메뉴 자동 도출·승인 흐름 13·Platform 범위 8·Grade"),
 ("설계", "개발 표준 정의서", "MD", "완료", "v0.1", "docs/EDIM_개발표준정의서.md", "언어중립 핵심+FastAPI·React/TS 기준, 개정 트리거 5건 명시"),
 ("설계", "공통코드 정의서", "Excel", "진행", "-", "DB정의서 §10 포함", "분리 여부 협의 — 코드 16그룹"),
 ("설계", "배치(Job) 정의서", "MD", "완료", "v0.2", "docs/EDIM_배치보고서양식정의서.md (A부)", "공통 규약 + 업무 잡 JOB-01~08(Run·AI·재고단가·MRP·정산) + 운영 배치 BAT-01~06 — 통합본"),
 ("설계", "보고서·양식 정의서", "MD", "완료", "v0.2", "docs/EDIM_배치보고서양식정의서.md (B부)", "렌더 통제(Print Set-up·Grade·워터마크)·RPT-01~07 구현·FORM-01~04(고객 양식 확정 대기) — 통합본"),

 # ---- 구현·시험 ----
 ("구현·시험", "단위테스트 계획·결과", "문서", "예정", "-", "-", "P1 개발과 병행"),
 ("구현·시험", "통합테스트 시나리오", "Excel", "예정", "-", "-", "E2E: 코드 등록→CPQ→Run→견적 (개요 §7 워크플로우 기반)"),
 ("구현·시험", "기능확인서 (FVT)", "Excel(자동)", "완료", "v0.1", "docs/03_기능확인서_FVT/EDIM_기능확인서.xlsx", "기능 179·비기능 22 자동 생성 — 판정 기입 후 버전 고정"),
 ("구현·시험", "성능시험 계획·결과", "문서", "예정", "-", "-", "REQ-N-001~005 목표 검증 (산출물 1시간·BOM 30초 등)"),
 ("구현·시험", "결함관리대장", "Excel", "예정", "-", "-", "시험 단계 운영"),

 # ---- 이행·운영 ----
 ("이행·운영", "데이터 이행 계획서", "MD", "완료", "v0.1", "docs/EDIM_데이터이행계획서.md", "9대상·5단계·검증 6기준 — 조사(2.3) 후 v0.2 갱신"),
 ("이행·운영", "설치·배포 매뉴얼", "MD", "진행", "-", "컴포넌트정의서 §8 구축 현황 + 서버 구축 이력", "Self-managed 패키지용 정식 매뉴얼 필요"),
 ("이행·운영", "운영자 매뉴얼", "MD", "완료", "v0.1", "docs/EDIM_관리자가이드.md", "계정·권한/문서 통제/승인 운영/서버 운영/장애 대응 — 개발 서버 기준, 운영 이관 시 확장"),
 ("이행·운영", "사용자 매뉴얼", "MD", "완료", "v0.1", "docs/EDIM_사용자가이드.md", "공통 조작(F-Key·탭·검색)+모듈별(CPQ/PLM/Code/ERP/Toolbox/공통)+FAQ — GENERAL·SETUP 대상"),
 ("이행·운영", "교육 계획·자료", "문서", "예정", "-", "-", "Set-up 사용자(코드·Macro 작성) 교육이 핵심"),
 ("이행·운영", "검수·인수 문서", "문서", "예정", "-", "-", "FVT 승인 연계"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 산출물목록"
    ws["B2"].font = TITLE_FONT
    counts = {}
    for row in DOCS:
        counts[row[3]] = counts.get(row[3], 0) + 1
    info = [
        ("문서 버전", "v0.2"),
        ("작성일", "2026-07-07 (v0.2 갱신 2026-07-11 — 착수 문서 3종 완료 반영)"),
        ("목적", "SI 표준 산출물 체계 대비 작성 현황·계획 관리 (문서 레지스터)"),
        ("총 산출물", f"{len(DOCS)}종 — " + " · ".join(f"{k} {v}" for k, v in sorted(counts.items()))),
        ("상태 정의", "완료 = 작성됨(버전 관리) / 진행 = 부분 존재 / 템플릿 = 양식만 존재 / 예정 = 미착수"),
        ("우선 권고", "① 설치·배포 매뉴얼(Self-managed 정식본) ② 현행분석·제안서(고객사·사업 조건 확정 후) ③ 구현·시험 문서(테스트 계획·결함대장) — 착수 3종·배치/보고서양식·매뉴얼 2종 완료"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 100

    ws = wb.create_sheet("산출물목록")
    headers = ["No", "단계", "산출물", "형식", "상태", "버전", "작업자", "파일/위치", "비고·선행조건"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for n, row in enumerate(DOCS, 1):
        phase, name, fmt, status, ver, loc, note = row
        ws.append([n, phase, name, fmt, status, ver, ASSIGNEES.get(name, "미정"), loc, note])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 4, 5, 6, 7) else WRAP
        ws.cell(row=ws.max_row, column=2).fill = PatternFill("solid", fgColor=PHASE_FILL.get(phase, "FFFFFF"))
        ws.cell(row=ws.max_row, column=5).fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "FFFFFF"))
    for i, w in enumerate([5, 11, 26, 11, 8, 7, 9, 50, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (산출물 {len(DOCS)}종: {counts})")


if __name__ == "__main__":
    main()
