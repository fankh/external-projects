# -*- coding: utf-8 -*-
"""EDIM 기능정의서.xlsx 생성 스크립트.

기준: docs/EDIM_개요.md §13 기능 코드, EDIM_컴포넌트_정의서.md, EDIM_DB_정의서.md
실행: py docs/tools/make_feature_xlsx.py  (저장 위치: docs/EDIM_기능정의서.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "EDIM_기능정의서.xlsx")

# ---------------------------------------------------------------- 스타일
HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
SUB_FONT = Font(name="맑은 고딕", size=10, color="555555")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
MODULE_FILLS = {}  # 모듈별 교차 색상
PALETTE = ["FFF8E7", "E8F4FD", "F0FFF0", "FFF0F5", "F5F0FF", "FFEFC2",
           "E0F7FA", "FDF2E9", "EAF2F8", "F9EBEA", "F4ECF7", "E8F8F5"]

# ---------------------------------------------------------------- 데이터
# (기능ID, 기능코드, 기능명, 설명, 주요 처리 규칙, 권한, 컴포넌트, DB, Phase, 비고)
MODULES = []  # (모듈키, 모듈명, [rows])

MODULES.append(("SYS", "시스템 공통", [
 ("SYS-001", "-", "로그인", "ID/PW 인증 및 토큰 발급", "실패 횟수 제한, 비밀번호 정책", "전체", "SVC-01", "sys_user", "P1", ""),
 ("SYS-002", "-", "토큰 갱신·세션 관리", "Refresh 토큰 갱신, 세션 만료 처리", "테넌트 claim 포함 JWT", "전체", "GW-01, SVC-01", "sys_user", "P1", ""),
 ("SYS-003", "-", "사용자 관리", "사용자 등록/수정/상태 변경(ACTIVE/LOCKED/RETIRED)", "user_level 4단계(PLATFORM/ADMIN/SETUP/GENERAL)", "ADMIN", "SVC-01, FE-04", "sys_user", "P1", ""),
 ("SYS-004", "-", "역할 관리", "역할 생성 및 사용자-역할 매핑", "", "ADMIN", "SVC-01, FE-04", "sys_role, sys_user_role", "P1", ""),
 ("SYS-005", "-", "권한 부여", "Head Tab/Hierarchy/기능/Table 단위 권한 설정", "VIEW/EDIT/APPROVE/SETUP 4개 액션, 권한 없는 항목 미표시", "ADMIN", "SVC-01, FE-04", "sys_role_permission", "P1", "개요 §4"),
 ("SYS-006", "E-3", "Hierarchy Tree 조회·검색", "제품/일반DB/설정 Tree 표시, 이름·비고·심볼 검색", "권한 필터 적용", "전체", "SVC-02, FE-01", "sys_hierarchy", "P1", ""),
 ("SYS-007", "E-3", "Hierarchy 노드 편집", "노드 생성/편집/삭제 (승인 후 반영)", "is_system Tree는 편집 불가, DB 상호관계 점검 후 저장", "SETUP↑", "SVC-02, SVC-10", "sys_hierarchy", "P1", ""),
 ("SYS-008", "E-3", "Hierarchy 이동·Address 추적", "노드 이동/개명 시 데이터 연결 유지", "이전 address를 이력 보관, 참조 자동 추적", "SETUP↑", "SVC-02", "sys_hierarchy, sys_history", "P1", "개요 §4"),
 ("SYS-009", "-", "승인 요청", "Set-up 자산 변경의 승인 요청 생성", "DRAFT→PENDING 전이", "SETUP↑", "SVC-10", "sys_approval_request", "P1", ""),
 ("SYS-010", "-", "승인함 처리", "요청 목록 검토, 승인/반려 결정", "승인 시 대상 approval_status=APPROVED 트랜잭션 전이", "ADMIN", "SVC-10, FE-01/04", "sys_approval_request", "P1", ""),
 ("SYS-011", "-", "승인 이력 조회", "대상별 승인 이력 타임라인", "", "전체", "SVC-10", "sys_approval_request", "P2", ""),
 ("SYS-012", "-", "변경 이력 조회", "테이블/대상별 Audit 로그 조회", "before/after JSON 비교 표시", "ADMIN", "SVC-02, FE-04", "sys_history", "P2", ""),
 ("SYS-013", "-", "알림", "승인·To-do·일정·이상경고 알림 목록/읽음", "", "전체", "SVC-13", "sys_notification", "P2", ""),
 ("SYS-014", "-", "실시간 알림", "웹 SSE/WebSocket, 모바일 푸시", "", "전체", "SVC-13, FE-05", "sys_notification", "P4", ""),
 ("SYS-015", "H-1/E-1", "Main Work Frame", "Head Tab + 3판넬(좌 Tree/중앙 작업/우 Templet) 공통 프레임", "사용자 권한 항목만 표시, Breadcrumb 표시", "전체", "FE-01", "-", "P1", "개요 §4"),
 ("SYS-016", "E-1", "화면 개인화", "Frame·Canvas 색상, 언어, 로고, 판넬 크기(아코디언/펼침)", "", "전체", "FE-01", "-", "P4", "슬라이드 57"),
 ("SYS-017", "E-3", "일정·To-do 위젯", "Schedule management, To-do list, Done items, 승인요청 List", "ERP 이벤트와 연동", "전체", "FE-01, SVC-09", "erp_process_event", "P2", ""),
 ("SYS-018", "E-3", "업무 소통 SNS", "좌측 하단 판넬 — 사용자 일반 업무·일정·업무용 SNS(Project 중심 대화)", "모바일 APP-003과 동일 채널", "전체", "FE-01, SVC-13", "-", "P5", "슬라이드 57"),
 ("SYS-019", "H-1", "Head 메뉴 편집", "Head Item 추가/편집/삭제", "기업 사용자는 ERP 하부 Item만 편집 가능, 나머지는 Platform", "ADMIN", "FE-04, SVC-02", "sys_hierarchy", "P2", "슬라이드 57"),
 ("SYS-020", "H-1", "Head-Templet 연결", "각 Head Item과 Hierarchy·(좌/우/중앙) Templet 연결 설정", "화면 구성이 Head 선택에 따라 전환", "SETUP↑", "SVC-02, SVC-06", "tbx_templet", "P2", "슬라이드 57"),
]))

MODULES.append(("DOC", "문서 관리", [
 ("DOC-001", "-", "문서 이력·Released Status", "문서별 Progress(Person/date/Released Status/Approver/App.Date/Version/DOC No.) 관리", "개정 Version(KD-0.2 등) 자동 증가", "전체", "SVC-11, SVC-10", "doc_control", "P3", "슬라이드 58"),
 ("DOC-002", "-", "관리 등급(보안 등급)", "문서 Management Grade(S-2 등) 설정 — 등급별 열람/출력 통제", "권한 없는 등급 문서 접근 차단", "ADMIN", "SVC-01, SVC-11", "doc_control, sys_role_permission", "P3", "슬라이드 58"),
 ("DOC-003", "-", "Document Code 채번", "문서 코드 자동 생성 (예: EU-3-2020-450-6-21-4-SR-7)", "Product Code 기반 규칙 채번", "시스템", "SVC-11", "doc_control", "P3", "슬라이드 4/8"),
 ("DOC-004", "-", "문서 보안 솔루션", "Authorization Settings·Security Solution 연계(워터마크·DRM·출력 통제)", "상세 요건 미정 — 고객 협의", "ADMIN", "SVC-11", "-", "P5", "슬라이드 58 placeholder"),
]))

MODULES.append(("CODE", "RCCS 코드 관리", [
 ("CODE-001", "S-1-1", "코드 그룹 등록", "Specification 그룹(KOF 등) 생성, Hierarchy 위치 지정", "group_code 중복 불가, 승인 후 사용", "SETUP", "SVC-03, FE-02", "code_group", "P1", ""),
 ("CODE-002", "S-1-1", "코드 자릿수 정의", "그룹 내 Item Slot(A:Fan Model, B:Size…) 정의", "Slot 문자 그룹 내 유일", "SETUP", "SVC-03", "code_item", "P1", ""),
 ("CODE-003", "S-1-1", "자릿수 값 등록", "Sub Item List(350/400…, Gal./CU…) 등록", "중복검토 실행 후 승인 요청", "SETUP", "SVC-03", "code_item_value", "P1", ""),
 ("CODE-004", "S-1-2", "Table 참조 값 연결", "값이 Table 참조인 경우 tbl_data_table 연결", "", "SETUP", "SVC-03, SVC-05", "code_item_value", "P1", ""),
 ("CODE-005", "S-1-2", "자재·일반구매품 코드 등록", "Raw Material/GPI 그룹의 Sub Code 등록 (Motor 등)", "가격(G:Price) Slot은 단가 Table 연계", "SETUP", "SVC-03", "code_group, code_item_value", "P1", ""),
 ("CODE-006", "S-1-3", "Product Code 등록", "Main Code(KDCR 3-13) 생성, 설명·Hierarchy 지정", "Main Code 중복성 검토 필수", "SETUP", "SVC-03", "product_code", "P1", ""),
 ("CODE-007", "S-1-3", "Product Code 자릿수 구성", "제품 코드가 사용할 Slot과 값 출처(Sub Code) 연결", "Code Hierarchy에서 Item List 탐색·설정 반복", "SETUP", "SVC-03", "product_code_item", "P1", ""),
 ("CODE-008", "S-1-4", "Code Relationship 설정", "Mother-Child 관계 + 수량 + Slot 매핑 조건 등록", "Slot 매핑 테이블(slot_map)로 Mother/Child 자릿수 연동", "SETUP", "SVC-03", "code_relationship, code_relationship_slot_map", "P1", "관계형 BOM 핵심"),
 ("CODE-009", "S-1-4", "Part List Running Test", "Mother 조합 선택 → 조건 일치 Child 전개 검증", "검증 통과 후 승인 요청 가능", "SETUP", "SVC-03", "code_relationship", "P1", ""),
 ("CODE-010", "D-2", "BOM 전개", "완성품 Code → Code Relationship 재귀 전개 → 전체 Part List", "재귀 CTE, 깊이 제한, 순환 차단", "GENERAL↑", "SVC-03, ENG-02", "code_relationship, cpq_selection_item", "P1", ""),
 ("CODE-011", "S-1-5", "Arrangement Code 등록", "표준 구성 조합(Double Deck 2 등) 등록", "Fan Direction(L0~R270)·설치방식 옵션", "SETUP", "SVC-03", "arrangement_code", "P2", ""),
 ("CODE-012", "S-1-6", "Arrangement 구성 설정", "구성 Module 배치·결합조건(Macro) 설정, DWG 연결", "결합 조건은 tbx_macro 참조", "SETUP", "SVC-03, SVC-04", "arrangement_component", "P2", ""),
 ("CODE-013", "S-1-x", "코드 승인", "코드 자산 승인 워크플로우 (Pending→Approved)", "미승인 코드는 CPQ에서 선택 불가", "ADMIN", "SVC-10", "code_*", "P1", ""),
 ("CODE-014", "S-1-x", "등록 Code Table 조회", "그룹별 등록 코드 목록·상세·승인상태 조회", "", "전체", "SVC-03, FE-02", "code_*", "P1", ""),
 ("CODE-015", "S-1-x", "코드 자산 연결", "Sub/Product Code에 DWG(2D/3D)·Data Up-Load·Table 연결 관리", "우측 판넬에서 코드별 도면/자료 즉시 확인", "SETUP", "SVC-03, SVC-04, SVC-05", "product_code, dwg_file, tbl_data_table", "P1", "슬라이드 33~38 우측판넬"),
 ("CODE-016", "S-1-x", "코드 Excel Import/Export", "Registered Code Table의 값 목록 Excel 일괄 등록/내보내기", "형식 검증 후 반영", "SETUP", "SVC-03, INT-03", "code_item_value", "P2", "슬라이드 32"),
]))

MODULES.append(("DWG", "도면·PLM", [
 ("DWG-001", "S-4-1-1", "도면 등록", "도면 기본정보(도번·유형·축척·크기) 등록", "도번 중복 불가, 표준도면은 프로젝트 미연결", "SETUP", "SVC-04", "dwg_drawing", "P2", ""),
 ("DWG-002", "S-4-1-1", "부품 호출·배치", "Hierarchy에서 하부 부품 도면 호출, Canvas Drag 배치", "Block 상태로 호출(기본 치수)", "SETUP", "FE-02, SVC-04", "dwg_document", "P2", "노트 워크플로우 1"),
 ("DWG-003", "E-2", "CAD 편집 명령", "복사/이동/반전/연장/삭제/회전, 좌표 이동, 색상·선형식", "CAD와 동일한 명령 Toolbar", "SETUP", "FE-02", "dwg_document", "P2", ""),
 ("DWG-004", "E-2", "선택·스냅", "1차 블록 선택, 2차 세부(끝점·중앙·모서리·중심) 선택", "정보 선택 시 관계·검증·조립순서 색상 구분", "SETUP", "FE-02", "-", "P2", ""),
 ("DWG-005", "S-4-1-1", "치수선·지시선", "치수선 삽입, 지시선 자동 번호(A~N) 부여", "번호 임의 변경 시 중복 검사, 치수선 감추기", "SETUP", "FE-02, SVC-04", "dwg_dimension", "P2", ""),
 ("DWG-006", "S-4-1-1", "치수 Macro 바인딩", "치수마다 Macro 연결 또는 Variant 직접 입력", "KEY/DETAIL 치수 구분(승인도면용)", "SETUP", "SVC-04, ENG-01", "dwg_dimension, tbx_macro", "P2", "파라메트릭 핵심"),
 ("DWG-007", "S-4-1-1", "치수·부품 동기화", "입력 치수값에 의한 Parametric Design 반영", "치수와 부품 일체화, 항목 선택 시 도면 활성화", "SETUP", "FE-02, ENG-03", "dwg_document", "P2", ""),
 ("DWG-008", "S-4-1-1", "설계·자료 우선순위", "Design/Data Priority 설정, 순환 참조 점검", "우선순위 불일치 시 경고 제공", "SETUP", "SVC-04, ENG-01", "dwg_dimension", "P2", ""),
 ("DWG-009", "E-4", "부품 상호관계 설정", "Block A/B 선택, 조건1(수직·수평·중심)·조건2(접촉·좌표·각도), 관계값 Macro", "관계 진행 우선순위로 데이터 혼선 차단", "SETUP", "SVC-04", "dwg_part_relation", "P2", ""),
 ("DWG-010", "E-4", "설계 검증 규칙", "설계 조건 Macro 설정, 불일치 시 경고 문구", "공학식·Table 활용", "SETUP", "SVC-04, ENG-01", "dwg_verification", "P2", ""),
 ("DWG-011", "E-4", "조립 순서 Set-up", "조립품 부품에 조립 순서 번호·주의사항 정리", "", "SETUP", "SVC-04", "dwg_bom", "P2", ""),
 ("DWG-012", "S-4-1-1", "Block 저장·호출", "Block 단위 저장(임시저장→승인→사용), 상위 호출 시 Block 상태", "저장 장소 확인 절차", "SETUP", "SVC-04", "dwg_document", "P2", ""),
 ("DWG-013", "-", "도면 개정 관리", "Rev A/B/C 개정, 사유·내용 기록", "", "SETUP", "SVC-04", "dwg_revision", "P2", ""),
 ("DWG-014", "-", "도면 승인", "작성→검토→승인→발행 단계 관리", "슬라이드 26-7 스키마 준용", "ADMIN", "SVC-04, SVC-10", "dwg_approval", "P2", ""),
 ("DWG-015", "-", "도면 Import", "DXF/DWG/PDF/STEP 업로드 → DrawingDocument 변환", "DWG는 ODA 변환(INT-04), PDF→DXF 지원", "SETUP", "SVC-04, INT-04", "dwg_file, dwg_document", "P2", "프로토타입 승계"),
 ("DWG-016", "-", "도면 Export", "DXF(R2010)/PDF 출력", "", "전체", "SVC-04, ENG-03", "dwg_document", "P2", ""),
 ("DWG-017", "-", "부품 마스터", "부품 등록(규격·재질·단위·단중·공급처·표준품)", "RCCS Product Code 연결 가능", "SETUP", "SVC-04", "prt_part", "P2", ""),
 ("DWG-018", "-", "재질 마스터", "재질 코드(SS400 등)·밀도·규격 관리", "밀도는 원가 무게 계산 입력", "SETUP", "SVC-04", "mat_material", "P2", ""),
 ("DWG-019", "-", "도면 BOM 관리", "도면 부품 구성표(품번·수량·비고)", "", "SETUP", "SVC-04", "dwg_bom", "P2", ""),
 ("DWG-020", "-", "첨부파일 관리", "도면별 파일(DWG/PDF/STEP/PNG) 첨부", "실체는 Object Storage", "전체", "SVC-04, SVC-12", "dwg_file", "P2", ""),
 ("DWG-021", "S-4-1-2", "Work Process 공정 데이터", "Product Code별 자재/공정/조립 정보(작업장·인원·스킬·시간·창고·최소재고)", "제조비 계산 입력", "SETUP", "SVC-09", "erp_work_process", "P3", "슬라이드 45"),
 ("DWG-022", "S-4-1-1", "도면 Templet 관리", "설정 도면 Templet(도면 Frame·비율·형식) 등록·호출", "신규 도면은 Templet에서 시작 (Free CAD)", "SETUP", "SVC-04, SVC-06", "tbx_templet", "P2", "슬라이드 41"),
 ("DWG-023", "-", "CAD Mapping", "도면 항목과 외부 CAD 파일(2D/3D) 매핑 연결", "Sub Material/Variant List에 CAD Mapping 체크", "SETUP", "SVC-04, INT-04", "dwg_file", "P3", "슬라이드 44"),
 ("DWG-024", "E-4", "설계 Simulation", "설정된 관계·검증 규칙 기반 도면 동작 시뮬레이션(우측 판넬)", "검증 경고 시각화", "SETUP", "FE-02, ENG-01", "-", "P4", "슬라이드 63"),
 ("DWG-025", "E-2", "측정·특성 도구", "치수 측정(수평/수직/정렬/각도/호길이/면적/무게), 특성 변경(색상/투명도/선굵기/레이어), 자르기(Trim), 그룹화/해제, 등각·단면 View", "CAD와 동일 형식 명령·단축키(DI/CH/CO/RO/E/TR/REG)", "SETUP", "FE-02", "-", "P2", "슬라이드 61"),
 ("DWG-026", "E-2", "Referencers 역참조 조회", "도면/부품의 Where-used — 어느 상위 도면·Arrangement에서 참조되는지 조회", "Viewer 탭: Variants/Viewer/Referencers/Supersedure/Attachment", "전체", "SVC-04", "dwg_bom, code_relationship", "P3", "슬라이드 5/60 Viewer 탭"),
 ("DWG-027", "E-2", "Variants·Supersedure 관리", "도면 변형(Variants) 목록과 대체 관계(Supersedure — 신규정이 구버전 대체) 관리", "대체된 도면 사용 시 경고", "SETUP", "SVC-04", "dwg_supersedure, dwg_revision", "P3", "슬라이드 5/60 Viewer 탭"),
]))

MODULES.append(("TBL", "데이터 Table", [
 ("TBL-001", "E-4", "Table 정의", "Variant/Tech/Material Table 생성(열 정의·부서·Hierarchy 주소)", "DB로서의 명확한 구성 요구, 승인 후 사용", "SETUP", "SVC-05", "tbl_data_table", "P1", ""),
 ("TBL-002", "E-4", "Table 행 편집", "행 추가/편집/일괄 입력 (Key + A~N 값)", "row_key 유일", "SETUP", "SVC-05", "tbl_data_row", "P1", ""),
 ("TBL-003", "-", "Excel Import", "정해진 양식의 Excel에서 Table 행 일괄 등록", "Excel 연동 형식 정리(슬라이드 69)", "SETUP", "SVC-05, INT-03", "tbl_data_row", "P1", ""),
 ("TBL-004", "-", "Table 범위 조회", "Macro 참조용 범위 조회 API(TableN(col,range))", "GIN 인덱스, 캐시 적용", "시스템", "SVC-05, ENG-01", "tbl_data_row", "P1", ""),
 ("TBL-005", "E-4", "Data Up-Load", "Data/File/Image 자료 등록(부서·유형·이름·설명)", "중복검토", "SETUP", "SVC-05, SVC-12", "tbl_data_table, dwg_file", "P2", ""),
 ("TBL-006", "-", "Table 승인", "Table 정의·대량 변경 승인", "", "ADMIN", "SVC-10", "tbl_data_table", "P1", ""),
]))

MODULES.append(("TBX", "Toolbox", [
 ("TBX-001", "S-2-1", "UI Designer 위젯 배치", "Widget Palette(Button/Label/Entry/Text/Canvas/Frame/Menu/Combo/Table) Drag&Drop", "Qt Designer 방식 폼 빌더", "SETUP", "FE-03, SVC-06", "tbx_ui_form", "P4", ""),
 ("TBX-002", "S-2-1", "동작 Templet 바인딩", "위젯별 동작(저장/삭제/복사/등록/찾기)·대상 Data 연결", "Set-up 전용 창(Templet)에서 설정", "SETUP", "FE-03, SVC-06", "tbx_ui_form, tbx_templet", "P4", ""),
 ("TBX-003", "S-2-1", "Form 저장·버전·게시", "layout_def 저장, 버전 관리, 승인 후 게시", "게시된 Form은 FE-01 동적 렌더러가 실행", "SETUP", "SVC-06, SVC-10", "tbx_ui_form", "P4", ""),
 ("TBX-004", "S-2-1", "AI UI 제안", "Application 설명 입력 → 용도/항목/필요 DB Table 정리된 UI 초안", "제안 후 사용자 Customizing", "SETUP", "AI-04, FE-03", "tbx_ui_form", "P4", ""),
 ("TBX-005", "S-2-2", "Prompt 입력", "자연어로 계산 요구 기술", "예: SS Fan 샤프트 길이 계산", "SETUP", "FE-03, AI-03", "tbx_macro", "P4", ""),
 ("TBX-006", "S-2-2", "Macro 수식 편집", "Excel 문법 계산식 작성(TableN/Var/PreC 참조)", "문법 특허 검토 결과 반영", "SETUP", "FE-03, ENG-01", "tbx_macro", "P2", "엔진은 P1"),
 ("TBX-007", "S-2-2", "Flowchart 편집", "흐름도 기반 로직 편집", "", "SETUP", "FE-03", "tbx_macro", "P4", ""),
 ("TBX-008", "S-2-2", "4-Way Sync 변환", "Prompt↔Macro↔Flowchart↔Coding 실시간 상호 변환", "AI 변환, Description 동시 갱신", "SETUP", "AI-03, ENG-01", "tbx_macro", "P4", "핵심 차별점"),
 ("TBX-009", "S-2-2", "함수·그래프 마법사", "함수 검색/설명/삽입, 그래프 유형 선택 마법사", "공학 함수 Templet 포함", "SETUP", "FE-03", "-", "P4", ""),
 ("TBX-010", "S-2-2", "Data Information Call", "Enterprise DB(Table/Chart/Formula) 탐색·주소 지정", "Hierarchy 주소 기반", "SETUP", "FE-03, SVC-05", "tbl_data_table", "P4", ""),
 ("TBX-011", "S-2-2", "Macro Test Run", "조건값 입력 → 계산 실행 → 결과 확인", "결과·입력 저장(test_input/result)", "SETUP", "ENG-01", "tbx_macro", "P2", ""),
 ("TBX-012", "S-2-2", "Macro 승인·버전", "생성/편집/삭제의 엄격한 검증·승인 절차, 버전 관리", "미승인 Macro는 Run에서 사용 불가", "ADMIN", "SVC-10", "tbx_macro", "P2", ""),
 ("TBX-013", "S-2-1", "Templet 관리", "Command button/Combo/Data Table/Print/Graph Templet 등록·재사용", "is_system Templet은 EDIM 제공", "SETUP", "SVC-06", "tbx_templet", "P4", ""),
 ("TBX-014", "S-2-2", "기능 찾기·도움말", "자연어로 필요 기능/함수 검색(기능 찾기), User Help·Manual 연동", "AI 기반 함수 추천", "전체", "AI-03, FE-03", "-", "P4", "슬라이드 29"),
 ("TBX-015", "S-2-2", "인터넷 검색·AI 질의", "Toolbox 내 인터넷 검색 기능, AI 질의응답(내부 자료 검색·응답)", "발표자 노트 명시 요건", "전체", "AI-05, FE-03", "-", "P5", "슬라이드 27 노트"),
]))

MODULES.append(("CPQ", "CPQ 제품선정", [
 ("CPQ-001", "C-1", "제품 선정 화면", "Product Tree에서 제품 선택, Selection 시작", "승인된 코드만 표시", "GENERAL", "FE-01, SVC-07", "cpq_selection", "P2", ""),
 ("CPQ-002", "C-1", "Arrangement 구성", "모듈 이미지 Drag 조합, 이동/삭제/분할/통합, 치수 DB 표시", "화면·도면 비율 자동 조절, 기준 치수 입력 시 하부 반영", "GENERAL", "FE-01, SVC-07", "cpq_selection, arrangement_*", "P2", "노트 워크플로우 2"),
 ("CPQ-003", "C-1", "사양 입력", "Spec List 입력 Table(풍량·정압·온도 등)", "Slot 값 선택 → 제품 Code 부분 결정", "GENERAL", "FE-01, SVC-07", "cpq_selection", "P2", ""),
 ("CPQ-004", "C-1", "고객 사양 Excel Import", "정해진 Excel 양식 Import → Selection 입력 항 자동 적용", "", "GENERAL", "INT-03, SVC-07", "cpq_selection", "P3", ""),
 ("CPQ-005", "C-2", "기술 데이터 화면", "Fan 선정 상세(Impeller/Motor/Casing 옵션), 성능 곡선, 기술 Table", "선정 분류·SL No 관리", "GENERAL", "FE-01, SVC-07", "cpq_selection, tbl_*", "P3", ""),
 ("CPQ-006", "C-3", "Document Template", "입력값(온도·습도)→Macro 계산→출력값(밀도), 그래프 표시", "그래프 전용 Data Table Templet", "GENERAL", "FE-01, ENG-01", "tbx_ui_form", "P3", "노트 워크플로우 3"),
 ("CPQ-007", "C-1", "완성품 Code 생성", "Slot 선택 조합 → Finished Goods Code 확정", "미완성 Slot 존재 시 확정 불가", "GENERAL", "SVC-07, SVC-03", "cpq_selection", "P2", ""),
 ("CPQ-008", "C-1", "Sub Item List 조회", "완성품 Code 연결 하위 Code 전개 목록, Click 시 세부 정보", "BOM Code Relationship의 Mother 연결 Sub 추출", "GENERAL", "SVC-07", "cpq_selection_item", "P2", ""),
 ("CPQ-009", "D-2", "X-Code 비규격 처리", "비표준 Option 선택 시 X-Code 표시→검토요청→R&D 신규설계→신규코드 등록→복귀", "X-Code 상태 추적", "GENERAL/R&D", "SVC-07", "cpq_selection", "P3", ""),
 ("CPQ-010", "C-1", "DWG View", "선정 제품 도면 확인, 3각법 6면·전체 View", "SVG 렌더, 팬줌·레이어", "GENERAL", "FE-01", "dwg_document", "P2", ""),
 ("CPQ-011", "S-3-1", "Selection Set-up", "선정 UI 구성 — Product Tree 가져오기, Item Image·Arrangement 정의", "Call Form으로 구성", "SETUP", "FE-02, SVC-06", "tbx_ui_form", "P3", ""),
 ("CPQ-012", "S-3-2/3", "Technical·Document Set-up", "기술 Data/서류 Form 구성(입출력 항목·그래프 Templet)", "", "SETUP", "FE-02, SVC-06", "tbx_ui_form", "P3", ""),
 ("CPQ-013", "S-3-4", "Print Set-up", "양식 배치·Data 위치·그래프·워터마크·Font·용지·머리글/바닥글·내보내기(PDF/Office)", "Print Test 기능", "SETUP", "FE-02, SVC-11", "tbx_ui_form", "P3", ""),
 ("CPQ-014", "E-2", "Selection Toolbar 구성", "사용자 명령 버튼 Templet 제작 — Arrangement/Move/Delete/Add/Copy/DWG View, 사양 입력 Combo, 정형화 사양 호출(Excel), 도구 이미지 설정", "EDIM Toolbar로 제작, 다양한 Macro 준비 필요", "SETUP", "FE-02, SVC-06", "tbx_templet", "P3", "슬라이드 60"),
 ("CPQ-015", "E-2", "Module 이미지 제작", "제품 구성용 Module 그림 제작 도구(그림 제작 Modul) — 저장 이미지 호출·제어", "등록 Arrangement의 구성 제품 호출과 연동", "SETUP", "FE-02, SVC-04", "dwg_document", "P3", "슬라이드 60"),
]))

MODULES.append(("RUN", "EDIM Run", [
 ("RUN-001", "D-2", "Run 실행 요청", "Selection에서 BOM/DWG/PRICING/TECH/ALL 실행", "비동기 잡 큐, 부분 실행 지원", "GENERAL", "SVC-07, ENG-02", "cpq_run", "P2", ""),
 ("RUN-002", "D-2", "BOM·Part List 생성", "Main Code 시작 Code Relationship 전개 → BOM 저장", "Project BOM folder 저장", "시스템", "ENG-02, SVC-03", "cpq_selection_item, cpq_output", "P2", ""),
 ("RUN-003", "D-2", "치수 Macro 일괄 실행", "Sub-code 조건별 Macro 계산 → 계산값 Table 생성·저장", "우선순위 위상 정렬, 실패 목록 기록", "시스템", "ENG-02, ENG-01", "cpq_run.dimension_values", "P2", "노트 Runtime"),
 ("RUN-004", "D-2", "승인도면 생성", "CPQ Main Concept Drawing → 승인도면(CAD, 필요시 PDF)", "Project DWG folder 저장", "시스템", "ENG-03", "cpq_output, dwg_file", "P2", ""),
 ("RUN-005", "D-2", "제작도면 생성", "각 Code 연결 도면의 Macro Run → 치수 반영 도면 추출", "BOM 관련 내용만 Run 가능", "시스템", "ENG-03", "cpq_output", "P2", ""),
 ("RUN-006", "D-2", "기술자료 계산·생성", "Technical Data Calculation → Project Data folder 저장", "", "시스템", "ENG-02, ENG-01", "cpq_output", "P3", ""),
 ("RUN-007", "D-2", "일반 서류 생성", "General Document — 각 Code 연결 서류 계산·생성", "", "시스템", "ENG-02, SVC-11", "cpq_output", "P3", ""),
 ("RUN-008", "D-2", "Run 상태·오류 조회", "진행률·단계별 상태·Macro 오류/검증 경고 목록", "재시도 지원", "GENERAL", "SVC-07", "cpq_run", "P2", ""),
 ("RUN-009", "D-2", "산출물 조회", "Project Folder(DWG/Price/Data/BOM)별 산출물 목록·다운로드", "", "GENERAL", "SVC-12", "cpq_output, dwg_file", "P2", ""),
 ("RUN-010", "E-4", "BOM Export·부분 Run", "BOM 화면의 EBOM Run(BOM만)/EDIM Run(전체)/Export(Excel) 구분 실행", "BOM 관련 내용만 Run하는 부분 실행", "GENERAL", "ENG-02, INT-03", "cpq_run", "P2", "슬라이드 70"),
]))

MODULES.append(("CST", "원가·견적", [
 ("CST-001", "D-3", "단가 관리", "견적/구매이력/재고단가/견적적용 4종 단가 Table", "Code·부품 기준, 적용기간 관리", "SETUP", "SVC-08", "cst_price", "P3", ""),
 ("CST-002", "D-4", "자재비 계산", "원자재 정보(면적·무게·부피)×단가 → Material Cost", "밀도(mat_material) 활용", "시스템", "SVC-08, ENG-02", "cst_calc", "P3", ""),
 ("CST-003", "D-4", "제조비 계산", "Work Process 제조 정보(시간·임율·장비) → Manufacturing Cost", "", "시스템", "SVC-08", "cst_calc, erp_work_process", "P3", ""),
 ("CST-004", "D-3", "직접비 집계", "자재비+제조비+기타 직접비 → Direct Cost", "Project Price folder 저장", "시스템", "SVC-08", "cst_calc", "P3", ""),
 ("CST-005", "D-3", "PCR 작성", "Business Type별 원가 전개(조달·외주·판관비)→ EBIT", "Table/Direct 구분 산식", "GENERAL", "SVC-08, FE-01", "cst_pcr", "P3", ""),
 ("CST-006", "D-3", "견적서 생성·출력", "PCR 기반 견적서(수량·단가·조건) 생성, PDF 출력", "VAT·유효기간·납품/지급조건", "GENERAL", "SVC-08, SVC-11", "cst_quotation", "P3", ""),
 ("CST-007", "-", "견적 이력 관리", "견적 상태(DRAFT/SENT/ORDERED) 및 이력", "", "GENERAL", "SVC-08", "cst_quotation", "P3", ""),
]))

MODULES.append(("ERP", "ERP 업무", [
 ("ERP-001", "S-3-5", "프로젝트 등록", "Project No 자동, 영업단계(기술제안~종료), Client·담당자·Pain Point·접수자료", "PS 프로세스 시작점", "GENERAL", "SVC-09, FE-01", "prj_project", "P2", ""),
 ("ERP-002", "§11", "프로세스 이벤트 관리", "프로세스 코드별 이벤트 생성·상태 전이(TODO→진행→완료)", "선행 완료 조건 검사", "GENERAL", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-003", "§11", "부서별 업무함", "부서·담당자별 To-do/진행/완료 목록", "이상 경고(기한 초과) 표시", "GENERAL", "SVC-09, FE-01", "erp_process_event", "P5", ""),
 ("ERP-004", "§11", "영업 프로세스", "견적검토(PCR)→견적(QCR)→수주(OR) 흐름 처리", "CPQ·견적 연동", "영업", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-005", "§11", "승인도서·고객승인", "AP 생성→APP 고객 승인 처리", "산출물(승인도면) 연결", "영업/기술", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-006", "§11", "제작 프로세스", "제작의뢰(MR)→제작점검(MRR)→Part List(PL)→BOM", "EDIM Run 산출물 연동", "기술", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-007", "§11", "자재 프로세스", "발주요청(PR)→발주(PO)→입고(MI)→출고(MO)→수령(RM)→검사(ER)", "재고·창고 관리 연계", "자재", "SVC-09", "erp_process_event", "P5", "상세 v0.2"),
 ("ERP-008", "§11", "생산 프로세스", "생산계획(MP)→작업지시(WR)→입고(SFI)→검사(EF)→완료(FF), Stock 생산(SPP)", "부적합(NCF) 분기", "생산", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-009", "§11", "출고·납품 프로세스", "출고요청(FDR)→고객출고(SFO)→납품(DF)→기성청구(IR)", "", "영업/자재", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-010", "§11", "시운전·CS", "시운전 요청(CR)→시운전(CF), 인증서 발급", "", "CS", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-011", "§11", "재무 프로세스", "세금계산 매입/매출(II/IO), 출금/입금(PA/PI), 정산(RR/RA), 재무관리(FM)", "", "재무", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-012", "§11", "부적합 처리", "NCF 요청·처리, 재작업 요청(WR)", "", "QC", "SVC-09", "erp_process_event", "P5", ""),
 ("ERP-013", "§11", "평가", "업체 평가(CE), Project 평가(PE)", "", "ADMIN", "SVC-09", "com_company", "P5", ""),
 ("ERP-014", "§11", "Dashboard", "전후 공정·Project 일정·Event 상황·이상 경고(시간/자금)·손익·자금 흐름", "프로젝트별 집계", "ADMIN/경영", "SVC-09, FE-01", "erp_process_event, cst_*", "P5", ""),
 ("ERP-015", "-", "회사 마스터", "고객/공급/파트너/은행 등록, 담당자, 평가등급, 지불조건", "", "SETUP", "SVC-09", "com_company", "P2", ""),
 ("ERP-016", "S-3-5", "사용자 정의 ERP Toolbar", "ERP Toolbar 항목 사용자 편집, 권한별 표시", "User Customizing ERP", "SETUP", "SVC-06, FE-01", "tbx_ui_form", "P5", ""),
 ("ERP-017", "§11", "구매 조건 관리", "공급처(국가/본사)·납기·납품조건(EXW/FOB/CIP, 설치/상차/하차)·운송(트럭/배/항공)·지불조건·화폐·물품형식(물건/서비스/인력/협력사)·최소구매수량·인증서·단위(부피/체적/면적/수량)", "Right quality/quantity/time/price/source 원칙", "자재", "SVC-09", "com_company, erp_process_event", "P5", "슬라이드 46"),
 ("ERP-018", "§11", "제품코드 매핑", "사용자용 코드 ↔ 공급자용 코드 매핑 관리", "발주서에 공급자 코드 표기", "자재", "SVC-09, SVC-03", "prt_supplier_code_map", "P5", "슬라이드 46"),
 ("ERP-019", "§11", "재고 관리", "적정 재고량, 유통기한 관리", "최소 재고(min_stock) 경고", "자재", "SVC-09", "erp_work_process", "P5", "슬라이드 46"),
 ("ERP-020", "§11", "창고 관리", "자재 분류(원자재/가공/구매), 물성 위험등급(액체-독극물/가스-폭발성/고체-판·파이프·개), 저장위치 계층(지역/공장/창고/Storage/Sector), 보관품질(정기점검·유통기한), 입출고 통제·기록", "위험물 보관 규정 준수", "자재", "SVC-09", "erp_warehouse, mat_material", "P5", "슬라이드 46"),
 ("ERP-021", "§11", "재고 단가 관리", "재고 단가 4종 산출 — 최고/최저/평균/최근", "cst_price(STOCK) 원천", "자재", "SVC-08", "cst_price", "P5", "슬라이드 46"),
 ("ERP-022", "§11", "MRP 자재 소요 계획", "제품 단위당 BOM 기반 자재 소요량·소요 시기 산출", "생산계획 연동", "자재/생산", "SVC-09", "erp_process_event", "P5", "슬라이드 46"),
 ("ERP-023", "§11", "생산 Scheduling·Capacity", "주문량·가용 자원 기반 생산 일정, 생산능력(Capacity Planning) 분석", "", "생산", "SVC-09", "erp_process_event", "P5", "슬라이드 46"),
 ("ERP-024", "§11", "작업지시 관리", "작업 지시서(양식) 발행, 작업 단계별 진행 상황 기록·모니터링", "양식은 Print Form 활용", "생산", "SVC-09, SVC-11", "erp_process_event", "P5", "슬라이드 46"),
 ("ERP-025", "§11", "공정 관리", "공정별 데이터 실시간 수집·분석 — 작업장/기계종류/작업인원/필요스킬등급/작업시간/조립공정/전·후공정 DB", "erp_work_process 확장", "생산", "SVC-09", "erp_work_process", "P5", "슬라이드 46"),
 ("ERP-026", "§11", "자재흐름·물류 관리", "물류 창고, 물류 방식 관리", "외주·물류 관리와 연계", "자재", "SVC-09", "-", "P5", "슬라이드 46"),
 ("ERP-027", "§11", "품질 관리", "제품 검수, 하자 관리", "NCF 부적합 프로세스 연동", "QC", "SVC-09", "erp_process_event", "P5", "슬라이드 46"),
 ("ERP-028", "§11", "원가 DB 관리", "자재 단가 DB, 인건비 DB, 공정비용 DB 유지", "제조비 계산(CST-003) 원천", "재무/생산", "SVC-08", "cst_price", "P5", "슬라이드 46"),
 ("ERP-029", "§11", "불량·대체자재 관리", "불량품·폐품 처리, 자재 표준화·단순화, 대체 자재 연구·적용", "", "자재", "SVC-09", "-", "P5", "슬라이드 46"),
 ("ERP-030", "§11", "외주 제조 관리", "3rd party 제조업체 — 제조 의뢰, 재작업 요청, 인증서 발급", "Partner 실시간 Code 연결", "생산", "SVC-09, INT-01", "com_company", "P5", "슬라이드 10/77"),
]))

MODULES.append(("DUCT", "건축 설비 설계", [
 ("DUCT-001", "E-4-2", "건축도면 학습", "건축도(램프·빔·소방구역·실 구획) AI 판독 — 설치 불가 지역 구분", "설비(장비 도면·소화 배관·전기)·건축 기호 학습 포함", "PLATFORM", "AI-02", "(RAG 인덱스)", "P5", "슬라이드 68"),
 ("DUCT-002", "E-4-2", "건축 도면 호출", "층별·대공간 복수 도면 호출, 연결 Point(XYZ) 탐색", "", "SETUP", "SVC-04", "dwg_document", "P5", "슬라이드 68"),
 ("DUCT-003", "E-4-2", "Duct 자재 DB", "Duct 종류/손실/무게/Size별 최대길이, Pitting/Hanger/Joint/Insulation, 최대 Sizing·블록화 설계(수동변경·분할합체), 실 용도별 설계기준(풍속), 장비 기초 도면", "", "SETUP", "SVC-05", "tbl_data_table", "P5", "슬라이드 68"),
 ("DUCT-004", "E-4-2", "설비 설계 조건 설정", "풍량 계산(실별 용도·환기횟수/국소 급배기/임의입력), 장비 리스트 풍량, 건축 정보(층고·보·텍스높이·마감), 용도(급기/배기/순환/국부), 입출구 Point, Diffuser 기준, 출발-종착점(장비↔실 Combo 연결), 공기 조건(GasType·온도·습도·밀도), Point별 유속(Size)/Std 기준, Duct Option(점검구·Turning)", "설계 전용 Templet 사용", "GENERAL", "FE-01, ENG-01", "tbl_data_table", "P5", "슬라이드 68"),
 ("DUCT-005", "E-4-2", "Duct 자동 배치", "가장 빠른 길 찾기(유체 흐름 고려 설계), Diffuser 자동 배치(위치·수량 조정), 수동 조정(Drag·Click)", "경로 탐색 알고리즘", "GENERAL", "ENG-03", "dwg_document", "P5", "슬라이드 68"),
 ("DUCT-006", "E-4-2", "설비 기술자료 계산", "압력 손실/Leak율/온도 변화/결로/하중 계산, 장비 풍량 vs 설계 풍량 비교, 냉난방·배기 자료 기반 장비 선정", "공학식 Macro 활용", "GENERAL", "ENG-01", "tbx_macro", "P5", "슬라이드 68"),
 ("DUCT-007", "E-4-2", "설비-EDIM 연결", "설계 결과의 EDIM 연동 — 기술(BOM/Code), 제조(최소 스크랩), 자재(구매), 영업(견적)", "EDIM Run 파이프라인 재사용", "시스템", "ENG-02", "cpq_*", "P5", "슬라이드 68"),
]))

MODULES.append(("AI", "AI 기능", [
 ("AI-001", "§9", "도면 학습 파이프라인", "CAD 변환→추출→구조화→RAG 인덱싱 잡 관리", "Platform 권한 전용, 비용 계측", "PLATFORM", "AI-02, FE-04", "(RAG 인덱스)", "P5", ""),
 ("AI-002", "§9", "메타데이터 추출", "타이틀블록(도번·Rev·재질·축척)·레이어·블록 속성·BOM 규칙 분류", "난이도 下 — 우선 적용", "PLATFORM", "AI-02", "-", "P5", ""),
 ("AI-003", "§9", "2D 엔티티 파싱·OCR", "DWG/DXF 엔티티(치수·공차·기호) 추출, 스캔 도면 OCR+도면인식", "도면 표준 상이 시 튜닝", "PLATFORM", "AI-02", "-", "P5", ""),
 ("AI-004", "S-2-2", "Macro 자동 생성", "Prompt→Macro/Flowchart/Description/Coding 생성", "ENG-01 문법 검증 통과 필수", "SETUP", "AI-03", "tbx_macro", "P4", ""),
 ("AI-005", "S-2-2", "4-Way 상호 변환", "표현 간(Macro↔Flowchart 등) AI 변환", "변환 결과 사용자 검토·승인", "SETUP", "AI-03", "tbx_macro", "P4", ""),
 ("AI-006", "S-2-1", "UI 자동 제안", "설명→용도/항목/필요 DB Table→UI Form 초안", "", "SETUP", "AI-04", "tbx_ui_form", "P4", ""),
 ("AI-007", "§9", "사내 자료 챗봇", "도면·기술자료·서류 RAG Q&A, ERP 연동 서류 생성 지원", "내부 자료 검색·응답용", "전체", "AI-05", "(RAG 인덱스)", "P5", ""),
 ("AI-008", "-", "AI 생성물 승인 게이트", "AI 생성 Macro/UI/자료의 검토·승인 절차 강제", "미승인 생성물 사용 차단", "ADMIN", "SVC-10", "sys_approval_request", "P4", ""),
]))

MODULES.append(("APP", "모바일 App", [
 ("APP-001", "§12", "QR 정보 열람", "QR 스캔→도면/서류/Project History/처리할 업무", "권한 검사 후 표시", "전체", "FE-05, INT-05", "-", "P5", ""),
 ("APP-002", "§12", "모바일 승인", "업무 승인 처리(승인함·결재)", "", "ADMIN↑", "FE-05, SVC-10", "sys_approval_request", "P5", ""),
 ("APP-003", "§12", "Project 소통", "Project 중심 대화, History 관리", "", "전체", "FE-05, SVC-13", "-", "P5", ""),
 ("APP-004", "§12", "자재 입출고", "현장 자재 입출고 처리(MI/MO)", "QR 기반 자재 식별", "자재", "FE-05, SVC-09", "erp_process_event", "P5", ""),
 ("APP-005", "§12", "검수", "자재·완성품·설치완료 검수 처리", "사진 첨부", "QC", "FE-05, SVC-09", "erp_process_event", "P5", ""),
 ("APP-006", "§12", "유지보수", "Client 유지보수 접수·처리", "", "CS", "FE-05, SVC-09", "erp_process_event", "P5", ""),
 ("APP-007", "§12", "공지·푸시", "공지(알림) 수신, 푸시 알림", "", "전체", "FE-05, SVC-13", "sys_notification", "P5", ""),
 ("APP-008", "§12", "AR 뷰", "증강현실 — 3D 모델·Digital Twin 데이터 오버레이", "INT-02 연계", "현장", "FE-05, INT-02", "-", "P5", "후순위"),
]))

MODULES.append(("ADM", "관리 콘솔", [
 ("ADM-001", "-", "테넌트 관리", "테넌트 생성·설정·상태, 요금제(SaaS)", "", "PLATFORM", "FE-04", "sys_tenant", "P3", ""),
 ("ADM-002", "-", "사용자·권한 관리 UI", "SYS-003~005의 관리 화면", "", "ADMIN", "FE-04, SVC-01", "sys_*", "P1", ""),
 ("ADM-003", "-", "승인 관리", "승인 요청 일괄 처리·위임·규칙", "", "ADMIN", "FE-04, SVC-10", "sys_approval_request", "P2", ""),
 ("ADM-004", "-", "AI 학습 관리", "학습 데이터 등록, 파이프라인 실행·품질 확인", "Platform 전용", "PLATFORM", "FE-04, AI-02", "-", "P5", ""),
 ("ADM-005", "-", "감사·사용량", "감사 로그, API 사용량, AI 토큰 비용 조회", "", "PLATFORM/ADMIN", "FE-04, INF-07", "sys_history", "P3", ""),
]))

HEADERS = ["No", "기능 ID", "모듈", "기능코드", "기능명", "기능 설명", "주요 처리 규칙",
           "사용자 권한", "관련 컴포넌트", "관련 DB", "Phase", "비고"]
WIDTHS = [5, 10, 13, 9, 22, 46, 40, 12, 16, 24, 7, 16]


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    wb = Workbook()

    # ---- 표지 시트 ------------------------------------------------
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 기능정의서"
    ws["B2"].font = TITLE_FONT
    ws["B4"] = "EDIM (Enterprise Digital Integration Management) — CTO/ETO 통합 비즈니스 플랫폼"
    ws["B4"].font = SUB_FONT
    info = [
        ("문서 버전", "v0.2 — 발표자료 전 슬라이드 재검토 반영 (문서관리·건축설비 모듈 신설, ERP 자재/생산 상세, PLM Viewer 기능 추가)"),
        ("작성일", "2026-07-07"),
        ("기준 문서", "EDIM_개요.md / EDIM_DB_정의서.md / EDIM_컴포넌트_정의서.md"),
        ("원천 자료", "EDIM Tool System EP2.pptx (NOVA Solution, 78 슬라이드) — 전 슬라이드 시각 검토 완료"),
        ("기능 수", str(sum(len(rows) for _, _, rows in MODULES)) + "개"),
        ("모듈 수", str(len(MODULES)) + "개"),
        ("Phase 정의", "P1 RCCS코어 / P2 설계·Run / P3 원가·문서 / P4 Toolbox·AI / P5 ERP·모바일"),
        ("권한 표기", "PLATFORM > ADMIN > SETUP > GENERAL, ↑ = 해당 등급 이상"),
    ]
    r = 6
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80

    # ---- 기능목록 시트 --------------------------------------------
    ws = wb.create_sheet("기능목록")
    ws.append(HEADERS)
    style_header(ws, 1, len(HEADERS))
    no = 0
    for mi, (mkey, mname, rows) in enumerate(MODULES):
        fill = PatternFill("solid", fgColor=PALETTE[mi % len(PALETTE)])
        for row in rows:
            no += 1
            fid, code, name, desc, rule, user, comp, db, phase, note = row
            ws.append([no, fid, f"{mkey} {mname}", code, name, desc, rule, user, comp, db, phase, note])
            for c in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=ws.max_row, column=c)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = CENTER if c in (1, 2, 4, 8, 11) else WRAP
            ws.cell(row=ws.max_row, column=3).fill = fill
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # ---- 모듈요약 시트 --------------------------------------------
    ws = wb.create_sheet("모듈요약")
    phases = ["P1", "P2", "P3", "P4", "P5"]
    ws.append(["모듈", "모듈명", "기능 수"] + phases)
    style_header(ws, 1, 3 + len(phases))
    for mi, (mkey, mname, rows) in enumerate(MODULES):
        counts = {p: 0 for p in phases}
        for row in rows:
            counts[row[8]] = counts.get(row[8], 0) + 1
        ws.append([mkey, mname, len(rows)] + [counts[p] or "" for p in phases])
        for c in range(1, 4 + len(phases)):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c != 2 else WRAP
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=PALETTE[mi % len(PALETTE)])
    total = sum(len(rows) for _, _, rows in MODULES)
    ws.append(["합계", "", total] + [
        sum(1 for _, _, rows in MODULES for row in rows if row[8] == p) for p in phases])
    for c in range(1, 4 + len(phases)):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = Font(name="맑은 고딕", size=10, bold=True)
        cell.border = BORDER
        cell.alignment = CENTER
    for i, w in enumerate([8, 18, 8, 6, 6, 6, 6, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (기능 {total}개, 모듈 {len(MODULES)}개)")


if __name__ == "__main__":
    main()
