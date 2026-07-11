# -*- coding: utf-8 -*-
"""EDIM 위험관리대장.xlsx 생성 — 착수 시점 초기 위험 등록 + 운영 규칙.

등급 = 영향(상3/중2/하1) × 발생가능성(상3/중2/하1) → 6~9 심각 / 3~4 주의 / 1~2 관찰.
실행: py docs/tools/make_risk_xlsx.py  (저장: docs/EDIM_위험관리대장.xlsx)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "EDIM_위험관리대장.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

LEVEL = {"상": 3, "중": 2, "하": 1}
GRADE_FILL = {"심각": "F4C7C3", "주의": "FFF3C4", "관찰": "E3ECF7"}
STATUS_FILL = {"Open": "FFF3C4", "감시": "E3ECF7", "종결": "C8E6C9"}
CAT_FILL = {"범위": "FFF8E7", "요구": "E8F4FD", "데이터": "F0FFF0", "기술": "FFF0F5",
            "일정": "F5F0FF", "자원": "FFEFC2", "법무": "FDE9F0", "외부": "F0F1F3"}


def grade(impact, prob):
    s = LEVEL[impact] * LEVEL[prob]
    return ("심각" if s >= 6 else "주의" if s >= 3 else "관찰"), s


# (ID, 분류, 위험 내용, 근거, 영향, 가능성, 대응전략, 대응 방안·트리거, 담당, 상태, 식별일)
RISKS = [
 ("R-01", "범위", "DUCT 건축 설비 모듈 포함 여부 미확정 — 완결된 수직 모듈로 규모가 큼 (별도 견적 사안)",
  "보완노트 §1.2·§3.3", "중", "상", "회피",
  "M1(요구 확정) 전 사업 범위 결정 — 미포함 시 후속 사업으로 분리, WBS 영향 없음 확인", "PM", "Open", "2026-07-11"),
 ("R-02", "범위", "ERP 자체 구현 vs 기존 ERP 연계(INT-01) 경계 미확정 — P5 규모(51기능)에 직접 영향",
  "보완노트 §2.3·§3.3", "상", "상", "회피",
  "M1 전 결정 필수. 연계 시 대상 ERP·어댑터 스펙 확정, 자체 구현 시 P5 일정 재검토", "PM·아키텍트", "Open", "2026-07-11"),
 ("R-03", "요구", "보안 솔루션(DRM 등) 연계 범위 미정 (DOC-004) — 문서 통제 완성도에 영향",
  "보완노트 §3.3 · 보안관리계획서 §4", "중", "중", "완화",
  "자체 워터마크·Grade 통제 선행 구현(완료), 외부 솔루션은 Print/다운로드 경로 어댑터로 격리", "PM", "Open", "2026-07-11"),
 ("R-04", "법무", "Macro Excel 호환 문법의 특허 침해 가능성 (REQ-N-021, 슬라이드 27 노트)",
  "DB정의서 §12 #4", "상", "하", "완화",
  "법무 검토 선행. 자체 파서(eval 미사용)라 문법 조정 용이 — 검토 결과에 따라 함수·표기 변경", "PM·법무", "Open", "2026-07-11"),
 ("R-05", "데이터", "고객 기존 코드 체계가 RCCS 구조와 불일치 → 코드 표준화 재설계 필요 (이행 최대 위험)",
  "데이터이행계획서 §6", "상", "중", "완화",
  "W5~6 이행 대상 조사에서 표본 분석 조기 실시 — 불일치 시 코드 표준화 과업 별도 협의", "DBA·컨설턴트", "Open", "2026-07-11"),
 ("R-06", "데이터", "원천 데이터 품질 저하(중복·결측·기간 중복 단가) — 이행 공수 증가",
  "데이터이행계획서 §6", "중", "상", "완화",
  "표본 300건 품질 분석(W5), DB 제약(XOR·EXCLUDE·UQ)을 검증기로 활용, 정제 공수 별도 산정", "DBA", "Open", "2026-07-11"),
 ("R-07", "기술", "DWG 변환 손실·ODA 라이선스 미확정 — DWG 경로는 501 대기 상태",
  "구현우선순위 P4-3", "중", "중", "완화",
  "DXF 경로 우선(구현 완료). ODA 라이선스 확정 시 표본 변환 검증·손실 유형 목록화", "아키텍트", "Open", "2026-07-11"),
 ("R-08", "일정", "착수일 가정(2026-08-03) — 계약 지연 시 전 일정·투입 계획 재산정",
  "WBS 문서정보", "중", "상", "수용",
  "WBS는 START 상수만 변경해 재생성하는 체계 — 계약 확정 즉시 재산정·재배포", "PM", "Open", "2026-07-11"),
 ("R-09", "기술", "DBMS 고객 표준 상이(Oracle/MSSQL) 시 DDL·재귀 CTE·EXCLUDE 제약 이식 필요",
  "DB정의서 §12 #1", "상", "하", "완화",
  "착수 협의에서 확정. PG 전제 기능(EXCLUDE 등) 대체 설계안 사전 식별", "DBA", "Open", "2026-07-11"),
 ("R-10", "기술", "멀티테넌시 전략(컬럼/스키마/DB 분리) 미확정 — 보안(테넌트 격리)·운영에 영향",
  "DB정의서 §12 #2 · 보안관리계획서 §2.2", "중", "중", "완화",
  "설계 단계 확정. tenant_id+RLS 기본안으로 진행하되 SaaS 규모·보안 요건 확인 후 결정", "아키텍트", "Open", "2026-07-11"),
 ("R-11", "기술", "성능 목표 미충족 위험 (REQ-N-001~005: 산출물 1시간·BOM 30초·응답 2/3초)",
  "요구사항정의서", "중", "하", "완화",
  "PoC에서 BOM 전개·Run 파이프라인 실측 완료 — W35 성능시험으로 최종 검증", "QA·인프라", "감시", "2026-07-11"),
 ("R-12", "자원", "투입 인력(9~11명, AI·CAD 특수 역량 포함) 미확정 — 협의 대상",
  "WBS 투입(안)", "상", "중", "완화",
  "계약 시 확정. AI·CAD 등 희소 역량은 조기 확보, PoC 코드 인수인계로 램프업 단축", "PM", "Open", "2026-07-11"),
 ("R-13", "외부", "Digital Twin(DTDesigner) 연계 스펙·외부 ERP 어댑터 대상 미정",
  "구현우선순위 협의 대기", "하", "중", "수용",
  "P5/후속 사업 범위로 관리 — 스펙 협의 완료 전 인터페이스 지점만 예약", "아키텍트", "감시", "2026-07-11"),
 ("R-14", "데이터", "이행 기간 중 원천 데이터 변경 — 이행본과 운영본 불일치",
  "데이터이행계획서 §6", "중", "중", "완화",
  "기준일(cut-off) 합의 + 증분 이행 1회 계획 반영", "DBA·고객사", "Open", "2026-07-11"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM 위험관리대장"
    ws["B2"].font = TITLE_FONT
    counts = {}
    for r in RISKS:
        g, _ = grade(r[4], r[5])
        counts[g] = counts.get(g, 0) + 1
    info = [
        ("문서 버전", "v0.1"),
        ("작성일", "2026-07-11"),
        ("목적", "사업 위험의 식별·평가·대응·감시 — 주간 보고 시 갱신, 심각 등급은 운영위원회 상정"),
        ("평가 기준", "등급 = 영향(상3/중2/하1) × 발생가능성(상3/중2/하1) → 6~9 심각 / 3~4 주의 / 1~2 관찰"),
        ("대응 전략", "회피(원인 제거) / 완화(영향·확률 축소) / 전가(제3자 이전) / 수용(감시만)"),
        ("현황", f"위험 {len(RISKS)}건 — " + " · ".join(f"{k} {v}" for k, v in sorted(counts.items()))),
        ("운영 규칙", "신규 위험은 본 스크립트 RISKS에 추가 후 재생성 — xlsx 직접 편집 금지. 종결 위험도 행 유지(이력)"),
        ("관련 문서", "사업수행계획서 §4.4 · 보안관리계획서 · 데이터이행계획서 §6 · DB정의서 §12 미결정"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 110

    ws = wb.create_sheet("위험대장")
    headers = ["ID", "분류", "위험 내용", "근거(문서)", "영향", "가능성", "점수", "등급",
               "대응전략", "대응 방안·트리거", "담당", "상태", "식별일"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for row in RISKS:
        rid, cat, desc, src, impact, prob, strategy, plan, owner, status, found = row
        g, score = grade(impact, prob)
        ws.append([rid, cat, desc, src, impact, prob, score, g, strategy, plan, owner, status, found])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 5, 6, 7, 8, 9, 11, 12, 13) else WRAP
        ws.cell(row=ws.max_row, column=2).fill = PatternFill("solid", fgColor=CAT_FILL.get(cat, "FFFFFF"))
        ws.cell(row=ws.max_row, column=8).fill = PatternFill("solid", fgColor=GRADE_FILL[g])
        ws.cell(row=ws.max_row, column=12).fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "FFFFFF"))
    for i, w in enumerate([6, 7, 46, 18, 5, 6, 5, 6, 8, 46, 11, 6, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    wb.save(OUT)
    print(f"saved: {os.path.abspath(OUT)}  (위험 {len(RISKS)}건: {counts})")


if __name__ == "__main__":
    main()
