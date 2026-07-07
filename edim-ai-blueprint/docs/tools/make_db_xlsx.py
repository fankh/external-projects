# -*- coding: utf-8 -*-
"""EDIM DB정의서.xlsx 생성 — EDIM_DB_정의서.md를 파싱하여 Excel로 변환.

실행: py docs/tools/make_db_xlsx.py  (저장: docs/EDIM_DB정의서.xlsx)
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..")
SRC = os.path.join(BASE, "EDIM_DB_정의서.md")
OUT = os.path.join(BASE, "EDIM_DB정의서.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1A1A40")
DOM_FILL = PatternFill("solid", fgColor="2E8B57")
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1A1A40")
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def parse_md():
    text = open(SRC, encoding="utf-8").read()
    lines = text.split("\n")
    tables = []      # {domain, section, name, title, cols: [(컬럼,타입,Null,키,설명)]}
    common_codes = []  # (코드그룹, 값, 사용테이블)
    open_issues = []   # (#, 항목, 내용)
    domain = ""
    i = 0
    cur = None

    def md_row(line):
        return [c.strip() for c in line.strip().strip("|").split("|")]

    while i < len(lines):
        line = lines[i]
        m = re.match(r"^## \d+\. (.+)$", line)
        if m:
            domain = m.group(1).strip()
        m = re.match(r"^### ([\d.]+) `(\w+)` — (.+)$", line)
        if m:
            cur = {"domain": domain, "section": m.group(1), "name": m.group(2),
                   "title": m.group(3).strip(), "cols": []}
            tables.append(cur)
            # 다음 md 표 탐색
            j = i + 1
            while j < len(lines) and not lines[j].startswith("| 컬럼"):
                if lines[j].startswith("### ") or lines[j].startswith("## "):
                    break
                j += 1
            if j < len(lines) and lines[j].startswith("| 컬럼"):
                j += 2  # 헤더·구분선 스킵
                while j < len(lines) and lines[j].startswith("|"):
                    cells = md_row(lines[j])
                    if len(cells) >= 5:
                        cur["cols"].append(cells[:5])
                    j += 1
                i = j
                continue
        # 공통 코드 (§10)
        if line.startswith("## 10. 공통 코드"):
            j = i + 1
            while j < len(lines) and not lines[j].startswith("| 코드 그룹"):
                j += 1
            j += 2
            while j < len(lines) and lines[j].startswith("|"):
                cells = md_row(lines[j])
                if len(cells) >= 3:
                    common_codes.append(cells[:3])
                j += 1
        # 미결정 (§12)
        if line.startswith("## 12. 미결정"):
            j = i + 1
            while j < len(lines) and not lines[j].startswith("| #"):
                j += 1
            j += 2
            while j < len(lines) and lines[j].startswith("|"):
                cells = md_row(lines[j])
                if len(cells) >= 3:
                    open_issues.append(cells[:4] if len(cells) >= 4 else cells[:3] + [""])
                j += 1
        i += 1
    return tables, common_codes, open_issues


def clean(s):
    """markdown 서식 제거"""
    s = re.sub(r"`([^`]*)`", r"\1", s)
    s = re.sub(r"\*\*([^*]*)\*\*", r"\1", s)
    s = s.replace("→", "->").replace("&amp;", "&")
    return s


def style_header(ws, row, cols, fill=HDR_FILL):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    tables, common_codes, open_issues = parse_md()
    wb = Workbook()

    # ---- 문서정보 --------------------------------------------------
    ws = wb.active
    ws.title = "문서정보"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "EDIM DB정의서"
    ws["B2"].font = TITLE_FONT
    info = [
        ("문서 버전", "v0.1 (초안)"), ("작성일", "2026-07-07"),
        ("대상 DBMS", "PostgreSQL 16 (가정 — 미확정)"),
        ("기준 문서", "EDIM_DB_정의서.md (본 파일은 MD에서 자동 생성)"),
        ("테이블 수", f"{len(tables)}개"),
        ("명명 규칙", "snake_case, 도메인 접두어, PK=<엔티티>_id BIGINT IDENTITY"),
        ("공통 컬럼", "tenant_id, created_by/at, updated_by/at (본 문서 표에서는 생략)"),
        ("승인 패턴", "approval_status: DRAFT→PENDING→APPROVED/REJECTED"),
    ]
    r = 4
    for k, v in info:
        ws.cell(row=r, column=2, value=k).font = BOLD_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        r += 1
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80

    # ---- 테이블목록 ------------------------------------------------
    ws = wb.create_sheet("테이블목록")
    ws.append(["No", "도메인", "테이블명", "테이블 설명", "컬럼 수", "섹션"])
    style_header(ws, 1, 6)
    for n, t in enumerate(tables, 1):
        ws.append([n, clean(t["domain"]), t["name"], clean(t["title"]), len(t["cols"]), t["section"]])
        for c in range(1, 7):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 5, 6) else WRAP
    for i, w in enumerate([5, 34, 24, 44, 8, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    # ---- 컬럼정의 (전체 단일 시트) ---------------------------------
    ws = wb.create_sheet("컬럼정의")
    headers = ["테이블명", "No", "컬럼명", "타입", "Null", "키/제약", "설명"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for t in tables:
        # 테이블 구분 행
        ws.append([t["name"], "", clean(t["title"]), "", "", "", f"도메인: {clean(t['domain'])}"])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.fill = DOM_FILL
            cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            cell.border = BORDER
            cell.alignment = WRAP
        for n, col in enumerate(t["cols"], 1):
            name, typ, nul, key, desc = (clean(x) for x in col)
            ws.append([t["name"], n, name, typ, nul, key, desc])
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=c)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = CENTER if c in (2, 5) else WRAP
    for i, w in enumerate([22, 5, 24, 16, 6, 26, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    # ---- 공통코드 --------------------------------------------------
    ws = wb.create_sheet("공통코드")
    ws.append(["No", "코드 그룹", "값", "사용 테이블"])
    style_header(ws, 1, 4)
    for n, row in enumerate(common_codes, 1):
        ws.append([n] + [clean(x) for x in row])
        for c in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c == 1 else WRAP
    for i, w in enumerate([5, 22, 56, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- 미결정사항 ------------------------------------------------
    ws = wb.create_sheet("미결정사항")
    ws.append(["#", "항목", "내용", "담당/기한"])
    style_header(ws, 1, 4)
    for row in open_issues:
        ws.append([clean(x) for x in row])
        for c in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c == 1 else WRAP
    for i, w in enumerate([5, 20, 70, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    total_cols = sum(len(t["cols"]) for t in tables)
    print(f"saved: {os.path.abspath(OUT)}  (테이블 {len(tables)}개, 컬럼 {total_cols}개, "
          f"공통코드 {len(common_codes)}행, 미결정 {len(open_issues)}건)")


if __name__ == "__main__":
    main()
