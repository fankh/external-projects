# -*- coding: utf-8 -*-
"""거버넌스 정의서 드리프트 게이트 (요구 #71).

문서가 코드와 어긋나면 실패한다. 손으로 고친 문서는 다음 커밋에서 바로 낡기 때문에,
docs/EDIM_거버넌스정의서.xlsx 는 tools/gen_governance.py 의 산출물이어야 한다.

  py tests/check_governance.py          # 코드와 문서가 같은지 검사
  py tools/gen_governance.py            # 어긋났을 때 재생성
"""
from __future__ import annotations

import pathlib
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

from gen_governance import OUT, build  # noqa: E402


def read_committed() -> dict[str, list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(OUT, read_only=True)
    out: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        rows = []
        for r in wb[name].iter_rows(values_only=True):
            if r is None:
                continue
            rows.append(["" if c is None else str(c) for c in r])
        out[name] = rows
    return out


def main() -> int:
    if not OUT.exists():
        print(f"FAIL — 문서 없음: {OUT.relative_to(ROOT)} → py tools/gen_governance.py")
        return 1
    expected = build()
    actual = read_committed()

    problems: list[str] = []
    for sheet, rows in expected.items():
        if sheet not in actual:
            problems.append(f"시트 누락: {sheet}")
            continue
        got = actual[sheet]
        if len(got) != len(rows):
            problems.append(f"{sheet}: 행 수 불일치 (문서 {len(got)} · 코드 {len(rows)})")
        for i, (a, b) in enumerate(zip(got, rows), start=1):
            if [str(x) for x in a] != [str(x) for x in b]:
                problems.append(f"{sheet} {i}행: 문서={a[:3]} / 코드={b[:3]}")
                if len([p for p in problems if p.startswith(sheet)]) > 3:
                    break
    for sheet in actual:
        if sheet not in expected:
            problems.append(f"코드에 없는 시트: {sheet}")

    if problems:
        print("FAIL — 거버넌스 정의서가 코드와 어긋납니다 (#71)")
        for p in problems[:12]:
            print(f"  · {p}")
        if len(problems) > 12:
            print(f"  … 외 {len(problems) - 12}건")
        print("\n  재생성: py tools/gen_governance.py")
        return 1

    counts = " · ".join(f"{k} {len(v) - 1}" for k, v in expected.items())
    print(f"PASS — 거버넌스 정의서가 코드와 일치 ({counts})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
