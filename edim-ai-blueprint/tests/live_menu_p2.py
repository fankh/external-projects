# -*- coding: utf-8 -*-
"""메뉴정의서 P2 라이브 — 공지 발송(v34.47) · 번역 일괄 Export/Import(v34.48).

실행: PYTHONUTF8=1 py tests/live_menu_p2.py
정리: 공지 알림 psql 삭제, 번역은 원값 재저장 원복, 임시 xlsx 삭제.
"""
import io
import json
import os
import subprocess
import tempfile
import urllib.request
from urllib.parse import quote

import openpyxl

BASE = os.getenv("BASE", "https://edim.seekerslab.com/").rstrip("/")
API = f"{BASE}/api/v1"
ANN = "TEST-ANN 공지 검증"
TESTVAL = "EN-IMPORT-TEST"
n = 0


def ok(label, cond):
    global n
    assert cond, f"FAIL {label}"
    n += 1
    print(f"PASS {label}")


def psql(sql):
    subprocess.run(["ssh", "edim-server",
                    f"sudo docker exec edim-postgres psql -U edim -d edim -tAc \"{sql}\""],
                   capture_output=True, text=True, timeout=40)


def login(uid, pw):
    r = urllib.request.Request(f"{API}/auth/login",
        data=json.dumps({"userId": uid, "password": pw}).encode(),
        headers={"Content-Type": "application/json"}, method="POST")
    return json.loads(urllib.request.urlopen(r).read())["token"]


def req(tok, method, path, body=None, raw=False):
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(API + quote(path, safe="/?=&%"), data=data,
        headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}, method=method)
    with urllib.request.urlopen(r) as resp:
        payload = resp.read()
        return payload if raw else json.loads(payload or b"null")


TOK = login("edim", "edim")
psql("DELETE FROM sys_notification WHERE notify_type='ANNOUNCE' AND title LIKE 'TEST-ANN%'")
d0 = req(TOK, "GET", "/notifications/digest")["unread"]

rows0 = req(TOK, "GET", "/i18n/data/COMPANY?locale=en")
assert rows0, "COMPANY 행 필요"
target = rows0[0]
orig = target["value"]

# Export 형식 (API 직접)
wb = openpyxl.load_workbook(io.BytesIO(req(TOK, "GET", "/i18n/data/COMPANY/export.xlsx", raw=True)))
ws = wb.active
ok("번역 Export 헤더", [c.value for c in ws[1]] == ["ID", "원문", "en", "ja", "zh"])
ok("번역 Export 행 수", ws.max_row - 1 == len(rows0))

wbi = openpyxl.Workbook()
wsi = wbi.active
wsi.append(["ID", "en"])
wsi.append([target["entityId"], TESTVAL])
imp_path = os.path.join(tempfile.gettempdir(), "edim_i18n_import_test.xlsx")
wbi.save(imp_path)

from playwright.sync_api import sync_playwright  # noqa: E402

try:
    with sync_playwright() as pw:
        b = pw.chromium.launch()
        p = b.new_page(viewport={"width": 1600, "height": 900})
        p.goto(f"{BASE}/login", wait_until="networkidle")
        p.fill("input[name=userId]", "edim")
        p.fill("input[name=password]", "edim")
        p.get_by_role("button", name="로그인 (Enter)").click()
        p.wait_for_url("**/erp/**", timeout=15000)

        # 1) 공지 — 벨 📢 작성 → 발송 → 본인·타 사용자 수신 → GENERAL 403
        p.locator("button[title='알림']").click()
        p.wait_for_timeout(400)
        ok("ADMIN 공지 버튼", p.locator("[data-announce-open]").count() == 1)
        p.locator("[data-announce-open]").click()
        p.wait_for_selector("[data-announce-form]", timeout=5000)
        p.locator("input[aria-label='공지 제목']").fill(ANN)
        p.locator("[data-announce-send]").click()
        p.locator("text=공지 발송 ✓").wait_for(timeout=10000)
        ok("공지 발송", req(TOK, "GET", "/notifications/digest")["unread"] == d0 + 1)
        tok2 = login("jang.s", "edim")
        ok("타 사용자 수신", any(x["title"] == ANN and x["type"] == "ANNOUNCE"
                            for x in req(tok2, "GET", "/notifications?limit=5")))
        try:
            req(tok2, "POST", "/notifications/announce", {"title": "x"})
            ok("GENERAL 발송 403", False)
        except urllib.error.HTTPError as e:
            ok("GENERAL 발송 403", e.code == 403)

        # 2) 번역 일괄 Import — UI 업로드 → 반영
        p.goto(f"{BASE}/code/data-i18n?entity=COMPANY&locale=en", wait_until="networkidle")
        p.wait_for_timeout(600)
        ok("⬇⬆ 버튼", p.locator("[data-i18n-export]").count() == 1 and p.locator("[data-i18n-import]").count() == 1)
        p.locator("[data-i18n-import] input[type=file]").set_input_files(imp_path)
        p.locator("text=일괄 Import ✓").wait_for(timeout=15000)
        got = next(x for x in req(TOK, "GET", "/i18n/data/COMPANY?locale=en")
                   if x["entityId"] == target["entityId"])
        ok("Import 반영", got["value"] == TESTVAL)
        b.close()
finally:
    req(TOK, "PUT", "/i18n/data", {"entityType": "COMPANY", "entityId": target["entityId"],
                                    "locale": "en", "value": orig})
    psql("DELETE FROM sys_notification WHERE notify_type='ANNOUNCE' AND title LIKE 'TEST-ANN%'")
    try:
        os.remove(imp_path)
    except OSError:
        pass
    print("정리 — 번역 원복·공지 psql·임시 파일 삭제", flush=True)

print(f"\nlive_menu_p2: {n}/{n} PASS")
