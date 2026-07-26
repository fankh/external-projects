# -*- coding: utf-8 -*-
"""API 계약 드리프트 게이트 (18.6).

`docs/EDIM_컴포넌트정의서.xlsx` 의 API 시트는 **배포된 스펙(as-built)** 에서 생성한다.
손으로 관리하면 낡는다 — 실제로 종전 API목록 108건 중 **42건만** 라우터에 남아 있었고,
그 목록에서 만든 OpenAPI(95경로)도 같은 괴리를 물려받았다. 문서를 보고 연동하면 404 다.

두 가지를 본다.
1) 문서 API 시트 = 생성기 산출물
2) as-built 스냅샷 = 배포된 스펙 (서버에 닿을 때만. 스냅샷만 맞추면 "문서는 스냅샷과 맞는데
   스냅샷이 낡은" 상태를 통과시킨다 — 종전이 바로 그 상태였다)

  py tests/check_api_contract.py
  py tests/check_api_contract.py --offline
  py tools/gen_api_contract.py --dump && py tools/gen_api_contract.py   # 갱신
"""
from __future__ import annotations

import json
import pathlib
import subprocess
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

from gen_api_contract import METHODS, OUT, SNAP, build  # noqa: E402


def read_committed() -> dict[str, list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(OUT, read_only=True)
    out: dict[str, list[list[str]]] = {}
    for name in ("API목록", "설계-구현 대조"):
        if name not in wb.sheetnames:
            continue
        rows = []
        for r in wb[name].iter_rows(values_only=True):
            if r is None:
                continue
            rows.append(["" if c is None else str(c) for c in r])
        out[name] = rows
    return out


def live_op_count() -> int | None:
    code = ("import json;from app.main import app;"
            "s=app.openapi();"
            "print(sum(len([m for m in v if m in "
            "('get','post','put','patch','delete')]) for v in s['paths'].values()))")
    try:
        r = subprocess.run(
            ["ssh", "edim-server", f'sudo docker exec edim-backend python -c "{code}"'],
            capture_output=True, text=True, timeout=120, encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        return None
    for line in reversed((r.stdout or "").splitlines()):
        if line.strip().isdigit():
            return int(line.strip())
    return None


def main() -> int:
    if not OUT.exists():
        print(f"FAIL — 문서 없음: {OUT.relative_to(ROOT)}")
        return 1
    expected = build()
    actual = read_committed()
    problems: list[str] = []
    for sheet, rows in expected.items():
        if sheet not in actual:
            problems.append(f"시트 누락: {sheet}")
            continue
        got = actual[sheet]
        if len(got) != len(rows):
            problems.append(f"{sheet}: 행 수 불일치 (문서 {len(got)} · 생성기 {len(rows)})")
        for i, (a, b) in enumerate(zip(got, rows), start=1):
            if [str(x) for x in a] != [str(x) for x in b]:
                problems.append(f"{sheet} {i}행: 문서={a[:4]} / 생성기={b[:4]}")
                if len([p for p in problems if p.startswith(sheet)]) > 3:
                    break

    spec = json.loads(SNAP.read_text(encoding="utf-8"))
    snap_ops = sum(len([m for m in v if m in METHODS]) for v in spec["paths"].values())
    note = ""
    if "--offline" in sys.argv:
        note = " · 배포 스펙 대조 생략(--offline)"
    else:
        live = live_op_count()
        if live is None:
            note = " · 배포 스펙 대조 못함(서버 미도달 — 통과로 읽지 말 것)"
        elif live != snap_ops:
            problems.append(f"스냅샷이 배포 스펙과 다름 — 배포 {live} · 스냅샷 {snap_ops} "
                            "→ py tools/gen_api_contract.py --dump")
        else:
            note = f" · 배포 스펙 일치(오퍼레이션 {live})"

    # 미분류가 남으면 서비스 구분 규칙을 보완해야 한다 — 문서·생성기가 똑같이 미분류면
    # 위 비교는 통과하므로 별도로 잡는다(17.2 의 경계정의와 같은 이유).
    un = [r[3] for r in expected["API목록"][1:] if r[1] == "(미분류)"]
    if un:
        problems.append(f"API목록 미분류 {len(un)}건: {', '.join(un[:6])}"
                        f"{' …' if len(un) > 6 else ''} → gen_api_contract.py 의 SVC 보완")

    if problems:
        print("FAIL — API 계약 문서가 배포 스펙과 어긋납니다 (18.6)")
        for p in problems[:12]:
            print(f"  · {p}")
        print("\n  갱신: py tools/gen_api_contract.py --dump && py tools/gen_api_contract.py")
        return 1
    print(f"PASS — API 계약 문서가 배포 스펙과 일치 (경로 {len(spec['paths'])} · "
          f"오퍼레이션 {snap_ops}){note}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
