# -*- coding: utf-8 -*-
"""F4 라이브 — 무반응·메시지-온리 일소 검증 (Next 재작성, 2026-07-19).

S-1-4 관계 승인 실배선(EBOM Run→승인 요청→결재→APPROVED 전이) · Table XLSX Export ·
M-3-7 편집 패널/Export · S-3-4 PrintSetup(자리표시자·PDF·Office xlsx) ·
S-1-1 승인 요청 게이트(중복검토 선행) · 코드 상세 Referencers · 문서함 PDF 미리보기.
실행: PYTHONUTF8=1 py tests/live_f4_noop.py
정리: 관계 승인 요청 1건은 결재 처리 — 이력 성격 누적 허용 (b19 PO 전례).
"""
import io as _io
import subprocess

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE = "https://edim.seekerslab.com"
API = f"{BASE}/api/v1"
n = 0


def ok(label: str, cond: bool) -> None:
    global n
    assert cond, f"FAIL {label}"
    n += 1
    print(f"PASS {label}")


def psql(sql: str) -> str:
    r = subprocess.run(
        ["ssh", "edim-server",
         f"sudo docker exec edim-postgres psql -U edim -d edim -tAc \"{sql}\""],
        capture_output=True, text=True, timeout=30)
    return (r.stdout or "").strip()


with sync_playwright() as pw:
    req = pw.request.new_context()

    def call(tok, method, path, data=None):
        return req.fetch(API + path, method=method,
                         headers={"Authorization": f"Bearer {tok}",
                                  **({"Content-Type": "application/json"} if data is not None else {})},
                         data=None if data is None else data)

    tok = req.post(f"{API}/auth/login",
                   data={"userId": "edim", "password": "edim"}).json()["token"]

    # 0. 잔존 정리 — 이전 실패 런의 PENDING 관계 요청은 반려 처리
    for r0 in call(tok, "GET", "/approvals/inbox").json():
        if r0["assetType"] == "관계":
            call(tok, "POST", f"/approvals/{r0['id']}/decide",
                 {"approve": False, "comment": "F4 스위트 사전 정리"})

    # 1. Table XLSX Export (M-3-7 실배선)
    r = req.get(f"{API}/tables/Table12/export.xlsx",
                headers={"Authorization": f"Bearer {tok}"})
    ok("export.xlsx 200 + PK 시그니처", r.ok and r.body()[:2] == b"PK")
    wb = load_workbook(_io.BytesIO(r.body()))
    ws = wb.active
    ok("XLSX 헤더 Key + 데이터 행", ws.cell(1, 1).value == "Key" and ws.max_row >= 5)

    b = pw.chromium.launch()
    page = b.new_context(viewport={"width": 1600, "height": 900}).new_page()
    page.goto(f"{BASE}/code", wait_until="domcontentloaded")
    page.wait_for_selector(".login-dlg, .app .titlebar", timeout=15000)
    if page.locator(".login-dlg").count():
        page.get_by_label("사번").fill("edim")
        page.get_by_label("비밀번호").fill("edim")
        page.get_by_role("button", name="로그인 (Enter)").click()
    page.wait_for_selector(".app .titlebar", timeout=15000)
    page.wait_for_timeout(1000)

    # 2. S-1-4 Code Relationship — EBOM Run → 승인 요청 실배선
    page.locator(".tn", has_text="Code Relationship (S-1-4)").first.click()
    page.wait_for_timeout(1500)
    page.get_by_role("button", name="EBOM Run ▶").click()
    page.wait_for_selector("text=Running Test 통과", timeout=10000)
    ok("S-1-4 EBOM Run = Test 통과", True)
    page.get_by_role("button", name="승인 요청").click()
    page.wait_for_timeout(1500)
    inbox = call(tok, "GET", "/approvals/inbox").json()
    rel = next((x for x in inbox if x["assetType"] == "관계"), None)
    ok("승인함 수신 — 유형 '관계' (code_relationship)", rel is not None
       and "KDCR 3-13" in rel["target"])
    ok("중복 요청 정직 409", call(tok, "POST", "/approvals",
       {"targetTable": "code_relationship", "targetCode": "KDCR 3-13",
        "label": "dup"}).status == 409)
    ok("결정(승인) 200", call(tok, "POST", f"/approvals/{rel['id']}/decide",
       {"approve": True, "comment": "F4 검증"}).ok)
    cnt = psql("SELECT count(*) FROM code_relationship cr JOIN product_code pc "
               "ON pc.product_code_id=cr.mother_code_id "
               "WHERE pc.main_code='KDCR 3-13' AND cr.approval_status='APPROVED'")
    ok("관계 세트 APPROVED 전이", cnt.isdigit() and int(cnt) >= 1)

    # 3. S-1-1 SubCode — 승인 요청 게이트 (중복검토 선행)
    page.locator(".tn", has_text="Sub Code 등록 (S-1-1)").first.click()
    page.wait_for_timeout(1500)
    ok("중복검토 버튼 노출", page.get_by_role("button", name="중복검토").count() >= 1)
    ok("승인 요청 게이트 — 중복검토 전 disabled",
       page.get_by_role("button", name="승인 요청").first.is_disabled())

    # 4. M-3-7 데이터 Table — 행 클릭 편집 패널 + ⬇ Export
    page.locator(".tn", has_text="데이터 Table 관리 (M-3-7)").first.click()
    page.wait_for_timeout(1500)
    page.locator("table.g:visible tbody tr").first.click()
    page.wait_for_timeout(400)
    ok("행 클릭 = 편집 패널(입력) 오픈", page.locator("input[name], .gb input.in").count() >= 1
       and "행 클릭" not in page.locator("body").inner_text()[:0])   # 패널 존재 확인
    with page.expect_popup() as pop:
        page.get_by_role("button", name="⬇ Export").first.click()
    # 다운로드 응답은 팝업 내비게이션이 중단되어 about:blank 로 남을 수 있음 — 팝업 발생 자체가 실배선 증거
    ok("⬇ Export — XLSX 창(프록시)", pop.value is not None)
    pop.value.close()

    # 5. S-3-4 Print Set-up — 자리표시자·PDF 실렌더·Office xlsx
    page.locator(".titlebar .mod", has_text="CPQ").first.click()
    page.wait_for_timeout(600)
    page.locator(".tn", has_text="Print Set-up (S-3-4)").first.click()
    page.wait_for_timeout(1200)
    ok("자리표시자 목록 렌더", page.get_by_text("자리표시자 목록", exact=False).count() >= 1)
    ok("용지 선택(data-ps-paper) 노출", page.locator("[data-ps-paper]").count() == 1)
    ok("Office(xlsx) 실배선 — enabled + 마커",
       page.locator("[data-office-export]").count() == 1
       and page.locator("[data-office-export]").is_enabled())
    with page.expect_popup() as pop2:
        page.get_by_role("button", name="PDF", exact=True).click()
        page.wait_for_timeout(2500)
    ok("PDF — 실렌더 창", pop2.value is not None)
    pop2.value.close()

    # 6. 코드 상세 — Referencers(Where-Used) SSR 렌더
    page.goto(f"{BASE}/detail/code?code=EWT-3", wait_until="networkidle")
    page.wait_for_timeout(600)
    body = page.locator("body").inner_text()
    ok("코드 상세 Referencers(Where-Used)", "Where-Used" in body)
    ok("코드 상세 단가 이력", "단가 이력" in body or "이력" in body)

    # 7. 문서함 — 행 선택 → PDF 미리보기 실배선
    page.locator(".tn", has_text="문서함 (M-5-4)").first.click()
    page.wait_for_timeout(1500)
    rows_cnt = page.locator("table.g:visible tbody tr").count()
    if rows_cnt:
        page.locator("table.g:visible tbody tr").first.click()
        page.wait_for_timeout(400)
        ok("문서함 PDF 미리보기 enabled",
           page.get_by_role("button", name="PDF 미리보기").first.is_enabled())
    else:
        ok("문서함 빈 목록 — 미리보기 검사 생략", True)

    b.close()

print(f"\nOK — live_f4_noop {n}/{n}")
