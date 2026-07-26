# -*- coding: utf-8 -*-
"""DB 정의서 드리프트 게이트 (18.2).

두 가지를 본다.

1) **문서가 생성기 산출물과 같은가** — 손으로 고친 DB 정의서는 다음 마이그레이션에서 바로
   낡는다. 거버넌스 정의서(#71)와 같은 규칙이다.
2) **스냅샷이 실제 스키마와 같은가** — 서버에 닿을 때만 검사한다. 스냅샷만 맞추면
   "문서는 스냅샷과 일치하는데 스냅샷이 낡은" 상태를 통과시키게 된다. 종전 DB 정의서가
   정확히 그 상태였다(설계 초안 기준 54 테이블, 실제 107).

  py tests/check_db_dict.py            # 문서 = 생성기 산출물 (+ 서버 도달 시 스키마 대조)
  py tests/check_db_dict.py --offline  # 스키마 대조 생략(네트워크 없는 환경)
  py tools/gen_db_dict.py --dump && py tools/gen_db_dict.py   # 어긋났을 때 갱신
"""
from __future__ import annotations

import json
import pathlib
import subprocess
import sys

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

from gen_db_dict import OUT, SNAP, build  # noqa: E402


def read_committed() -> dict[str, list[list[str]]]:
    import openpyxl
    wb = openpyxl.load_workbook(OUT, read_only=True)
    out: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        rows = []
        for r in wb[name].iter_rows(values_only=True):
            if r is None:
                continue
            rows.append(["" if c is None else str(c) for c in r])
        out[name] = rows
    return out


def live_counts() -> tuple[int, int] | None:
    """서버 실제 테이블·컬럼 수. 닿지 않으면 None (그 사실을 호출부가 밝힌다)."""
    sql = ("SELECT count(DISTINCT table_name)||' '||count(*) FROM information_schema.columns "
           "WHERE table_schema='public' AND table_name<>'alembic_version'")
    try:
        r = subprocess.run(
            ["ssh", "edim-server",
             f'sudo docker exec edim-postgres psql -U edim -d edim -tAc "{sql}"'],
            capture_output=True, text=True, timeout=60, encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        return None
    for line in (r.stdout or "").splitlines():
        parts = line.strip().split()
        if len(parts) == 2 and all(p.isdigit() for p in parts):
            return int(parts[0]), int(parts[1])
    return None


def main() -> int:
    if not OUT.exists():
        print(f"FAIL — 문서 없음: {OUT.relative_to(ROOT)} → py tools/gen_db_dict.py")
        return 1
    expected = build()
    actual = read_committed()
    problems: list[str] = []
    for sheet, rows in expected.items():
        if sheet not in actual:
            problems.append(f"시트 누락: {sheet}")
            continue
        got = actual[sheet]
        if len(got) != len(rows):
            problems.append(f"{sheet}: 행 수 불일치 (문서 {len(got)} · 생성기 {len(rows)})")
        for i, (a, b) in enumerate(zip(got, rows), start=1):
            if [str(x) for x in a] != [str(x) for x in b]:
                problems.append(f"{sheet} {i}행: 문서={a[:3]} / 생성기={b[:3]}")
                if len([p for p in problems if p.startswith(sheet)]) > 3:
                    break
    for sheet in actual:
        if sheet not in expected:
            problems.append(f"생성기에 없는 시트: {sheet}")

    snap = json.loads(SNAP.read_text(encoding="utf-8"))
    snap_t = len({c["t"] for c in snap["columns"]})
    snap_c = len(snap["columns"])
    live_note = ""
    if "--offline" in sys.argv:
        live_note = " · 스키마 대조 생략(--offline)"
    else:
        live = live_counts()
        if live is None:
            live_note = " · 스키마 대조 못함(서버 미도달 — 통과로 읽지 말 것)"
        elif live != (snap_t, snap_c):
            problems.append(
                f"스냅샷이 실제 스키마와 다름 — 실제 테이블 {live[0]}·컬럼 {live[1]} / "
                f"스냅샷 {snap_t}·{snap_c} → py tools/gen_db_dict.py --dump")
        else:
            live_note = f" · 실 스키마 일치(테이블 {live[0]} · 컬럼 {live[1]})"

    if problems:
        print("FAIL — DB 정의서가 스키마와 어긋납니다 (18.2)")
        for p in problems[:12]:
            print(f"  · {p}")
        print("\n  갱신: py tools/gen_db_dict.py --dump && py tools/gen_db_dict.py")
        return 1
    print(f"PASS — DB 정의서가 스키마와 일치 (테이블 {snap_t} · 컬럼 {snap_c} · "
          f"제약 {len(snap['constraints'])}){live_note}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
